
# =========================================================
# SAÚDE KIDS - BETA (VERSÃO COMPLETA, ESTÁVEL)
# - Login (SQLite) + Cookie opcional
# - Admin: usuários, métricas, sugestões, backup/restore
# - Glicemia: momentos extras (personalizado) + CRUD (editar/excluir no histórico)
# - Excel: Glicemia_Tabela (tudo) + Glicemia_Resumo (pivot) + Nutrição
# =========================================================

import os
import re
import random
import shutil
import sqlite3
import string
import uuid
import zipfile
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

import pandas as pd
import plotly.express as px
try:
    import matplotlib.pyplot as plt
    HAS_MPL = True
except Exception:
    HAS_MPL = False

import pytz
import streamlit as st
from supabase import create_client
from email.mime.text import MIMEText
from openpyxl.styles import Alignment, PatternFill
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    HAS_REPORTLAB = True
except Exception:
    HAS_REPORTLAB = False


# =========================================================
# (OPCIONAL) LOGIN PERSISTENTE POR COOKIE (NÃO BUGA SE NÃO TIVER)
# Para ativar no Streamlit Cloud: requirements.txt -> extra-streamlit-components==0.1.60
# =========================================================
try:
    import extra_streamlit_components as stx  # type: ignore
    _cookie_manager = stx.CookieManager()
    HAS_COOKIES = True
except Exception:
    _cookie_manager = None
    HAS_COOKIES = False

COOKIE_KEY = "SK_LOGIN_EMAIL"
COOKIE_DIAS = 30

# ================= CONFIGURAÇÕES INICIAIS =================
fuso_br = pytz.timezone("America/Sao_Paulo")
st.set_page_config(page_title="Saúde Kids BETA", page_icon="🧪", layout="wide")

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
BACKUP_DIR = BASE_DIR / "backups"
DATA_DIR.mkdir(parents=True, exist_ok=True)
BACKUP_DIR.mkdir(parents=True, exist_ok=True)

ARQ_G = DATA_DIR / "dados_glicemia_BETA.csv"
ARQ_N = DATA_DIR / "dados_nutricao_BETA.csv"
ARQ_R = DATA_DIR / "config_receita_BETA.csv"
ARQ_M = DATA_DIR / "mensagens_admin_BETA.csv"
ARQ_M_ROOT = BASE_DIR / "mensagens_admin_BETA.csv"
DB_USERS = DATA_DIR / "usuarios.db"
APP_DB = DATA_DIR / "saude_kids.db"

# ================= SUPABASE =================
def _get_supabase_client():
    try:
        url = st.secrets["SUPABASE_URL"]
        key = st.secrets["SUPABASE_KEY"]
        return create_client(url, key)
    except Exception:
        st.error("Configure SUPABASE_URL e SUPABASE_KEY em Settings > Secrets do Streamlit.")
        st.stop()

supabase = _get_supabase_client()

APP_TO_DB = {
    "glicemia": {"ID": "id", "Usuario": "usuario", "Data": "data", "Hora": "hora", "Valor": "valor", "Momento": "momento", "Dose": "dose", "Dose_Rapida": "dose_rapida", "Dose_Longa": "dose_longa"},
    "nutricao": {"ID": "id", "Usuario": "usuario", "Data": "data", "Momento": "momento", "Info": "info", "C": "c", "P": "p", "G": "g"},
    "receita": {
        "Usuario": "usuario",
        "manha_f1_min": "manha_f1_min", "manha_f1_max": "manha_f1_max", "manha_f1_dose": "manha_f1_dose",
        "manha_f2_min": "manha_f2_min", "manha_f2_max": "manha_f2_max", "manha_f2_dose": "manha_f2_dose",
        "manha_f3_min": "manha_f3_min", "manha_f3_max": "manha_f3_max", "manha_f3_dose": "manha_f3_dose",
        "noite_f1_min": "noite_f1_min", "noite_f1_max": "noite_f1_max", "noite_f1_dose": "noite_f1_dose",
        "noite_f2_min": "noite_f2_min", "noite_f2_max": "noite_f2_max", "noite_f2_dose": "noite_f2_dose",
        "noite_f3_min": "noite_f3_min", "noite_f3_max": "noite_f3_max", "noite_f3_dose": "noite_f3_dose",
        "glargina_cafe_ui": "glargina_cafe_ui", "glargina_janta_ui": "glargina_janta_ui"
    },
    "sugestoes": {"ID": "id", "Usuario": "usuario", "Data": "data", "Sugestão": "Sugestão"},
}
DB_TO_APP = {t: {v: k for k, v in m.items()} for t, m in APP_TO_DB.items()}

def sb_select(table: str, filters=None, order=None):
    q = supabase.table(table).select("*")
    for col, val in (filters or {}).items():
        q = q.eq(col, val)
    if order:
        q = q.order(order)
    resp = q.execute()
    return resp.data or []

def sb_insert(table: str, payload):
    return supabase.table(table).insert(payload).execute()

def sb_upsert(table: str, payload):
    return supabase.table(table).upsert(payload).execute()

def sb_update(table: str, filters: dict, payload: dict):
    q = supabase.table(table).update(payload)
    for col, val in filters.items():
        q = q.eq(col, val)
    return q.execute()

def sb_delete(table: str, filters: dict):
    q = supabase.table(table).delete()
    for col, val in filters.items():
        if isinstance(val, (list, tuple, set)):
            q = q.in_(col, list(val))
        else:
            q = q.eq(col, val)
    return q.execute()

def _df_to_db_records(table: str, df: pd.DataFrame):
    if df is None or df.empty:
        return []
    mapping = APP_TO_DB[table]
    out = []
    for rec in df.to_dict(orient="records"):
        row = {}
        for app_col, db_col in mapping.items():
            if app_col in rec:
                val = rec[app_col]
                if pd.isna(val):
                    val = None
                row[db_col] = val
        out.append(row)
    return out

def _records_to_df(table: str, records):
    cols = list(APP_TO_DB[table].keys())
    if not records:
        return pd.DataFrame(columns=cols)
    mapping = DB_TO_APP[table]
    out = []
    for rec in records:
        row = {}
        for db_col, val in rec.items():
            row[mapping.get(db_col, db_col)] = val
        out.append(row)
    df = pd.DataFrame(out)
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    return df

# ================= NORMALIZAÇÃO =================
def norm_email(x: str) -> str:
    return (x or "").strip().lower()

def norm_senha(x: str) -> str:
    return (x or "").strip()

# ================= TEMPO =================
def agora_br() -> datetime:
    return datetime.now(fuso_br)

# ================= BACKUP / RESTORE =================
BACKUP_STATE_FILE = BACKUP_DIR / "last_auto_backup.txt"

ARQUIVOS_BACKUP_MAPA = {
    "usuarios.db": DB_USERS,
    "saude_kids.db": APP_DB,
    "dados_glicemia_BETA.csv": ARQ_G,
    "dados_nutricao_BETA.csv": ARQ_N,
    "config_receita_BETA.csv": ARQ_R,
    "mensagens_admin_BETA.csv": ARQ_M,
}

def criar_backup_zip_em_bytes():
    ts = agora_br().strftime("%Y-%m-%d_%H-%M-%S")
    nome = f"backup_saude_kids_{ts}.zip"
    out = BytesIO()

    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for nome_zip, caminho_real in ARQUIVOS_BACKUP_MAPA.items():
            caminho_real = Path(caminho_real)
            if caminho_real.exists():
                z.write(caminho_real, arcname=nome_zip)

        manifest = {
            "criado_em": agora_br().strftime("%d/%m/%Y %H:%M:%S"),
            "tipo": "backup_completo_global",
            "inclui_todos_usuarios": True,
            "arquivos": list(ARQUIVOS_BACKUP_MAPA.keys()),
            "app": "Saúde Kids BETA",
        }
        z.writestr("manifest.json", pd.Series(manifest).to_json(force_ascii=False, indent=2))

    out.seek(0)
    return out.getvalue(), nome

def criar_backup_zip_em_disco():
    zip_bytes, nome = criar_backup_zip_em_bytes()
    caminho = BACKUP_DIR / nome
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    with open(caminho, "wb") as f:
        f.write(zip_bytes)
    return caminho

def restaurar_backup_zip_bytes(zip_bytes: bytes):
    tmp_dir = BACKUP_DIR / "_tmp_restore"
    if tmp_dir.exists():
        shutil.rmtree(tmp_dir)
    tmp_dir.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(BytesIO(zip_bytes), "r") as z:
        z.extractall(tmp_dir)

    restaurados = []

    for nome_zip, destino in ARQUIVOS_BACKUP_MAPA.items():
        src = None

        candidato = tmp_dir / nome_zip
        if candidato.exists():
            src = candidato
        else:
            encontrados = list(tmp_dir.rglob(nome_zip))
            if encontrados:
                src = encontrados[0]

        if src and src.exists():
            Path(destino).parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, destino)
            restaurados.append(nome_zip)

    manifest = tmp_dir / "manifest.json"
    if manifest.exists():
        restaurados.append("manifest.json")

    shutil.rmtree(tmp_dir)
    return restaurados

def backup_automatico_diario_3h():
    """
    Streamlit não roda 24/7.
    Regra: após 03:00, se ainda não fez backup HOJE => faz 1.
    """
    agora = agora_br()
    hoje = agora.strftime("%Y-%m-%d")
    if agora.hour >= 3:
        ultima = ""
        if BACKUP_STATE_FILE.exists():
            ultima = BACKUP_STATE_FILE.read_text(encoding="utf-8").strip()
        if ultima != hoje:
            criar_backup_zip_em_disco()
            BACKUP_STATE_FILE.write_text(hoje, encoding="utf-8")

def listar_backups() -> pd.DataFrame:
    backups = []
    for p in sorted(BACKUP_DIR.glob("backup_saude_kids_*.zip")):
        stat = p.stat()
        dt = datetime.fromtimestamp(stat.st_mtime, tz=fuso_br)
        backups.append({
            "arquivo": p.name,
            "caminho": str(p),
            "data_hora": dt,
            "tamanho_mb": round(stat.st_size / (1024 * 1024), 2),
        })
    if not backups:
        return pd.DataFrame(columns=["arquivo", "caminho", "data_hora", "tamanho_mb"])
    df = pd.DataFrame(backups).sort_values("data_hora", ascending=False).reset_index(drop=True)
    return df

def apagar_backups_antigos(dias_manter=7) -> int:
    limite = agora_br() - timedelta(days=dias_manter)
    apagados = 0
    for p in BACKUP_DIR.glob("backup_saude_kids_*.zip"):
        dt = datetime.fromtimestamp(p.stat().st_mtime, tz=fuso_br)
        if dt < limite:
            try:
                p.unlink()
                apagados += 1
            except Exception:
                pass
    return apagados

backup_automatico_diario_3h()

# ================= DESIGN =================
st.markdown(r"""
<style>
    .stApp { background-color: #0e1117; color: #ffffff; }
    .card { background-color: #1a1c24; padding: 22px; border-radius: 18px; border: 1px solid #30363d; margin-bottom: 18px; }
    .metric-box { background: #262730; border: 1px solid #4a4a4a; padding: 14px; border-radius: 12px; text-align: center; }
    .dose-destaque { font-size: 34px; font-weight: 700; color: #4ade80; }
    label, p, span, h1, h2, h3, .stMarkdown { color: white !important; }
    .stTextInput>div>div>input, .stNumberInput>div>div>input {
        background-color: #262730 !important; color: white !important; border: 1px solid #4a4a4a !important;
    }
    .stTabs [data-baseweb="tab-list"] { background-color: #0e1117; }
    .stTabs [data-baseweb="tab"] { color: white; }
</style>
""", unsafe_allow_html=True)

# ================= EMAIL (RESET) =================
def gerar_senha_temporaria(tamanho=6):
    caracteres = string.ascii_letters + string.digits
    return "".join(random.choice(caracteres) for _ in range(tamanho))

def enviar_senha_nova(email_destino, senha_nova):
    """
    Recomendado (seguro):
      - No Streamlit Cloud: st.secrets["GMAIL_APP_PASSWORD"] = "SUA_SENHA_DE_APP"
    """
    import smtplib

    meu_email = "ewerlon.osbadboys@gmail.com"
    minha_senha = (st.secrets.get("GMAIL_APP_PASSWORD", "") or "").strip()
    if not minha_senha:
        return False

    corpo = f"<h3>Saúde Kids</h3><p>Sua nova senha de acesso é: <b>{senha_nova}</b></p>"
    msg = MIMEText(corpo, "html")
    msg["Subject"] = "Sua Nova Senha - Saúde Kids"
    msg["From"] = meu_email
    msg["To"] = email_destino

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(meu_email, minha_senha)
            smtp.send_message(msg)
        return True
    except Exception:
        return False

# ================= DB USERS =================
# ================= DB USERS + PLANOS =================
TESTE_DIAS = 20
MENSALIDADE_DIAS = 30

def init_db():
    if not sb_select("users", {"email": "admin"}):
        sb_insert("users", {"nome": "Administrador", "email": "admin", "senha": "542820"})
    if not sb_select("plans", {"email": "admin"}):
        now = agora_br().isoformat()
        sb_insert("plans", {"email": "admin", "created_at": now, "trial_end": now, "paid_until": "2099-12-31T23:59:59"})

def _db_conn():
    return None

def garantir_plano(email: str):
    email = norm_email(email)
    if not email or email == "admin":
        return
    if not sb_select("plans", {"email": email}):
        now = agora_br()
        trial_end = now + timedelta(days=TESTE_DIAS)
        sb_insert("plans", {"email": email, "created_at": now.isoformat(), "trial_end": trial_end.isoformat(), "paid_until": ""})

def get_plano_status(email: str):
    email = norm_email(email)
    if email == "admin":
        return {"allowed": True, "motivo": "admin", "trial_end": None, "paid_until": None, "dias_restantes": None}
    garantir_plano(email)
    rows = sb_select("plans", {"email": email})
    r = rows[0] if rows else {}
    now = agora_br()
    trial_end = None
    paid_until = None
    try:
        trial_end = datetime.fromisoformat(r.get("trial_end")) if r.get("trial_end") else None
    except Exception:
        pass
    try:
        paid_until = datetime.fromisoformat(r.get("paid_until")) if r.get("paid_until") else None
    except Exception:
        pass
    if paid_until and now <= paid_until:
        dias = (paid_until.date() - now.date()).days
        return {"allowed": True, "motivo": "pago", "trial_end": trial_end, "paid_until": paid_until, "dias_restantes": dias}
    if trial_end and now <= trial_end:
        dias = (trial_end.date() - now.date()).days
        return {"allowed": True, "motivo": "teste", "trial_end": trial_end, "paid_until": paid_until, "dias_restantes": dias}
    return {"allowed": False, "motivo": "expirado", "trial_end": trial_end, "paid_until": paid_until, "dias_restantes": 0}

def registrar_mensagem_mensalidade(email: str, mensagem: str):
    email = norm_email(email)
    if not email:
        return
    sb_insert("billing_msgs", {"email": email, "data_hora": agora_br().strftime("%d/%m/%Y %H:%M"), "mensagem": (mensagem or "").strip(), "status": "novo"})

def listar_mensagens_mensalidade():
    try:
        rows = sb_select("billing_msgs", order="id")
        df = pd.DataFrame(rows)
        if df.empty:
            return pd.DataFrame(columns=["id", "email", "data_hora", "mensagem", "status"])
        return df.sort_values("id", ascending=False).reset_index(drop=True)
    except Exception:
        return pd.DataFrame(columns=["id", "email", "data_hora", "mensagem", "status"])

def marcar_mensagem_status(msg_id: int, status: str):
    sb_update("billing_msgs", {"id": int(msg_id)}, {"status": status})

def excluir_mensagem(msg_id: int):
    sb_delete("billing_msgs", {"id": int(msg_id)})

def ativar_mensalidade(email: str, dias: int = MENSALIDADE_DIAS):
    email = norm_email(email)
    if not email or email == "admin":
        return
    garantir_plano(email)
    now = agora_br()
    paid_until = now + timedelta(days=int(dias))
    sb_update("plans", {"email": email}, {"paid_until": paid_until.isoformat()})

init_db()

if "logado" not in st.session_state:
    st.session_state.logado = False
if "user_email" not in st.session_state:
    st.session_state.user_email = ""
if "pending_email" not in st.session_state:
    st.session_state.pending_email = ""

# ====== COOKIE (OPCIONAL) ======
def cookie_get_email():
    if not HAS_COOKIES or _cookie_manager is None:
        return ""
    try:
        v = _cookie_manager.get(COOKIE_KEY)
        return norm_email(v)
    except Exception:
        return ""

def cookie_set_email(email: str):
    if not HAS_COOKIES or _cookie_manager is None:
        return
    try:
        _cookie_manager.set(COOKIE_KEY, norm_email(email), expires_at=timedelta(days=COOKIE_DIAS))
    except Exception:
        pass

def cookie_clear():
    if not HAS_COOKIES or _cookie_manager is None:
        return
    try:
        _cookie_manager.delete(COOKIE_KEY)
    except Exception:
        pass

if not st.session_state.logado:
    ck = cookie_get_email()
    if ck:
        stt = get_plano_status(ck)
        if stt.get('allowed'):
            st.session_state.logado = True
            st.session_state.user_email = ck
        else:
            st.session_state.logado = False
            st.session_state.user_email = ''
            st.session_state.pending_email = ck

# ================= ARQUIVOS (DADOS) =================
TABLE_BY_FILE = {
    str(ARQ_G): "glicemia",
    str(ARQ_N): "nutricao",
    str(ARQ_R): "receita",
    str(ARQ_M): "sugestoes",
    str(ARQ_M_ROOT): "sugestoes",
}


def _app_conn():
    return None

def _table_for_file(arq) -> str:
    return TABLE_BY_FILE.get(str(arq), "")

def _read_table_df(table: str, where: str = "", params=()):
    filters = {}
    if where == "Usuario=?" and params:
        db_col = APP_TO_DB.get(table, {}).get("Usuario", "usuario")
        filters[db_col] = params[0]
    rows = sb_select(table, filters=filters)
    return _records_to_df(table, rows)

def _append_df_table(table: str, df: pd.DataFrame):
    recs = _df_to_db_records(table, df)
    if recs:
        sb_insert(table, recs)

def _replace_user_rows(table: str, user_email: str, df_new: pd.DataFrame):
    db_col = APP_TO_DB.get(table, {}).get("Usuario", "usuario")
    sb_delete(table, {db_col: user_email})
    recs = _df_to_db_records(table, df_new)
    if recs:
        sb_insert(table, recs)

def _delete_user_app_data(user_email: str):
    for table in ["glicemia", "nutricao", "receita", "sugestoes"]:
        db_col = APP_TO_DB.get(table, {}).get("Usuario", "usuario")
        sb_delete(table, {db_col: user_email})

def init_app_db():
    return

def migrar_csvs_para_sqlite():
    init_app_db()

    if ARQ_G.exists():
        qtd = len(sb_select("glicemia"))
        if qtd == 0:
            try:
                df = pd.read_csv(ARQ_G)
                if not df.empty:
                    if "Usuario" not in df.columns:
                        df["Usuario"] = ""
                    if "ID" not in df.columns:
                        df["ID"] = [f"GL-{uuid.uuid4().hex[:12]}" for _ in range(len(df))]
                    if "Dose" not in df.columns:
                        df["Dose"] = ""
                    if "Dose_Rapida" not in df.columns:
                        def _parse_r(x):
                            s = str(x or "")
                            m = re.search(r"(?:Rápida|Rapida)\s*:?\s*([^|]+)", s, flags=re.IGNORECASE)
                            if m:
                                return str(m.group(1)).strip()
                            if s.strip() and ("longa" not in s.lower()):
                                return s.strip()
                            return ""
                        df["Dose_Rapida"] = df["Dose"].apply(_parse_r)
                    if "Dose_Longa" not in df.columns:
                        def _parse_l(x):
                            s = str(x or "")
                            m = re.search(r"Longa\s*:?\s*([^|]+)", s, flags=re.IGNORECASE)
                            return str(m.group(1)).strip() if m else ""
                        df["Dose_Longa"] = df["Dose"].apply(_parse_l)
                    cols = ["ID", "Usuario", "Data", "Hora", "Valor", "Momento", "Dose", "Dose_Rapida", "Dose_Longa"]
                    for c in cols:
                        if c not in df.columns:
                            df[c] = ""
                    _append_df_table("glicemia", df[cols])
            except Exception:
                pass

    if ARQ_N.exists():
        qtd = len(sb_select("nutricao"))
        if qtd == 0:
            try:
                df = pd.read_csv(ARQ_N)
                if not df.empty:
                    if "Usuario" not in df.columns:
                        df["Usuario"] = ""
                    if "ID" not in df.columns:
                        df["ID"] = [f"NT-{uuid.uuid4().hex[:12]}" for _ in range(len(df))]
                    cols = ["ID", "Usuario", "Data", "Momento", "Info", "C", "P", "G"]
                    for c in cols:
                        if c not in df.columns:
                            df[c] = ""
                    _append_df_table("nutricao", df[cols])
            except Exception:
                pass

    if ARQ_R.exists():
        qtd = len(sb_select("receita"))
        if qtd == 0:
            try:
                df = pd.read_csv(ARQ_R)
                if not df.empty:
                    if "Usuario" not in df.columns:
                        df["Usuario"] = ""
                    df = df.drop_duplicates(subset=["Usuario"], keep="last")
                    keep = [
                        "Usuario",
                        "manha_f1_min", "manha_f1_max", "manha_f1_dose",
                        "manha_f2_min", "manha_f2_max", "manha_f2_dose",
                        "manha_f3_min", "manha_f3_max", "manha_f3_dose",
                        "noite_f1_min", "noite_f1_max", "noite_f1_dose",
                        "noite_f2_min", "noite_f2_max", "noite_f2_dose",
                        "noite_f3_min", "noite_f3_max", "noite_f3_dose",
                        "glargina_cafe_ui", "glargina_janta_ui"
                    ]
                    for c in keep:
                        if c not in df.columns:
                            df[c] = ""
                    _append_df_table("receita", df[keep])
            except Exception:
                pass

    for sugest_path in [ARQ_M, ARQ_M_ROOT]:
        if sugest_path.exists():
            qtd = len(sb_select("sugestoes"))
            if qtd == 0:
                try:
                    df = pd.read_csv(sugest_path)
                    if not df.empty:
                        if "Usuario" not in df.columns:
                            df["Usuario"] = ""
                        if "Data" not in df.columns:
                            df["Data"] = ""
                        if "Sugestão" not in df.columns:
                            alt = [c for c in df.columns if c.lower().startswith("sugest")]
                            df["Sugestão"] = df[alt[0]] if alt else ""
                        if "ID" not in df.columns:
                            df["ID"] = [f"SG-{uuid.uuid4().hex[:12]}" for _ in range(len(df))]
                        _append_df_table("sugestoes", df[["ID", "Usuario", "Data", "Sugestão"]])
                except Exception:
                    pass
            break


migrar_csvs_para_sqlite()

def carregar_dados_seguro(arq: str) -> pd.DataFrame:
    table = _table_for_file(arq)
    if table:
        return _read_table_df(table, "Usuario=?", (st.session_state.user_email,))
    if not os.path.exists(arq):
        return pd.DataFrame()
    df = pd.read_csv(arq)
    if "Usuario" not in df.columns:
        df["Usuario"] = ""
    return df[df["Usuario"] == st.session_state.user_email].copy()

# ================= IDs + CRUD GLICEMIA =================
def _ensure_id_column(df: pd.DataFrame, col_name="ID", prefix="GL") -> pd.DataFrame:
    if df is None or df.empty:
        return df
    if col_name not in df.columns:
        df[col_name] = [f"{prefix}-{uuid.uuid4().hex[:12]}" for _ in range(len(df))]
    else:
        mask = df[col_name].isna() | (df[col_name].astype(str).str.strip() == "")
        if mask.any():
            df.loc[mask, col_name] = [f"{prefix}-{uuid.uuid4().hex[:12]}" for _ in range(int(mask.sum()))]
    return df

def carregar_glicemia_com_id() -> pd.DataFrame:
    df_all = _read_table_df("glicemia")
    if df_all.empty:
        return pd.DataFrame(columns=["ID", "Usuario", "Data", "Hora", "Valor", "Momento", "Dose", "Dose_Rapida", "Dose_Longa"])

    if "Usuario" not in df_all.columns:
        df_all["Usuario"] = ""

    df_all = _ensure_id_column(df_all, "ID", "GL")
    if "Dose" not in df_all.columns:
        df_all["Dose"] = ""

    if "Dose_Rapida" not in df_all.columns:
        def _parse_r(x: str) -> str:
            s = str(x or "")
            m = re.search(r"(?:Rápida|Rapida)\s*:?\s*([^|]+)", s, flags=re.IGNORECASE)
            if m:
                return str(m.group(1)).strip()
            if s.strip() and ("longa" not in s.lower()):
                return s.strip()
            return ""
        df_all["Dose_Rapida"] = df_all["Dose"].apply(_parse_r)

    if "Dose_Longa" not in df_all.columns:
        def _parse_l(x: str) -> str:
            s = str(x or "")
            m = re.search(r"Longa\s*:?\s*([^|]+)", s, flags=re.IGNORECASE)
            return str(m.group(1)).strip() if m else ""
        df_all["Dose_Longa"] = df_all["Dose"].apply(_parse_l)

    def _mk_dose_display(r) -> str:
        dr = str(r.get("Dose_Rapida", "") or "").strip()
        dl = str(r.get("Dose_Longa", "") or "").strip()
        out = ""
        if dr:
            out = dr
        if dl:
            out = (out + " | " if out else "") + f"Longa: {dl}"
        return out

    df_all["Dose"] = df_all.apply(_mk_dose_display, axis=1)
    return df_all[df_all["Usuario"] == st.session_state.user_email].copy()

def salvar_registro_glicemia(valor: int, momento: str, dose: str, dt: datetime, dose_rapida: str = "", dose_longa: str = ""):
    dr = (dose_rapida or "").strip()
    dl = (dose_longa or "").strip()

    if not dr and dose:
        m = re.search(r"(?:Rápida|Rapida)\s*:?\s*([^|]+)", str(dose), flags=re.IGNORECASE)
        if m:
            dr = str(m.group(1)).strip()
        elif str(dose).strip() and ("longa" not in str(dose).lower()):
            dr = str(dose).strip()
    if not dl and dose:
        m = re.search(r"Longa\s*:?\s*([^|]+)", str(dose), flags=re.IGNORECASE)
        if m:
            dl = str(m.group(1)).strip()

    dose_display = ""
    if dr:
        dose_display = dr
    if dl:
        dose_display = (dose_display + " | " if dose_display else "") + f"Longa: {dl}"

    novo = pd.DataFrame([{
        "ID": f"GL-{uuid.uuid4().hex[:12]}",
        "Usuario": st.session_state.user_email,
        "Data": dt.strftime("%d/%m/%Y"),
        "Hora": dt.strftime("%H:%M"),
        "Valor": int(valor),
        "Momento": (momento or "").strip(),
        "Dose": dose_display,
        "Dose_Rapida": dr,
        "Dose_Longa": dl,
    }])
    _append_df_table("glicemia", novo)


def aplicar_edicoes_e_exclusoes_glicemia(df_editado: pd.DataFrame):
    if df_editado is None or df_editado.empty:
        return

    df_editado = df_editado.copy()
    if "Excluir" not in df_editado.columns:
        df_editado["Excluir"] = False

    df_editado["ID"] = df_editado["ID"].astype(str)
    df_editado["Data"] = df_editado["Data"].astype(str).str.strip()
    df_editado["Hora"] = df_editado["Hora"].astype(str).str.strip()
    df_editado["Momento"] = df_editado["Momento"].astype(str).str.strip()
    df_editado["Dose_Rapida"] = df_editado.get("Dose_Rapida", "").astype(str).str.strip()
    df_editado["Dose_Longa"] = df_editado.get("Dose_Longa", "").astype(str).str.strip()

    def _mk_dose_display(r) -> str:
        dr = str(r.get("Dose_Rapida", "") or "").strip()
        dl = str(r.get("Dose_Longa", "") or "").strip()
        out = ""
        if dr:
            out = dr
        if dl:
            out = (out + " | " if out else "") + f"Longa: {dl}"
        return out

    df_editado["Dose"] = df_editado.apply(_mk_dose_display, axis=1)
    df_editado["Valor"] = pd.to_numeric(df_editado["Valor"], errors="coerce").fillna(0).astype(int)

    ids_excluir = df_editado.loc[df_editado["Excluir"] == True, "ID"].astype(str).tolist()
    if ids_excluir:
        for rid in ids_excluir:
            sb_delete("glicemia", {"usuario": st.session_state.user_email, "id": rid})

    df_upd = df_editado.loc[df_editado["Excluir"] != True].copy()
    for _, r in df_upd.iterrows():
        sb_update(
            "glicemia",
            {"usuario": st.session_state.user_email, "id": str(r["ID"])},
            {
                "data": r["Data"],
                "hora": r["Hora"],
                "valor": int(r["Valor"]),
                "momento": r["Momento"],
                "dose_rapida": r.get("Dose_Rapida", ""),
                "dose_longa": r.get("Dose_Longa", ""),
                "dose": r.get("Dose", ""),
            }
        )

# ================= RECEITA (RÁPIDA/LONGA) =================
def _schema_receita_nova(rec: pd.Series, periodo: str) -> bool:
    need = [
        f"{periodo}_f1_min", f"{periodo}_f1_max", f"{periodo}_f1_dose",
        f"{periodo}_f2_min", f"{periodo}_f2_max", f"{periodo}_f2_dose",
        f"{periodo}_f3_min", f"{periodo}_f3_max", f"{periodo}_f3_dose",
    ]
    return all(k in rec.index for k in need)

def calc_insulina(v: int, momento: str):
    df_r = carregar_dados_seguro(ARQ_R)
    if df_r.empty:
        return "0 UI", "Configurar Receita"

    rec = df_r.iloc[0]
    periodo = "manha" if momento in ["Antes Café", "Após Café", "Antes Almoço", "Após Almoço", ] else "noite"

    try:
        if not _schema_receita_nova(rec, periodo):
            return "0 UI", "Receita inválida"

        f1_min = float(rec[f"{periodo}_f1_min"]); f1_max = float(rec[f"{periodo}_f1_max"]); f1_dose = float(rec[f"{periodo}_f1_dose"])
        f2_min = float(rec[f"{periodo}_f2_min"]); f2_max = float(rec[f"{periodo}_f2_max"]); f2_dose = float(rec[f"{periodo}_f2_dose"])
        f3_min = float(rec[f"{periodo}_f3_min"]); f3_max = float(rec[f"{periodo}_f3_max"]); f3_dose = float(rec[f"{periodo}_f3_dose"])

        if v < 70:
            return "0 UI", "Hipoglicemia!"

        if f1_min <= v <= f1_max:
            return f"{int(f1_dose)} UI", f"Faixa 1 ({int(f1_min)}-{int(f1_max)})"
        if f2_min <= v <= f2_max:
            return f"{int(f2_dose)} UI", f"Faixa 2 ({int(f2_min)}-{int(f2_max)})"
        if f3_min <= v <= f3_max:
            return f"{int(f3_dose)} UI", f"Faixa 3 ({int(f3_min)}-{int(f3_max)})"
        return "0 UI", "Fora das faixas"
    except Exception:
        return "0 UI", "Erro na Receita"

def calc_insulina_rapida(v: int, momento: str):
    return calc_insulina(v, momento)

def calc_glargina(momento: str):
    df_r = carregar_dados_seguro(ARQ_R)
    if df_r.empty:
        return "0 UI", "Longa: Configurar"

    rec = df_r.iloc[0]
    try:
        cafe = float(rec.get("glargina_cafe_ui", 0) or 0)
        janta = float(rec.get("glargina_janta_ui", 0) or 0)
        if momento == "Antes Café":
            return f"{int(cafe)} UI", "Longa (Antes Café)"
        if momento == "Antes Janta":
            return f"{int(janta)} UI", "Longa (Antes Janta)"
        return "—", "Longa: não aplicável"
    except Exception:
        return "0 UI", "Longa: erro"

# ================= PRÓXIMA MEDIDA (+2h) e WHATSAPP =================
def proxima_medida_apos(momento: str, dt_base: datetime):
    mapa = {
        "Antes Café": "Após Café",
        "Antes Almoço": "Após Almoço",
        "Antes Janta": "Após Janta",
    }
    if momento not in mapa:
        return "", ""
    dt_apos = dt_base + timedelta(hours=2)
    return mapa[momento], dt_apos.strftime("%H:%M")

def link_whatsapp_lembrete(momento: str, valor_glicemia: int, dose_rapida: str, dose_longa: str) -> str:
    dt_agora = agora_br()
    momento_apos, hora_apos = proxima_medida_apos(momento, dt_agora)

    linhas = [
        "🧪 Saúde Kids",
        "",
        f"✅ Medida AGORA: {momento}",
        f"📍 Glicemia: {int(valor_glicemia)}",
    ]

    if dose_rapida and dose_rapida != "—":
        linhas.append(f"⚡ Rápida: {dose_rapida}")
    if dose_longa and dose_longa != "—":
        linhas.append(f"🩸 Longa: {dose_longa}")

    if momento_apos and hora_apos:
        linhas.extend(["", f"⏰ Próxima medida: {momento_apos} às {hora_apos} (2h após)"])

    mensagem = "\n".join(linhas)
    return "https://wa.me/?text=" + quote(mensagem)

def link_whatsapp_nutricao(momento: str, alimentos: list, c_tot: int, p_tot: int, g_tot: int) -> str:
    dt = agora_br().strftime("%d/%m/%Y %H:%M")
    itens = "\n".join([f"• {a}" for a in alimentos]) if alimentos else "• Nenhum alimento"
    msg = (
        f"🍽️ Refeição - {momento}\n"
        f"👧 Coly comeu:\n"
        f"{itens}\n"
        f"\n📊 Carboidratos: {int(c_tot)}g"
        f"\n🥩 Proteínas: {int(p_tot)}g"
        f"\n🥑 Gorduras: {int(g_tot)}g"
        f"\n📅 {dt}"
    )
    return "https://wa.me/?text=" + quote(msg)





# ================= MOMENTOS / ALIMENTOS =================
MOMENTOS_ORDEM = [
    "Antes Café",
    "Após Café",
    "Antes Almoço",
    "Após Almoço",
    "Antes Janta",
    "Após Janta",
    "Madrugada",
]

# =========================
# CATÁLOGO DE ALIMENTOS (400+)
# Formato: "Nome": [carboidrato, proteína, gordura]  (valores aproximados por porção)
# Observação: para aumentar a cobertura sem inflar manualmente, geramos variações de porção:
#   (P) ~ 1/2 porção, (M) porção padrão, (G) ~ 1,5 porção.
# =========================

def _round_macro(x: float) -> int:
    return int(round(float(x)))

def _gerar_catalogo_alimentos_400plus():
    base = [
        # CAFÉ DA MANHÃ / LANCHES
        ("Pão francês (1 un)", 28, 4, 1),
        ("Pão integral (2 fatias)", 24, 5, 2),
        ("Pão de forma (2 fatias)", 23, 4, 2),
        ("Torrada (4 un)", 18, 3, 1),
        ("Biscoito cream cracker (5 un)", 19, 3, 2),
        ("Biscoito maizena (6 un)", 22, 2, 3),
        ("Bolo simples (1 fatia)", 30, 4, 10),
        ("Bolo de chocolate (1 fatia)", 35, 4, 12),
        ("Granola (30g)", 20, 4, 5),
        ("Aveia (30g)", 19, 5, 3),
        ("Cereal matinal (30g)", 24, 3, 1),
        ("Manteiga (1 colher chá)", 0, 0, 4),
        ("Requeijão (1 colher sopa)", 1, 1, 4),
        ("Geleia (1 colher sopa)", 13, 0, 0),
        ("Geleia zero açúcar (1 colher sopa)", 4, 0, 0),
        ("Mel (1 colher sopa)", 17, 0, 0),
        ("Ovo cozido (1 un)", 1, 6, 5),
        ("Ovo mexido (1 un)", 1, 7, 6),
        ("Omelete (2 ovos)", 2, 12, 12),
        ("Queijo mussarela (1 fatia)", 1, 7, 6),
        ("Queijo minas (1 fatia)", 1, 6, 5),
        ("Presunto (1 fatia)", 1, 5, 3),
        ("Peito de peru (1 fatia)", 1, 6, 2),

        # LEITES / DERIVADOS
        ("Leite integral (200ml)", 10, 6, 6),
        ("Leite desnatado (200ml)", 10, 7, 0),
        ("Leite sem lactose (200ml)", 10, 6, 6),
        ("Leite de soja (200ml)", 4, 7, 4),
        ("Leite de amêndoas (200ml)", 2, 1, 2),
        ("Leite de coco (200ml)", 3, 1, 5),
        ("Iogurte natural (170g)", 9, 5, 3),
        ("Iogurte grego (170g)", 8, 8, 5),
        ("Iogurte grego zero (170g)", 4, 10, 0),
        ("Danone tradicional (170g)", 15, 3, 3),
        ("Danone zero açúcar (170g)", 6, 5, 2),

        # CAFÉS / ADOÇANTES / BEBIDAS
        ("Café puro", 0, 0, 0),
        ("Café com açúcar (1 xícara)", 10, 0, 0),
        ("Café com adoçante (1 xícara)", 0, 0, 0),
        ("Café com leite (200ml)", 10, 5, 4),
        ("Chá sem açúcar", 0, 0, 0),
        ("Chá com açúcar", 10, 0, 0),
        ("Adoçante (stevia/sucralose)", 0, 0, 0),
        ("Xilitol (1 colher chá)", 4, 0, 0),
        ("Água", 0, 0, 0),
        ("Água de coco (200ml)", 9, 0, 0),
        ("Suco de laranja (200ml)", 22, 1, 0),
        ("Suco de uva (200ml)", 25, 0, 0),
        ("Suco natural (200ml)", 20, 0, 0),
        ("Refrigerante (350ml)", 35, 0, 0),
        ("Refrigerante zero (350ml)", 0, 0, 0),
        ("Coca-Cola Zero (350ml)", 0, 0, 0),
        ("Guaraná Zero (350ml)", 0, 0, 0),
        ("H2OH (350ml)", 5, 0, 0),
        ("H2OH Zero (350ml)", 0, 0, 0),

        # FRUTAS
        ("Banana (1 un)", 23, 1, 0),
        ("Banana prata (1 un)", 21, 1, 0),
        ("Maçã (1 un)", 19, 0, 0),
        ("Pera (1 un)", 20, 0, 0),
        ("Manga (1 fatia)", 28, 1, 0),
        ("Melancia (1 fatia)", 8, 0, 0),
        ("Melão (1 fatia)", 9, 0, 0),
        ("Abacaxi (1 fatia)", 13, 0, 0),
        ("Uva (1 cacho pequeno)", 18, 0, 0),
        ("Laranja (1 un)", 15, 1, 0),
        ("Tangerina (1 un)", 12, 1, 0),
        ("Morango (10 un)", 8, 1, 0),
        ("Kiwi (1 un)", 15, 1, 0),
        ("Mamão (1 fatia)", 12, 1, 0),
        ("Goiaba (1 un)", 14, 1, 0),
        ("Pêssego (1 un)", 15, 1, 0),
        ("Ameixa (2 un)", 12, 0, 0),
        ("Abacate (1/2 un)", 9, 2, 15),

        # VERDURAS / LEGUMES
        ("Alface (1 prato)", 2, 1, 0),
        ("Tomate (1 un)", 4, 1, 0),
        ("Cenoura (1/2 un)", 10, 1, 0),
        ("Beterraba (1/2 un)", 13, 2, 0),
        ("Pepino (1 un)", 3, 1, 0),
        ("Brócolis (1 concha)", 7, 3, 0),
        ("Couve (1 concha)", 6, 3, 0),
        ("Espinafre (1 concha)", 4, 3, 0),
        ("Abobrinha (1 concha)", 4, 1, 0),
        ("Berinjela (1 concha)", 6, 1, 0),
        ("Chuchu (1 concha)", 4, 1, 0),
        ("Vagem (1 concha)", 8, 2, 0),
        ("Couve-flor (1 concha)", 5, 2, 0),

        # PRATOS
        ("Arroz branco (3 colheres)", 28, 3, 0),
        ("Arroz integral (3 colheres)", 25, 3, 1),
        ("Feijão (1 concha)", 14, 5, 1),
        ("Macarrão (1 prato raso)", 30, 5, 1),
        ("Macarrão integral (1 prato raso)", 28, 6, 1),
        ("Purê de batata (3 colheres)", 20, 2, 4),
        ("Batata cozida (1 un média)", 20, 2, 0),
        ("Batata frita (1 porção pequena)", 30, 3, 15),
        ("Mandioca cozida (1 porção)", 28, 1, 0),
        ("Pão de queijo (2 un)", 18, 3, 6),
        ("Cuscuz (1 porção)", 25, 3, 1),
        ("Tapioca (1 un média)", 28, 1, 1),
        ("Farofa (2 colheres)", 18, 1, 6),

        # CARNES / PROTEÍNAS
        ("Frango grelhado (1 filé)", 0, 30, 3),
        ("Frango cozido (1 porção)", 0, 28, 3),
        ("Frango empanado (1 filé)", 15, 25, 8),
        ("Carne bovina (1 porção)", 0, 26, 10),
        ("Bife grelhado (1 bife)", 0, 26, 9),
        ("Carne moída (1 porção)", 0, 24, 12),
        ("Porco (1 porção)", 0, 24, 15),
        ("Linguiça (1 un)", 2, 15, 25),
        ("Salsicha (1 un)", 3, 12, 20),
        ("Hambúrguer (1 un)", 4, 15, 18),
        ("Fígado (1 porção)", 4, 24, 5),

        # PEIXES
        ("Peixe grelhado (1 filé)", 0, 24, 5),
        ("Peixe assado (1 filé)", 0, 23, 6),
        ("Tilápia (1 filé)", 0, 26, 3),
        ("Salmão (1 filé)", 0, 22, 13),
        ("Atum (1 porção)", 0, 25, 2),
        ("Sardinha (1 porção)", 0, 23, 10),

        # LEGUMINOSAS
        ("Lentilha (1 concha)", 20, 9, 1),
        ("Grão-de-bico (1 concha)", 22, 8, 3),
        ("Ervilha (1 concha)", 16, 5, 0),

        # DOCES
        ("Chocolate (25g)", 13, 2, 8),
        ("Chocolate meio amargo (25g)", 10, 3, 8),
        ("Brigadeiro (1 un)", 20, 3, 6),
        ("Doce de leite (1 colher sopa)", 20, 4, 6),
        ("Sorvete (1 bola)", 18, 3, 5),

        # FAST
        ("Pizza (1 fatia)", 30, 12, 10),
        ("Pastel (1 un)", 25, 6, 12),
        ("Coxinha (1 un)", 22, 7, 10),
        ("Pão com queijo (1 un)", 30, 12, 10),

        # TEMPEROS
        ("Azeite (1 colher sopa)", 0, 0, 14),
        ("Maionese (1 colher sopa)", 1, 0, 10),
        ("Ketchup (1 colher sopa)", 5, 0, 0),
        ("Mostarda (1 colher sopa)", 1, 0, 0),
    ]

    alimentos = {}
    for nome, c, p, g in base:
        alimentos[nome] = [_round_macro(c), _round_macro(p), _round_macro(g)]
        nome_p = nome.replace(")", ") (P)") if ")" in nome else f"{nome} (P)"
        alimentos[nome_p] = [_round_macro(c * 0.5), _round_macro(p * 0.5), _round_macro(g * 0.5)]
        nome_g = nome.replace(")", ") (G)") if ")" in nome else f"{nome} (G)"
        alimentos[nome_g] = [_round_macro(c * 1.5), _round_macro(p * 1.5), _round_macro(g * 1.5)]

    if len(alimentos) < 400:
        extra = {}
        for k, v in list(alimentos.items()):
            if "(G)" in k:
                extra[k.replace("(G)", "(2x)")] = [
                    _round_macro(v[0] * 1.33),
                    _round_macro(v[1] * 1.33),
                    _round_macro(v[2] * 1.33),
                ]
            if len(extra) + len(alimentos) >= 420:
                break
        alimentos.update(extra)

    return alimentos

ALIMENTOS = _gerar_catalogo_alimentos_400plus()

# =========================================================
# LOGIN UI
# =========================================================
if not st.session_state.logado:
    st.title("🧪 Saúde Kids - Acesso")

    if not HAS_COOKIES:
        st.caption("ℹ️ Login persistente desativado (extra-streamlit-components não instalado).")

    abas_login = st.tabs(["🔐 Entrar", "📝 Criar Conta", "❓ Esqueci Senha", "🔄 Alterar Senha"])

    # -------- ENTRAR --------
    with abas_login[0]:
        u = norm_email(st.text_input("E-mail", key="l_email"))
        s = norm_senha(st.text_input("Senha", type="password", key="l_pass"))

        if st.button("Acessar Aplicativo", use_container_width=True):
            rows = sb_select("users", {"email": u, "senha": s})
            ok = rows[0] if rows else None
            if ok:
                stt = get_plano_status(u)
                if stt.get('allowed'):
                    st.session_state.logado = True
                    st.session_state.user_email = u
                    cookie_set_email(u)
                    st.rerun()
                else:
                    st.session_state.logado = False
                    st.session_state.user_email = ''
                    st.session_state.pending_email = u
                    st.rerun()
            else:
                st.error("E-mail ou senha incorretos.")

    # -------- CRIAR CONTA --------
    with abas_login[1]:
        n_cad = (st.text_input("Nome Completo") or "").strip()
        e_cad = norm_email(st.text_input("E-mail para Cadastro"))
        s_cad = norm_senha(st.text_input("Senha para Cadastro", type="password"))

        if st.button("Realizar Cadastro", use_container_width=True):
            if not n_cad or not e_cad or not s_cad:
                st.warning("Preencha nome, e-mail e senha.")
            else:
                try:
                    sb_insert("users", {"nome": n_cad, "email": e_cad, "senha": s_cad})
                    garantir_plano(e_cad)
                    st.success("Conta criada com sucesso!")
                except Exception:
                    st.error("Este e-mail já está cadastrado.")

    # -------- ESQUECI SENHA --------
    with abas_login[2]:
        email_alvo = norm_email(st.text_input("Digite seu e-mail cadastrado"))
        if st.button("Recuperar Acesso", use_container_width=True):
            rows = sb_select("users", {"email": email_alvo})
            user = rows[0] if rows else None

            if user:
                nova = gerar_senha_temporaria()
                sb_update("users", {"email": email_alvo}, {"senha": nova})

                if enviar_senha_nova(email_alvo, nova):
                    st.success("Nova senha enviada para seu e-mail!")
                else:
                    st.warning("E-mail não configurado no app (sem GMAIL_APP_PASSWORD).")
                    st.info("Use a senha temporária abaixo para entrar e depois altere sua senha:")
                    st.code(nova)
            else:
                st.error("E-mail não encontrado.")

    # -------- ALTERAR SENHA --------
    with abas_login[3]:
        alt_em = norm_email(st.text_input("Confirme seu E-mail", key="alt_em"))
        alt_at = norm_senha(st.text_input("Senha Atual", type="password", key="alt_at"))
        alt_n1 = norm_senha(st.text_input("Nova Senha", type="password", key="alt_n1"))

        if st.button("Confirmar Alteração", use_container_width=True):
            ok_rows = sb_select("users", {"email": alt_em, "senha": alt_at})
            ok = ok_rows[0] if ok_rows else None
            if ok:
                sb_update("users", {"email": alt_em}, {"senha": alt_n1})
                st.success("Senha alterada com sucesso!")
            else:
                st.error("Dados atuais incorretos.")


    # ====== TELA DE TESTE EXPIRADO / MENSALIDADE ======
    if st.session_state.get("pending_email"):
        email_p = st.session_state.pending_email
        stt = get_plano_status(email_p)
        st.markdown("---")
        st.subheader("⏳ Seu teste acabou")
        te = stt.get("trial_end")
        if te:
            try:
                st.caption(f"Teste terminou em: {te.strftime('%d/%m/%Y')}")
            except Exception:
                pass
        st.info("Para continuar usando, solicite a mensalidade de **30 dias** pelo aplicativo. O admin vai receber sua mensagem.")

        msg = st.text_area("📩 Mensagem para o Admin (mensalidade)", value="Olá! Quero ativar minha mensalidade de 30 dias.")
        c1b, c2b = st.columns(2)
        with c1b:
            if st.button("✅ Enviar solicitação", use_container_width=True):
                registrar_mensagem_mensalidade(email_p, msg)
                st.success("Solicitação enviada! Aguarde retorno do admin.")
        with c2b:
            if st.button("↩️ Voltar para login", use_container_width=True):
                st.session_state.pending_email = ""
                st.rerun()
        st.stop()

    st.stop()

# =========================================================
# APP PRINCIPAL (ADMIN / USER)
# =========================================================
if st.session_state.user_email == "admin":
    st.title("🛡️ Painel Admin - Gestão Estratégica")
    t_usuarios, t_metricas, t_sugestoes, t_mensal, t_backup = st.tabs(
        ["👥 Pessoas Cadastradas", "📈 Crescimento e App", "📩 Sugestões", "💳 Mensalidades", "💾 Backup & Restauração"]
    )

    _users_rows = sb_select("users", order="email")
    df_users = pd.DataFrame(_users_rows)[["nome", "email"]] if _users_rows else pd.DataFrame(columns=["nome", "email"])

    with t_usuarios:
        st.subheader("Lista de Usuários")
        st.dataframe(df_users, use_container_width=True)
        st.metric("Total de Cadastros", len(df_users))
        st.markdown("---")

        st.markdown("---")
        st.subheader("🗑️ Excluir Usuário")
        st.caption("Remove o usuário do login e apaga os dados dele (Glicemia / Nutrição / Receita).")
        del_email = st.selectbox("Selecione o usuário para excluir", df_users["email"].tolist(), key="del_user_email")
        confirmar = st.checkbox("Confirmo que quero excluir este usuário e todos os dados", key="del_user_confirm")
        if st.button("🗑️ Excluir Agora", use_container_width=True, disabled=not confirmar):
            if del_email == "admin":
                st.error("Não é permitido excluir o admin.")
            else:
                sb_delete("users", {"email": del_email})
                sb_delete("plans", {"email": del_email})

                _delete_user_app_data(del_email)

                st.success(f"Usuário {del_email} excluído.")
                st.rerun()

        st.subheader("🔑 Alterar Senha de Usuário (Admin)")
        user_selecionado = st.selectbox("Selecione o E-mail do Usuário", df_users["email"].tolist())
        nova_senha_admin = norm_senha(st.text_input("Nova senha para este usuário", type="password"))
        if st.button("Confirmar Alteração de Senha", use_container_width=True):
            if nova_senha_admin:
                sb_update("users", {"email": user_selecionado}, {"senha": nova_senha_admin})
                st.success(f"Senha de {user_selecionado} alterada com sucesso!")
            else:
                st.warning("Digite uma senha antes de confirmar.")

    with t_metricas:
        c1, c2 = st.columns(2)
        with c1:
            st.write("### Distribuição de Registros (Glicemia)")
            df_uso = _read_table_df("glicemia")
            if "Usuario" in df_uso.columns and not df_uso.empty:
                uso_por_user = df_uso["Usuario"].value_counts().reset_index()
                uso_por_user.columns = ["Usuario", "Registros"]
                fig_pizza = px.pie(uso_por_user, values="Registros", names="Usuario", hole=.3)
                st.plotly_chart(fig_pizza, use_container_width=True)
            else:
                st.info("Sem dados.")
        with c2:
            st.write("### Cadastros (snapshot)")
            dados_c = pd.DataFrame({
                "Mês": ["Jan", "Fev", "Mar"],
                "Usuários": [max(1, len(df_users)//2), max(1, int(len(df_users)/1.1)), len(df_users)]
            })
            st.plotly_chart(px.line(dados_c, x="Mês", y="Usuários", markers=True), use_container_width=True)

    with t_sugestoes:
        st.subheader("📩 Sugestões recebidas")
        df_sug = _read_table_df("sugestoes")
        if not df_sug.empty:
            st.dataframe(df_sug[["Usuario", "Data", "Sugestão"]], use_container_width=True)
        else:
            st.info("Sem sugestões.")

    
    with t_mensal:
        st.subheader("💳 Mensalidades (Teste 20 dias + Plano 30 dias)")

        st.markdown("### Status de planos")
        try:
            df_pl = pd.DataFrame(sb_select("plans", order="email"))
            if df_pl.empty:
                df_pl = pd.DataFrame(columns=["email","trial_end","paid_until"])
        except Exception:
            df_pl = pd.DataFrame(columns=["email","trial_end","paid_until"])

        def _fmt_iso(x):
            try:
                if not x or str(x).strip()=="":
                    return ""
                return datetime.fromisoformat(str(x)).strftime("%d/%m/%Y")
            except Exception:
                return str(x)

        if df_pl.empty:
            st.info("Nenhum plano encontrado.")
        else:
            df_show = df_pl.copy()
            df_show["trial_end"] = df_show["trial_end"].apply(_fmt_iso)
            df_show["paid_until"] = df_show["paid_until"].apply(_fmt_iso)
            st.dataframe(df_show, use_container_width=True)

        st.markdown("---")
        st.subheader("✅ Ativar / Renovar mensalidade")
        emails = df_users["email"].tolist()
        alvo = st.selectbox("Usuário", emails, key="pay_user")
        dias = st.number_input("Dias de acesso", min_value=1, max_value=365, value=30, step=1, key="pay_days")
        if st.button("Ativar / Renovar", use_container_width=True):
            if alvo == "admin":
                st.warning("Admin já tem acesso liberado.")
            else:
                ativar_mensalidade(alvo, int(dias))
                st.success(f"Mensalidade ativada por {dias} dias para {alvo}.")
                st.rerun()

        st.markdown("---")
        st.subheader("📩 Solicitações recebidas no app")
        df_msgs = listar_mensagens_mensalidade()
        if df_msgs.empty:
            st.info("Sem solicitações ainda.")
        else:
            st.dataframe(df_msgs, use_container_width=True)

            msg_id = st.selectbox("Selecionar mensagem (ID)", df_msgs["id"].tolist(), key="msg_id_sel")
            colm1, colm2, colm3 = st.columns(3)
            with colm1:
                if st.button("Marcar como Visto", use_container_width=True):
                    marcar_mensagem_status(int(msg_id), "visto")
                    st.rerun()
            with colm2:
                if st.button("Marcar como Resolvido", use_container_width=True):
                    marcar_mensagem_status(int(msg_id), "resolvido")
                    st.rerun()
            with colm3:
                if st.button("Excluir Mensagem", use_container_width=True):
                    excluir_mensagem(int(msg_id))
                    st.rerun()

    with t_backup:
        st.subheader("💾 Backup Manual / Automático / Restauração")

        st.write("### 📦 Gerar Backup Manual")
        st.caption("Este backup é completo e global: leva todos os usuários cadastrados e todas as informações de todos os usuários.")
        if st.button("📦 Gerar Backup Agora", use_container_width=True):
            caminho = criar_backup_zip_em_disco()
            b, nome = criar_backup_zip_em_bytes()
            st.success(f"Backup completo gerado com sucesso: {caminho.name}")
            st.download_button(
                "⬇️ Baixar Backup Completo (.zip)",
                data=b,
                file_name=nome,
                mime="application/zip",
                use_container_width=True,
                key=f"download_backup_manual_{nome}",
            )

        st.markdown("---")
        st.write("### ♻️ Restauração Manual")
        st.caption("Ao restaurar, o sistema repõe usuários, glicemias, nutrição, configurações e mensagens do backup enviado.")
        up = st.file_uploader("Enviar arquivo .zip de backup", type=["zip"], key="backup_restore_uploader")
        if up is not None:
            if st.button("✅ Restaurar Agora", use_container_width=True):
                restaurados = restaurar_backup_zip_bytes(up.getvalue())
                if restaurados:
                    st.success("Restauração concluída com sucesso: " + ", ".join(restaurados))
                else:
                    st.warning("Nenhum arquivo do backup foi restaurado.")
                st.rerun()

        st.markdown("---")
        st.write("### ⏰ Backup Automático")
        st.info("Config: 1 backup por dia após **03:00** (Brasília).")
        if BACKUP_STATE_FILE.exists():
            st.caption(f"Último dia registrado: {BACKUP_STATE_FILE.read_text(encoding='utf-8').strip()}")

        st.markdown("---")
        st.write("### 🗂️ Backups existentes (pasta backups/)")
        df_bk = listar_backups()
        if df_bk.empty:
            st.warning("Nenhum backup encontrado.")
        else:
            df_show = df_bk.copy()
            df_show["data_hora"] = df_show["data_hora"].dt.strftime("%d/%m/%Y %H:%M:%S")
            st.dataframe(df_show[["arquivo", "data_hora", "tamanho_mb"]], use_container_width=True)

            st.markdown("#### Ações")
            selecionado = st.selectbox("Selecionar backup", df_bk["arquivo"].tolist())
            p_sel = BACKUP_DIR / selecionado

            colx1, colx2, colx3 = st.columns(3)
            with colx1:
                if p_sel.exists():
                    with open(p_sel, "rb") as f:
                        st.download_button(
                            "⬇️ Baixar Selecionado",
                            data=f.read(),
                            file_name=selecionado,
                            mime="application/zip",
                            use_container_width=True,
                            key=f"download_backup_existente_{selecionado}",
                        )
            with colx2:
                if st.button("🗑️ Apagar Selecionado", use_container_width=True):
                    if p_sel.exists():
                        p_sel.unlink()
                        st.success("Backup apagado.")
                        st.rerun()
            with colx3:
                if st.button("🧹 Apagar Antigos (7 dias)", use_container_width=True):
                    apagados = apagar_backups_antigos(dias_manter=7)
                    st.success(f"Apagados: {apagados}")
                    st.rerun()

else:
    tab1, tab2, tab3, tab4 = st.tabs(["📊 Glicemia", "🍽️ Nutrição", "⚙️ Receita", "📩 Sugerir Melhoria"])

    # ====== GLICEMIA ======
    with tab1:
        st.markdown('<div class="card">', unsafe_allow_html=True)

        dfg = carregar_glicemia_com_id()

        c1, c2 = st.columns([1, 2])
        with c1:
            v_gl = st.number_input("Valor Glicemia", 0, 600, 100, step=1, key="g_valor_glicemia")

            # ✅ Momentos extras (personalizado)
            MOMENTOS_BASE = MOMENTOS_ORDEM + ["Outro (personalizado)"]
            m_sel = st.selectbox("Momento", MOMENTOS_BASE)

            momento_extra = ""
            if m_sel == "Outro (personalizado)":
                momento_extra = (st.text_input("Digite o momento (ex: 'Antes Academia', 'Após Treino', 'Após Remédio')") or "").strip()
            m_gl = momento_extra if m_sel == "Outro (personalizado)" else m_sel

            # Regras de exibição (rápida/longa)
            MOMENTOS_RAPIDA = ["Antes Café", "Antes Almoço", "Antes Janta"]
            MOMENTOS_LONGA = ["Antes Café", "Antes Janta"]

            # Próxima medida (+2h) apenas para os "Antes ..." padrão
            dt_agora = agora_br()
            momento_apos, hora_apos = proxima_medida_apos(m_gl, dt_agora)
            if momento_apos and hora_apos:
                st.info(f"⏰ Próxima medida: **{momento_apos}** às **{hora_apos}** (2 horas após)")
            # RÁPIDA / LONGA (SUGESTÃO + EDIÇÃO MANUAL)

            # ===== RÁPIDA =====
            dose_r_sug, msg_r = "—", "Rápida não aplicável"
            dose_r_final = ""

            if m_gl in MOMENTOS_RAPIDA:
                dose_r_sug, msg_r = calc_insulina_rapida(int(v_gl), m_gl)  # ex: "3 UI"
                try:
                    sug_num = int(str(dose_r_sug).split()[0])
                except Exception:
                    sug_num = 0

                # --- RÁPIDA: mostra sugestão primeiro; editar é opcional ---
                dose_r_final = dose_r_sug  # padrão = sugestão

                st.markdown(
                    f'<div class="metric-box"><small>Rápida (sugestão)</small><br>'
                    f'<span class="dose-destaque">{dose_r_final}</span></div>',
                    unsafe_allow_html=True
                )

                editar_r = st.checkbox("✍️ Editar dose Rápida antes de salvar", value=False, key="editar_rapida")
                if editar_r:
                    dose_r_edit = st.number_input(
                        "Dose Rápida (UI)",
                        min_value=0, max_value=50, value=int(sug_num),
                        key="dose_rapida_edit"
                    )
                    dose_r_final = f"{int(dose_r_edit)} UI"

                    # mantém o card igual ao valor final
                    st.markdown(
                        f'<div class="metric-box" style="margin-top:10px;"><small>Rápida (ajustada)</small><br>'
                        f'<span class="dose-destaque">{dose_r_final}</span></div>',
                        unsafe_allow_html=True
                    )
            else:
                st.caption("Rápida: não aplicável neste momento.")
                dose_r_final = ""

            # ===== LONGA =====
            dose_l_sug, msg_l = "—", "Longa não aplicável"
            dose_l_final = ""

            if m_gl in MOMENTOS_LONGA:
                dose_l_sug, msg_l = calc_glargina(m_gl)  # ex: "8 UI"
                try:
                    sug_l = int(str(dose_l_sug).split()[0])
                except Exception:
                    sug_l = 0

                st.markdown(
                    f'<div class="metric-box" style="margin-top:10px;"><small>{msg_l} (sugestão)</small><br>'
                    f'<span class="dose-destaque">{dose_l_sug}</span></div>',
                    unsafe_allow_html=True
                )

                # --- LONGA: mostra sugestão primeiro; editar é opcional ---
                dose_l_final = dose_l_sug  # padrão = sugestão

                editar_l = st.checkbox("✍️ Editar dose Longa antes de salvar", value=False, key="editar_longa")
                if editar_l:
                    dose_l_edit = st.number_input(
                        "Dose Longa (UI)",
                        min_value=0, max_value=100, value=int(sug_l),
                        key="dose_longa_edit"
                    )
                    dose_l_final = f"{int(dose_l_edit)} UI"

                    st.markdown(
                        f'<div class="metric-box" style="margin-top:10px;"><small>Longa (ajustada)</small><br>'
                        f'<span class="dose-destaque">{dose_l_final}</span></div>',
                        unsafe_allow_html=True
                    )
            else:
                st.caption("Longa: não aplicável neste momento.")
                dose_l_final = ""

            # WhatsApp da glicemia separado da nutrição
            dose_r_msg = dose_r_final if dose_r_final else (dose_r_sug if dose_r_sug else "—")
            dose_l_msg = dose_l_final if dose_l_final else (dose_l_sug if dose_l_sug else "—")

            link_wpp = link_whatsapp_lembrete(
                momento=m_gl,
                valor_glicemia=int(v_gl),
                dose_rapida=dose_r_msg,
                dose_longa=dose_l_msg,
            )
            st.link_button("📲 Enviar glicemia no WhatsApp", link_wpp, use_container_width=True)

            if st.button("💾 Salvar Glicemia", use_container_width=True):
                # salva doses separadas (Rápida / Longa) para o histórico e para o PDF
                dose_r_save = ""
                dose_l_save = ""

                if m_gl in MOMENTOS_RAPIDA and "dose_r_final" in locals() and dose_r_final:
                    dose_r_save = str(dose_r_final).strip()

                if m_gl in MOMENTOS_LONGA and "dose_l_final" in locals() and dose_l_final:
                    dose_l_save = str(dose_l_final).strip()

                # mantém a coluna "Dose" (texto) para compatibilidade, mas salva também em colunas separadas
                dose_texto = ""
                if dose_r_save:
                    dose_texto = dose_r_save
                if dose_l_save:
                    dose_texto = (dose_texto + " | " if dose_texto else "") + f"Longa: {dose_l_save}"

                salvar_registro_glicemia(int(v_gl), m_gl, dose_texto, agora_br(), dose_rapida=dose_r_save, dose_longa=dose_l_save)
                st.rerun()
                st.rerun()

        with c2:
            if not dfg.empty:
                st.write("### Tendência")

                # Selecionar dia específico (inclui 'Geral (últimas 25)')
                datas = sorted(dfg["Data"].astype(str).unique().tolist())
                opcoes = ["Geral (últimas 25)"] + datas

                hoje_str = agora_br().strftime("%d/%m/%Y")
                ontem_str = (agora_br() - timedelta(days=1)).strftime("%d/%m/%Y")

                if ontem_str in opcoes:
                    idx_default = opcoes.index(ontem_str)
                elif hoje_str in opcoes:
                    idx_default = opcoes.index(hoje_str)
                else:
                    idx_default = len(opcoes) - 1

                dia_sel = st.selectbox("📅 Ver tendência do dia:", opcoes, index=idx_default, key="trend_day_sel")

                if dia_sel == "Geral (últimas 25)":
                    df_plot = dfg.copy().tail(25)
                    titulo = "Tendência (últimas 25)"
                else:
                    df_plot = dfg[dfg["Data"].astype(str) == dia_sel].copy()
                    titulo = f"Tendência do dia {dia_sel}"

                # Ordena por data/hora real e mostra eixo em HH:MM
                try:
                    df_plot["DT"] = pd.to_datetime(
                        df_plot["Data"].astype(str) + " " + df_plot["Hora"].astype(str),
                        dayfirst=True,
                        errors="coerce"
                    )
                    df_plot = df_plot.dropna(subset=["DT"]).sort_values("DT")
                    x_col = "DT"
                except Exception:
                    x_col = "Hora"

                fig = px.line(df_plot, x=x_col, y="Valor", markers=True, title=titulo)
                fig.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", font_color="white")
                if x_col == "DT":
                    fig.update_xaxes(tickformat="%H:%M")
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("Sem dados para gráfico ainda.")

        st.markdown("### ➕ Adicionar medida manual")

        with st.expander("Adicionar medida manualmente", expanded=False):
            colm1, colm2, colm3 = st.columns(3)

            with colm1:
                data_manual = st.date_input(
                    "Data",
                    value=agora_br().date(),
                    key="g_data_manual"
                )

            with colm2:
                hora_manual = st.text_input(
                    "Hora (HH:MM)",
                    value=agora_br().strftime("%H:%M"),
                    key="g_hora_manual"
                )

            with colm3:
                valor_manual = st.number_input(
                    "Valor Glicemia",
                    min_value=0,
                    max_value=600,
                    value=100,
                    step=1,
                    key="g_valor_manual"
                )

            colm4, colm5, colm6 = st.columns(3)

            with colm4:
                momento_manual = st.selectbox(
                    "Momento",
                    options=MOMENTOS_ORDEM + ["Outro"],
                    key="g_momento_manual"
                )

            with colm5:
                dose_rapida_manual = st.text_input(
                    "Dose Rápida",
                    value="0 UI",
                    key="g_dose_rapida_manual"
                )

            with colm6:
                dose_longa_manual = st.text_input(
                    "Dose Longa",
                    value="0 UI",
                    key="g_dose_longa_manual"
                )

            if st.button("💾 Salvar medida manual", use_container_width=True, key="btn_salvar_manual"):
                novo = pd.DataFrame([[
                    st.session_state.user_email,
                    data_manual.strftime("%d/%m/%Y"),
                    hora_manual.strip(),
                    int(valor_manual),
                    momento_manual,
                    dose_rapida_manual.strip(),
                    dose_longa_manual.strip(),
                    f"GL-{uuid.uuid4().hex[:10]}"
                ]], columns=[
                    "Usuario", "Data", "Hora", "Valor", "Momento",
                    "Dose_Rapida", "Dose_Longa", "ID"
                ])

                novo["Dose"] = novo.apply(lambda r: f"{r['Dose_Rapida']} | Longa: {r['Dose_Longa']}" if str(r['Dose_Longa']).strip() else str(r['Dose_Rapida']), axis=1)

                _append_df_table("glicemia", novo[["ID", "Usuario", "Data", "Hora", "Valor", "Momento", "Dose", "Dose_Rapida", "Dose_Longa"]])

                st.success("Medida manual salva com sucesso!")
                st.rerun()

        # ✅ HISTÓRICO EDITÁVEL (editar/excluir)
        st.markdown("### 🧾 Histórico (editar / excluir)")

        if not dfg.empty:
            df_hist = dfg.copy()

            if "Excluir" not in df_hist.columns:
                df_hist["Excluir"] = False

            cols_order = ["Excluir", "ID", "Data", "Hora", "Valor", "Momento", "Dose_Rapida", "Dose_Longa"]
            for c in cols_order:
                if c not in df_hist.columns:
                    df_hist[c] = ""

            df_hist = df_hist[cols_order].tail(80).reset_index(drop=True)

            # Normaliza tipos para o st.data_editor não quebrar
            df_hist["Excluir"] = df_hist["Excluir"].fillna(False).astype(bool)
            for col in ["ID", "Data", "Hora", "Momento", "Dose_Rapida", "Dose_Longa"]:
                if col not in df_hist.columns:
                    df_hist[col] = ""
                df_hist[col] = df_hist[col].fillna("").astype(str)
            if "Valor" not in df_hist.columns:
                df_hist["Valor"] = 0
            df_hist["Valor"] = pd.to_numeric(df_hist["Valor"], errors="coerce").fillna(0).astype(int)

            df_edit = st.data_editor(
                df_hist,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Excluir": st.column_config.CheckboxColumn("Excluir"),
                    "ID": st.column_config.TextColumn("ID", disabled=True),
                    "Data": st.column_config.TextColumn("Data", disabled=True),
                    "Hora": st.column_config.TextColumn("Hora", disabled=True),
                    "Momento": st.column_config.TextColumn("Momento"),
                    "Valor": st.column_config.NumberColumn("Valor", min_value=0, max_value=600, step=1),
                    "Dose_Rapida": st.column_config.TextColumn("Dose Rápida"),
                    "Dose_Longa": st.column_config.TextColumn("Dose Longa"),
                },
                key="glicemia_editor",
            )

            col_a, col_b = st.columns(2)
            with col_a:
                if st.button("✅ Salvar alterações do histórico", use_container_width=True):
                    aplicar_edicoes_e_exclusoes_glicemia(df_edit)
                    st.success("Histórico atualizado!")
                    # ✅ limpa o cache/estado do data_editor para recarregar do CSV
                    st.session_state.pop("glicemia_editor", None)

                    st.rerun()
            with col_b:
                st.caption("Marque 'Excluir' e clique em salvar para remover.")
        else:
            st.info("Sem registros ainda.")

        st.markdown("</div>", unsafe_allow_html=True)

    # ====== NUTRIÇÃO ======
    with tab2:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        dfn = carregar_dados_seguro(ARQ_N)

        # ---------- util: normalização (sem acento) ----------
        import unicodedata
        import difflib

        def _norm_txt(s: str) -> str:
            s = (s or "").strip().lower()
            s = unicodedata.normalize("NFD", s)
            s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
            s = " ".join(s.split())
            return s

        st.subheader("🍽️ Nutrição")
        m_nutri = st.selectbox("Refeição", MOMENTOS_ORDEM, key="n_m")

        # estado persistente da seleção (mantém vários itens sem “sumir”)
        # Fonte da verdade: widget multiselect (nutri_sel_mult)
        if "nutri_sel_mult" not in st.session_state:
            st.session_state.nutri_sel_mult = []

        # ✅ Evita erro do Streamlit: limpar multiselect só ANTES de criar o widget
        if st.session_state.get("_nutri_clear_pending", False):
            st.session_state["nutri_sel"] = []
            st.session_state["nutri_sel_mult"] = []
            st.session_state["nutri_add_sel"] = None
            st.session_state["_nutri_clear_pending"] = False

        busca_alim = st.text_input(
            "🔎 Buscar alimento",
            value="",
            placeholder="Digite para filtrar (ex: banana, frango, iogurte). Não precisa acento.",
            key="nutri_busca",
        )

        opts_all = list(ALIMENTOS.keys())
        qn = _norm_txt(busca_alim)

        if qn:
            # 1) contém (rápido)
            opts = [o for o in opts_all if qn in _norm_txt(o)]
            # 2) fallback por similaridade (quando a pessoa digita “quase certo”)
            if len(opts) < 20:
                mapping = {_norm_txt(o): o for o in opts_all}
                close = difflib.get_close_matches(qn, list(mapping.keys()), n=25, cutoff=0.55)
                for k in close:
                    o = mapping.get(k)
                    if o and o not in opts:
                        opts.append(o)
        else:
            # sem busca: mostra os mais usados (se houver) + restante
            try:
                if not dfn.empty and "Info" in dfn.columns:
                    usados = []
                    for s in dfn["Info"].astype(str).tail(80).tolist():
                        for part in [p.strip() for p in s.split(",") if p.strip()]:
                            if part in ALIMENTOS and part not in usados:
                                usados.append(part)
                    opts = usados + [o for o in opts_all if o not in usados]
                else:
                    opts = opts_all
            except Exception:
                opts = opts_all

        # limite para manter o app leve
        if len(opts) > 250:
            opts = opts[:250]

        colA, colB = st.columns([3, 1])
        with colA:
            pick = st.selectbox("Escolher alimento para adicionar", options=opts, key="nutri_pick")
        with colB:
            if st.button("➕ Adicionar", use_container_width=True):
                cur = list(st.session_state.get("nutri_sel_mult", []) or [])
                if pick and pick not in cur:
                    cur.append(pick)
                    st.session_state["nutri_sel_mult"] = cur  # força UI manter + adicionar

        # lista atual (permite remover)
        st.multiselect(
            "Alimentos selecionados (pode remover clicando no X)",
            options=opts_all,
            key="nutri_sel_mult",
        )

        sel = list(st.session_state.get("nutri_sel_mult", []) or [])


        c_tot = sum(ALIMENTOS[x][0] for x in sel) if sel else 0
        p_tot = sum(ALIMENTOS[x][1] for x in sel) if sel else 0
        g_tot = sum(ALIMENTOS[x][2] for x in sel) if sel else 0

        col1, col2, col3 = st.columns(3)
        col1.metric("Carbos", f"{c_tot}g")
        col2.metric("Proteínas", f"{p_tot}g")
        col3.metric("Gorduras", f"{g_tot}g")

        cbtn1, cbtn2, cbtn3 = st.columns([2, 2, 1])
        with cbtn1:
            if st.button("💾 Salvar Refeição", use_container_width=True, disabled=not bool(sel)):
                dt = agora_br()
                novo_n = pd.DataFrame(
                    [[st.session_state.user_email, dt.strftime("%d/%m/%Y"), m_nutri, ", ".join(sel), c_tot, p_tot, g_tot]],
                    columns=["Usuario", "Data", "Momento", "Info", "C", "P", "G"]
                )
                novo_n["ID"] = [f"NT-{uuid.uuid4().hex[:12]}"]
                _append_df_table("nutricao", novo_n[["ID", "Usuario", "Data", "Momento", "Info", "C", "P", "G"]])

                st.success("Refeição salva!")
                st.session_state["_nutri_clear_pending"] = True
                st.rerun()

        with cbtn2:
            link_nutri = link_whatsapp_nutricao(
                momento=m_nutri,
                alimentos=sel,
                c_tot=c_tot,
                p_tot=p_tot,
                g_tot=g_tot,
            )
            st.link_button(
                "📲 Enviar nutrição no WhatsApp",
                link_nutri,
                use_container_width=True,
                disabled=not bool(sel),
            )

        with cbtn3:
            if st.button("🧹 Limpar seleção", use_container_width=True):
                st.session_state["_nutri_clear_pending"] = True
                st.rerun()
        st.markdown("### Últimas refeições")

        # Mostra e permite excluir linhas do usuário logado
        if dfn is None or dfn.empty:
            st.info("Sem refeições registradas ainda.")
        else:
            dfn_all = dfn.copy()
            if "Usuario" in dfn_all.columns:
                dfn_user = dfn_all[dfn_all["Usuario"] == st.session_state.user_email].copy()
            else:
                dfn_user = dfn_all.copy()

            if dfn_user.empty:
                st.info("Sem refeições registradas ainda.")
            else:
                # Mantém referência ao ID original para exclusão segura
                if "ID" not in dfn_user.columns:
                    dfn_user["ID"] = [f"NT-{uuid.uuid4().hex[:12]}" for _ in range(len(dfn_user))]
                dfn_user["_row_id"] = dfn_user["ID"].astype(str)
                view = dfn_user[["Data", "Momento", "Info", "C", "P", "G", "_row_id"]].tail(12).copy()
                view.insert(0, "Excluir", False)

                edited = st.data_editor(
                    view,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Excluir": st.column_config.CheckboxColumn("Excluir"),
                        "_row_id": st.column_config.TextColumn("_row_id", disabled=True),
                    },
                    disabled=["Data", "Momento", "Info", "C", "P", "G", "_row_id"],
                    key="nutri_last_editor",
                )

                cdel1, cdel2 = st.columns([1, 2])
                with cdel1:
                    if st.button("🗑️ Excluir selecionadas", use_container_width=True):
                        to_del = edited[edited["Excluir"] == True]["_row_id"].tolist()
                        if not to_del:
                            st.warning("Marque pelo menos 1 linha em 'Excluir'.")
                        else:
                            # Remove do arquivo completo (todas as pessoas), baseado no índice original
                            for i in to_del:
                                sb_delete("nutricao", {"id": str(i), "usuario": st.session_state.user_email})
                            st.success(f"Removido: {len(to_del)} registro(s).")
                            st.rerun()
                with cdel2:
                    st.caption("Dica: marque 'Excluir' e clique em **Excluir selecionadas** para remover do histórico de Nutrição.")

        st.markdown("</div>", unsafe_allow_html=True)

    # ====== RECEITA ======
    with tab3:
        st.markdown('<div class="card">', unsafe_allow_html=True)

        df_r_all = _read_table_df("receita")
        if not df_r_all.empty and "Usuario" in df_r_all.columns:
            r_u = df_r_all[df_r_all["Usuario"] == st.session_state.user_email]
        else:
            r_u = pd.DataFrame()
        v = r_u.iloc[0] if not r_u.empty else {}

        st.subheader("⚡ Receita Rápida (por faixas)")

        st.markdown("**🌞 Manhã**")
        cm1, cm2, cm3 = st.columns(3)
        with cm1:
            m1_min = st.number_input("Faixa 1 - Mín", value=int(v.get("manha_f1_min", 70)), key="m1_min_u")
            m1_max = st.number_input("Faixa 1 - Máx", value=int(v.get("manha_f1_max", 150)), key="m1_max_u")
            m1_dose = st.number_input("Dose Faixa 1 (UI)", value=int(v.get("manha_f1_dose", 3)), key="m1_dose_u")
        with cm2:
            m2_min = st.number_input("Faixa 2 - Mín", value=int(v.get("manha_f2_min", 151)), key="m2_min_u")
            m2_max = st.number_input("Faixa 2 - Máx", value=int(v.get("manha_f2_max", 300)), key="m2_max_u")
            m2_dose = st.number_input("Dose Faixa 2 (UI)", value=int(v.get("manha_f2_dose", 5)), key="m2_dose_u")
        with cm3:
            m3_min = st.number_input("Faixa 3 - Mín", value=int(v.get("manha_f3_min", 301)), key="m3_min_u")
            m3_max = st.number_input("Faixa 3 - Máx", value=int(v.get("manha_f3_max", 600)), key="m3_max_u")
            m3_dose = st.number_input("Dose Faixa 3 (UI)", value=int(v.get("manha_f3_dose", 8)), key="m3_dose_u")

        st.markdown("---")
        st.markdown("**🌙 Noite**")
        cn1, cn2, cn3 = st.columns(3)
        with cn1:
            n1_min = st.number_input("Faixa 1 - Mín", value=int(v.get("noite_f1_min", 70)), key="n1_min_u")
            n1_max = st.number_input("Faixa 1 - Máx", value=int(v.get("noite_f1_max", 150)), key="n1_max_u")
            n1_dose = st.number_input("Dose Faixa 1 (UI)", value=int(v.get("noite_f1_dose", 3)), key="n1_dose_u")
        with cn2:
            n2_min = st.number_input("Faixa 2 - Mín", value=int(v.get("noite_f2_min", 151)), key="n2_min_u")
            n2_max = st.number_input("Faixa 2 - Máx", value=int(v.get("noite_f2_max", 300)), key="n2_max_u")
            n2_dose = st.number_input("Dose Faixa 2 (UI)", value=int(v.get("noite_f2_dose", 5)), key="n2_dose_u")
        with cn3:
            n3_min = st.number_input("Faixa 3 - Mín", value=int(v.get("noite_f3_min", 301)), key="n3_min_u")
            n3_max = st.number_input("Faixa 3 - Máx", value=int(v.get("noite_f3_max", 600)), key="n3_max_u")
            n3_dose = st.number_input("Dose Faixa 3 (UI)", value=int(v.get("noite_f3_dose", 8)), key="n3_dose_u")

        st.markdown("---")
        st.subheader("🩸 Longa (dose fixa)")
        gl1, gl2 = st.columns(2)
        with gl1:
            glargina_cafe_ui = st.number_input("Longa - Antes Café (UI)", value=int(float(v.get("glargina_cafe_ui", 0) or 0)), key="gl_cafe")
        with gl2:
            glargina_janta_ui = st.number_input("Longa - Antes Janta (UI)", value=int(float(v.get("glargina_janta_ui", 0) or 0)), key="gl_janta")

        if st.button("💾 Salvar Receita", use_container_width=True):
            nova_rec = pd.DataFrame([{
                "Usuario": st.session_state.user_email,

                "manha_f1_min": m1_min, "manha_f1_max": m1_max, "manha_f1_dose": m1_dose,
                "manha_f2_min": m2_min, "manha_f2_max": m2_max, "manha_f2_dose": m2_dose,
                "manha_f3_min": m3_min, "manha_f3_max": m3_max, "manha_f3_dose": m3_dose,

                "noite_f1_min": n1_min, "noite_f1_max": n1_max, "noite_f1_dose": n1_dose,
                "noite_f2_min": n2_min, "noite_f2_max": n2_max, "noite_f2_dose": n2_dose,
                "noite_f3_min": n3_min, "noite_f3_max": n3_max, "noite_f3_dose": n3_dose,

                "glargina_cafe_ui": glargina_cafe_ui,
                "glargina_janta_ui": glargina_janta_ui,
            }])

            _replace_user_rows("receita", st.session_state.user_email, nova_rec)
            st.success("Receita salva com sucesso!")

        st.markdown("</div>", unsafe_allow_html=True)

    # ====== SUGESTÃO ======
    with tab4:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        txt = st.text_area("Sugestão de Melhoria:")
        if st.button("Enviar Sugestão", use_container_width=True):
            if txt.strip():
                dt = agora_br().strftime("%d/%m/%Y %H:%M")
                novo_m = pd.DataFrame([[f"SG-{uuid.uuid4().hex[:12]}", st.session_state.user_email, dt, txt.strip()]], columns=["ID", "Usuario", "Data", "Sugestão"])
                _append_df_table("sugestoes", novo_m)
                st.success("Enviado com sucesso!")
            else:
                st.warning("Digite uma sugestão antes de enviar.")
        st.markdown("</div>", unsafe_allow_html=True)



def gerar_pdf_bytes(df_g: pd.DataFrame, df_n: pd.DataFrame) -> bytes:
    """
    Gera um PDF com:
      - Resumo
      - Tabela no formato 'Data x Momento' (igual ao Excel Glicemia_Resumo)
      - Gráfico de tendência (últimas medições)
    """
    desired_order = ["Antes Café","Após Café","Antes Almoço","Após Almoço","Antes Janta","Após Janta","Madrugada"]
    if not HAS_REPORTLAB:
        return b""

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=1.3*cm,
        leftMargin=1.3*cm,
        topMargin=1.2*cm,
        bottomMargin=1.2*cm
    )
    styles = getSampleStyleSheet()
    story = []

    # ===== Cabeçalho =====
    titulo = f"Relatório Saúde Kids - {st.session_state.user_email}"
    story.append(Paragraph(titulo, styles["Title"]))
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"Gerado em: {agora_br().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 10))

    # ===== Resumo Glicemia =====
    story.append(Paragraph("Glicemia - Resumo", styles["Heading2"]))
    if df_g is None or df_g.empty:
        story.append(Paragraph("Sem registros de glicemia.", styles["Normal"]))
        story.append(Spacer(1, 10))
    else:
        vals = pd.to_numeric(df_g.get("Valor", pd.Series([], dtype=float)), errors="coerce").dropna()
        resumo = [
            ["Registros", str(len(df_g))],
            ["Mínimo", str(int(vals.min())) if not vals.empty else "-"],
            ["Máximo", str(int(vals.max())) if not vals.empty else "-"],
            ["Média", f"{vals.mean():.1f}" if not vals.empty else "-"],
        ]
        t = Table(resumo, colWidths=[5.2*cm, 11.0*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
            ("FONTSIZE", (0,0), (-1,-1), 10),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("PADDING", (0,0), (-1,-1), 6),
        ]))
        story.append(t)
        story.append(Spacer(1, 10))

    # ===== Tabela no formato Excel (Data x Momento) =====
    story.append(Paragraph("Glicemia - Tabela (Data x Momento)", styles["Heading2"]))
    if df_g is None or df_g.empty:
        story.append(Paragraph("Sem dados para tabela.", styles["Normal"]))
        story.append(Spacer(1, 10))
    else:
        # Pivot igual ao Excel
        pivot = df_g.pivot_table(index="Data", columns="Momento", values="Valor", aggfunc="last")
        pivot = pivot.reindex(columns=ordenar_colunas_momentos(list(pivot.columns)))
        pivot = pivot.sort_index()

        # Limitar para caber no PDF (últimas 20 datas)
        if len(pivot) > 31:
            pivot_show = pivot.tail(31).copy()
        else:
            pivot_show = pivot.copy()

        # Separar colunas base vs extras
        base_order, extras_cols = separar_momentos_extras(list(pivot_show.columns))
        base_cols_present = [c for c in desired_order if c in pivot_show.columns]

        # ===== Tabela principal: só momentos base (na ordem correta) =====
        cols = ["Data"] + base_cols_present
        data_tbl = [cols]
        for idx, row in pivot_show.iterrows():
            line = [str(idx)]
            for c in base_cols_present:
                v = row.get(c, "")
                if pd.isna(v):
                    line.append("")
                else:
                    try:
                        line.append(str(int(v)))
                    except Exception:
                        line.append(str(v))
            data_tbl.append(line)

        # larguras
        ncols = len(cols)
        total_w = 18.0 * cm
        w_data = 3.0 * cm
        w_rest = (total_w - w_data) / max(1, ncols - 1)
        col_widths = [w_data] + [w_rest] * (ncols - 1)

        table = Table(data_tbl, repeatRows=1, colWidths=col_widths)
        style_cmds = [
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("FONTSIZE", (0,0), (-1,-1), 7),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN", (0,0), (-1,0), "CENTER"),
            ("ALIGN", (0,1), (0,-1), "LEFT"),
            ("ALIGN", (1,1), (-1,-1), "CENTER"),
            ("PADDING", (0,0), (-1,-1), 3),
        ]

        # cores por célula
        for r in range(1, len(data_tbl)):
            for c in range(1, len(cols)):
                txt = data_tbl[r][c]
                if not txt:
                    continue
                try:
                    val = int(float(txt))
                except Exception:
                    continue
                if val < 70:
                    style_cmds.append(("BACKGROUND", (c, r), (c, r), colors.HexColor("#FFFFE0")))
                elif val > 180:
                    style_cmds.append(("BACKGROUND", (c, r), (c, r), colors.HexColor("#FFB6C1")))
                else:
                    style_cmds.append(("BACKGROUND", (c, r), (c, r), colors.HexColor("#C8E6C9")))

        table.setStyle(TableStyle(style_cmds))
        story.append(table)
        story.append(Spacer(1, 10))

        # ===== Medidas extras (abaixo) =====
        if extras_cols:
            story.append(Paragraph("Medidas extras", styles["Heading3"]))
            # pega últimos 30 registros extras (independente do pivot)
            df_extras = df_g.copy()
            df_extras = df_extras[~df_extras["Momento"].isin(desired_order)].copy()

            # Ordenar por datetime e pegar últimas 30
            df_extras["DT"] = pd.to_datetime(df_extras["Data"].astype(str) + " " + df_extras["Hora"].astype(str),
                                             dayfirst=True, errors="coerce")
            df_extras = df_extras.dropna(subset=["DT"]).sort_values("DT").tail(30)

            if df_extras.empty:
                story.append(Paragraph("Sem medidas extras registradas.", styles["Normal"]))
            else:
                cols_e = ["Data", "Hora", "Momento", "Valor"]
                data_e = [cols_e] + df_extras[cols_e].astype(str).values.tolist()
                tE = Table(data_e, repeatRows=1, colWidths=[3*cm, 2*cm, 9*cm, 2*cm])
                tE.setStyle(TableStyle([
                    ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
                    ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
                    ("FONTSIZE", (0,0), (-1,-1), 7),
                    ("VALIGN", (0,0), (-1,-1), "TOP"),
                    ("PADDING", (0,0), (-1,-1), 3),
                ]))
                story.append(tE)

        # ===== Doses separadas (Longa e Rápida) no formato da Glicemia (Data x Momento) =====
        story.append(Spacer(1, 8))

        def _extrair(tag: str, s: str) -> str:
            try:
                import re
                m = re.search(rf"{tag}\s*:\s*([^|]+)", str(s), flags=re.IGNORECASE)
                return (m.group(1).strip() if m else "")
            except Exception:
                return ""

        def _mk_pivot_doses(df_src: pd.DataFrame, col_val: str):
            # df_src precisa ter Data, Hora, Momento e Dose; col_val é "R" ou "L"
            if df_src is None or df_src.empty:
                return None

            pivot = df_src.pivot_table(index="Data", columns="Momento", values=col_val, aggfunc="last")
            pivot = pivot.reindex(columns=ordenar_colunas_momentos(list(pivot.columns)))
            pivot = pivot.sort_index()

            # Limitar (últimas 31 datas)
            if len(pivot) > 31:
                return pivot.tail(31).copy()
            return pivot.copy()

        def _render_pivot_table(title: str, pivot_show: pd.DataFrame):
            story.append(Paragraph(title, styles["Heading3"]))
            if pivot_show is None or pivot_show.empty:
                story.append(Paragraph("Sem registros.", styles["Normal"]))
                story.append(Spacer(1, 10))
                return

            # manter só momentos base (igual glicemia)
            base_order, _extras_cols = separar_momentos_extras(list(pivot_show.columns))
            base_cols_present = [c for c in desired_order if c in pivot_show.columns]

            cols = ["Data"] + base_cols_present
            data_tbl = [cols]
            for idx, row in pivot_show.iterrows():
                line = [str(idx)]
                for c in base_cols_present:
                    v = row.get(c, "")
                    if pd.isna(v):
                        line.append("")
                    else:
                        line.append(str(v))
                data_tbl.append(line)

            ncols = len(cols)
            total_w = 18.0 * cm
            w_data = 3.0 * cm
            w_rest = (total_w - w_data) / max(1, ncols - 1)
            col_widths = [w_data] + [w_rest] * (ncols - 1)

            table = Table(data_tbl, repeatRows=1, colWidths=col_widths)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
                ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
                ("FONTSIZE", (0,0), (-1,-1), 7),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                ("ALIGN", (0,0), (-1,0), "CENTER"),
                ("ALIGN", (0,1), (0,-1), "LEFT"),
                ("ALIGN", (1,1), (-1,-1), "CENTER"),
                ("PADDING", (0,0), (-1,-1), 3),
            ]))
            story.append(table)
            story.append(Spacer(1, 10))

        try:
            # Base de doses vem do mesmo DF (coluna "Dose" salva no app)
            df_dose = df_g.copy()
            df_dose["Dose"] = df_dose.get("Dose", "").astype(str)
            df_dose = df_dose[df_dose["Dose"].str.strip().ne("")].copy()

            if df_dose.empty:
                story.append(Paragraph("Doses de insulina", styles["Heading3"]))
                story.append(Paragraph("Sem doses registradas.", styles["Normal"]))
                story.append(Spacer(1, 12))
            else:
                df_dose["DT"] = pd.to_datetime(
                    df_dose["Data"].astype(str) + " " + df_dose["Hora"].astype(str),
                    dayfirst=True, errors="coerce"
                )
                df_dose = df_dose.dropna(subset=["DT"]).sort_values("DT")

                # Extrai valores do texto salvo: "Rápida: X UI | Longa: Y UI"
                # usa colunas separadas se existirem; senão, faz fallback para o texto legado "Dose"
                if "Dose_Rapida" in df_dose.columns:
                    df_dose["R"] = df_dose["Dose_Rapida"].astype(str)
                else:
                    df_dose["R"] = df_dose["Dose"].apply(lambda s: _extrair("Rápida", s) or _extrair("Rapida", s))

                if "Dose_Longa" in df_dose.columns:
                    df_dose["L"] = df_dose["Dose_Longa"].astype(str)
                else:
                    df_dose["L"] = df_dose["Dose"].apply(lambda s: _extrair("Longa", s))

                # Mantém só linhas com algo
                df_dose = df_dose[(df_dose["R"].astype(str).str.strip().ne("")) | (df_dose["L"].astype(str).str.strip().ne(""))].copy()

                # Tabelas separadas: primeiro Longa, depois Rápida
                df_l = df_dose[df_dose["L"].astype(str).str.strip().ne("")].copy()
                df_r = df_dose[df_dose["R"].astype(str).str.strip().ne("")].copy()

                pivot_l = _mk_pivot_doses(df_l, "L")
                pivot_r = _mk_pivot_doses(df_r, "R")

                _render_pivot_table("Insulina Longa - Tabela (Data x Momento)", pivot_l)
                _render_pivot_table("Insulina Rápida - Tabela (Data x Momento)", pivot_r)

                # ===== Nutrição - Últimos registros (abaixo das doses) =====
                story.append(Spacer(1, 6))
                story.append(Paragraph("Nutrição - Últimos registros", styles["Heading3"]))

                try:
                    if df_n is None or df_n.empty:
                        story.append(Paragraph("Sem registros de nutrição.", styles["Normal"]))
                        story.append(Spacer(1, 12))
                    else:
                        dfn_pdf = df_n.copy()

                        # garante colunas
                        for col in ["Data", "Momento", "Info", "C", "P", "G"]:
                            if col not in dfn_pdf.columns:
                                dfn_pdf[col] = ""

                        # ordena por data (e hora, se existir)
                        if "Hora" in dfn_pdf.columns:
                            dfn_pdf["DT"] = pd.to_datetime(
                                dfn_pdf["Data"].astype(str) + " " + dfn_pdf["Hora"].astype(str),
                                dayfirst=True, errors="coerce"
                            )
                        else:
                            dfn_pdf["DT"] = pd.to_datetime(dfn_pdf["Data"].astype(str), dayfirst=True, errors="coerce")

                        dfn_pdf = dfn_pdf.dropna(subset=["DT"]).sort_values("DT").tail(20)

                        if dfn_pdf.empty:
                            story.append(Paragraph("Sem registros de nutrição.", styles["Normal"]))
                            story.append(Spacer(1, 12))
                        else:
                            cols_n = ["Data", "Momento", "Info", "C", "P", "G"]
                            data_n = [cols_n] + dfn_pdf[cols_n].astype(str).values.tolist()
                            tN = Table(data_n, repeatRows=1, colWidths=[3*cm, 3*cm, 7.5*cm, 1.5*cm, 1.5*cm, 1.5*cm])
                            tN.setStyle(TableStyle([
                                ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
                                ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
                                ("FONTSIZE", (0,0), (-1,-1), 7),
                                ("VALIGN", (0,0), (-1,-1), "TOP"),
                                ("PADDING", (0,0), (-1,-1), 3),
                            ]))
                            story.append(tN)
                            story.append(Spacer(1, 12))
                except Exception:
                    story.append(Paragraph("Não foi possível montar a seção de nutrição.", styles["Normal"]))
                    story.append(Spacer(1, 12))

        except Exception:
            story.append(Paragraph("Não foi possível montar as tabelas de doses.", styles["Normal"]))
            story.append(Spacer(1, 12))


    # ===== Tendência por dia (últimos 30 dias) =====
    story.append(Paragraph("Tendência por dia (últimos 30 dias)", styles["Heading2"]))
    if df_g is None or df_g.empty:
        story.append(Paragraph("Sem dados para gráfico.", styles["Normal"]))
    elif not HAS_MPL:
        story.append(Paragraph("Matplotlib não disponível para gerar gráfico.", styles["Normal"]))
    else:
        try:
            df_plot = df_g.copy()
            df_plot["DT"] = pd.to_datetime(
                df_plot["Data"].astype(str) + " " + df_plot["Hora"].astype(str),
                dayfirst=True,
                errors="coerce"
            )
            df_plot = df_plot.dropna(subset=["DT"]).sort_values("DT")

            # Filtra último mês (30 dias)
            dt_max = df_plot["DT"].max()
            dt_min = dt_max - pd.Timedelta(days=30)
            df_plot = df_plot[df_plot["DT"] >= dt_min].copy()

            if df_plot.empty:
                story.append(Paragraph("Sem dados no período de 30 dias.", styles["Normal"]))
            else:
                # Agrupa por dia (data)
                df_plot["DIA"] = df_plot["DT"].dt.strftime("%d/%m/%Y")
                dias = df_plot["DIA"].unique().tolist()

                # Para não ficar gigante, limita a 31 dias (máximo)
                dias = dias[-31:]

                import matplotlib.dates as mdates

                for dia in dias:
                    dfd = df_plot[df_plot["DIA"] == dia].copy()
                    if dfd.empty:
                        continue

                    story.append(Paragraph(f"Dia {dia}", styles["Heading3"]))

                    fig = plt.figure(figsize=(7.5, 2.6), dpi=150)
                    ax = fig.add_subplot(111)

                    vals = pd.to_numeric(dfd["Valor"], errors="coerce")
                    ax.plot(dfd["DT"], vals, marker="o")

                    ax.set_ylabel("Glicemia")
                    ax.set_xlabel("Hora (HH:MM)")

                    # Eixo X só com hora
                    locator = mdates.AutoDateLocator(minticks=4, maxticks=8)
                    formatter = mdates.DateFormatter("%H:%M")
                    ax.xaxis.set_major_locator(locator)
                    ax.xaxis.set_major_formatter(formatter)

                    ax.tick_params(axis="x", rotation=45, labelsize=8)
                    ax.tick_params(axis="y", labelsize=8)
                    ax.grid(True, alpha=0.3)

                    fig.tight_layout()

                    img_buf = BytesIO()
                    fig.savefig(img_buf, format="png")
                    plt.close(fig)
                    img_buf.seek(0)

                    img = Image(img_buf)
                    img.drawWidth = 18.0 * cm
                    img.drawHeight = 6.0 * cm
                    story.append(img)
                    story.append(Spacer(1, 10))

        except Exception:
            story.append(Paragraph("Falha ao gerar gráficos por dia.", styles["Normal"]))

    story.append(Spacer(1, 10))

    # ===== Nutrição (opcional) =====
    story.append(Paragraph("Nutrição - Últimos registros", styles["Heading2"]))
    if df_n is None or df_n.empty:
        story.append(Paragraph("Sem registros de nutrição.", styles["Normal"]))
    else:
        ult_n = df_n.copy().tail(15)
        cols_n = [c for c in ["Data", "Momento", "Info", "C", "P", "G"] if c in ult_n.columns]
        data_tbl_n = [cols_n] + ult_n[cols_n].astype(str).values.tolist()
        col_widths = [3*cm, 3*cm, 7*cm, 1.5*cm, 1.5*cm, 1.5*cm][:len(cols_n)]
        t3 = Table(data_tbl_n, repeatRows=1, colWidths=col_widths)
        t3.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("FONTSIZE", (0,0), (-1,-1), 7),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("PADDING", (0,0), (-1,-1), 3),
        ]))
        story.append(t3)

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    return pdf


# ================= EXCEL (SIDEBAR) =================
st.sidebar.markdown("---")
if st.sidebar.button("📥 Gerar Excel Completo"):
    df_e_g = carregar_glicemia_com_id()
    df_e_n = carregar_dados_seguro(ARQ_N)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 1) Aba Glicemia (Tabela completa, incluindo momentos extras)
        if not df_e_g.empty:
            df_e_g = _ensure_id_column(df_e_g.copy(), "ID", "GL")
            df_e_g.to_excel(writer, sheet_name="Glicemia_Tabela", index=False)
            wsT = writer.sheets["Glicemia_Tabela"]
            for cell in wsT[1]:
                cell.alignment = Alignment(horizontal="center")

            # 2) Aba Glicemia (Resumo pivot por Data x Momento)
            pivot = df_e_g.pivot_table(index="Data", columns="Momento", values="Valor", aggfunc="last")
            pivot = pivot.reindex(columns=ordenar_colunas_momentos(list(pivot.columns)))
            pivot.to_excel(writer, sheet_name="Glicemia_Resumo")
            ws1 = writer.sheets["Glicemia_Resumo"]

            f_ok = PatternFill("solid", fgColor="C8E6C9")  # verde claro
            f_hi = PatternFill("solid", fgColor="FFB6C1")  # vermelho claro
            f_lo = PatternFill("solid", fgColor="FFFFE0")  # amarelo claro

            for row in ws1.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    if cell.value is None or str(cell.value) == "nan":
                        continue
                    try:
                        val = int(cell.value)
                        cell.alignment = Alignment(horizontal="center")
                        if val < 70:
                            cell.fill = f_lo
                        elif val > 180:
                            cell.fill = f_hi
                        else:
                            cell.fill = f_ok
                    except Exception:
                        pass

        # 3) Aba Nutrição
        if not df_e_n.empty:
            df_e_n.to_excel(writer, sheet_name="Nutrição", index=False)
            ws2 = writer.sheets["Nutrição"]
            for cell in ws2[1]:
                cell.alignment = Alignment(horizontal="center")

    st.sidebar.download_button("⬇️ Baixar Agora", output.getvalue(), file_name="Relatorio_Saude_Kids.xlsx", use_container_width=True)


# ================= PDF (SIDEBAR) =================
st.sidebar.markdown("---")
st.sidebar.subheader("🧾 Relatório em PDF")

if not HAS_REPORTLAB:
    st.sidebar.warning("Para gerar PDF, adicione **reportlab** no requirements.txt e faça deploy novamente.")
else:
    if st.sidebar.button("🧾 Gerar PDF Relatório", use_container_width=True):
        df_pdf_g = carregar_glicemia_com_id()
        df_pdf_n = carregar_dados_seguro(ARQ_N)
        pdf_bytes = gerar_pdf_bytes(df_pdf_g, df_pdf_n)
        if not pdf_bytes:
            st.sidebar.error("Não foi possível gerar o PDF.")
        else:
            nome_pdf = f"Relatorio_Saude_Kids_{agora_br().strftime('%Y-%m-%d_%H-%M')}.pdf"
            st.sidebar.download_button("⬇️ Baixar PDF", pdf_bytes, file_name=nome_pdf, use_container_width=True)


# ================= SAIR =================
if st.sidebar.button("🚪 Sair"):
    st.session_state.logado = False
    st.session_state.user_email = ""
    cookie_clear()
    st.rerun()
