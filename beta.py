# =========================================================
# SAÚDE KIDS - BETA (VERSÃO COMPLETA, ESTÁVEL)
# - Login (SQLite) + Cookie opcional
# - Admin: usuários, métricas, sugestões, backup/restore
# - Glicemia: momentos extras (personalizado) + CRUD (editar/excluir no histórico)
# - Excel: Glicemia_Tabela (tudo) + Glicemia_Resumo (pivot) + Nutrição
# =========================================================

import os
import random
import shutil
import sqlite3
import string
import uuid
import zipfile
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

import pandas as pd
import plotly.express as px
import pytz
import streamlit as st
from email.mime.text import MIMEText
from openpyxl.styles import Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

# =========================================================
# (OPCIONAL) LOGIN PERSISTENTE POR COOKIE (NÃO BUGA SE NÃO TIVER)
# Para ativar no Streamlit Cloud: requirements.txt -> extra-streamlit-components==0.1.60
# =========================================================
try:
    import extra_streamlit_components as stx  # type: ignore
    _cookie_manager = stx.CookieManager()
    HAS_COOKIES = True
except Exception:
    _cookie_manager = None
    HAS_COOKIES = False

COOKIE_KEY = "SK_LOGIN_EMAIL"
COOKIE_DIAS = 30

# ================= CONFIGURAÇÕES INICIAIS =================
fuso_br = pytz.timezone("America/Sao_Paulo")
st.set_page_config(page_title="Saúde Kids BETA", page_icon="🧪", layout="wide")

ARQ_G = "dados_glicemia_BETA.csv"
ARQ_N = "dados_nutricao_BETA.csv"
ARQ_R = "config_receita_BETA.csv"
ARQ_M = "mensagens_admin_BETA.csv"

# ================= NORMALIZAÇÃO =================
def norm_email(x: str) -> str:
    return (x or "").strip().lower()

def norm_senha(x: str) -> str:
    return (x or "").strip()

# ================= TEMPO =================
def agora_br() -> datetime:
    return datetime.now(fuso_br)

# ================= BACKUP / RESTORE =================
BACKUP_DIR = Path("backups")
BACKUP_DIR.mkdir(exist_ok=True)
BACKUP_STATE_FILE = BACKUP_DIR / "last_auto_backup.txt"

ARQUIVOS_BACKUP = [
    "usuarios.db",
    ARQ_G,
    ARQ_N,
    ARQ_R,
    ARQ_M,
]

def criar_backup_zip_em_bytes():
    ts = agora_br().strftime("%Y-%m-%d_%H-%M-%S")
    nome = f"backup_saude_kids_{ts}.zip"
    out = BytesIO()
    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for arq in ARQUIVOS_BACKUP:
            if os.path.exists(arq):
                z.write(arq)
    out.seek(0)
    return out.getvalue(), nome

def criar_backup_zip_em_disco():
    ts = agora_br().strftime("%Y-%m-%d_%H-%M-%S")
    nome = f"backup_saude_kids_{ts}.zip"
    caminho = BACKUP_DIR / nome
    with zipfile.ZipFile(caminho, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for arq in ARQUIVOS_BACKUP:
            if os.path.exists(arq):
                z.write(arq)
    return caminho

def restaurar_backup_zip_bytes(zip_bytes: bytes):
    tmp_dir = BACKUP_DIR / "_tmp_restore"
    if tmp_dir.exists():
        shutil.rmtree(tmp_dir)
    tmp_dir.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(BytesIO(zip_bytes), "r") as z:
        z.extractall(tmp_dir)

    for arq in ARQUIVOS_BACKUP:
        src = tmp_dir / arq
        if src.exists():
            shutil.copy(src, arq)

    shutil.rmtree(tmp_dir)

def backup_automatico_diario_3h():
    """
    Streamlit não roda 24/7.
    Regra: após 03:00, se ainda não fez backup HOJE => faz 1.
    """
    agora = agora_br()
    hoje = agora.strftime("%Y-%m-%d")
    if agora.hour >= 3:
        ultima = ""
        if BACKUP_STATE_FILE.exists():
            ultima = BACKUP_STATE_FILE.read_text(encoding="utf-8").strip()
        if ultima != hoje:
            criar_backup_zip_em_disco()
            BACKUP_STATE_FILE.write_text(hoje, encoding="utf-8")

def listar_backups() -> pd.DataFrame:
    backups = []
    for p in sorted(BACKUP_DIR.glob("backup_saude_kids_*.zip")):
        stat = p.stat()
        dt = datetime.fromtimestamp(stat.st_mtime, tz=fuso_br)
        backups.append({
            "arquivo": p.name,
            "caminho": str(p),
            "data_hora": dt,
            "tamanho_mb": round(stat.st_size / (1024 * 1024), 2),
        })
    if not backups:
        return pd.DataFrame(columns=["arquivo", "caminho", "data_hora", "tamanho_mb"])
    df = pd.DataFrame(backups).sort_values("data_hora", ascending=False).reset_index(drop=True)
    return df

def apagar_backups_antigos(dias_manter=7) -> int:
    limite = agora_br() - timedelta(days=dias_manter)
    apagados = 0
    for p in BACKUP_DIR.glob("backup_saude_kids_*.zip"):
        dt = datetime.fromtimestamp(p.stat().st_mtime, tz=fuso_br)
        if dt < limite:
            try:
                p.unlink()
                apagados += 1
            except Exception:
                pass
    return apagados

backup_automatico_diario_3h()

# ================= DESIGN =================
st.markdown(r"""
<style>
    .stApp { background-color: #0e1117; color: #ffffff; }
    .card { background-color: #1a1c24; padding: 22px; border-radius: 18px; border: 1px solid #30363d; margin-bottom: 18px; }
    .metric-box { background: #262730; border: 1px solid #4a4a4a; padding: 14px; border-radius: 12px; text-align: center; }
    .dose-destaque { font-size: 34px; font-weight: 700; color: #4ade80; }
    label, p, span, h1, h2, h3, .stMarkdown { color: white !important; }
    .stTextInput>div>div>input, .stNumberInput>div>div>input {
        background-color: #262730 !important; color: white !important; border: 1px solid #4a4a4a !important;
    }
    .stTabs [data-baseweb="tab-list"] { background-color: #0e1117; }
    .stTabs [data-baseweb="tab"] { color: white; }
</style>
""", unsafe_allow_html=True)

# ================= EMAIL (RESET) =================
def gerar_senha_temporaria(tamanho=6):
    caracteres = string.ascii_letters + string.digits
    return "".join(random.choice(caracteres) for _ in range(tamanho))

def enviar_senha_nova(email_destino, senha_nova):
    """
    Recomendado (seguro):
      - No Streamlit Cloud: st.secrets["GMAIL_APP_PASSWORD"] = "SUA_SENHA_DE_APP"
    """
    import smtplib

    meu_email = "ewerlon.osbadboys@gmail.com"
    minha_senha = (st.secrets.get("GMAIL_APP_PASSWORD", "") or "").strip()
    if not minha_senha:
        return False

    corpo = f"<h3>Saúde Kids</h3><p>Sua nova senha de acesso é: <b>{senha_nova}</b></p>"
    msg = MIMEText(corpo, "html")
    msg["Subject"] = "Sua Nova Senha - Saúde Kids"
    msg["From"] = meu_email
    msg["To"] = email_destino

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(meu_email, minha_senha)
            smtp.send_message(msg)
        return True
    except Exception:
        return False

# ================= DB USERS =================
def init_db():
    conn = sqlite3.connect("usuarios.db")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            nome TEXT,
            email TEXT PRIMARY KEY,
            senha TEXT
        )
    """)
    if not conn.execute("SELECT 1 FROM users WHERE email='admin'").fetchone():
        conn.execute("INSERT INTO users VALUES ('Administrador', 'admin', '542820')")
    conn.commit()
    conn.close()

init_db()

if "logado" not in st.session_state:
    st.session_state.logado = False
if "user_email" not in st.session_state:
    st.session_state.user_email = ""

# ====== COOKIE (OPCIONAL) ======
def cookie_get_email():
    if not HAS_COOKIES or _cookie_manager is None:
        return ""
    try:
        v = _cookie_manager.get(COOKIE_KEY)
        return norm_email(v)
    except Exception:
        return ""

def cookie_set_email(email: str):
    if not HAS_COOKIES or _cookie_manager is None:
        return
    try:
        _cookie_manager.set(COOKIE_KEY, norm_email(email), expires_at=timedelta(days=COOKIE_DIAS))
    except Exception:
        pass

def cookie_clear():
    if not HAS_COOKIES or _cookie_manager is None:
        return
    try:
        _cookie_manager.delete(COOKIE_KEY)
    except Exception:
        pass

if not st.session_state.logado:
    ck = cookie_get_email()
    if ck:
        st.session_state.logado = True
        st.session_state.user_email = ck

# ================= ARQUIVOS (DADOS) =================
def carregar_dados_seguro(arq: str) -> pd.DataFrame:
    if not os.path.exists(arq):
        return pd.DataFrame()
    df = pd.read_csv(arq)
    if "Usuario" not in df.columns:
        df["Usuario"] = ""
    return df[df["Usuario"] == st.session_state.user_email].copy()

# ================= IDs + CRUD GLICEMIA =================
def _ensure_id_column(df: pd.DataFrame, col_name="ID", prefix="GL") -> pd.DataFrame:
    if df is None or df.empty:
        return df
    if col_name not in df.columns:
        df[col_name] = [f"{prefix}-{uuid.uuid4().hex[:12]}" for _ in range(len(df))]
    else:
        mask = df[col_name].isna() | (df[col_name].astype(str).str.strip() == "")
        if mask.any():
            df.loc[mask, col_name] = [f"{prefix}-{uuid.uuid4().hex[:12]}" for _ in range(int(mask.sum()))]
    return df

def carregar_glicemia_com_id() -> pd.DataFrame:
    df_all = pd.read_csv(ARQ_G) if os.path.exists(ARQ_G) else pd.DataFrame()
    if df_all.empty:
        return pd.DataFrame(columns=["ID", "Usuario", "Data", "Hora", "Valor", "Momento", "Dose"])
    if "Usuario" not in df_all.columns:
        df_all["Usuario"] = ""
    df_all = _ensure_id_column(df_all, "ID", "GL")

    # tenta persistir IDs (migração tranquila)
    try:
        df_all.to_csv(ARQ_G, index=False)
    except Exception:
        pass

    return df_all[df_all["Usuario"] == st.session_state.user_email].copy()

def salvar_registro_glicemia(valor: int, momento: str, dose: str, dt: datetime):
    novo = pd.DataFrame([{
        "ID": f"GL-{uuid.uuid4().hex[:12]}",
        "Usuario": st.session_state.user_email,
        "Data": dt.strftime("%d/%m/%Y"),
        "Hora": dt.strftime("%H:%M"),
        "Valor": int(valor),
        "Momento": (momento or "").strip(),
        "Dose": (dose or "").strip(),
    }])
    base = pd.read_csv(ARQ_G) if os.path.exists(ARQ_G) else pd.DataFrame(columns=novo.columns)
    if "Usuario" not in base.columns:
        base["Usuario"] = ""
    base = _ensure_id_column(base, "ID", "GL")
    pd.concat([base, novo], ignore_index=True).to_csv(ARQ_G, index=False)

def aplicar_edicoes_e_exclusoes_glicemia(df_editado: pd.DataFrame):
    if df_editado is None or df_editado.empty:
        return

    base = pd.read_csv(ARQ_G) if os.path.exists(ARQ_G) else pd.DataFrame()
    if base.empty:
        return
    if "Usuario" not in base.columns:
        base["Usuario"] = ""
    base = _ensure_id_column(base, "ID", "GL")

    df_editado = df_editado.copy()
    if "Excluir" not in df_editado.columns:
        df_editado["Excluir"] = False

    # normalização
    df_editado["ID"] = df_editado["ID"].astype(str)
    df_editado["Data"] = df_editado["Data"].astype(str).str.strip()
    df_editado["Hora"] = df_editado["Hora"].astype(str).str.strip()
    df_editado["Momento"] = df_editado["Momento"].astype(str).str.strip()
    df_editado["Dose"] = df_editado.get("Dose", "").astype(str).str.strip()
    df_editado["Valor"] = pd.to_numeric(df_editado["Valor"], errors="coerce").fillna(0).astype(int)

    ids_user = set(base.loc[base["Usuario"] == st.session_state.user_email, "ID"].astype(str).tolist())
    ids_excluir = set(df_editado.loc[df_editado["Excluir"] == True, "ID"].tolist()).intersection(ids_user)

    # atualiza
    df_upd = df_editado.loc[df_editado["Excluir"] != True].copy()
    for _, r in df_upd.iterrows():
        rid = str(r["ID"])
        if rid not in ids_user:
            continue
        mask = (base["Usuario"] == st.session_state.user_email) & (base["ID"].astype(str) == rid)
        if mask.any():
            base.loc[mask, "Data"] = r["Data"]
            base.loc[mask, "Hora"] = r["Hora"]
            base.loc[mask, "Valor"] = int(r["Valor"])
            base.loc[mask, "Momento"] = r["Momento"]
            base.loc[mask, "Dose"] = r["Dose"]

    # exclui
    if ids_excluir:
        base = base[~((base["Usuario"] == st.session_state.user_email) & (base["ID"].astype(str).isin(ids_excluir)))].copy()

    base.to_csv(ARQ_G, index=False)

# ================= RECEITA (RÁPIDA/LONGA) =================
def _schema_receita_nova(rec: pd.Series, periodo: str) -> bool:
    need = [
        f"{periodo}_f1_min", f"{periodo}_f1_max", f"{periodo}_f1_dose",
        f"{periodo}_f2_min", f"{periodo}_f2_max", f"{periodo}_f2_dose",
        f"{periodo}_f3_min", f"{periodo}_f3_max", f"{periodo}_f3_dose",
    ]
    return all(k in rec.index for k in need)

def calc_insulina(v: int, momento: str):
    df_r = carregar_dados_seguro(ARQ_R)
    if df_r.empty:
        return "0 UI", "Configurar Receita"

    rec = df_r.iloc[0]
    periodo = "manha" if momento in ["Antes Café", "Após Café", "Antes Almoço", "Após Almoço", ] else "noite"

    try:
        if not _schema_receita_nova(rec, periodo):
            return "0 UI", "Receita inválida"

        f1_min = float(rec[f"{periodo}_f1_min"]); f1_max = float(rec[f"{periodo}_f1_max"]); f1_dose = float(rec[f"{periodo}_f1_dose"])
        f2_min = float(rec[f"{periodo}_f2_min"]); f2_max = float(rec[f"{periodo}_f2_max"]); f2_dose = float(rec[f"{periodo}_f2_dose"])
        f3_min = float(rec[f"{periodo}_f3_min"]); f3_max = float(rec[f"{periodo}_f3_max"]); f3_dose = float(rec[f"{periodo}_f3_dose"])

        if v < 70:
            return "0 UI", "Hipoglicemia!"

        if f1_min <= v <= f1_max:
            return f"{int(f1_dose)} UI", f"Faixa 1 ({int(f1_min)}-{int(f1_max)})"
        if f2_min <= v <= f2_max:
            return f"{int(f2_dose)} UI", f"Faixa 2 ({int(f2_min)}-{int(f2_max)})"
        if f3_min <= v <= f3_max:
            return f"{int(f3_dose)} UI", f"Faixa 3 ({int(f3_min)}-{int(f3_max)})"
        return "0 UI", "Fora das faixas"
    except Exception:
        return "0 UI", "Erro na Receita"

def calc_insulina_rapida(v: int, momento: str):
    return calc_insulina(v, momento)

def calc_glargina(momento: str):
    df_r = carregar_dados_seguro(ARQ_R)
    if df_r.empty:
        return "0 UI", "Longa: Configurar"

    rec = df_r.iloc[0]
    try:
        cafe = float(rec.get("glargina_cafe_ui", 0) or 0)
        janta = float(rec.get("glargina_janta_ui", 0) or 0)
        if momento == "Antes Café":
            return f"{int(cafe)} UI", "Longa (Antes Café)"
        if momento == "Antes Janta":
            return f"{int(janta)} UI", "Longa (Antes Janta)"
        return "—", "Longa: não aplicável"
    except Exception:
        return "0 UI", "Longa: erro"

# ================= PRÓXIMA MEDIDA (+2h) e WHATSAPP =================
def proxima_medida_apos(momento: str, dt_base: datetime):
    mapa = {
        "Antes Café": "Após Café",
        "Antes Almoço": "Após Almoço",
        "Antes Janta": "Após Janta",
    }
    if momento not in mapa:
        return "", ""
    dt_apos = dt_base + timedelta(hours=2)
    return mapa[momento], dt_apos.strftime("%H:%M")

def link_whatsapp_lembrete(momento: str, valor_glicemia: int, dose_rapida: str, dose_longa: str) -> str:
    dt_agora = agora_br()
    momento_apos, hora_apos = proxima_medida_apos(momento, dt_agora)

    linhas = [
        "🧪 Saúde Kids",
        "",
        f"✅ Medida AGORA: {momento}",
        f"📍 Glicemia: {int(valor_glicemia)}",
    ]

    if dose_rapida and dose_rapida != "—":
        linhas.append(f"⚡ Rápida: {dose_rapida}")
    if dose_longa and dose_longa != "—":
        linhas.append(f"🩸 Longa: {dose_longa}")

    if momento_apos and hora_apos:
        linhas.extend(["", f"⏰ Próxima medida: {momento_apos} às {hora_apos} (2h após)"])

    mensagem = "\n".join(linhas)
    return "https://wa.me/?text=" + quote(mensagem)

# ================= MOMENTOS / ALIMENTOS =================
MOMENTOS_ORDEM = [
    "Antes Café", "Após Café",
    "Antes Almoço", "Após Almoço",
        "Antes Janta", "Após Janta",
    "Madrugada",
]

ALIMENTOS = {
    "Pão Francês (1un)": [28, 4, 1],
    "Pão de Forma (2 fatias)": [24, 4, 2],
    "Pão Integral (2 fatias)": [22, 5, 2],
    "Tapioca (50g)": [27, 0, 0],
    "Arroz Branco (servir)": [10, 2, 0],
    "Arroz Integral (servir)": [8, 2, 1],
    "Feijão (concha)": [14, 5, 1],
    "Carne Boi (100g)": [0, 26, 12],
    "Frango (100g)": [0, 31, 4],
    "Peixe (100g)": [0, 20, 5],
    "Ovo Cozido (1un)": [1, 6, 5],
    "Macarrão (pegador)": [30, 5, 1],
    "Batata Doce (100g)": [20, 2, 0],
    "Banana (1un)": [22, 1, 0],
    "Maçã (1un)": [15, 0, 0],
}

# =========================================================
# LOGIN UI
# =========================================================
if not st.session_state.logado:
    st.title("🧪 Saúde Kids - Acesso")

    if not HAS_COOKIES:
        st.caption("ℹ️ Login persistente desativado (extra-streamlit-components não instalado).")

    abas_login = st.tabs(["🔐 Entrar", "📝 Criar Conta", "❓ Esqueci Senha", "🔄 Alterar Senha"])

    # -------- ENTRAR --------
    with abas_login[0]:
        u = norm_email(st.text_input("E-mail", key="l_email"))
        s = norm_senha(st.text_input("Senha", type="password", key="l_pass"))

        if st.button("Acessar Aplicativo", use_container_width=True):
            conn = sqlite3.connect("usuarios.db")
            ok = conn.execute("SELECT * FROM users WHERE email=? AND senha=?", (u, s)).fetchone()
            conn.close()
            if ok:
                st.session_state.logado = True
                st.session_state.user_email = u
                cookie_set_email(u)
                st.rerun()
            else:
                st.error("E-mail ou senha incorretos.")

    # -------- CRIAR CONTA --------
    with abas_login[1]:
        n_cad = (st.text_input("Nome Completo") or "").strip()
        e_cad = norm_email(st.text_input("E-mail para Cadastro"))
        s_cad = norm_senha(st.text_input("Senha para Cadastro", type="password"))

        if st.button("Realizar Cadastro", use_container_width=True):
            if not n_cad or not e_cad or not s_cad:
                st.warning("Preencha nome, e-mail e senha.")
            else:
                try:
                    conn = sqlite3.connect("usuarios.db")
                    conn.execute("INSERT INTO users VALUES (?,?,?)", (n_cad, e_cad, s_cad))
                    conn.commit()
                    conn.close()
                    st.success("Conta criada com sucesso!")
                except Exception:
                    st.error("Este e-mail já está cadastrado.")

    # -------- ESQUECI SENHA --------
    with abas_login[2]:
        email_alvo = norm_email(st.text_input("Digite seu e-mail cadastrado"))
        if st.button("Recuperar Acesso", use_container_width=True):
            conn = sqlite3.connect("usuarios.db")
            c = conn.cursor()
            user = c.execute("SELECT email FROM users WHERE email=?", (email_alvo,)).fetchone()

            if user:
                nova = gerar_senha_temporaria()
                c.execute("UPDATE users SET senha=? WHERE email=?", (nova, email_alvo))
                conn.commit()
                conn.close()

                if enviar_senha_nova(email_alvo, nova):
                    st.success("Nova senha enviada para seu e-mail!")
                else:
                    st.warning("E-mail não configurado no app (sem GMAIL_APP_PASSWORD).")
                    st.info("Use a senha temporária abaixo para entrar e depois altere sua senha:")
                    st.code(nova)
            else:
                conn.close()
                st.error("E-mail não encontrado.")

    # -------- ALTERAR SENHA --------
    with abas_login[3]:
        alt_em = norm_email(st.text_input("Confirme seu E-mail", key="alt_em"))
        alt_at = norm_senha(st.text_input("Senha Atual", type="password", key="alt_at"))
        alt_n1 = norm_senha(st.text_input("Nova Senha", type="password", key="alt_n1"))

        if st.button("Confirmar Alteração", use_container_width=True):
            conn = sqlite3.connect("usuarios.db")
            ok = conn.execute("SELECT * FROM users WHERE email=? AND senha=?", (alt_em, alt_at)).fetchone()
            if ok:
                conn.execute("UPDATE users SET senha=? WHERE email=?", (alt_n1, alt_em))
                conn.commit()
                conn.close()
                st.success("Senha alterada com sucesso!")
            else:
                conn.close()
                st.error("Dados atuais incorretos.")

    st.stop()

# =========================================================
# APP PRINCIPAL (ADMIN / USER)
# =========================================================
if st.session_state.user_email == "admin":
    st.title("🛡️ Painel Admin - Gestão Estratégica")
    t_usuarios, t_metricas, t_sugestoes, t_backup = st.tabs(
        ["👥 Pessoas Cadastradas", "📈 Crescimento e App", "📩 Sugestões", "💾 Backup & Restauração"]
    )

    conn = sqlite3.connect("usuarios.db")
    df_users = pd.read_sql_query("SELECT nome, email FROM users", conn)
    conn.close()

    with t_usuarios:
        st.subheader("Lista de Usuários")
        st.dataframe(df_users, use_container_width=True)
        st.metric("Total de Cadastros", len(df_users))
        st.markdown("---")
        st.subheader("🔑 Alterar Senha de Usuário (Admin)")
        user_selecionado = st.selectbox("Selecione o E-mail do Usuário", df_users["email"].tolist())
        nova_senha_admin = norm_senha(st.text_input("Nova senha para este usuário", type="password"))
        if st.button("Confirmar Alteração de Senha", use_container_width=True):
            if nova_senha_admin:
                conn = sqlite3.connect("usuarios.db")
                conn.execute("UPDATE users SET senha=? WHERE email=?", (nova_senha_admin, user_selecionado))
                conn.commit()
                conn.close()
                st.success(f"Senha de {user_selecionado} alterada com sucesso!")
            else:
                st.warning("Digite uma senha antes de confirmar.")

    with t_metricas:
        c1, c2 = st.columns(2)
        with c1:
            st.write("### Distribuição de Registros (Glicemia)")
            if os.path.exists(ARQ_G):
                df_uso = pd.read_csv(ARQ_G)
                if "Usuario" in df_uso.columns and not df_uso.empty:
                    uso_por_user = df_uso["Usuario"].value_counts().reset_index()
                    uso_por_user.columns = ["Usuario", "Registros"]
                    fig_pizza = px.pie(uso_por_user, values="Registros", names="Usuario", hole=.3)
                    st.plotly_chart(fig_pizza, use_container_width=True)
                else:
                    st.info("Sem dados.")
            else:
                st.info("Sem dados.")
        with c2:
            st.write("### Cadastros (snapshot)")
            dados_c = pd.DataFrame({
                "Mês": ["Jan", "Fev", "Mar"],
                "Usuários": [max(1, len(df_users)//2), max(1, int(len(df_users)/1.1)), len(df_users)]
            })
            st.plotly_chart(px.line(dados_c, x="Mês", y="Usuários", markers=True), use_container_width=True)

    with t_sugestoes:
        st.subheader("📩 Sugestões recebidas")
        if os.path.exists(ARQ_M):
            st.dataframe(pd.read_csv(ARQ_M), use_container_width=True)
        else:
            st.info("Sem sugestões.")

    with t_backup:
        st.subheader("💾 Backup Manual / Automático / Restauração")

        st.write("### 📦 Gerar Backup Manual")
        if st.button("📦 Gerar Backup Agora", use_container_width=True):
            b, nome = criar_backup_zip_em_bytes()
            st.download_button("⬇️ Baixar Backup (.zip)", b, file_name=nome, use_container_width=True)

        st.markdown("---")
        st.write("### ♻️ Restauração Manual")
        up = st.file_uploader("Enviar arquivo .zip de backup", type=["zip"])
        if up is not None:
            if st.button("✅ Restaurar Agora", use_container_width=True):
                restaurar_backup_zip_bytes(up.getvalue())
                st.success("Restauração concluída! Recarregue o app (F5).")

        st.markdown("---")
        st.write("### ⏰ Backup Automático")
        st.info("Config: 1 backup por dia após **03:00** (Brasília).")
        if BACKUP_STATE_FILE.exists():
            st.caption(f"Último dia registrado: {BACKUP_STATE_FILE.read_text(encoding='utf-8').strip()}")

        st.markdown("---")
        st.write("### 🗂️ Backups existentes (pasta backups/)")
        df_bk = listar_backups()
        if df_bk.empty:
            st.warning("Nenhum backup encontrado.")
        else:
            df_show = df_bk.copy()
            df_show["data_hora"] = df_show["data_hora"].dt.strftime("%d/%m/%Y %H:%M:%S")
            st.dataframe(df_show[["arquivo", "data_hora", "tamanho_mb"]], use_container_width=True)

            st.markdown("#### Ações")
            selecionado = st.selectbox("Selecionar backup", df_bk["arquivo"].tolist())
            p_sel = BACKUP_DIR / selecionado

            colx1, colx2, colx3 = st.columns(3)
            with colx1:
                if p_sel.exists():
                    with open(p_sel, "rb") as f:
                        st.download_button("⬇️ Baixar Selecionado", f.read(), file_name=selecionado, use_container_width=True)
            with colx2:
                if st.button("🗑️ Apagar Selecionado", use_container_width=True):
                    if p_sel.exists():
                        p_sel.unlink()
                        st.success("Backup apagado.")
                        st.rerun()
            with colx3:
                if st.button("🧹 Apagar Antigos (7 dias)", use_container_width=True):
                    apagados = apagar_backups_antigos(dias_manter=7)
                    st.success(f"Apagados: {apagados}")
                    st.rerun()

else:
    tab1, tab2, tab3, tab4 = st.tabs(["📊 Glicemia", "🍽️ Nutrição", "⚙️ Receita", "📩 Sugerir Melhoria"])

    # ====== GLICEMIA ======
    with tab1:
        st.markdown('<div class="card">', unsafe_allow_html=True)

        dfg = carregar_glicemia_com_id()

        c1, c2 = st.columns([1, 2])
        with c1:
            v_gl = st.number_input("Valor Glicemia", 0, 600, 100, step=1)

            # ✅ Momentos extras (personalizado)
            MOMENTOS_BASE = MOMENTOS_ORDEM + ["Outro (personalizado)"]
            m_sel = st.selectbox("Momento", MOMENTOS_BASE)

            momento_extra = ""
            if m_sel == "Outro (personalizado)":
                momento_extra = (st.text_input("Digite o momento (ex: 'Antes Academia', 'Após Treino', 'Após Remédio')") or "").strip()
            m_gl = momento_extra if m_sel == "Outro (personalizado)" else m_sel

            # Regras de exibição (rápida/longa)
            MOMENTOS_RAPIDA = ["Antes Café", "Antes Almoço", "Antes Janta"]
            MOMENTOS_LONGA = ["Antes Café", "Antes Janta"]

            # Próxima medida (+2h) apenas para os "Antes ..." padrão
            dt_agora = agora_br()
            momento_apos, hora_apos = proxima_medida_apos(m_gl, dt_agora)
            if momento_apos and hora_apos:
                st.info(f"⏰ Próxima medida: **{momento_apos}** às **{hora_apos}** (2 horas após)")

            # RÁPIDA
            if m_gl in MOMENTOS_RAPIDA:
                dose_r, msg_r = calc_insulina_rapida(int(v_gl), m_gl)
                st.markdown(
                    f'<div class="metric-box"><small>Rápida: {msg_r}</small><br><span class="dose-destaque">{dose_r}</span></div>',
                    unsafe_allow_html=True
                )
            else:
                dose_r, msg_r = "—", "Rápida não aplicável"

            # LONGA
            if m_gl in MOMENTOS_LONGA:
                dose_l, msg_l = calc_glargina(m_gl)
                st.markdown(
                    f'<div class="metric-box" style="margin-top:10px;"><small>{msg_l}</small><br><span class="dose-destaque">{dose_l}</span></div>',
                    unsafe_allow_html=True
                )
            else:
                dose_l, msg_l = "—", "Longa não aplicável"

            # WhatsApp
            link_wpp = link_whatsapp_lembrete(
                momento=m_gl,
                valor_glicemia=int(v_gl),
                dose_rapida=dose_r,
                dose_longa=dose_l,
            )
            st.link_button("📲 Abrir WhatsApp com mensagem pronta", link_wpp, use_container_width=True)

            if st.button("💾 Salvar Glicemia", use_container_width=True):
                dose_para_salvar = dose_r if m_gl in MOMENTOS_RAPIDA else ""
                salvar_registro_glicemia(int(v_gl), m_gl, dose_para_salvar, agora_br())
                st.rerun()

        with c2:
            if not dfg.empty:
                fig = px.line(dfg.tail(20), x="Hora", y="Valor", markers=True, title="Tendência (últimas)")
                fig.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", font_color="white")
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("Sem dados para gráfico ainda.")

        # ✅ HISTÓRICO EDITÁVEL (editar/excluir)
        st.markdown("### 🧾 Histórico (editar / excluir)")

        if not dfg.empty:
            df_hist = dfg.copy()

            if "Excluir" not in df_hist.columns:
                df_hist["Excluir"] = False

            cols_order = ["Excluir", "ID", "Data", "Hora", "Valor", "Momento", "Dose"]
            for c in cols_order:
                if c not in df_hist.columns:
                    df_hist[c] = ""

            df_hist = df_hist[cols_order].tail(80).reset_index(drop=True)

            df_edit = st.data_editor(
                df_hist,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Excluir": st.column_config.CheckboxColumn("Excluir"),
                    "ID": st.column_config.TextColumn("ID", disabled=True),
                    "Valor": st.column_config.NumberColumn("Valor", min_value=0, max_value=600, step=1),
                },
                key="glicemia_editor",
            )

            col_a, col_b = st.columns(2)
            with col_a:
                if st.button("✅ Salvar alterações do histórico", use_container_width=True):
                    aplicar_edicoes_e_exclusoes_glicemia(df_edit)
                    st.success("Histórico atualizado!")
                    st.rerun()
            with col_b:
                st.caption("Marque 'Excluir' e clique em salvar para remover.")
        else:
            st.info("Sem registros ainda.")

        st.markdown("</div>", unsafe_allow_html=True)

    # ====== NUTRIÇÃO ======
    with tab2:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        dfn = carregar_dados_seguro(ARQ_N)

        m_nutri = st.selectbox("Refeição", MOMENTOS_ORDEM, key="n_m")
        sel = st.multiselect("Alimentos", options=list(ALIMENTOS.keys()))

        c_tot = sum(ALIMENTOS[x][0] for x in sel)
        p_tot = sum(ALIMENTOS[x][1] for x in sel)
        g_tot = sum(ALIMENTOS[x][2] for x in sel)

        col1, col2, col3 = st.columns(3)
        col1.metric("Carbos", f"{c_tot}g")
        col2.metric("Proteínas", f"{p_tot}g")
        col3.metric("Gorduras", f"{g_tot}g")

        if st.button("💾 Salvar Refeição", use_container_width=True):
            dt = agora_br()
            novo_n = pd.DataFrame([[st.session_state.user_email, dt.strftime("%d/%m/%Y"), m_nutri, ", ".join(sel), c_tot, p_tot, g_tot]],
                                 columns=["Usuario", "Data", "Momento", "Info", "C", "P", "G"])
            base = pd.read_csv(ARQ_N) if os.path.exists(ARQ_N) else pd.DataFrame(columns=novo_n.columns)
            if "Usuario" not in base.columns:
                base["Usuario"] = ""
            pd.concat([base, novo_n], ignore_index=True).to_csv(ARQ_N, index=False)
            st.rerun()

        st.dataframe(dfn.tail(12), use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    # ====== RECEITA ======
    with tab3:
        st.markdown('<div class="card">', unsafe_allow_html=True)

        df_r_all = pd.read_csv(ARQ_R) if os.path.exists(ARQ_R) else pd.DataFrame()
        if not df_r_all.empty and "Usuario" in df_r_all.columns:
            r_u = df_r_all[df_r_all["Usuario"] == st.session_state.user_email]
        else:
            r_u = pd.DataFrame()
        v = r_u.iloc[0] if not r_u.empty else {}

        st.subheader("⚡ Receita Rápida (por faixas)")

        st.markdown("**🌞 Manhã**")
        cm1, cm2, cm3 = st.columns(3)
        with cm1:
            m1_min = st.number_input("Faixa 1 - Mín", value=int(v.get("manha_f1_min", 70)), key="m1_min_u")
            m1_max = st.number_input("Faixa 1 - Máx", value=int(v.get("manha_f1_max", 150)), key="m1_max_u")
            m1_dose = st.number_input("Dose Faixa 1 (UI)", value=int(v.get("manha_f1_dose", 3)), key="m1_dose_u")
        with cm2:
            m2_min = st.number_input("Faixa 2 - Mín", value=int(v.get("manha_f2_min", 151)), key="m2_min_u")
            m2_max = st.number_input("Faixa 2 - Máx", value=int(v.get("manha_f2_max", 300)), key="m2_max_u")
            m2_dose = st.number_input("Dose Faixa 2 (UI)", value=int(v.get("manha_f2_dose", 5)), key="m2_dose_u")
        with cm3:
            m3_min = st.number_input("Faixa 3 - Mín", value=int(v.get("manha_f3_min", 301)), key="m3_min_u")
            m3_max = st.number_input("Faixa 3 - Máx", value=int(v.get("manha_f3_max", 600)), key="m3_max_u")
            m3_dose = st.number_input("Dose Faixa 3 (UI)", value=int(v.get("manha_f3_dose", 8)), key="m3_dose_u")

        st.markdown("---")
        st.markdown("**🌙 Noite**")
        cn1, cn2, cn3 = st.columns(3)
        with cn1:
            n1_min = st.number_input("Faixa 1 - Mín", value=int(v.get("noite_f1_min", 70)), key="n1_min_u")
            n1_max = st.number_input("Faixa 1 - Máx", value=int(v.get("noite_f1_max", 150)), key="n1_max_u")
            n1_dose = st.number_input("Dose Faixa 1 (UI)", value=int(v.get("noite_f1_dose", 3)), key="n1_dose_u")
        with cn2:
            n2_min = st.number_input("Faixa 2 - Mín", value=int(v.get("noite_f2_min", 151)), key="n2_min_u")
            n2_max = st.number_input("Faixa 2 - Máx", value=int(v.get("noite_f2_max", 300)), key="n2_max_u")
            n2_dose = st.number_input("Dose Faixa 2 (UI)", value=int(v.get("noite_f2_dose", 5)), key="n2_dose_u")
        with cn3:
            n3_min = st.number_input("Faixa 3 - Mín", value=int(v.get("noite_f3_min", 301)), key="n3_min_u")
            n3_max = st.number_input("Faixa 3 - Máx", value=int(v.get("noite_f3_max", 600)), key="n3_max_u")
            n3_dose = st.number_input("Dose Faixa 3 (UI)", value=int(v.get("noite_f3_dose", 8)), key="n3_dose_u")

        st.markdown("---")
        st.subheader("🩸 Longa (dose fixa)")
        gl1, gl2 = st.columns(2)
        with gl1:
            glargina_cafe_ui = st.number_input("Longa - Antes Café (UI)", value=int(float(v.get("glargina_cafe_ui", 0) or 0)), key="gl_cafe")
        with gl2:
            glargina_janta_ui = st.number_input("Longa - Antes Janta (UI)", value=int(float(v.get("glargina_janta_ui", 0) or 0)), key="gl_janta")

        if st.button("💾 Salvar Receita", use_container_width=True):
            nova_rec = pd.DataFrame([{
                "Usuario": st.session_state.user_email,

                "manha_f1_min": m1_min, "manha_f1_max": m1_max, "manha_f1_dose": m1_dose,
                "manha_f2_min": m2_min, "manha_f2_max": m2_max, "manha_f2_dose": m2_dose,
                "manha_f3_min": m3_min, "manha_f3_max": m3_max, "manha_f3_dose": m3_dose,

                "noite_f1_min": n1_min, "noite_f1_max": n1_max, "noite_f1_dose": n1_dose,
                "noite_f2_min": n2_min, "noite_f2_max": n2_max, "noite_f2_dose": n2_dose,
                "noite_f3_min": n3_min, "noite_f3_max": n3_max, "noite_f3_dose": n3_dose,

                "glargina_cafe_ui": glargina_cafe_ui,
                "glargina_janta_ui": glargina_janta_ui,
            }])

            if not df_r_all.empty and "Usuario" in df_r_all.columns:
                df_r_all2 = df_r_all[df_r_all["Usuario"] != st.session_state.user_email].copy()
            else:
                df_r_all2 = pd.DataFrame()

            pd.concat([df_r_all2, nova_rec], ignore_index=True).to_csv(ARQ_R, index=False)
            st.success("Receita salva com sucesso!")

        st.markdown("</div>", unsafe_allow_html=True)

    # ====== SUGESTÃO ======
    with tab4:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        txt = st.text_area("Sugestão de Melhoria:")
        if st.button("Enviar Sugestão", use_container_width=True):
            if txt.strip():
                dt = agora_br().strftime("%d/%m/%Y %H:%M")
                novo_m = pd.DataFrame([[st.session_state.user_email, dt, txt.strip()]], columns=["Usuario", "Data", "Sugestão"])
                base_m = pd.read_csv(ARQ_M) if os.path.exists(ARQ_M) else pd.DataFrame(columns=novo_m.columns)
                if "Usuario" not in base_m.columns:
                    base_m["Usuario"] = ""
                pd.concat([base_m, novo_m], ignore_index=True).to_csv(ARQ_M, index=False)
                st.success("Enviado com sucesso!")
            else:
                st.warning("Digite uma sugestão antes de enviar.")
        st.markdown("</div>", unsafe_allow_html=True)


def gerar_pdf_bytes(df_g: pd.DataFrame, df_n: pd.DataFrame) -> bytes:
    """
    Gera um PDF simples com resumo + últimas medições.
    """
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = []

    titulo = f"Relatório Saúde Kids - {st.session_state.user_email}"
    story.append(Paragraph(titulo, styles["Title"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"Gerado em: {agora_br().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 12))

    # Resumo Glicemia
    story.append(Paragraph("Glicemia - Resumo", styles["Heading2"]))
    if df_g is None or df_g.empty:
        story.append(Paragraph("Sem registros de glicemia.", styles["Normal"]))
    else:
        try:
            vals = pd.to_numeric(df_g["Valor"], errors="coerce").dropna()
            resumo = [
                ["Registros", str(len(df_g))],
                ["Mínimo", str(int(vals.min())) if not vals.empty else "-"],
                ["Máximo", str(int(vals.max())) if not vals.empty else "-"],
                ["Média", f"{vals.mean():.1f}" if not vals.empty else "-"],
            ]
            t = Table(resumo, colWidths=[6*cm, 9*cm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
                ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
                ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ]))
            story.append(t)
        except Exception:
            story.append(Paragraph("Não foi possível gerar resumo de glicemia.", styles["Normal"]))

        story.append(Spacer(1, 10))
        story.append(Paragraph("Últimos registros de Glicemia", styles["Heading3"]))
        ult = df_g.copy().tail(20)
        cols = [c for c in ["Data", "Hora", "Momento", "Valor", "Dose"] if c in ult.columns]
        data_tbl = [cols] + ult[cols].astype(str).values.tolist()
        t2 = Table(data_tbl, repeatRows=1, colWidths=[3*cm, 2*cm, 5.5*cm, 2*cm, 2.5*cm])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
        ]))
        story.append(t2)

    story.append(Spacer(1, 14))

    # Resumo Nutrição
    story.append(Paragraph("Nutrição - Resumo", styles["Heading2"]))
    if df_n is None or df_n.empty:
        story.append(Paragraph("Sem registros de nutrição.", styles["Normal"]))
    else:
        try:
            story.append(Paragraph("Últimos registros de Nutrição", styles["Heading3"]))
            ult_n = df_n.copy().tail(20)
            cols_n = [c for c in ["Data", "Momento", "Info", "C", "P", "G"] if c in ult_n.columns]
            data_tbl_n = [cols_n] + ult_n[cols_n].astype(str).values.tolist()
            # larguras aproximadas
            col_widths = [3*cm, 3*cm, 6*cm, 1.5*cm, 1.5*cm, 1.5*cm][:len(cols_n)]
            t3 = Table(data_tbl_n, repeatRows=1, colWidths=col_widths)
            t3.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
                ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
                ("FONTSIZE", (0,0), (-1,-1), 8),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
            ]))
            story.append(t3)
        except Exception:
            story.append(Paragraph("Não foi possível montar tabela de nutrição.", styles["Normal"]))

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    return pdf


# ================= EXCEL (SIDEBAR) =================
st.sidebar.markdown("---")
if st.sidebar.button("📥 Gerar Excel Completo"):
    df_e_g = carregar_glicemia_com_id()
    df_e_n = carregar_dados_seguro(ARQ_N)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 1) Aba Glicemia (Tabela completa, incluindo momentos extras)
        if not df_e_g.empty:
            df_e_g = _ensure_id_column(df_e_g.copy(), "ID", "GL")
            df_e_g.to_excel(writer, sheet_name="Glicemia_Tabela", index=False)
            wsT = writer.sheets["Glicemia_Tabela"]
            for cell in wsT[1]:
                cell.alignment = Alignment(horizontal="center")

            # 2) Aba Glicemia (Resumo pivot por Data x Momento)
            pivot = df_e_g.pivot_table(index="Data", columns="Momento", values="Valor", aggfunc="last")
            pivot.to_excel(writer, sheet_name="Glicemia_Resumo")
            ws1 = writer.sheets["Glicemia_Resumo"]

            f_ok = PatternFill("solid", fgColor="C8E6C9")  # verde claro
            f_hi = PatternFill("solid", fgColor="FFB6C1")  # vermelho claro
            f_lo = PatternFill("solid", fgColor="FFFFE0")  # amarelo claro

            for row in ws1.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    if cell.value is None or str(cell.value) == "nan":
                        continue
                    try:
                        val = int(cell.value)
                        cell.alignment = Alignment(horizontal="center")
                        if val < 70:
                            cell.fill = f_lo
                        elif val > 180:
                            cell.fill = f_hi
                        else:
                            cell.fill = f_ok
                    except Exception:
                        pass

        # 3) Aba Nutrição
        if not df_e_n.empty:
            df_e_n.to_excel(writer, sheet_name="Nutrição", index=False)
            ws2 = writer.sheets["Nutrição"]
            for cell in ws2[1]:
                cell.alignment = Alignment(horizontal="center")

    st.sidebar.download_button("⬇️ Baixar Agora", output.getvalue(), file_name="Relatorio_Saude_Kids.xlsx", use_container_width=True)

# ================= SAIR =================
if st.sidebar.button("🚪 Sair"):
    st.session_state.logado = False
    st.session_state.user_email = ""
    cookie_clear()
    st.rerun()
