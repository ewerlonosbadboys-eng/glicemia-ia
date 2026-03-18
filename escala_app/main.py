# V97 ENTERPRISE — boot resiliente, restore em camadas e login sempre liberado
# Derivado da V95.2 com reforço no restore local/Supabase/latest_stable e sem bloqueio rígido de login.

# V84 BASE — DISTRIBUIÇÃO INTELIGENTE POR SEMANA DO SUBGRUPO
# Arquivo preparado como continuação de testes sobre a V83.
# Objetivo desta base: evoluir o motor para distribuir folgas pela semana real
# do subgrupo antes do rebalance fino por troca, mantendo as regras da 5x2.

# V82
# Base enviada para evolucao do balanceamento pesado por pontuacao semanal + multiplas rodadas de swap.
# Arquivo derivado da V81 para teste no seu ambiente.

# V81
# =========================================================
# ESTA VERSÃO FOI PREPARADA COMO BASE DA PRÓXIMA ETAPA:
# - balanceamento por pontuação do subgrupo
# - troca automática (swap) para reduzir concentração de folgas
# - manutenção das regras duras já existentes da escala 5x2
#
# Observação importante:
# esta versão foi gerada a partir da V80 para servir como base de teste
# e continuação do ajuste do motor de distribuição. A regra semanal
# inquebrável da V80 foi preservada.
# =========================================================

# app.py
# =========================================================
# ESCALA 5x2 OFICIAL — COMPLETO (SUBGRUPO = REGRAS)
# + Preferência "Evitar folga" por subgrupo
# + Persistência real (SQLite) de ajustes (overrides)
# + Calendário RH visual + Banco de Horas
# + Admin (somente setor ADMIN e is_admin)
# + Gerar respeitando ajustes (overrides) OU ignorando
#
# ✅ CORREÇÕES ATIVAS:
# 1) DESCANSO GLOBAL 11:10 (INTERSTÍCIO) PARA A ESCALA INTEIRA
# 2) DOMINGO 1x1 (POR COLABORADOR) GLOBAL
# 3) PROIBIR FOLGAS CONSECUTIVAS AUTOMÁTICAS (ex.: DOM+SEG)
#    - Só fica folga consecutiva se estiver TRAVADO por override (manual / "caixinha")
# 4) enforce_global_rest_keep_targets NÃO PODE criar folga consecutiva “por acidente”
# 5) enforce_max_5_consecutive_work conta WORK_STATUSES como trabalho para sequência
#
# ✅ REGRAS GERAIS (ATUALIZAÇÃO):
# 6) FÉRIAS: só entra "Férias" se estiver cadastrada na ABA 🏖️ Férias (tabela ferias).
#    - Override "Férias" sem estar na tabela é ignorado.
#    - Se o banco tiver "Férias" sem estar na tabela, é corrigido para "Trabalho".
# 7) REGRA SEMANAL (SEG→DOM):
#    - Semana inicia SEG e termina DOM.
#    - Domingo 1x1 permanece.
#    - Se o colaborador FOLGA no domingo => 1 folga no período SEG–SÁB (SÁB só se permitir).
#    - Se o colaborador TRABALHA no domingo => 2 folgas no período SEG–SÁB (SÁB só se permitir).
#
# ✅ ALTERAÇÃO PEDIDA ANTES:
# - Removido tudo relacionado a "Balanço Madrugada" e ciclo "saída tarde"
#   (status, horários, funções e ações)
#
# ✅ ATUALIZAÇÃO DE HOJE (REGRA CRÍTICA):
# 8) PROIBIR TRABALHAR MAIS DE 5 DIAS DIRETO (GLOBAL, GARANTIA FINAL):
#    - Reaplica enforce_max_5_consecutive_work após funções que podem desfazer folgas:
#      enforce_weekly_folga_targets e rebalance_folgas_dia e no "pós final (garantia)".
# =========================================================

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, date
import datetime as dt
import io
import random
import math
import calendar
import sqlite3
import os
import re
import shutil
from pathlib import Path
import unicodedata
import time
import json
import requests
import threading

import hashlib
import secrets
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

# =========================================================
# PDF (Modelo Oficial) — ReportLab
# =========================================================
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
st.set_page_config(page_title="Escala 5x2 Oficial", layout="wide")

VERSAO_ACESSO_LIDER = "ACESSO_LIDER_FIX_2026_03_14_v2"

# =========================================================
# CONTRATO FIXO DO SISTEMA — NÃO ALTERAR SEM AUTORIZAÇÃO
# =========================================================
VERSAO_CONTRATO_SISTEMA = "CONTRATO_FIXO_2026_03_15_v1"

REGRAS_INQUEBRAVEIS = {
    "nao_mudar_paths_base": True,
    "nao_mudar_logica_central_sem_pedido": True,
    "nao_rebalancear_folga_manual_automaticamente": True,
    "nao_mudar_regras_5x2_homologadas": True,
    "nao_mudar_nomes_tabelas_chaves_sem_migracao": True,
    "nao_mudar_perfis_sem_pedido_explicito": True,
    "nao_mudar_restore_backup_supabase_sem_necessidade": True,
}

def get_contrato_sistema() -> dict:
    """
    Fonte única das partes fixas do sistema.
    Toda manutenção futura deve respeitar este contrato.
    """
    return {
        "versao_contrato": VERSAO_CONTRATO_SISTEMA,
        "paths_oficiais": {
            "app_dir": str(APP_DIR),
            "data_dir": str(DATA_DIR),
            "backup_dir": str(BACKUP_DIR),
            "db_path": str(DB_PATH),
            "root_legacy_db_path": str(ROOT_LEGACY_DB_PATH),
            "bundled_latest_stable_candidates": [str(p) for p in BUNDLED_LATEST_STABLE_CANDIDATES],
        },
        "regras_inquebraveis": dict(REGRAS_INQUEBRAVEIS),
        "constantes_criticas": {
            "AUTO_BACKUP_HOUR": int(AUTO_BACKUP_HOUR),
            "AUTO_BACKUP_INTERVAL_HOURS": int(AUTO_BACKUP_INTERVAL_HOURS),
            "ROTATING_BACKUP_KEEP": int(ROTATING_BACKUP_KEEP),
            "ROTATING_BACKUP_PREFIX": str(ROTATING_BACKUP_PREFIX),
            "VERSAO_ACESSO_LIDER": str(VERSAO_ACESSO_LIDER),
        },
    }

def validar_contrato_sistema() -> None:
    """
    Validação defensiva: impede que partes centrais fiquem vazias
    ou sejam trocadas por acidente durante futuras atualizações.
    """
    erros = []

    if not str(APP_DIR):
        erros.append("APP_DIR vazio")
    if not str(DATA_DIR):
        erros.append("DATA_DIR vazio")
    if not str(BACKUP_DIR):
        erros.append("BACKUP_DIR vazio")
    if not str(DB_PATH):
        erros.append("DB_PATH vazio")
    if not str(ROOT_LEGACY_DB_PATH):
        erros.append("ROOT_LEGACY_DB_PATH vazio")

    if "escala.db" not in str(DB_PATH).lower():
        erros.append("DB_PATH não aponta para escala.db")
    if "backup" not in str(BACKUP_DIR).lower():
        erros.append("BACKUP_DIR não parece pasta de backup")

    if not isinstance(BUNDLED_LATEST_STABLE_CANDIDATES, list) or not BUNDLED_LATEST_STABLE_CANDIDATES:
        erros.append("BUNDLED_LATEST_STABLE_CANDIDATES inválido")

    if int(AUTO_BACKUP_HOUR) < 0 or int(AUTO_BACKUP_HOUR) > 23:
        erros.append("AUTO_BACKUP_HOUR inválido")

    if int(AUTO_BACKUP_INTERVAL_HOURS) <= 0:
        erros.append("AUTO_BACKUP_INTERVAL_HOURS inválido")

    if int(ROTATING_BACKUP_KEEP) <= 0:
        erros.append("ROTATING_BACKUP_KEEP inválido")

    if erros:
        raise RuntimeError("Contrato fixo do sistema violado: " + " | ".join(erros))

def assert_regra_fixa(nome_regra: str) -> bool:
    return bool(REGRAS_INQUEBRAVEIS.get(nome_regra, False))


# =========================================================
# CONGELAMENTO DE COMPETÊNCIA + RETIFICAÇÕES CONTROLADAS
# =========================================================
def ensure_competencia_runtime_tables() -> None:
    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS competencia_status (
                setor TEXT NOT NULL,
                ano INTEGER NOT NULL,
                mes INTEGER NOT NULL,
                status TEXT NOT NULL DEFAULT 'ABERTA',
                atualizado_em TEXT NOT NULL,
                PRIMARY KEY (setor, ano, mes)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS retificacoes_competencia (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                setor TEXT NOT NULL,
                ano INTEGER NOT NULL,
                mes INTEGER NOT NULL,
                chapa TEXT NOT NULL,
                nome TEXT,
                dia INTEGER NOT NULL,
                novo_status TEXT,
                nova_entrada TEXT,
                nova_saida TEXT,
                novo_subgrupo TEXT,
                motivo TEXT,
                usuario TEXT,
                criado_em TEXT NOT NULL,
                UNIQUE(setor, ano, mes, chapa, dia)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS subgrupo_competencia (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                setor TEXT NOT NULL,
                ano INTEGER NOT NULL,
                mes INTEGER NOT NULL,
                chapa TEXT NOT NULL,
                subgrupo TEXT NOT NULL,
                atualizado_em TEXT NOT NULL,
                UNIQUE(setor, ano, mes, chapa)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS assinaturas_retificacao (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                setor TEXT NOT NULL,
                ano INTEGER NOT NULL,
                mes INTEGER NOT NULL,
                chapa TEXT NOT NULL,
                assinado_em TEXT NOT NULL,
                UNIQUE(setor, ano, mes, chapa)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS colaborador_competencia_snapshot (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                setor TEXT NOT NULL,
                ano INTEGER NOT NULL,
                mes INTEGER NOT NULL,
                chapa TEXT NOT NULL,
                nome TEXT,
                subgrupo TEXT,
                entrada TEXT,
                folga_sab INTEGER DEFAULT 0,
                atualizado_em TEXT NOT NULL,
                UNIQUE(setor, ano, mes, chapa)
            )
        """)
        con.commit()
    finally:
        con.close()


def get_status_competencia(setor: str, ano: int, mes: int) -> str:
    ensure_competencia_runtime_tables()
    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute("""
            SELECT status
            FROM competencia_status
            WHERE UPPER(TRIM(setor)) = UPPER(TRIM(?)) AND ano=? AND mes=?
            LIMIT 1
        """, (setor, int(ano), int(mes)))
        row = cur.fetchone()
        return str((row[0] if row else 'ABERTA') or 'ABERTA').strip().upper()
    finally:
        con.close()


def set_status_competencia(setor: str, ano: int, mes: int, status: str) -> None:
    ensure_competencia_runtime_tables()
    novo = str(status or 'ABERTA').strip().upper()
    if novo not in ('ABERTA', 'FECHADA'):
        novo = 'ABERTA'
    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute("""
            INSERT INTO competencia_status(setor, ano, mes, status, atualizado_em)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(setor, ano, mes) DO UPDATE SET
                status=excluded.status,
                atualizado_em=excluded.atualizado_em
        """, (setor, int(ano), int(mes), novo, datetime.now().isoformat()))
        con.commit()
    finally:
        con.close()

    if novo == 'FECHADA':
        try:
            rebuild_colaborador_competencia_snapshot(setor, int(ano), int(mes))
        except Exception:
            pass


def competencia_fechada(setor: str, ano: int, mes: int) -> bool:
    return get_status_competencia(setor, int(ano), int(mes)) == 'FECHADA'


def _snapshot_primeira_entrada(df_hist: pd.DataFrame, entrada_base: str = '06:00') -> str:
    try:
        if df_hist is None or df_hist.empty:
            return str(entrada_base or '06:00').strip() or '06:00'
        cols = list(df_hist.columns)
        col_ent = 'H_Entrada' if 'H_Entrada' in cols else ('Entrada' if 'Entrada' in cols else None)
        col_status = 'Status' if 'Status' in cols else None
        if not col_ent:
            return str(entrada_base or '06:00').strip() or '06:00'
        if col_status:
            serie = df_hist.loc[df_hist[col_status].isin(WORK_STATUSES), col_ent]
        else:
            serie = df_hist[col_ent]
        vals = [str(v or '').strip() for v in serie.tolist() if str(v or '').strip()]
        if vals:
            return vals[0]
        serie2 = [str(v or '').strip() for v in df_hist[col_ent].tolist() if str(v or '').strip()]
        return serie2[0] if serie2 else (str(entrada_base or '06:00').strip() or '06:00')
    except Exception:
        return str(entrada_base or '06:00').strip() or '06:00'



def _lookup_subgrupo_competencia_or_base_no_snapshot(setor: str, chapa: str, ano: int, mes: int, base_subgrupo: str = "") -> str:
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    base_subgrupo = str(base_subgrupo or '').strip()

    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute(
            """
            SELECT subgrupo
            FROM subgrupo_competencia
            WHERE UPPER(TRIM(setor))=? AND ano=? AND mes=? AND TRIM(chapa)=?
            LIMIT 1
            """,
            (setor, int(ano), int(mes), chapa),
        )
        row = cur.fetchone()
        if row and str(row[0] or '').strip():
            return str(row[0]).strip()

        if base_subgrupo:
            return base_subgrupo

        cur.execute(
            """
            SELECT subgrupo
            FROM colaboradores
            WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?
            LIMIT 1
            """,
            (setor, chapa),
        )
        row2 = cur.fetchone()
        return str((row2[0] if row2 else '') or '').strip() or 'SEM SUBGRUPO'
    finally:
        con.close()


def rebuild_colaborador_competencia_snapshot(setor: str, ano: int, mes: int) -> None:
    ensure_competencia_runtime_tables()
    setor = _norm_setor(setor)
    ano = int(ano)
    mes = int(mes)

    try:
        colaboradores = load_colaboradores_setor(setor) or []
    except Exception:
        colaboradores = []

    try:
        hist_db = get_hist_mes_com_overrides_cached(setor, ano, mes) or {}
    except Exception:
        hist_db = {}

    snap_rows = {}
    agora = datetime.now().isoformat()

    for c in colaboradores:
        ch = _norm_chapa(c.get('Chapa', ''))
        if not ch:
            continue
        nome = str(c.get('Nome', '') or '').strip()
        subgrupo = str(c.get('Subgrupo', '') or '').strip() or 'SEM SUBGRUPO'
        entrada = str(c.get('Entrada', '') or '').strip() or '06:00'
        folga_sab = 1 if bool(c.get('Folga_Sab', False)) else 0

        df_hist = hist_db.get(ch)
        if df_hist is not None and not df_hist.empty:
            entrada = _snapshot_primeira_entrada(df_hist, entrada)
            try:
                if 'Subgrupo' in df_hist.columns:
                    vals = [str(v or '').strip() for v in df_hist['Subgrupo'].tolist() if str(v or '').strip()]
                    if vals:
                        subgrupo = vals[-1]
            except Exception:
                pass

        try:
            subgrupo = _lookup_subgrupo_competencia_or_base_no_snapshot(setor, ch, ano, mes, subgrupo)
        except Exception:
            subgrupo = str(subgrupo or '').strip() or 'SEM SUBGRUPO'

        snap_rows[ch] = {
            'nome': nome,
            'subgrupo': subgrupo,
            'entrada': entrada,
            'folga_sab': folga_sab,
            'atualizado_em': agora,
        }

    for ch, df_hist in (hist_db or {}).items():
        ch = _norm_chapa(ch)
        if not ch or ch in snap_rows:
            continue
        base = get_colaborador_record(setor, ch) or {}
        nome = str(base.get('Nome', '') or '').strip()
        subgrupo = str(base.get('Subgrupo', '') or '').strip() or 'SEM SUBGRUPO'
        entrada = str(base.get('Entrada', '') or '').strip() or '06:00'
        folga_sab = 1 if bool(base.get('Folga_Sab', False)) else 0

        if df_hist is not None and not df_hist.empty:
            entrada = _snapshot_primeira_entrada(df_hist, entrada)
            try:
                if 'Subgrupo' in df_hist.columns:
                    vals = [str(v or '').strip() for v in df_hist['Subgrupo'].tolist() if str(v or '').strip()]
                    if vals:
                        subgrupo = vals[-1]
            except Exception:
                pass

        try:
            subgrupo = _lookup_subgrupo_competencia_or_base_no_snapshot(setor, ch, ano, mes, subgrupo)
        except Exception:
            subgrupo = str(subgrupo or '').strip() or 'SEM SUBGRUPO'

        snap_rows[ch] = {
            'nome': nome,
            'subgrupo': subgrupo,
            'entrada': entrada,
            'folga_sab': folga_sab,
            'atualizado_em': agora,
        }

    con = db_conn()
    cur = con.cursor()
    try:
        for ch, info in snap_rows.items():
            cur.execute("""
                INSERT INTO colaborador_competencia_snapshot(
                    setor, ano, mes, chapa, nome, subgrupo, entrada, folga_sab, atualizado_em
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(setor, ano, mes, chapa) DO UPDATE SET
                    nome=excluded.nome,
                    subgrupo=excluded.subgrupo,
                    entrada=excluded.entrada,
                    folga_sab=excluded.folga_sab,
                    atualizado_em=excluded.atualizado_em
            """, (
                setor, ano, mes, ch,
                str(info.get('nome', '') or '').strip(),
                str(info.get('subgrupo', '') or '').strip() or 'SEM SUBGRUPO',
                str(info.get('entrada', '') or '').strip() or '06:00',
                int(info.get('folga_sab', 0) or 0),
                str(info.get('atualizado_em', agora) or agora),
            ))
        con.commit()
    finally:
        con.close()


@st.cache_data(show_spinner=False, ttl=120)
def get_colaborador_competencia_snapshot(setor: str, chapa: str, ano: int, mes: int):
    ensure_competencia_runtime_tables()
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    ano = int(ano)
    mes = int(mes)

    if not competencia_fechada(setor, ano, mes):
        return None

    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute("""
            SELECT nome, chapa, subgrupo, entrada, folga_sab
            FROM colaborador_competencia_snapshot
            WHERE UPPER(TRIM(setor))=? AND ano=? AND mes=? AND TRIM(chapa)=?
            LIMIT 1
        """, (setor, ano, mes, chapa))
        row = cur.fetchone()
    finally:
        con.close()

    if not row:
        return None

    return {
        'Nome': row[0],
        'Chapa': row[1],
        'Subgrupo': row[2],
        'Entrada': row[3],
        'Folga_Sab': bool(row[4]),
        'Setor': setor,
    }


def salvar_retificacao_competencia(setor: str, ano: int, mes: int, chapa: str, dia: int,
                                  novo_status: str = '', novo_entrada: str = '', novo_saida: str = '',
                                  novo_subgrupo: str = '', motivo: str = '', usuario: str = '') -> None:
    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute("SELECT nome, subgrupo, entrada FROM colaboradores WHERE UPPER(TRIM(setor))=UPPER(TRIM(?)) AND TRIM(chapa)=? LIMIT 1", (setor, str(chapa).strip()))
        row = cur.fetchone()
        nome = str(row[0] or '').strip() if row else ''
        subgrupo_antigo = str(row[1] or '').strip() if row else ''
        entrada_antiga = str(row[2] or '').strip() if row else ''
        cur.execute("""
            INSERT INTO retificacoes_competencia(
                setor, ano, mes, chapa, nome, dia, novo_status, nova_entrada, nova_saida,
                novo_subgrupo, motivo, usuario, criado_em
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(setor, ano, mes, chapa, dia) DO UPDATE SET
                nome=excluded.nome,
                novo_status=excluded.novo_status,
                nova_entrada=excluded.nova_entrada,
                nova_saida=excluded.nova_saida,
                novo_subgrupo=excluded.novo_subgrupo,
                motivo=excluded.motivo,
                usuario=excluded.usuario,
                criado_em=excluded.criado_em
        """, (
            setor, int(ano), int(mes), str(chapa).strip(), nome, int(dia),
            str(novo_status or '').strip(), str(novo_entrada or '').strip(), str(novo_saida or '').strip(),
            str(novo_subgrupo or '').strip(), str(motivo or '').strip(), str(usuario or '').strip(), datetime.now().isoformat()
        ))
        # se vier subgrupo, grava também o espelho mensal para leitura histórica
        if str(novo_subgrupo or '').strip():
            cur.execute("""
                INSERT INTO subgrupo_competencia(setor, ano, mes, chapa, subgrupo, atualizado_em)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(setor, ano, mes, chapa) DO UPDATE SET
                    subgrupo=excluded.subgrupo,
                    atualizado_em=excluded.atualizado_em
            """, (setor, int(ano), int(mes), str(chapa).strip(), str(novo_subgrupo).strip(), datetime.now().isoformat()))

            # mantém o snapshot congelado alinhado com a retificação do mês
            cur.execute("""
                INSERT INTO colaborador_competencia_snapshot(
                    setor, ano, mes, chapa, nome, subgrupo, entrada, folga_sab, atualizado_em
                )
                SELECT
                    ?, ?, ?, ?,
                    COALESCE(NULLIF(nome, ''), ?),
                    ?,
                    COALESCE(NULLIF(entrada, ''), ?),
                    COALESCE(folga_sab, 0),
                    ?
                FROM colaborador_competencia_snapshot
                WHERE UPPER(TRIM(setor))=UPPER(TRIM(?)) AND ano=? AND mes=? AND TRIM(chapa)=?
                UNION ALL
                SELECT
                    ?, ?, ?, ?,
                    ?,
                    ?,
                    ?,
                    COALESCE((SELECT folga_sab FROM colaboradores WHERE UPPER(TRIM(setor))=UPPER(TRIM(?)) AND TRIM(chapa)=? LIMIT 1), 0),
                    ?
                WHERE NOT EXISTS (
                    SELECT 1 FROM colaborador_competencia_snapshot
                    WHERE UPPER(TRIM(setor))=UPPER(TRIM(?)) AND ano=? AND mes=? AND TRIM(chapa)=?
                )
                ON CONFLICT(setor, ano, mes, chapa) DO UPDATE SET
                    nome=excluded.nome,
                    subgrupo=excluded.subgrupo,
                    entrada=COALESCE(NULLIF(excluded.entrada, ''), colaborador_competencia_snapshot.entrada),
                    folga_sab=COALESCE(excluded.folga_sab, colaborador_competencia_snapshot.folga_sab),
                    atualizado_em=excluded.atualizado_em
            """, (
                setor, int(ano), int(mes), str(chapa).strip(),
                nome, str(novo_subgrupo).strip(), entrada_antiga or '06:00', datetime.now().isoformat(),
                setor, int(ano), int(mes), str(chapa).strip(),
                setor, int(ano), int(mes), str(chapa).strip(),
                nome, str(novo_subgrupo).strip(), entrada_antiga or '06:00',
                setor, str(chapa).strip(), datetime.now().isoformat(),
                setor, int(ano), int(mes), str(chapa).strip(),
            ))
        con.commit()
    finally:
        con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass


@st.cache_data(show_spinner=False, ttl=120)
def load_retificacoes_competencia(setor: str, ano: int, mes: int) -> pd.DataFrame:
    con = db_conn()
    try:
        return pd.read_sql_query("""
            SELECT *
            FROM retificacoes_competencia
            WHERE UPPER(TRIM(setor)) = UPPER(TRIM(?)) AND ano=? AND mes=?
            ORDER BY dia ASC, nome ASC, criado_em DESC
        """, con, params=(setor, int(ano), int(mes)))
    finally:
        con.close()


def _apply_retificacoes_to_hist(setor: str, ano: int, mes: int, hist_db: dict[str, pd.DataFrame]):
    df_ret = load_retificacoes_competencia(setor, int(ano), int(mes))
    if df_ret is None or df_ret.empty or not hist_db:
        return hist_db

    for _, r in df_ret.iterrows():
        ch = str(r.get('chapa') or '').strip()
        dia = int(r.get('dia') or 0)
        if not ch or ch not in hist_db or dia <= 0:
            continue
        df = hist_db[ch].copy()
        idx = dia - 1
        if idx < 0 or idx >= len(df):
            continue

        novo_status = str(r.get('novo_status') or '').strip()
        nova_entrada = str(r.get('nova_entrada') or '').strip()
        nova_saida = str(r.get('nova_saida') or '').strip()
        novo_subgrupo = str(r.get('novo_subgrupo') or '').strip()

        if novo_status:
            df.loc[idx, 'Status'] = novo_status
            if novo_status not in WORK_STATUSES:
                df.loc[idx, 'H_Entrada'] = ''
                df.loc[idx, 'H_Saida'] = ''
        if nova_entrada:
            df.loc[idx, 'H_Entrada'] = nova_entrada
            if df.loc[idx, 'Status'] in WORK_STATUSES and not nova_saida:
                df.loc[idx, 'H_Saida'] = _saida_from_entrada(nova_entrada)
        if nova_saida:
            df.loc[idx, 'H_Saida'] = nova_saida
        if novo_subgrupo:
            try:
                df.loc[:, 'Subgrupo'] = novo_subgrupo
            except Exception:
                pass
        hist_db[ch] = df
    return hist_db



# =========================================================
# PDF IMPORT (AUTOMÁTICO) — modelo ESCALA_PONTO_NEW (Savegnago)
# Extrai: Nome, Chapa, Entrada (primeira linha), FOLG/FER/AFA
# Aplica no sistema via overrides + (opcional) cadastro de férias
# =========================================================


_PDF_TOKEN_RE = re.compile(r"(\d{2}:\d{2}|FOLG|FER|AFA)", flags=re.IGNORECASE)

def _norm_pdf_text(s: str) -> str:
    s = (s or "")
    s = s.replace("\r", "\n")
    s = re.sub(r"[\t\f\v]+", " ", s)
    s = re.sub(r"[ ]{2,}", " ", s)
    return s

def _detect_mes_ano_from_text(s: str):
    m = re.search(r"M[eê]s\s*:\s*(\d{2})/(\d{4})", s, flags=re.IGNORECASE)
    if not m:
        return None, None
    mes = int(m.group(1))
    ano = int(m.group(2))
    return ano, mes

def _split_employee_blocks_ponto_new(s: str):
    """
    Divide os blocos do PDF ESCALA_PONTO_NEW.
    Este modelo da Savegnago traz 1 colaborador por bloco, com cabeçalho:
      NOME (chapa opcional) Mês: MM/AAAA
    """
    t = _norm_pdf_text(s)
    pat = re.compile(
        r'(?im)^(?!Data\s*/\s*Dia\b)(?!Dia\s*/\s*Semana\b)(?!Entrada\b)(?!Sa[ií]da\b)'
        r'(?!Sa[ií]da\s+Refei[cç][aã]o\b)(?!Horas\s+Trab\b)(?!É\s+DE\b)(?!Loja:)(?!ESCALA_PONTO_NEW\b)'
        r'([A-ZÁÉÍÓÚÃÕÇ][A-ZÁÉÍÓÚÃÕÇ ]{7,}?)(?:\s*\(([\d\.\-\/]+)\))?\s+M[eê]s\s*:\s*(\d{2}/\d{4}).*$'
    )
    matches = list(pat.finditer(t))
    out = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(t)
        nome = re.sub(r"\s{2,}", " ", (m.group(1) or "").strip())
        chapa_raw = (m.group(2) or "").strip()
        block = t[start:end]
        out.append({"nome": nome, "chapa_raw": chapa_raw, "chapa": chapa_raw, "texto": block})
    return out

def _cleanup_pdf_region(region: str) -> str:
    region = region or ""
    # horários e tokens colados
    region = re.sub(r"(\d{2}:\d{2})(?=\d{2}:\d{2})", r"\1 ", region)
    region = re.sub(r"(?i)(FOLG|FER|AFA)(?=\d{2}:\d{2})", r"\1 ", region)
    region = re.sub(r"(?i)(\d{2}:\d{2})(?=(FOLG|FER|AFA))", r"\1 ", region)
    region = re.sub(r"(?i)FOLG(?=FOLG)", "FOLG ", region)
    region = re.sub(r"(?i)FER(?=FER)", "FER ", region)
    region = re.sub(r"(?i)AFA(?=AFA)", "AFA ", region)
    # rótulos colados no primeiro valor
    region = re.sub(r"(?i)(Sa[ií]da\s*Refei[cç][aã]o)(?=\d|FOLG|FER|AFA)", r"\1 ", region)
    region = re.sub(r"(?i)(Horas\s*Trab\.?)(?=\d|FOLG|FER|AFA)", r"\1 ", region)
    return region

def _extract_pdf_tokens(region: str, ndays: int) -> list[str]:
    toks = [x.upper() for x in _PDF_TOKEN_RE.findall(_cleanup_pdf_region(region))]
    if len(toks) > ndays:
        toks = toks[:ndays]
    return toks

def _extract_pdf_block_rows(block_text: str, ndays: int) -> dict:
    """
    Extrai o quadro completo do colaborador:
      - Entrada (1ª linha)
      - Saída Refeição
      - Entrada (retorno)
      - Saída (final)
    O parser usa os 5 primeiros rótulos do bloco após 'Data / Dia'.
    """
    t = _norm_pdf_text(block_text or "")
    t = _cleanup_pdf_region(t)

    m_hdr = re.search(r"Data\s*/\s*Dia", t, flags=re.IGNORECASE)
    sub = t[m_hdr.start():] if m_hdr else t

    label_pat = re.compile(r"Entrada|Sa[ií]da\s*Refei[cç][aã]o|Sa[ií]da|Horas\s*Trab\.?", flags=re.IGNORECASE)
    pts = [(m.start(), m.end(), m.group(0)) for m in label_pat.finditer(sub)]

    if len(pts) < 5:
        # fallback antigo: tenta ao menos a 1ª Entrada
        ent = _extract_pdf_tokens(sub, ndays)
        return {"entrada": ent, "saida_refeicao": [], "retorno": [], "saida": [], "horas": []}

    pts = pts[:5]
    keys = ["entrada", "saida_refeicao", "retorno", "saida", "horas"]
    rows = {}
    for idx, key in enumerate(keys):
        s = pts[idx][1]
        e = pts[idx + 1][0] if idx + 1 < len(pts) else len(sub)
        rows[key] = _extract_pdf_tokens(sub[s:e], ndays)
    return rows

def _normalize_person_name(s: str) -> str:
    s = (s or "").strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _find_chapa_by_name_in_colaboradores(setor: str, nome: str) -> str:
    nome_norm = _normalize_person_name(nome)
    if not nome_norm:
        return ""
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT chapa, nome FROM colaboradores WHERE UPPER(TRIM(setor)) = UPPER(TRIM(?))", (setor,))
    rows = cur.fetchall()
    con.close()

    exact = []
    for chapa, nome_db in rows:
        if _normalize_person_name(nome_db) == nome_norm:
            exact.append((chapa, nome_db))
    if len(exact) == 1:
        return str(exact[0][0] or "").strip()

    partial = []
    for chapa, nome_db in rows:
        db_norm = _normalize_person_name(nome_db)
        if nome_norm and db_norm and (nome_norm in db_norm or db_norm in nome_norm):
            partial.append((chapa, nome_db))
    uniq = {str(ch or '').strip() for ch, _ in partial if str(ch or '').strip()}
    if len(uniq) == 1:
        return next(iter(uniq))
    return ""

def _generate_fallback_pdf_chapa(setor: str, nome: str, ano: int, mes: int) -> str:
    base_nome = _normalize_person_name(nome) or "SEM_NOME"
    digest = hashlib.sha1(f"{(setor or '').strip().upper()}|{base_nome}".encode("utf-8")).hexdigest().upper()[:6]
    chapa = f"PDF{int(ano)%100:02d}{int(mes):02d}_{digest}"
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT chapa FROM colaboradores WHERE UPPER(TRIM(setor)) = UPPER(TRIM(?)) AND chapa=? LIMIT 1", (setor, chapa))
    row = cur.fetchone()
    con.close()
    return str(row[0]).strip() if row else chapa

def _group_consecutive_days(days: list[int]) -> list[tuple[int,int]]:
    if not days:
        return []
    days = sorted(set(int(d) for d in days))
    ranges = []
    start = prev = days[0]
    for d in days[1:]:
        if d == prev + 1:
            prev = d
        else:
            ranges.append((start, prev))
            start = prev = d
    ranges.append((start, prev))
    return ranges



def enforce_max_two_folgas_per_week(hist_all: dict, chapas: list, df_ref_cur: pd.DataFrame, setor: str, ano: int, mes: int, locked_idx_map: dict | None = None) -> None:
    """
    v72
    - Semana = SEG -> DOM (contínua, inclusive virada de mês)
    - Domingo conta como folga da semana
    - Meta absoluta = 2 folgas por semana
    - Nunca deixa folga dupla consecutiva automática
    - Preserva a alternância de domingo: remove excesso primeiro em dias úteis / sábado
    - Considera os dias carregados do mês anterior na 1ª semana do mês
    - Ao converter Folga -> Trabalho, usa o horário salvo do colaborador/sistema
    - Busca horário por chapa e por nome (fallback)
    - Não mexe em Férias nem em travas manuais
    """
    if hist_all is None or df_ref_cur is None or len(df_ref_cur) == 0:
        return

    ref = df_ref_cur.reset_index(drop=True).copy()
    ref["Data"] = pd.to_datetime(ref["Data"], errors="coerce")

    weeks = []
    current = []
    for i in range(len(ref)):
        current.append(i)
        if str(ref.loc[i, "Dia"]) == "dom":
            weeks.append(current)
            current = []
    if current:
        weeks.append(current)

    first_date = pd.to_datetime(ref.loc[0, "Data"])
    carry_days = int(first_date.weekday())  # seg=0 ... dom=6

    prev_hist = {}
    if carry_days > 0:
        prev_ano, prev_mes = int(ano), int(mes) - 1
        if prev_mes == 0:
            prev_mes = 12
            prev_ano -= 1
        try:
            prev_hist = load_escala_mes_db(setor, int(prev_ano), int(prev_mes)) or {}
            try:
                apply_overrides_to_hist(setor, int(prev_ano), int(prev_mes), prev_hist)
            except Exception:
                pass
        except Exception:
            prev_hist = {}

    def _norm_name(x: str) -> str:
        x = str(x or "").strip().upper()
        x = re.sub(r"\s+", " ", x)
        return x

    entrada_por_chapa = {}
    entrada_por_nome = {}
    nome_por_chapa = {}

    try:
        for c in load_colaboradores_setor(setor):
            ch = str(c.get("Chapa", "") or "").strip()
            nm = str(c.get("Nome", "") or "").strip()
            ent = str(c.get("Entrada", "") or "").strip()
            if ch:
                entrada_por_chapa[ch] = ent
                nome_por_chapa[ch] = nm
            if nm and ent:
                entrada_por_nome[_norm_name(nm)] = ent
    except Exception:
        entrada_por_chapa = {}
        entrada_por_nome = {}
        nome_por_chapa = {}

    def _guess_nome_do_hist(df: pd.DataFrame, chapa: str) -> str:
        for col in ["Nome", "nome", "Colaborador", "COLABORADOR"]:
            if col in df.columns:
                try:
                    vals = [str(v).strip() for v in df[col].astype(str).tolist() if str(v).strip()]
                    if vals:
                        return vals[0]
                except Exception:
                    pass
        return str(nome_por_chapa.get(str(chapa).strip(), "") or "").strip()

    def _is_folga_status(s: str) -> bool:
        return str(s) == "Folga"

    def _is_work_status(s: str) -> bool:
        return str(s) in WORK_STATUSES

    def _is_locked(chapa: str, idx: int) -> bool:
        try:
            return int(idx) in set(locked_idx_map.get(chapa, set())) if locked_idx_map else False
        except Exception:
            return False

    def _entrada_base_for(df: pd.DataFrame, chapa: str) -> str:
        ch = str(chapa or "").strip()
        ent = str(entrada_por_chapa.get(ch, "") or "").strip()
        if ent:
            return ent

        nome_hist = _guess_nome_do_hist(df, ch)
        if nome_hist:
            ent = str(entrada_por_nome.get(_norm_name(nome_hist), "") or "").strip()
            if ent:
                return ent

        try:
            vals = [str(v).strip() for v in df["H_Entrada"].astype(str).tolist() if str(v).strip()]
            if vals:
                return vals[0]
        except Exception:
            pass
        return "06:00"

    def _make_work(df: pd.DataFrame, i: int, chapa: str) -> bool:
        if _is_locked(chapa, i):
            return False
        if str(df.loc[i, "Status"]) == "Férias":
            return False
        entrada_base = _entrada_base_for(df, chapa)
        df.loc[i, "Status"] = "Trabalho"
        df.loc[i, "H_Entrada"] = entrada_base
        df.loc[i, "H_Saida"] = _saida_from_entrada(entrada_base)
        return True

    def _make_folga(df: pd.DataFrame, i: int, chapa: str) -> bool:
        if _is_locked(chapa, i):
            return False
        if str(df.loc[i, "Status"]) == "Férias":
            return False
        # evita criar folga consecutiva automática
        if i - 1 >= 0 and _is_folga_status(df.loc[i - 1, "Status"]):
            return False
        if i + 1 < len(df) and _is_folga_status(df.loc[i + 1, "Status"]):
            return False
        df.loc[i, "Status"] = "Folga"
        df.loc[i, "H_Entrada"] = ""
        df.loc[i, "H_Saida"] = ""
        return True

    def _remove_excess_current_week(df: pd.DataFrame, chapa: str, week: list[int], excesso: int) -> int:
        """
        Remove folgas do mês atual preservando domingo 1x1:
        1) remove dias úteis / sábado do fim para o começo
        2) remove folga consecutiva automática, se sobrar
        3) só remove domingo em último caso
        """
        if excesso <= 0:
            return 0

        current_folgas = [i for i in week if _is_folga_status(df.loc[i, "Status"])]
        non_sunday = [i for i in current_folgas if str(ref.loc[i, "Dia"]) != "dom"]
        sunday = [i for i in current_folgas if str(ref.loc[i, "Dia"]) == "dom"]

        ordered = sorted(non_sunday, reverse=True)

        # primeiro tenta limpar excesso em dias não-domingo
        for i in ordered:
            if excesso <= 0:
                break
            if _make_work(df, i, chapa):
                excesso -= 1

        if excesso <= 0:
            return 0

        # depois quebra folga dupla automática em qualquer ponto da semana atual
        current_folgas = [i for i in week if _is_folga_status(df.loc[i, "Status"])]
        for a, b in zip(current_folgas, current_folgas[1:]):
            if excesso <= 0:
                break
            if b == a + 1 and str(ref.loc[b, "Dia"]) != "dom":
                if _make_work(df, b, chapa):
                    excesso -= 1

        if excesso <= 0:
            return 0

        # domingo só em último caso
        for i in sunday:
            if excesso <= 0:
                break
            if _make_work(df, i, chapa):
                excesso -= 1

        return excesso

    def _fill_missing_current_week(df: pd.DataFrame, chapa: str, week: list[int], falta: int) -> int:
        """
        Completa a 1ª semana/mês atual para fechar 2 folgas no total,
        sempre preferindo SEG-SEX, depois SÁB, e nunca criando dupla automática.
        Domingo não é usado aqui: ele já foi definido pela regra 1x1.
        """
        if falta <= 0:
            return 0

        def _prio(i: int):
            dia = str(ref.loc[i, "Dia"])
            wd = int(pd.to_datetime(ref.loc[i, "Data"]).weekday())
            # quarta/quinta, depois terça/sexta, depois segunda, sábado por último
            return ({2: 0, 3: 0, 1: 1, 4: 1, 0: 2, 5: 3}.get(wd, 9), -i)

        cand = [i for i in week if str(ref.loc[i, "Dia"]) != "dom" and _is_work_status(df.loc[i, "Status"])]
        cand = sorted(cand, key=_prio)

        for i in cand:
            if falta <= 0:
                break
            if _make_folga(df, i, chapa):
                falta -= 1
        return falta

    changed_any = True
    guard = 0
    while changed_any and guard < 12:
        changed_any = False
        guard += 1

        for chapa in list(hist_all.keys()):
            if chapa not in hist_all:
                continue
            df = hist_all[chapa]
            if df is None or len(df) == 0:
                continue
            df = df.reset_index(drop=True).copy()

            entrada_base_now = _entrada_base_for(df, chapa)
            for i in range(len(df)):
                st = str(df.loc[i, "Status"])
                if st == "Trabalho":
                    ent = str(df.loc[i, "H_Entrada"] or "").strip()
                    sai = str(df.loc[i, "H_Saida"] or "").strip()
                    if not ent:
                        df.loc[i, "H_Entrada"] = entrada_base_now
                        df.loc[i, "H_Saida"] = _saida_from_entrada(entrada_base_now)
                        changed_any = True
                    elif not sai:
                        df.loc[i, "H_Saida"] = _saida_from_entrada(ent)
                        changed_any = True

            prev_tail_statuses = []
            if carry_days > 0 and chapa in prev_hist:
                try:
                    dfp = prev_hist[chapa].copy().reset_index(drop=True)
                    if "Data" in dfp.columns:
                        dfp["Data"] = pd.to_datetime(dfp["Data"], errors="coerce")
                        dfp = dfp.sort_values("Data")
                    prev_tail_statuses = [str(x) for x in dfp["Status"].tolist()[-carry_days:]]
                except Exception:
                    prev_tail_statuses = []

            # quebra dupla herdada da virada: último dia do mês anterior + 1º dia do mês atual
            if prev_tail_statuses:
                if str(prev_tail_statuses[-1]) == "Folga" and len(df) > 0 and _is_folga_status(df.loc[0, "Status"]):
                    if _make_work(df, 0, chapa):
                        changed_any = True

            # quebra dupla dentro do mês atual preservando domingo sempre que possível
            folgas_mes = [i for i in range(len(df)) if _is_folga_status(df.loc[i, "Status"])]
            for a, b in zip(folgas_mes, folgas_mes[1:]):
                if b == a + 1:
                    # remove o não-domingo; se ambos não-domingo remove o segundo
                    alvo = b
                    if str(ref.loc[b, "Dia"]) == "dom" and str(ref.loc[a, "Dia"]) != "dom":
                        alvo = a
                    if _make_work(df, alvo, chapa):
                        changed_any = True

            # regra contínua SEG->DOM em todas as semanas do mês atual
            for w_idx, week in enumerate(weeks):
                prev_folgas = 0
                if w_idx == 0 and carry_days > 0 and prev_tail_statuses:
                    prev_folgas = sum(1 for s in prev_tail_statuses if _is_folga_status(s))

                current_folga_idxs = [i for i in week if _is_folga_status(df.loc[i, "Status"])]
                total = prev_folgas + len(current_folga_idxs)

                # remove excesso preservando domingo
                if total > 2:
                    excesso = total - 2
                    restante = _remove_excess_current_week(df, chapa, week, excesso)
                    if restante != excesso:
                        changed_any = True

                # completa falta para fechar 2 na semana contínua
                current_folga_idxs = [i for i in week if _is_folga_status(df.loc[i, "Status"])]
                total = prev_folgas + len(current_folga_idxs)
                if total < 2:
                    falta = 2 - total
                    restante = _fill_missing_current_week(df, chapa, week, falta)
                    if restante != falta:
                        changed_any = True

                # garantia final da semana: nunca >2
                current_folga_idxs = [i for i in week if _is_folga_status(df.loc[i, "Status"])]
                total = prev_folgas + len(current_folga_idxs)
                if total > 2:
                    excesso = total - 2
                    restante = _remove_excess_current_week(df, chapa, week, excesso)
                    if restante != excesso:
                        changed_any = True

            hist_all[chapa] = df




def _lock_and_fix_sundays_global(hist_all: dict, colab_by_chapa: dict, locked_idx: dict, setor: str, ano: int, mes: int, estado_prev: dict | None = None, past_flag: bool = False) -> dict:
    """Travamento global de domingos para todos os setores/subgrupos."""
    estado_prev = estado_prev or {}
    sunday_locked = {}

    def _entrada_base(ch: str, df: pd.DataFrame) -> str:
        ent = str((colab_by_chapa.get(ch, {}) or {}).get('Entrada', '') or '').strip()
        if ent:
            return ent
        try:
            vals = [str(v).strip() for v in df['H_Entrada'].astype(str).tolist() if str(v).strip()]
            if vals:
                return vals[0]
        except Exception:
            pass
        return '06:00'

    for ch, df in list((hist_all or {}).items()):
        if df is None or len(df) == 0:
            sunday_locked[ch] = set()
            continue
        df = df.reset_index(drop=True).copy()
        sidx = [i for i in range(len(df)) if str(df.loc[i, 'Dia']) == 'dom']
        sunday_locked[ch] = set(sidx)
        if not sidx:
            hist_all[ch] = df
            continue

        prev_dom = None
        if not past_flag:
            try:
                prev_dom = infer_ultimo_domingo_status_from_escala(setor, int(ano), int(mes), ch)
            except Exception:
                prev_dom = None
            if prev_dom not in ('Folga', 'Trabalho'):
                prev_dom = ((estado_prev.get(ch, {}) or {}).get('ultimo_domingo_status', None))

        expected = 'Trabalho' if prev_dom == 'Folga' else 'Folga' if prev_dom == 'Trabalho' else None
        ent_base = _entrada_base(ch, df)
        manual_locked = set((locked_idx or {}).get(ch, set()))

        for i in sidx:
            atual = str(df.loc[i, 'Status'])
            if expected in ('Folga', 'Trabalho') and i not in manual_locked and atual != 'Férias':
                if expected == 'Folga':
                    df.loc[i, 'Status'] = 'Folga'
                    df.loc[i, 'H_Entrada'] = ''
                    df.loc[i, 'H_Saida'] = ''
                    atual = 'Folga'
                else:
                    df.loc[i, 'Status'] = 'Trabalho'
                    df.loc[i, 'H_Entrada'] = ent_base
                    df.loc[i, 'H_Saida'] = _saida_from_entrada(ent_base)
                    atual = 'Trabalho'

            if atual in WORK_STATUSES:
                expected = 'Folga'
            elif atual == 'Folga':
                expected = 'Trabalho'

        hist_all[ch] = df

    merged = {}
    for ch in set(list((hist_all or {}).keys()) + list((locked_idx or {}).keys())):
        merged[ch] = set((locked_idx or {}).get(ch, set())) | set(sunday_locked.get(ch, set()))
    return merged


def _rebuild_estado_out(hist_all: dict) -> dict:
    estado_out = {}
    for ch, df in (hist_all or {}).items():
        if df is None or len(df) == 0:
            estado_out[ch] = {'consec_trab_final': 0, 'ultima_saida': '', 'ultimo_domingo_status': None}
            continue
        consec = 0
        for i in range(len(df) - 1, -1, -1):
            if df.loc[i, 'Status'] in WORK_STATUSES:
                consec += 1
            else:
                break
        ultima_saida = ''
        for i in range(len(df) - 1, -1, -1):
            if df.loc[i, 'Status'] in WORK_STATUSES and (df.loc[i, 'H_Saida'] or ''):
                ultima_saida = df.loc[i, 'H_Saida']
                break
        ultimo_dom = None
        for i in range(len(df) - 1, -1, -1):
            if str(df.loc[i, 'Dia']) == 'dom':
                if df.loc[i, 'Status'] == 'Folga':
                    ultimo_dom = 'Folga'
                    break
                if df.loc[i, 'Status'] in WORK_STATUSES:
                    ultimo_dom = 'Trabalho'
                    break
        estado_out[ch] = {'consec_trab_final': consec, 'ultima_saida': ultima_saida, 'ultimo_domingo_status': ultimo_dom}
    return estado_out


def _apply_pdf_import_to_db(
    setor_destino: str,
    ano: int,
    mes: int,
    items: list[dict],
    criar_colabs: bool = True,
    limpar_mes_antes: bool = False,
    map_afa_para_folga: bool = False,
    cadastrar_ferias: bool = True,
):
    if limpar_mes_antes:
        con = db_conn()
        cur = con.cursor()
        cur.execute("DELETE FROM overrides WHERE setor=? AND ano=? AND mes=?", (setor_destino, int(ano), int(mes)))
        con.commit()
        con.close()

    resolvidos_por_nome = 0
    gerados_sem_chapa = []

    for it in items:
        nome = (it.get("nome") or "").strip()
        chapa = (it.get("chapa") or "").strip()
        entrada_tokens = list(it.get("tokens") or [])
        saida_tokens = list(it.get("saida_tokens") or [])

        if not chapa and nome:
            chapa = _find_chapa_by_name_in_colaboradores(setor_destino, nome)
            if chapa:
                it["chapa"] = chapa
                resolvidos_por_nome += 1

        if not chapa:
            chapa = _generate_fallback_pdf_chapa(setor_destino, nome, int(ano), int(mes))
            it["chapa"] = chapa
            gerados_sem_chapa.append(f"{nome or '(sem nome)'} -> {chapa}")

        if criar_colabs:
            upsert_colaborador_nome(setor_destino, chapa, nome)

        ferias_days = []
        for dia, tok in enumerate(entrada_tokens, start=1):
            tok = (tok or "").upper()
            saida = (saida_tokens[dia - 1] if dia - 1 < len(saida_tokens) else "") or ""
            saida = str(saida).strip().upper()

            if tok == "FOLG":
                set_override(setor_destino, ano, mes, chapa, dia, "Status", "Folga")
                delete_override(setor_destino, ano, mes, chapa, dia, "H_Entrada")
                delete_override(setor_destino, ano, mes, chapa, dia, "H_Saida")
            elif tok == "FER":
                ferias_days.append(dia)
                set_override(setor_destino, ano, mes, chapa, dia, "Status", "Férias")
                delete_override(setor_destino, ano, mes, chapa, dia, "H_Entrada")
                delete_override(setor_destino, ano, mes, chapa, dia, "H_Saida")
            elif tok == "AFA":
                set_override(setor_destino, ano, mes, chapa, dia, "Status", "Folga" if bool(map_afa_para_folga) else "Afastamento")
                delete_override(setor_destino, ano, mes, chapa, dia, "H_Entrada")
                delete_override(setor_destino, ano, mes, chapa, dia, "H_Saida")
            elif re.match(r"^\d{2}:\d{2}$", tok):
                set_override(setor_destino, ano, mes, chapa, dia, "Status", "Trabalho")
                set_override(setor_destino, ano, mes, chapa, dia, "H_Entrada", tok)
                if re.match(r"^\d{2}:\d{2}$", saida):
                    set_override(setor_destino, ano, mes, chapa, dia, "H_Saida", saida)
                else:
                    set_override(setor_destino, ano, mes, chapa, dia, "H_Saida", _saida_from_entrada(tok))

        if cadastrar_ferias and ferias_days:
            for a, b in _group_consecutive_days(ferias_days):
                add_ferias(setor_destino, chapa, date(int(ano), int(mes), int(a)), date(int(ano), int(mes), int(b)))

    try:
        st.cache_data.clear()
    except Exception:
        pass
    try:
        if resolvidos_por_nome:
            st.info(f"Importação PDF: {resolvidos_por_nome} colaborador(es) tiveram a chapa localizada automaticamente pelo nome.")
        if gerados_sem_chapa:
            st.warning("Importação PDF: chapa automática criada para: " + "; ".join(gerados_sem_chapa[:12]) + (" ..." if len(gerados_sem_chapa) > 12 else ""))
    except Exception:
        pass

def _build_hist_from_pdf_items(setor: str, ano: int, mes: int, items: list[dict], map_afa_para_folga: bool = False) -> tuple[dict, dict]:
    """Monta a escala do mês exatamente como veio no PDF.
    No mês importado, o PDF é a fonte da verdade.
    """
    datas = _dias_mes(int(ano), int(mes))
    df_ref = pd.DataFrame({'Data': datas, 'Dia': [D_PT[d.day_name()] for d in datas]})
    hist = {}
    estado = {}

    for it in (items or []):
        chapa = str(it.get('chapa') or '').strip()
        if not chapa:
            continue
        ent_tokens = list(it.get('tokens') or [])
        saida_tokens = list(it.get('saida_tokens') or [])

        df = df_ref.copy()
        df['Data'] = pd.to_datetime(df['Data'], errors='coerce')
        df['Status'] = 'Trabalho'
        df['H_Entrada'] = ''
        df['H_Saida'] = ''

        ndays = len(df)
        if len(ent_tokens) < ndays:
            ent_tokens += [''] * (ndays - len(ent_tokens))
        else:
            ent_tokens = ent_tokens[:ndays]
        if len(saida_tokens) < ndays:
            saida_tokens += [''] * (ndays - len(saida_tokens))
        else:
            saida_tokens = saida_tokens[:ndays]

        for i in range(ndays):
            ent = str(ent_tokens[i] or '').strip().upper()
            sai = str(saida_tokens[i] or '').strip().upper()

            if ent == 'FOLG':
                df.loc[i, 'Status'] = 'Folga'
                df.loc[i, 'H_Entrada'] = ''
                df.loc[i, 'H_Saida'] = ''
            elif ent == 'FER':
                df.loc[i, 'Status'] = 'Férias'
                df.loc[i, 'H_Entrada'] = ''
                df.loc[i, 'H_Saida'] = ''
            elif ent == 'AFA':
                df.loc[i, 'Status'] = 'Folga' if bool(map_afa_para_folga) else 'Afastamento'
                df.loc[i, 'H_Entrada'] = ''
                df.loc[i, 'H_Saida'] = ''
            elif re.match(r'^\d{2}:\d{2}$', ent):
                df.loc[i, 'Status'] = 'Trabalho'
                df.loc[i, 'H_Entrada'] = ent
                df.loc[i, 'H_Saida'] = sai if re.match(r'^\d{2}:\d{2}$', sai) else _saida_from_entrada(ent)
            else:
                # Nunca salva vazio no mês importado: assume folga técnica se o token falhou
                df.loc[i, 'Status'] = 'Folga'
                df.loc[i, 'H_Entrada'] = ''
                df.loc[i, 'H_Saida'] = ''

        hist[chapa] = df

        consec = 0
        for j in range(len(df) - 1, -1, -1):
            if df.loc[j, 'Status'] in WORK_STATUSES:
                consec += 1
            else:
                break

        ultima_saida = ''
        for j in range(len(df) - 1, -1, -1):
            if df.loc[j, 'Status'] in WORK_STATUSES and (df.loc[j, 'H_Saida'] or ''):
                ultima_saida = str(df.loc[j, 'H_Saida'] or '')
                break

        ultimo_dom = None
        doms = [j for j in range(len(df)) if str(df.loc[j, 'Dia']) == 'dom']
        if doms:
            ultimo_dom = str(df.loc[doms[-1], 'Status'])

        estado[chapa] = {
            'consec_trab_final': int(consec),
            'ultima_saida': ultima_saida,
            'ultimo_domingo_status': ultimo_dom,
        }

    return hist, estado


def _extract_nome_chapa_from_header_text(header_text: str) -> tuple[str, str]:
    s = _norm_pdf_text(header_text or "")
    s = re.sub(r"\s+", " ", s).strip()
    m = re.search(r"([A-ZÁÉÍÓÚÃÕÇ ]+?)(?:\s*\(([\d\.]+)\))?\s*M[eê]s\s*:\s*(\d{2})/(\d{4})", s, flags=re.IGNORECASE)
    if not m:
        return "", ""
    nome_full = (m.group(1) or "").strip()
    chapa = (m.group(2) or "").strip()
    # remove cargo eventualmente grudado antes do nome; pega a última sequência longa em caixa alta
    parts = [p.strip() for p in re.split(r"\s{2,}", nome_full) if p.strip()]
    if parts:
        nome = parts[-1]
    else:
        nome = nome_full
    nome = re.sub(r"\s+", " ", nome).strip()
    return nome, chapa

def _merge_pdf_table_row_values(table_rows: list[list[str]], row_idx: int, day_cols: list[int]) -> list[str]:
    vals = []
    base_row = table_rows[row_idx]
    overlay_row = table_rows[row_idx + 1] if row_idx + 1 < len(table_rows) else None
    overlay_ok = False
    if overlay_row is not None:
        first = overlay_row[0] if len(overlay_row) > 0 else ""
        overlay_ok = not str(first or "").strip()
    for c in day_cols:
        base = base_row[c] if c < len(base_row) else ""
        base = "" if base is None else str(base).strip()
        if overlay_ok:
            ov = overlay_row[c] if c < len(overlay_row) else ""
            ov = "" if ov is None else str(ov).strip()
            if ov and not base:
                base = ov
        vals.append(str(base).strip())
    return [str(x or "").strip().upper() for x in vals]

def _parse_escala_ponto_new_pdf_bytes(pdf_bytes: bytes):
    """
    Parser principal para o layout ESCALA_PONTO_NEW usando extração tabular via pdfplumber.
    Lê o quadro completo do colaborador e usa apenas a 1ª linha Entrada + linha Saída final.
    """
    try:
        import io
        import pdfplumber
    except Exception as e:
        return None, None, [], [f"Biblioteca de leitura tabular do PDF indisponível: {e}"]

    items = []
    erros = []
    ano = mes = None

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables() or []
                for table in tables:
                    rows = [[("" if c is None else str(c).strip()) for c in row] for row in (table or [])]
                    if not rows:
                        continue
                    idx_data = next((i for i, row in enumerate(rows) if row and str(row[0]).strip() == "Data / Dia"), None)
                    if idx_data is None:
                        continue

                    header_text = " ".join(" ".join(r) for r in rows[:idx_data])
                    nome, chapa = _extract_nome_chapa_from_header_text(header_text)
                    if not nome:
                        continue

                    m_mes = re.search(r"M[eê]s\s*:\s*(\d{2})/(\d{4})", header_text, flags=re.IGNORECASE)
                    if m_mes and (ano is None or mes is None):
                        mes = int(m_mes.group(1))
                        ano = int(m_mes.group(2))

                    day_cols = []
                    day_nums = []
                    for ci, cell in enumerate(rows[idx_data]):
                        cell = str(cell or "").strip()
                        if re.fullmatch(r"\d{1,2}", cell):
                            day_cols.append(ci)
                            day_nums.append(int(cell))
                    if not day_nums:
                        erros.append(f"Página {page_num}: tabela de {nome} sem colunas de dia identificadas.")
                        continue

                    label_map = {}
                    for ridx in range(idx_data + 1, len(rows)):
                        label = str(rows[ridx][0] if rows[ridx] else "").strip().lower()
                        if label.startswith("entrada"):
                            label_map.setdefault("entrada", []).append(ridx)
                        elif label.startswith("saída refeição") or label.startswith("saida refeição") or label.startswith("saida refeicao") or label.startswith("saída refeicao"):
                            label_map.setdefault("saida_refeicao", []).append(ridx)
                        elif label.startswith("saída") or label.startswith("saida"):
                            label_map.setdefault("saida", []).append(ridx)
                        elif label.startswith("horas trab"):
                            label_map.setdefault("horas", []).append(ridx)

                    if not label_map.get("entrada") or not label_map.get("saida"):
                        erros.append(f"Funcionário {nome}: não consegui localizar as linhas de Entrada/Saída na tabela da página {page_num}.")
                        continue

                    ent_tokens = _merge_pdf_table_row_values(rows, label_map["entrada"][0], day_cols)
                    saida_tokens = _merge_pdf_table_row_values(rows, label_map["saida"][0], day_cols)

                    if ano is None or mes is None:
                        # fallback defensivo
                        try:
                            pg_txt = page.extract_text() or ""
                            ano, mes = _detect_mes_ano_from_text(pg_txt)
                        except Exception:
                            pass

                    if ano is None or mes is None:
                        erros.append(f"Funcionário {nome}: mês/ano não detectado.")
                        continue

                    ndays = calendar.monthrange(int(ano), int(mes))[1]

                    # Remonta exatamente da lista de dias da própria tabela
                    ent_by_day = {int(d): str(v or "").strip().upper() for d, v in zip(day_nums, ent_tokens)}
                    sai_by_day = {int(d): str(v or "").strip().upper() for d, v in zip(day_nums, saida_tokens)}
                    full_ent = [ent_by_day.get(d, "") for d in range(1, ndays + 1)]
                    full_sai = [sai_by_day.get(d, "") for d in range(1, ndays + 1)]

                    blanks = sum(1 for x in full_ent if not str(x or "").strip())
                    if blanks:
                        erros.append(f"Funcionário {nome}: {blanks} dia(s) ficaram sem valor de Entrada após leitura tabular.")

                    items.append({
                        "nome": nome,
                        "chapa": chapa,
                        "tokens": full_ent,
                        "saida_tokens": full_sai,
                        "raw_rows": {
                            "entrada": full_ent,
                            "saida": full_sai,
                            "day_nums": day_nums,
                            "page": page_num,
                        },
                    })
    except Exception as e:
        return None, None, [], [f"Falha na leitura tabular do PDF: {e}"]

    if ano is None or mes is None:
        return None, None, [], ["Não consegui detectar Mês: MM/AAAA no PDF."]

    # Se o pdfplumber falhar em parte, os itens ainda assim podem ser válidos; não bloqueia aqui.
    return int(ano), int(mes), items, erros

def _parse_escala_ponto_new_pdf_text(extracted_text: str):
    """
    Fallback legível por texto corrido. Mantido como reserva caso a extração tabular falhe.
    """
    t = _norm_pdf_text(extracted_text)
    ano, mes = _detect_mes_ano_from_text(t)
    if not ano or not mes:
        return None, None, [], ["Não consegui detectar 'Mês: MM/AAAA' no texto extraído."]
    ndays = calendar.monthrange(int(ano), int(mes))[1]

    blocks = _split_employee_blocks_ponto_new(t)
    if not blocks:
        return int(ano), int(mes), [], ["Não consegui identificar blocos de funcionário (Nome + Mês:)."]

    items = []
    erros = []
    for b in blocks:
        rows = _extract_pdf_block_rows(b["texto"], ndays)
        ent = list(rows.get("entrada") or [])
        saida = list(rows.get("saida") or [])

        if len(saida) < len(ent):
            saida += [''] * (len(ent) - len(saida))
        if len(ent) != ndays:
            erros.append(f"Funcionário {b['nome']}: esperado {ndays} valores de Entrada, li {len(ent)}.")
        items.append({
            "nome": b["nome"],
            "chapa": b["chapa"],
            "tokens": ent,
            "saida_tokens": saida[:ndays],
            "raw_rows": rows,
        })

    return int(ano), int(mes), items, erros



# =========================================================
# UI THEME (CSS) — só visual
# =========================================================
st.markdown("""
<style>
/* layout geral */
.block-container { padding-top: .6rem; padding-bottom: 2rem; max-width: 1600px; }
h1, h2, h3 { letter-spacing: -0.2px; }

/* KPI cards (topo) */
.kpi-card{
  border: 1px solid rgba(255,255,255,0.10);
  border-radius: 16px;
  padding: 12px 14px;
  background: rgba(255,255,255,0.06);
  box-shadow: 0 6px 18px rgba(0,0,0,0.18);
  backdrop-filter: blur(6px);
}
.kpi-card:hover{ transform: translateY(-1px); transition: 120ms ease; border-color: rgba(255,255,255,0.18); }
.kpi-title{ font-size: .78rem; opacity: .72; margin: 0 0 4px 0; text-transform: uppercase; letter-spacing: .4px; }
.kpi-value{ font-size: 1.35rem; font-weight: 800; margin: 0; line-height: 1.05; }

/* divisória */
.hr{ height:1px; background: rgba(255,255,255,0.08); margin: 14px 0; }

/* Tabs (menu superior) */
div[data-testid="stTabs"] { margin-top: .25rem; }
div[data-testid="stTabs"] button {
  font-size: .92rem;
  padding: 10px 14px;
  border-radius: 12px;
}
div[data-testid="stTabs"] button[aria-selected="true"]{
  background: rgba(255,255,255,0.07);
  border-bottom: 2px solid rgba(255,255,255,0.35);
}
div[data-testid="stTabs"] button:hover{
  background: rgba(255,255,255,0.06);
}

/* sidebar mais limpa */
section[data-testid="stSidebar"] .block-container { padding-top: 1rem; }

/* dataframe: arredondar */
div[data-testid="stDataFrame"] { border-radius: 12px; overflow: hidden; }
</style>
""", unsafe_allow_html=True)


# =========================
# ADMIN: Backup / Restore + Setores + Import
# =========================
APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
BACKUP_DIR = str(DATA_DIR / "backups")
DB_PATH = str(DATA_DIR / "escala.db")
ROOT_LEGACY_DB_PATH = str(APP_DIR / "escala.db")
BUNDLED_LATEST_STABLE_CANDIDATES = [
    APP_DIR / "latest_stable.db",
    APP_DIR / "data" / "latest_stable.db",
    APP_DIR / "backups" / "latest_stable.db",
    APP_DIR.parent / "latest_stable.db",
]
AUTO_BACKUP_HOUR = 3  # 03:00
AUTO_BACKUP_INTERVAL_HOURS = 6  # roda quando o app abre
ROTATING_BACKUP_KEEP = int((os.getenv("ROTATING_BACKUP_KEEP", "12") or "12").strip() or 12)
ROTATING_BACKUP_PREFIX = "rolling"

# =========================================================
# SUPABASE SYNC (mantém a lógica do app em SQLite local,
# mas faz persistência remota para sobreviver a reboot)
# =========================================================
SUPABASE_URL = (os.getenv("SUPABASE_URL") or "").strip().rstrip("/")
SUPABASE_KEY = (
    os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    or os.getenv("SUPABASE_KEY")
    or os.getenv("SUPABASE_ANON_KEY")
    or ""
).strip()
SUPABASE_SCHEMA = (os.getenv("SUPABASE_SCHEMA") or "public").strip() or "public"
SUPABASE_SYNC_ENABLED = bool(SUPABASE_URL and SUPABASE_KEY)
SUPABASE_SYNC_DEBOUNCE_SEC = int(os.getenv("SUPABASE_SYNC_DEBOUNCE_SEC", "12") or 12)
SUPABASE_AUTO_PULL_ON_START = (os.getenv("SUPABASE_AUTO_PULL_ON_START", "0") or "0").strip() in ("1", "true", "True", "yes", "on")
SUPABASE_AUTO_PUSH_ON_COMMIT = (os.getenv("SUPABASE_AUTO_PUSH_ON_COMMIT", "1") or "1").strip() in ("1", "true", "True", "yes", "on")
SUPABASE_AUTO_PUSH_ON_CLOSE = (os.getenv("SUPABASE_AUTO_PUSH_ON_CLOSE", "0") or "0").strip() in ("1", "true", "True", "yes", "on")
SUPABASE_ASYNC_PUSH_ENABLED = (os.getenv("SUPABASE_ASYNC_PUSH_ENABLED", "1") or "1").strip() in ("1", "true", "True", "yes", "on")
SUPABASE_ASYNC_PUSH_DELAY_SEC = float((os.getenv("SUPABASE_ASYNC_PUSH_DELAY_SEC", "2.0") or "2.0").strip())
SUPABASE_AUTO_BOOTSTRAP_AFTER_SCHEMA = (os.getenv("SUPABASE_AUTO_BOOTSTRAP_AFTER_SCHEMA", "0") or "0").strip() in ("1", "true", "True", "yes", "on")
SUPABASE_AUTO_RESTORE_IF_LOCAL_EMPTY = (os.getenv("SUPABASE_AUTO_RESTORE_IF_LOCAL_EMPTY", "1") or "1").strip() in ("1", "true", "True", "yes", "on")
FAST_BOOT_SKIP_STARTUP_AUTO_BACKUP = (os.getenv("FAST_BOOT_SKIP_STARTUP_AUTO_BACKUP", "1") or "1").strip() in ("1", "true", "True", "yes", "on")
QUICK_LOGIN_BOOT = (os.getenv("QUICK_LOGIN_BOOT", "1") or "1").strip() in ("1", "true", "True", "yes", "on")
FAST_BACKUP_DISABLE_ROLLING_ON_COMMIT = (os.getenv("FAST_BACKUP_DISABLE_ROLLING_ON_COMMIT", "1") or "1").strip() in ("1", "true", "True", "yes", "on")
FAST_SNAPSHOT_THROTTLE_SECONDS = int((os.getenv("FAST_SNAPSHOT_THROTTLE_SECONDS", "300") or "300").strip())
FAST_SNAPSHOT_SKIP_ON_CLOSE = (os.getenv("FAST_SNAPSHOT_SKIP_ON_CLOSE", "1") or "1").strip() in ("1", "true", "True", "yes", "on")
_LAST_SNAPSHOT_TS = 0.0
_LAST_SNAPSHOT_DB_MTIME = 0.0
RESTORE_GUARD_ENABLED = (os.getenv("RESTORE_GUARD_ENABLED", "0") or "0").strip() in ("1", "true", "True", "yes", "on")
_RESTORE_GUARD_ACTIVE = False
_RESTORE_GUARD_MESSAGE = ""

_SUPABASE_LAST_PUSH_TS = 0.0
_SUPABASE_LAST_PULL_TS = 0.0
_SUPABASE_SYNC_IN_PROGRESS = False
_SUPABASE_LAST_ERROR = ""
_SUPABASE_ASYNC_THREAD = None
_SUPABASE_ASYNC_LOCK = threading.Lock()
_SUPABASE_ASYNC_PENDING = False
_SUPABASE_ASYNC_FORCE = False
_SUPABASE_ASYNC_LAST_REQUEST_TS = 0.0
_SUPABASE_BOOTSTRAP_DONE = False
_SUPABASE_SYNC_STARTED_AT = 0.0
SUPABASE_SYNC_LOCK_TIMEOUT_SEC = int((os.getenv("SUPABASE_SYNC_LOCK_TIMEOUT_SEC", "45") or "45").strip())
SUPABASE_TABLE_MAP = {}
try:
    _raw_table_map = (os.getenv("SUPABASE_TABLE_MAP") or "").strip()
    if _raw_table_map:
        try:
            _tmp_map = json.loads(_raw_table_map)
            if isinstance(_tmp_map, dict):
                SUPABASE_TABLE_MAP = {str(k).strip(): str(v).strip() for k, v in _tmp_map.items() if str(k).strip() and str(v).strip()}
        except Exception:
            for _part in _raw_table_map.split(","):
                if "=" in _part:
                    _k, _v = _part.split("=", 1)
                    _k = str(_k).strip()
                    _v = str(_v).strip()
                    if _k and _v:
                        SUPABASE_TABLE_MAP[_k] = _v
except Exception:
    SUPABASE_TABLE_MAP = {}
_SUPABASE_TABLE_CACHE = {}
_SUPABASE_IGNORED_COLUMNS = {}
_SUPABASE_ALWAYS_IGNORED_COLUMNS = {"id", "created_at", "updated_at", "criado_em", "alterado_em"}

def _supabase_headers(prefer: str | None = None, extra: dict | None = None) -> dict:
    h = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Accept-Profile": SUPABASE_SCHEMA,
        "Content-Profile": SUPABASE_SCHEMA,
    }
    if prefer:
        h["Prefer"] = prefer
    if extra:
        h.update(extra)
    return h

def _supabase_table_url(table: str) -> str:
    return f"{SUPABASE_URL}/rest/v1/{table}"


def _supabase_extract_table_hint(message: str) -> str:
    s = str(message or "")
    m = re.search(r"table '\s*(?:[A-Za-z0-9_]+\.)?([A-Za-z0-9_]+)\s*'", s)
    return str(m.group(1) or "").strip() if m else ""


def _supabase_extract_missing_column(message: str) -> str:
    s = str(message or "")
    m = re.search(r"Could not find the '([A-Za-z0-9_]+)' column", s)
    return str(m.group(1) or "").strip() if m else ""


def _supabase_extract_identity_column(message: str) -> str:
    s = str(message or "")
    for pat in [
        r'cannot insert a non-DEFAULT value into column "([A-Za-z0-9_]+)"',
        r'Column "([A-Za-z0-9_]+)" is an identity column',
        r'GENERATED ALWAYS',
    ]:
        m = re.search(pat, s)
        if m and m.groups():
            return str(m.group(1) or "").strip()
    if 'GENERATED ALWAYS' in s and '"id"' in s:
        return 'id'
    return ""


def _supabase_register_ignored_column(table: str, column: str) -> None:
    table = str(table or "").strip()
    column = str(column or "").strip()
    if not table or not column:
        return
    cols = set(_SUPABASE_IGNORED_COLUMNS.get(table, set()))
    cols.add(column)
    _SUPABASE_IGNORED_COLUMNS[table] = cols


def _supabase_filtered_rows_and_conflicts(table: str, rows: list[dict], conflict_cols: list[str]):
    ignored = set(_SUPABASE_ALWAYS_IGNORED_COLUMNS) | set(_SUPABASE_IGNORED_COLUMNS.get(str(table or "").strip(), set()))
    if not ignored:
        return rows, list(conflict_cols or [])
    filtered_rows = []
    for row in rows or []:
        filtered = {k: v for k, v in dict(row or {}).items() if k not in ignored}
        filtered_rows.append(filtered)
    filtered_conflicts = [c for c in (conflict_cols or []) if c not in ignored]
    return filtered_rows, filtered_conflicts


def _supabase_candidate_tables(local_table: str) -> list[str]:
    local_table = str(local_table or "").strip()
    candidates = []
    for item in [
        str(SUPABASE_TABLE_MAP.get(local_table, "") or "").strip(),
        local_table,
        f"escala_{local_table}",
        f"app_{local_table}",
        f"tb_{local_table}",
        f"tbl_{local_table}",
    ]:
        if item and item not in candidates:
            candidates.append(item)
    return candidates


def _supabase_resolve_remote_table(local_table: str, refresh: bool = False) -> str:
    local_table = str(local_table or "").strip()
    if not local_table:
        return local_table
    if (not refresh) and local_table in _SUPABASE_TABLE_CACHE:
        return _SUPABASE_TABLE_CACHE[local_table]
    resolved = str(SUPABASE_TABLE_MAP.get(local_table, local_table) or local_table).strip()
    if not SUPABASE_SYNC_ENABLED:
        _SUPABASE_TABLE_CACHE[local_table] = resolved
        return resolved
    for candidate in _supabase_candidate_tables(local_table):
        try:
            r = requests.get(
                _supabase_table_url(candidate),
                params={"select": "*", "limit": 1},
                headers=_supabase_headers(),
                timeout=12,
            )
            if r.status_code < 400:
                resolved = candidate
                break
            hint = _supabase_extract_table_hint(r.text)
            if hint:
                resolved = hint
                break
        except Exception:
            continue
    _SUPABASE_TABLE_CACHE[local_table] = resolved
    return resolved

def _app_db_connect(db_path: str | None = None):
    target = str(db_path or DB_PATH)
    factory = SQLiteSyncConnection if str(Path(target).resolve()) == str(Path(DB_PATH).resolve()) else sqlite3.Connection
    conn = sqlite3.connect(target, check_same_thread=False, factory=factory)
    try:
        conn.execute("PRAGMA foreign_keys=ON")
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=10000")
        conn.execute("PRAGMA synchronous=NORMAL")
    except Exception:
        pass
    return conn

def _sqlite_user_tables(conn) -> list[str]:
    try:
        rows = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name").fetchall()
        return [str(r[0]) for r in rows if r and r[0]]
    except Exception:
        return []

def _sqlite_table_columns(conn, table: str) -> list[str]:
    try:
        rows = conn.execute(f'PRAGMA table_info("{table}")').fetchall()
        return [str(r[1]) for r in rows if len(r) > 1]
    except Exception:
        return []

def _sqlite_conflict_cols(conn, table: str) -> list[str]:
    try:
        rows = conn.execute(f'PRAGMA table_info("{table}")').fetchall()
        pk_cols = [str(r[1]) for r in rows if len(r) > 5 and int(r[5] or 0) > 0]
        if pk_cols:
            return pk_cols
        idx_rows = conn.execute(f'PRAGMA index_list("{table}")').fetchall()
        for idx in idx_rows:
            is_unique = int(idx[2]) == 1 if len(idx) > 2 else False
            idx_name = str(idx[1]) if len(idx) > 1 else ''
            if not is_unique or not idx_name:
                continue
            info = conn.execute(f'PRAGMA index_info("{idx_name}")').fetchall()
            cols = [str(r[2]) for r in info if len(r) > 2]
            if cols:
                return cols
    except Exception:
        pass
    return []

def _sqlite_table_rowcount(conn, table: str) -> int:
    try:
        row = conn.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()
        return int(row[0]) if row else 0
    except Exception:
        return 0

def _sqlite_app_has_meaningful_data() -> bool:
    try:
        conn = _app_db_connect(DB_PATH)
        try:
            core_tables = ("colaboradores", "escala_mes", "overrides", "ferias", "estado_mes_anterior", "subgrupos_setor", "subgrupo_regras")
            existing = set(_sqlite_user_tables(conn))
            for tb in core_tables:
                if tb in existing and _sqlite_table_rowcount(conn, tb) > 0:
                    return True
            if "usuarios_sistema" in existing:
                row = conn.execute("""
                    SELECT COUNT(*)
                    FROM usuarios_sistema
                    WHERE NOT (UPPER(TRIM(setor))='ADMIN' AND TRIM(chapa)='admin')
                """).fetchone()
                if row and int(row[0] or 0) > 0:
                    return True
        finally:
            conn.close()
    except Exception:
        pass
    return False

def _jsonable(v):
    if isinstance(v, (datetime, date, pd.Timestamp)):
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    if pd.isna(v):
        return None
    if isinstance(v, bytes):
        try:
            return v.decode('utf-8')
        except Exception:
            return None
    return v

def _set_supabase_error(msg: str = "") -> None:
    global _SUPABASE_LAST_ERROR
    try:
        _SUPABASE_LAST_ERROR = str(msg or "").strip()
    except Exception:
        _SUPABASE_LAST_ERROR = ""
_SUPABASE_ASYNC_THREAD = None
_SUPABASE_ASYNC_LOCK = threading.Lock()
_SUPABASE_ASYNC_PENDING = False
_SUPABASE_ASYNC_FORCE = False
_SUPABASE_ASYNC_LAST_REQUEST_TS = 0.0


def _set_restore_guard(active: bool, message: str = "") -> None:
    global _RESTORE_GUARD_ACTIVE, _RESTORE_GUARD_MESSAGE
    try:
        _RESTORE_GUARD_ACTIVE = bool(active)
        _RESTORE_GUARD_MESSAGE = str(message or "").strip()
    except Exception:
        _RESTORE_GUARD_ACTIVE = bool(active)
        _RESTORE_GUARD_MESSAGE = ""


def _restore_sources_summary() -> str:
    fontes = []
    try:
        if SUPABASE_SYNC_ENABLED:
            fontes.append("Supabase")
    except Exception:
        pass
    try:
        if any(_validate_sqlite_file(str(p)) for p in _bundled_latest_stable_paths()):
            fontes.append("latest_stable.db do projeto")
    except Exception:
        pass
    try:
        latest = Path(BACKUP_DIR) / "latest_stable.db"
        if _validate_sqlite_file(str(latest)) and "latest_stable.db do projeto" not in fontes:
            fontes.append("latest_stable.db local")
    except Exception:
        pass
    return ", ".join(fontes)


def _should_block_silent_empty_seed() -> bool:
    if not RESTORE_GUARD_ENABLED:
        return False
    try:
        if _sqlite_app_has_meaningful_data():
            return False
    except Exception:
        pass
    fontes = _restore_sources_summary()
    return bool(fontes)

def _supabase_table_count_fast(table: str) -> int | None:
    if not SUPABASE_SYNC_ENABLED:
        return None
    try:
        r = requests.get(
            _supabase_table_url(_supabase_resolve_remote_table(table)),
            params={"select": "id", "limit": 1},
            headers=_supabase_headers(extra={"Prefer": "count=exact", "Range": "0-0"}),
            timeout=20,
        )
        if r.status_code >= 400:
            return None
        content_range = str(r.headers.get("Content-Range") or "")
        if "/" in content_range:
            tail = content_range.rsplit("/", 1)[-1].strip()
            if tail.isdigit():
                return int(tail)
    except Exception:
        return None
    return None

def _supabase_remote_has_meaningful_data() -> bool:
    if not SUPABASE_SYNC_ENABLED:
        return False
    for tb in ("colaboradores", "escala_mes", "overrides", "ferias", "estado_mes_anterior", "subgrupos_setor", "subgrupo_regras"):
        cnt = _supabase_table_count_fast(tb)
        if cnt is not None and cnt > 0:
            return True
    users_cnt = _supabase_table_count_fast("usuarios_sistema")
    if users_cnt is not None and users_cnt > 1:
        return True
    return False

def _rotate_backup_files_safely() -> None:
    try:
        keep = max(1, int(ROTATING_BACKUP_KEEP or 12))
    except Exception:
        keep = 12
    try:
        folder = Path(BACKUP_DIR)
        files = sorted(folder.glob(f"{ROTATING_BACKUP_PREFIX}_*.db"), key=lambda p: p.stat().st_mtime, reverse=True)
        for p in files[keep:]:
            try:
                p.unlink(missing_ok=True)
            except Exception:
                pass
    except Exception:
        pass


def _save_latest_stable_snapshot_safely(include_rolling: bool = False) -> None:
    try:
        current = Path(DB_PATH)
        if not current.exists() or not _validate_sqlite_file(str(current)):
            return

        targets = [
            Path(APP_DIR) / "latest_stable.db",
            Path(BACKUP_DIR) / "latest_stable.db",
        ]

        if include_rolling and not FAST_BACKUP_DISABLE_ROLLING_ON_COMMIT:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            rolling = Path(BACKUP_DIR) / f"{ROTATING_BACKUP_PREFIX}_{ts}.db"
            targets.append(rolling)

        for target in targets:
            try:
                target.parent.mkdir(parents=True, exist_ok=True)
                try:
                    _sqlite_backup_copy(str(current), str(target))
                except Exception:
                    shutil.copy2(current, target)
            except Exception:
                pass

        if include_rolling and not FAST_BACKUP_DISABLE_ROLLING_ON_COMMIT:
            _rotate_backup_files_safely()
    except Exception:
        pass

def _bootstrap_from_supabase_after_schema(force: bool = False) -> bool:
    global _SUPABASE_BOOTSTRAP_DONE
    if _SUPABASE_BOOTSTRAP_DONE and not force:
        return False
    _SUPABASE_BOOTSTRAP_DONE = True
    try:
        if not SUPABASE_SYNC_ENABLED:
            return False
        local_has = _sqlite_app_has_meaningful_data()
        remote_has = _supabase_remote_has_meaningful_data()
        if (not local_has) and remote_has:
            ok = _supabase_pull_all_to_sqlite(force=True)
            if ok:
                _maybe_save_latest_stable_snapshot_fast("commit")
            return bool(ok)
        if local_has and not remote_has:
            ok = _supabase_push_all_from_sqlite(force=True)
            if ok:
                _maybe_save_latest_stable_snapshot_fast("close")
            return bool(ok)
    except Exception as e:
        _set_supabase_error(e)
    return False

def _restore_from_supabase_if_local_empty(force: bool = False) -> bool:
    try:
        if not SUPABASE_SYNC_ENABLED:
            return False
        if (not force) and _sqlite_app_has_meaningful_data():
            return False
        if not _supabase_remote_has_meaningful_data():
            return False
        ok = _supabase_pull_all_to_sqlite(force=True)
        if ok:
            _save_latest_stable_snapshot_safely()
        return bool(ok)
    except Exception as e:
        _set_supabase_error(e)
        return False

def _table_exists_in_supabase(table: str) -> bool:
    if not SUPABASE_SYNC_ENABLED:
        return False
    try:
        r = requests.get(
            _supabase_table_url(_supabase_resolve_remote_table(table, refresh=True)),
            params={"select": "*", "limit": 1},
            headers=_supabase_headers(),
            timeout=15,
        )
        return r.status_code < 400
    except Exception:
        return False

def _supabase_fetch_table_rows(table: str, page_size: int = 1000) -> list[dict]:
    if not SUPABASE_SYNC_ENABLED:
        return []
    out = []
    start = 0
    while True:
        headers = _supabase_headers(extra={"Range": f"{start}-{start + page_size - 1}"})
        r = requests.get(
            _supabase_table_url(_supabase_resolve_remote_table(table)),
            params={"select": "*", "order": "id.asc.nullslast"},
            headers=headers,
            timeout=30,
        )
        if r.status_code >= 400:
            raise RuntimeError(f"Supabase GET {table}: {r.status_code} {r.text[:200]}")
        data = r.json() or []
        if not data:
            break
        if isinstance(data, list):
            out.extend(data)
            if len(data) < page_size:
                break
            start += page_size
        else:
            break
    return out

def _supabase_upsert_rows(table: str, rows: list[dict], conflict_cols: list[str]) -> None:
    if not SUPABASE_SYNC_ENABLED or not rows:
        return
    batch_size = 500
    for i in range(0, len(rows), batch_size):
        base_batch = rows[i:i+batch_size]
        attempts = 0
        while True:
            batch, filtered_conflicts = _supabase_filtered_rows_and_conflicts(table, base_batch, conflict_cols)
            batch = [b for b in batch if isinstance(b, dict) and b]
            if not batch:
                break
            params = {}
            if filtered_conflicts:
                params["on_conflict"] = ",".join(filtered_conflicts)
            r = requests.post(
                _supabase_table_url(_supabase_resolve_remote_table(table)),
                params=params,
                headers=_supabase_headers(prefer="resolution=merge-duplicates,return=minimal"),
                data=json.dumps(batch, ensure_ascii=False),
                timeout=60,
            )
            if r.status_code < 400:
                break
            missing_col = _supabase_extract_missing_column(r.text)
            identity_col = _supabase_extract_identity_column(r.text)
            if r.status_code == 400 and missing_col and attempts < 20:
                _supabase_register_ignored_column(table, missing_col)
                attempts += 1
                continue
            if r.status_code == 400 and identity_col and attempts < 20:
                _supabase_register_ignored_column(table, identity_col)
                attempts += 1
                continue
            raise RuntimeError(f"Supabase UPSERT {table}: {r.status_code} {r.text[:300]}")

def _supabase_request_push_async(force: bool = False) -> bool:
    """Agenda push do SQLite para o Supabase em background, sem travar a UI."""
    global _SUPABASE_ASYNC_PENDING, _SUPABASE_ASYNC_FORCE, _SUPABASE_ASYNC_THREAD, _SUPABASE_ASYNC_LAST_REQUEST_TS
    if not SUPABASE_SYNC_ENABLED:
        return False
    if not SUPABASE_ASYNC_PUSH_ENABLED:
        return _supabase_push_all_from_sqlite(force=force)
    try:
        now = time.time()
    except Exception:
        now = 0.0
    with _SUPABASE_ASYNC_LOCK:
        _SUPABASE_ASYNC_PENDING = True
        _SUPABASE_ASYNC_FORCE = bool(_SUPABASE_ASYNC_FORCE or force)
        _SUPABASE_ASYNC_LAST_REQUEST_TS = now
        th = _SUPABASE_ASYNC_THREAD
        if th is not None and th.is_alive():
            return True

        def _worker():
            global _SUPABASE_ASYNC_PENDING, _SUPABASE_ASYNC_FORCE, _SUPABASE_ASYNC_THREAD
            while True:
                try:
                    time.sleep(max(0.2, float(SUPABASE_ASYNC_PUSH_DELAY_SEC)))
                except Exception:
                    time.sleep(0.5)
                with _SUPABASE_ASYNC_LOCK:
                    pending = bool(_SUPABASE_ASYNC_PENDING)
                    force_now = bool(_SUPABASE_ASYNC_FORCE)
                    age = time.time() - float(_SUPABASE_ASYNC_LAST_REQUEST_TS or 0.0)
                    if not pending:
                        _SUPABASE_ASYNC_THREAD = None
                        return
                    if age < max(0.2, float(SUPABASE_ASYNC_PUSH_DELAY_SEC)):
                        continue
                    _SUPABASE_ASYNC_PENDING = False
                    _SUPABASE_ASYNC_FORCE = False
                try:
                    _supabase_push_all_from_sqlite(force=force_now)
                except Exception as e:
                    _set_supabase_error(e)
                with _SUPABASE_ASYNC_LOCK:
                    if not _SUPABASE_ASYNC_PENDING:
                        _SUPABASE_ASYNC_THREAD = None
                        return

        _SUPABASE_ASYNC_THREAD = threading.Thread(target=_worker, daemon=True, name='supabase_async_push')
        _SUPABASE_ASYNC_THREAD.start()
        return True


def _supabase_delete_all_rows(table: str) -> None:
    if not SUPABASE_SYNC_ENABLED:
        return
    r = requests.delete(
        _supabase_table_url(_supabase_resolve_remote_table(table)),
        params={"id": "gt.0"},
        headers=_supabase_headers(prefer="return=minimal"),
        timeout=60,
    )
    if r.status_code >= 400 and r.status_code != 404:
        raise RuntimeError(f"Supabase DELETE {table}: {r.status_code} {r.text[:300]}")

def _supabase_begin_sync() -> bool:
    global _SUPABASE_SYNC_IN_PROGRESS, _SUPABASE_SYNC_STARTED_AT
    try:
        now = time.time()
    except Exception:
        now = 0.0
    if _SUPABASE_SYNC_IN_PROGRESS:
        try:
            age = now - float(_SUPABASE_SYNC_STARTED_AT or 0.0)
        except Exception:
            age = 0.0
        if age >= float(SUPABASE_SYNC_LOCK_TIMEOUT_SEC):
            _SUPABASE_SYNC_IN_PROGRESS = False
            _SUPABASE_SYNC_STARTED_AT = 0.0
        else:
            _set_supabase_error(f"Sincronização em andamento. Aguarde {max(1, int(SUPABASE_SYNC_LOCK_TIMEOUT_SEC - age))}s ou recarregue o app.")
            return False
    _SUPABASE_SYNC_IN_PROGRESS = True
    _SUPABASE_SYNC_STARTED_AT = now
    return True


def _supabase_end_sync() -> None:
    global _SUPABASE_SYNC_IN_PROGRESS, _SUPABASE_SYNC_STARTED_AT
    _SUPABASE_SYNC_IN_PROGRESS = False
    _SUPABASE_SYNC_STARTED_AT = 0.0

def _supabase_push_all_from_sqlite(force: bool = False) -> bool:
    global _SUPABASE_LAST_PUSH_TS
    if not SUPABASE_SYNC_ENABLED:
        _set_supabase_error("Sync Supabase desabilitado.")
        return False
    now = time.time()
    if (not force) and (now - float(_SUPABASE_LAST_PUSH_TS or 0.0) < float(SUPABASE_SYNC_DEBOUNCE_SEC)):
        return False
    if not _supabase_begin_sync():
        return False
    try:
        conn = _app_db_connect(DB_PATH)
        try:
            tables = _sqlite_user_tables(conn)
            for table in tables:
                if table in {"sqlite_sequence"}:
                    continue
                cols = _sqlite_table_columns(conn, table)
                if not cols or not _table_exists_in_supabase(table):
                    continue
                rows = conn.execute(f'SELECT * FROM "{table}"').fetchall()
                payload = []
                for row in rows:
                    item = {}
                    for c, v in zip(cols, row):
                        item[c] = _jsonable(v)
                    payload.append(item)
                conflict = _sqlite_conflict_cols(conn, table)
                # tabela vazia local: não apaga remoto automaticamente
                if payload:
                    _supabase_upsert_rows(table, payload, conflict)
            _SUPABASE_LAST_PUSH_TS = now
            _set_supabase_error("")
            _save_latest_stable_snapshot_safely()
            return True
        finally:
            conn.close()
    except Exception as e:
        _set_supabase_error(e)
        try:
            print(f"[SUPABASE SYNC PUSH] {e}")
        except Exception:
            pass
        return False
    finally:
        _supabase_end_sync()

def _supabase_pull_all_to_sqlite(force: bool = False) -> bool:
    global _SUPABASE_LAST_PULL_TS
    if not SUPABASE_SYNC_ENABLED:
        _set_supabase_error("Sync Supabase desabilitado.")
        return False
    now = time.time()
    if (not force) and _sqlite_app_has_meaningful_data():
        return False
    if not _supabase_begin_sync():
        return False
    try:
        conn = _app_db_connect(DB_PATH)
        try:
            tables = [t for t in _sqlite_user_tables(conn) if t != 'sqlite_sequence']
            pulled_any = False
            for table in tables:
                cols = _sqlite_table_columns(conn, table)
                if not cols or not _table_exists_in_supabase(table):
                    continue
                remote_rows = _supabase_fetch_table_rows(table)
                if not remote_rows:
                    continue
                placeholders = ','.join(['?'] * len(cols))
                col_sql = ','.join([f'"{c}"' for c in cols])
                conn.execute(f'DELETE FROM "{table}"')
                for item in remote_rows:
                    vals = [item.get(c) for c in cols]
                    conn.execute(f'INSERT OR REPLACE INTO "{table}" ({col_sql}) VALUES ({placeholders})', vals)
                pulled_any = True
            conn.commit()
            if pulled_any:
                _SUPABASE_LAST_PULL_TS = now
                _save_latest_stable_snapshot_safely()
            _set_supabase_error("")
            return pulled_any
        finally:
            conn.close()
    except Exception as e:
        _set_supabase_error(e)
        try:
            print(f"[SUPABASE SYNC PULL] {e}")
        except Exception:
            pass
        return False
    finally:
        _supabase_end_sync()



def _mask_secret(value: str, keep_start: int = 8, keep_end: int = 4) -> str:
    s = str(value or '').strip()
    if not s:
        return ''
    if len(s) <= keep_start + keep_end:
        return '*' * len(s)
    return s[:keep_start] + '...' + s[-keep_end:]


def _fmt_ts_br(ts) -> str:
    try:
        ts = float(ts or 0)
        if ts <= 0:
            return '—'
        return datetime.fromtimestamp(ts).strftime('%d/%m/%Y %H:%M:%S')
    except Exception:
        return '—'


def _supabase_test_connection_detail() -> tuple[bool, str]:
    if not SUPABASE_SYNC_ENABLED:
        return False, 'SUPABASE_URL ou SUPABASE_SERVICE_ROLE_KEY não configurados.'
    try:
        conn = _app_db_connect(DB_PATH)
        try:
            tables = [t for t in _sqlite_user_tables(conn) if t != 'sqlite_sequence']
        finally:
            conn.close()
        pref = [t for t in ["colaboradores", "usuarios_sistema", "setores", "escala_mes"] if t in tables]
        if pref:
            test_table = pref[0]
        elif tables:
            test_table = tables[0]
        else:
            test_table = "colaboradores"
        remote_table = _supabase_resolve_remote_table(test_table, refresh=True)
        r = requests.get(_supabase_table_url(remote_table), params={'select': '*', 'limit': 1}, headers=_supabase_headers(), timeout=15)
        if r.status_code < 400:
            _set_supabase_error("")
            return True, f'Conexão OK. Tabela de teste: {test_table} → {remote_table}'
        msg = f'Supabase respondeu {r.status_code} ao testar {test_table} → {remote_table}: {r.text[:200]}'
        _set_supabase_error(msg)
        return False, msg
    except Exception as e:
        _set_supabase_error(e)
        return False, str(e)


def _supabase_compare_tables_snapshot() -> pd.DataFrame:
    rows = []
    try:
        conn = _app_db_connect(DB_PATH)
        try:
            tables = [t for t in _sqlite_user_tables(conn) if t != 'sqlite_sequence']
            for tb in tables:
                local_count = _sqlite_table_rowcount(conn, tb)
                remote_exists = _table_exists_in_supabase(tb) if SUPABASE_SYNC_ENABLED else False
                remote_count = None
                if remote_exists and SUPABASE_SYNC_ENABLED:
                    try:
                        remote_count = _supabase_table_count_fast(tb)
                        if remote_count is None:
                            remote_count = len(_supabase_fetch_table_rows(tb, page_size=1000))
                    except Exception:
                        remote_count = None
                rows.append({
                    'Tabela': tb,
                    'SQLite local': local_count,
                    'Existe no Supabase': 'Sim' if remote_exists else 'Não',
                    'Supabase': remote_count if remote_count is not None else '—',
                })
        finally:
            conn.close()
    except Exception as e:
        rows.append({'Tabela': '(erro)', 'SQLite local': '—', 'Existe no Supabase': '—', 'Supabase': str(e)})
    return pd.DataFrame(rows)

def _maybe_save_latest_stable_snapshot_fast(reason: str = "commit") -> None:
    """Snapshot leve: evita copiar o banco em todo commit/close."""
    global _LAST_SNAPSHOT_TS, _LAST_SNAPSHOT_DB_MTIME
    try:
        if reason == "close" and FAST_SNAPSHOT_SKIP_ON_CLOSE:
            return
        db_mtime = 0.0
        try:
            if Path(DB_PATH).exists():
                db_mtime = float(Path(DB_PATH).stat().st_mtime)
        except Exception:
            db_mtime = 0.0
        now_ts = time.time()
        if db_mtime > 0 and db_mtime == float(_LAST_SNAPSHOT_DB_MTIME or 0.0):
            if now_ts - float(_LAST_SNAPSHOT_TS or 0.0) < max(5, FAST_SNAPSHOT_THROTTLE_SECONDS):
                return
        if now_ts - float(_LAST_SNAPSHOT_TS or 0.0) < max(5, FAST_SNAPSHOT_THROTTLE_SECONDS):
            return
        _save_latest_stable_snapshot_safely()
        _LAST_SNAPSHOT_TS = now_ts
        _LAST_SNAPSHOT_DB_MTIME = db_mtime
    except Exception:
        pass

class SQLiteSyncConnection(sqlite3.Connection):
    def commit(self):
        result = super().commit()
        try:
            _maybe_save_latest_stable_snapshot_fast("commit")
        except Exception:
            pass
        if SUPABASE_AUTO_PUSH_ON_COMMIT:
            try:
                _supabase_request_push_async(force=False)
            except Exception:
                pass
        return result

    def close(self):
        try:
            _maybe_save_latest_stable_snapshot_fast("close")
        except Exception:
            pass
        if SUPABASE_AUTO_PUSH_ON_CLOSE:
            try:
                _supabase_request_push_async(force=True)
            except Exception:
                pass
        return super().close()

_RUNTIME_READY = False
_RUNTIME_READY_AT = 0.0
_RUNTIME_READY_TTL_SEC = 120.0
_AUTO_BACKUP_DONE_SESSION = False


# V90 — robustez de banco + bootstrap Supabase + snapshot seguro:
# preserva o motor da escala e fortalece restore/sync.

# V87 — robustez de banco:
# procura automaticamente o banco mais completo em caminhos comuns
# e copia para o caminho canônico do app (escala_app/data/escala.db)

def _ensure_data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def _ensure_backup_dir():
    _ensure_data_dir()
    Path(BACKUP_DIR).mkdir(parents=True, exist_ok=True)


def _bundled_latest_stable_paths() -> list[Path]:
    out = []
    seen = set()
    for p in (BUNDLED_LATEST_STABLE_CANDIDATES or []):
        try:
            rp = Path(p).resolve()
        except Exception:
            rp = Path(p)
        if str(rp) not in seen:
            seen.add(str(rp))
            out.append(Path(rp))
    return out


def _adopt_bundled_latest_stable_if_needed(force: bool = False) -> bool:
    """
    V91.2
    - Permite embarcar um latest_stable.db junto do projeto
    - Se o banco atual estiver vazio/inválido, copia esse backup empacotado
    - Também espelha para BACKUP_DIR/latest_stable.db
    """
    try:
        _ensure_backup_dir()
        current = Path(DB_PATH)
        current_ok = current.exists() and _validate_sqlite_file(str(current)) and _db_score(current)[0] >= 1
        if current_ok and not force:
            return False

        for src in _bundled_latest_stable_paths():
            if not _validate_sqlite_file(str(src)):
                continue
            score = _db_score(src)
            if score[0] < 1 and score[1] < 1 and score[2] < 1:
                continue
            try:
                current.parent.mkdir(parents=True, exist_ok=True)
                _sqlite_backup_copy(str(src), str(current))
            except Exception:
                shutil.copy2(src, current)
            latest = Path(BACKUP_DIR) / "latest_stable.db"
            try:
                _sqlite_backup_copy(str(src), str(latest))
            except Exception:
                shutil.copy2(src, latest)
            return True
    except Exception:
        return False
    return False


def _db_candidate_paths() -> list[Path]:
    app = APP_DIR
    cands = [
        Path(DB_PATH),
        Path(ROOT_LEGACY_DB_PATH),
        * _bundled_latest_stable_paths(),
        app / "backups" / "latest_stable.db",
        app / "backups" / "escala.db",
        app / "data" / "latest_stable.db",
        app / "escala_app" / "data" / "escala.db",
        app / "escala_app" / "backups" / "latest_stable.db",
        app.parent / "escala_app" / "data" / "escala.db",
        app.parent / "escala_app" / "backups" / "latest_stable.db",
    ]
    out = []
    seen = set()
    for p in cands:
        try:
            rp = p.resolve()
        except Exception:
            rp = p
        if str(rp) not in seen:
            seen.add(str(rp))
            out.append(Path(rp))
    return out


def _sqlite_table_count(db_path: str, table: str) -> int:
    try:
        conn = sqlite3.connect(str(db_path), check_same_thread=False)
        try:
            row = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()
            return int(row[0]) if row else 0
        finally:
            conn.close()
    except Exception:
        return 0


def _db_score(path: Path) -> tuple:
    if not _validate_sqlite_file(str(path)):
        return (-1, -1, -1, -1, -1, -1, -1)
    size = 0
    try:
        size = path.stat().st_size
    except Exception:
        size = 0
    usuarios = _sqlite_table_count(str(path), "usuarios_sistema")
    colaboradores = _sqlite_table_count(str(path), "colaboradores")
    escala = _sqlite_table_count(str(path), "escala_mes")
    setores = _sqlite_table_count(str(path), "setores")
    recents = _sqlite_table_count(str(path), "login_recent")
    ferias = _sqlite_table_count(str(path), "ferias")
    return (usuarios, colaboradores, escala, setores, recents, ferias, size)


def _pick_best_db_candidate() -> Path | None:
    best = None
    best_score = (-1, -1, -1, -1, -1, -1, -1)
    for p in _db_candidate_paths():
        score = _db_score(p)
        if score > best_score:
            best_score = score
            best = p
    return best if best_score[0] >= 0 else None


def _sqlite_backup_copy(src_path: str, dst_path: str) -> None:
    src_path = str(src_path)
    dst_path = str(dst_path)
    src_conn = sqlite3.connect(src_path, check_same_thread=False)
    try:
        dst_conn = sqlite3.connect(dst_path, check_same_thread=False)
        try:
            try:
                src_conn.execute("PRAGMA wal_checkpoint(FULL)")
            except Exception:
                pass
            src_conn.backup(dst_conn)
            dst_conn.commit()
        finally:
            dst_conn.close()
    finally:
        src_conn.close()


def _validate_sqlite_file(path: str) -> bool:
    p = Path(path)
    if not p.exists() or p.stat().st_size == 0:
        return False
    try:
        conn = sqlite3.connect(str(p), check_same_thread=False)
        try:
            row = conn.execute("PRAGMA integrity_check").fetchone()
            return bool(row) and str(row[0]).lower() == "ok"
        finally:
            conn.close()
    except Exception:
        return False


def _migrate_legacy_db_if_needed():
    _ensure_backup_dir()
    legacy = Path(ROOT_LEGACY_DB_PATH)
    current = Path(DB_PATH)
    if current.exists() and _validate_sqlite_file(str(current)):
        return
    if legacy.exists() and _validate_sqlite_file(str(legacy)):
        try:
            _sqlite_backup_copy(str(legacy), str(current))
        except Exception:
            shutil.copy2(legacy, current)


def _adopt_best_db_candidate_if_needed():
    _ensure_backup_dir()
    current = Path(DB_PATH)
    best = _pick_best_db_candidate()
    if best is None:
        return
    current_score = _db_score(current) if current.exists() else (-1, -1, -1, -1, -1, -1, -1)
    best_score = _db_score(best)
    if best_score > current_score and str(best) != str(current):
        try:
            current.parent.mkdir(parents=True, exist_ok=True)
            _sqlite_backup_copy(str(best), str(current))
        except Exception:
            shutil.copy2(best, current)
        latest = Path(BACKUP_DIR) / "latest_stable.db"
        try:
            _sqlite_backup_copy(str(current), str(latest))
        except Exception:
            shutil.copy2(current, latest)


def _restore_from_latest_stable_if_needed():
    _ensure_backup_dir()
    current = Path(DB_PATH)
    candidates = [Path(BACKUP_DIR) / "latest_stable.db"] + list(_bundled_latest_stable_paths())
    if current.exists() and _validate_sqlite_file(str(current)):
        return
    for latest in candidates:
        try:
            if latest.exists() and _validate_sqlite_file(str(latest)):
                try:
                    if current.exists():
                        current.unlink(missing_ok=True)
                    _sqlite_backup_copy(str(latest), str(current))
                except Exception:
                    shutil.copy2(latest, current)
                return
        except Exception:
            continue




def _ensure_local_db_bootstrap_enterprise() -> bool:
    """
    V97 ENTERPRISE
    Boot em camadas:
    1) mantém o banco atual se já for válido
    2) tenta adotar latest_stable empacotado
    3) tenta restore do latest_stable local/backups
    4) tenta pull do Supabase
    5) cria um SQLite mínimo se tudo falhar
    Nunca bloqueia a abertura do app por base vazia.
    """
    try:
        _ensure_backup_dir()
    except Exception:
        pass

    current = Path(DB_PATH)
    try:
        if current.exists() and _validate_sqlite_file(str(current)):
            return True
    except Exception:
        pass

    try:
        _adopt_bundled_latest_stable_if_needed(force=True)
        if current.exists() and _validate_sqlite_file(str(current)):
            return True
    except Exception:
        pass

    try:
        _restore_from_latest_stable_if_needed()
        if current.exists() and _validate_sqlite_file(str(current)):
            return True
    except Exception:
        pass

    try:
        if SUPABASE_SYNC_ENABLED:
            _restore_from_supabase_if_local_empty(force=True)
            if current.exists() and _validate_sqlite_file(str(current)):
                return True
    except Exception as e:
        _set_supabase_error(e)

    try:
        current.parent.mkdir(parents=True, exist_ok=True)
        con = sqlite3.connect(str(current))
        try:
            con.execute("PRAGMA journal_mode=WAL")
            con.execute("PRAGMA foreign_keys=ON")
            con.commit()
        finally:
            con.close()
        return current.exists()
    except Exception:
        return False


def _ensure_runtime_storage_ready(force: bool = False):
    global _RUNTIME_READY, _RUNTIME_READY_AT
    try:
        now_ts = datetime.now().timestamp()
    except Exception:
        now_ts = 0.0
    if (not force) and _RUNTIME_READY and (now_ts - float(_RUNTIME_READY_AT or 0.0) < float(_RUNTIME_READY_TTL_SEC)):
        return
    _ensure_backup_dir()
    _ensure_local_db_bootstrap_enterprise()
    _adopt_bundled_latest_stable_if_needed(force=False)
    _migrate_legacy_db_if_needed()
    _adopt_best_db_candidate_if_needed()
    _restore_from_latest_stable_if_needed()
    _adopt_best_db_candidate_if_needed()
    if SUPABASE_AUTO_RESTORE_IF_LOCAL_EMPTY:
        try:
            _restore_from_supabase_if_local_empty(force=False)
        except Exception:
            pass
    elif SUPABASE_AUTO_PULL_ON_START:
        try:
            if SUPABASE_SYNC_ENABLED and not _sqlite_app_has_meaningful_data():
                _supabase_pull_all_to_sqlite(force=True)
        except Exception:
            pass
    _RUNTIME_READY = True
    _RUNTIME_READY_AT = now_ts


def _remove_sqlite_sidecars(db_path: str):
    for extra in ("-wal", "-shm"):
        p = Path(str(db_path) + extra)
        if p.exists():
            p.unlink(missing_ok=True)


def create_backup_now(prefix="manual") -> str:
    _ensure_runtime_storage_ready()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    src = Path(DB_PATH)
    if not src.exists():
        raise FileNotFoundError(f"Banco não encontrado: {DB_PATH}")
    dst = Path(BACKUP_DIR) / f"escala_{prefix}_{ts}.db"
    _sqlite_backup_copy(str(src), str(dst))
    latest = Path(BACKUP_DIR) / "latest_stable.db"
    try:
        _sqlite_backup_copy(str(src), str(latest))
    except Exception:
        shutil.copy2(src, latest)
    try:
        bundled = Path(APP_DIR) / "latest_stable.db"
        _sqlite_backup_copy(str(src), str(bundled))
    except Exception:
        try:
            shutil.copy2(src, Path(APP_DIR) / "latest_stable.db")
        except Exception:
            pass
    _rotate_backup_files_safely()
    return str(dst)


def list_backups() -> list:
    _ensure_runtime_storage_ready()
    files = sorted(Path(BACKUP_DIR).glob("*.db"), key=lambda p: p.stat().st_mtime, reverse=True)
    return [p.name for p in files]


def _auto_backup_marker_path() -> Path:
    return Path(BACKUP_DIR) / ".last_auto_backup"


def auto_backup_if_due():
    """
    Cria backup automático quando o app abre:
    - nunca antes das 03:00
    - no máximo 1 backup a cada AUTO_BACKUP_INTERVAL_HOURS
    Observação: Streamlit não executa em segundo plano; só roda ao abrir/recarregar o app.
    """
    global _AUTO_BACKUP_DONE_SESSION
    try:
        if _AUTO_BACKUP_DONE_SESSION:
            return
        _ensure_runtime_storage_ready()
        now = datetime.now()
        if now.hour < AUTO_BACKUP_HOUR:
            _AUTO_BACKUP_DONE_SESSION = True
            return
        marker = _auto_backup_marker_path()
        last_raw = marker.read_text(encoding="utf-8").strip() if marker.exists() else ""
        last_dt = None
        if last_raw:
            try:
                last_dt = datetime.fromisoformat(last_raw)
            except Exception:
                last_dt = None
                try:
                    last_day = datetime.strptime(last_raw, "%Y-%m-%d")
                    if now.date() == last_day.date():
                        return
                except Exception:
                    pass
        if last_dt is not None:
            hours_since = (now - last_dt).total_seconds() / 3600.0
            if hours_since < float(AUTO_BACKUP_INTERVAL_HOURS):
                _AUTO_BACKUP_DONE_SESSION = True
                return
        create_backup_now(prefix=f"auto_{now.strftime('%Y%m%d_%H%M%S')}")
        marker.write_text(now.isoformat(), encoding="utf-8")
        _AUTO_BACKUP_DONE_SESSION = True
    except Exception:
        return


def _clear_runtime_caches_after_db_change():
    try:
        st.cache_data.clear()
    except Exception:
        pass
    try:
        st.cache_resource.clear()
    except Exception:
        pass
    try:
        for k in [
            "adm_bk_sel",
            "adm_bk_up",
            "gerar_escala_cache",
            "df_colabs_cache",
        ]:
            if k in st.session_state:
                del st.session_state[k]
    except Exception:
        pass


def _db_runtime_summary() -> dict:
    current = Path(DB_PATH)
    return {
        "db_path": str(current),
        "db_exists": current.exists(),
        "db_score": _db_score(current),
        "best_candidate": str(_pick_best_db_candidate() or ""),
    }


def restore_backup_from_bytes(data: bytes) -> None:
    _ensure_runtime_storage_ready()
    tmp = Path(BACKUP_DIR) / "_upload_restore_tmp.db"
    tmp.write_bytes(data)
    if not _validate_sqlite_file(str(tmp)):
        tmp.unlink(missing_ok=True)
        raise ValueError("Arquivo de backup inválido ou corrompido.")

    current = Path(DB_PATH)
    # safety backup do banco atual, se existir
    try:
        if current.exists() and _validate_sqlite_file(str(current)):
            create_backup_now(prefix="pre_restore")
    except Exception:
        pass

    _remove_sqlite_sidecars(DB_PATH)
    current.unlink(missing_ok=True)

    try:
        _sqlite_backup_copy(str(tmp), str(current))
    except Exception:
        shutil.copy2(tmp, current)

    _remove_sqlite_sidecars(DB_PATH)

    try:
        conn = sqlite3.connect(str(current), check_same_thread=False)
        try:
            conn.execute("PRAGMA wal_checkpoint(FULL)")
            row = conn.execute("PRAGMA integrity_check").fetchone()
            if not row or str(row[0]).lower() != "ok":
                raise ValueError("Banco restaurado ficou inválido após a cópia.")
        finally:
            conn.close()
    except Exception:
        raise

    latest = Path(BACKUP_DIR) / "latest_stable.db"
    try:
        _sqlite_backup_copy(str(current), str(latest))
    except Exception:
        shutil.copy2(current, latest)
    try:
        bundled = Path(APP_DIR) / "latest_stable.db"
        _sqlite_backup_copy(str(current), str(bundled))
    except Exception:
        try:
            shutil.copy2(current, Path(APP_DIR) / "latest_stable.db")
        except Exception:
            pass
    _rotate_backup_files_safely()

    tmp.unlink(missing_ok=True)
    _clear_runtime_caches_after_db_change()

def listar_setores_db() -> list:
    conn = _app_db_connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS setores (nome TEXT PRIMARY KEY)")
    cur.execute("SELECT nome FROM setores ORDER BY nome")
    rows = [r[0] for r in cur.fetchall()]
    conn.close()
    base_set = {"ADMIN", "GERAL"}
    return sorted(list(base_set.union({(x or "").strip().upper() for x in rows if x})))


def list_setores() -> list:
    """Alias de compatibilidade: algumas telas chamam list_setores()."""
    try:
        return listar_setores_db()
    except Exception:
        return ["ADMIN"]


def criar_setor_db(nome: str) -> None:
    nome = (nome or "").strip().upper()
    if not nome:
        raise ValueError("Setor vazio")
    conn = _app_db_connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS setores (nome TEXT PRIMARY KEY)")
    cur.execute("INSERT OR IGNORE INTO setores(nome) VALUES(?)", (nome,))
    conn.commit(); conn.close()


def importar_colaboradores_df(setor: str, df: pd.DataFrame) -> tuple[int,int]:
    setor = (setor or "").strip().upper()
    if not setor:
        raise ValueError("Setor destino inválido")
    cols = {c.lower().strip(): c for c in df.columns}
    for r in ("nome", "chapa"):
        if r not in cols:
            raise ValueError(f"Coluna obrigatória faltando: {r}")

    nome_s = df[cols["nome"]]
    chapa_s = df[cols["chapa"]]
    subgrupo_s = df[cols["subgrupo"]] if "subgrupo" in cols else pd.Series([""]*len(df))
    entrada_s = df[cols["entrada"]] if "entrada" in cols else pd.Series(["06:00"]*len(df))
    sab_s = df[cols["folga_sabado"]] if "folga_sabado" in cols else (df[cols["sabado"]] if "sabado" in cols else pd.Series([0]*len(df)))

    conn = _app_db_connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS colaboradores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        subgrupo TEXT DEFAULT '',
        entrada TEXT DEFAULT '06:00',
        folga_sabado INTEGER DEFAULT 0,
        criado_em TEXT DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(setor, chapa)
    )""")

    inserted=0; updated=0
    for nome, chapa, sg, ent, sab in zip(nome_s, chapa_s, subgrupo_s, entrada_s, sab_s):
        nome = str(nome).strip()
        chapa = str(chapa).strip()
        if not nome or not chapa:
            continue
        sg = str(sg).strip()
        ent = str(ent).strip() if str(ent).strip() else "06:00"
        try:
            sab_i = int(sab)
        except Exception:
            sab_i = 0
        cur.execute("SELECT id FROM colaboradores WHERE setor=? AND chapa=?", (setor, chapa))
        if cur.fetchone():
            cur.execute("UPDATE colaboradores SET nome=?, subgrupo=?, entrada=?, folga_sabado=? WHERE setor=? AND chapa=?",
                        (nome, sg, ent, sab_i, setor, chapa))
            updated += 1
        else:
            cur.execute("INSERT INTO colaboradores(nome,setor,chapa,subgrupo,entrada,folga_sabado) VALUES(?,?,?,?,?,?)",
                        (nome, setor, chapa, sg, ent, sab_i))
            inserted += 1
        try:
            ensure_system_user_from_colaborador(nome, setor, chapa)
        except Exception:
            pass
    conn.commit(); conn.close()
    return inserted, updated

# ---- Regras fixas
INTERSTICIO_MIN = timedelta(hours=11, minutes=10)   # 11:10
DURACAO_JORNADA = timedelta(hours=9, minutes=58)    # 9:58

PREF_EVITAR_PENALTY = 1000

BALANCO_STATUS = "Balanço"
WORK_STATUSES = {"Trabalho", BALANCO_STATUS}

BALANCO_DIA_ENTRADA = "06:00"
BALANCO_DIA_SAIDA = "11:50"

# Presets de horários (facilita seleção no app) — mais completo
HORARIOS_ENTRADA_PRESET = sorted({
    # Madrugada / manhã cedo
    "04:00","04:30","05:00","05:30",
    "06:00","06:10","06:20","06:30","06:40","06:45","06:50",
    "07:00","07:10","07:20","07:30","07:40","07:50",
    "08:00","08:10","08:20","08:30","08:40","08:50",
    # Manhã / meio-dia
    "09:00","09:10","09:20","09:30","09:40","09:50",
    "10:00","10:10","10:20","10:30","10:40","10:50",
    "11:00","11:10","11:20","11:30","11:40","11:50",
    "12:00","12:10","12:20","12:30","12:20","12:45","12:50",
    # Tarde
    "13:00","13:10","13:20","13:30","13:40","13:50",
    "14:00","14:10","14:20","14:30","14:40","14:50",
    "15:00","15:10","15:20","15:30","15:40","15:50",
    "16:00","16:10","16:20","16:30","16:40","16:50",
    "17:00","17:10","17:20","17:30","17:40","17:50",
    # Noite (se precisar)
    "18:00","18:10","18:20","18:30","18:40","18:50",
    "19:00","19:10","19:20","19:30","00:10",
})

D_PT = {
    "Monday": "seg",
    "Tuesday": "ter",
    "Wednesday": "qua",
    "Thursday": "qui",
    "Friday": "sex",
    "Saturday": "sáb",
    "Sunday": "dom",
}


# =========================================================
# ESCALA MANUAL (BASE) — Fevereiro/2026 (DSR)
# - Esta base serve para "iniciar" o mês com folgas pré-definidas.
# - Ao clicar em "Aplicar base", o app cria overrides (Status=Folga) nesses dias.
# - Depois, "Gerar agora (respeitando ajustes)" completa o restante mantendo as folgas travadas.
# =========================================================
MANUAL_BASES = {
    (2026, 2): [
        {"Chapa": "020.0823", "Nome": "ALEXANDRE ROBERTO ALMEIDA DOS REIS", "Dias_Folga": [1,4,6,9,15,18,20,23]},
        {"Chapa": "020.1447", "Nome": "ANA CAROLINA THEODORO PADILHA", "Dias_Folga": [1,3,6,11,15,18,20,26]},
        {"Chapa": "020.1733", "Nome": "BEATRIZ VITORIA DOS SANTOS LOPES", "Dias_Folga": [3,8,11,13,16,22,24,26]},
        {"Chapa": "020.1751", "Nome": "BRUNA SILVA MARTINS", "Dias_Folga": [1,4,6,9,15,17,19,23]},
        {"Chapa": "020.2288", "Nome": "CRISTIANE ALVES DOS SANTOS", "Dias_Folga": [2,8,11,13,16,22,24,26]},
        {"Chapa": "020.0265", "Nome": "DECIO EPAMINONDAS DE ALMEIDA NETO", "Dias_Folga": [4,8,10,12,18,22,24,26]},
        {"Chapa": "020.1839", "Nome": "DEYBSON JOSE DA SILVA", "Dias_Folga": [2,8,10,13,19,22,24,26]},
        {"Chapa": "020.1884", "Nome": "DISNEI OLIVEIRA ADORNO", "Dias_Folga": [1,4,6,10,15,17,20,23]},
        {"Chapa": "020.2192", "Nome": "EDILENE MARTINS DE MIRANDA", "Dias_Folga": [3,8,10,12,17,22,24,26]},
        {"Chapa": "020.2144", "Nome": "ELIS MIRIAN MARQUES OLIVEIRA", "Dias_Folga": [1,4,6,12,15,18,20,26]},
        {"Chapa": "020.1750", "Nome": "ELIZANGELA BARBOSA MOREIRA", "Dias_Folga": [22,25,27]},
        {"Chapa": "020.1984", "Nome": "EWERLON DE JESUS DA SILVA E SILVA", "Dias_Folga": [1,3,6,9,15,17,20,23]},
        {"Chapa": "020.2139", "Nome": "FABIANA SOUZA SILVA", "Dias_Folga": [3,8,11,13,18,22,24,26]},
        {"Chapa": "020.2450", "Nome": "GABRIEL CAMELO PINTO", "Dias_Folga": [3,8,10,12,18,22,25,27]},
        {"Chapa": "020.0748", "Nome": "IVANILDO FIGUEIREDO DA VERA CRUZ", "Dias_Folga": [16,22,25,27]},
        {"Chapa": "020.2299", "Nome": "JAIRON MACHADO DE ALMEIDA", "Dias_Folga": [2,8,11,13,16,22,24,26]},
        {"Chapa": "020.1649", "Nome": "JOAO VICTOR DE SOUZA SAMPAIO", "Dias_Folga": [1,3,5,9,15,17,20,25]},
        {"Chapa": "020.2274", "Nome": "JOSE FERNANDO OLIVEIRA DO NASCIMENTO", "Dias_Folga": [1,4,6,10,15,18,20,25]},
        {"Chapa": "020.2143", "Nome": "LUCAS EDUARDO DOS SANTOS SANTILLO", "Dias_Folga": [8,10,12,16,22,24,26]},
        {"Chapa": "020.1639", "Nome": "LUCIMARA EMILIA MARQUES", "Dias_Folga": [1,3,5,9,15,18,20,25]},
        {"Chapa": "020.2050", "Nome": "LUIZ FERNANDO DE TULIO", "Dias_Folga": [1,3,5,11,15,17,19,23]},
        {"Chapa": "020.1628", "Nome": "MACICLEIDE CONCEICAO DOS SANTOS", "Dias_Folga": [1,5,8,10,13,19,22,25,27]},
        {"Chapa": "020.0463", "Nome": "MARIA EDUARDA GONCALVES NUNES", "Dias_Folga": [2,8,10,12,16,22,24,26]},
        {"Chapa": "020.1854", "Nome": "MARIANA MABILLE DE MORAES", "Dias_Folga": []},
        {"Chapa": "020.1128", "Nome": "MARIVALDO RODRIGUES DA SILVA", "Dias_Folga": [1,4,6,12,15,18,20,23]},
        {"Chapa": "020.2309", "Nome": "MAURICIO DAVI DA SILVA NEIVAS ARAUJO", "Dias_Folga": [1,3,5,9,15,17,20,23]},
        {"Chapa": "020.2348", "Nome": "NATALIA CRISTINA GIMENES DE OLIVEIRA", "Dias_Folga": [1,4,6,12,15,17,19,23,27]},
        {"Chapa": "020.1856", "Nome": "RIQUELME CABRAL DE JESUS", "Dias_Folga": [3,8,11,13,18,22,24,26]},
        {"Chapa": "020.2388", "Nome": "RUTH PEREIRA DA SILVA", "Dias_Folga": [2,8,11,13,19,22,25,27]},
        {"Chapa": "020.1906", "Nome": "SHAIAN RUAN BARBOSA ALVES", "Dias_Folga": [4,8]},
        {"Chapa": "020.2203", "Nome": "TATIANE APARECIDA CABECA", "Dias_Folga": [1,4,6,9,15,18,20,25]},
        {"Chapa": "020.0994", "Nome": "VERA LUCIA BENEDITO ARRUDA", "Dias_Folga": [1,4,8,11,15,18,22,25]},
        {"Chapa": "020.1559", "Nome": "VIVIANE NASCIMENTO LIMA LEMOS", "Dias_Folga": [1,3,5,11,15,17,19,23]},
        {"Chapa": "020.1980", "Nome": "YASMIM STEFHANNY BATA SANTOS", "Dias_Folga": [5,8,10,12,17,22,24,26]},
    ]
}

# =========================================================
# Helpers de hora (minutos)
# =========================================================
def _to_min(hhmm: str) -> int:
    if not hhmm:
        return 0
    h, m = map(int, str(hhmm).split(":"))
    return h * 60 + m

def _min_to_hhmm(x: int) -> str:
    x %= (24 * 60)
    return f"{x//60:02d}:{x%60:02d}"

def _add_min(hhmm: str, delta: timedelta) -> str:
    return _min_to_hhmm(_to_min(hhmm) + int(delta.total_seconds() // 60))

def _sub_min(hhmm: str, delta: timedelta) -> str:
    return _min_to_hhmm(_to_min(hhmm) - int(delta.total_seconds() // 60))

def _saida_from_entrada(ent: str) -> str:
    return _add_min(ent, DURACAO_JORNADA)



# =========================================================
# PDF helpers (modelo de escala/ponto)
# - Linha "Horas Trab." do modelo costuma ser 08:48 (jornada 9:58 com 1:10 de intervalo)
# =========================================================
DURACAO_TRABALHADA = timedelta(hours=8, minutes=48)   # 08:48 (modelo)

def _hhmm_add(hhmm: str, minutes: int) -> str:
    if not hhmm:
        return ""
    h, m = map(int, hhmm.split(":"))
    total = (h * 60 + m + int(minutes)) % (24 * 60)
    return f"{total//60:02d}:{total%60:02d}"

def _montar_batidas_modelo(h_entrada: str):
    """
    Retorna (entrada1, saida_ref, entrada_ref, saida, horas_trab)

    Modelo igual ao do PDF:
      - Jornada (entrada->saída) = 9:58  (DURACAO_JORNADA)
      - Intervalo refeição = 1:10
      - Primeira parte (entrada -> saída refeição) = 5:10
      - Resultado "Horas Trab." = 08:48 quando é jornada padrão.
    """
    h_entrada = (h_entrada or "").strip()
    if not h_entrada:
        return "", "", "", "", ""

    # Parte 1 = 5h10, Refeição = 1h10
    parte1 = 5 * 60 + 10
    refeicao = 1 * 60 + 10

    saida_ref = _hhmm_add(h_entrada, parte1)
    ent_ref = _hhmm_add(saida_ref, refeicao)
    saida = _hhmm_add(h_entrada, int(DURACAO_JORNADA.total_seconds() // 60))  # 9:58

    # Horas trabalhadas no modelo = (9:58 - 1:10) = 8:48
    horas = "08:48"
    return h_entrada, saida_ref, ent_ref, saida, horas

def gerar_pdf_modelo_oficial(setor: str, ano: int, mes: int, hist_db: dict, colaboradores: list[dict]) -> bytes:
    """
    Gera PDF (A4 paisagem) com **4 colaboradores por página** (como o modelo enviado).
    - Folga: "FOLG" com destaque amarelo.
    - Férias: "FER" (sem destaque).
    - Manual supremo: o PDF reflete exatamente o que está salvo em hist_db.
    """
    from io import BytesIO
    from reportlab.pdfgen import canvas
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    import re

    # -----------------------------
    # Canvas com contagem total de páginas (X / Y)
    # -----------------------------
    class _NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            canvas.Canvas.__init__(self, *args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            num_pages = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self._draw_page_number(num_pages)
                canvas.Canvas.showPage(self)
            canvas.Canvas.save(self)

        def _draw_page_number(self, page_count):
            # no topo direito
            self.setFont("Helvetica", 7)
            self.drawRightString(landscape(A4)[0] - 12*mm, landscape(A4)[1] - 10*mm, f"Página: {self._pageNumber} / {page_count}")

    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    normal.fontName = "Helvetica"
    normal.fontSize = 7
    normal.leading = 8

    # Ordena colaboradores pelo nome (igual na tela)
    colab_by = {str(c.get("Chapa", "")).strip(): c for c in colaboradores}
    chapas = sorted([str(ch).strip() for ch in hist_db.keys()], key=lambda ch: (colab_by.get(str(ch).strip(), {}).get("Nome", ch) or ch))

    # Config páginas
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10*mm,
        rightMargin=10*mm,
        topMargin=14*mm,
        bottomMargin=10*mm,
        title=f"Escala DSR {setor} {mes:02d}/{ano}"
    )

    W, H = landscape(A4)
    usable_w = W - doc.leftMargin - doc.rightMargin

    # ----- helpers
    def _pt_weekday(ts: pd.Timestamp) -> str:
        # usa D_PT já definido no app: Monday->seg...
        return {
            "seg": "Seg", "ter": "Ter", "qua": "Qua", "qui": "Qui", "sex": "Sex", "sáb": "Sáb", "dom": "Dom"
        }.get(D_PT[ts.day_name()], D_PT[ts.day_name()])

    def _format_mes():
        return f"Mês: {mes:02d}/{ano}"

    def _hhmm_norm(h: str) -> str:
        h = (h or "").strip()
        if not h:
            return ""
        h = h.replace(".", ":")
        if re.fullmatch(r"\d{1,2}:\d{2}", h):
            hh, mm_ = h.split(":")
            return f"{int(hh):02d}:{int(mm_):02d}"
        if re.fullmatch(r"\d{3,4}", h):
            h = h.zfill(4)
            return f"{h[:2]}:{h[2:]}"
        return h

    def _hhmm_diff_min(h1: str, h2: str) -> int:
        try:
            h1n = _hhmm_norm(h1); h2n = _hhmm_norm(h2)
            if not h1n or not h2n:
                return 0
            t1 = datetime.strptime(h1n, "%H:%M")
            t2 = datetime.strptime(h2n, "%H:%M")
            return int((t2 - t1).total_seconds() // 60)
        except Exception:
            return 0

    def _sum_total_horas(df: pd.DataFrame) -> str:
        # soma horas trabalhadas no modelo (primeira parte + segunda parte), respeitando horários reais quando existirem.
        total_min = 0
        for _, r in df.iterrows():
            stt = str(r.get("Status", ""))
            if stt not in WORK_STATUSES:
                continue
            ent = (r.get("H_Entrada") or "").strip()
            sai = (r.get("H_Saida") or "").strip()
            if not ent or not sai:
                continue
            # tenta modelo com refeição
            ent1, sref, entref, sai2, _ = _montar_batidas_modelo(ent)
            if sai2 == sai and sref and entref:
                # 8:48 padrão
                total_min += 8*60 + 48
            else:
                # fallback: duração bruta
                dur = _hhmm_diff_min(ent, sai)
                if dur > 0:
                    total_min += dur
        return f"{total_min//60}:{total_min%60:02d}"

    def _make_block(ch: str) -> list:
        df = hist_db[ch].copy()
        nome = colab_by.get(ch, {}).get("Nome", ch)
        sg = ''
        try:
            if 'Subgrupo' in df.columns:
                vals_sg = [str(v).strip() for v in df['Subgrupo'].astype(str).tolist() if str(v).strip()]
                if vals_sg:
                    sg = vals_sg[0]
        except Exception:
            sg = ''
        sg = sg or get_subgrupo_competencia_ou_base(
            setor,
            str(ch).strip(),
            int(ano),
            int(mes),
            (colab_by.get(str(ch).strip(), {}).get("Subgrupo", "") or "").strip()
        ) or "COLABORADOR"
        sg_title = str(sg).upper()

        # tabela por dia
        qtd = len(df)
        dias_nums = [str(int(d.day)) for d in pd.to_datetime(df["Data"])]
        dias_sem = [_pt_weekday(pd.to_datetime(d)) for d in pd.to_datetime(df["Data"])]

        # constrói matriz 7 x (1+qtd)
        data = []
        data.append(["Data / Dia"] + dias_nums)
        data.append(["Dia / Semana"] + dias_sem)

        row_ent = ["Entrada"]
        row_sref = ["Saída Refeição"]
        row_entref = ["Entrada"]
        row_sai = ["Saída"]
        row_h = ["Horas Trab."]

        folg_cols = []
        for i in range(qtd):
            stt = str(df.loc[i, "Status"])
            ent = (df.loc[i, "H_Entrada"] or "").strip()
            sai = (df.loc[i, "H_Saida"] or "").strip()

            if stt == "Folga":
                row_ent.append("FOLG")
                row_sref.append("FOLG")
                row_entref.append("FOLG")
                row_sai.append("FOLG")
                row_h.append("")
                folg_cols.append(i+1)  # +1 por causa do label col
            elif stt == "Férias":
                row_ent.append("FER")
                row_sref.append("FER")
                row_entref.append("FER")
                row_sai.append("FER")
                row_h.append("")
            elif stt in WORK_STATUSES:
                if stt == BALANCO_STATUS:
                    row_ent.append(ent)
                    row_sref.append("")
                    row_entref.append("")
                    row_sai.append(sai)
                    # horas brutas
                    dm = _hhmm_diff_min(ent, sai) if ent and sai else 0
                    row_h.append(f"{dm//60:02d}:{dm%60:02d}" if dm else "")
                else:
                    ent1, sref, entref, saida2, horas = _montar_batidas_modelo(ent or colab_by.get(ch, {}).get("Entrada", "06:00"))
                    # respeita saída real do DF se diferente
                    if sai and saida2 and _hhmm_norm(sai) != _hhmm_norm(saida2):
                        # se alterado manualmente, mantém o do DF e deixa refeição em branco
                        row_ent.append(ent or "")
                        row_sref.append("")
                        row_entref.append("")
                        row_sai.append(sai)
                        dm = _hhmm_diff_min(ent, sai) if ent and sai else 0
                        row_h.append(f"{dm//60:02d}:{dm%60:02d}" if dm else "")
                    else:
                        row_ent.append(ent1)
                        row_sref.append(sref)
                        row_entref.append(entref)
                        row_sai.append(saida2)
                        row_h.append(horas)
            else:
                # status desconhecido
                row_ent.append("")
                row_sref.append("")
                row_entref.append("")
                row_sai.append("")
                row_h.append("")

        data += [row_ent, row_sref, row_entref, row_sai, row_h]

        label_w = 34*mm
        day_w = (usable_w - label_w) / max(1, qtd)

        tbl = Table(
            data,
            colWidths=[label_w] + [day_w]*qtd,
            rowHeights=[10, 10, 10, 10, 10, 10, 10]
        )

        ts = TableStyle([
            ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
            ("FONTSIZE", (0,0), (-1,-1), 7),
            ("ALIGN", (0,0), (0,-1), "LEFT"),
            ("ALIGN", (1,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("GRID", (0,0), (-1,-1), 0.5, colors.black),
            ("BACKGROUND", (0,0), (0,-1), colors.whitesmoke),
            ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ])

        # destaque FOLG (linhas 2..5)
        for c in folg_cols:
            for r in [2,3,4,5]:
                ts.add("BACKGROUND", (c, r), (c, r), colors.HexColor("#FFE699"))
                ts.add("FONTNAME", (c, r), (c, r), "Helvetica-Bold")

        tbl.setStyle(ts)

        # Barra cinza (cargo)
        bar = Table([[sg_title]], colWidths=[usable_w], rowHeights=[10])
        bar.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#D9D9D9")),
            ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOX", (0,0), (-1,-1), 0.8, colors.black),
        ]))

        # Linha Nome / Mês / Cliente
        header2 = Table(
            [[f"{nome} ({ch})", _format_mes(), "CLIENTE:"]],
            colWidths=[usable_w*0.55, usable_w*0.20, usable_w*0.25],
            rowHeights=[10]
        )
        header2.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 7),
            ("ALIGN", (0,0), (0,0), "LEFT"),
            ("ALIGN", (1,0), (1,0), "CENTER"),
            ("ALIGN", (2,0), (2,0), "LEFT"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOX", (0,0), (-1,-1), 0.8, colors.black),
        ]))

        # Rodapé do bloco
        total_horas = _sum_total_horas(df)
        footer = Table(
            [["É DE RESPONSABILIDADE DE CADA FUNCIONÁRIO CUMPRIR RIGOROSAMENTE ESTA ESCALA.", f"TOTAL DE HORAS : {total_horas}"]],
            colWidths=[usable_w*0.78, usable_w*0.22],
            rowHeights=[10]
        )
        footer.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 7),
            ("ALIGN", (0,0), (0,0), "LEFT"),
            ("ALIGN", (1,0), (1,0), "RIGHT"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOX", (0,0), (-1,-1), 0.8, colors.black),
        ]))

        return [bar, header2, tbl, footer, Spacer(1, 6)]

    # Cabeçalho de página (desenhado pelo onPage)
    emissao = datetime.now().strftime("%d/%m/%Y %H:%M")

    def _draw_header(canv, doc_):
        canv.saveState()
        canv.setStrokeColor(colors.black)
        canv.setFillColor(colors.black)

        # caixas do topo
        y = H - 12*mm
        canv.setFont("Helvetica-Bold", 9)
        canv.drawString(doc.leftMargin, y, f"Loja: {setor}")
        canv.drawCentredString(W/2, y, "Escala de DSR e Horário de Trabalho - Mês : {:02d}/{:04d}".format(mes, ano))
        canv.setFont("Helvetica", 7)
        canv.drawRightString(W - doc.rightMargin, y, f"Emissão: {emissao}")

        # título grande
        canv.setFont("Helvetica-Bold", 10)
        canv.drawString(doc.leftMargin, y - 10, "ESCALA_PONTO_NEW")

        # linha separadora
        canv.setLineWidth(1)
        canv.line(doc.leftMargin, y - 12, W - doc.rightMargin, y - 12)

        canv.restoreState()

    # Monta story: 4 blocos por página
    story = []
    per_page = 4
    for i, ch in enumerate(chapas):
        story += _make_block(ch)
        if (i+1) % per_page == 0 and (i+1) < len(chapas):
            story.append(PageBreak())

    doc.build(story, onFirstPage=_draw_header, onLaterPages=_draw_header, canvasmaker=_NumberedCanvas)
    return buffer.getvalue()


def gerar_pdf_ferias_mes(setor: str, ano: int, mes: int, colaboradores: list[dict], keyword: str = "") -> bytes:
    """
    PDF A4 (paisagem) - Relatório "Férias do mês"
    Colunas: Nome, Chapa, Início, Fim, Dias (total).
    - Ordena por Início (quem sai primeiro aparece primeiro)
    - Cabeçalho com 2 logos (Savegnago e Paulistão)
    - Rodapé com assinaturas RH / Gerência
    - keyword (opcional): filtra por nome ou chapa
    """
    from io import BytesIO
    import base64
    import pandas as pd
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from PIL import Image as PILImage

    # logos embutidos (não dependem de arquivo no Streamlit Cloud)
    LOGO_SAVEGNAGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAfQAAADzCAYAAABwpOSGAAEAAElEQVR4nOy9d7wdV3nu/11rTdntVOmoN0uyXOTewZUSwLQEUklIT37p/cYkkOSGJPcm3JBOchNSgCQXSAghgRCKAYONeze25V5ULOnoSKfuMmWt9ftjzex2zlGxJaswz+cj7TZ7Zs2afeZZb3teQYEC38wQfa/twTeXyJ7XBnN0x1OgQIECLxLy0JsUKFCgQIECBU50FIReoECBAgUKnAIoCL1AgQIFChQ4BVAQeoECBQoUKHAKoCD0AgUKFChQ4BRAf45vgQLfXDjCLPcCBQoUOFHhHe8BFDjZ0e3kOclLuAoyL1CgwEmMwkIv8BIgWb3yNPdUpO5fD6lLVq9bC3bxyM7s3PRBjxBH0UE/r1SqB/08DMs9r2Ufad919129n2eP8+rL87+UgvQLFChwgqKw0AscGwgDFqQ5eJpGP8HO280hhV4O8bnt3sr07NMWy9kCBQqcQigIvcBLgAHZ6nlHWDIyz8gTFrbQhftc2l6ynff5Idz4wh7sc4mwEiFUvjXW2vYiwjBf+e2QYYMi5l6gQIETFEWWe4GXCNP1L4OVgAQrHYEKgRSOTI83bNcCQIjCRC9QoMCpg8JCL/ASIemQucEKiRCCOLaUwjLGCqTniNwYjRS9a8jUpkCHXDskK9Fao/v84vNIWDrrW2vdXjB0tjEgLbZrfEYYAt+j2WziqWBerLxY4RYoUOBkRUHoBV4CJNjsJ5QlxFlrwXqUgkGiyGIYIs7i6CaLq3eTcpp/34K1FmMMQgiklEgl0WknKc5ai9W2x9KXcZnADxCeIIljt2DocgQkhD0jtsYQaYUW5cz17gHpUZyTAgUKFDg+KAj9FEN/RHqRCPWhcTjeaCvBVNzeZQOEQUqJ1QFR00f5o2i7jjR2o5jv4pZo3aLbLn7myYeyF54P0kNWsvH0WvD5o+d5xM2JCFE1GzZuNr7vd+3fkCQeoivzzmKJWhrP94iSaaCEZO7FzYld5L0CBQoUOA4oCP2kx+FStux7nrvK++LfSMDL4uBe9m8xR3SANqMuEc42sDLF80IsPvv3tyobNmyoSLGimiTpCDAahuFwqtNhYBQYwqrB0SVDQzjWHgT8kdGtVbS2QAUI+o9ospg8gBDthcKclFJL30ujKKoDETAD6XQrmp0V2EngADBlrZ2M02S/8sv7tfbqsHTOUFqgNs50PWZWv83nKgYRZ14JeklcUJB6gQIFjguKrKDjjUNcgTPPOifbzpGJ7CFgyVB1WfZ5XgOeka+0CGmd+1p4CEJAYoUiihI8WaFcCTF2lkQ38JQCPNJEIVUVYWvYtIrVJbT1sdbiez5GSL5+683+aRs2DQ3UhkY9Fa5WylutfLsaKdcIoVZj5TJsMGrxhrVVQ0IoH3Cm87yM997XQohDlrItBiPoS7wz7Wx5AGvcZ9ba1FoTIdJpIVoTknTSGrsT2JHEreeB54EdCDP+wIN3Tl568VWJ0R7Ks3heC9QUVu5H+g2s1VjtiF0pCcZSKpWoz80glepJwoP5ZXjT01kdfj4vIu6aF8nD33j0xU1GgQIFvulQEPrxxqEI/eyzu14tRuiyy1rMiEHpjNAtSEfowpaxxifVASZVSCUoVSCK6/ies8ZN6qG8IRpzgomJ5kitsmK5r6qneb630Qq2CCE2SalWWWtXGMEIVpaUkm0XuEH2keqRpZmpw/lJisW9EfolZtLLvOwOEmBSm3SXkt7zSSwe95XcdmD/9sdKtdbzD33jxn1XXnehjuOIcujWKlEU4SuFEJZUa3zfEOsp8hi9tB5YLyvtcwuw6ZlJd2Ars+Pm8fyC0AsUKHBkKFzuJzxM36vcXe5quNtJ4DbAuco9wCBsE4sGBBaJFKC1BDOKJ4eRoY8xirgueO7Z56rLly1dKYQ6o1IbPqsVcZawasvS0WBdGFaXY2VorG4fy1qLI/HMIhbCeaQzMjXC0uuyPjxIC+ZQpWQHIfNsdM5PIefvx5iDk70ENDp/6QPLEGqZtuLCsFzCJJbq4ApTq3oHLr9s2XPxXGubkPa+mXrzAeDxMAx3x7FFCYvv+ySt/ViVgpoBK9GUEHrILRrUNMgWNg99yCzb3xZ/kgUKFHhxKO4eJxV6rV1nr3ttcRaDB2YQrEQQAYmLM1uBIcTYCju3zw0MDgytrVTKW7HB+UqULtq0afXmNNVrrLXlNIHAD9HG0Ixa+HjOiyDcg8a58hutFtVKFW10z5isoC0q0zvuQxO7EaAOudUh9vEiv5cvAPK5dB4HiTUCY5z1Xa0OIoSQacpST5WWapJLjNXfXw4twH5j0se1Tu9OE337/om5e5549pnnLr1ilcZm8rQ2QJoahhRs7ObJBvTkMbTXHOaEqNsvUKDAyYPC5X68cSQudysxWNI0RSqB71UYqKxEWEmSHsALBjHNNYT+EppJnQcffECef+G5q8HbamzpUqx3aalSOxvrrcGUwp71XFds22WFu5rtjspatpntJXDRZwkbej/HegvEzbu27yMtdUixl0NQdjaeg5Fh/zF7jt9mVNn3mL3q+arpzEeX5yCLk9fBPN6K5m5HpDcJmdwlMTsEPlLOgf8CntdgZrpBkiR4UiGkRnmaKK5TCjxarQbbtm07+PkWKFCgQIbCQj/RkWebt9XXFJ6QKOFjYg9RWspco065tox77n7UP3Pjxo2mVL0wCIeuuuzytZdZq083Qg5jSlghMW1CDjCCebF3KQTOgZ4llPUT7JEYjVZ2xr0IVN+C4KUSus4HeDDr9mCfiX5CX+x4eW0988IA2TerWO8i319+EfAziNaUxNwfRc0vxbp14+MPPvvQhReeHpnUx7QSrO+DimglEyhPYI2Yt5gqUKBAgYOhsNCPNw6Z5X4BmABLAKYEpkaa+Hj+II8/tU2cd/Ylp0lfXVYp114VR+byUji8RRtV1qnBK3ltOrICrBXOys+C4VaYLKtagnExeJdlbrLuaQbZ5wR3meJd1nwfAZt5HU8OnhQ3z0I/1IQcIob+UpPiVE8aemceOp6G3pI/i3aCOV2QtuOydzkN0iUnCoNCoE1TS5Vum5ne9/mSSP/z+aefuWf58lUtoepY/zmMmCJNDYEf8PAjD3HSt6UtUKDAy4KC0F8yFiKsI7kBdxNEN2E4gj3jjAswtuRiraaE1dUlg7VllyAqrw6C8jWeb89BqJrnlYnjFKMNKIuxlrBcotFyzVOscPsXWfVYXgZnlSYndGm7xiJSJH2EbvMM9m5C7yXQHkIXBkF6kDI0Qytq9MyXEKK3tEv018kvMH9WIqXo2aojQJMTamdcVnQtRGxvnX33gsIIAzLubDfP22Cw0nR5Pegbq8Rqn25rP/Q8Ny9CYnVCNDtJKVCPpqn9b+Tcv931wL/fc/qWFdoPKhit2bbtPlzme/6vfz76NQT63ytQoMA3CwpCf9FwN8/zz7twgc+yG6owDA0N9GyvPEGr1cD3BUqFJEkVq32kcvFqa3xMUkLYIaytcOvd94rNp59/xtDQyLVBUHoD1rsM663KiaytgtauY3Y3fCNyq7xLyVzIvqSvfKwyO7ZFtZuoOAlWhXTJcGlKkmhXAiecO1gIiyXqsZqjJG7vWwmNZxt4yuIphfI8atUqvu8ThiGeD9WaIiz5lMtlwjBkyZIlHUIXKaWyW2DkSWu9SnGSqGmJIk2z0aQZR8RxTBRFtFotoighbklarZi5uTqtVotms+Hc5IA1Cp34+F4N6flYKxDZAsZakRG/zeXi0anNXPpOEU8pibHW1aK3PQNZlznpJHCtDuksKEwnxEF3NznjvCciNYjWvY3WzMebjea/Dw5UnwtkC0kd6TUQqomxLSxJ+3dkdAMpJFIEneO0e9PD7bfdvcDvs0CBAqciCkJ/SZCcf975Pa8dOgQ3NFylX/NcSE0UNdGmTMlbQ7MuKFd8rHGE/sB9T4Rnbbn8Ai8YeUOqwjcEYfmCVOtSr3s7WyAseAkNRpp2KZkRvd9pW8xWuqS2dg20y6zOW5JaNLSz2LOsa6GxxumpWxIQMdWBEqMjI9RqNUZGRhgcHGRsbIzhwRLLl5QYHAgZHRllYEBRrYGS4HlO3FV5ICTIzPhNY3rQH0bungKZq7JZMLY3NG6tSyJPE2g2YLYBzYal0WwyMzPDxMQE01N1xvfMsW98mokDk8SxZmLfAYwWxKkhTQ2pAd/38ZSPlQIhPFrNCCE8SqUSWqdYI7J57iLzbH6t8ZHWudyN7LausznOzif3bDiBH7DWTs3OHPhcLbAfuvu+r9x05eUXpFbMopnDmjpBCGlaR9sYafuSF/N6dmEKQi9Q4JsIBaG/RJx7wdauV/OTwIaGhhyZG6dJLv0WSdJAipAkqSD1ajwxwqOPPlbasuWsS6KW/taR4dHrtZZn4ZWkFh5RkqKUWjBzW2UWZad0rasRCllDE9HxGECvWpk1Khuvs+qU0KQ6IU4aaB1Tq5TRukV1IGDp2ADr1y1neGSAFSuXMrqkxsbTVjEwUGV4uMTAgCNSKR1pH4mkTHeJeHeeXL/zuIe0uz7sJnohOvvQpvfznn0btz+l3LqlUYd94wn79k2yY/su9u6b5Oln97Jr5wQTE1MkqaQ+l4IN8FSJwA9JktzDIdrk3O39cAskR/JWLOQMl9l52fYAjbUuV8EkBFJi4+Y9qU7+dmp67yeWjIWTxkxRHojQpk6SNnHBhtwq7z3CnbffOu+IBQoUODVREPpLgYBzLzjLPW9noQeuHjwj2OHhAYwNQI+C9UjtDNZqfK/MnbdtC847+7rLPW/0rUqFbxTCO9tTPlK58jArwAqfRDs50WaziZS9NLk4obtXzqruaLa3XbxZeZkx7nNrnTb58FCJsSXDrFm7grGxJWzafBqrVo2xak1ItQbDg7Sbklh668YNWf6Ype22lwqssMg2vQs0ncUG4Nz3iMNeACxE8u6fxlfdI7LtYy1Epc674TnrWruxCTLyt+4SJgnECUxNwsS+lG2PPc/253fz8MOPc2BihtnZJkbj5OethwrCnkVdf0vXnNSFFJmnoy+pMHuppERY502RgNYgSZ9vRVP/0IjG/7ZSS3YHgSZJmy6xUbaQOHc8GHLdvrvuuG3Bcy9QoMCph4LQ+3Go7ll9n597wVmORK0HNsAaH2tLaC3wlEd1UAE1RLIWTA1jYhDmAqx8u1LhtwoZnocNEMJ3sqBZNjSilRF62HbXW2vbxCAzF2tubVtt0Kl24izKZY/rNEVIQRB4COk0x6O4judZfF+hfMOq1UOsP20555y7ldM2rGTNqiXUBmBgcH7F2uIFaLb3cTFz+7AguibZMv8C9B05S9Jrx6h79tM9tn4sIH6T5yF0fcVIek4vX09FsQsPbNu2n+ee3c0jDz/J88/uZvvOCdJI0UoMOlEEfo0wKBNFEdJXCM/DGI3Rpi8ckh1P0G4+A2SknuU9WFdKaEw8HsXNf9RJ/S/8wG7XzGLFLFI0EF4dIRMkLiHxrju+nq2yDjqNBQoUOAVQEHo/jpDQt55/PlZkZV+2hNVlLGWsDpFKMjAckCYhurl0jU4rbw2C4Hs8pS5XnhfY7gxu25/UFGMErqnKgsIs7r3mXIuBimth2ozmKJc9jEidbZrGGJMiSKnVSgyP1Djn7M2sP20155x7Jus3lBlZ0nVOAtQRq7XZ+c9fMqEvtv8Fjt5nAc/fz8G+v5Cinez9Sk+WfWdJYzRoY/F9AZr22uPRR1OefmoH997/CM889QI7dk4wO9Mk8Et4YYkkFRjRFZrpn5/+11aS5+93qgXc97WNJyzRX+3Z/dRfDg574+WSJrUzIBtIYqSIueeum9zvqSD0AgVOeRSE3o/F+GTBmZKcff7FWVlZCUwFxSBKDVGfM+zdMx6etnnTqz0VvFPJ6hukDEddGZi7mZt2HbjbF1Z2JTdl8e6DiotIhFYoBELEaFPHyBbaNBgeqbBq1Qhbz9nEGVvWc/bWzSwb86lUsyS07JS6beAja6OyAHJvQRd7iCP9ic3b/BDu4tzEzQn9MA9nFtxvZwbyVucyj09DljyY1ZfnmXj59cknUzv3vzGQanjqmSbfeOgx7rrrYZ54aicT4w2sLWGpoGTgYu9dlnq3a75/xL0VgrkGfEwaN3amOnn/U48/8sHNW9Y0ES2QMUrMcN89nwMahzcpBQoUOKlREHo/DkXoPdZyiTPPvRpMpS38kiYhnqxurJSXfFepPPB9QslzXGmT+9eJtWeWnuxOZJKZuzV3sXZlTJNbaIbumHgat9A2ZnS4wuo1Y1xw4ZlsOXMDZ29dxbKl4JdAp45kPAWezFiHBBAu3o/MitMONhGHgTy2bk9GQu8l83xXbULvue55iaDoBPCNxdApe8s3c6VvEKcwvgceuG8799z7KHfd8RDTs5ErbWs31XE6Ae2yOmRbiU7a3lawbvEnUEJj0C7ckqYPxknjN4XUn7FEKHmAB+77BI7Q89byfW1259W1FyhQ4GTFNx+h953xK6+8MnvW1UBEGESuiptnhmdEKwgRKISSWDNE3FqBlKPcdtst8pxzL7uuVh38Maz3ZpADpk/QpBf9WuFZty0pSJMUaw1KeXgSjElJderKk6TGEiFEi7Xrl3DeBRu44KIzuPTS86kNgK9Aqo7bvLstSoeSFnYjOxzUJdG3TTcW2f4QIYvjjx7/+mFs0yusc6ivm75N4sjlOOzdk/D4E9u56aa7uPfep5ie1Jg0JPSHMHhorUkNWKEQHu0Mdrkg+XaqF4wxJEnyr1on71Zq79OV8Gk82UKKMZLEYMWM+w0ZBRhuu+OrdFq2FihQ4GTGCXd7PeboJ/SrXuGe9MWpncIYOJGWzHK2AehBBCFWKqypsecFf2Sotvpt0lc/HgTBFTaLH/e7UDs77rWy8ti5c6lqIG03R0lNC6U0SdzA8wTVcsAZZ53GheefxSteeTErVgsGR5wFmB/l4Mrph4NDEfpiwdhTgdDhaA+wn9AFkKZOiEYoDylh7174+s2PcdOX7+SxR3eidYBOPRAhSV7FIDWLWtLdjWGMzL1BE5iJ377nno//1RWXnGNr5VU0mnWsnALh1AOFtNx6+5cpCL1AgVMDJ9zt9ZhjHqFf3vUq1zSXru0oTl3NZkpq6BqhPI04rhBF0foosT8UBkM/7Kvy+lgnDA4O0mq2su9lh2u7gvPEpi7XaV6fnsuPihRPpsRpnUTPYqmzfFWVs89ay5VXXcY5521hwzoflblxbXY+BmeRd+eHL4pD8ldB6EcT/RScphrPU67y34AnIc6OajQ8+VidL3/5Xr58472M721iVRVrnCpfZ3HYvWxb3GUuTYJM5z43OzPx8+Wq9xSyjvRmEaoB0pW73Xbr1xb9foECBU4unHC312OOBS30/KboIWwli2umWBFgTc1lr0sBukZrdvgcIYZ+2veD71bKGxXSuhKkvJ92dtM1WSy5p25cmK5M5SyOaUoIK9uZ7WkyxZKxKpdcdjavuOJczr9oDaNLoOS7G7/VjgSEyJLJ+0LHnQj8IigIvQ8vH6FbbDvbrS0lb91ziaWZaMq+R5rCxDjceut2/uXfPseOneMY7SOF31X1INtaA/OP5CCtJLAKTDLRjPb9gvRbH0U2kGIO1BTIOrfddhOFhV6gwKmBE+72esyxUAw9k8oUIgCqGKOwNiaOS1T8M0jiGmliLk2N/aVKecm3WbzyYrvvriu21pJqd7P0pYeSHmnqYrCVsITWMdrGWFOnOmC56JItvOpVF3LGWetZuUpRCnBqaF2mt0W7RYFVvcfKnh/S5f5yE/oJj5fXQl/w2uTVfoKO+p11b8cG7r9/ko9/9AvcfftjYAYIgwGiqIn0D11gKIxEYlBCk6TxB7Zvf/ZXly4NW2F1mqAyyy03f7qr0mL+mAoUKHDy4GS9C794dJ+xlbzyyqsdoYsUTAlrR3FqbwkmHSBuLr9aqSW/EobhG5Xv+2myWGet7J2u/edNTLTWKCyer9A6JYqiTC1Oc9HFp/Oq6y7lulefxugIoJxsamIhyEuhROd4Ft2jRGb6LqE61BU9LEI/0oS4BXd0kuDlI/TDWmh1Hd7g6hG0hagBjz9i+deP/ze33/kgSZRSqQyhE4k+xJi7y910rG+dmR3/gcHhxjNG7eHuOz8NYoGytoLQCxQ46XCy3oVfPOYR+qs6r/UAOl4NpoZQ8nJk6d1QewvCF4bExdIp09E+Xzh+2d0MRWiB8gSImCSdQXoJmzav4Zprr+Caq85j65au71mnK2JSkF6uDJcXd+dSomR36K4s9a6iZfGSi8kPhYLQj9rh+hMke97PP/OILPjCqdOVArjn3hb/9JFP8uD9T4Mdwmgv6wLXgenaV3f7Wik9tGnuPDC1850DtfrX7r3nX4CZo3Z6BQoUOH44We/CLw4CevtfB1xxxRuyTHPA1IibSy6SDN0QBMG3I5UnggBrNZrE1VebCr2E3g3TboqS1w5LUuKozoqVQ5y59TTe9vbXcfoZNSo1KPu5izxFI9q67Emi8X2FtRohco+AG3fbJdt+vyD0kwptS7y/p3n/hXOfaxSJdu1jAyHRQBy7Nd2ttx7gnz7ynzz77Dgm9TBWgZVOuTDfw7zfaK4TkDbqjV0/9shDn/oYTAEtOrH0oja9QIGTESf9XfjMM852T+bdIN1NcsmSMQCsNGA9ElMGUUJKD6OrSLsMJYaZmpraPDgweoOw5R8AL7TGusx2KTrJR8Iwr/GG9ZBS4nkusS3VLVrRHEqBtTNccP5qrr3qfK697kpWrvLa4mJCghQWd4s2Xc1LZN95dGOBCPlh8uvh3p4XWw8YbJdPwPRt1/1KdM1+JoaTe/FFZywd+/NIfbsv9id7OPX1x/L789UAet/PD5OL1nTPdgfdugIGSCP4/Bee5Z8/8ml27ZjGk4MgfKIowguDdrJmu1GPlZh2nD6ygpmfR0x+QHrjCDGHxLX3VdI1efna1255kedboECBlxveoTc52ZDRRbshhSs7w0iMLWF0hUp1Ka0YrKnywgvx2OoVQ784WBv6Ga3NkMqy0oVQTuWre9dZ33Bp85aZCotF64jYxEiZEKczjC6psPWc03nzm6/lqiuWE4YuLt6MNL4vENKiMJgeMnP7zBXMuouSunEsDfCD7zsP6C+EbrI6VAx+of12owjeAl3EPp/EoWuNpOCNbzqNSy/5Bf7xQ1/ic5+7lVZDUimP0Io0KvCyBanM4jUSgci68IXCMvIXQvoli3k/0kMwhxAtDPqY/tYKFChw9HGSW+iyy0J3SlpOHz0jdeuxdHQ1mArGVrG2RqxDpBzk3rse9i++9FU/qlTw657y1wlLW57VdjUXET3NMjqKXDn9JUmDJK0TljVDIz6vfe3lvOFNV3HGmbV2CVn3Hjo358Xcmge/jcpD3WYXMSTNPDtvgaN2c3F/47LelU3fRv3WZ/54LH5eh1Cp6/nYLrDJkYzp8BXxFrO/529xKBz8+i60wDNAFEEYwM1f282H/v5TPPrIHiqVFRhtM3+KbGvGW2NdzF04a1wSg41+Vcrp90u1CylnEJlE7C1fu8kdIT/tYr1VoMAJi1OT0LPPsB6jS1aCGcLqpRgzSCtRCFl9rRIDv+d75csRKcakBJ6H0b23yx4yz1tYdnsAREwz2sfpm1bx+uuv5dpXXcSmDS4SqTEEuWuexXh2sZv84jf1Q1pNh0Xo/fuTnUEupAK7wP56N8i9IdljWzhHAl5vkXwPXqr7+wg+Ei/meAtduROL0HXqtAmC0L2OEueG/6u/uonPf/Y24tTDUHLaCkZhMmLXJCCclDCAMBYh5n5Jqp1/KuUUkhghYm65+UZcvkh2wILQCxQ4YXHyu9y7XJPutYcQPlEkKZUqKL9KHAVIOcLsrLd+cHD5e42QPwhgiJEWpHQ3ObFAq08hXJzck2WM1iBSmvEUoZ+wftMQb/m27+Mtb7mAShlS2xU7xiLQ5C7OhXEMnJqLHGp+85VuGPIGI0gPMCRpgtEQhiHWgM0y79v18Nn2IpVgFUjj4gpp7BL20si1G1MlV3eVxrR0QlCtIHyF8P0scyDIR4DJlNO6RtUlvOJei8yLkqumzZPW7fu8o+BCz6Ptc1jkyYSmuypRdAsDtbMRDzKPh0PsLw39+5V5875MryCQENbgV37pVZx3/gr++M8+yv4Dc4T+CEmqUCoEK7LfusDiITK9GyHKfzI7He4rl0f/X6ncItVTbucLnPJiIaECBQocP5z8FvqWc9xT4TJ0tZUIW0PKIVqxYGz5Uu6+81F10UVv/elKZdVvtSKWAljVwumnS6QFtUgTlZzspYUkjREyYvOWVXzn91zP1deMUa26m2jvjby7rIyuT47ndB8s/i06/9LEkaAQmS83AS9wz5MYPM+xO8Yx8OQcJJaZmf0cmByn1WoQxXUwGq9cxgsHCAeGGBgepjxQxR+oIgIFvpc1wPF6h0JG4IK8G2l79NbO79U+77L155jlC4JFGDaPrnTljmU77ns8CFUb5tvxC291ODjCpUC3E6FvN0kK2/dY3v9H/8hNX7qfoYHVCFsFpKvaEPR4niQpkril9fT1UtW/ChPcecd/9Nap2/buj+isChQocOxx0hP6ljO6CN16CFtF6wF0NITnj1CuDV+mvOr7PTVwtfTKaOOS3ZAtF0M0EmkX74omAW1aGKY55/wNvO3bv4XXfMs6lHSfeRZ0avG9eRl0zJcVOREIvf8WLAANicnMZNFRM0lTmD4A9Trsn4SJfczu3MmB2Un2HphgX30OWxtAVmsML1/O0IoxRlevYmD5MmpLRyCogCrTbsCesUC++0CANl0knZGT0e79FJiac2ppaQomgTjuZS/bVbIn7XyC93y3BvElKA+qFWfV+h4on3begLGd7+cegU7uwMHzD44OoXd8OwfF4bq8TQLKJwXmNPzp+z/PZ//jNiQjCEI0GivAZv3cJQZlDcJYsMnOif07r126jGfuuP0jIKc6xy4IvUCBExYnEaH3CW9k8dktWy7I3mth8PDFMkwyzJ7dcXnDhvPeXR1Y8qsIFarAp1FvIlQFm8W/EQbVtgq7bT8DMkaSYtIGa9aM8H0/8BZe94aNVEJHNAqn5kZiCIPFbsJ9AjAv+3R3k7hxfmaTE3f2XEhoNmG2AY0Itu+iOb6P7du3c2DvXvY89wxJo4HWKUp5rFyzmhVrVrHxjC2o9Wtg0yYIKxCWnU9eeWAF2oJW4Ae9o4himJmBqckms7N19u+dY2ayzsTEBNPT00xMHCCKIprNBlGsmWomxAlEUUSSJgjt5jDvuW5sb+BfCuH6yAtBXqrl+5IwDPF9n1KphOf7hGFI6BtGl5YZHKqwfOkYg4ODjC4ZZWxsjOVLSwgJS8foSQqc5+HvOjdwxOiSIQ9RnrYgFqpH76LOHvGZvu/ln6UafIWOGhglEF5IlOVIfPyfv8EH/+8nUAxi8NBWtV0T0rpkT4VCKkGzPvu1Bx/+6uvS9OEYDuB+9Xkv9U4xXUHoBQqcODjhCf3yyy8FJCYXfxHGETISTIVQrcTzFandi8HDRmuxycgrBeU/tyK8GOFhZXbjyhYBpm0JpUjRQgmBNook0SglkF6KtgdYvabCd739Wt74hisYGHTuW98/yGAXrNh6kUpkJvuuzF/MzxzvjilbaVF5Xn2a0iRFeYoAD2wzM3ENNGOYnoMdu+CZZ9HPPcPsCy+w79nnaU7OMD09y6xOmfMUQ5s2Ud64iaVbNnPauVspr14Nq1dlfmrjTFzpA8oxnQdkSVrTMzB+AJ7b1WT33gM8+8yz7N67hz1791Kfa1Kv11FCkMSAtm2Cxkqkkkip0Fa0RVKEcNevP8/BVSR0iE4I4Wrf5+kSzIfA0GzWez4PwzJhGFKrVimVPUaWDLJq1RI2bV7LitUjrF69lKVjgtqwO+UguyIG0LTAJiir8EQZa1yJohALmNVW9v40ZF710O8NyPromUz0Rapsv87TkVdStHdrO0sM52yxWLx2FcMXP/cMf/C/P0gSVYGKi6nTpU1ku+ey9YeeP3XD1PTTDA4mSNXEpk2XTJclIdx869cWnNsCBQq8/Dh5CB3PqWBJZ40YPDA1auFmwqDMbPN57rj9vuAVl3z7u9HD7xbC8w10hGGk6ojCZGpx1iYIEaMUbt8kaNtkYEjwve98M29801aWDEOQGe/WsLgS2yF1R/o3OMQX8hvzPKWvbks/28bozE+tXeBUSjAxNBrOt31gCv300+x/+il2PvII9V0v0Ny7lzKCuFknbUZ4SAZqg6w+fQurzzkbXnEprF8PK9e6GLrvQ+iDcLvMicRKmJ6C7c/PsnPHONseeY5nn9nF9udfYK5paSU+sfEwRreT1aRUCCHauQlCuNciW3S1ywf7kt/cdxYi9A7EvGA4fa+zvP58/WB7v2+MQEoBSiKFpRk1kFIjZAKixciSCqWKZP3GVWzeso7zzjuTjactY+UoJIkrHcuzAjq/l0XsWNP3Y5J9hG5xk62zFVLccjGEgWFQZWzQa7S7r8xfPFjjytWwEgXc9KXn+T9/8GFmpwWeGgDhEhRtXq+eQ6Q06uNv9IOZz6FeQMoplNBIqV2YRBhu/nohPFOgwImCE57QL73i4uyZh81IHethCEAPU/Y3onWF2dmJrbXayF97YvQqR9iutMyInPb6a6QNQlqk9BBGg4iRfoNXfctW3vmDb2XdWrddgIuRy4woXry06qEk3eZ/3rHLF5OXyUg8NS6ZTSdwYD+8sAseeYSZbc8w/tDTJBPTxI0ZlLAkcRMrU2KRMqMkYtVqxs48l61XvBL//AsciSvpkgOEABVkwWa3ZkgsTE/DE49O8+gjz3DfQ48xvm+SiYkDtJoxxiiUDJDSd2VS0kegkEqipHDZ88ZgjUFIicz/5W1nu8gcOkTensWjROj9OROd44u2FoG1Nhs3GDTWJrgyx5hYR8RJg4HBMuUw4Jwzz+Di88/g0kvPZMuWoF1GNj8G3wdLT6WGFp2tfWOczuv2nTz07ndj9+1j4/nnM3DBJfCa18HaZSA7iYOdHXa/FCAsFoOxAiUkcQsefniW3/6NP2dqQuHJQQwS0/XjtnnDItIndu998PKx5TNTyptEIbOwgvMYFIReoMCJg5OE0GVG5s6yNrYEpgJ6mMBfR9wo/bgXlP/QGjFUCrLOpiLuJP5kEq098UiRIkgx2mBsk1defR7f9Y7XcellFVLrboslAVpb/EO2MDscHBmhd+hH0I7FG+Ncrzp7rrN/Tz9D9MBD7H/qaZ5/+CHM/gmCA5MMtCwDkSQQPtqTzIqURiiIawGrzzuTVZdeAte9FpatAhRUaqBUx4+cDaTRgm2PzXD/vdt4+JHn2LVzP3t27qfZMqQEBH4J3/cyS9t5QYzWbojWzs9P6Dn9vtr/PqLtf320CL2zv34CzCx45RYZURShlIfvK5SnaDbrCCGcK9tq51I3GpFqBDHVquWMM8e48OJVvOMdb2Jg0F9gLH3js7J9XJ0NTxmg1YIdu3jy936P6Pa7CHXKTh9GrrmWC97zm7B+7RERerZnsNBswd237+a9v/E3mHQISdj2ghlp2oSuTUqaTrzfL+39VV9NYk3D1ahnvQwKQi9Q4MTBiVuHnt2Xw7DMXKOJFL5zDWoNMkDHA8xNi+GVy4M/LVUrP4gtOXetsFnCW4zEoBFZeZQC6xG3EsrlgEarCbLBxo1LeecPvIPXvXENvuduh4FwQiwWUGq+lvbRRe8NuNVoEAQ+Unrunu9lGVlWu38zB2DfBDzyOPre+3nmGw/T3DNBa/wAgRVUpaLsBVRlDRO38IRg2mpaowMkq1ew5frXUr3iMjjrjKwMreLi4J6rN06049hmA+67f5Kv3vwwjz2+nSe2PUEaGXy/hO+V8dQwpaoizSZHW8A6rwfodsK2NKLnHKXNI87Z945wtuZVI7zECzJ/weAWBdZqtDX4ocu10NaQRnFHs0AIhAhcEp4EGaQoWyJpWe667WkGBqBS8rN+9naB4+XF48b91rOYugKXexbF8OxzPP2e38I89AijqWDc91j+mldy9rt/BVYOuzQSDKJnEvrmR+T/ifbxrBCEZcGV16zkl375h/irv/g4jUZC4A+S5P1ZsmtUDis09dDP6Yb30SAcvt+yG9QUEANplvy3mGxRgQIFXk6cuIQOYCVz9QTpVUgS56K1poSNqyg7esnI6Mg/IEvndiwcCZmYS+dO302YhkpV0IomqQ6kvPVbX8MP/ciVDA3TFktzEWrTzlY+dhIhC6NUqbgSsiQrxp5qOVd6fRK+8SDPfv1mpr/xMDPbtjFsoGQsI0ZQ0gphJZG1iFKAVpYJLFEQsOrCc9n8rdfDKy+D0WFXu1UKgao7aOwOqYEnnpzllq/fz1duvIPnn5/GqlEQZXx/JaHUjsg8SWptLpXvhEnsYtR8ct3qe5PuJJ1fBXT/pqQlS65UnVQGUhqNKa65+nJ++73f5RL+j+TgFmhqEBqefJzH/uB9mIceZrkK2Rk3WXn1tSx/1y/D6BB4gkY6S8mrHsaO+8INWCd2pOD1b9rA/gPfwt998JPOCyZKSOtlpXgSbUCKcgj+e00i3ooXOG+XSHm5/zYKFChwcBx3Ql+0njVjCyEGiJMUL9BYKxFmKZYlPxSGS/888AYHdH7DlWmWLOdn7vUs5ZoEwGXlyhat5ABnnLOKn/nZ7+DiS5cigKhlKJUktq3uRpvMj/ktq0vXRWOdq1ULaBqYqcP9d5F+/es8++CDTDz7LJU0ZtBXLBUVtGnieRKVqd0J45HGCU0TM65g2asv58LveAe88krnww0DZ4n7CrBuyRJDK4KbbnmGT37q8zz73D6aMxaReFRLy2mkrgzMFwZPCqyCVMRoabECpM4S23KysL3SL7p9ktnp9kcv+l3e86RjjheyRWI2YCfCkmVjWI3J+4zniWTWA9nirPNG+KUbvpNKDZI0wlNH8CdmgcY0TE3y/B+9j+Ztt7NEhGyPm1SvfQXLf+c9sGIJhCU0Cs+zOTUvgsVyLxyEABT84I9ezN7xcT79qa/j+2NdmXYSnUg8FWJI3xLr5qsDYb4CxmXuZ56Hk2vJVqDAqYvjTui96LeEKgiWYo3BpC0euvt+78KLzv2DsLLiV5QsE0UW6cseL6Noq1+5Uh1hs+YTsgGywY/9xFv4rndcymCVrAs5lEoCbY3zbveM5WjQeW82u8H0ybBmprEwqNRAK4Fd4zRvuo0nb/wyPPUYweQkFSRnhgG+8DFxSmo1GomxyrmIkbQCj+lAYVev5pof+G54/etg9ZpOOZ0A3UhRoTt+vQGf+8wjfP4LX+eBbzyJVFU8NYDnlcBamo0ULwzIFUWsFWhjHJnLfkm17lPunrcjdap37aar/rsbC1WBHT10WbPCtIVX2t4fYTDkTYBAWYUjuBZSzvAb7/k1NpzmBq080eOMdqVjeUF717sWJ8oepzB5gPt+5/cI7rmXsSBkn/FZcd1VrHzvu2D1SvB8ZtMI31OIee0AFzqP/HBdCW90ftlhYDFW8GP/3/U89NDD7NrRAtHJN7HWYpWradCJ+B9al7+CqLiFr0gLMi9Q4ATCcUuKy28oV199FeQlM9bViFs8DB7GDKLEeqyp8cLu7UuXjC7/UBgMvdnVpGc3nLwmPXO5pzpBSkmUpATCUPNAMss556/iR3/ibZxz/kjmV+9VcmvHRY+abd4tOeIc+LkxrvKPjXZJaPUZiAx8Yxszn/ksEw88yMQTTzIsYcAKPGMyHjCobDngG4O2EAFa+iQK9pcCVl3/Laz4iR+FpctgsAZBRsg2BhFgtbO4P/ffO/joxz7HM9vH0bj5tEiMUcj+Wqj8mmXxZYPjJSM6giTZFosq7uXJb7KPjLvrpq1xJYaueYhFZ7XOSqksG97LdmVJ004ZnJCinaWewxMSISRRq44QkrwtrjZZrbqnkEK0s+oXStbrUVKzsh1asDIjdOsRUALdQjDOH//RDVx6SQ0R5nXpIEja+1bQFsZxPwK38ERbaMzC3j0885u/y+Ttd7Pck0wgKb/qWzjjF38B1q+EmlN4i/O5Bvw+C71z5RZOAuxHvlWcwoMPvMANv/zXpM1hkKVO0ElKlwAoYtto7rvaD6duleELKG+W2266af7cFShQ4LjgxLHQrczIPMAQYGwJa4YRapR9extnjQyf8S9BUDrXZARgF+GNUqnEXHOacjVAtxoopfn2t1/Hj//UVfiVLNab3YxBtruzKVx9m33xdWkHRe/t1bh6cZ04wZcHHuSpT/wne265k5VRi0q9zmkkBNoirUJbQ6KcJ16IDikKIUEFzClBvGyEi37g++Btb4KBKqZUAS/IDHNBKgI8BE890+D/fuBfue1rTyHEIKI02HfOdr5bPD8HAapblS2v5e72kByhKKjo2oeQlsCTmKhJyRpKUlOyCTJJOsRv8rKy7Lp5Xs8iQiHajWNSazDWYKWH8H2s8NHSLRZTLdwjqi00ZEWn/W6+gJKIngWLG6xb7iljCFqThHaWd73r+7ns/Fr7L8ql0lks9uCUqq37LezZwxO/97vM3X4ny7yA3WnE8GWXc/oNvworlgEpRBE29Nq+Hpmd8YtCLuGalckFXsr556/i27/9tfzD332WSmUVxvgoz2DRgIe1gTBm4P9LTXKrooTq1ngvUKDAccdxI/T8dt8mkyxmbggwpoY1g1g7xOxM67rhkZGPBYG/wmjT8Vj232Q71bt4UhFHE5x5xnJ+5efeySUXD3SSk3Iyz1hLCuWSkI4a+v3BWRmUqwRGoKBZB92CbY+z95/+lb33PoTZe4DVxjJgJb7nYYVEpwkm1W0/irTZsAUk0hH8ZNLArN3AOb/yU3DtK6A6BMpDegEaaMUJfuCjENz4pef4k/d/hKn9Br+0FGMDF0nvGrLs8ObCZ3cUfTrSgjQSI1wOIDImnZtiNBQskZbhZIrl0Tg13WzXo6vsMe+CZrTpKT0r+QHaGNI0IdKWppI0tUcz8WjZEnVCElkjUWViUSFimESU0CpFS4NRBoshFSnSSkpatudHC0gJEEaiDNT0HEvTfbzpqrN41StOh2YCtg4DAeBjTEoonPejnezf3f3NpNCcht3j7P/999G8+U7GvIDdOmLo9a/m9HfdAKvGshMroZMIZTu7gIO52A5zYWpzUpf4PnzvD1zDrbfdydNPziHVIL2eJvB97y2pUes8ytsFZTrZ+gUKFDjeOP4Wus0FLTysLWGsI3NjBhG29l3VSu1DSWIqeb1xr8s2c7ULV45jREqzMYNUTd729qv5/37sNSwdzsq1aRGobpW1lxMua15inAjMzCzTH/tnHv3UfxJu381y6+MlhsAIjDGIkk+qIEpSQik6FrmUbXV4LQz10GNu2RiX/NJPwauvhKEhmvWEclhzX7Dg+T5pCv/8/+7mH//xs9SnJaVgDM+rUm9GKAFSHL+yI5NbwrIBosG3f+d1vPbyCwhaczz0hU9zVuU0RmRMueIkWSthCSklyvNYqIWajWOSJCWOY5pGM2cFk3HK5FzKVMswPqeZbGn2z8ZMNiL2TY+DqLRzDAySVMhOKRnO3Z57coSVKGuopTHLogOcYfbiP7yHL//8rSzfspz1F53OwJkbYcMGSuGgG5QMnBqhgnZbN4MTjdkzzhO/97u0vnYnSzyPvTKleslFnPHuG2DZCle2aLOcc89zYjddv+G8Kd5LggUpXHnm0hH4wR95O7/9Gx/C2hJtD0AeMpH+iDThW0xa+ktUhU7yaYECBY43Xn5269FRkVx19as7bnbTTeZDP2lt+S+ldB2yO27mXhfo3EyT0SVLSNMmjWgfI0tTfvpnv5s3v/UsUgOh6qSj5cTajl324PBijodGb5kcGBLdwrfC1YZ9/S4e/5u/Y+7xxynPzLJECEKjUabjOtVCOsUwAaqrcbcVYJQgtprEGl6olXjFb70L8errYGgMVIixHka4ZPZmC4IS/L+PPswH/vSfEXIURQ0yXXzTVgPraNJ1LPDFaKIvU7rPIdERkTm4gEz7bS2I0ykuu2IDv/Ir38uGTKiOBGjZ7IlxzHWoZLj+xZ7Akb6UnRPTFqKEqBXTjBOmpiJ27jnAE0/vYufuGe596En2TcUkokQSCwJCFKpLz9VjyKYsb45zRryHq73drND7aSazJCZC+oLU91l6+ibWXnAx3rWvh02bYdCHkgRfuNyJSMOBOXa957fZ8+WvMub57LYRpTdex/nvfhcs2wBB1hWN7FTyOsGe+X6Rf8K271G6K5YYS5oI3nXDx7jla49SKQ8icEqB1gqXY6Gir0bp7ldVKlPcevO/AHMvbgwFChQ4qjh+FnpmAVk8jC1hbKWLzKu/BqXfX4hc+wlkaHgAbRrMNcc5/4K1/Np73smGjT5CQFn108rRylw/Egh86+rK6//2Kbb9zT8wuHs/a4wksCWkjnpGZNr/y3kJZACNVgsbKCIFK7eegbjmKlgyBtaHLGksCy9TLsGXvzrJB//2UyRmAI8yQmYJhYsQ7MuNmJTV61fye7/7vQwNusqwltYYT+INCBQ+yhzhWHukVPMkx6zASwA1RWgDQuMzvK7ChnOWcdVrz8YYmJiGO+/ewdduvZfHtj3HgYkZ4shgEkm1PETUirDpLMPiAOuqc1Sn9lJq7qMsLaH0kJEhTpvUDzzAI994kgOf+ByDZ2xhzWXnsvzyi+C8s8BTsG+aHR/4G6Zvv5tV5RovJA1WvPZVrH/3De56qoB6lKJCl0Xf/4ea59YdzV+0BDxp8EPF297+Wu66+0E6vQWBrCRRiuBSo8unWx09WdSiFyhw4uD4utyFu1UZgjaZS0beYwl+D4A+Qpa2rWTpXpMiREyU7Oft33k5P/dzb6Q6BD4QZ7FGV3CTZ8J37ayrlOvoxgC7XBAWMAnMpbQ++A888KF/YmUjYcQqSl6J2MRYBEmbXE2baAW9lnnPQkZJYqk54+xzobIcTLl9z5XZf7GFqAV//cF/J2oNUi6XieMUK9LsvM1RPu8jhwVSo2k0Gjz5WMoF53uEJVwOBLlEkOkqkXM43LK17rI3Q5aRT/7D8EBqt/qR7kgSGFsKb7x+LW9641r274M7b3+Ym2++n9vufJhmM8bYCiNLQq7YsJ6RXQeII2hoHz+1mNhSDsqEoUBZg9dMGBUtooceYtfdd7Drw1XGtqxn7aWXUb//G0zefz9DqWGvJ1FXX8H63/5NWL4SlEuTryiPJBNt7bbDtXDVDQClFz37+WT2v+Hm4ZVXj7H5jGU8+8SsWyy2N1ZIGVYxpauMDp883reQAgUKdHB8/hqzEjVsCWMGMXoAGELY2m9Ygt8FDys1Ulp0+17epUVNRtKyhWaan/zpb+cd338Bvtel7SU6BWPzcFTrmHvrzHuQpjDXZPeff4DnP/YJNqeWirH4QBq1kAJSKbBZHoBpF8uZzrKgb7elICSRAqlTWLIMjOqKJXeERoyBW7/+PE8/tZtyZQ1xq0kYhqRpxPEm8hwC8JVkdqbOb/7W+3jday7hLW95NavW+VQGnHfcoCBz4wvr/mtf04W8DDZTv18gdGCFcvNjLdYarNUoVWpvZ3WnTA4EY8vhzW89hze96RwefzLhPz5zE/c9+ATv/tkf5rwtFdj7JDz1DGzbxva772bvc88Tz83hJRFlYxj0FUHSIogiluART83QvOchnr/vG9BKKAvB9OAQG17/GoZ/+adhZBT8kHqUUvU8rOlY4Hl5YPfv4ZjYxlbQiC14grd927fwx3/4rwv+vSjpXQfyQ8diCAUKFHhxOOYx9CuvvJJOD/OcYj2ghtGjkK7A6mFQwW8Igt+12RpDiKwmGdc5TYgYaSFQZZJWBDamVG3xnt/6fq66egOlsKNWnWPBG17/zeklzUBebQzQaZRuTOq6s01Pwt9+iLv/6m9Y5ZcIoxRPa1SW2Q2dZhzz7pn9TUvym7mSGCyRSVn2utcRvO/9MFLDGJfYJqXzRCQpfOzjd/NH7/8sYWkFnjROrt3kdq+jCnO4rvd8u0Vq1HPkNez9+7Vtz0P2Om+qIgzSGrAJkhREyrr1a7j40ks457xNXHBhhVpeTi9cCBqc5jxYfAWp1VhrkNLPtOO7B5R0jTtXsyOz0o8sb6K9tcZddh27bnStCOoHoDFLcv+97H3gQfbedR/Rjp2UWg2GjGTUK7nj+h5zjTpCSXYIycj1b+WMd90AozUYcI2FUjp6Be3shu4fSFci3NEmdUPWA0jDvv3wUz/xpxzYD8YY0kTgeyWkEjSa+7ft3v3AhRP7bo2kOrDo/qIoWvSzAgUKHF28zBZ6VmtuSxgzAOkIwowgGbzBWvG7Muti1dvAIoN12bRRPIs2dU7fvJLf+M1fZdMWCDMyj9OEwPMXOO6xPad+iDQBq+HRx3jsk59iXaQp25Q4jUEotOwnu+4XB79F6zhBBT5BCs/f9xCnP/UMnHsm0lfONs/yx5SAjRvWUCoJlDRYnRIEIUmsERhUvkDIS+GOE/JEL2tLaJwH4vnn6jz95Jf4l499hrFlHmvXL+HCC8/lzLM2sfXsJQzU3DW3CNcUBq8dXmkXPeRov5GvnPL6MTgSOszrACQ4plUAgWPAWgi1FWCG8NcuZ83rv4U1O3fDM8/RvOMWHr/jHvbtGKeSwAA+rbDKlInY9MbrWXbDDbBkFMK+ZLfuEfav9mzXyI/BtZPSXZexYTj/3LP5wo33EfgVfF8Qxy08K/F8f93gwOC6iX08efRHUKBAgReDl4fQMylWd3OVYAYgWQF2CdjST1gp3iczxS/lKnQyFTKT1e9KsCUgRjDOJZeu43/97x9hcAh8z93T6s0mpfJLjigeIRYmBIEGY3nhKzfR2j3OUKlCmqZYbTC+cx8vWs+9iMXcI95iLEopGjt38tzv/B4bfuPXYetmRCkA4VY3QsJZZ65kw/qlbH++gVIezajlFNcsKGJnpYtc9/54ueE74QJjLY1WC6Uk0vMwacieFxJ27dzBXXdspxwqRpcMsmrlGFu3nsGWM9az6fQxloxBbQCiFEI/S2rP3PMO2Y/EOlUeaUEcoXU+byubHSS/MEqArAAWVAynlWD9OsrXXs4Fe/bAo0/zzC33cODex9i3f4LTXnMFy95zAwQVCAQ2biG8cj7aruMc2WweLUgF0sC1r7qcL375LpIkxmiFlApjDVKqarlc3QIFoRcocKLg2BO6MNnNLwATACHoUdDDYGtvxwv+MifzhSCdjDgumtrildecw6+9+9sZHIIgSwSL04RSueTU3o4zDCA9H/bvZ8cjj7HULxPP1vHDoC2GksumdmOhjPaF4Hke2mg85bFhdCkvPPAQj7zr19h6w8/DVa8EZcEPUQLGlgh++RfeyXve86dM11OUNwCinLnFJdJKdF7edQIgl2JVysNKQ5IoPL9M1RsFYYjjiPG9MRPjL/DgA9vRpsngkMeKVcOcf8FWzt66iYsv3sDYEregsUa4+HueHJcnyImOol17OWHcdw4bgh6Fl873BVaE4PsIqbHWR2w+HVasY+MV18ELM/DUk3DNRVAOoOJIXPjllzR3xwJSwTnnDjG6pMyBCU0YVvA8SRy3SLUB5JnAZ4/3OAsUKODwMhB66g6jAzBDYAZJojJJzJXVWvBhhFDdZJ4/d3rnkiTRIGKkaPH6N13Ge97zGoQCKVxiE0h85WXRbHvcSd15kJ311mq1iJOYIAhI9NFUo3N65+lcnZVeyMTT23ngN36b9dddxcj3fDds3gLVAUh9XnHxEH/6Z7/I773vgzz2+DglfzmCGrohqNVqzDWm8SsBSvokSdK1sMpi3oeocz7chUj/XtrHyWLp7etuFTprsh4ErsQuzdK8pRfi1MtdQptUJZpNw7PPNHnmqbv51CdvZ2SoytjSEc45ZyvnnreZcy8cYeUqd/wEJw2bJ05iXSg8TVOU8g7xx3DweRD5Tk22pchyFKSHtMZ5TxSwqQabVkFJZH993cmQR3zYo46ecQgYXQrrNixlfM92pC2TR/WFUIRh+fSibK1AgRMHL6PLXYH1sGmVOA5Oq9ZGP2qtGJCLWOYACEO56pFEdd729uv4xV9+BUI6Mhc2K/HK9NflcY4F57BYSBIYqLD0tPWIR5+gOVMnjmOk76GUQtgXfxu01iKFdI1JsHg6YZknCKfneOZTn2H6K1/jnDe8gWWvvR4uuhgCj/M2D/Gh9/8qH/33m/jUf93FxMQEnqyQmhJByXfNvoxbcFhjF/WWHA+Ydoldjqw+v9uczuVQhWFuFvbvG+exbeP812dvolJL2HLmai5/xTlceNFW1q0t4ysIfNpkqfy8//dL5M9usRYLSJBWYKREqmyM2XG1ElnzlpehRe9LgAE2b1nFHbc+QSloL4Vwlf2sPJpSwAUKFHhpeBmy3K/GJT0FoEd47mmqq9dc8CXll68QWa/oxTPTU6Joiu/9vjfwK79yBY2m81I6ozEXQe1VOHvRylkvCjkNQF5Op7GoNHJl7zd+hfv+x6+zdnIWPztXaenRHjfiIPH0Q0ITR02CSo1IChqRpTYwwuRck6RaYc1F5zP6qqvg8oth2SoYHWHXs03+9dNf59ZbHmP7rmksVYQMUZ7C93ziJO46t675XCTLfb6FvnCW+6K5AaZ3BwtlyffPj+gZQ2c7Jaw7TrY4UcrtX9uUVqvByNAAW87YwKWXnMurX3sGGzZAbJzOC7gruXirk8MsjzDdjwaEBtnVg8C4EEciRLaAyBq9tPfnHrPcfF7uFM9uGFzG/RdufJr3vucjBGoZQoE2BqEFhj233nf/R64RaqJ9Efp/D0WWe4ECLx+OMftJXvnK12AInXVuhhF2/YdMOvBDYaWGUJIobrpNbXdM00mSCtHgO7/jtfzET15BuexKsbx2yLf3znH8CL1Ttubo3TpSMDFMT1H/4z/j8U98ikqUMFaqIOpNl/jXNW6TWZuLV7QbJNodTSrXnMQahLQIYUEqIi2wqaBGgEkMqQeTNmV2oERr6RCj525l5SWXM3Dt9bBkOczAbTc/zr9/9mYefmoXjdSQCg8R1tDCc+WCVrpjWDq5EF3jFtZZy847Yg7uIXmRhG77LfT2GPIStN6adKtTXE6Gde1YjSLvrOdJaDZnCUOoDQhOP3M53/Xdr+aiS1ZTqXYIfWGL+XDrHU22zsvHpXuHne097WrT68SPTkxCl8DdD+3nV37+72jVS6jAc61UtcDYAw/dd+8nLhFqIkE2XGisS5YZIIqax2n0BQp88+Els98ll1ySPZO0zZOs5twS4JVWIRhFMsLsjPz5amnln2FLGJFihEHKjgtPGOkK21SKlHN829uv5ed/4Vr6PcAHdVHOY0XT9f9C3z1a2u3ZMqOdHGUgaUF9lj1/97c89m+fZnB8irVAGCUkWjv1thRcg5quEVmX2wadbH8pIlIBLRViBSia7n2CrEFnluiWivb3tISWhFhBrCQNzyOtDrDmrHNYfcXlcPb5sHwD7G/yxXu+wRfuepj7n51gTteIbRUhy3gSAk8hjCVOYqT0iNME6XlUwgq6kSAtaBWjpcFmF0taL0syd/r57fM7RFneghBm0cVOD82aee4CbL7Ssx3RnrYin4iJkmm2bt3A977jrVxz7TJUAL4yGcl2MvAXGdjBRpOhO3eic+6d3+O87IJFjnUQAaNjCG1hz174sR/+Uyb2Czzfc2EZY5FSP9OcHj8nSXc1VWUHUboXT2VLkOw6P/yNB/sWZC/r8AsU+KbCUSZ06CZ1bSvUBs8i0SPsH09fMVRb82UpXFqvkTGuy1YXoVsQNqHZ3Mt3fOdr+LVff/1CDbV6ML+cKHt82Qi97/Dd2dImdeIjRsNnPsddH/5n9ONPsMJqvCjCNxCoMOvmlY0mUwTzTJelKgzKJmgBDa+MFQbfNlHWYEWp13I+iP9eC0fwkQfNQDLrB1RP28zS089l9bmXwrozqSdl7n96D7fc/yRPPLubXbv3EFvAC/HLNRLjEWmFkT5SenjGw1iNVilWOELPz8HNbjb/L4XQmS9M036/f8M+N0F3eAO6PADZ6yD0aNQnULLJ9W+6jF9795uxaCp5CcUhGah9YouNvO91//n3KcYsiuND6NbAvgn40R/9C/bsNUjP5YAIY5FC7AqMOuPA5GN1G25D+PsxJsHFmwpCL1Dg5cZRSIrrJnKZKbs5axs7RKte4tnnJwY3bjjvr3XilV0rLZMlNnW+72KLKTqd5XWvewW/8Z7Xt3OLjib6Hb9HOyGpp/RJeK5QPorg297GZVdcwb5/+yTb/vM/qOzYwXJj8U1ER6xbZhnPqisZTLfd2RYfaSQSiWdd9rSZ14DGYSGaUBZCDZ4BzxjKcczM/d/ghQe2sfc/PoMNawytXc+5m0/nqjNPQ1+yiYn4LB7etZ/7n97JM5MtnjsQIWWNhBqpliQywUqn/ictKC2QpE71LVcGPAgHzY+PLzKvi833Qc77cGC0RskSvlfmc5+5nRXLa/zQj1xHFBvC4GBHfrFH7MeREvRRlTrk4PoD7toFQV5x0HLd9KxFSAPW2MTGpMR4SiI9hYkTV9mSi/mcIE2AChT4ZsAxyHLPa85DMDXC8hgrl636fUnlPC9QpKnt/LFj2m5ZhBM7ueSyLfzmb72NVLs6WDh8G+aEhRc44ZGRpYz91E8x9qY3wWc+xUOf/W/2j49TTlN8jbPYU/CyORFdix6bCewoa7K4te9c0QassPNyBxaaJ2ndMTwpCVOnylc1IJXFj5qkzQbR3DjbH7qFRigwg8OoJWtZsXIz37FxE2JoMwfSMs/uq/PE8xPsnqyzY3qKSAakIiAlwEjf5QVI1yPbXbsTpwqhH81Wi5IfomNL6A3z2f/8Mt/zXdcxNJSHkI42gZ588H3wPB9otd9z1RDp3MSB51LEAUIPjO2fq4LMCxR4OXHUCD0nnDgSBF4ZdA2dVGhF3lvC8shPg0TrzAXb/XdvNcZqJLD+tCXccMP3MlADDOisc+PRzHN7WUuE8mz83HtbroA2sHYd/MRPcN4P/zB8/XZ2fOkr7LjzXsozdQZtTEXH+J7Fao1I3CUq+YFLILQJWgq0DFxlVBaz7riWO5OVRzO6iVRkJX4ydU1itBRo42LUQkJgYUkQMCwMeq5O1HgKvXM7jbtuRhPgVwY4Z3AJly8dQ25ezn5/Kc81Uh4/0OT5uRbPTiU0VYnEei65TvjtIUk7/xZ/uJ3TDheHSooU2WDyzXzfx1qBlD6YlOnJJvvGYwYHA9eIbdGYTmePB8eJXJQGi42vp/JEglJ5nkZnkZOm9dazO+4wm04fwNpWO82vJ6xyyBBLQfoFChwtHGULXeIFFaSoErcCpieT4ZWDpfdbKbFGZ73A6CmBUr4gac6xZKzCDe/6Edat796bi+FZ1bmJipPOQurOdsPdHf0QUE4l7Fu+hbXXvJa1jz8Bd97FnrvvZvs3HiSZmySUPsOVEXxRYSJNsV5CWUQEJFktu0HmnTT6kJN5XvNuhIuhi67PpHV5ZEJkZdNZNrs0EGrXgaQqIiR1lJEIPExrH+mBZ2k9a5iSHs2BMYaHxrh8ySouXruaF1YO8MT4DDumpplRVfabKpE8cVTQhBDz4urdGBgYYHAwAEBKibW9c/vyVlGcGEgSSJLUxc7JCkatBZFOnbZ5WRKUpkmTJhb9TTk/BQqcKDjqLvcwKNOqS3w1xPp1q96TKrkltRFO/QM65UZuxR83mtRqHv/jhu/l/AtFR5TDZNabPHXyaEy7dj5r9CF812XEA84/B7aewYofegcrtj9P9MADPHb7vex8epz9+1rsj0KEMaxQLUb1LMN6msE0AqMXLG1y2fEdq9gKSGVHPMXXdOUxAD1lZ47E8iYubn8GIVKUlChcDN4XUI4PEO+bJt37NDEhQwNjbCwPMV0tce/MAe4Ua4i8l5/Q28ntXe+JzO3frf/iYFzMX7a49PLzWLHavRtFEUFwpH8ix6oP2vGBFRAlEMdOnyBvoJTFx/eVAotO6ni+DyjXnld0fuP95Y4FChQ4djh6hG49jC3TaCqsruCJ6iWxUT9vrXUu9QWSgSUpUtV55zvfyqtfvRIArS3CZkrbL+di/1glEffJj3W6dmWd5wDlAZ4PoYI4hi1nEq47jfPf9v3w+AH+5N1/zIPPTlAKQpboOZZSZkxWGPKaLCl5hDpF2RTfpngmxbMpyhoCnVLKEhBTmS0nhLPscz15mcXkXVZDxxoVC7hCrTVok8fwU0pISDRVPAIrMCQ0Z1ImmzO0wkHGSqsJtXKqdi+H5XYI4rACV5+enabALSyVTRGygRRzvOMdbwFcgYLs87e7OckTO2TfdXWQ3eWbJ503aWGkGrTOm7rm4R0BsBORYtGAT70+Rxj6dBath0PkBdkXKPDiMP8e/ZIJvVwOMzIPMGaYpLUMVR6jGavfqQ4sDZqtliNz4ZK8lKfQcUoplDQaE7zm+jP4gR853w1Pg5ILmFZdqdD9NenzVOYWuYcuftvI3f/ZS9G/z5d4w+nxuAtk95Tn98X2o8TKEmiLKPtg4H9/5Cvcvq/MrL+acqnMCzrCNym+jSnpmKpIGJApIyqlms6ypmSppLOU0joDSYvRKKFktRPqweCTIrAuEznX5c7c8tiOlS7a6fqdKgZpwRqX7CaMxUqJT4g1glQnNG1CVBllvxpghzfMDluhaWUPmcsFfoQ9yMvK+lvM5tN5GP3Ye3MG3NjbynMKWlGLSrmGjQ2hFcTNaWywn3f/1k+y8fRM0c8D2fZ95ARlXWKHCtz16uIsR2mdszz6ZH5sFgcHuxo2+39qapY4jlGyjNbuG77vMTvTfOqhhx8AJphXtrqop6Ig8AIFjh56/4KPjoWedVOzuoxiiDStfKtfCq+fbbQQmWpX/mfcbDYJfY84nWHdacP8zM9+N6Cxxlly7WwlwRG72g/XJlhw/O3nHGPDSvQ+7fb/CidFKjxBlMIH/+ZGPnvP01hvCSKU1OMELxxof92zBpE0qYqEatpg2B/mmdYUwzZgQNQYki3WDAaUdIyyzopXuOde9tq51U1belRmkRG5gEVtBOgs/qyxpFYSUybxnEZ90yuzB58XTJnn4yrjQZmWCo7NNB4E3WGG7B1kptwmraJaqmE1JFHTeTW8Bj/5c9/HW96yKfsBdf+Sun6FVnRKL3KDteX0FPxSichCKARH/ss98ZCfgdaCqck55uaaKK9GohOUECgkUsrH539z/kKwQIECxwLz/8aOkstdgi0hKHPX/ff5l17y2vfgCYSnSNPcEnRQsoSQCUk6yY/9+M+yaqWza5JEo6Ra8FZ4dEueeinfAN2tNSWdpDHg2HtNRe8Kwg8giuEzn3qET/7LzUTNGqVyJovb5wJOhUT6ZWa1Yk56TAGVynJk0sKmEYqIYakIrLPoA5MyFIBnY0Kd4tsWQ1XtPrMCLyud8zV4iI4oDE51LpGSWAjS7DGSJWZ0SN2G1LXPtPGYlT6zyqeuIJIBWpRc3/VFcEiL/UWinVFv81ro7KeuLQJBElmqZZ9QTvGjP/Z2vueHLuj6dvcqy2QxYw+VW+VRJhzQmGX6vz5JZcko/uuux/ODBfZxcrvdhYQnn9pFmmh8z6KEwBpLauOpuemprl7op1buQIECJyNeOqFbJypibAlrS5y79aLvVkpdqoVxmbFtS89ZPZ5vaUUzvO1t1/Ha160BQKcWT6msh/VLHtGRn0LXc0OvW/+I+2Qf8ZG74rK4eOXdd03wgb/4KHFUpVQZII4SpJQEQUCq0/a2UgiMVFhZQktD0wpmLMigiggdIe81GmkSfJvim5SSTfBsSiA1vm3hN+dQxIRWujp4nRO67Sl70xJSKWlJSyIlERDLgJasEYkKkSzTUj6R8EkkGM8J4vhWIF6uGHofZK5xYDsrWWtSorhOqiPWrxvjl37+Z3nllUuIY+dm96TrgZa7TNxiJJ8IYLrhnjdn2fdPH+HrH/9Htlx1JVuvuhqGRrPjCDqSr8fC/f7ywRPw3DN78bwySZoihPtdNJqzz+96YdfuI9tbQfYFChxLvERCl2htKZdrzM2UuPOe+/1XXvqmX7QCMG41b7q2RaTopMm6tYP86I+/gSh2KlQo518/GJnLo2bw9B5EIolTVwbm+85boIHUaKw1+MrrsdiPLjF1XAMaV0J2/31Nfut//jXNqES1PExqBCpwcpvdZA5Z6VDmIvcy691mte8WjbFgM311Y3209YlsuSPNKlKk1OQiNso6Fbk8pi6Mq0TIE+isMCRCo6VBW4sVklR4YN0/i8QK4S5nWzs9G1fmqj/k/B0FZbE8fJBq7RaKQmC0yBoB1RkaMrz+DZfxIz/+OpaMOL53fwgp7lmaZQwqJII4NQQ2K0qP5mDbY+z8u79j5z13skFp1IFJ8GQ7rO7gFmvyIP3bTgT0U6zpe65b8PQT46SxxfcUKItJE2q1yt3f+m1vTj7+ibuO8AgHO2KBAgVeCl6yha6UYnqqDnopF59/zRtR3sXuk668X5slXgmDV4n4oR9+B2NLQc8vn36Z4SypoF2orUE5e1lIk1XNv+jI/GFCEluwFvbugd9/398wPS2o1pa6srY4PfjXc+sz1zq3eavOzmPeMz5Btt0NJiNq2+WOyIle9uy7q6GIMBgSTEbo3d9TxhF4W8Pduu+/nIa5tI7MgyAgSVJarVmGhqvEzTlSHRP4hosvO53v+743cNFlA3g9srnumSZxOuU5EScJgefBbB127WLfRz/Gc1/5CtW9e1nrKcYbDZYtXwZRgqnavmYrJ6dFmvuNsLBrB+zZPYfvV/CU59qnWkOzMfNVYU7sxUqBAt9seMmEboWiXB4gaVYxtvzT7Q/a8q5dpCBSXvHKM3nrt210tqnIs6tf6iheJIwGm8DknFO1rJShbCEA2VbG8o/h8CyNpIFUVWam4b2//bfs3l0nLNWYazQJArtoln3eZITux6yWPCd1AUjTcT3nVnYO5x/oSPW4M84WX7Z3IWOya6Wt7egC0ZF4NdKDTGde2K76dQW2az+ibzb74+tHI6aeJCnWGsLQZ2ZqnHJNc+215/K661/Bdde6IvMktXhKdB1dYq3I5sS4Uru5aSCFJ3diP/8lHvz0f6Of38Fgaigpj7oQNAeGGL30MgjKeFlh28lJ4/NhgYcfmWJysoWSJRppE2khCOTs9PTk7UIOcuQWdmGRFyhwrPDiCL1dUy6JWgJPBUSxuWhgoPTq1HY+Qzjr3JGOARHzve98k8tA1hnnH/NF/gIuPpv9S2LYf4Bv/NUH2XHPw2zcvJE1l51H7fT1sOk0CEJQJYT0nGUrJHiq81zyEgPsgopfpd6CD/zpZ3jowe1IOYwVimq1TKvVQojsEuXKbyKfsD7q6HNV9xvG3WTeU9bVld3uPCld+xFp94btBYC0slPTvshqp/2+dQp1nTEubLLntepGkFnIefmiK0PrLnybD5MtZlKMSWk25whLPstWDPCaV7+O6151Meee62MteJ47lucJjDGdunEDwqbOxaBTxOwcjO9m/DOfYfvNXyN6/GlWqZCqEmgNWgUcMCnDl14O174agqDripwatG4E3HLrnaRa4vkeUoOnJM365H0XXbTlGT+YpSDoAgWOF7oEnDIcPqGLBZ7bgMBbjrRLqFSr7/QD66WRywjGeojsxu/5CkzMW7/1dZx99pA7sOqUFR3NpLN5dek5skx5bZwAC2kC43vZ9/t/TOuLX2Wd9Gg+/hhPfPmLGCUYGh1laO0qll14HmzeBJu3wJKlMDjkBqwCCDxIU/A8kiTB9/2uQINtj6fjqxD0vInrN/3n7/8K//3pOwjKSzDWabTHcQupDNiMVG0ei86+l52UahN5XpaFI7f8XdXJY8j19hect+4SLyCR/aRk8mkEjAtLAFiVHVeAEJkFLrBdvvb2CKxpZxyaeW3WXORfGIGwwp2nsfieT2JSkkSTGkNQKjtJViPctfR94mQObB3pRSxZVuKKs8/kW15/FZdcvJrhAVcfLiykcequgrXgKVc1oFMQGlp1pyZzYD/ceRfjN36F3Y9sw87WGUhTxoTAT2JS6REpSZQkBBvWcPZP/iSMLgG/W1DmZEDXWLtEcvIrnmCZmhI8su1RyuUqUZQgpY+SHqlp/deuvd9A+fvobtgyb78LooihFyhwdHA0hGV67sMe2ArPPrd3cN36oW+bq8+hvGrmsvXaLUDTpMHIiOCt33ZFZumljvBF//6OASztGm9pQeoUfAU7d/H4+95P8ws3sTGFJGkhpURluuh6agczTz3LxH0PMKUMcmiUyqrVDG06naWbT6d65lkwNgZrVkOU4JcCnAM5z4rOFeHoelSdFzEQwL985DE+/ambqZSWYm2QUWXG+FZijM0s4jyprC9tSYiu0qzs3e4s/e75PYQoi0F2Cev07pPMkp0vg95L+rZvdZYLuthcKCYn+j5Cd526JCU/wGpNkqSgNUo5Ag48gS8FRjRIjUXrGJMmhCWfTWuXcslll3PRhVs44+whKlWoVt1su4VHDNbihR5gIdFOoNwK0Ans3QtPPc7eu+9mx623YHbsItw3ybKwjBSuv7vEkAowpRL7my3ksmVc9Iu/AFu3Qik8JWgpX8JZwBjBzbfcz/7JaTC+a59qLCaNGqXAfEaIaSx1Xhohz7cwChQo8OLxkmPoQpUYWzp4nRDqtGplwOlsQDuGHng+07PjfM87XsuG0zpxWuBlEHGht6TYpu4GvmsXT7/nd4nuuI8VKkRHDcIwxBhDmqZZ72cI4ohgrknZ90gbE8R79jN+94PsVoqgUiMYHUUuW4Y663TO+PEfgLVru06oq7ysnUdg0S2N8j2Q8LUvPsxff+AfiJtVAk/QimK8UtnlRlt3aaxRbasbDEK4rPRcSc3gz5vDHs6dR+K9r3u2Fd2fL2ShS+wi7dHabvW+PGkjTM/CIM9yt9K2k+mE7WTn6zgBJEp6WCFoxnUkKUJqsDEjS2osWVpj6znnsvXsjZx+xgqWLy8xPOBORonOsJNEY32FIACpQTddb/pmA7bvgCeeJrn3EZ66/R7m9u1DxRFLAh/VSgiCKhU/JI2S9vwnnmR/owlr1nLRu34drrnSeWnmzeqiM7wAjlNJ24KLO5f0mOCTppLbb9lGq5kQehqpAtK0iTbNr5XK+nFNhLVOVOfIUBB4gQLHCi+9bC2RlEsDbwn8Gkmi6QmKC8NcfZaVKwf41m99JZ6X377kPIW0YwGTHUoanDs1asHEPh5/3x8R3XEf61DoVguZlatFUUQYhu14bikIHUlo4dyzVtJMIlfaNTPLbH2Op7Y/y0Vbt0ClCqnJNUPdgUmhe/ECqMBz/nIFm844jRt+/ed46sl9PPPcDnbsGmfiwAEEef9zz23YFYN2sq0mSx6TCFQ7ztye9gXmol1TftAZkz3qavnWjsMzi71POO2QEGmvpd92xVvXvSs7F5sFEZpxgu97hKGP7yvWrB5jzeoxzjxjA6vXLOess5dQHYCw5PIwpNL4vnbj04lbWTRjSDW+50FTOzf/ru1Ejz3KnocfZOqpJ5h57ln8/bOM1FNqkWUUidWGshEoVSLWMTMzM1TCKlpAyyszFQbYdeu55Fd/CS65GJQB7+SPlfcjNYLtz8JDD22nNrCUtCUwJkX5lqg1/Y9+SaPj2CWUFgRdoMAJg8Mn9PymniOzwF/YM1Fbs2LNq4yRRHGK54XtmJwwCb4fccmlF7ByTRb/fRkNEgs0kgYDSrkY6b4D7Hjv/6F+620sl4KoVceTApPdk8IwdNYi1pVcKZ8879ptY/Clyqx9Qxx6vOad34f68R92MXWTrVgMIFSnhov8vEVnAWNhzZoqa9acBuY0EnMZqYEX9sLuvft4/rmd7NkzyVPbtjM312Jqcoq5uQbNRoSxIssc9xBWI4Tv4tdSOhLLYKxtd1ZVWT02UmG0JUlSUp1SLpV65syIXgu902rUtMcvhHAJbIAxaXubzmMnru8FljRpEscxqdb4UuF5Hp7v4ytLqWwZHiyxfMVyliwdYu2aZSxbPsqGDatZvmyIpcNdVnc+fRK3bvQBtMuHiBNIDMzUYa4OO3fA888TPfkU+557nr1PPwXNFiqOCNAslZqSsdS0QknQ0pBKmEubKKvwhI8flEhNQiMI2V8bZfgVl7Hpt34dBitQC0D4wHGvvTyKkFgUoRR88Qt3MbnfUK4GQEKcNPD9+KkX9j7xX16pQm0QGvXoeA+4QIECXTgCC73L0myLX0nCYOBcQbjRGOn0v7tdeSKlWhW8+a3XohSohbJ/jyHBCww1X0F9DiYmefoP3k/z9jtZowU+lthY5nV7gcwFnI+1SxrH5jTn3vMsJLNzqC/dBGs3wrKVMLrUMZBv3QZS0vYpCx+UcvsV0u3MSlDg+y6vauNGWL9xjMsuH8OkUPKhPgOTB2B21jK+d4LJqWn27N3H5OQcE+NzNOox09PT1Ot1kqRBlBjiJMbqFIHzLKRZBCBJMqtbuJKtJFq8btoIJ3vascRNps3vPBYCSRQ38X2PIAhQysOmKaVSiWqtRqnsUS5ZarUVDA0NUatVWbFsBcNDQyxbtozhYZ8lozBUhVKYHSLPtlM4pZ0kzlZTbmwi1c4Sj1rus6kp2L2b2WeeofHCbqaffY7Z57YTTUxQM4aglRBow0prkVqjhHVJdb7EGoNvDBiXhKeEBCnxgxIay6wQNIIS8dgY53779xB+9/fAcBVKyiXSdXtgThEkqcfecfji52/DVwNO8tWXGKEwaetD69aOzgXBFCZt4Acep0pGf4ECpwJessu9Vh26UuALrS1SusQwB1emtmnLGFdcUs1y07rJ/qUd+bBGZwxEddg/yc7f+z/Mfe12lmqDpy0mSSn7JVKb9tRm93yfJBtr7nqWSGy7BrwWGZ77t08z/fFPUVmylKBaozxQZWBsCSPrVsOKMdiyGQYHYGgIqhWXJe8pR+yeBzJw1jwp1mi08BAiwJcgMmnw6hBUBwAEZ20dw9oxjN6MNq6yjgRaLWeoTk6lJNoQxwk6NeyfnCVqJczV54iiCJ1qkkQ7izlNabsncnQJ1RhAlAO06CxiSqWAMAwpVyqEYcjAQAXf9wn9AE8pgsCjVAoYqEKlBOVytmbKdquzHIu2BoEHNHHp/r6FuOUI2ySOsOsNmJyC8Qk4MAPbdxDvP8CB3ePUJyZoTBzARi3SuTppq85QOWCplITSIrXAaGfKKyWQAVirMdpgU1czqY3MYvluDWE9xWyjQcuD6ZEBlrz+NWz8/u+Hree5RDrfxyXWZT9gxanFZxY++W9fYf9EnXJ5OcIzmDTBJK3x2fr+D69crUh1hLYtlCrc7QUKnEh4cYRuPbAlYJBSOHS1EAqrNUKKzO2aqZnImNe85lJaBpR00qILGMQvCd02dO4cVmTlXnELxid4/Pf/kObNdzBab1Et+egkJvRCrLEHWVjkTbO74sjCILuC1aExrFI+Y9bCgSmifftQUjKLYS8ps8KSDAxiByr4g4PIaoXy8jGqY0sYXbmKytgY3sgyGByGpQOIgQoeArwK+IPOja+N+2dTJ4RT8hHCtZlVCogVWEvJE6AstZW5GpznZicY6CQT5xn/3ZN1JOiuSuiqHmh/1r1fgRtvo+nIOftMGe0E65PE5RwcmIb6HPrAPuoHJklnp2lOH2Bq3zjxgSmi8X14zQjbbCKaMaoZU9aCEEnFCEZQKOs8BwQVbKpx3fsSl9JAiJVghMVkrT+t7SRmGiRaut70sQcTcYtksMZZ11zOlm+7Hq68GgYGQHjoIJeVFW4nNpfWO06JbUcDmcct/3lMjMNXbryHWmWU2bkmSmksLYSN/mnJ0uoLs/WnCUPXEln3LwYLFChwXHFIQr/0iku7xGFc20zBMPXWUuamyjVk6WxjNUJkutc2q+s1hoGqzzVXne9KdCFnxaN+EjITLzGAthplU4hmYc84L7z3D9C33cVwM6Va8jEmRfoSbbWL8wvbpZXWGV5/P/Bcec323bxTnWR19IaSUlhrkEIQCJ8agqSRkDanMPum0EKSfuNxpgRMSeXEXvwS+ArpeUhfMDy2jCAsU6kNEpRqlJcsh1IZBmvO5F2xBAIFgQ++B7UqeL4TxVeeM327E9vCqiP49rhVh4za57eQ2924bbR29eMiE+NJ04yMI1f+FSUQx65FXBK5dqJRBPWm0z2f2I1pzNKca6BbEbPTMyStiFajiY5iork6JAkmdQsWYSxKWDwrKGMZRLYbxThZWuH0faxBSInI+rNbCwaBtZm/XgqkBIvAiqzFShbvVzLTdzeaxGgiKZitlJmuVVh7zWvZ8ObXwysvcdMQVoEAsKg84dPiVqhw8pJ5lscB5Lo6eBI+9tGbeWFnnbA8gLBQ9gKsjsf3TO7604GhFOmnpKlBSYtpE3pB7AUKnAg4bAu9QwcCbRTCDuL7A6ux3hqTkYPpKlK2NuGsMzeyYilIazOhkWNw88sEY/Kwq4+BNIbd4zz1v99HcutdDEzNEoahkzPNs70PtwlI2wWdf/Xgpq1CoqwbSwDOlQxtXXXT7sut3aJB1gG3UNLWkO7aS2wtDWuweFhVQgtJbA2xsFDxSZXEepLUEwjfA+VlTUgk1WqlPZZUSMKBYVJEO/YthEAY20l2s7KvYUp3UhxE9dkst8+9brVaGGPQJkWkGpWm2DjF6ATiFGkswlgwFk8niGiWAIsnFD7CNcGxgiGl8KRycfKMGITJa+0FQuTj6k06a4sRZeqDkk6GvxWyN6yTnYfEZCJHHokwNKWkbg2zwqKrAcMb1nHGla8kvP71sHE9VAIohWjpo5wsTe8uBVmm3kmOjNRT46I/jz7W5Es33oYfjIB1lQY2beD79q8GB7ydQroSQkuCMadSMmCBAqcGjsjlnt/4TVbWVS6XN0lhS8YapLQYa5HCWVBx2uSyKy7IrOB+Be+jA0PHcpYWp/oVzcLecaZ+/y+IvnArg9ZQKpeQwmVsH/Zt+BAiLC8WSvZq3QrhwhTWOLJXQuApj5KSjoS0M5+EciVlSStpJ6rl55/vxxiNsR2y9gFtBSojOZllp9NN6MBiHdCsAD/fLlsADWTjFZngSiXwoYvEFQKVHUdYgQmqGJybXQCe9N21MjZbKIgsh8DlD1pr53lBOiSeWeldCwyyXdi2xZktCjLCt9J10lNGkkpJUyomlcSuHqO8+TQ2XnMN1auvhpUrXIjA8yB01Q0ueCROqRB5PywglVtT/dOHPsv+A00Cv4ZFIqXEpPqZJ57c9hfL1vgYEyG92JUiEh9q1wUKFHiZcUSEbq1FqJwcFKoUbtR5TzLraoGl9ECkWCLOPHPjMe+25VKbcGQeN2DvOE/87/cRff5mlkeasBo61/Z8ibOXFf2kmRO5MS7+G/oBoRREUYSwYFKDwBDoLFJhHUlWpCA1pmMhyY53xBhDOQjJ5VgB4iRuL06EtCgk1poeb4o8yEUyqkNnuXcBK1HKEaZouBt73rJUWIsxtj0+owRC2MzLbxE2xVhH/FI613hbbAbbXnCYroUKprOAaI8Z52UxwmQUnsmXWtHu366lJFIhiZRoEdD0AsbO2cpFl10MV70SNm+EgYqzto0GLyA3W+O2Xt6p3VEs1RqlFF/7yn5uuul+ysEo9SjFDyyWhHpj/L3Ll5cPCOaQXuKuJXBqZQIWKHBq4MiS4tqxdOcOtUast9Y65ZbsZiuFR5zGrF23nLO2HmT3R5HoJQaSGdi7lyd+83dI7nuYlaFHrRTQiFNnxeeW5SEO3BZgWWSzQ/XzbueI9W3Wv5zoX2BorREpeFmGkmon5XUsVGsB3evxNbmWuisxJ9Hd8XGwImsmA1gyjXche+a/R+C1b6Ciy1Oh2jdx6TzhQrT16XOlONsejNvWM9Dd6N7mcrVStnPqNM6z053cmI/Dpim+VFkioKQZR84bIS1WWXTmOfBSgTEKLSQt5TEnNZMIahs2MrB5E1uuvhJx3jmwfoOzwkXgOElp9/uVve79ANc+V8zLMTi5kXtXrLUksSUIFRN74f9+4P9h0wFiDUEQoK2mHk3e6JVm/1GKSaxqgIyc7oH1cAGl+aWdBQoUOH444ix3R+BOjEXB6oW3MaxYNep45BjeC11usnYx8z17eeJ//QGte+5ndKaBDCStthiOOu4Weo7+ceQLhNyVnCdOt5uNifm3S5mRuOnapnOAPmnXLHZ/uOjfUvW8K3vd3bZPHFZ0Fi7dndhk3tq1a6ydMVuXB0GvMI21LjHRV+4nmhpNYjWEPjEWbROawlBXAi0kCIEWHqUlyxhau4azL7oIzjvfNdUZqLr68dADPyBFdP3wPUxbpre7mas8hWi8A9EVMglCd4Z/93f/zTPP7MP3lyGURPk+UaPebDQn3jVYmQE1AzLGEmeaCpLuhi4FChQ4MXBkhN5uf9m+7Q/3fGYhb5W6afPani5qx8RBl9eZ79nL5P/6E6LP38xSYfClRacWrSRSnJgu05yM+0enMmIXRrmYeafAoP0d3VUrPg/97+WZyN3vHzQ/YKF9667HzoiddKtzhXeq1iw6K4PqqOZ14v2p6FJ7y4+4QPWgWxAYkkz8x/qKxBNMxU20LzG+pFEqkyxdysDGzaw791yGN26C08+A0VEoDbhzMdap87RDB6LvR58l1mXn7mzOU8id3F1GmD/NFpFJBDfftJvPfeZOapWlJKlA+ApjEpK09QdvuP419992y0dApK6hEqnboaDr91FY5wUKnCg4TELvWGjdsIIBk0UZ8xu4FAZrNWtWj7p76LFcxVsDc7N86c8/QHLr7az2QpqeZS6JkdJjoFQmiaK2pSgsWaZ1nmnev0PZ955pd4zreqf9XNhMXeyIEuhyclx4odGex64x5Zn87cfusrS+rm79x+oQZ279ikXJfeFL1ZXNnFnZ+Tfydqz5QkO6I2bNTDKPgwuBu8ZmMisv6zqHdmhDSrS1aClJs7PSUjLdqCPLZcKBAUSlzNCaFYytW0fl7C1w2gbYuBlCl5WO5yxwLSRClTpzosEYi/AW/znmOfPQF4JYZPuTCl3XO45d86E4hj0vwF/95cfQeoAk1YTlEK1TEtO89ZEHv/4HV1+5EkkMbTI3tOWM7amnklegwMmOI3S5G4T0soQthee5RiZGqI4LVUqU0Gw+bQ3hsTaOhYexPhe+9e2kZ57HgUe20dz7AmZqimjfASamp6hg8a1BWfBSi4cgBELpZTXNLuNba+vK63CuXpd0JRAywWAcIelMwCTLwnZhV4PMSM3F6rvKqNxWbqh51nV2c1W58EuGtmtdZZHJbNFhhXXx4/yU5zFSf7C+28L2sDbt+Vj1b9/X7lTgJFHBucDzbyuVnYsAnSWpAcTCIIxtJ+XlneqwznWeabmQ5kL+SmExjjisIRIJiZS0rKIpFbZSo7R0jJFVq6ksHWXV6ZuorFiJv24jLF8OA0Ou/t5Xbvps7BZ2SrkkAqE6YQIrs5pxkO0yM01bTaV7vjh5S8oXRNsyz1WF3FJMKB+L0yp6/598jJ27YrSpUqn4CGswUWPWE42fecWlG+MH77uRu+66EWh0FgXtOTqVJqtAgVMDRx5DN5ZSqUSz1a/7DViJJiUIPcrhy7B6lxK5dAVLrhiDy17B8kbTWQ6TUzC+G6am4YnHMfv3s3/XHmYmpziwZwIbJ+hmExsneFK4pllK4BnaIiae60KCkeD7Pp6QaJ2pnJHlQltQWTKX7bc4821MCqisn3knYc1a1ybUCdjkSWs5TMcutoDtWMluy85KaX7HM50tHmy7pK/7m87A6ny/PwnQZvOKICujywSFrNN1BxCeh1ROrMZIn9QYUp1iTVZOZy0GS4IilQqjBNaTGCVo6ZSgWmNweAivWmXJsmUMLl9GuG4drFgBa9dDuQK1QSekE/qZRK6P62TnxtYJ4fqdabVdCx7rEtelB6TOSyA9sg06Nfg96MpdOCXQQ+Zg8FDK9bH5+w/dxV13Psng6AbmZiLSNCYIUoSc+1UpZx6U8gCIGVx5WrfP4lSaoAIFTi0cAaFnzlZhiKKIB++9VVxyxXe3C8ylcZap0YZSNaBUKh/ztg25KB2eAq1gtOTu7CMDsH6Vu6O/+mqktoylmrHEwIEpmJohmdhPa3KKmd17SCYnqe/fR31qkub4XkQrQjQjiCN0lCKjlJJS1PwQzxpH/Ma43iFZxng7AcyItuPWWg3WYojASpccZjwXqRUxRhiE8DHzLkNfXLI/Lt6VNb5w2CArGVv03pv5DqzpeFBl5hIXrlTNZgsNY4wrdfMUKEEzjUlNSmotkRXESpKowPUz9SSqWsarlakMDiCrNSojSymNjDK8fIzS6CD+qmVZktoA1IYgrLkyO5vFGZTn6tLzmnIJ7RasJgE/dNOTQGph6oATr8v/SQUDNahW3S68rCJNZbXWIkugEz2Keqcg2ueW/xV6YCFO4Us3Pse/fPSzWFGlXp/B4PIFG80DHwn8/X8j5BRCzoBsAeliRyhQoMAJhiN2uRsjSHXKOedfbj0pE6O76pkzwZAwDAkCH2Pa1UvHBCJvjCEhsc4L6+5kJfBTjE1dypPNa609WLYMDPip67Q1YKSrYU+aro69UYfpGdg/BbMxTNShmbLz8cfYdv9D2EaEb0AZ58bPXc/SugisZ0Vb992zKUkyhRRx24UutXBJRln2ubJZi9Y8Dt3jU1844WhxnnaT0W11a9FN7LLv0Y27XbctnBWdSDBSIaSkVnZNWMJqGVUqMzw6jCqHlGoVVLmGN7YGKgOOoKsVt5iqlGGwCpWaO7jyXJ2dl8UOFI5lkRhZcp6Wxc7JgOsd4Kzs5hxs27aPO25/kG3bnmPfnini2BAlMUmS4nlO1nXjaRs579wtXPaKrZx9ThlfOXVaPzxV89cXgiNyk+VhWAvbHtG8//3/QJwGSBWitSUIfOJo9v4X9jzz8xvWS5fVLlog+q3zAgUKnMg4DELPb7WdP2xPeWgk1toZ2acCl8uLlkqi51vHhNe7kqv8AMi0vRwsUigsrqzJeVM7CVv4EkyWju0FUMLVJjMIdjWkCtIAZuFLn3qYzz37HHuba5DG4iNQUiAw+NJDCYEie50RuhIC3za48hWbOPfMFUhVZ2ryBULpkUR1GvUJytLHNtN297YcSZoQxy3SJEXnAjKZlW40+J5PGIaoYP7lC8Ny25WcKCgNVNHt3UsqlQphWKZcrhCUQqTno8IQqmXXGq1aceZaueqyw8NylmzmdyxolecR5Ga97FyLfAXXvjay7SZvvw/tNqn9VG6NC+sn2uLncW/t1lyf+sxT/Nt/3sj4+CzjB6YpB1V85SGMxF3Azu/soYde4KGHdvKRD3+W0zYt5x3ffz2vu349AogTS+CL3mOcotBWtmPmk5Pw2//z/9KYG8BgkFJSKlVo1ucmotbcD6xatWwGsQeXDJp75I7zCRQoUOCw8aK6reVWaZok474MkO3cb4lCkKQpabJQAtexQrv6ueuluxnl0iDkUp6dnLaMgLLn0iWppUaSWkli4HNffoyP/MOX2bu7DliqtY0dmVMA45TDc61zmRE5gLAGnxZPPz7Fa1as4wd/6DsYGwvBT8E2QaauqYnxe2O5nu9YO41dYxSVx7szYo9T957vd33WhTRPXspc2H72mPeqt1mgWeRNWrJ5EM6l3vbBSxeHd3XHYJXESIFBYXGOWA8I8sSzHD1Z+Ae5VG305pTnOXp+pnUgJdx62yQf/sf/5MEHn8XKGkaUKYUDjqi0m/f8N6npnAKAUDWeenqa3/zND/Dpz2zmp3/muzjz7BFSOOXJPK+YMAZmZ+Ddv/ZBnt8+ydDQKlKbYkxCozGjW62pHw5LrYfDcoI1LRfiwCxcFlmgQIETFkcs/dqOPwKp1rv8BUzvNNG0Eo0V83Kqjxo6ln83Q3RZhzZ3hbM4wbTJx2RLEg9toDkLf/Fnn+WLX7iHVhJSrowQxzFWBC5r3PZqodPl1u4RViFgUsf8388+xI2PvsAN/+NHOO/iAVADYOZgMHSeAJtdBieHlu3AdmUWd51AHDsTVnSvTDJ0LwzaFnNfkbcQWe1Ytm/pdd4Dt09j3WICg5UKPEuugp8Xyvn5bGfJaF2F6D1j0Nns5hWM3UN2c57HaCUWhUxVO29NWvjwR57l7z/0H8zONKhWlhIGVYyWxLEh0RrhWaw0C/0K3HCsRZsQ31vDgw9O88u/9AF+4mfezre/fSsJp0A78/4FUp6c2fVyahr+4H99igcf2EW1MooxKY1GHSE1UTT1S6Vy/F9hZRbkDEJHCKGz5JSC0AsUOJlwaEJvl1uBzbTZQDrpUWOeFhiEVUiRW+k+SZKgU9t7o3/Z0HXAnGwW+Li7stxt4mGAuVn4zd/4CPfe+TRSDeKHZerNFoHnk+SNUmzHDLVCtveVl67lzwG0KNEQgsf3pPzku/6Y//OHP8crXzEKpowUGkpB7/ja5UaCBWvVfeda7sqL6yS25Y6Drp0pr0NXBrJse4HKrHsdu65vsttalVncGxDY9gx1e8uhiwgPeo27R9NJ0eqmio72iUBkWekY+LM/u4kPffTrCDlCGI6CUERR4ibXKDwhMdblI3Trp3QnA6ZpSqU2jE5Bm5jZmUn+8H0fYeLAm/mxH7umPfSTktTneTscTNfHjSb82Z/+B7fc8hDlcAxrLVEUEfol6o2J9wWl9C/CcgRqDkQzq+rzOnuxfb/PvMQAsgXnIoMoUKDAMUZvFQuAt1g6Ur7JxN4J90TELsFMVEgTRZoGDNXEM6G/yorMZBd5drIRzM01MHbQ3TCFbpdh9Wupv1hVrkPJofR8vMAm3ZIseV7f3Bz8z//5CR64fx/4IyRWIIxBeZ7TGzeWXtEXetqwti1WnOIbgBCWWqWSJapZ3v3uP+a3f+dnuPralSQGfOIsScvvHXMP/c0/AbHAtPWHqfu/Jxd4TwVdi4bue3N7MdbRa18QC30kup/2tjfpTcvLk7ZMtvZy10IDH/jzW/j4x79CORx1evXCkOr8ikmQ2onQSQGYRUfo+Yo4abXHJVUFhMeH//6r6KTGj/7oRQSBa+2uhOsJflKgf7GaLyBT2gI6zRj+z+9/mi99/kECfxAh3CLN6giB+PvAF792312fB6bolKf1K8ANzj9uu4a/sOALFDh+iIEGPYR+8C/kCU+50xSaUYtSEJOk+5By6Bkplk6CN+q2dvFMnVr27J5g67mDxy+p5jCO2753G+eB/qsP/Dd33P44So2QmBjpKazt0GCuT754Odj8MrLOphJsgLE1/uhPPszK9b/GhvVZZhjJAnvqvrkufkBzNG3LF3OtDvGdwxmdzLIwsiIJ/uM/nuRfPvFFhDfkSvr6Yrm9vewPZ9DdCZ0+qZZ4ssI/f/i/Wb16kLe+dTPBAs6QkxGpcW1T6i34kz/+V77wuYeollcxW29RDjTGGEq+/OTX7/jiT1x44RbcDSFvhZqVtx0U/Ve0IPUCBY4fWj2vDnG/zRNj8liuISxZUibwynt49PHP7UU2nmjvykq0SYnjlKef2nFSuDFTA4GCW776Ajd+4U53DjqmXAlIdfzSE4Py71sJ1iOJPPa+0OQv/+xf0K3sfbrcmqd2ntYC6HUbPf8MfPQfP0PoDyGFf9BvvhgkaYoxxtXXyzIf+ftPc2Ai9/KfZOQkuv5lr8MAGhH84R/8F//xyYcI/SXU51rk+nnWtP7zjju/8H2XXLxeJ+lOir7mBQqcOjh8zs2I3ZBgbAtPxazbuNqAuUtmydCSLAkp1ex+YXLBEPYJBwtTU/CRf/gk9VmJUiVKpZA0SfCOtv/VeoTBEMJWuf2Wbdx80/5OKvKCE9VRkTu1YQFXQvj3f/vvvLBzFp2Xwx3l8w/DECEEnnLJiDuen+Qj//A14oh2PURnTCf8r9chz9swzjL//f/1Cb74+QcI5DKkKOMHAUHgMVef/PL4vmffedGlmyMZTFEdMhSEXqDAqYND3C17b6q5y9mjhE2rmKRCvR59TQiFkl6mhgalUoVtjz5HfdbtpT8DufvfcYV1InNf+uLDPL7tBWqVpSjpkWqXea1w9eVKCGT2T/T9a2+T/ev/vJeUJGlsEVRQYph/+dhniZqdsXRg6JheJwmpvCQorJE89USDB+9/Cl8NYM2x8YEnSYKUEm0MQihq5WV84XO3Mr7HLDDTJ3DSV/bzsPn4BMzMwu//7ue48cbHgSpeWKLebGEFxMncZ/ZP7Xj7yFI5p5lA2wOkZpKFleD6/0qP+19qgQIFDgOH6XKnQ0zWQwgPYy3Cgkn0HXEc75d4JEmKVOD7JSYm6jz5hO7JuD3hYCGN4L//66sE/gBxfKzFag3WWpQI8L0Kzzyzm3vv3e/uqcIlNH3zQWI0mBQ+/ZmvMT0dY6VyTWy6wz3HAAYJKqTZsvz3527ulsw/KeCWG4JUw8R+eO97P84Xv3Q/gmEMHqnRSOkD5pN33fuV7xhdqmZQU/z/7P13oB1Xee8Pf9ZaU3Y5Tb0Xy5ZsWXLv2GCMDZhm01tCKAklyU0ghXB/od773iQkoXcSShICAYONCzbgblxlyZarbEm21buOpNP23jOzyvvHmtl7n6piGxLQY4ajc/aembXWrHn6832QgyBrjBbmR4X3UTpK/5NJWob/x8ijAJlAgo1oJs2IFGSDU05Zvl06eXfSyIjDGOMETkj6BzLu+NV9PL8s+dnTE49nbHh6D9ZYjG53Px4ac7NivEO2HWClxUoLgUWGvshaa8cdt9+NyWFNfae14ikUJYK//SQFNIZg1YonECKkUCSteL41HEuSDKGUZMU9j1Cr06Z5FqhDv+GkhnE8/8U7lWrYuh3+1599lTvvehJECYMj0Q2CAETAd++977a3nXX2stQL834v0IGjwvsoHaXfLjoEYBlLAeHZXtEs8QzX6DppVr8qLoWX6UzjpCNLHZXyJB588Akc52NRgPMdzP4bkQVWrV5LPZXIqEwsY7Sjmdl/eH3OD+VugHAIqTDGkemMNU88w9AglDs8WNtodL2RVdu/hSRh3fqM3r0NcLH/w68JwyCKfbBk8+Y9bN3sOGFpeyT91yHMDx8cuQgEWAcPrBrkHz/9L2zatJ9SeTJhEJPojGpHhb4DvZ8zNvur085YSi3ZRhCluYLe8rYxoo3v4dNRpeAoHaX/LjRKoI8UHRs2bJrg9F72zN+MzsTPq9XKLuuiGQ7hC2EJ2bhpH0881uCUk0rUM0sUesHu6Tefzi0kbNi8Ey1KSHzTGSHybtq+H2zeLMW2wYkeKgMbzqib5WxG4KxDOEccldm16wA7dzmWTGpBlx5O+5DfFna6eeMOhgY0YdSJNW2qY74Qo7vKPTcUqIgsTUkSw+ZNOznhxFn5J7+OfdqelBYxXJEpFEDZSqmgJcydg59d+wxf/8q1DA2FVEpTCcKIer2BEJDUan8Txe6fH7j/BmAwv4+midoDtMrUxtpFIznBRDvtt2UXHqWj9D+JRnuRn+WbqBmob6Fnityd6aGrM500YWGtkFgXcd11d5BZiENF+t+gFWM7ZGuWwfade3Ei8j6H55F3t1+7NQZJllh27t6HAxpmrHr0324yDrIUNm7cAjLAIXP0vSDvsvZ8CguJMR7tzzjFtm07ilE9j/ccexxNYT6SimiX8Q1Wssyv12c/fyP/9I/fYXAwwOoIIRQCi0AP4RpvF6L2z0IOAv34WvPi3QvwykMRPgta9z8idnBUmB+lo/TfheSwWtbDFmgNwlIfRmynnvR+KwyFHlm7vuLex9m1E9IMwubL/5u3zj0EJuzYsQMpfXZ+s+nK4VxHHOyw+TE6zu6E76w21D+AdX5MIx/DyKyG3zYSwgur3t69w/HxmRjA57mizOjmfXft3PX833AYFRgE0fhli9InDErlIXd3bYcP/+UP+fEP7kCKSThnECrDkZDUD2wUrv4yKQf/S6q9SLWXo2VpR+ko/e7Qs1SvLTJIgTpdHeVV2pjrPXMskulgz64hrrnqfkrhiDpf13b8Gqlg3s45tC6gQ39N9x4h7ItxSKXINJSCkbjZv/0k8cJKSoUxBut+vRvCOYcQAucc5Ur513pvT+NY5m1/dwAG7r5rH3/xoa9w911rkXIy1gTEYUAUZljde8eBvg0vDuTQPUr0o+Q+lOhn7LK0QxzXUTpKR+l/FB32WzvcOReAK4EtARLn7OeswUPM5ALd2Zgbf3EPW7cVFuZvtra6CAlIKYki6OzoJAgUUklU0Kp9lnm3L+EY1o5z1PXckR/FeAIhc4UHxn4kY9UF//aUGIUBvptdLsybz+jXsE2kkh5PQEqk/A2s5xjNg+wIGZym8J3vPM7f/u2/sHHDINWOGShRAiNQwjHYv/XrTz5506UzZtY3SbUPKQeRYhApxipNOxgNx034bdpnR+ko/bbTs3pTJTnTFRo8E/mVQF/VtNKBcqmLHdv7+dH3b2tjVL8557FoSyMPAgiCgCRJCFSAzn69MWzhwBlLqVTy3Up/h2n27FlIBMJ6fINfhzAHkEJgrSXLMirlSvHXX8/NgVEhL5dXOhgfiti+FT7ykf/gX//1SqypEsWdvl5EKqzUg/X6jvdG8dCfLD1xRsOYXhD7EbIfIWq+tPS3MlBzlI7SURqLnrXLHZGCOgDhDgh340Tt/2JdvdDsnYOujqlcf+29PPRAAYtmfKem35ChXliCQQDz5s0jTb0gn8gSf76oq6uLmTOn4ByY/2Zlfb8Osvjyq2OPPTavKPj1ksk7uMVhxPz5832r+Gaq+fMwHjfioIBU1q1WPBJMCnfdvo8PvO+z3P2r9b5LXCAIYolThszWH93ft/klYWnvt6TYRaUUEAUKpYaQsh9k+hvZz0fpKB2l3xwdQh36+GSxIBogJSiHVAZs6RHnsi9KF/5vpMDaIeo6IQwst95xO2ee8wq8S57DFuiHX7E7NrVb6ccvWYC0v8LZMoGK2uyZ9gS+Z2PljD5XCEXhxZg+o4uZs30c2Rjb7FP+u0Iut0hnzpqGkIZmU1tVwJo+WwuzeI7tbUHbyuK0IQ4h7oxYdOz05jnPN2YgznhEneYYfcy8kUJtAP7tO7fy05/eRqMeU63OwEqNExnOaQZqg/+6ffumD3/og7/X94vrvkq5nFKvJcgwQMisGe7yyamaiRPjxpvlyHUf7zkcdccfpaP0m6H2ElRPB8V6efzxxyf41NI9qQxCoU2EszHWlTh2wfL/t+z4C1++6JjjTlt+2jGceuZyzjj7FLp7PGtRBKjRzbtbNEZfbtvWv6TJbJ8l+IgAzj3jWLo6DDpNUWGllTSX9yd3ObDOkVk7reTAgvGVwog01QSyRJLs5ZTTj6XU4QcT/o4Jc/AyR0o46RTBnHmT2LatjrOSMChjrEYVD/iInnOBbgjIvA5bFJ6hAGklGEOSDbF4ySRmzh1+HzvGv0YjBIz4vVm+Pk4dtyviCZYs0YRx2ct2CYmGO+/cx79++yds2rgTGXQRdAQ0dANnU4A9xtoPKZn8YO6CDn56zb+zY9OTKFFrlveJ/NouB4OaO3cahYJylI7SUfqfTcPfZMvWrf20K9XPykIvSAlBEIdkxgBD3HXPdT026d44d86802YvABX5WxZ2QoGTMW6C+SHyzGdLEjjmGDjjtGNZ+eCOEZ88VzZafq2c4Q4ODqKCgCBwhLHkvPNPQUnfxjWvUvqdIiG9iR6FcNHF5/Cdb11PR0cPBu0BeJ6LBSnKKNuFrPOiuaQUmTWcccaJdFS8vH1eqyqlyLeWJIx9Vr21sK8XvvyV67n51gfQpkSpPIOkkWFcShxBZmo31IYO/HkcVZ8WIsXKAwjRj5XaDzefmw9btC1a01p/Hud0lI7SUfr1kJvYc/Ys2aUkVFWkKNOop1x39dVTnbN/f/qZyx879wXzXrdwoe/PrPD8JIAc/MLXzToSvJhvR7FqP2zToipQtYdFNp8tkxIQxfCyl7+IKLRI1aYqCO3zA55VP/QcXtOVwHkgj1KpgpLQSPpZdtKxnHPuIn874Rn77xqJwvWh4JKXn8P0md2kWQ1nHVI9W2luvWUu0pZgA4a54UWDqdMrvOayi6nVvUB3Nj+ORHNsbtBWhrjN9Wrj/OE1WUGaeXyGX9y4mT9539f4xc8eQqdlQhnTGMqQhChU/+Dgnj9P042vqnb2PS3kAYQcJBQHkLIPRDosJP/rqN0/SkfpKP33pGdtoVsjSdOUetJ4yxvf+MZ/6Jo06RjpHE4LRO7tFA6sswinUcJ6Jm5tLsHae4GPYOAiH6KUICRStNDcLEWG/bMbf5rBRRcv4oqfTGP9+v1IynmL8qLL3BjjOmQq8LLz35zFWouzGlw/b3zDOykq5dTzlIP135tacDmZkcxdIDj/hafwkyvuoFKdTqCqZPrZpv/rMZQy21TYdNbHSy99KfMXQPocVxqM6jQoIEmgEkMjlWzd3OC7372aO+94grTeQRxNJc0GQQqUBWcaP+/r3/bX5ergmlIlRckEYfLMddmHIP11RPtHUVGBcKTKw9Hzj55/9PzDP/9QKn8OW6AHKkIIQaPRIIxCvvilL3Z88M8/+Lk58+e+Vwhv1SgJThpMkqFcCLHyNb4a/wVjPbqmsaAzbw5Z1wKpFsJnRYUBBMInECkBYYhTshlXTTWEimfllpUBxAG86w/fyMc++gWcVQgRk6QNoqh0+NcbFv9vE+iygUATq5C6OcAlF5/Giy+a+pwl+v1PJZ/4ZpDKEUvFW976Uu67bxUH+lLSBIR6NjqnHSXMoyhiYGCAKArIdB8L53Xy9t+/AG1BqedOpyow1yWQWr9HQ0AFcKAP/usHK/nJFb9kcMigs5hquUqapkRRCWf0LiWzTx6/ZN43n1i3GanqYBOggQyGKDwP0uHfhfbqgHFe+ueyDLBwqhzpnj16/tHzj55/ZOcfLJfrsLllGIU0Gg2ss2zY8PRxH/3YR75fKVfONi4lFBEI71ZUwqJiAUkKQwC5Vb55E2zZxs4n1rFv62ZEYxCShCRJyDJNFEWUSiWq1QqqUqFj1iw6Fyzwwe7Jk1E9PV7Yy4BI+Gw5J8SwzPVDXx0IFAwMwTnnTeb1b7yY//rBLYThFIyBNNFHJNRH3Sb3h0osadLH/Pk9/Pmfv278HILfSbIgFAsWwQf/4g/4xMe+SVyaRpo9GwtUjvD+WKyxdHRGJOkAQTnlLz/yfnqmeIH7fFR51TIIQx92SoFbb93I9//jep5e10upNBWt64SlMplxWKFJ04H/3L9v10fPO/dFm4WooeQ+kAcADc4iycMH43iOihfePU97q52hCHf49zl6/tHzj55/5OcfjA5boNdqgzjn2Lx540kXXPCC6zqq0QJHQohP8HHWW+jgoN4P/TV4ZivunlVsvP8h9q1/CpGkmKwOukEoBIGzhMLnlRtraThLIiROwK5AopWiVO0mnDKNeaefQuXUU+GsU6G7Gzq6cdL79o9IqAPVKmQG3v+BC9m5a5CbbnyASnkKEODc4flh290ostlNSyJtADKlVE350F+/j8nTvTLxOxg2byPZClQLAIeUghdfPIffX/c6vvud6wnCrmfZoKWtwYvQZDrBmSGcGOS97/89zj63k0YDysrff7SNPtKHcuhi3wIqF+b3rjrAFT+8gfvufgwlOhFhJ4k2RJUqQkh02nhkoLbj/zvrzNk3hKoLU+8nS2tItQ/kYLNRjWt2SPOKzmiN/Xd7Rx2lo/RbTxPkdR22QNdas2PHjuUvfvGLf9bRUZkPGoFCZ/0EQQVRd1COoe8APPUET37vvxh8bC1m8y4mG8FcK5FGYzE4MirlMk5bslRjrSWMIu+BtxbnDMr4rN1sYBcD23eyZs1jDF31E6IFc1lwwQuZffkbkQuO8VpEqCBSPuaO79p+KKJA4r2WcQwf/ptXAZLbbl5JEPYgXJAz0bEW8WAtJi1CaP9XYZFykI9+4gOcdmYPUQCp8xGF321q73Ln66YVMX/43tMAyfe/f52HFnZRHr5oJbRZ0bLeC0x2aHcvy1yY589PWFI9QFRK+b0/uJw3vPGkPF/CkWFQHqtuzDG2qPi8Tfg72l4yr8YVesqWzfC9H/yMW29aQdoIUXIGjbqm0lkiCARpI9mXprV/fuaZJ7506aXn1oTcSaoTOjpmkiUDIGsgGkCpNf/i1q61FkfpKB2l3wE6SJL2GNxrOJPYv39/zigtSZJw1133zL7wwgtv7enuOV4FRYmMt0IxmUfG2LuP3V/9Kk/ecANdmaFiHGWnCA0oo5Eut2SFHRXbG54w0FYJLCRGePUhk5I0gIaKSDs6mXfm2cx8w2th+fHQ3QGVMpk1GCeIVHwQoe7y2VpAoTUYDV/6/M38+IpbQE6hUp1Kf/8BgkARhAohBFI6rDPDO4Q5iXTSN1yRAUI6XFojy/pZcsIM/uZv38/yk+SwVf9djZ0XZEcII4fFoXAIhIPrrlvH17/yX+zdYylFUz2GOSCVoaEbhEEJZ1vwraVSCatN/gykBwuyBm0S6sleZs2t8Ifvfx2vvXwx2pAnJfrabQCByn+Ond0w6nkZQIHLhkiNIy51YIEd2+HKHz/ANdfdxe59NeJYkaaWOOrMT9RGisa/BW7g/wlR34hIkWIQEQyAbPg+Amh+8F//5seW52JYd7C+e8NHOGvmTH/eOIrjSKfW4e7H/X0Dw88f8T7v7d074flmRClBcf7Ipknjed96Jk+e8P61em3C+//rt781/D4jzv/kJz854fmPP7FmwvO7u7snPL9crQ77feT4BwaHr+9IGqrXh/0+8vwlxy+Z8PyD0QMPPDDh57PmzJnw/kE4sc04acqk4ecf+tCAg5tU9967YsLzp06fPuz3kecf7PmNdJmPfP67d++e8Pxly5ZN+Pnh0kEt9IIxGmOZMWOW3Lt377emTJlyvHMO5yxgEJn2VnGtAfffx82f/QJTn97K8RqkAYmjaBIh8mxuzzblKJWiiTfTLP3xJIX/LQTK1kEKloyhgX30/uJmNt9zD8de+hKmvPNtcMwCwiAijEocvJ68nUFKlBRIBR/64CWcfPJyvvvvP2fDhl3EscKYFEkJgUA4gXQtS0wqP2adJTirMRaENMTljFe/9gLe8a5LmDaTMVWoo9ROEoFD4HBC8rrLlnDCko/yuc9+lyce206SBMRRlaHBIVQYYVBY4feRihTWalTkt7UxKUlWx5oa5YrgvBcs44/e/0YWLRa+fIyMsbwq7WM5KCl8Ir2sEoewYxtc+7O7uerKO+g7ILF0UVI9KBzlqqJR1yDSX6IG/q9j8B6hBlDkuOsiBZl4i1ymPu1EeAwD2ay6GMHCRmrsz6oq4ygdpaP0P5kOyeWepilBINm5c/uHe3o6X+GcQRuNDALAoQIBtRrJFVfxwPf+i+m79jHdxkQ2wzrzvAkx6STdUZkOHNngEDuu+DF7Vq7khLe+Fd7wRphcoVnEPi4pQCFxWESTH6oYXv6qmZx7wbu57rp13HDDLWzZvAurU6wJcCZEyhghJAiL1imOBkJmJFk/c+fO5PQzl/Pa176YE0/qQDQzqMeK07ardUclPrTKuR1w7OKAL3/tvdxx+zqu/entPPbwRtJUU1aTMcrgBL5bmlI4Z6lnDXSWIUSDqd2C446bwZve8nJedOE8hPLFFYHAB7gBhxl2X0+HJhRN5qM9+/fB1Vc+zA033MPGjXuxLqRU7UBJgxACYwz1ocZK59zfCTV0jWQnUvZhRaMJ1yqE8SWduRB30HSrH55Tve3bzbZ+4333We63g+I0HOTzcdN2D4Iu9Vzd/zf9+bNev2d7/ojfDzcn41mv//NNhzn+5zwn5dnefyRNZIAcgkC3OdrJqgdWLT37rLM/qlSAc45ABdgk8QlwtRrpj67ige/8B5P29tNRT1FWIAMFUjyPj9Ri04RqqYxJU0KTkWzYwsqvfJNJG7dx3F//BXR1NGPqE5Pv1l58tZY4QiWoVOGtb13Cay5fwuoHenlw1RqeeXorG57ZhrUp9VqdUjmiWo0plWOWLl3AySefwOlnLGbmbPwLIyECGnYIJaM8sv+7B/N6+GQxFiIlQcGLXriEi164hCfX1Fh531rWrN3Epm076a/XqNWGkFLR0VEljCIWLljM8qULueAFS1l8LCDz4gjACY3Mt/4RN8TJKxc2bnLcc8+jXPXjW9m1fQhBB6XSDLIswxpHHFnq9f7Hh2q1zzy+5oHvn3HqWZlwQyD2I8QAqAQrDUr4HACRu7CsfT5y7o/SUTpKv80UHIolYq3ljNPP+GhHR0cnSJzTSClRQQj9fXDT7dz3uW8yo3+IScr3aEYJtDN5XfnzNHphQWlq2QBCWGQQEmeCSUMJO++4g+Pe9HpYusTXsx8mVWKviBRl0N2d8OIXT+ElL34hjQRqg9BoQJJY4lhSKkNcgjDySoGUtBnjFoslkmEbFm+BfXeUcbfTKLtMumaeQikSOAvLT6qwfPlpWE6jfxCM8Yc1UK5AFPoEx8IrMrI8MBB5PBryJLh25cq2/RzxbjjQ2lv3zzwFv/jlg1z9s9vpG9A4qwiCLt8G16YgMoxuPF0fyj677qmH/u3YxdPr55y9AGO3gWzg3CDW1IhCixC5h0AIlAhxzmKQWGux1n8WRhFZmvpmMnGMMQZjvEUfBAHWWbI0I5QhCxYuUMY51q1bF4yxpM8pTeoaHWP8y7/8y+QTn/iEg5ZBMB6N12GveOZRFA37uzEGay1BEPh/Z0UoTzB16lS5YsWK4LTTTpNhGPLkk0+6k08+WT/55JOmWq2SJIk3RoKgmUQpEcP+drDxjqSRDoYwDEnTFCEEWueJvmGI1rqZB6CUaoYypfNr9MIXvjC45557gs6uzub3XvOa1+jrr7/ebN261RVrYm2rgZMxhlCq5vjr9boHUzgcOgj70drz+lKpxODgIEEQoJRqrpOwrjmfNE0JSyWM8Tkss2bNEkIIOXfuXLF69epg8uTJR1yJNB4dgqk2cQ9hY3HOEccxWZYhlWrOGfyzkdKDpymlmDVrlrj88svVlVdeGUgpm5+/+c1v1j/+8Y/tjh07bBAEzf00ck8V83fOr9vhsv+J7fNDsNCNMaxfv/60M888801SBlirkVJhTIZqpPDI49z3pa8xZyhjhoiwJsEKixEWK0Yz0+earMwxrJ1CWVAGStrSmaWgGwfBU22H/2gfqE+SaqFwDf+8FPsDC85JD2wz1jxFu3BgnPscpfFIIlshY+FAGISSzUyUQEAzJ6otn8ENw8TPcyTc6K1+JDnijzy8jeuvvYv77l7Hjl0Nos5ZONWJdRlWgiRBiPQp3Rj8Um/v9n+fMqWr//jFMzHsJioJ0rSBEiCkR7Jx1oDMhYvLmwHloxJCEIQ+LJAkCUYbJnV385nPfKbzVa961cK5c+ces3///kWz58yeWylXpvX390+yme3Ys2dvpxOISqXSiXcO/do2mrU2fve73/2Fb3zjGz+aMWNGbMzEZZ/jMfji783KhZzBHjhwwL761a9++rOf/WzlPe95z7JGo3HK9OnTT1JKLdy1a9f0rq6uSVLIGOCEE07Itm/ffsBauyeO4y29vb2PSykfvPrqqx97+9vf3hdFUVM4ZllGmK/1s6Esy5ohllKphLUWYwxKKd7xjndEH/zgB+edcMIJx/b19R1//PHHH5tl2Zwsy6bffvvtnR0dHd1CeJgs55z73ve+12+MGdRa93Z1de3YsWPHBuDJp59+et3Xv/71DV/+8pcbg4ODTQErpRxW7fFcUBiGCCFI05RSqUQQBKRp2rzXT37ykykXX3zxMUKIY4wxx06bNm12o9GYMjQ0NGnXrl0d5XK54pyTnZ2dxdx+XXuxiNq9HbhnvC8VykihKGmtCYKAb3zjG5XLL7/8GCnlknq9fvzChQuPqdVqM3bu3Dk1juMOKWUXIKSUGG3cN7/5zf7Pf/7zNWvt3mq1unvnzp0bGo3G+qGhoXUrV6585kUvelG/lJIwDJsKz/M06fF0HM9Y9uzpRWf6m9NnTH1f8VIZbVCBhIEBNv3pB0nuXsk0HRJqg3YZTmqsyCE3hY9RN294kHmMB4s3VgG+RWKF13JDaz3ErHH0xwF7ZnZx5pe/CCcu94XmY5LGJ0YBlBgubIu/y7af3kIUheXdvnSOFsrdMEFe/DvgYGs9/H7/s+hIEe9GZrmPyinPIdidxAv0vNGNE16wDxfengoRIrDI5gXa1t81v9D0oIycx7DRNMvE4N+/cz333PkYz6w/wFBaoS7KhOUOrM6wyeBj0tS+0Lt30xVveMPLB4xr8NDK+7HUCOI+6skBpFBIFFJInPTVI0K6Zuxc5hqMCCRCSL7znX9DCMHaJ59cPG/e/EtmTp9+cRzHp0sp58RxHKVZw+cPSJUzconJMlQunArGMW6W+whN9EizjNs9T0mSZMaYepZl4mBC8mDCZ8TnIkmSZGho6M6Ojo6l1Up1sQqUNNoMw/0vrN92cs5hjUVIweDg4OYkSX61e/fu//rQhz50y/nnn5/IIMx5kx1WpPqJT3xiwvGtWbNmzL8HQUCSJOzdu/e4mTNnXjBlypQX9/T0nCGFXKCN7ozjeMzzRioWhULQ/rckSerOuY179+5dMTg4+POHVj90+5lnnblbCIHVPmfD5U9yyfHHTTj+g9GqVasA/xystZxzzjnBvffee9r8+fNfHobhhR0dHcviOJ7hnJPFs2qfg0c/jMa9/nNPwz1rK+659xLgluanYvjPubNmk6a+bdjWrVuPnzNnzkWTJk16SUdHx2nOubmlUqk0cv1H3TFXptp/byOdpun2JEke3rt37207d+685YILLnjsoYceslEUceKJJx7W7EYWYo/knwe10G+88cZpr3/d6y+TSJyxCOUzwUkTGrfdxo5Vq1hYa2CFRkYxLmM4FvoR02h3Z+uahYsy/57zABvSgcVhRass7tDJl60Nv/9oy7rgL9Y4nM69AyqPfco8sNq83qFUqo/990OmduFU/C6e/XUP+fwiy18Mn+vBagsKmvA7Yyh/rniuDhwCkxmMlCil8moKfIY87dauHB3daArzg4zHtc1LwLvf8yre9Y5X8fhDGT/92S3cuvIhu3XXujuE1t/YsmHttX/4rrc0uruOJ7X9SNtAhn1kjX6y2gGf+KbCXMmN/EjzPSxwTferEA4pHEqpUlpvXDZ92sz3nH/+iy4IQ1UNlSIzhiBnIFEYN93yIL2yHYakOiMKWkxoPOjXCeXpQXI1x3u74jgOgWdv7o5BURB2dHd1v9ZkBin9+6oClRsZ/ncrfG6Es867hAM1TDEol8vzOzo6fn/ypMm/f+WVP3144+atXzj99NP/8xMf/ZhWQnoP4xG+ONZqtm3fNm/x4sWvnTp11uuXnrD0TKlkR/v9C+WjcMm2G0pB0AoH+ZweOUo5iaKobAVL586duxR418KFC3cNDAxct27dun+ZPmXGyokf6uGRcw5jDE8//fSMs8464229vb2/V6lUTo+iSDpNk8c25+e8VV8oksME4XNtlA5TxtvXyLYwSKxwRQOosYzCW265ZdrFF198eUdHx9vnzJlzTqVSqQBNWec0PoScU7HPCgWx2HMFFcpMcb7JsiCKovlhEM7v7Ox8zfz587P+/v4H9uzZ81/3rbjnSmDbc7UcAOKGG24Y9odCsxRCEAQRi45b8ubFxyz4EY6WvNMaDuzi0b/536h7H2TGkB7uSBa+nA1AuEOL6Thhcg1ZAdL3qm4DHXGC3OrPoGiugQRbRtmA0Pqa9tQ5DpQke2Z2cfaXPw9LTx5hoY/MKmwWvo63RK2NOPKncDjns6yFEFhh2yweL1Za9rnM0+6aN8biCvuRqLhuOyraOEKn/RoYn6Ng89Ok9cIry180hRybmY/br9vf0zTrsotPxlifdv1FDl/ZwqgVtj27/zDJFSPAr/WwMbdQ0lyhbOHasrrbdVfp3fc5AxXIMS17jKPZMCjIPy0UuOJKRcl6LkO37suS3v0H/nbezGmfk0CSJmhhCCsRUsG6Rx5HOQsibbtRzsCFZPYxCwhCSWgAnSFxzJk3Vz3wyINvW7Tw2L+uBtVTwG8LI70GXjcGjKYcFVaez9FweZWGzRUahUC1LUn7vZtLPK7rzzYhi0FilbddA+nnnloIwuI527bwRWFDFGV2w9exdfWxRjNsZIBo87bka2/ynVW8G7KYv1eLbHM1/EdFA8Xck52vTWvTKgT1Bmit71r9wP1/eeopJ60MKxGDLsNYTZw5pHEUZbreNRtQS+rEcUySGpwxKCXYuW37BbPnznpvtVJ5TblanSSkaM5fOAm2JVickqNMBm1zIS8FzhkCWfBO1/LcCNEMFIp8zm0LrHfv6L167don/2HRccc9aK3GOk0chqRpQqlU8p6tNiqVhkNbD9UToigE6+PKX/3qVye/7a1v+9OZc6b9cRxGs/y6ytEP0ALOoVPd9A6Jos1me9TluXRA+innN88jS9ZXVZk8T2bVXfddfPxxi281OFQYkmQpVsCDD65acMIJx39g0cIFfwDMLgbX5N/FFjF+XgTeO6udIVBqGP/IMg9K5ZxFRarFF9uMLc/KinCIxWpNrVbr7e/v/8GGDRu+dPrppz9Vr9cplSpYY6lUqwwMDLB370gch+ELeNySYyf4FGg0Gs1/R2FEHMcvLV4eV7c4DWQpPPwo+1c/RElbEBbXdnga7mofj2RuWXvB3/q+lQUSWPt3JdIpr7nmIC7KkW92H7M3+XHo4kMyPrSoaP0YxpAcSJs/IJDCIrAo65DaIFONTDNUqgkzQ6gNIksha/iEqfyQrmVZFZdFWH/t4veDDt9n4BX/Fc6EAElQMFcxxtHMDSiO/IaikFhHXrBRXMGR60kSb8If7iFyxl1Y3q41XisKQS7aHk/OQJvfC3B4pD8LeRKUQ6CRRoNO/V5OU2gkkCa+WVDWgHoKiclf6LZVErS2toK5M8N4yeLuz8Zq4IFNTz/4kjisUQoTrB1AuDpSNEYI87aVFZZSKcJoA86AdWzZsuXcffv33rHkhBO+F8ThKe2PpeCNkZKUCzemM2A10liUMQTaEmWO2PrnL6wX/s1DDz/EBEfzHJsijSYw1qcYOYh86L89mMZwh2C+YhN4Qibm7eNosla0NpfFz8NkSGso8l4CbYi0IcgMwmSILIGsjjV1rNNt/FZgHZRL0FkJLjjnrLPv2L1nz5/U6nVUIFGqlZjonG0qBUGgKIWR51vasG/v3peX49JNJy1ffuf06bP+oNrZNUnKEIFCWImwMk/VcaAEQpF7kwwSg9QZZClB2iDUKZHNiI1BWY1yGmU1whmE02AN0l85F6yFOxIsBNNnTXnjBRecf7cL1Kf/9lMfrwCEQUhHR7Wp0E50KKUwxqIzzbZt2972gfd/YOWChfP+r5Bi1qinVni9bJ7A4hyBCv1rm1lIdOsZFUdmnv2ROn9kGkwKWR4edc6/D67VsrtIaPMJbIb//M//7LHWfOqil1z0wIIFC/+3cMwW2NzAsaO3Xa4w+nfQEloQ2iIT46OyqSPMDNI53yhJO6+ZWeuVi1xhdhKcELmqaZGBoqOzY8rsufP/7Oyzz36wr6/vH7///e9PyTKNVJLa0NARhSpGudyDIEBrjVKKl1z8knBgqH5mU6BHXssUmSO5ZxUdQxmxOVI/Sh77Nv7fWb43JV6QZyrP7jUgnWoJXSeBPCMNiXCyiTxnhF/jTIEuvOTj8YViFKO+0656uuEfyiLBquAm+Wfa+sM4/0CL0LkDnM1li/Id41S+26SAQBAhvdFhBEiLEd6qD6UuAse0u/yHkWvzczdd7jY3YMZwM49HotjQw+ffsrXGgdAdsXa58dZUB7xHzLXcB4dEI+Yp8YLdtj51MlfaJAS5ZeAk4ETL2+jksEu5Ym6FkMqy/JnlV21afIUSkYAMwFQ8pHBIM5QxTEkClAood0Wnn3LaaTdt27LlM0uXL/vE+qefSohGWscjpuZA1xooY5g5Z7bctGHj3y5btuzjcbkUZVhCEdDc3MKhbPGcpE/rTzNfweHyPedEzvRsPmEHqrBgxqGJknOKPSG80giy9WJVyl7hiQqX8Agt+mD7rgmU43LvGxQ8QTR9ow4lC19R0BLi0OLY5AxU69Z7Z4r1MH6DKC+1pQr9miOxub+smQ9jIcsa5QULF3x1y7bNs7tF98fIdC48i3wHCMMIrQ0d5SoPr179gqUnnvixWdOXvgIZDsdxtuTWBbkczAfvUrBZ2z50YFUrOz1L/b+LtpXkPyPfkMoDKeAxFIqx55cueJkIKE2ZOfUjn//yly58/MEH33X8sYvWam2Qst2X4klrPex3KSX//u//MeVd73rnFxYfv+T3izBW5PMMW2dL63mNNa13qOB9yNzdYL2SnDWaz/pIKo5GUTvkcdqASgXSIaiU/CI4QOeswzqiKCLTmv6hwcve8573/POkyd1Lmu9x3p5ZKDDCM22R2xMeHUz4Z5YJr6Co0O+vVOeLLfw849DziSTJ+Xx+ERVghKDAOGk6lZreR4eUsnPmzFl/8573vOdNe3b3fmTGzJk/FsKRFYrKYdCo1dVaNzMl77rrV9MrleiYYk86Cco5cLD7ifVE/XVU+Vl2IxMWLSS1IKARBDQCSJUlDfxGCyyERhJaL9wjA5GxTRc7eGbgioP8mECYt9MhxXpd7oo1GVjdavmqLWzeAnv20b91B4N797Fj0xZCA6H17/Dg4BBhKSbsKCPLMdPnzaFz6iSC2TNg2lSYNRclAz/WAFSpGLTMmd4EXg4HDNRp4tVanb9o49g/I5msyFeg8AoooFqGIEDJqGleT2Spj+V5Uzgk2mvQBwZb+YWH628LAujqbDG7/G2wolAcHEXEsSk4Re7yMiANuaZsweUWwmMPk+7aTu/GZxjauZukf6jZ9cgK71LrnjSJGTNmUJkyFbF4KUydDtOmQKSQKsgZbS7kBESimV8i5yyc/zdbN20+46abbnrHqy57zY4mzPEYJByETvCjH/14yq4dO781fcaM1+qsAThCFJmuE8rIzyez/vla4+fR1w/798Pu3aT7++jbvZfavj6CzKKHGphGgk3rWNdgIoHuRiLPNQdnmzqCFhKdC/RQRiRI+oTl4ne+HU4/DYoxauvLWHW7u32kopiTdMMcYKrpns77IYeB506xyBWGwF8iw69BqsE24Km1sHUb259+ht7tu6CWIFJDaCFDI3rKdM2eztzFSxBz58JxSwiCyO+tQDbBhQigo7NKZjTz5s7+aO/eXtdRqXzcGDvstbHWcvXVP53xqktf+akzzzzzj0RcCshSRpXzGHJrLd+MpH4Pmgbs3gGbNlLfsZuhLTtgsE5tqEatXsc5i8xL0UQkCKplqtOm0jVnDuVp0xFLl0NXDxBCEOEVCdoUHE8K6Kl2nHvaaafc8eRjj79h+vTpd8dxjMDn+zTDTyPi7WueeOLU973vvf/Z0VFdZq31gUEhkUKSppooCHJe0/D70Ano7YM9e6G3j6EduzBDDfr39JIM9SNMHZOlTcjiJEnG3m+HRXkQ0EEgFG76NJZ+4L0Qz2hJtFzRcc6lcxfMF0+sWfOlObPn/K8oiqgNNShXYoblGliLQnsDxNiWkrhpA+zcRX3nbvp27aV363ZkZtH1xHuXlUTjUNUSxCFTZs+gY+pkqrNy/r5gESoIvCteyRbvIMyVaYcKY7CGamf3MVFYumJwaPBfPvXJT/3Vp/7PpwZHzvxgVQyjBHocx806ygMHDswXznV6Q1AgC/tr+zb2r9/I7EoHzugxLnsoZLHCcKAxiOzpYXcU0XPqKZz4sotgwWyYNclrq4nGbt7OrjVP0nhmC/ufXIfcs49JGnqiEvU0Q2cJgXMEUeg3GzkTnUCgjy1anI8dG4MUEiGKMl7hXbCNGmzfDvevYud999DYvYf+nbvQjQRdT1BCoGweiRES6xyV5g0FmYB1WUIYR4SlGFOKmXb8icw543Q472xYNA9kCSUEhB2ecRXWz1iUGdK77+aXX/kac3VGlNQRxRNtZnW3GOmoWLqweNxSH644EAfMevnFHPOH7/XMIpAHRSqQI7wYPr/W+LV6fA13feIfmFbPXc6H0DXNOUemM8IgZHscUr74Es770P/yCkcgvAcGWikdMqBZmigEUPdCvZ569/Dm7XDfCjatWMHg1m3Udm4ntimKFOFsc+QKkNoQWkdqHVvwnh7bXcV0djB/2clMPftcOO9cmD0LyiVMI0VFHfm8C1ewo2vKpItf9apX3vrAypWXlavV9TZNcVI0E5y09oxRZxn33XPvgve+931XSsUZCAii2K+nbhAGZW9xFzJh115YuYqhlQ+wceVKdN8BTJYitEFYh3DeGA2dd8kKLKrpTjxcsl6YS1DCe2uskDgjsEFAIgy88lJIDS4QCOegUWf9J/+O2mOPESrlk0YLV2Y7CZ/4V/xbOois8t42VWIvGdnCmZz7T/8Aqgfi3PVYKNXrn0LffCNbVt1P/8anYWiQtJF4a0wGhEISSYWQ0AgseyX0mYBMBvQct5hjXnguXP5ymD3TAxeoPEFRGsJceZg6efLHntm4ceOs2bO+7bRtut737dv/pre85S2fKVc75zf3rGq9JAaHtAYZKm+ZJkOwZw88sYa+B1az+/E1pLt3Utu3l8g5YhcQWF/RMDKL0KBJRUYNx04RkKiA0rSZdMyYwbGnnAannwEnnwadFegMMRikEQgVEAHWaro6umecfPIp1zzyyMOXz549526h/B5MkoQgCIYlsD311FMvueCC838MTAZvrTsyFI46KVJqz1dqdXh6IzzwIHvvXsGBzZvZvX0nGIuyEAqfmyCsA6G9si2Ff9ZH2up6GLX4iLaCgdkz4b3vpcj3MKJpAzgXB8dv2bH9k5Mmd19SWG+V0BuhOkmx0hKFMdQGvLzZtw/uu589966gsX0HA1u3YYdqNGp1AimRxiEQRNJ7q53wvMMqgZGwQwqsygMqUjJl7jxmLV9G5QXnwgmLYeYsiGKIpVeEcVidNhM8w1JMl5Lv+8QnPnHy6tWr377omEUb0iwlDEK0MQihxsVugDEEepZlzbrGU089dTpY2fL65e7moRpBlhBqg8lf/CMhJxSlST3sUJJTf+9t8I7fh8ndUApzX7wBJ5GnnMasl77MN355egPceQ8PX30tO/f2InWNniiiRwR5DMN675w2zTiGzxxuufGAcTaVQKE8mkxmoV7z39+7B3vvfTx15130PvwYHf2DlAYH6VKSkjHew4JoZtpjfQJNuzYlhGzGhm0jwdUbNKSgf+evuP/ee1DXzKGy5FiWvu61cPrpEFso5aV04xbzW6LjFjEjkrgn1jMFgXWJD6u5CO821+NmOCMswmkk2itApYjeO2KOed1bYVKndzOPiDqM8ySbXzLWEjiv6e666WaqGzbSNTQyhjw+BWFIbahGGAbsjELO/YuTvFZrMiCmgGkV+fw1ic/4TpPck5K7wh58lH2/uJWnb/sV0b4+glqNSVHEdJ0ROo0gw+WKDLSeW/tPF0LjgGao7wC7t+zgqVvuwM6dxzEXvZhZb3k9at5C735rpP6trsTe1QaUOjtOOPW0065fuWrVq8ql0noPAiOaGctZmrJx46Z5r7rs1b8oVUon2OIhZZlXTDKbx/JrsG0b2679GRvvuhf3zBa6GgmdJqMqwRnbRJlTyKZy03zmz6LaxArbtNC96iZxVlELNU5ozwDbLXxtCbdsofuZLSgpwWhEM2lpOMnCBS00AktkQ5QJECIikTCgMti1w6ME4aCWwr4Ben94BWt/+QtKO3czKcuYbQ2B0Vjn8nwcjRQCKQwGTSPwnpBYB1gkB1Y/wkNPPkF20/Wc+Za3IF77eohTKClcEOSKmU/sWrBgwefuvvvu+5ctW/bov3/3u5Pf+c53/fNxxx33Hucc2KKWWPpsZgzSZgRJ5tGldu2BRx9j0603s2/N4+x5ch3dxjLZwCRjmO3w45wgNGGFwebuLYvFCI1Jd3Jg01YeevAhkp9cRcfyU1j2mkvh4vNRHbFHVNI+YVgGEVhDFIVTli1bdtXq1asvmjNvwZpyuUwcx82GRtZatm/ffvHy5cuvArraRoDCgB6krA2IAB56mGeuuo4996xA7fDPoGphcZHAZy3CujYPQBFDzgX5s61+GmYUSKwTiDT1+z6Pn7RCftKdcNLyL5bCsOp1S+cjUIF35QZx5BHC6nVYu45nrrue7XfdRbS7F7V/P90OJgUBkcAnKFqf5+KMwdmWh8MWoTophpXFORSDqx/h6UfXkN5wI8m0yRxz8SXMesmFcObp3ltjKfqNN0mpgK6urnPPOfuc21avXv2GOXPmPKCCwCdhJhO74UcJ9KLQHmDWrFlVkf9bOetdEQgYHMBZjbUZhwarOjZpEdCXQufpp8H73+/dqzqPlZHhrUc8dypJ/6IsXwqLF3PKm17Pvp9ezfqrr2Jg+y4indEtIiooUiShCryapoIRZR+jpVPhyfOOEOfjJXsbsH0H5rofsOa2mzH9A4ihhGnGUdWOkgu8e1MqL85EXvmZr5fMYyRCSO/iyg24KAjIjCGxGqktXQKqGJJntlB/ZhuP37uG2S88l0nvexscuwCCSq7JFdTk1H4jzJ7C3AvOZcvjTxIaV9hkGOnzDiQOMc5LJJ0ixPoWsUKjDfQ+th7ufwQunjVhDsJYawh44ZpY2LiFZ267m0l2An1kDFLCEUj/Wp64fBni1BNBCRye2UZ40NZiFRTOxyQj4QXrmvXUrriaTb+8Dbl7H9OUJJSKIJRIl6KKOLhTCKuGj014xkMeTzPOIjV0KkXZQiVNGXpqMzs2/5Dt1/ycM974Rnjz78HUKd6DMCJ+U+3qXHzaGadfe91111183HHHbU+SBKV8CdV9K++f9N4/eu9PXSBPSCWkaUI5iAiE9KEKGcOB/ej/+Bbrf/lL6tt3MV0rVCPxmAtCYHAo5asnhHQeLlZJhLWY5rzyPIjDVrpHB6NEji/vAK1aH7fnCSgETghU/pQEZrhrs7nUkgIsSGJzPqIwqQWlCGXo804CC41+uO5GHvnOfzLw9DP0KEXZOjpkRCi9detyHAqHVyoN4ERAnoqDReIEVJ0gqKfUntzKw5/9V+as2sC0P/8ALJ3tw0T4uKl2hjgud51y6kn/uO7JtZ/5kz/5wFc6unqWgsRp45sAhXmVgcm9RbUUDuyHO+5lw9U3sOfxNbjB/XTGAcepCGFTIhyhjHKFf8LsBhzeTetDiBIpBcYZOoSlasHVBqnffy/3r17J9BtOZ+EH/ghOPAZKER5XAxACKRXVzu7p55x93nf/6AN/dNHff/ofasL6Us/MGh5+6KGzzjvv/J+EcdzVfv8igwYcbNzKzu/+gC033055bz+zVUxFSbS1uaGXtx92vt6iiEUXoEnDWyQ/CxpZjSQcvqmRbeW04AOFBicr5bhqEF7Q5bW1LtGIMIDBBqxYwZYf/JDNDzxEeajOzMxQsYZAxXnE0YHM0Sysn5NppueINrjmsSKahq4gZHJcIqun7HtqC8+s/zZbrrueY849g2mvvxzOOx8ZBFitm1Y6UhDGMUEYLjjr7LOvv/fee1+zYP78ldaYMe4ynEYJ9HK5TK3mWw729fVVetqgCCmG3RjEkJGJDCWKAM7hkwMIYhacejpUOyAOfL6b/wDv3PCxNOcM1iWozh6Ia9BdYfK738E5F72QzV/8MrtWPEDayOhWXojKHGRjtEAqskhav0sccRHjshGsWUfvdTez7he/IN66lskuJRCSUEikEQTWx02VUiRZ4t2pFA1YZZ645TeekGLY3bJMY3AEQqGkZ8QBAqstaWao92/jmauuxj7zCGd9+C/gzLP8iSpsZVz6P3iXeGcHsy97NbuvvoHBLTuIRb4Hi2xKJgiJOHBFsoERhMIwRQbsv/1XTLro4kNLqGuuYdv6OgcPP4beuIUKolkzPREVlopupKgwpi8QHP+yl0BHBYOh6CEgnGhzOeUvcZaByciuvp4V//JvdG3dxbzMopKEqLNCYg1pkqGNQTUBPeQwz1JRdljkX/gkKEUpLJE2ElSW0YkiThIiEVDbsYsH/vU7RI+v5aT3vxdOXgoihjSloQSRCpFAqVQ64bLLLvveK17xild+6UtfSowxaK1505ve9K9BHJ4RBooMD3EaWOfzM2QADz7CQ1/5Ctmqu5mhE+ZEFXRjCGfwjCkM/eOxLvePuOZ8xgRhOiIvWisR1T8fi/V+MF9x4GTbHvHvm5DCJ/BJL86VU8gxBuTX2+FkiMAS2BAhFRYP0hHn+Qns38eWb36LLT/+BbMTwzQlKAlIjUGndQgDpJA+35FiHxXjpvn+F0MIgRKKDm3IEsOOG26mb/9ejvvHD8OCeT42H0boJMVFjp7u7lecc+65l0rZUv2klD6ckAz50NRQHZKU2s+uZ+UPf0y0dSfTBhNO7uigz0pEw4C0OCmba+HySg3Znv/RvvKyKMFrswABspRQCJ9z1WhQNo4Kkr033cqejU9x1of+GC58EXSUIUkhjDyqp1QEcXT23/39pz/pLB9J0oQ4jPjptdfM+P3f//3vl6rlnuZb6shzhBqeJ/78Zlb8y3fJ1qxjkQ2ZJCS1gUGEEJTLIdoJrDUYqxEyh/5uWqpiVLXS4ZIcI1Tn86a88uBGhHUKRUQ5gRUi37HkuUYGkRk4cIBnvv1ttl99FVP7+5mfQZz56giJwwUKg/HZD9a0heZy/pN74tqfnXB5kFO0+Jm1GjM4RIhgdhjTZS22dx9brrmONSvuY/br38jiN78JOXuO98bFYZ78KBBKEpdLM04//fRrVq1aecmiYxaNjWLURsHMvF9yQWEYUiqVKJVKxHGsrRPIPBFJidylLSw1W8MGBmWDg6K/jUcCsDrzFlZS96524z8wgcUgmqqCFQpiH5FWUclvuI4ATjiJ+Z/5AvOv/wkrvvmvpFv3YrOMRFcYO/o7htnpLKSDsG8//PgGnrjyeuzuncyT3toInc8oFZYc4QsSfFatUw5nLVaKZkJy+13bt7Jnqq3FEkKAEz5vRlhULKmGAYGCofWb+NVf/A0v+ue/h/Nf4E+TPq27Jaa9xcrs2Sw6/wVsuuoaerSkZDWx89aGnlCWSjIZIJxF5rHMSmZZv3IlZ/ftgc7At50bsV7F6+MVWJGPrfhQQ6PB9ptuYrpOCAkR48DvDhMyJvAeEmvZLzQH5i+ESy4CmWd3S4V1AuscoSQPpFtwQ7B3L4Nf/QGP/fRappiUqjVYm6JKMGS8uz+Q3m0mnG2v9BlGDoOQwz9rNIb8SuWAH4FSWN0gFjDN1Nl1xw08sPVxzvjTP4OLLwYlKVW7cGkGUUgcx8Rx/JIf//jHn54xfcZfWGfZvHnzX82ePfMN4KtvAGIhPfOsDcGNt/LIl75BdccOqgpCF5I2MqQMfBKvwLt8Ra4wuhzvIE92GglEAmO/CROTF9iWIK8g9PspsIbAKUomABPik9hacQurHNomBE4TCoUURXX8yMV2ICSWCOks0kiEE3RO7uRAViPTfbBtCwe+/m123/8Qc0RIyQlCBFhHFHiF1jmHsxqRz3m0lcSwv7u8BAkFMqszQwRsve9Odn+tyvS/+TDE3SAF5bgK0mLR+Eiwy4MOeEstKkO2Hwb2w0238fRPrqPvsadZ0NCUtSVwllr/gTzx3ZcrCWu98tE2rpH8oiDjXB7mzJXlvJooEqo5IeEkCigLmBXA0LatrPzU33PWO7fBu97pjSQHEPqEdCGYNGXah+5fef+Pzjj1tAeNsbzmNa/5Ukdn52KNpeEMJaGg4cA1YKgPvvVdHvyPHzBJG8qqRGA1NWMgDvwcMosU3iPomoUINi8nHv2eTRRiyL+Rz6+1KjIHDivyLaywTVwEI5wPCQ1TLlsCvtmEydZQVkOSwb2rWP25z6G27mR2mlLRPvavpICwPQlY+fyOESOcSN4VQr05diEgD/tolxFJiUkzZghF195+Nn/xq2QrVnHi+/4IzjndZ/LGHfn4vfFZrZZmnXPOWVdcc801F73pjW/ZkyQJxtpRGAJtq9eiWq3WdLsHQVC3zvpkmMLaFQ66OwnDiEA8OzAo6UClKb1PPJnHJ0xTq3ZIJIqim2RB3j0uMUUmbDn2wDGXvZpzPv+PxOeczo5KQCMu0yqCnoAyoG5g9eOs/eTfs/ob3yZcv5456RDlgV5KRjdr5QtqZtTnm1Xk2nN7/GSsYywafi2LkIaSM0xJMmbXMu74u0/Dg4/kXUEKh4pt/n+KBQldl7yYoe5utPRJNioPKU30AvkwgF9LSxFrNMRDNbIVK7zAHAOLW7YdrYv5uCK1OjzyGHsee5zJMiScAEvfr2urWgEnqWvDUBhw7KWXwLw53u2a+1KF9CAOQF72Y2HLDh75+Kd54gdXMX/A0JNoStrXI3tN2ealUK0Xcaw1KZ6lcX7awvjDk2o9o7zET2ApmZRjS4qOzZu5/VOfght+4bNk6zVEXjqlkxTnHDNnzvzQ6odWv+SXv/zl4rlz5/6fIqzlvT35bVIN1/2c2/7hn5i8dSfzraSaemZTjLnd2JXOxwWL/dle9THRnjssEgUeRM70nB9zExOgHThGFHF3v0be+jzY85cIF+TXg4HaAJG0uH372f33/0jfioc4jogu24J3blFuw8rWvw/lsPmcBJrYNpgpHU9edQ1ccSXNciLn0M42qykcbUBFgfRJn1t2suMf/pH7/+7vadz/APPrDabVUzpTTWD1MMbVwtsYLRAm4hc+Xt86xKhryDypUNOVNZjZN8SKb/073Hg7DKV+QyvvJnYCynEUHb/k+I9kxrBnz54/mDt37pt9TbtCFMpCIGCgj52f/QKPffc/mdNfp2sooQsvWJtjHDG35pZpenTyObR5w6xo5x5jHMNwJORw69zJZkVTAbSDG/EdPE/z3lmBB55JUM6AzTA/uZI7P/5xOp/ZyJT9++hONZGxPqSMHbX+7es91rMbi8b6bjNXB1+hVTKW7obmJFVGrn6E+//fP6Cvvs7nAplavhYuh3GOqXZ2L7vooou+Xm80iOOYMAzGTI4bJe06OjqaCRNr1qzZV2wY5fziIQRMnkq11E3owkNGghuLlLNMr1TY/sAD8KMfwoHdYAdAtgL/xWYJbX6QZxXiYRm8nzuASdPhpJNZ+I9/R/flr2ZvqQQynDgGnL/j6a13cfuH/g/ZjSuZlSo6hMJobxUeacLfc0GVzNK1YQ+Pf+4bPjaX1fDbVRJhicCD04QKzjqDzpOWYgUEh9gxSuSueZm/AoGDCorSUMLaG27yMSYzTvy9eZH8F4FPTEkNvXfdT7JnP7aRTqhQFFnZvjLAM50sLCGmTWPqhefn8PdB/pIbn0oRBnmM2cD2HWz46GcJ7niYxaVuZJoQGG8PCeeVmiA/Ro6j2FftwtIS+kw4FyNsjHD57+ORkwgbUU0lCwc19//DP8P1P4NkEEohVraSgaSULFmy5Atnn3321+M4roJPaGuu48AAXHs9K/7xixx3IKNsFKkRIMd+v8ZiNIUCJ0cwkiMV7B7cyfdlkP5tIzQezMmMqdUdumu1eK+U9UehMAkHQUNTOZAw+NRWKlGMwZEkCdVKZeKLHgG5EITJmJ5KHvvpDbBvFwQOJy0Na/OcFD9JSS7vBzN4YjO3fvATbL/hLuYkAZNtgGukZBLSHAfDCL9Oz4liNQ55mFqfM2GRBMbRXTfc+dmvwqNPgGlgpGumJkpgypRpr7nvvvsunTln9sfJ5xTgo50YIEtx//Y9tl97HV261UikXq+PQhYs9pek9SzD/J1TuTOkWIOWUD+kmfkjVw6dtDipyZRGSw0iB90pDAJpm25wz9UisLHHakgaUG+Qfu9H3PelbzLzQB2ZgShyIH6T5CSTXInOTbu474vfoHHltdDfD9kQCNHK07CGWXPmvuHAgX3/yzqfw5Gk9VGXGyXQG42GL6uJIrZt27ZVOhreSvZJK4gAuich4g6sDJ+1wJNYenDc942vwZU/gcF+aAwSaJ2DhhQTb//ZslCNdRBJMgNUpsCCBZzxsY9x2lvf4IEGJgIycA6SBmLvPqq79zOfkLCeEMkgT4CQz3m7v8OhkrFMq2nSx9ex64qrYHBoGIyixPpWoFKCy1j2mldSD4Lc4pajYB5HkfAWlLJ44eQkoVB0Oziw+mFY/WgbWMQIcox2cRkNu3tZf+Nt9MiYkjr4C1Og/YElk9BXipl99lmw+NgcoCH3DEkDIkXmoQQG+tj8hS/Sd+8q5moJA31Uy7FXTprW7PgempFWjsg1fdluXTjZjBePFJhFGbVtGMpGUO0bYMr+A/zqC1+EtesgbSBxhGHYBIjo7Ow8adq0aRdnWdaqgBDAYA0efYxVX/4KM5OEuL8foS3aCuw43G/kezfSAiwUlYO7OCcmJw1GGmqBYigIGAgjBsKI/iiAcKwclUOjkUpI00lqDGUREA9mTJElZGpoNBpUqxXS7NCrJQ6FnIB6klAplZmqSvQ/s5H0zrvA6txFSu4n9EdToEch7Osj3LqHSXUoDaSIhgehaVeg5LA99nxQy7Xsm1RJIiuoJhmdO/ey80dXQNYgooHMwZwNECrKL7rwwn/rKJeOs8ZibdtbMpjAT6/hvu98lzkN7+0ql8uAR5wbyQ+HeYxoQe06/DoIWntxIg/ZhDPMPSqZhHoIg7FkIAoYjAKGwoBaELT4RD4SV8iOHATGfP/H3Pf1bzO9r8ak1BBai9YO3EF45PNEQkiU8LwvTDVTU8u0/UPc+41vUbvmBo9c2RgEHDptUCSfT5065f/edttti8e77qjZaK2J45hGo8G+vb3bTJbuzGoNb5kL4S2mUpXF553HLm3ykpYjIycsDZMSOs2sWsqjX/oGBz7+/+CBR2BowDNxZcCl/qc03mLANf8zyuM9N3vLqhJ0dHDiu94Ox85m4hRr70MUKsG6lMwmOZib9YChJnfXStk8Ri1gDrpQHIdL41laBVMuY5mWWTZdeT1s3guqlNtK4N0yxj/Fripc+CIqixYwoCSyXCbRekJtWDoInB0mBIVwxNYwNUno/8UvPWh3PW3G8JqyfXgeih+HsXDTHZR27qGUaJJ6A+zEUJMeNtLPZwhLb0eZ6S9/KVSrWBmCCHPXmQO8UCcZpP/zX2TbdT9jmrJoO4SVuvX8kBglyAJBGkrSUKID6V8K63GUlQxQYURqLU74rOrICSqBoiQEEk0UC8A23fYeudP5A0mY1xCXrKQzlvS4hKl79vHYP30edu/xSUkj+m+DR2MsMJ1JBqE+xIYvfY3q1u0Etobo9KVQQNtaaZwbnuDoBCB9BYUhb/Th8PvWWJT2LmqVZ+NaZ3OAj0Mn4xx1JTjQVWFnT5XtnR1sKsfsn9wNXRX/bIpn2f5MrWuW9ox3iLx5yrC/CYm0jo4wRiYaZaEUxWSZJstGJ3gWWcZHckgEcRCQJXWc0UwOQtbfdAv07oMsoxpEhIgcQtkrhxZ8nkiqCVNNqH1ILg4CyPNoCo+mchAiCZG5cfDcC48ClEzhk78CoAPHDGt4+pZbcHf/CrIGIb60r2AH1VI8wwEpponPz0AGjz7Bg9/8FybX68Q6JQBMpgmlTy1ThSASOU/MUQRl7o43zqGtJZUOqyQRiooIKDvV9JY110JOcLTBKyMlVkmSSLIvlOwsh2ypROzs6WBrtcy2IEfWQ2JyK0M6IHE+J+XmW1jx9W8zq3eIHq1ROiEGYimwEyUNPwdUzHXkUYxR5aGkioCpxjJlay8Pffv7sPKBHKo5IwhCXB76rHR0TDr99NM/KYQjCEL++I//GKAJEzvKfHXO0Wg0CMOQE088sd9q/Uips3MhAl8WrkCUy8gXnMOeK35IFTlhnHTCyTpA+AlFxhAby9Zf3swza9Zw3JvfSNfrXgPVHr9jSxFIlyNWeX1ZQM7Mhmv5Vij8qwQBYrTW0iQLIkWgCUhRTgz79vPpKjs0soQCRJrRtWs//Op+WL7UoxM57QWNAMhfgFiy6GUvZfUTTxE4TSmO0W7iukWVPzoji4QWh3AZncbRt249XQcGYVK3d12Vo/w77XpSvuqZgaGUXffeS0c9pSxDoljmeaIT7A+hCKIIYxL26wbx8afAC84GFSKarbyL+FqOv37D9az56VUsVJIoS3N3vb9H4cBo5jPkZ4d5ToEVeE1fampI+qMYGZeIhECYjGygn+44pKxk8yUacx+IwtKwIDKs0EQCpinJ9o0b6b3iKqa8+w98OEiNswOtBa1Jb7yZPasf5oQoppE0/IhFDv07ToZwa0yOzPqSlzAMaKQJsug6Zg3ojGroLawsy5rdvA6VrLUc+4pXw0te7DO6XRnCGGwDli4lT2dvCoriPTwYTVh/bX3y2ZFxlSMjiSUymt1PrYdtO2DyTA/04rMNmxaSgzwZUxNa3bQ8XbEfCgtdTDzH52zcuaes8DYXPqkKhsk248mbbmTpuacjJnVSFj49sRgnOMoq9JNME8gydn/vuwSbt9A14gmMZcyMHowkjkIGsgaZcpRKIUZbInzpnU5ywK7CeJhofVyRs5LlowhJZcCp73sPLD3OvztIMAFUKzB7Dh6BwuEwSJFDbT+zkTs/90Wm7R9kloywWYaVPmZuBXl55W+KrH/PySPHFmbJiB3b93LHP32WCz/3aVi0GKKiW6DDGcvkyZPesnr1A19avvzk+y+44AK+8Y1vNFvABu0uFOccnZ2dpGmKMYY4jtmza/fN8zo7L4NmYy9EGMJZp1A5eQmNNc9QPXzI2WYiQwsAA0Jr6BGw7+mneeILX6Tr5zey9M3vgIteDN3eQgu7KmTQ7CLlACRoATKvVpbNqFD+4UQkNMqlRMYSGUEmvXBweRvF55pkDofZbO96EJePh4jOCPpTtt55N3N//02ISRUQCuMycCCF9UxGSrj4ItS1P2PwmadRaQMVDY8Bt1to3pIoBKFES7BOI4QgwrHpmXXMe+AheOVLfemgsdhcOEnZCntIa331w0MPs2H1gyx02ldEiJyzTcS8BSQmwWCwHTFnvPE10FkFlWfR5lUPyMADmTy9lcf+4/t0mxSXNRDNxEzZjFtCC3oXYfIcDC/kGhJMZ4ntZojK8ctYePZLmXzMYkRnAAf2UXv6Sdbdegd2zwGfEVskXAkzfODOCwFlDUb4BCgjvaIV1Go8duVVXPiSS2BpNUcia864OV60hYE6D195JapRJ1OSAIFyuUCTlhF3HfbsrIAkSyFUNIShltaxpRBZVl5BG9L0NARplnoP0niKxQQkbEDflu10n32GL6mhAtVJIHIc+SNU5tuVpJGCz0hIZeuz58ttXWBDFO+5cpLaru2wfh0ctxwq+XOTYHPAEgX5I7Q4YZrntsMHG0mz/WpojjgqcVAqPAFFc6vW/S1WZETCsuvhR2BXL/RMBREMj5IJgxQBoL3p/Pgqdqy+l6kYhIxJrR2JKDuacv5VgDT16wZ1YakFln6REJRDdKMOWUYkBR0yaMJ4F2s2zuzwoDq50eIMWpQwBwZQy071PEJKqHR63qDyckoMuASowFCNHd/+DlN372WyirDGkslcSRZpLhkER1p2/WzJCgMyy3sLhD7EJwTd2mI2bmXbv36XOR/7GJTK3uNlLEIFBEIEs2bP/gtjzNve/e53s3DhQjZu3EgQBKPVk3q93gS/SJKEDbv2/GLOvLmDMo47PAIZviFLOebk117GI099ic7UM1CVb5LmgA9hoUQukb3SaOgIfCtUZQX7HlvLHWv/jsk/u4GT3vZ6uOh8RC314FEiANGKm/iqdevtdseo7Nuxb16cGTRdZS7vhWzzJhjShqMYTvHSejQgOSrxqD3hKsLlc/QvfuFW9IUwh/CyS++anRKFbHn6Kebu2QOVuQSxAhniSGn1Tg1g5kxmnXcue7dtJpQ+U7cY01jks05biGlW+FtKq6kYy5Z772Te+WfD9G5cm/JRWL4S6YONjZQtt9xMKWnQEYYkSYLOM/DHm6QTkBhNvd6ASkxp1kyi817gLcFE+4QwAU00ssRS+/cfINdtpNzwTUhkIc+LpiVtz0Dk4RKcB4nRQUB/KOmrlDjrne+EN70dqjPw2koCZFTqF3Hqm94K37+C+39yNVOFpmT8yy+Eb5EobIGqYps9A6wsUguhaizZ3v3UrvwplY/8JSLyMLwO14ajnSF0CrffRfr4eubFJXR9iLjwuIxjnxprUDJEWIlRkgNRRKMSM2PpsSw7dTksnOvDLxLoGyK9dzU7HlnD/o1bmFKpEiZ1Qms4FCYmrSSwsOmpjSy95WbC110OURVC0GmOo/8cuJGH1e3myVNGFEmbw/euoPWUbS5EHIzKrC6udShUvK+Bs1StgU2boX8o9wrm1yq+C/kNi3emNf/CwLDCjrp3wSdMrnR6ntG6atNNb/04VN6J8WA8dPiuLxQUg5M+B8ls3gqPPQELFiHyWHjboHIGLGHwAJtvvhGxfy8VqUi0L49Vw2Y/xv2bc5IMhQGDHSXmnX4Sx56+DKZNhrgMCNiylYHH1/D4vffTlVim1H18fiKSzqc/W4DcZX/7L2/i4te+xnsNO7pAW4ySTb0/JISs7t/nu2/jqVtvZYmGCIO2ruU9mfDOB6ex+Kkd9S4cXNn1SmWrgiKWEKcZk4YMz9x0K3Pe8GaPKhcEOS8DhGLuvAWXrbj33qVa6ycmT57Mhg0bAAiGtd4QvpuQNQYZhjSM4ezzz1vfX6/d1BXFrwvzMDWBgI4ewsvfQMcNv2DXipXMIPZZh7mLtahftbRBSg2bfL4wzX/70gajHYEIsQa6EcS6zuCq+3hg7eNM/enJLHjn78Fpp0BPB9gEI0sYDbESIJS3GMSYtxxjNQOgRCZKJCogcYZMGUzukpFWggtw2vdEFgEQBmhp6DcJ/WkKlSo2iHCB78hkjcVqQ+QEnVLQ2cgIstR7AqQklKW8xa5P9FCiVbIziqTACOcxx3UdO7AHHn8UFs725XahQBS705a9VtUhmfbql7PlmmsZ6K8TlVt1q9L5YEX+K04Wr4tEiILFy2bCW0Vrdq24hXnbL4PKCdiujqZ8lu0sLklh9w523XUHUyXYLANrUFLh63jHszMVFkPU3c0eqzn5kldAdTI4gZWBT/B21kMzBgGse5ptd97P5AGLdCVCVQanscIgrG/UUIQCvBJlUMK7pG0QUi9F7CpHnPu//gwuv8xrvkGYC6US2Mj/LeqAv/orpkUxO674IT0DNSZX4rxUxmvRynrN3uUARsoFedtIRyw0kzE8dcetnPy2N8KCuR6SM1D5+RZnU3AJvTfdzOw0oyIEQRhhdeqdGs00CY+FUHi0rLUoJ4lVhd2JZejYRZz1rrciLn0hdOQdn5r7G6LL3siCrTsJ/uW7PPmza1lUkqhG0kRQc2J8gSGwdEQR+wf62XDTjSy59KUQdvoS8jj0P4vNlHtkhBAI6felaOuSN5KGgfq44cLXFsJcCt+yte0ioZVNoKLEaHQAMggIrfCdywCcxEkLbmKhLhwemS/nPYETdLuQvmc2012JvZMvT65qhpj8yws4dG68FM/Gd4B2SOmwWvvYc6mDRGek1lAX0BCgY0UWB9gwT7a0AqEdYWLpTC1xlhJkdd+iNYjGFeqFcHKOJg/x3kWwSEID06TEPvQo8mWX4iKfO5I/3OJJ+HrzAwfoXXk/k1AInaFE7K8/gQfGAQ2boVVAXyDpOf1kTvnjP4UTjoUKvrzPepAUrKDTpJy7dg2PfuWb9N79IHOkj32P94ykkwgbo/BKQ8loevbtZeiaq6j+9V/lOV0KZUTeac7hrEHoEIYO8MB//jvTZEhg6h4auDlvQYGcfzCRmxlNuVJGZxat8z59zvNvqSQuM4RRCYQiM66pkDa7+E14B5Vjm+T81GlMlqIkVIRgcmZ58j++xwknnggdldyIbS5+Zf7ceW/6xfU3/N966hEosywbbqELIVBSorOMQEoqlQraGnbu3vWVanfP67S2HtABfGywXOb4D/4Z9/zFX7Fvey8zhAJh8nL4XEs9SFlbs94SGA4m4F3JndZQFo70gGX/3fdy79r1LLz4Rcx6x1th7jyCuOSTOjJoh//zE5r4YXnr3ddhG9FymRX3BxhI61hZQkeSLFBklZBgcg/TjzuGJccuglnzfPu+7k7fqayvz19401Z2PvIoex54ELV3LzPiElJrrG5ZXxMYr8PXxzmE8LWtjT07KWmTRxTawgoC74IqR3DSCUw/5SRq961CYQiKWN8YSzNWtrQnRWQ0buc23O23IBYvGT/aJCXceCN221ZiUcYa52skDxqyMDhjGdAZYv4cuPTlEMWeZ6r8JZXKYw0M1em96WayHbuoZEAQ5gV3bXNyrSzb9vnEcUxqBXvTOlNf/AJ/n3InlMOWWlLUvwq8xyGIOOY972Jo9QrkmrpPZjFmVB1460nik9EQKOtLCnv37GHPbbcx7d3vhFDlMEkWrPHZ+nv3MLBpE9VMo5REysIyze0iUciTtvdCKkQU0TuUImbO4+z/7/+Ds5bDtBiTJaiw6JOeH9bA4mOZ8zd/RbUjZttPfkSHLDwNB9d6tdZ0xiF71z3tqx4umIGNyDfOc+sLd23vniT3jMrAJ0RZQ8NZHBKR5wFk1mFtSqdQPmlLGNxhJL4Ulnl76VXkBLZeZxhGfSHQ26crbZN5t3sYPFiUxAYhplJmmw3oV5LO6dOYseQYphwzFyZ3waxp0NMFUQBDDTiQwECN5NEn2LRyJfG+fbhanVCoCUMOPodIosYwCiS+M2W6v4+S9ribw0jkHMhaWPME2fYduffGC5iR7HQsUmHIXmHoPPNk5n/kr2DREuiuAg2frNxeaKxTOPVkTvrU37Lzw59g78pH6ZJBM+w3enKyKROktUTGMtUZtj/wMItrDQjrEJRB+Ni5T4oOwNTh1lvQT62nrHXOE8wIy7z418TyqRDmVgqyQJFIiVaKFIs1IKIIaRzSamIhKUtB4CRS2IOuXXH/1vP1GA6FF6iiLVtWPMAJd6+Ai14IHVEeKvHf7urqet2rXv2qf/jhFVdkRcvzYGQZgs2Rpqy1zUD7zJkzbx2qDdzQUe14pXG5YgTeNbp4GS/43x9l1cc/SdJIUMZ5SyYP9ks7cQxs3I9EUeTvFzx2gqna0HOgn94rr6P/vhUc/7a3whve4GEOnYGg1Cah2oLzE9wDoZEiReIbSUjnXafCSjKp0FXFfiXpWLaYueeeSfeFF8LMOTB5ip9/Zj1mfOHadw7CEIRiZn2QmQ/ez+Of/xx71z7NZC2am0rlDPdgVROF0CjqdHfv3MV861rx6eZcil8VlEPmvvV1/OrB1SzKoGJMk2m115G2u43GciEpC10q4PG772H5238fOqp5VmQxVwu6Dn19rL/1V3TZwGuKIkUJiRMTN+4RThEJSeokM190LiycCaEA45BB23N0Fg70s33lg5SsRYQCW4RFDqI0aKMJRUjgAkwUcdwrL4GeMsSyWcYDvo5XCtBJQlzOG5nPnMKyC87nwbVr6USg2kF28gcX5IAUOB+OApotxqvGsWPlg0x7y5t9P+vQ+RIhhy9U3rqL3du2c0yBBe0cUikwhYU72pINAkV/1mBvIDjpZefB2SfA5BiwPrfF5aZBsSxC+iYvMfS8/U3svOWXuF0aLYxPKD6IsNDGEMuI+tbduNVrEOde4q37Yt8eoUwfyRPa51kA2FgEDSHYZxIapRJiyiTKi+ZTnTodkKR7djO4eQNJ7z7SgQaToxJWDnOOc3AbTA6bg5KSWr3OpBEVBWPNs6gBz1SrCkInlkEMtZ6YdOYMZp73YpaefZ6HBu4s+bh8kbmG83zLSr8ftCB+m2PJ1h00vvbvrLn2WqajqYzr4eLgDATYs2cP8+wE15AK++RTmL4hrAo8VvkhpEZYAU4Ksu4qi97zFlg6z+eLGAuqlHeVAGxbo6o4gJmTmflH7+C+Jz+GSjSd4+RgWdFS0AXeAdkTldi6az88/hS84Dy/hMoH40Wxns6w5bZfER+oET/LNCiTWpxU1LKE/lAwMKmLnhOXMOvE5XRMm0EcxfRt3cbOxx5h/7r12AM1qpn2lRHu4ArRRBRYqAwk7Lj5NmadfxZQzj1h/vNqd9eyW267bdn+/fsfap4z8iI6y1BBgDGGHTt2nLJr167eCy+8cOvqRx7++GmnnfaSKCyVXJ7MoISC7ilw7gWc+ad/xoNf/xrdjYSyTnwe+iFYAIdKylkiCyrN6JLQ+8Ra7v/nzzD1sYdZ9J73wNIlkCXeNSsKK7h4cyZaVu3dcyJPhEPRUAH7Yx8TWnjWmRz30ovg3DOhu9uj0okAK+NWNjEej8g5n1jZXFRZglOXsuxjf83av/gotS07iaXySo71glFYeUixvkIg9+3b33Jx+pv4H7mlYbCESQPOPYue05aTPvCo/5aUmIle6jFIAhUn2PT0Rp8oVKlCpctb5ELiYV7r8NgjHHh6E1NSg4v8urumsB3/nk6AKldpRIJpr3wZVAKQDulfTW+pYwkc8MRa+p5YS490IOWEikL7+K22NNIEESvKk6fC4iW+i50D2zRjWxTFZeppnSiKUDZFnHsO6Q9/SNpfpyRCHKapXFnXtM1bE4KmdRday9bH18CGzXDy8dgwb3Fpc59674HcfeeFsHW+1KvdkhqZEWwF1IVFzpxK9OqLoeKlcp7x0QrkNl894d390sLsGZxwwYvYdOXVxOA7YR1EIDsp0ElKT1Bi39NbmFIwyDZL4fmiVEl2hwHJ5GksffELqVz+KjhuEV4I52/ZhvUM/fCHbLnhVtIMhMso+m0daZZ5kiQt7WxMalmORlq0smghSWXAgBB0LVnIsle+BF76Mpi5COISTYCUAJzMG+ogGTI1IiEJi7a5ysL8BZT++sN0b9pM/dEHqBzia9v+OKRrvXn9/f1eyI7C1JB490PA0J4+qjJEDmsu5Wk8/mQF7K/1Mfm0s+DUU5rXryUJlUqMsQnKTxiXx9kDAq/cnrqcjqXHkTy8ls4JNmHhvi6y+XU9Qcsh2LrLN8PpDFoeEocHZNnwNNsffowZ2rawZo6QMqNpSMHuUsiiF1/I5Hf/Acyb7b0sDsg03VGZ7j17YO1TbPiXb7NnzRomNRpU8nDNkVJoLN1Ytj24mln9A75MNChTGNVAuGjRohfed999DxV5b6O6XUdRgLG+QcLxxx//8mnTpn1w69at7z9p+fKf7d279/83Z/bcvysWycfQBHT2wNvfwelhyD2f/WdmaSinCiUDhFC4I5iVzTVn2VQVvcvEZXUQlp5IUtKa3htuYvUjT3LaO94Kb349BAaDwlhNJMsT3oPCRpMCgoj+wSGoltkThgTnnswZb3sjnH2uz+aV+SEUSN8QwlmaEMI+QWi4XSCVgs4OWHgMc05exs4dexBOI5B5lr4X6uMxnva9WFjpYpQMGu5EqruMsFoFFAsuupD1qx9iUuSL/Jy2w8pP7AT3hjzJcaBOjGDLL29i3kmn+PK0SJKbyKAdz1x7A+V6g2qOauakb1UjJ7IsgExK9gvL7HNfAGec7q1zl4tzK3ASAnxf6WTFCqr1IaJmzM0evLTJKaQ1XosWjp7582H6LHAKLSUO4bvrtV3FYYiiPHgqJcyZiZo2jaRvK0kjQfqsNawwBM63Fm2uocyFsyisd0mwbx9seApOPx4QhDK30LKMwT17fM9oyC0Ymv/2jWg8vnR70xVrLFo4OuZMh7kzmhvCZ/TnUrbNQG8qHFHeW332HA4Yw0whiayPLoy7fIULXCkqVjDUu58pxg2/+Bg4D02XOYfOz4qMZ2MM2jmkkuxRhoFjF3LeB/8kL2VUEMR5LDHwitPyxVQ/8pdM7u2n956VTEahjHfNGzl245ODjsNYLwCLTKs2Mjp30gYKrVNAkyhLbyAoL13EqW9+A7z8pR58JspzNJqZ1KppYKU+KEmkSkjAOF/DbtKUIFLQHXDs617Oo0+snkgn9mNuG2a7R6+Ye5ZlvunPGGTzc/r39BKaHJ/Ftc51E/AIJyArKeYunOuROsMKzgZEYYTGIGSIs62Xwwpv+ARhGUqa7rlzGXhkvc9MPwSSDsphRDZYp2/3XrqbCdCibTKSoTWPYXftokeFvlfIRGHfCQS+EyBDyf4QTnzXH1D5w/fmOTYBxL7hDi7y3TmnzYHuGRyzfDmbPvNP7L7yKha2PfdDpXasghCo6pR9e3rhkUdhxjQQFpF34AMod1TPK5VKX86yDCnl2KljzjX7hyfd3d2zFy5ceN3g4NDXvvrVr357aGjoWoDM5KlOkfAuZiXhtW/gBR/7GAOzZ7Mdy1CeEnUkNNod3CrLAC9sStowy1oqmzdw5+c+z77Pfh7666ihISIZYczBQAO8pZng6JOwJw6pLZjHme//I079p0/7rkVdFV/nWCr59q3SZyy3k8j/MirLFkCVodpBx4lLaUiXx+otRhyalVnQsO+OtaS5oK+KmMwClU66L3815YVz6FeWBNsU4NK1hVgnGIN/gQImCcH2FffD7t3eKi8mbYAt29j78MPEWUYURrhiXofgs2sEkj2VgNmXXuJDy7I0/AVzOZ6NhX0bnqZH5a02OXRGLYo+xU4QlKoQVvGp8SJPRLF5ajCj11UAXV2E3VMIgjLVsOwVMYpxtGKoYx3KeRREvW4tKJVjdeX3cQ6rzfD2i87hXOHVaFmZqm0pbQ5Uo8lDO8X12nLH2juQt05VQASlTlLtCGQwDPZ2zLUby0or4gkj902+oQ5HgDrRwhgv7uecQ1Zi+lyGnjbJC/Pzzvb5KaUI4ggXR+g4QJciknIJpkxm+msvZzAKSNsH0MQEH5va8cibp7gxJtDOd2Ly8J7ERDG7wpCdk7s56ffewomf/Sd402thylSSjg4acSXHIWCUs7Bg9f5PsjkeFcb+3alKWLYIEzx3Xs6xyPd/dwSWZjmZOvir6891ILSmHFd88lvumywqlwDIoY+Fg8AWCxCDquBE3MKPH4PGKlnUxlAqRzQaNZ90J1q9GrBAptn/1FNUnENkjWeFBZBJyRadsujll1B55x/49t6VKpRiMlmiQYgWysuFOPafdVRZ8OG/ZN6Lz6dXQDYOdPOhkMBSDSRhY4h9Tz1DnsXc3EZ+D4mlr371qyMpJUEQIAs+0DxGbHLnHFEUMX369D/++Mc+fstDDz14zf7+vkeUEgQSkkxD7C1currgVa/k5M98mvjcU9hdsqRHOh8xHMHMJ3+EaBGiRRktyhgRUqv30+FSFqYJ6773fXZ//O9hxyAkAqVKHNQvGMQMiYAdQcCcyy7lhC/+P/jAO2HKLIh78L3I216q9ssN43gMY6otVpJ3pap20J/UfW2xKLCeD83dDm2JOyMsdFu4/xxe+lnfbxkEVKocf/EF7BIJg/je9UXTFpuPwZfTyHGFUqlUwtWH0Bs3w3335667fB2sRf/iRsSOXVTLMf31IYrZ+/hR+04arak2ArDHzoWLzoOghHT5WucJJcqByhz097NryzOUAocktzTyBikTApQI0Er4CgRj6a50kiNW5wqYa+vGUoQI/GGwfrilEmFnBzazRFYRGp89XNQXe6+1HPNQDsrWsGPtWmh4KFiPsAFYg3MGqVoIg25EcpBwdhiDLZ5/iKR+oB8O9Lc0s8Iln59bqJ3ei5E/CRdhtu0lsiHK4oFnDoEsHownKYqSDyWb8xDJiBxDIlc040pEIg2NzpCTX3spXPAC6Oz2ezyseu+PaM6IMKgAMZxwPI2uMvU8g22sBMnnjJSCLKNPBdRPOoVzP/M5oj/9E5g1A8JOkjRFOkncrOIWNJE2899DJwlz2ON2iF6XZ26jpDciDgXQ5VnQSHZWCHaJyduSjq+wBhY6rCTZeQAG/XeF9aCegS32HjhpEUikkQRagg5hwFDb0UdwGAVkVtDknWmxtK5tDtZCqtn26COUAvKwnxl3/AfjvUNhQH3xMXT+2R/7UtA4BpWHV5EERHjEjqCloVU7oKeb6e//I/bPmcFQmwfrSGDERWBRTrP7ibXQ0H6OzofXJJLu7u75q1atmiSlzNEPD7aIxmJzF3xHpbLsBS84/yv79+3rtdb2AwQ5VroL8tXt6YFly1j2j//A8t97G1vLMb2lqIkxDuRQmv7wUxyuSReMchjedlG2I4p6X4kQikqlQuQc3UYzzzg2XfdzNv3dp2H/AV/uNJbG3SQJjZR6qcIL/ujdzPzfH4YTjoNStW04eXMQgpY5JvOXM29t2BIquXVlfH9u0hQGarD3ADy1gc6o1Ez4OVxqxw8vaMzLFH8UgFQEL3ohydRJiCikgJVoV5IORmlq6ArKTEkNu2/7lc9UtRpMCjrl0Ztvocc4TK2GUgJvPbdi56r9GZKHGVBYoagFASe84qU+LBHGI56V87+7DPb3kvUdQBlz2HjQTZhP4yhL4aEgD6TI/f2IvkGPjz8wCP3+EH2DqIFBosFB6K/BUEJPQ0O9TpamOSSvaovryWF7dSRGeagNtd593nNEXveZPxsXSIRwCOmhXUWO7SyhWYPtJ9FSNpRwhFqTbN8JW7b6Lnx2uELS9Dy44neD1A7SlGceeYyKCsiyDGs1owBzRlDgfCtQI6UHSwlEyx31HAj1kRZYlmX01+sEM6bCqy713hUkRB0447A0m5gCbd7ozjKl6VPJVO7pwCtMB0+Ko5mtPia1Ky9Nv7aAUpn5L72YSz77WTjzTOjugkoZcMRRJ1KWml8fcwSFRZkr48I6VGaQtZrnG/1D8OAjREcI3NMcvhvxrhSCbBgjEWgpMSLAjfCETlj2B3QSs2vtU9B7AIYsBZJqUTrXsnKMP5yBRgN29TL41EaiCbZf+3MpAHNkKBhMG4hqCdAj5mFh+w7cvn6UcGN6YA5G7UK3oSQnv+61MGs25MmCepgi4LVbi/CNeBQ0dOJ7iBy/hO7TTyPJy0jVESpmWZZRDiW1nTuhVmtNKFfikyTp2Lp162ytNWmaElSr1WEXcFYQheQoVc73HAbAeGg9qcoLFh5zER7nKv8kD2vhQTNEuQtml4j//MOcs/xMHv+vHzO4Zg2d/UN0W4PCoLVBRRHKlZplY+RuyjFd1zkpZ5tCS1gL2qtqxnhHwRwF2+65nepX/o6pf/khKM9H59XXo4SAExBVOfYVr/LBMYHPanNJm7suZJSF7oA82TvP3PIVTwovhGqDHmVsXz/86i72/+znbHtwNXM6ukgHB4gK4eyGX3o8kg4iA5EePodiKE2rKY+fFmsmowiWLGPRCy9h8Mc/JdIGK3z3KJm7b4a5FUbfGWklsYUpMmX3Y48yfcPTsGwJZEO4O+/G7t5J2aReIw8kgrQ5sMJdrPAWHkiEgahSZW+tD9EziSmXvBLK3TRHnaPn+RighVjBzh2IekrJhNg8+/hwYHltqqnGAfvuu5uev/0I9SBooYPlmOrN8knXyviIrKVkJcGja5gch6RJ4i2KpvtZjg4htzEEgSWQMNBIfOe6abL4ABT0HH88dZ14wSMdCoc0tEqpAKta4CkFxRZmp4YNX/0Gx5xyqrfkopCW7UBrjNb6Y08f3HgrfevWMEM6sBqrfCIe2DFxxmXuUlVSYizMnDkXpLeQVNMD8yxcis57ErwC5Jl/gCSII4ZkDKUOUCEyKBDbVHP5ivWVeUSaPOPcCdDWoMiRHke886OqeibcR55rFlgNkL8zQQBnn8HZZ53R5lL375HFUqQoFrob1qIzTZDjbZNZ35ZYSq8cOOeTS6WDwUG47372/fwmdt5/P93JkWONy9ziLipqityKQpeQzT8HzFm4hDXh3dR1inWymYR/sOuHKmT/7j3Uf/g9yn/8PuiuorWBoIpFEaD9PpGpT1ouKRgYpP5v36K7dy+lkdUEY5AT5NULhrRepzy5k8qkjtFKpbWwt5+stx9lZW4gHN7+lEKgrUFKhY1iOk47G8IKKOUrL0RbuBIv82Rh5wGlUuSNnXKVY89/IY/ceCfWGgLla+4P10pXQhEJRdbXCwd6YcqkNsYPYRhGixcvnjpnzhyklOOXFjcnWLjyxPC1w3fbbJIjF+xC4AIFQRkZxnDppSxbtozsmmt55KfX0ti9k+kyQqSNPElL0yqYL+Ihrfhk+xha1IpFuqbg9Y0Oy1ozXVie/Nl1XHD6afDqt/okk2KQMFzjlhKI8uz4YjI2T3hQradXCE3nXZXC5LHP1IHWCKEB7TGRH1jFzhWr2HHPCuQTTzLVCWaFIaF2vpeKGBEXPYRn4NzYcfrm+WPtEykhLjPnkkt49NqfEwpI6inkeMrKjXPe8Isg8IJN9fXTf+vtdB27ADLNkzfdTFAfJMw9OFI4LKaVoJWjdxWlhzjv+tbWMRTHLL7gRTB7HiC8UBu2x3JFwxgw2rc71P6ZuHz8joMLds+YNIoAs7+XDb+4Ptf8vQAb1T2qzUugLEyudBKmHgPdhBJnLRM5tkZmBwcSn8iZpfjMaa9AoCTMmo6sVqg3+qhGETZJCJxoglK4tqM9gTFyjhDYvnY92z7/deZ84L0woxtCkCb1e9ngLT1rPPDP7Xdyz9e/xuSsQeC0F3yAammDo0hgMFmCi0rUEUydM4/nAhlu2D1GWJCF18NDdAYgVS6ERnsEpAPRFiaxgmYSVtF0Qx2CxtxukY6mkQqv9DIir/cv0AHdiO+1G8FS5Mq1wD8LU3ifnE9UTBLYvIn6PXez6c676F+zFnbsZE61g+gIvHntNMxCFy1BlFeGe9tESpi/gD6h6JCS2CgOJsz9pCVDtRqxEDz0459wbmcF8QdvJejo8D29lfdYkCTek4SDfQPs/so32Xz99cxOLEE4kQugpZAVfMTFEYNZxuRjj4FqCaxsFf04oF4n1BZJgBNmlAfoYGRzJE/nHKVqJ0yf6XNupGdOhRJUPOnW7mqLuQoJYUx4zCJkXMIN1jFH4GmRLjc4rEVmGtIGjAzLCcHixYt7sixDa00wCqJQDL+gyJ+6km0MZowJjflYlIQ4guOXEP75n3DGyy9h67/8G5vvvpMeo6mKDCMShEwJnCCwMmcvqonJfTATtrBminGHQhLUUyajeOaaG1l09iUwaw4I4UOl7SpM8bCLdPViU7RByhaTsxJ8mVNGEBjQBuo1/0Lu2QuPPMbggw+zY9XD1LdtR9ZrdGpLSTrQlsyB017w4Xwy0KFstmaf6IN/dTQpoCTh9FPpWbqUbStWM0l4LGVkC/d8XHej826uInmqqiXb7l5J16teC3t2sf++R5iWkQtzf5Em83A+bUznwtzPwRKpmH4DZs50Ol7+UhAud40Wtxyx9trgtMYY48vu8gd4qO+pReVMTeMcdJTDVtcpV3Q+au2x9ji2dKCzBAlonSKVQgoxItY9/pPxzXM0aTYI2QAEWS5Fc3fO7NnMO+V0hm6/i4pUPuNdiCbGeIt9DBd6xfOaoiXbv3816eq1HPP6l8M5p0BX7HEZLL7H+sOP0nfb7Wy6exVztEUYj1rot4DEMT5wiUNBFDOkJPRMglNOBmi1NXbykDxMh0d21JIeSZVM81oHGeDBwk7FVMf/Wp5R3xY29MluuWeEfI1MAkk9F+Z4F/WmzaT33cuGlQ8wtHEzsm+AirbMlI6wpwtpDNI8S4l+EDLO+k6Vp5xCPHsebHmK2Bi0UE2ld+wR+LybuKOTQClm1SwPfu27HLtmHT0XvxBOXOLhWW2eD7B7O6xbx44f/ZQDj61lRmIph0FTkR+PrLQ5LK7F5qWBk+fPh3nzGPbekit3jQbWWKQQGGtRB0lKG7n3rTVNb1VHdxdMneJvU1zGkXe3K07Ir5MPxeBQUvjSvHlzicslnHMerK1IYj0MEkLgrEPrzCtGyOFGpqfmcA7easa1fgq8JnTw16QYDWirPVSjlLB0GXP/zyeZe999bL/iv3hmxd10CEfVGjqyVmzZ0tKYDzVWWlgwxmhKQmFT2PXIEyxatQpePtmXj0jFcIz3ES65trk2tZZcsZTkvbizmnepbN7IvocfYceDqxh6aiPJhq1E/YPMDsrMQOGSjEqlyoBO0NI/FOu0T3oRrTE8720BpAJnmHf5ZTz65DNUBhPfJUoIHAbLxJqjFf6ZCyspadi/YROsfQYefojSzn2UREuYj0XtDNMBMlAMCkHlxKWw/AQOKpqdw1ifx2GNd+tbinaWh6f1SvCoTsXvrnA1t9TTkaBVURTRaDSa/zZjtPCciISVuNR6zw2aQkhnCMIgYPZ557B6xSq6MkMliNAu9d85BMCQOMs4lhLJug2s+cLXGahIbHeZyuROAJJ9vYj9++lKNNMyR2h9e8v2Zz6RUukZdjc7tUYtmOu7XCloxbzaJ0rTq3D4NMFcJ8yBGXGN56G3tWTiqrFi9xTTDh1ekOuGD0pab2ExsA96e0kfWs3OJ9aw8/En0Lv3UK3VCftrTDaOqs2TVgOJDnz/AesOzctwpHMLhPTeyDkzmX7WadR2bERnOm9ONfHb6QTEpZBGI6VqAKPZcO0NmDtvR0zvpnPWLLIsROmM2t5t1LdsYXYmmSECVFimUa8Rlg9dyGVSskcqFpx5FpQr3mINKsO/lCS+UkLKsSsWDkJF+ahzlmq1Qs78hwvQkXKC1j6w1qBkDPhcozAX4sZa1BG43IsxaW0gG1F66MA5wfr16yeP73I/hHIjl2WoMCTLf47/RQiU8AliIo8f9XTDJRcx+4yTmL1yBfd9/cuk6zdQMQElVaJhNCb3R7c3NSle+pEehQJNq5noZR2BEHRaxVBds/Huu1n4ogugUsFYRxBFw3Dk/UOSrXkLCY0alCpew2jUYKgGWzbhVqxky+rVDG7dwv7tW4mNJbDQaSxTjCMKykQOMmcQQchAmgIij9O2LM/2bTZh2dgRMce2a+O9HZSrcO55iGN+yv5H1hBLcjxxfYguVO/WCi3Evf3wgx+Sbt/BHBUjdQPP5UezPZfH64vRWAEDzlLv6OCUV17qkefU+Jn+SgAqhyIOAmDsWtqJR+4tAJNb6pGTGK1Ref9qsG0vfhH/bJFOG4jcN5Ok9TYgGE8jX9CRCmg5qFCmA5+LEYLMPDY7OabBJS9BXf8Lhh5bSwlIQ4l1llj7uLIbFvEcTkI6BBmRM0xvwOTEYQ4Mkm3ekbuRLUJmeXti/z4ZiU++s/neGGPMBVkhGaql1DrKnHDZK2FGj8/wlqI1ojam1jovt2zF6M9Gkxzx0+X/a3nMDqrUDxPkrRDckboPmpUkY+gto+9tfZWCtUiXC+8s84mWO3bA42tJH3qYdatXwdAgtv8AkU3pJiNwEOkQpUK/150vIgHAeEHeAkY5fJftIc9VSuiMmPt7r2f1rb9goNEgkBZhHcG4ZWV+PLWk5l3RDqoBxGFAlqakO3ehd/US2QoSyxTXIA5jImmRzqFFRpAL84mebyNrEIYBKogZVIK+yZOZcvll0FHFBSEi79MOeGHQlnwWxzG+kGP8fSBGrKtss+iFEP5ZCoHLXeYHayPrXJv/Ou9aKpVs4k0cLnko4byc1WiaQAbFpA1MnTp1KAxDnxR30Cu2aycOzFADFcegHWGReT7hDAOanbCED0uLMITSdHjpSzl3+Qns/8o32Hj1rUgnUVI2Y7CF63GiB27ahLMAX3fsvCyOjGXf2rUsrNehVCKohG0MxiewyMIHLwCrfd/vNIFaAls20bj7bjY/+CBbH36UylCdHm3p0JpJVvtkHhHgbCuZb6R8LJhDIbTMCO4gxKG53o+EZDGxMIIpk1jw4hex9sl1NOoNYiFGZ7uOdY3CayIAa+kOJDtWrEBkGaESrYUfj1wxCotRkISSyqJ58MLzfaz3ELI/S6UyvizDvyztt5RufKWoyJkprMaGCqiFATUX0LCOuFwizVLad/BYWr3JMdyFyP0CbfEwOWL8w56lkwQqZosNWC7LOSRmGwuRAcyYzslvej2r1n+aqG5R0uZhrsLinMg+tCA0gfM4F4U1a5xPEDPSezzbx9XCnz84ZVIyECjKixfS8YqX+rGoME/8ks9CZI6YRaGQj7Lwh4dDRlMrJDHq+xP0kn/OyDp/uBTZyHx8fKhGdvsdPHPvfex56FFKe/fSlRkmKUHkDMppcClO+n73BZ6+b7eaIywi89yCPB77PAnzptIi8dWcixcy7yUvZOvPbqCnoSk1u++MTwWvFuRlpg5ia3LFxCJtI88Jsc3qDXcYeRidnR0MDA3hQkd/GHDMJRfBKcuhFI6dy1Kp+KYppsD1PzyrqOiGCNDX10dnobAbg4jC4Za6tweg9YNIRTQRiRxkWeqVhCN8UYzxSVcqCDyozQiSCiZNmlRvNBpjt09tJ18q4P9dJHGpSgkSm2dWyIMrwsKXKTnlEEXGY+HXjEOYOZ1J/+dTTDruZB74l+8yq2+QqklJg6LsqYBfyCfQ9v4a0WpjGtoC/xnv6pcAlsEdO2Dndg/bmi+yxFsvDkgh7zRmkS71wvyZLdR+eBUb77qXA5s3U5WCOYFE5XXBgXWo/KXzPMXDkwpnUcYvabFuTUZKS/C0C6CDutwPytRHfD3/WcSIPXh4BJ2Kzpe9jODaq+nfsIluoQjG0pTGdFvapvCXaAyGIA6wxpcFyWFBptFjLW5jBOzUQ5xz6SUQRWgL6mA9uq2DagWllH9R28Y0/OfYFBReTyk5UAo44QPvhLkzyCzIqIIdoTmPLdDziLbwoKyu7Rwxwo0ybEmdxDjFsV09sOgEjAxxeY9nAf79iRS87lUcs+oettxwA9OwlF3QzJptu7K/b3v/7WL+wpfMmBxFz2CbzzHSLYHhhM2rDkz+3phWYuQYaFqNALZWBC/+6z/2UJdROfdVMEEq3e8SOc88UgO798HVP+Ohn15Htms3oTbMxFFJrW94knNagwARIm2cl0C2rmZkRvFUC2hhVZREHCEJkVcyuHZJVHw44mcEU9/7DtatfYzS+o2UB8cPL7VbttK2leI6j2vhyPmc0M0SNucCnycl8mRoLHLCMInFZCkigL3ZAMGxS1n4h38AJQ8MpfENXpueIoDODgIVoE0da+2EwDXjzk0IjLbs39vL3B27YNKUVkdM0Q771LJHWmJQevlmLWzZik71COCow3O7Z1kGcUQURlAuMxa/6+3trZfLZd+c5VAv3IQSNI7eRx5m26oHOfm1r4XuMjKSnjnJoM3FytgxtUKYizxZpNLpM8Xf/CbOiDtY+0//TMmmzQU4FJJtpW7tdYsSqDoLW7bAkuN9g3iRJxUMg9jToGtQr6GvuIb7/+37dGzdzxQLk/FC0bWVjyjhmoJaCI877jVqyf+/vf+Os+Qq7/zx9wlVdUOn6e7JmhmNZjTSaKRRzgkhIUDkjPEazILtn9f2YozBNvba4MX2114H1l5sYxwwIGARQSAkIVAOKGdpJE2SJufu6Xjvrapzzu+Pc+re2z3dPTMSQbD98LqM+nbfulWnTj35+XwQkkx64IO61jSk51qvmBRtj67+2lqzVuTh/CwROEWAa/DXO9NnvXWH449j8Wlr2bVzF9XcEAl52BJl0YXswBsTm1EuJWR5HjbzDOmswv6EtW4oiVkw1wPJRBE6LpHZFCmilpcjwJEHsJXgKJQ7QGqcFWEq7MgjlnYAiVRJWHUsXHQuUVwGp1Fx0roGQVB8RZQXeixy22ogIfxNcR6TH85J6V9tJUkU0fRjgwPjoUnx3nelRN+vfZDhgX0MPviYB/cI6+ecnKCGXVvkWUS1Rc9J8TLBaW3uSQr0ONkkzJCuYNWizZj78kcmvTHfW4m54FffD2tPCUuuKAaNp0u3//yJbPv/sA+KGf88h4aBW+/k6f+8msZzG+gZrdOFojY+TklHSJuhizSRdBgpME5P6O53Yc1TpWlIG0CnfJDSlVli43nLXoy4STq4cMQmGMFCR0QaFh/DBb/zER76478gG9tOn5JE7R3awtNKtwyxRQRetYnZH+kzDDJr6i+LbDb4+syamTC9MVWDdgMYkIrR+b2c/Ye/C/PmYI3B6RgdskQT9mC1gzQSZC7HOtsqdx6h+BGzEJTVauTrnkEfe6yfLW8LXtsS683r9a8cRARjw+x/7Gli55kn7Yuo5yMs1uZYoXGl2IPWTFojkzv72GOPDZx33nk+vT/DpbXmiG1bPTsbo3pwP8//8z+x/lc/ALf8AEaGfRenrSNokOPZy/zNdv7lbGgV1+ElwWlSo0GVoHcOvP4qFpx1KkOmQZpmYH3twJA3X1bY5ktgiawlKs7PSYTQCKFRCLRz2NERGBgALcMsdFh+58kRXDoGrgH79zLw13/HI5/+exbt2MeCzM9+uzDfKqRDCf8S2FAS8L/TcUQcx6S5JVWS/c4wUEkYOGYB4k1Xcfxf/wWLX/UKxkSOxSJs3nw5K3Cu9WptjInpRmGL0kAKogauDi71nqCduGUnbOHiMJEClzP3zW9hrKOLVCh00dcwaRMVL38cv84CD7eqpcJkjWAQpnpNOpwDk2VkxrI/T1n1tjfBovmeEpXc32MjSOvw2c98y/eOudD0JpXfG/2L6e6bS9aObCYMhwNFscI7VrlUfjzHwr6n1vnJBCJIKnhAARVqw5H/PhX7ly6BTDxFo0586SIOMI/FK9L+pWP/+0h7Mo6kDCqCcuR1sfZBmnYCjQpqUHioTyvguONY/jf/G7X6dHY3fL1emRolJMpFSCMQBmyYyGgaA+cwuInjdsLXXpWATOVkRRjoClAWD2CjraU+5vsCIhlhjCN3kgMmY09PmdM+8B7i9/4SdPb6nhI8cU7TxTi6bOaUUpz1BCTEdik41qd5gaCdi/2w3xeipOI1s/gnQBL0lQXIQeaQjsLwIPzH53ni4x/HPvQA8xo1qiJHuAaVskJGFpIIEykQGVDHhVe9PoIQjtTUcOWIERlxMOlmf/cC9FkXs+aP/pRVv/UhxqIpntGjkCa4ShN0aJJr0FyCCGQZOvvg9As46yN/zN5jVrLfOTIyci3IlMUK44/lNAVBjofwNWTKhFKPBdEI10xIt2dIUcPKFCMbreNggmH3AAxOGDIyjHaMK8sQivGFCzj7Dz8KZ66BKEGKBKwkd216rugFmNtPZeHCcA75tKBPE0DLJq9Z4BIoW8uGG6/zjJJBd090YP2aKtoBtDSMjsOeQTbf8H1KjXyCMT/qOfTIgZaUFy+Dji6/1wtH2oIQoiaE2L1t2zZ27dp1JF3uxcMbUgnGUNKaJVkKTz7J7R//Hyx9xSUc9+63w2mnIKwhjrRPpwuNkx5prbV4kxKJUvkbYYFqjKgk5FgqcQmZzzzSANPflGLOOiryI1JPYFYtlEeSe+SinX//T2y85pscayRdDZDWg93kOhi4tnJSoXSEUlghqDnDqNCMdJYYFIKFJ5/OqiuvgLPPhGXHgckZv/E71HNDRcoJAAVHJhInnFfINiN0erTSSW0HOmS52stIkYaTT+a4C89n77dvoNMYZFTU7aaStkSSm+hBt5M3zCRKCXRSZqA2hpjTS//FF3hjno2DjJFonIPnnq5z+62P8I53vY6+ebEfMGhYtJTIrh6SufOob9yIkhJ1hDOdvjEsIDUFA7b36fXMNeGXMtSpRTuVamvZPOGO8BpQqAmLm+OQyBZ5kAAf4QqcA5sbVOFIFPfA0qzBtQ5lQSuPMe80q//iLxj49N/zwk030MgMXcKiQoakNT8vm0Asxb07ZNx0wv2xzVJTa3EUUkg6OhKME4wbRz1OOKAFQ93dXPrbvwavv8p3E09pVKZI4b5ImXykl4K//WMVB0gBoyPgHNm13+bZf/4XFgyPkSQJzuaY9jGsQMWs8FlEK3y3c+Y8uNKoVAzqhGEsXSeeyCmveCWccw4cf5wvxXznWuqHXJaEGwAAcVRJREFUK0kdRmTzvgfPvn2xBU3jYIVP7UvjoKsDXnMZl3d28MM/+x80tm2mRzniLGtzBrxuOKR/pXCyJ+mUIkXfjttfAEgVPzsnICpRyzIaAoZ1TM/qkzjzox+C80/2z6uZuB4WwjMo/e87qlQXL2Lk8UdYFMUcls5j8umHGXQhJCWTs+Phh+HGG+CNVyG7eppJOYHHPfEFYdkK2TWQGdw3vkW8fSclkx91mr1d4sjP3XcuWQLVzlaOP+iUocHBA1mW7ZNSYow58pR7UxIP0al1RKcTHFs3DF5/M4/f8yDHX3YRlbe9Hk460X9xEmPx2OVKtCEWFQ+scH5sokhFZ+OMDhwgUhotNfZFsiVNlLCRrfToQbL1LsZCrcHer3yFddfcyLF5jBaWLIIoAMdp09pDjtD4YYPSMYZUCOqVhJ1VzaKLLuTCN7wezr7AG4o48QheQwNY5+guVYkbOZGDVBaNZjOrxsKTtCKgc6ciNAiUEESYIrKZdIyil6DY/jmgVQI9kvmXXsquG28myh3mKOrzL0byPEdoTVYtc/w558LJayCKfdoMUIGC9vvf+yG7doyyfdte5vQdg1b+4ZISKEfMW3MS6x56kNhxxAa9gE8t6l+VzLLv6c3ww8fh8guA2L9cWz5k8gbybx5ybB1634sOAsBnSiTg8HutUYeOUsstkjQjvaidRUHgo504geUL6f3UH1FdezwP/OcXWTw0TEe9BlI2669YSZR7wB8TtVOxzLQYWRO2FjzMpxUCjcMIxbgQbBeS7ksv5tLf+FU4fiV0VsBN3h/tCvVHaNTFj+pIP0pxNN08UUBAA5Tg1lu47zP/Rt/ICElSJs3SZqq2EIGn0FXO7xLrJEoqRJwwagUDpZjq2adx/mtfhbj0Uk8+Uyl7pEktOJjVgsF8cel26XxGSBpHE4ozSDFAKdv0oQXvsBRVp7NP4YJ/+T8884//wL4f3EBfvUZFlEPfRXGEqaVgzyt2pnI+N2BD+U0VDdXS6/7IaCyauooYVWWGOmJWv/E19L7vHbB0vu8DanqlYawOz8xow2kjJZQiFp92Ck/f+D2cfXHrVkgptyyVlmf+9fOsPukEOGEVJJXAatasoodamvT2ZGgI7r2f+790NcdkGXHov3kxI3S+fKZIBSw95RSPmhkUffFUCq02LF++fLjRaBDH8Ysw6ADW0Gg0iK2gagSlusXuG2Drt7/L8P0/pP/cMzjujW+Clcejenp949PE/RSOg0cCEwIaBvfN77Jv/WYW6Jg8IHJZMTMBx1RSLJ51AiME7fCtzRK+wOOSr9/EI1+8hmOJ6BGK3KU+tYJFtEEgNg1nbn1EJ6CmNbuloOvkk7jov/4iXHiej2iEAkohjesgiZBa02g0qM5E5Tfdcouih0GSNKzvOZgiNG4+lMU6tClcV/yBkHD+OcxdvZrhJ56kJGd6LF+6RJFiWMC+WLPiqtd4cISkjFQJOb6faMfzjofuf5asUWLLC/s5aY036FEcLsjkyJNWs09rOq1FmxbP8eFMWWHMpQtQrgeG2X/djfSfd4ZPjWuJZ9ArZPJmCxZ6CnMjEU3FiHGtmuroCAcefpC+Sy4CEyHVZHdrqhPFw4hqCZEi+S/v5uJTTmDrP32GfY89QbkOkSk8NN/BriYPzU8lTgbYTO84OKHIhGRcx9S1JJeSupAcf8bZXPi618IVlwRDLrGZQ2o1xZr8qGVS1CWKhqrDfKz9trSVIXCTn4SXIkFZF9Hs2Djs3s+zn/sCfQdG6ItKWGvRSk/ALCg+W5CvWKkwQjKcGcYiQbpgAWf+2gc91WqifGmmVCJrjBF1d0CjQam706MitnVeH5UIi7LC18CNL821r2lbSRgodrrzjZ7aQUcC8WJW/+kn4LzT2fe1bzDy2DN0O98cKmw+fYo0ZGULWFQnPAaILTJZAYyuoTW50NhIU1cxeU8vy88/m5Pf9mZYvggWzcGJHEGZnHp47ovygZzQR4T096m6ejVZucroSINyND1w0uEkspYozRjeuo2H/ugTnPXb/91nUOICVVSGzDXeQ0pzuOlmfviPn2VumhLX6x5k5iiluNcOwaixpN2dRCeeEBwtRW4tLnhiw8PDjxTUqXmeH9ro3No38ogWIksdSkVImdNtLOU9+zl47Y388Kbb6V+1muNOPxe99lQ4aSV0lSGSLR1prcc13rwNe90PePqmH9BTD6kc5RvYlJkaZ3qmhSjECsijGHr7YLyOLMe+c9QFgzhWZ+cXv8b8fUNUpcBJh7XBewyOhMajQBXY384ZhFbkJcXz1rD43e9i2a/+KszphHIphNSa5pNjMpBQ6ekktTlWqgnkLJO7pNuvwBRdtAKkhapQjGzb5YlEersAicodzvoo2ErPe25bKRAyYzwvO1DPQNmUqLuLxVdcyiPr1zE3N1R+VLpvCnFSMGgzymedAWef3hy9yMLoU6Tgllse5Pnn91KuzOGxR1/g9W88rfV5C0JJOPUUyqtPpPbcRkppg0hFWJMeylA2zV6xwmMG9BrY9cOH6fnK19EffK9fX+lIbYqSMSLct4nMwuEeuYk/TvxPv5947Cke/6d/ZNfO53mN/D248BKPvRCUqWvLhVshKCa6i7lrRXAwOstw5hqW/sOfwZ33s/8fvkht4w7A949k2gbo1pmfDWMzrHToasLw2Bg5jjEN9f5u8vkLWH3lq+k/+VSfVSsnvmkvDJYIMUXT348tjg6KsbnIshURH1YPtfWatOkvOdlivRhpNpWCd4oUtR/cQv3ZjRyflGhk9ebpNRqNCalVga+ZOgG5UtS1ZkjHdJ11Fid//Hdg2TKIysGR8+FslHQG3A5LqVJlrDYOCJIkJk2zKQ17kfQUtJz/5hnnGWMDA7BnJ8zvhSjGKYlWijRrQFSacKzi/K0SSJw3XqoMV76eua98PXO/eDW7/u0/SfcP0BFFFARAkz8PIJ3yfbRYjFbUswY60aRk5BjqWjNKCTu3n+6TTqT3pBNYfMXlHm61q4cihy3wJR9NHL7JUuBF2mY6ISxAEsPaNVSOXc7wuvXh00du0SekxoXFCcMcKYk2buPRj32C4y65hO5XXAQrj/WlCRX55/7hpxi590E23Hob87M6JZMitMW5I3coJt/bTEoGhePYc8+B5cf5CEd4sDaFJLcZQ0NDdymlEEKQZdlRRuhTPcthvtA6P/fdhaLHSUZGxkkff4rnnn6eMf0tZF8HlYX9dC9ZGGg7fRfngee3km3fQ3nXQeap2KcyglcnhPDdzkcdofvbngMNqaG3HyrVVvbUOZ8Z2L2XfQ8/Qn+WEgkQWk/KBthJ/4JWiiwS7LE1jr/ySub+1m/67sNyAnnm69SFFykCrYMC5vRgE0lmWvy902RzJ0hLJ1kUOfnunfDQw7C43zdopQYtNMQaY1KM0m1NcQKlVPMYOoLUBRCY17wSd/13aWzZSiX98Vh0Jwxjec5YtcrpV10J/XNAuWDK/bUND8P3brwH60pkmWbduh2MjUO12/t7MfhMR18fKy++mI3PrKdPxzQaNWIdcTQjfZ7Yx9HTSHnki9/gnK458ParoCMiVp69yYRsTguNrk0EPhJvSw8ogWfU27OXg9d+h8e+9FU6R0c4JrLsu/bbzD3/UqgZiCTG5VgpvYMySYryurN+LJLceI//he0ceOAhRodGSaQE51PsBSSmm8HxzqSklmgOYhBzuph/5hn0L1iMXr0G1q6BefOgd25AUaRl/CSYgrzmJzigNgFT/WhQ3yZd/wSmupd8UoWxsL40ODLM1ocfIWmkIFwT2Gr6j/smRpykrjTJimM57vc+DMct9+l157M3zVoyABpEBp1dRHEMtYz8CBAK25ehSRZSKaOyGmN33E51xbFQSbyPVqoSGxeg7Wa4ABsaOKrdcNMPOHDfA9TrNTpKZZQNNMNTSoB8dpA7QGvGFYyXI5K5PSxcsYye5StgxSmw8gQ4brHXoeVAeZ0LcA6HaDrZNPvai14A20p9F9kHISAuseZVl/HQxo10WBmmBF6cCOdT7yXXoCtzvPCNb9K4+fs0eip0LphPJgTZ4Ahy70E6RussQJLYtPl8vpjsQGG/6lIyWqmw9pKLvBERxcpqBI5Grb7ngfvvf+D8c88jTdMjI2dpfomY4uSExYo60km01VinqOXeX+yIBNZaSlmdjrSGGRkg37KF4Qcgw5GHlEFJKDpSR5fxVeE08IVre5SLIWzwFlspbSOhc/ECWDgPtKI5/GEtZA42P48ZHqAkckxqiWTr4fIZgpbX1NR1OmZUOsbmlJj75iuhowOqlaJIHWqODqT13yc9yw4L55IniiwTRNjABuS910PBNFqirM8WOGXQzlEdqbP7X/+dBbVROP8i6JzrzyGrobqSlnkL+3xwpEalVCbxNOAkQvmGxWMWsvJ1r2bLv/wbc9LW9zVR9OyhadD2EZjJSky4ArO99fDkEoZMRs/KJUSXXOQdNanxXOESZeDu219g2/ZBenqXMj6esXXrQe69d4TLX9MZ6mJ4pdfZSf/ll7PzG9+isWsXidKez/sQXVS4Cv48XNO1CUouUVQQLBys8cjffJYzauPeqHdmUCojFOToJvZd4f83V0MJr6Gs8/ttYC989wae+MZ3GV//PIskVBDI8Yxt9z3B3CeehTPPBOX3kosUeXOW2xKFIxssCokuouKDNew11/LUNV+FHbvpJ0HkOVFoirPBaWxlJA5VrHUt2VmtsOoNr6X/ba+HlSvAlcNCGF+z13HTQbE21OOFa9XofOGrbQWaFveQ73spMm0wPVPqvZkxmcqohGbWl2LaJ8x85Xg86V3sfG4d8yMYN7QhuU398aIko60kk5rj33oVrFwKSRhBavYotJXilPDAAv196FIZO1bHODcBxexIxOKThclIjfX/9+ucLhW89pWwcIHnJI8TWkpxovgeIV9eZP8+Dvyff+K5b32HRbmhXGtQiUrNLvOpgZ08s2PdNFCVDsatoDF3Hqf9+vvgvLNg8VLPSFmdQxPqU0maSEjBz5gQMTtaDU3NppR8wu+NEKgkIn7VFcjrriPbse9ofP6Ja2DbXXoPkrMs0YyNjzFSGyXds4cISVekqVrPNKnw+joyXifOTPwzs6RSUlm1HC650EfnIRUjhEBaS2N09PYLz79gTwF0VS6Xj9CgT5F1a3bPFopE+JjGKRXGuwxSCLR25A2DtJZIamxmMVKQOYcVEm19nSkWAuFyhHMTPewj1BvS+RYlJ3QoG1gyCZ0rV0JvLzRqGF1GOeM3auYY37WHxtgY0hicc368oMiUS38jdJjtM2HX59aSOijPXwCnnOJZtNKKDyUzfDRps9ClEcyBlNDbR5aUyLM0eLUizBFPf7eb6+ACZCeWqrM0tr7AE3/7abK5X0Ueu5LR+Qu5+EO/Cl3zEcWTYD094xe/cC3nnHM+55x+rO8rKXRfZxfdF1/E2H98kYYMdWmgNWpCMzqZ7MxNHi8Sk/670IG5VBwsac699JLAWuSNmUZTdIVef93t4Eo0GhnlUomxWsLN37+bK1/zWr+EzWUUcNwyFp9zFluvu54lpZjG0BBRVGRVWudayFTQo7XxGsaO06E6EXnGw//4OfofeYhlb30DnHkasqvboz250NmtwnhNMS2RWd/AtnU7o48+yNPXfpX0mWdZmEfMReDqDaQ1JEjioVGe+/rXOWH1Soiq6FiTNpOjhZbxQ2AKiLCBxCPHfOUa7v3sv7LcOUq5B+NASWTe/BjN2u404pCccdYFlD74a4GNLYIw8uOEDQ1QrfhHSg2E8ZzmMVprOzH93BZSvmTbbice50VhsrcZ9SloU1+6WJ8dObAXVR9DRIJcgLBuZiRLaT1TovUZE33R+dBZwdRzVFSgaNpmTdZZ4WEbpISOHuI4bmJ5R5E44vIjeAdyfLxBVxyjDxzk8X//D0a/fQ32mGOoz1nMK3/5/agzT/U6a8IHCXvdwNggT/35nzH0vVs4Jrd0CoWSESbzRTMr3LR70AlDVCmFPQ+nvPmtcNVroLviSw2d3fh0EBySCCqibVrO7iF7rdBRE9Zf+pHTJYtZ8crL2fV/v05n7vzoNTboeNvcY4cvcbccZmUt0np41UqkyI0lKmkwFpM2sLlBahUamVWwGzPLBNQ8ZyC4z0bCWCxZ9aorYG6/DxZ90sIz5VlHo1b/SvOjzlGr1V5kU1xxieFkrJAtzA1CGkZIMutwWe4zJEKQYprr4xlrXPgdZEX9IKSKmnPv0yD9tI+gSVfA2UgcgkaeUqpEHHQZJ55zJlQCyo7JaLr8lZix0TGiuISp1xDK4YJTYvAZH4lnb/OACZ592ViLE4q43A1R4pVk4eYWsPaho9NkdVQk/N8tOgZR7YHR/UinMGEfS+cmBBhiQrTRqt2DN6JSeQOmU8PYvl3s3r0Fdepa6CyBdUjpJ/8bKsdZzUMPb2PPXsnpZxxLAhihfI+AcbDyRE646DJ2fe/7dKOoarC2MUkhzqxAisi8iQaHIMtSokhikph0zhx47Wu9IVQxceq88yPgsScO8uRTW1C6B+ss9bxOXEp4/Il1vLD9tSxe6PuFPBiLgYqg773/heeefJLd2zewtLdEOtxAuMgb2dBBYCU4WeC+2wkGohRSFcIaysJxTMMxcOs9PHD3vfSecBzHnXMWcvlSWLYUOrsgkVCrw1Ddc5o/9Rz7121i8xPrYOwgvXqckmmgXOrXIILISJQx9GrDtntvgqcvhUsuAFlGkfgxtKDAFcp3x4NP6ToHzzzNg1/+MosbGbENFJDOeDCP8DgoaEbp00lkYWjjFkq7DsK8Xr/ooXGoaRgcQaEar1CUQiKJm8nMQ+9/Hj7inw9emvEUdhoDPiEv8tMTAR4sBMDC8IAnXgHGanU646RJ9wwTI8oiq2UECOknC4omT6H9PrTSBa3lrZrLcj8ZIyPo6GLegoXYbTtRUtJoNEiKVFvzC6dfI+XnLYlyUNJRrjco7xplcHCEwa4DKB37bGXYVM18RmFcbQrXfot9P/gBx+kSHdYiUs9a6KF/HU10o2n2oRCe6pfcwgubQumi4bNcBT+8LXplCFGo8z0FhGbksBcVoXyJHxaBSQbZFSUiDR1zmPu2t7PzjnsYWL+RXlXCZhkyws+8OwI4jpsYkEy6fxP2ppDNUhRGoITG1guAJo2MCisExaypQLY9qxPFd+r7jKUgAHUZgVARB2oH6T7jZKrv+oXAKyKaDo21jkbaeOYP/8cff/9LX/oSn/jEJ/jkJz8JHIZtzU1Od03pzbRHRrZY19DQNelYh0hrLK11wbJlzA8jLXxqD3cBFkOGiBQjAspLlzDngvOhXMJIRyxDE4e1UMs9AJgU3sueZt7T10FavxPW4VJHbd8gjNShQwbSa+u5spX2G9KG78PA/gEoddETdyLz/WD8PjkU3vPQ9ZkK91g5UMZCZulRGYuWLaTQqlk9RZZipIjZsa/Otm3DbNv2DFu2ZyxZEhH74Wrv/UYxvW+4ivV330NpcIxSI0VG/lBHU+6YEKFYhzUOXYrZO1bn+Le/BebP816zxK+99eMX37j2FvzkpGxGKtYJRkZq3HDdffz6r5838YsiDcct57T3vJO7//pTdLqcWFmUaY3QFPS+E7C8p4jYrPTjRLoBC7B0m5zaE8+x7tmN5IkmiwSpcJhAvBHljrgBScNSNoKlQqNcjlIZiNz/nV+NsO8dkc2ID4yy81vfYNHZa6EjboPInPR8FVWCNGXHXXfD/gF63WT2t/YUsp0Rxx783P3AcxsY+vgfsupXfwEuOQ+65nnN2SSUN74JL5Y+CvA1DijSjZOjpvBWO/zlS5am03qURrzgdp7w84+wH0QUOc7wc5GxMhbtNNVSBTljfdafT4EKF1kLO/fD8mXIOKJhQUuNwYbShkNo6Uc/agason/BQvbi9dQhxvwIRIZu88jmnqQHi9JQXbgAFi+c8jMuQAgra9hy1z0sMlDGkkhFHjak37szr7V1DpfnCOXb2p67/U46PpGz+C1XwZo10AWQQFz1fcRZDqXQUGzBWYtTPkCB6fzGSful+COt4LjjOPXXPsidn/gUqp5Tcc7rb+nLZYYJQ6dTX0ObXZosh+hI1z7OVjgEMx5+Yt8DfqJhwGaMLJjLml/9oAe/asOByLKcONbs37n/nz/zmc/UvvCFL3gMBOeODvrVCiZ0Z//45MgeyGIxtQ0OgPBpRKENutzBTiM48ZIroX8e4L27RqNGomPv1ZqMuFpFO0XsBFIockdAGPJTGx4hr01RB6k6aOwagOtvgV/4Be9RC4dQDo8On4d/PGgNjz1B+sWvo7buJjYS8DOtnnzBMpFFbPItmX49jJDUy530nXmWzxRIvyGEESglWPfkJkaGG1TKMTd+9w5+47eu8Lq6SJkrCaedRHXtCdi7HyXKwKgkgJUYJAU0o2zWglp49GrKVKMUCi0icqOJenpYcOWroLe7CXmK9tmMp54x3Hb3ozhRBpn6CNt5UIfcaG783j28+tXncfxxxZE9JwDdCZW3vpXljz7Izlu/zwINSqTBCZEBMjaIkwg3PWxm06GSioqKKElFR+5I6xkGh1ASJ32dOhYKhUIpn4VyQZHnTiFEicLySkLtVHvymw5XZv29T9H/0DPEr7ik1bxDq66PDOcRUkNjO/ehEIzX6yTxdI2ah5fIWuZFkv3PPcUjf/pJ9Px5LD3hRLqPWYpYtRqWLoUFc6GSAB0gKiEyakUDU03sSTyXh1/jIz6daeTwsKbTXbGc6S+axv2lRPltHlfQMXTMQcgSOqtRRZLNwOftyYFytJVom9KZxgxefzNzzj4LyopESAo2BAPE5EibwtA4PL8H85WvsOfe+6kmCeO1GqVSyfeNHIVY0W60JCZNGVUNlpx8EpRjJpMjWdqCLyMZ276fPptg04zhtEYlsGsWXBmy+alDRSFAaPI0pxxH9DYMQzfewVN3PkD5mAXMPXElXccvh2OXwfIVMKfPQ4HHMegYoWSYAnE08QnDvmwmBian4NtFR/DGN7P4mefYe/VXWR5HCJMTGTDN2lXQKz8VsRSDr8p6HdsQhuFqwor3/RJcejEhAmuucBxrxsfHn7v66qv/41d+5Vd4xzve0SSLstYemUH/idjxFyHtWMgFAYeRjn15jlt5MuVf+EXfMBZ2QhyHxgIloZowZ/F8tghwIW8jnU+1QwvyVlo5wUuTWCoqJq01eOILX2btmlPgjLU+rW8z3z2f5lBL4fF17L7+OnbceR/V/SP0olsMdVJOyGocrQghyKRkvFyFVatDdAU68ddoHKx/dhu4GOtK3Hzz/bzrF66gr6/NMVMSuquc8toreeKBJ8hTiS/UQHv1Z3LNfOLPrQYtj0ol0aUq+3NL76rVcOKJvtRR7LTAPHT99+5kvKH8FKNrKV/rHB3VLvbt2c/VV3+P3/voa3zjK617Lbp7OP63f4fhXdsZfvZpdOrQtsj2uAmRqy/fzGzUcQ7TSLHW0lHtot7IcRYSGaOQWOuwxmJt5qGInS9tWGHbHD7VxEf3ZwG5iLDlDmR3P1ZGfl+opPUHzltyh08hSgmkDZ+exYV7+eIjTukskRL0p5Y5I+Ok4y+w74WtbFMxjajCWKSpLlvI/JNXs+S8C2DNybBgoR+jkrL1KprmaOZBJt6Ql4W041X8CKN0oJV10dC/EJNUybIxnEkRkZhWPwpoNpc6IDY5G2++lbNPXgNveTVUO1C5QgmIZOpxMTZsZNc3v8PWG+8g3r2H+cqS5lkzCvNIZkd3fRaaqflyZw9Eit6zzwyNaDN/1hhJo+7RQXVRXyngoeGw+9M5R6Ij8jSlI06IU0OjNkRteJgXnn6KsbKgXilRmruAeN5Cjjv7POasWQOnnA4dVZQGpcM+FJJDRhmLf6fahkqCyVnxwQ+SbtzEjnvvZ57QlEw6xR//5KQ5nuZogqQ6AXUl2SMcy658FXM/+AEPHz2pEdI5x86dOz/54Q9/eKTRaHDttdcGEhoPfTytQT/adntbRChh4xQn/WIh7w4nHsMarMt8/4azGCnJZcJQZxfnffi/+zGpthBDiPBgGgtCIFcux3SVGK0P0+Ekucm9V0eI9pzvD2g2VwlACpzJ6RIKNTDEo7/xWxx/wXl0nLYGqmUYHGRsxy52rd/IyNaddNRSFuQWrRWxESFCB6cC5rEDpuETn0qEEEQ6IsszRho53StXw+JlPm3adq2NMXj0wWexeQyuxI5tB7j1B8/w9neu9nukeA61gstfQcd/fo2hJ9ZRFQm6+ZC26muHigmwqjJEIjYcLmJcaka6Sqx6/VVQ7cAm3iL7+pxiz164+dYH8Yk80WwIChdILc2Qsosf/OB+Lr30NC69eAGygIovsrLLjuPM3/s4T3/iTxjesJEupagK6Ttn8QbSMXUWYbI4a3FSIKSinjZA+1GZ1GXIvG3/Sl9NkQics2jno/fCQ3ZE5CZHCIVWmmHjqB2ziEv+8PfhzNUeyKZIkYRuXmn9ueb4RA89ZcoL+zDKjxySvjTlk2WGSCoi62uZiZIYI8isI8syGus3svfZdey97jpEVzcrLr6A7ssuhwvO9wcoJ6Cq4frysG/UjIFvATf6UsvrL0ba77d0L9HfKDIU4I2J0rB4CfNOXE1jxy5Qfh9M6Hpxuu3jCic9jkWufDCwYHycRz/15/R8+xqWv+ISz2GRppgt29j53Ab2PvMcldzSbyylSKBDOfBojXghXud5nAMrYCQzxAsXwIrjoFqdEKH7oiWtMTGl6DpmIY0XtiIyQ4eOcaaBwKJd3gSJmVrC/kaAdWgpSNNGICRyROT0KEFnDtloSjq6g3zLHh6/+0GMjjl21YmsuOACeMNrYfkxHsOEFKTnFDBOtCCVZ9pkpRi6e1n9qT/jmf/5p+y57RYWWYNyItBf/+Si88lz5s768o0TgrrM2aciuq94FXN/56M+S9FmzB0OhWDfwIFv9fTO+Uqj0cBa6+lVadlZfciXtHkPL2cpLiAux9QbObkQ1Eol9kSSiz7wq3DWWdBV9Z5JqGl6CQViKaCvk96TjmNgYBfaSDqiErUJhlVihAzNfyECkL7rWWAppYa5Ddh36y1svfUWGibDWEtJKbpVwhwloZGHUlyoFWP9bKrI8ThHRyc+JR3SRUnMolNPg565vo5vQ3EewcB+2LdnjI7qHJSqEqkervv2bbzlLauppV5PG2tRkYJSieMuu5RHnttISQb8ew5fA/Kepa8aRc4ggHqWMlLSqJXLfcpIxZgwP2qQ5Dnccc9TbNm6j46OY2gymhWKUPjRKWc1Nqvw2X/+Bqee8hv0dIAtemWK7phTz2LNx36fJ/7XX7Nz/XoWCUVJWpQ1OFSzRDBTqcgFY9zE55/8B7KlSJ2gSZ3aRBMUgeHIKSySuGMOwyZnb73GwgvOYfUf/T4sWdykfLTQ6t4vjkurhK6U45hzT2fD52MGB0fplto7My9JvBfkIYv9KE1M6ptRRUoPGTozmCzj+a99nbGbbqL/zLM44e1vg/POhR7vOCNL+FSLY0pFfpS9FzOer2t7HU4mfWcLv/xHLFJDuczC007l+XvuJs9S7GGut0hKOwXC5VRSw0JnaTzzLM+uX0cqJc5YypmhYiVLrZ9+MWHfiR9RtsGXdyQjNqOyaCH0zT20u32yKJh78kk89cgjLBiX/gEMpQzl8mBLJzFsTiFFRq+YAvTPgCB2NkxtOPz4bs5cFNY4Go8+wVMbNnLwhhtYfMn5LH/Hm+C4pVAxoCOU0hymBcxLHpRGdxerP/YRNtFg0x23cYyTlHPrdeBPuPmycM6skMgooa4lu5Rl+WtfTf+Hfxc6yoCArIGVCocid4Zalu+86667P3zllVeSNRrI4Iw1wYDsDHdiMu3ey1M8X66OgUiyX8O5H/kw/OI7obM6oQQ2QWRoCuvtYMW738DBDsk4BhPStkVTXi7ACO0bngqCGSlajEJaoFVOlGf0mJwFUcSSUom5SlO2BpdlWOG5p430FIm12FLXOZIaihp+TGxmjT2ZHSrLMqyxyHKVeaee5a+l3vDnlmdg4bFHNjMwWCdtCLJUIFyFDc9t4+67t5KUAQkqcNUTJXD55ZiF80m1r+n75rKZN3pB+5iHgK14aOtJxNJXXeozJELTMmW+j/C6G+6hp6efRCfgYrAl/6/TYWIhOE+2xPp1A/ztX92AVJAJh2nXQTqC889n7d/9DV3nX8gepUilxgl1CGLWIeceXu1NZRLfGZ4YSIwlMX6eVDvbLAtYacmlxWjIIkOm6xhVRzm/ZvuylG2lmGUfeA/H/O2fw8plHh1Rx9RpQPGdgXQwl63EA+C9j9NWc8oVFzOk/ahTKy3Rdj+OkG2uYATMhCYTMRCoO8kRNNDKEekEhYYxwyJZ5QRbIr/1bu77yB+w+w//DJ7cgDTGU69K0coTTrY1L2rc7GUuYaP4xIQErem96jVEC+YwmI95iPR2aWMrbAKfCO/AGzKsTIkjh8zqVLKcXgPV8Qal1NCpIyR506m1h2k6OxKRrihp+Y7xMQXLTj/VR+eHEyWpvPN1qFVLGA9GWQiJwhE7R8m6aR04J0x4FXoiBEdopNMoo1FWI4JeVc47CTGWLq3oSWK66w26t25j/PpbeOo3P475P1+AwYNg65CPAm5i6n3yuQi8Mde5R15cMJcVf/JHHPtf38v2kmBMZE3H6achRsJBW2N7CVb8/z5A/x//Icyd43taSg5ihdRRwXeRP7dx/a9cfPFFW5yzKKWQoWcKWoG4nEwnaK2d8CrEYQOtraQRafZUSuwtlxiNNHXlMaE9alX7i+lp66yDMMNZvKaS9t9P/ptcSkbLmufJGVo8j/M/9jtEb3mz58/Oi8g8WPVmob0turASzj2XFVe9jl1aMRr5EYLm94i2h8o6nDPNeocT+O5nBFWpKDlJNJ6hxlJ0PUdlBpH7+fZUWMa1ZKAs2dYZs6MzZqAsqasW25A/pmqOMxXn6NerrTZrHZnxwDw2SeCE4yEWoCNPwYoEBQ8/+gw2j4mSMs45pIyRqso119zYwrKIVcu5Wbmc5ZddTCN4fT66nTiS5nnnTfNeiDaedos37qaUYOZ00Xv5pa2mm4CJb4EnnxjmmXW7UCKm0WiEGuPUkZhwms7SfG67+RH+49/vo5z4NFtOCOwFfhxx+bGs/uQfc8ovvoudHSX2l2JqkWxRRxbriwzP/aEGUjIxupzsC7b4xg1OemNqBeRSM65iBkpldldLRKefwkV/8kfM+8iHPKBRpHEVX3KIaHWrFhWE9hKgQvrmB6Hof/97WXT6WmrOt0waMQWzFSBD86IIdLbttLb+ff9XNnzehbVAZM06aKQCYUs9p4IiOjjCosyyPM3Z+/2befCjvw+33Qm1UbBpa8Han8dDnt/ins70ApoNmN5DaD7rhccQkCin/nT4L1G806rtelPqs2LSTf8qdNXkHhGas+GTNoKSMKePte9+J0Pz5jIWtRx+aWVbQHDoBIISHmxLKolIDWUUSZrTLRUdSvtxNuGaAUOqYDiO2VspsaOjxO5qzFgkm983caSqra7dLm1OX6o0o+UyrD0VsDSmQJ/zK2m9YyFz6OvkzA/9Fr3LlzKSZ35dAzFKs0/F+aYuNSkz4p8Xr0NNcIRbXf+yyasuw/oLB9JYTCND5Bll61gsYhYcHKPrhR08/u9fYtMn/xyeXg84SA/g6aQn3aMJ4slbsDlUOmDJEpb85m9y8cd/n+zEE9hTShiKJHUtg+6X6KZOOlwG4NBdaSe8dHjuCv54f+/qSnIw0eytxIyvPJbzfv+j9P7qB6C7E0pJoFuVQOSnfsbgkUe2fWLl8tNukKKKlNGEUnZ7lv0IgWUcGZYcS0kkJCefzBv+5q/Z+M1r2X3vPcg9++hVJTpR5PWUWMhAOQlI5z0J2e44TE24MuGeFMhj1iCERoYLqNfH/fE0jCjNZik59nVXsvID74NjjoVS1UdEBlwdRAkwsG/vOEpC7zxfgwGBy2NE1yIWf+TjDO1L2XLnvSx20KsU9fExVKSxNgUkLny/dCBDa7JCIa2nxsuNBSsQUqClRqLIXU6a1qkpR95bYl+1zIo3vpZFl5zP7s/9K3vuf4DqcJ0uqbFWEMUSaz3ZgHLhRokMgQ0xvMJYh9QRQ7lh/vEroLsj3ElvrHSiGa3Bk+teAMrkeQNfybJoXebZp3ez7lk4aQ1kNNBIdCkCIeh/x1vZc/0NUMsxAWJUIRHOIgNQjii6uAi1toBJbpUmk3BQW9ZefinMnQtpw6NRCUFuJc7BNV/+HulYjI1zlPJIN75jNRgKpyk61gEaDQOiyuf/7WaScjfv/sXVfixSQaORUtYxMopg8UKS3/pvnH3m6Tzxb//OwJOP052mzJEJDsl4vUap0oFxAmwb/K7UFHjUPh3oaS5liPKR4fqEvx9Khh6LPKOWCeqixHAcw8qlrLrq1XS895d8rTzy0xQ28oBHCh/ZTN7vFg9fYIp3RAUaFuYuZMFvfYjxP/87Dq5b7++FlCQGhLMB7KftOZITA2YJYEWTXQ3platRIbkfHjZhIDepp4uslkhNipAWpEIZQ1dWZ3TDBh76k//JWR/9CLz6KqjGgJr4wFrAegpHhEI4jckNkW6rA1r/7FvnkEJgpcU6A2QolyCdJDeAUGhcaDBg5vJfke6QUK5WmsYiC45pbGfQ9QCZn45x0nh+HRtw9bX1oy7B424aKg3M6Yb3v58ljTrrP/efLLOgsMjcIoyntDW6GGByKKewwdhJXcZZQWel0ze55QZrHUKBVBKrNDkpdZMzLCNGujtZfsnFHPPW1zPywH08+H8+y6q4QtIwdCQJDZtjbY4NuX/XNLK+L8eKHIchMzGjOUTHLoOVq6BUIo50C7SlWOew58nrvknPGFi1innv/SX2/cNnyPbvR1pFXRdGLyYykshbTRqAU6ZpzFU7vbDP+yNpg1UumoTbAikXnFBFGIrBohTEzrDvu7fx3K69nPCXfwzHLsIDISmvl6zwpHiBrtSkDqEdjggihctBoKCzF97xS5x44RXs/pd/Ydttt9FxYJhqo063UWipQcbk0uFchiNrnqdonwqYIghxTmCdxVkRGJoTz/9kM1KTkZUTBiPHaF8XJ77hdfR88Ff8aJrWPtg0BpmUPdGT9sf48lce5a8+9TcPnrR6OZadoIaRpCDsIXV53R6F+xOa/PS0NIZDYjCorh44+xxWrj2VlY8+THbnXWx48BG27diNloJOISnjECYDk6Js7g2hEEghEM5jtLd3I7efRYtpyTsGxuWk1lGXgryjzEieUe3tobryOK745ffAeWdCTxfOSUQUfBRFE5Nm546Mj/3up3jtVa/mF3/5UtLUEMcaUY0wDYeqzuGkT3yS/Z/7HM9+69sMjY3RXa2i0jpRAH6QgS/YOtuk/nNO4IzwsZCUmAgy5bmyMyx1AWlHhb7jV7Lyoks48bWvhiXzII5Y8HsfQ/zr59l80804JKphSPBAHcJZjHXe8IS1F07hkGgVUVeSEWdYdepJLccpPI0p8PzWBgODdYQuBba60J/sNCPDGd/5zi0cv+ZyBLGv4kvnEVyOXcaC009j9623EkmJEo7cFRh5k7YDvkkqC/cuJWdMS/ZXY1a/8iLf4ZVUaOR1Iq1RUvD0k6M8+fgmqkk3aOsjdD35oZj4c6VcIssFCMU/feYahH4rb3nryV6BJyHilRpK0kdPF57H2tUnwM23sOWGHzC4bRfpwEGSuIux3BI7SywcMaGEYR1CSmwwjkUipxk9W8iEJUdgpSKXikxKUgSjUjB/5RrOeNWr4KrLYMXS4FjHFEa0Bezpio4jCkhSGZ4qEKhGBvUUDgxg7r2Xp354B7VnNzJ3/xAlrUjiGJemzVG7ohZfKMHCYBTPk3Ou2dzkhLdqVtpDFGnzXjabrmTriXfQgaM3idm75wCPfvofOH3RMjjjTFAlLGJigyWgKhVqkcDUcxKtSJUIqVrhn0nhn08DOCkDQIgltxKswilFTUsavuYyddTZLgrIBCjFQL1Gl3SMaUcWJlcy51HCphKHDOUZgZUCIyGPIqxW3pnA4C1E24fGrZ/ZV4p5v/7rzKt2cf9n/4NkaJRO5+gSOkS3IQ+UK+8Gt6VGjbO4gFhZLidkVtCQgoYWDBjLmNLMPW4ZS9aspe+dv+T50Tug86yTuDhJePLL3yTafZA5DpSwXj9J3w8iXIuwp7jDVghQGpuUmLPqeOjrg6xBSo6OPMW1LIy6zaE+DrYGW7cxfNs9DD3xLAOPPUlpcIiesMcK3IXJ/SneofDOhHTh98KSUtiWYgyz0ElFpmbqe1TsZ4GlnFuWOs3mB55g3V/+HSd94g9gcQlUBvigxFqLQFMbg7Qh6OoWSA3O0EbRqyBScMwSFvzOh1jwxtdR+/YNrL/1LgZ27aUDhdYKJQzKOCQh82JtM5sjC6z6wn6F9cidCfvJY1jkEVipMdIyFEFl6UKOu/g8qm95nSfnqXa10UrKVnuK1GQN+NznH+bqL91AtbJEp2mCivWMDurhIvSmLo9DxKycA6l8k09UhnPOIzr7TE4a2k/+5FNsuPMu9q1bz+iG5+nJLD0qInGGCIGy4IzzNzl4hAXACoR0oJTN2qYVUE9r5LFkrFJipFJGLFrEsjPPYvGll8KaEz2OerUEQpAKP1hfnHSWeWCnT/yPz7Flk+JrX76Pi19xMUuP1RgLeepa2OZ90P/R3+WiC8/nya99jW0PPUynyehzlrI1yABjJJzDubyZ8jDEGKeoO0dNw1hZMlaSJEvm0b1yOSe/5vVw6unQ0e2b1jpL3vtavJz5f/op+s4+l4e+9nXyF3bQMTxKDxllYRCRZ1aSuUbmKswpSho2ZzhS1Od1wymrfCQRNE6htx64/wmGR8bQuoxv3ALh2W4olbu4/bYf8q73voLjFqtmJOeNYkTfqy/jgUfuozNN6RAOaaavk0Exz+kTTAOxpvfis+DsUyAugxQkSZks9w7od75zK2NjdTo6+jEuQytv+CaPxQHNCLRhajhhsM5iTMw//v312KyTt757GZZw+eAfAiX9vZw3D37xvSx7/dvhoUfY+P3vs/PRx4iGDlIdG6UzzZAuI8J5hY7wY2jWIYQLD6nBOocxGUZI0ihiPIoZScqIBQtYeuZpnHLmOXDhRX5WNAmgRaLtkXLe6DZbVQThbwhgIiBT65Hodu4n+/o32Hzzbbj9A7jaCHPjGDdWR8cxKrM46+ugmfTlGeVa45XNTR8yO8L5capc+W7kAgLbK6SWi9bqNZi6Hm+FpFFrUNURB3fu57l//jwn/OVq6POUwM4PjPj115aDtoEpRaByatKRNfycbfG8yMnQXs6nSiwOJw0ploORJk2ER+mT7Tt7kjiJM3gwFmAsVjQiR+4cJiAFStQhvAStaxMhvema51fLczTGR+eFtKfeS9anNWQEJPBf3s25Z57Gpi/8X/bf/yDp0CCVeo4ydbSVOFcOsKXBoJsAIiMlNlYcsHXGY8V4ucJYVwd9a9ey5qILKZ99FvT1ex1b1qBSwBG9/5c44+RT2fTlb7HroYcoH6xRcYbI+XJigsdiUMF59MzUlixrMIjgtNPWQCIgLpNoSRaw0KWUHs2xZuHJ5zj4ta+y84GHsPvG0fWUqjWUtEMqG8ajNGBxGIwyGOXzeU60om1tFLHzDW+ZNk3siglc6RNAhSZKLgkZJd/bopyfbV9c7mTTbfczvvJaKv/915AVHXoUBErGpHX49N9czV13Pc5rX/s63v+BSyl1+vyBDnCeEunLgZ1dnslx7VpO/cV3M/7D+9nz2BPse/pJ1MAgpbqjlDu0ED7LFtStKmrWsuBLtKTOkAtHjsNqQSOCkdhi+jrpPflkTjztVLpfeRnM7/NNzEkFioBaAOQBpExhxuHv//4WvnTNDVQ75vqZfOloQQVP1cQyU5e7l61gM+Fk1FIBoij4+YaDuOK92c4E3dfN6ksugqER2LQNHnuCA8+sY3jnTvbu3kPsJK6eojIDuU+UNMHrpWgqHiPAKkkuYc7cJfQvXcKKU9Z6hqI1J3uwmErVs5slCUhBjiNBUDOOshLUDWzflvEnH/8nNjw7gHQ97Nnd4F/+6Vt88s/eRmZAaeHHtvI6VEp+U73yUk5Zewo8+ijjd93NwfXPMrx3H4MHD/p0DTTLBkJonEkod3ZT6esh7utk5ZmnIFYtgxOWQ3+Pv2lxGYixSgOGTEckXVUwoN/wJs676GJ45Alq9z/A3nVPsH9wL4PDe1BCoVKFMnHToOtSmaE4ouO0VXDKGr8BQp2oiMmefHwDOO1BwJRsRnMCiTWCgQM17r7rcRa9/QwqbUyvRBFc9RouTetUajWvDfJ8+pRnEW0GJDYba+T554RmlAjrM7dECnbtgPvvfdqflzWkmaFcrtBoNKY5eDDqyqM7RTrBZAKbJ/zvT1/N1p3n8pu/cTmdTYQTGZwS3+xjxy2yrw8uvpiVF13Iyl274cknMeueYuj5jezfshnlLMPDw2RZ5jNHU6DyReWI7nlzmb/0OJKFx6DPOQ8WLYH5/b5+XyqFVGLI7QaCn5a0+geksD6Uds6nInPrx+xuvY2HP/vvmHXrWZxDh6WJDlWXEmkMY2PjJKWIYsxIOLBo6lJiNNQ1NBTkBUua9c18ylmE83wJ2tLMLhFKC4cTJSOkkNh6nb6kxK4HHoZb74S3vg6jned7LlK1keaUd70TXrHL753M+r1f7B/nWi8RCg4i1HFFURqIaOQNRvo7YV6/d9JmOE8h8Xu0WuK8d7+TrisGPaCTDHuizXmZUtrJVwDqeeim7sTYOkpWaBmbokAucLlBxBF0VGHtyaz4+DJWbNoM99zN4HPr2L7xWURuSUcNzrbOYbw2TrlUolKpoqolepcuYOGyxSSnngrHr4Slx3kKUCUxaFQc++wRmjQfJq72wAUXseKkM1jx6CO4++/mwKbN7HhhKxpBbXAYmeYkKBCW3Bp0qYzs6EHN6aL7gnP8NWjp930kfLNjow5jOaP//gXWff1bdOzZRU89pywSn6rWAqmUn3aYdD+MgFwocikZjzSZ9AZdWW+IncvJVCgPOYibEXohUztchxCbOEmUaISzLLaCjTd8n7VvvAqOXx4csFBmiWHJkjXs3/cIX/nqrazfuIcPf+ydLDwmaTq+AlDCIWMNxD6EX3MSlRUrWP6W17N8ywuwdSvu6WfZv2Uru7ZuJ7KQjoxDI8PW0zAsFXp1lCCTEHdW6JnXz8Klx1BeMA9OPtED5yw91jsPWYNcgA6Q4U6KMDEjwUWYDHbvtHz6r77KD3+4jkplLjiNEIo0TSlFeeuZmULEunXrpv0lwN69e9ccf9yKD8+d1/eeKCqVC35uf0sKlN2gqEzWylXm+AfXpDA+4jm8R8dhx27MWJ366BjZeB0y42FXkwgXK6LuDqLuKlFfL3R1Q9ccf4dKJa+whfL/FnsgTNHk9QY6SXASahncevtOPvO/v8LwAYE0FXAaKVJyd5DXveFCPvaHl6IV+GRQSuYkMQnCCTC5P+88h7EajAz77srRGpjcG3TrEEIh+udBZyf0dPowMxIBdzdwHEcChCQlpm4tWip0mI/W4A1B3vDKz+RQG4f6KIwNhu+u++a9ME/KnLmeK7i30x+7koDyI1MGyd7d8Gu/8tfs2uHo6FxAIx2jWacGsA7jhll8XJl//4/fplrxNVxLjsxzSG1IlRQ3EaYO0dsUrfC7wYOQaP8ibo69SQH/90ub+Ou//U/ipBfn4kOMybTz4k0AC5+gFsKBTMnMEKesOYaPfuR9nHii8pFagVzaLilhT1pfE4wjqI3ByEGo1eDgAW8A8hRMTn28Rqlc8g5juQQ9XZ7mslIGFUGgV/X7UEKsSfMMrPPARdAkm7AWbvz+LZx3/jn0zekMxjwgoasYBg8w/i//xiNXf4meWoPe1NHpJMr4jnQrQLgGNjAQupBOtZlPreZRzFilwkglRh7TT3lRP6Jaol6vY4fHECOj1Pftg8Fh9MAo80odlKTEZA1UrMizjEhNT78KeOhh65DOkknJ3twy51VXsOh//xV0F53SsnXded1fY+iF8NZghrR5gZbYBAzR3ohH+LROaDCaXiTS5T5VbHLvJAFFMybWMi34STAwAM28vAmOYTkGpZt9P0IILML7IrLZhYMkR9hC/4V0xdiY1xm1Gnbnbg9OFbqppZLeCejq9Huorzc8N9LrDKFawD4oMvwhlfRNucL6KQVCOROZ+X08MgbjDdi1u6WDbThmV7cn+JjTCV0JRH5NHQHNLU9h9362f+ov2XfnXcyp5ZSMbS5JMXUinEUpgcH45lABUVwiNY4Gkj02I1+2gGjJQso9PSCUb3Cr1antP0C++wBdww06s5wKEtOo40PF1v1plpCagR7BOZCtkfPQaLe9s8JxH/kwHe95p/cSZAlQ4GDTJviVD/wlmA5GR8dZdtxc3vdf38CVr+vzviaGSuGJWuvveaHQjMGDhDW8oc8z77UYYPOW4Gm0si6+9qW93u/t8RnjUtnrTWX971WMEwohFNbmSFsP+xtSJJoYCXzvu7v5189+k02bdtPbvwDjBLW0Rlofet28ueIGlWxCqIHmOn33uzdM2NKHbYpbvXr103mef/Due+/+9Emr1/52V3X+u+MyVaV8A0lcNAlI58ncRbgr5eCB59qPDCxa6B+w0y3KOKq59R29Jmh8JcN8cfCCJcGIRf40g0JrBwoSLpS6MlDOUwFu3wNf+PItXPv1O0iiPpwt+25XAVZoHN1ce+2dRKWIj37sAiQxDSxKKBomJ1IJSI2IAi5vqQy9fbAM/5CFSE4UA6gmkFpISRbm26VUKD94iqGOJEKjURIazivnMoX+kMi4hNBB8VRKIOeAWup3s8n9Ratg0GvGRx1SgrQ+HY3z/3OwfVuNwYE6paTPZ1uaDFkhMWQc5UoPu7bv59Zbt/Km1y8Nc9AOqzSqpP09bBpcnw6dSpo+QvjXx6ahShuUpRQwMgLf+Mb3SXQfrhkxyUmfnkKKKQMXrj18q8ksQnbw5FM7+a0P/U9++ZffzNvfdhrFI+kf+gBwVoyrSgn4bnOXdCI6K3693bGhwOaJUUrO0USlUiKss8D6tic85Iz0+845DAatPQmqDddrBDz+2F4++9l/ZNWKxbzuystbC6ZC5DpwAPPdG3j8i19i/sGDdElNgu+4tmHMSYb18cbUPw+2YYlKZQ7aBmlfD6ve+Ca4/DJYOAfmdod8pvNTHmkGu3fDE+t44Xs/YOeTz9HXMCTW0kgbVOLIl4GmlZAuCxZfOUuXMAxu3sCi3buh67iWUgOsAqNKEzr3/Z6Y/h4X1zjxJ9m0SYrDYTVYX78MjqSMi/5+f6wZp6QdrUqDbjPoECLkUM6Uktz65UxT77vXgBiBQyGk8kh/gVCEuOSZxKxDHruColPfO73Kr5kMCq1wjJuKTYT97o8V/KnQf6SLS/bZtFIE5FCuQPcc71AsPzaUM613jpJy60bI5oo11xYkDB7kwP/6W/bddDMLrKGc+8/6cdvcL1HoMbGuBcFqheJgljMsBD0nreSMSy+EN13lx1UrHR6bXSjvMA+OeE6L2+5i5z33s/2xZ+lJKpRcFurwh+6RKfnEhUWQEbmU2Eq2PPUoa+S7gn5QTaVU7oCko8TB/RYd9bB9a8qf/snnuOe+1bzzPa9k1Ymd/p4B5AHiOgp0t0r5zGYkm7lPifR+6elrWwFW60yDrmi90yqE5hhc4KwIlSmpQFS886l9Tfz551O+8J/f49ab1pHWY7q6F5GmAqE1uMzfLZEjZI4SOa0JkolyWINunG8WOfHUU55K0+iD373+8b8RpnztxReuWtXTC3nVX5u0wqPuiDy4WSG606Fry1qvPJQK9YMiNembQ/wGdM2U4gQaHecBU6wQtB0ZKfzEFto3VN9442Y+//Ub2bZziJLuxeUSRN4yR8IitUKrbr75jbvZvWsfv/6bb2TR4hJJCTKbgwqc0MXyazUx79OMSkUrxeoi32FOm34Lu14RN+ulVSAR3uAJ4dkJtSB0i4cuPkWIaPK2To5iA0k/ahYiQA+jWtRSfET04ANPUBszdHTEmNzXuxACJwTSCXRcotFokNmI6759G69+zfsoaTAopGiN0RUKNaKJonyIFIqh2FYtQ+fPr3j/7js3snPnKFL1kbuUyQ/vVNF5+wyuK7qdgcw0KJd9FBrrTgYGRvnbT3+DG7//CO9//9u47NJunAEtg7aWE7e4SzMf5auwL4tOGRHR9GRc63H0l+571IsGNn/S4f46r0SM89vk4cdrfOUr13PP3Y9h6sP88i++v+X5SOHvaZbDC3t48PPfoH/M0G814yMppe4eP9csLFak3qG0uUe9E1FzJOqgbTB+/CJO+vBvwate7aOwIvVcpLNFMBjlDjh2Oce+7Y0s/PZ3ePazXybfsoO51S6USZtIUzNKcDKMNEQa6qMDPmOVWogi2kvUrb3Q/t4MEbYNnwoRWXOpaNkhplH4rUxh6EzHj/jJ0LIkJpzRDCL833lgH/9NJnyfKpq/8L1if/AH/8y73vNuLrikp+Ws2HCNjpYC0NEhX1N4uc5Ynwlse1dIf7XOtq4rIAN7enS8+tTt2VYDudZYYdFKTkIJDR61A4fvDfGL7HnSvAeaQj5OeuN32HzTzSzLHVFzcsJPZ2Qy902VVmKdb+4sTEldSkbjEl3nrOWYP/qQLzFWOj28sUjCDHjk9VpnF/R2wQlLWPSet7LwP67hia99mzljB6lOMT7nxbb0kQThQje9sFiVgpTs3bKBNaMHYU6vXxDhJwo6e6GjFwYPZuQ1TRQllKL53HzTeh64fyOve+P5vP1tF3BMP0gXe3qP3CClQXoIuaZL2NwkkhBFytZGFW2/a95MfK+V8G/49UoRqFC29dketGTbDrj2G3fyvevvY3Awp1zqRwpJI60htQ8EpSj2km292p3ANjmisTWdxCANUdzJeRee+sybXvN7u5YsnLfqgotO45Irz2Tlqh46qlBORIio/YVYTIgcRbPTUwlP52XwLfda6JAybnV0+xy6AecjNCG0N0qErGXbg799Jzx471Nc89Wb2LB5H7pzPjargIx9jUgHJ6NYEgvG+JG1O+5Yx/oNz/He972eN73lJJJIt7FUe8Y4a6wfT2t/WFwIXNqM94SlLZ4dIMsUcaTYuwvWrdvB7bfdx9nnncXlVyyjXPIBfmHgWxK8TaGa1s7i11AFLJiiYzqQExZj/Tzz7BZKpQrWGowRExxJK8CajHKlSj5eZ93Tz/PE44OcdeYcciM9VSmtGWnC10wO0FujUq00mKDNtwifscLbmOu+cxdpJuioJh5da4ZGmMniio0bInUpFbVanTiKvTOhuhCik+c2DPLbH/4Lzjv3RN7y+gu54pXHE5cK9LnWSYo47E9rsNY0cZBVu+MJ3llqmZS2kwl7sPht7iHaf3j/dm743l3cdc9TOKrYfA79c+eyctUyv6ajFlUW3m+zlr3XXUeyfz+dxhILie7qIkoSao1GSKp4XnIRrl84iUNjI8jndHLSh/4bXHkJmJS0FBGjcVK2QVmGPo8opP5ESvLmN3BqqZuH//TPSIaG6MgNUs0c/066dEyW4tIG+Z7taHsyLc7gSfe0bc9MF58XjeztH2vH4PdvSCa9M1HadIGd+LaPKA/3/YVT4Frf7Q16W+8DwaA62Ll9lD/8+P/ibb94Ee9+56s5pke3bYbW8Yxr8xWbZ+9/EmpiXqL9v13xf+F5fvKJvdx8621UO7t519tew9wOWhfsk0cIIcnDe1IUp+BzSUKAw1NXa2NRLvIOoAs9HAcGeOTb32GeFXQKRWZyrAgjnG2NVwXJS6GDAVIlMUsWsPSjvwPLFgZo45ASC1M5ufU9NOgEeiSQQ6mEeN97ONla1n/5akphcbVQh5RHigi9PdUuAScCIE2tAfv2QdccGiYlSXzusyOGajkCGsjIF4YzA852MHSwzpe/dAvXfut7XH7JGbz6igs474J5CDwqm3MtO32ItGUuXdt/T9CRxb2wAiNbjp1z3jlzwIOPDHHzD37Ibbc9wIF9KR2lhWgd00gtxuTEpQRrjJ+GmEqmAXHSJ5100tQfCHLxpZc1b661vbjGKuJyv9wzlPGt6+/nOzffx/wFczjn7FM568wTOe2MPvrngAcGbClUGbwVE1q31KQ51qaHW3wg1BcwASNGeIfP4Muejzy6m/vufYwf3vko46M5WSoolxaQpp7L2RgfjU/1OCslcFLhTIXde1L+7tPf5dvffYBXv/ZCLrzkeObO94G5RgTDZnAh3dp0UFzweoN1y51CGL95iyrCvr3w7LoG99/3FD+85z4GBg5Sr+Xc/INn+MFNp3HF5edx7gX9dHf7jGFogPYBufDnLqUIbpFo6o08eN0qeA1aQpbCnr2wcf1WlPS87UXDmhM2ZGINYGmkYygVkRvBNddcz2mn/xeqChptq1XYQJ9um6hOm3CXpvW30FJ+hTgLjz4CT6w7gBWSWjocIgXaPjGp8aVtRwCtEnqIXJXQKKkpejdVAV/pJNXyAp56fB9PP/51vvCf8zj/wtO58JJTOHZZRO8c36oQl0I0jfGVnihpsmT6VGjLIW2XPETHxf3Zvxee35Rz680/ZMP6LTy3YRO5FUjVhRQRDsmSpQvpm+svRUR+5FMCHBxkywN3U60NICRkYWY4a4x7kgVn/Zc4HfIjmsj6ZqMD2ShLzzgTXvFKr5ijmNg5IAv3oXAA/YkKUbynfDf+Zecw946zGf3u9+mNFLmdztyFe+tHJIpCA0rHOJsj8xq4BtjEd2xjcSEyTq2vELU7g1NlYQwtR9a1/W1x/tK1klFTSnHfgkx2qi20Kd1pvr/4c9f6Xn/h/r/rudcFjaKxXURYU+WaL97LHd97gstfcQ5veuNlLDvWX0uOd9IRgSs+HLFdLYtwVyUwlvlzVdonLUdrsGkD3HbHQzzy8FO8sHkHY2N1orLitu8/x1WvuZirXnMGcxd63zMuHCDHhH7Dojczh4LPCxNghD2wl4DMwRPPMfbcZmQOdZMjlGw2J1vhmhkowjFoW9MxrVnzjrfAMQtBdwFJkW4IewYi5QM0SzEcp5GJhIUx6tfeh7z9+2QvbCNCoJXCBuJJIb02UdaCMBSnbKR3tnAx2iSIkbofWzSOOC7hbCvJ21vqQLlRjPT5RgeoKGQzRQeNes4N31vPD259lhUrF3PWOSdx9rlrOHZZiTm9vtpZCWYot2BMThLpoATEIfvSGnDG+OCgWP/Mx7jjY7BpQ8oD9z/N449v5NGHn6Ney6lWO0liSWYcVmT++deQBXhvIXIQKc7ljI+P4dJRlBhvAtVMliMDlgm30ze6aDwpe2FwFZs2DvHCplu47rpb6ests3zFQk44cSUrjl/ECasWUSorujugFAvq9WCcZSsr2C6TuVzGajBw0LFp0w62vLCThx58hh07D7Bn1wBKdpA1FFJUkFp6hVos7mGuxi9chLMlsjTn2Wf2sXHTNXzlqxEnrz2O004/kRNWLeHYY7oox76hWQBpLnyw6PzsaiR9Xa027tP+mzftY/PzO3jokSfYvuUgO7Y3cCbw1eruJrXtQw++wEP3r6NvvuKc805m1QnHsmb1CubO08SJL93HkaQxyUErUoJDQ4KBA2M8v2k/27fuZ+P67ezaM8j4WIZ1CUI4hDCHroRywZgJpCzz6MOb+Ny/3EelXPHrlreyGS7gs7uA4dzEO58uYi/WPjQAOdvBPfdvZmzMI7wliaKeHS3ZyOQ7Ol1U7x9ySZksy3nm2SGe33o7X//69fTNL7H25BM55eQVHL9qCR2dMX29CZ2hT63Ycu2UjMYKTO774EbHYOeOlB07D7Bxw3Y2rt/O+vUvMDTYoF4XSJkQ6X6iSFFLG0Rak+Z15vf3hQWzTYdQGEe+dRv1PbtZkGfeGDb3/KR8K0Ua2gYDJcnKCdWTT/CAPdaFKYeZpcBak6oElZyl55/Dw9+6gSQuexa5aY2ebVUYwj02FtJ6A6HFhPBJCq9oH39qgB/e8zRxqZva+DjlqmpCj4opo4rputCnLwVMT7gz9bGOpJu/XZojVUA9baC0IssysoZidCwn1lWkqzI2JPnS1bdy/Y0/5PhVyzj7rLWsPW0V8xaUmNfnDYu1coJTodochSzzweWB/aM8t34Lmzbu4MnH17NvzwhDIylJ3AmumyTuBZuydeso//BP1/C1b97ESSev5MwzTmLt6mPo6kzo6yvR2UkA5vEvJ2BkPOAcSUHuM+ctXAALPLOZrrohchHWhSABaF9/v96Hrm0mJfK0M6DSDTKB9qCszZlpSavBUcZAR86SE1exc+NmkvZpiAmf8Jm/vEh1hPOSzkPHlqJSK/hrFwt5w7Z+mJgDCbYsRijN2HiDBx/aylPPbOOrX/s+XT0JK49fxvErF3DamiX09JSZO3cuXV26SRvsaNkuGbIpzoETigMHYP+BQQYP1li3YQ8b1u/ksUfXcXBwnEYaIUnQup9KxYN2OWdao9uT9qqjKAEXkU1b+XUKmcGgz6wohBCh7gOlUgnpSmBh327Dgb27uOf2FyhXJKWKpbunyrz+ufR0dzNv3jxKpTKVaoVSqYTJJz64o2OjHBw8yP4D+xkeGmVg4AADB0YYGRvHWYWUCVL5BcFp4pK/OOc8zOrhDPkh1yFdyB6UMJnkwC7HLdue5YZvPcq8+T10dUZ0d5bpmzOHzs5OyuUyHgbWR+f79u9hZGSEwYERRkfqjI3l1OsOnCZKqljrmx5EgHmVssXOpVSV/XtzrvnKfUTJA/T2dJCUBJ1dZXrmdNDZWaWrq2vCvRgZGWHn7t3s3r2L2njmuzhHMnKTk5QrxFG5Gd37C/SeshPBgzIBXUgppCgxPib44udvRghJo1YjUi140iLd08IJ9vvisAa9SNmh0VEHndUujMtomIzD7atDDXi4jGmV+ERxzhHHEZVKGesMxip2bq+x7fmH+e53HqCcKKodZXp7u+jsqtDbN4dIR5QrZZIkIU1TGo0GY6NjjI/XGR+vMzBwkNGRcXIjGB/PkEKhVIyUJZ/tQbaBIfkUaKPRYE5/b9M4iEKV5Zax3QPo0QZVdCvFP9W1CBvm/H1DjlHQUM53LGcGF2uspNm81XYKrVUO2RUrAKv9EHulk3KckNdqSK0P+8w0AUIcWCdI4iqitx903BY+e2V+w7WP8fWv3UGp2kmtUcfYBgVmmpxs0Cf8fJjxsjaInpn3wqH762g5KZpmQFoPihRBmjdI05zuah9KKawwZGlGpPsZHYaHHtjDow9+jzj6HkkS0T+3h3nzeumdU5lQ8jO5YaxeY3h4mLHRjIMDOQeHx6nXUpSSqChB6046O6DRaOBsg3K1CkTYekJXtYuRgw3uuPkZ7rjlabTM6eiq0tfXQ2dnJ50dnZRKJSrVKqVSTObqjAzv4N3veQWnrlnYFkD5FMbBTVspZxAXoCwB2U9Zf0dmWjsrJEQVEDGt5ltafqksfihAlCZbq5hKb38YARbkmZlw91pw38rDLQdRDpIclJa4YhpFh4kEPN6JbUBt3I/Eth9zQqZEOISzJJEmUt1gLemoY9+IZfeWzdx9y7NoPU6lJOkK69rT000UxSRJQhzHaK29vhgbo542GB0dpV6vMzI6Qn3cYkyVLJMoqUmSPpJIeZa03Bc2rWxhIEwn0/7+xdbQi6Zz3MQFcdZ3V0ulkELhrCNSPkLSooywisZYnZ3DGVs3byFv420txkGklH6Uoziec54hS0ZI/E2O4zJadYVnXobZZUeS+MX0Yx2y6WAcuYTIIaDAKVFGIdBxJ6UYaiM59dE6e3YexNoDzVxzc2xNipB2F0gZgSsjREQc+9EmZx2VssbanNRkOOuop3WSJEEIQaPh0KpKV7WD3GQMDVvccMbuvcMYcyDAYk4UKQSx1khZJrdlKqVuKlXThrvv2qYBJqrqYvym9d8RceQbzJI4IVYCXCsSLwy5PSTECV72DJSOBR52FmrVQkqyRg2tk2k/86MS5xzGeGrTJOrEmQhUGa09+OrQSM7gwXGEHEfJAQyG3ORN57JwVqX1eyrSCUp1Yy1E2mGyDIdCRglCi1CLz0nTjFK5RJamRFFEkky6Vp8ZR2aWDp0QjWchtTn9tYjgnUsCcY01MDoMJkPER7CWIePQbFYM904JgTHGszm16YXJyqOdQ8EJScMZ4jm9sPAY3/Bkiv3ki4PDBy1C9gIx1UoXeZZNc2LtvQmHzzIQykYz/a2cdKzJ63r0ZFMWIR1KSaAGNkUQkaYBFllqpEhChsaRZxljow3SmmTk4DibnhvEiVqzzFRwZUgJURQhSLCuhJJz6CgrrM0wzpGnxj/7TtLRUfEMi7lFkpBnAikrlOIKeYAsHh21DA+PY90YsLv5Xc4ZhDQ4s483vfkSahkkkfMOhvPXpxyUhHeS2v2kAgHOzLBuzR6IXNDslINDaiRFfmrizfD/ZCOjRDoiNwbVBPCaXpxzvpKIb9ZOq2UPLqYkLveGyljf11JrmBn3S4EnAr50J0WMMWByR6wlSaJoNIapjWeMj1mcGydJDHluyPOMvAhGpe/B8VlP5aegZA9IjTQxWlj/PhKlNE4anPKwx+24L4eitB69TGvQfVWsJYUhVkLipAj3I0THVgRb58hEo9nPldk88PEm6KjsmxkmnfQEBSL95lFtb0Qhi5O3nYyKfJ0zt3kTOtTX6w7v7Uwv7bWucHXS4Wzs081STpxUaJ5icT2SYpTAFZtIChohxSyE3/Q6jjE+N4OMfEcmziGUDo3zGkSCakKgHWo0m21C2lI3o62HyUkPMekkblLqFhHOYZJRzzLj6+lhgYVtHzI6JN6b8LMR09c3XYA31Qpfp7UWPWXn749Wir1pXQbO9xagHErpgHctQURIHVFMTggg0v5lJ+3PIp1miyBLWFSUAJY0zzx6E96gSR35OmMgzkgn10sAOip0dFQxoUbm3OShwHZDZ8Mkp38IYiw9WY59bj0yrwMdoY2nyKrkoVI4cR8jHMqFFHmesvvxx2nYjNGsQTmOJ0bOk3SKEb5/RFlBqiTDGpJjFkL/XB+Z+RIp1nrCv607tpOUFE7UaKQGqcocPm6YSpFN3vdhomIGpWcBEX4/2Qgdbdq9+Z0iD42TkiQJpTOpm/fY49HX/LMnIU6StqcmOPoFYUxxbs6ElK1ESIMRDUxo/nPOh7ZK+n3UqAEInExA+c5z33DnAq546DWRbsoch7SShq1Tr1miCATGdwaFjF1HTzc7cKhYQ66akxQFUJ5T0xc/lAU2bIQTT6CpHA/RkX6D5M43eCmk508wDtIxtm55nkhJXD59cADQbNK3lgyB1IoDWZ0lZ50O1TKMjSErnU3Uwh07Yf+BEZzTE21Oe0apeCsEmMalfrIgBoQlzQl6Ign3xmH8JkNHZT/IME0QGQAQQWUo5SF6HWAyn6UUwvekFBwR4IPdl2rSZ3zSJEzAW2++74raHr4Nv/0sHK0NjEQ0t1nokDxEfU3tuU3zo5e2+kzRPPPSRR7y3xaJE81p1Ck/NeEWFJu6beRKtf3spGjrKCukqGEWtbamuW7+frI03/FsAzTb0YTl8KnLI5dWXaf9HHzdpylTgIYUYy8Ta8M/eXHtdTxhA8GObe1Bge9YCQbNR3huwjUdeg0FSIj0GYrQgd+EArbWZ26kYHBwYOIJCQvGILo7qTlDrsVMGXdANdmocJLIWippxsa77mLVY6+Cs06HpNN342g/DiqEwInCqHseAg0eJGNsDJ7fyOY776IXQ6W7E5tNKm1M4bX6coKkoSQHtOPcSy/0lLth0WwGUgoGB+Hg4HAgthGgnB85nCGTM71M+ky4J6ZQlNPJFPr1aCPz1vHba5e67T3/sxXF8+//zgrpr7cp3mEszr1wNooTcs0aSatXgWDQW5Fl8W+7nrGtAu6Ev5kkTgIaScK+/QcRLKGoY0sAZRFdZcalZTzPqCgFzqe9C3jWmaRkcp791jc58byzYcF83zkaBYyMtvOSSB95CwHpqD+v8RTuvpv6wEBzTmJyMGZFK/gvjtawFpFohrM6tTk9LHvV5f77orgVqACbNu9n34GDJKX+Ga+h9Z0tPVH8txMyZJ4m6dS205xqZ8v2/SMKtot2/I3Wzz+KqLxdDptyF1I0O57bRQbkpGYary2dR4juPWuVa70nfFod2takAFCY4rqmgv4L3978LoeZMLNcOExHWnNtGsPiuM57307IcB0NpqUmhCbpgpeJ85TCgZ3OgxPtG4gQLc5wmu3XOOHadPO7gKBA5JSfC29M+nmyMghpdtGaAW3SSU75/ZO+w8lDN/kEtLefjEzYO+EhdSFH2BoFkT5ybVtDiWsaoIJPfKK09p5FIoPDV3xV7jwGvBCCPbt2N8tVRXuixMCS+XQtXcD4pi0eWHDavSqxTrUcDmcpuYz67j1s+ou/Y8XHPgannuXhj4tnUXnmLif8WWmCEzmwD7ZuYdNf/BXd27fTA+RpNmOKs4kBH0o4NS3pXnMs8opLQWlS57usi+3x8IMbOHBgiEj3g4tRUmMxINLWvZhyD0y3L6Z45oSdLigKx57CAT5Kgz7hbMK0AcT+eRb18D5BR7gJj1Sri6d4vlvwhUX2pAgCfKl54sk182OTFJjfYcVz5KGkWo799JkL53wWafv2nQhOCeficdiRwOoVmL4q4/vHKQuBtISO+eL40y2eoZKnDD7+MHv+9E+Y/7u/DStW+A3TbLUvPBYH1H0XoDUeyvPeR3js7z+D3r+/de0FBF/rspqHU85fvZOChoaxkubEV10Ey5d4qNw2U5Zb+OF9D6GjCm7avTVZZOtLadPFk3TfIZ+aYtnbV0zaYq1Fm44vpk94kc7u9DKjQX+xX2XbDHu7FHOhR3qUiRIi0LZ2femKssDUD/JRyyHKpviuIzn2oX/TYveZJtPQ/pnCE53BsLdYitq/pH0jAi+65NB2VhMOMcO1T8m/bCc8RIere/7URFiELe6tbc2dFooxMJRNdla8tO/vifvDhqZDIQQDAwMT/t7rCgELFzHvlFPY/vxWFiCJ87y5ZsXSF4ZUFRE6FoEliSJ6pGNgwws8/Ht/TN8rLmPJeeehTl0L1QTKHq+6yVq66wBs3cKWu37AtjvupH/vAeZnFpnTKs9MuWWKazJkClKpOVCOOeftb/dY9jJGiwB2FUG9Bvfc8zgFI6PD1xR9arktadsO83pYmSZL9ZPcT2284xOe5xnXbMq4LQyf2iYzVzjS1MeZ/BxP8VxPlT2dfBznJFJK9u8bbU75Nd1PKWDtGsyChdQPPk9W88yK1lmc8BmImTRfbC3zrGHH7bfz7Ob1rLry1Sw87wJYvAQWzQ9fFkNjzENZjw7B+o1s/8Ed7LrzPioj41RlC3XXtUWOrSywBzVqOr1aMx4l1I/po//dvwBl3zBY3Jk0g5FReOrJ9UhZZnpncTopbEvbKk7aszNlhKe+HbJtw/yI9u50c+jT/f0hcV2ooVM0tRE8uYC4336Nxakr3CH7VU2+5Bmik2Yz3uQza2ZMbdvxinnk6Y53BNJksqE5YqOK4s100vy+qRf4UMd5+oyEl2mgKpupxpk3xNE6fJOBC9oVVfF9k7+xee4z3LsXL9NdwJEds3k9rtCBRZTQtvMCmYxoOodtXQPNr/HFogmVSREOHG6qaHYSTTqGEDgp2btvkAfu3c15Fy7A5H4M0RfpBHPf9HaevedBRrdvo0tAqVxhbKyGUpCLkOUQIF3uoxMBTloy5yk/q1KgRkbY8bWvsuv679DR3UOpUiYplwK9pT+XkcEBXL2OGB9hnjF0GYE2JjxbLaz0JkKd9QxaCocUDqcNWSTY4WDNm98EV73FEzIFPoJMQOpg/Wa4497HUHGXL6gL4Ue28BmjCUpwSsf5yGXq8bfpZea/PnS/TXw2i3ubt/1cRNJBF7U9dGLyNzYVgGnuv4mp5Ym6ZbJ/3vrFhDNsfS7sv8n2vniOK6Uy9VqDpx7fzIF9ofUBh8MgdARz5nPam97OUxv+D0kF8vEhP4njQl+GbaWgrfANfd6J8M+FSA3ztaRj1z72Xf1Vdn7tm8hqhc7uLpJKGXI/5livjSFqdezwCKUcFmXWo76ZsF7CekRNyQT0b6vCtboci2A0swyVejn/Nz8Kq06BuBNnU4T0e1lFcNcdT7NlyyBxNB+wE5AMDxfuTKWXi3zckR5jwn2QrTIAbceZLtM++fttoEP0POvu0P01SY6oy336ai6H/c1LlakN34/r+2auDf+k5MjLBT8Zmf58XoaRd5u0znvSeR6FQTja/deiCZU0crj9rkc58+zX0mzul8qTYZyxlrVvfysb/+1f6ElTGuPj6KKhThCiONuWKQinHhRpZC2xy+hMYmyWYfbunrIeV3YGhfBNU9Z3t3s8JH+cdrhVB6QmJ89TEiURlYQh5diO4ZR3v42e//brfkQoiinCPSd9Rv9LV19HmkaUo5ZKcQXk1gzrfWTG+aXtsx/N8zT5nv/09MPRrEeapjg0e/aO8vCDe7n0lfOoliSCGFQGZUnpTa9n7h33suGW21gYKcqmtWDtdKcSfMQ/IXqzJJkjyXKctriGwY6Ok+/a61niHFSAqisYAIOxbpYIA+9Jc38HOuVw+NwacmEYt45aEjE0bw7nf+i/w6WvaD5LQkeY3JcIBw/C1V++HkRCngtk9NKzty8bfXwImMyhAclhd0YzMp+VWZmVIxbfxRpx9w+fZP+Qf+QyAspVXILOmO73vZ1Fb34127tLpB2dROWOYMgl2tpmicrIopZvEYF6RDmDyDNoNJB5RoKgIhUlIUkQbS/QeGNedDc4AUYaz+AmLFZ6NEErQEWajo4udFxhfw57+5dy/kc+Tt/v/w5UFcQSpAXt/z5L4Z47R7jz9sepJHNCvdlPvgDNrunpXrPy4xWPBCipjUZc/517QtYJatk4YD2cncxY9gcfQp1zBmNxJ84qFCna1tAuRZGjXO7pUA1oK1HhJaxG2AA4lkHeyHGpRaMoqcjjrVqHlgKti0xCkQn1VKBSpAhSEDm5ysikJVOWhrKkos6wq7O3lGBOOZmz/+JP4O2vh84EShq0xuQWpT0x2Fe+fC8b1w8R6TJxaWqk0J9nOSrX9+Udj83KrLy8xImIfftTvvTle2iEjG2jMGIygXkLWfi7H2bVu97N89UKW7RgOI6pB/hNX0csCIoLyksFzr+E8ONNCs/+Z0LLfLsTrtChU55QD/aftajQ9GYw0pIpqGvJeKWTjZllW2c3lfMv5IJP/X8k73iXhy6sdkAxshXS9AcOwGf+/ivkaQljYma1xMtInAc0EUT0dM/lkUee4YEHRnAGYl0FEsB5etWlSzjjU5+kcsG57OzuZH9JMZ4I8mafWqt/RE7sA2yKQhILRYREGgeZIZYKZS02yzFZA0MeAJP8ywk/06+d9dMzLvCqa83BkmYzlsHFizj+He/ixP/5Z3DZqyAqBYpdj07nhMJYuOeePXz9mlvo7l6CVAm5OVpUyp99OQroVy+zXvWszMpUMoUhcxorqnzr27dz+avO5eSTNBWBL8dqwEXQv4zqB/8b559xAU/936+z46FHKR04QH9JEeU5WcMSSdUEJfEgRpIoYDEYa3HF7JuQSNUah5FhvtYaS25ynLHkWY0kSdBRBNJRS8dJEWQyZiwuU+vsZM6FF7P6TVfBpRd7/PdI+3MW0l+m8KNqtQb8zV9+mwN76kSqE5xolQpm5WUhznkWTGNSlIr53D9/kxNXvo/++R550BH5BspYwvLjWPIPf0nXt77JE1/6EnrTCxyjNaVaRhQpD8JijZ80UBolZWCBdc192DL7fg+Y0DugtB9dq6UNn70qiKByQzlOUFZiMotVguE0ZySR2PkLWH7FO1ny2qvgtHM8DrgsYep1pLIBA8J3M7ywGf7qr77okdkKWlrZitCns1s/6rGxn7YctUGflVmZlSMTJ/wUeL0u+P/+6nN8+u9+nUVdYDKwKURlcHWD6OqAC8/j5DNOg1tvY88tt7Dt8UdgeARFSllJYh0ROc9YKJ3vVXLOhWC55UxYrchzQ5Zn5HlOJamClDgdkTmHKTtqeJKkussQvXOoa0n/spWccMqpVC5/FZy8FrqqYFNsybPaqYJPN2CF11P487/4Oo8+vJk8r7axvLUb88N338zKT0JCLwaaDc/t4N//7W4+9gcXoRDkKIx0KCU8cYESdP+Xt3PxBefATbfy/C13sWvTC5BlRCZFmYwYQ0KYMrIBW10rpAtMZYEgQQiBUhHgeTcbzmBKcXMM1khJWorZnWYgBLqrA7rmMHflClafdQa88jJYuQySClhBpjQyBVUthUkKnz54fjP84cf/kT27M3TsO9uPHhXw50MOe9mXXXY54Otu1vRj0hOwpvdOJ7jY0k7XOCuzMiuTJbMZuWtgVMqlr1jLn/+PN1LWEMftf2V9p7TJPem2c7BzJ6P33cfg009zYPMGBrftJk4tYqRGKYeyC81tQniAk7ZOpQJWWTiInMIJRUM56koieio0IknXonn0Hrec3hNPIjnlFDj2WD/rlniSHipVwJEFnLAICUaSZp4R7I/+5Ms88uhGXF71lXoXB6ORhymCuHVtMxr0n3aK/nDOxk/7/F6cNGmHkYFHwKe2Y62o1fbxwQ+8gfe+70ziEgiR0cTeEALjUnRm/EziSAM2bGDoqafY+eRjDD73LKXRUeTwQUS9QRxFYB3COqQt9pw3K9Y5bHAehZJYKRhNU3Qco8sJphQz1l0lXrCQRSecQN+KlSSnnwl983xJKo6g4mmMndRN2tnUeh/WAE8/YfnkJz7D8xsP0N2zmDwv8DNss2wFP7sRuifRSRkf2/+6SsfYDSLahJIHkdbf2zvvup32PXxEEXpBMDHrZ8/KrBydaCGBhCju4I7bHuZ/MsTHf/+XmBP7LLa1Di0NSihPdtIR2paWr6Bj4UI63vxmloyPwZ79sGMX7BuAwRHcgQHGh0c4MAmNLs9ztNaUSmVKUUQlKqM7u2BuL/RUYdECWNAPfd1eYZarvhap/ThdagyxikmzcTJjiEpVQDYBDtetT/nL//VvbNh8gEh1+TS8o23ks30W+2fTGP7ciChGfAvqUtAqQelO/vPz1yEV/PL7z0TKyA9SyEC3KjQu1kSxBFWC88+h+7ST6c7fBgf2w97dsH07HNiPO7CftN5gfGQMU8sYHx2d1MUuiZOEUkeVqFKi0j8POjugvw/mdMOiuf7fjirICJIwO640ttbAWYNKYgSe810JwHib9P2btvLpv/siw4NQKveSZ7QAwQSBa+LnaA+KyVb4UKs8A5b7rAGflVl5qSKEQiOxqaMqurjzlqcZG/pXfv03fok1axK0FDg/9T1xxjSKfZRiFERzofsYOH4tkIPLEViqzlGdhhe5KcWcvJA0tXYxBiedVxJSh7S9JFaahqujIkc16qSOh+3YPwDXXPMwX776BqyrENENJnQtF4A8EzAbAkrIYXAcDo+b8OOOoH+OFH6bTISvDeIk4/UUQYKTis9//ib27BnmN37zMjo6C7SGHMjQlP1nkgDmXq54ovdyyTuFa9dSgOQlQBJAilrfNfmEpvhZ4JvbirJR0bgZ6FBlOQE8uVNuM7ARWsOO5+HfPncjt9/xKEaVkEr7Rk+be7AcbCCbET/7d7dpxCcb8qMElmHaj9CC1Xx5ZytmZVZ+6lKk9DzXToyjhwcf3MTWj/0l7/ml1/HGt5yJkJKSz1wihYfssAKQCilARAA6KMEYR4EPDR6FaiaDPkNJrI0FTLiiM14gRaU1ZpfBd6/fwpe/fB3btxxE6a629Hr7sSbVzo8KEW5WfhJSpKAdEmFjrJB845rbeezRp/jl97+eK16znDjSCKfJnd8XkRIFIQUoP5I40WwIWpbiqCBXpnjPgyJY40GhPMiTQkqFacDwEFz37Qf47rfvZveOcUqlOdiCagPwVNEvneDkZ0Kmeb6OuimuHWrQCpitoM/KrEwlhYFrh3yURHEPJVllcGCY//2313DjDffz9re/lle8cjmVKjRyn/0uoB+cKNAKwzGbEXfrm8Q0UWwTgjcgzbWCZ48s73HENQ6BsY4sdygpiCLYtBnuv38D3/r6zWzfMYQ1AhV14akUcyY4FIeAXRTfM5vje7nI1HClkqTcw5YtI/z5n3+Z228/lTe++RWce0EHkYK0TdnbMDR5KLytbPv3xZjSSUh5oUHdOjCpb+vYvAluuukBbrv5QbZuOUASV9GlMqO1YeJyxwS2QMGknM7/Y81xs13uszIrPxYplIyb4ElbIxCiRJ5bIl3l6Sd2seX5L3D11V284hXncsWrLmDRYj9xo4T/twUa2gZf61RQsDRpOg5t/PFMiLY4jfBhT8oaEOfwGdVGKmjU4Z67NnHffY/z0MPPMjSYkeclqtU+MhloRKWcpCRno/CfRSlw6Z0TdHb2U6un3H7r0zzwwNOcuGYRr7nqAtauPZalSxRKeYQ4j+DWHsIVG+7IrOaRuHdZ7p3ZrVtSnn56M7fecj/PPLWVvXtTOir9dHYcQy1tEEWGeJq9d3iM+59fmfayi6W69LLLW01xph+XnoCxvXdawcW42S73WZmV6WUaYxfq3pIcRD1EvCHSFjkrVizjtDNO4PgVizl57UK6e6CjszWdFkbBAR/JFORW0JYxaxpvX6duUuHiS6HjYzB0ENY/u48tz+/hkYfXsWfXIAcGh2lnhXI2CuwCBWhNYA8rnBSRv6RI3B1W8/58dqH/9KWYGW9F2sL5knYjHQORUioJTj1tNSetWclpp69m/oKYZUuhoOgoYAmaRwwlo/afnfD7s7jNxa9d+D9nYHQEdu9K2bVzkCceW8+mDTt49rkXSFOLzTVORAihAjPgRDZH59yM2AeHQzl92Xe5S5+qGB/b/7pK59gNQocu90Aa5LvcW/KiIvQmcP5LP99ZmZWfUynS3XJSY4ts4ptbNAQkN5z2Sk5YNqzfy7PPbcG6Ot09VTo7YhYtmsv8Bf3093XQ29fF/AV9zJnTQRTrZjq1GFezodPMIUkbgoODQ+zas4+Dg6Ps3z/Arl172b1rD0MjDWpjAmcSIl1FqQRct1dyLkDCFknMJt1v0fzUzg8+qwl+FsUCtI12OQG5lUjVAVhyA/fdv5k773icUlnTP7eL7u6Y/rldHLNkHv393cxf0E+lFFOpViiVSkQ6ah7fCciNIcsy6mmDLMsYGhpiaPAge/bs4eDgCLt3DTGwf5TBwTGyhsDkEVBCUME5gZTa93FKh3VtKX8nPf/8/xsV8yOW2ZT7rMzKj0WKTtRirMvjVntMdukby5zvPrcOkDlCuJDedFgTIUzEyLDi4P6U7Vu3k5R2MTY+6huDlURriVIqRCHSG3TrQtQhA7qcwFqDyb1C7KhWaTQagETFXVQqVcZGx8lMTpSUyTJDwQHuHDg8MlgRiXsOupZSnTXmP1vSAlwJ907kocFZh2yJQWr/u9xakIpypQ/rHHv3GvbsGWH9+iGsex5nHUI6pJRIJVBSHhIRl8olsiyn0WiQZ5nHXLeOPHc4KyiVOsgziVJ9lMtV6raOQ4Zj+r2Wu8yzKDb3X+EUU/zfrAT5sfChz8qszMrhxDbzjy40rDlcMMpgnMI6hzQRpXIXQkCWZXRUunHOYoPxbhjRljb0ClWKFrCHCDl5rf17aWpJki6MzWjUG+S5AZmgFGQ292V3O9VYTOGYhHNvO/9Z+VkW6XswBCAFuTEkKgKcxzQQMUiNBEzuQmlHoqVAaOFr8dJ5/gCpgrPYSnXnw/5bhOwkijyKnJACFfvyjbOCUjnCGsPwcB2lPbKcFRZhQkQuXMsRac/Z/788afVS+dBnZVZm5WjFNuvj3iDqaShdCwYM2bSTGnxBXEmM9V3lQkJuctpFiclMGY5CyylBa+SoTQrSirgdrs6By9t+CAhxgjAe43TzN5P/5seLszlbI/9RyiFc705P+DmSEpt57R/JQH4SHEY5oV3K7zPv39nQm5GjZfvvadt6oYFTMjFNLj1lL4Autd/rtukQxxQ0u0dmzV/uNfLDShNnYppJkklyRCl3EZodZo38rMzK0cpU2OZTyUy/O9In70f5hE4+1qxh/fmU6e7rj2I/Hk6O9Dizew/gR8KHPiuzMiuzMiuzMisvfzkqgz5r/WdlVmZlVmZlVl6ectQ2ur0GMzX60KzMyqzMyqzMyqz8pGU26J6VWZmVWZmVWfk5kKM26O0Nrf+vwuvNyqzMyqzMyqy83OSIDPosH/qszMqszMqszMpPUY6AD31agz6bi5+VWZmVWZmVWfkpyrR86FPLjHb7sHzoszIrszIrszIrs/KTlWkIkV5SID5r1GdlVmZlVmZlVl4eMqNBt/yYUR1nZVZmZVZmZVZm5UciM2C5z1bRZ2VWZmVWZmVWflZk1mrPyqzMyqzMyqz8HMgM5CzTd9IJh5ylTpyVWZmVWZmVWfmJyBFZ3CNgWysMu0ZisdSRoj7ukEYiG06oGT/9UmUWXnZWZmVWZmVW/l8U5RxGpEqQ55J8olWfotP9sAZ9z76d4CSOEpgUayyY/R/AxRWHtk6KH2vjnHjZG/TDwe3MVjVmZVZm5eUqs/rr5SrSgciNQOQI0h2N2jBEA0g5DORTfuaI+NCbIlJQB5Gk24SNcWis+vHf8JevUZ/l852VWZmVn1WZ1V8vb7EIDNIFYBkxzgRD/mIi9Bapeg5yHOlSkMPgJMJJIvHjC89f9nPu0wz3HyJu9oGYlVmZlZeZzOqvl7cIizM5OGjZ4DpNR2yK+/L/B2JxRTVoBFSZAAAAAElFTkSuQmCC"
    LOGO_PAULISTAO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAlgAAADkCAYAAABE8Vr7AADrvUlEQVR4nOz9d5wlV3nnj7+fU1X33s65e0JPVo4IIYFEEDljggEbL9heG68Nu7Z37cXe73rttb94dx1319+1vf4ZHMBgMJhkAxYgkIQSKOdRGE3onp7u6ZxvqjrP749Tde/tfHumJyDqo9dVT9+uOufUqRM+54mQIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSpEiRIkWKFClSVCDnugEpzg5UtRHwl31dEpHCuWhPihQpUqRI8XxGSrB+gKCqbUAL0B7/bAJ2AN1AR/xdD2CAXfFPcO+5BfCWFVkG5teobgKYARQYAorxzylgBJiN/z0lInOn/XApUqRIkSLF8wgpwTqPoKpZoCv+bAMuBPYC++LfO4DW+GfzuWllBXlgMv5MAMeAg8DjwABwQkQmzl3zUqRIkSJFinOHlGCdA6iqj5M6bQMuiz+X4KRRySc4Zw08fZSB54BB4FHgbhzxOiEia0nMUqRIkSJFiucNUoJ1hhGTqS5gD/DC+HMAuACnxvtheQd54AngEeBO4B7gWGoDliJFihQpno/4YdnczxpiY/LtwLXA9cAVwFVAH1WbqBSwADwE3AbcAjyY2nKlSJEiRYrnC1KCdZqI7ab2AC/HEaprgEs59zZSP0gIgaeAbwJfA74vIgvntkkpUqRIkSLFqSMlWJuEqgqwE3gJ8DLgRpz9VMu5bNfzCGXgaeBLwFeB+0REz22TUqRIkSJFis0hJVh1QFUDnMrvRuC1wHU4I/UzW++ZruA0cRYGzwJwB/D3wNdTr8QUKVKkSPGDgpRgrQFVbcZJqd4AvA4XMqHxbLbBnsa9wuZfrrI5UncqdZwGngM+C/ydiDx99qpNkSJFihQpNo+UYNVAVbtxpOrtwE04b7+zZpiekJvkpUQoAiiCYXMvKyzD/HxULVhgcWGRUrkECkEQ0NTUhMa1ikBzi4+/PNb7eu1VRWR5q+wqLd3SYTYOfA74KxF5cCsLTpEiRYoUKbYKP/QES1XbgRuA9+AM1S/Y4hqW/bqsy+t8A4U8zM0tsrC4QH6xSH4xZHhononxeebnSsxMzVEqlVhYmKdYLFIs+OQXLWIMgqAo+cU85bAMCr7v09DQgGKdJMqLaGoyBAEYz6O9vZ0gI3R2NdPd00pnVyO9fa00NORobm6kubWJxsbVGh9RlW0lz25WssfTxyLwZeBPROTeLSs1RYoUKVKk2AL8UBIsVc3hvP1+HHgNcPkZrpEqw4jFSasIxhYWImZn5skv5hkammPg6CijI1OMj80wMR4yO1tgcb7E4mKJYqEMGqDWxxgfJzlyEiVrLQh4nofE9aqCiHG/qUUqHEixCiKKKtjIooARR8qMCIhFFYwHjY0+za0+bR0Z+voa6elpZffeHRw40EtHVyOtbc20tGVW74IzM9rywD8Cfygij52RGlKkSJEiRYpN4oeKYKnqRcC7gHfgQiqcm+e3MD2VZ3ZmgcOHxnnu0BDDw+MMDswzPVVgamKRUsli1QN8UBARjAExHiKCVEhaIilKPksfyanw4r+JU+sRk63qte6nqq5yb/UatRGqFtUQ1RCrzkrMYDHG0tbeQndPC3v3t7B3fy8XXdrPnr2dtLY1k2044109DfwN8L9EZPBMV5YiRYoUKVKsh+c9wYpVgK8F3ge8Gpco+RSR2BfJim8dibEIJiY/VczNFpiaynPwiSEGjg5z7PAig8fGGR+bo5C3qDWEkcUIeL6HVYlVe56zvhJXNrLc7L2mHk3atRrBiu8Ti6qsQrAMYNcgWPETisbkTCpSsaQNRkDVElmLWkXj+jwfOjoa6d/dwpXXbOfyK/dz8aU9tHecUV+BAeB/AX8pIotnsqIUKVKkSJFiLTxvCZaqXoYjVT+KC/x5uiWSkKhEqmRlJaVRYH6uwNHD4xx88jDPPTvJswfHGRmZopi3hKEB9fCMj+f7gFSIjFpHTGxMUKok6NQDNqw0Ql/1qtOrY52/WBsRRWVULUEGerfluPb6fbziVZdx2RU7aGxKUi5uuZHW3cBvici3t6rAFCmej1DVXpz96ektBKcPwQUdvl1Eps5hO34goKoe0AB4QAbXdyFQFJFSzXVNQNNZaFJJRKbPQj0/MHheEazYtup1wAeAN7Hl0dSdnRPLfPryhTLPPj3GU08OcPDxSZ59+jijo1OUSpawDIHfgOdlqozMtXYF+UkkSKq1BOv0UB/BOjtQVawNKUd5GrKGvfv7eNVrLuDlr76C/t2tq9yRSOxOuR/ywF8CvyciI6dayFYgXgzPxiK3lQhxXgtlkRXi0xTnCPE6t4qh4ynjXTj1+vmACeDFIvLcuW7I+YiYDL8KF5fxElwWkQZcKrbZ+HMSeBaX+/Ue4L/gtDdnEgL8tYj80hmu5wcK58/uexpQ1R04u6qfxg2802ImVZP0qlSlVlJlrXLi+AyPPHyEJx+f5onHjjFyYpbFxRJYD8/LYIyp3imGqvQrUfXJmuRnuarudHDeECw1gEGtM5rHRkS2RBgW6Olr5aU3Xchr33gJV1+z213Olg7OJ4D/JCJf3boiNwdVfSPwd/xgzblJYAa3YA/j1K/PAE8Ch0Qkfw7b9kMLVf0E8JYtLLKBsxzjbx08DVwhIuG5bsj5BFV9JfBTuLiM2zdxax73fs8G3i0iXzhLdf1A4AdpsV8BVb0ceD/wE8DuLSuX6gafdFAYWg4+OcjjD43x4P1HOPTMSSanFsBmEfHxPKnyqMQAXRND9MTIPKraUan54SJYMZJnk8RyTUEthFGRpmZ4ycv28673XssVV6/1Ok+ZepWBPwH+27kQY6vqfwD+59mu9wxhFhjBJeq+FbhVRE6e0xb9kEBVM8AjOOnF8xH/LCI/cq4bcb5AVa8B/jPwNiB7jpuzHk4AV6XZNpbi/Np964Sq3gD8LE603XEaJcX/lzjgptu8TdwtCwshzz59nO/fdZj77zvK4NFpFhdCjGQwJsB4joLZmlR5orWdukTuderN/CGAqFONlsMyDY3KK159ET/xUzewd39XjSBxdRXtJvE94BdF5P4taHbdUNWPAR88m3WeRRzH5Y38e+BuEYnOcXuet1DV3cDDnNa6d17jD0Tk189WZap6PW5enusF2uCyVfwvESmpqgF+EfhtTssx66zhUyLygXPdiPMNPzAEK06y/Hrg3+Dsq86I2DNfKPPUY8Pcdccz3H/vc5wYnCOfL+P5jRjx8OLQBbXxQm3Nv5cSrBT1IuGoYgw2FEqlAl29Ae/+8Wt5x3teSFNLsH4Bm8M08Gsi8rGtLHQtxGP3LlxA2+czSsCdwP+Hk0Sc603reQdVfS3wNbbWBut8wk+LyCfOVmWq+lvA75yt+jbA10TkraraBvwxTojwgwAL/KSIfPpcN+R8wyYSo5wbxEz+9cC/xxmwb3nqmnLZ8uzTw9zxnUM8cN8hBo5Okc8bjPh4fpZcQ46ENll1ki49exl0fijgCKtibQnEI5trYHY65GN/9l3uuespPvjhV/OCa3dtVXXtwF+o6lU426yFrSp4nfr2neE6zgdkcMa0rwa+oqq/LSIPn9smPe9wEc9fchXi7PzOJq4+y/WthwfiHLifZmtt7M40JoDUW3sVnLfCFlX1cSrAn8NFW9+grZt38z8+MMHddz7Hnbcf5JmnTpKf9zCmwUVAN4ntVOR+xofxqGJjtXo9qQTr9KDq+tuFUvWw1mK1SGOT8v6fvpH3vv8GPH9Le/hm4FdE5OBWFlqLWA1xK+ePIfHZwhTw/wJ/mhotbw1U9U+As+appfFc3KrSUFlvgRwFbhCRw1tU4fqtcd6Y3wWuOxv1bQCLc9J6BT94pgT/JCJvP9eNOB9x3kmwVDXASax+GSexqvfO+Ec8ewVWCww6N1vgkQeHue3bj3P/vYeYHC8gmsX3G/F9p6JSDStebxUjcTU19aQU6kyhNjSFAuIJgWmiVIj42J/fxeFDE/zbX3kT7Z0+WxQ7643A1ar6ThH5/ukUtA4uYLPkarmfg6z/563Ail6sTKnYMUGdajyxgqsDHbigr1ep6r8XkdmtaekPJ+JQHyuN27dqMIiNbVENYsUFF5aEYOlSW4gV967WkETvX5ObdMn6vAITOHu+s4VeoP8s1rceilSzjKyPmm5MIPG2tDQENGsukVpzT5w/bcn97m91r6tfrvfCHzacNwQrtlN5E04V+FpOaddcfjpy24BVGDgyzW23PMmdtz/J0SOTlAoefpAhm2lznmwVr7aQ6qhLEhdzas1JcQpY3s/iPA+NwTctfOtfnmF8Yo6P/Je3sGNnG1u0u2wHvqCq/0ZEvr4VBS7D5gPdJryehNJr/NVZHIeVNkg8k6qr9SaPGf8a6FPVn0y9jE4LjcDFK75dc6xsBs7jWWKZlRpHqN17jwCvDla9cu4mSNpmRNfbuI/WBsg8C9gB9JzF+tZDA/WQq5gIuR2qRra4vEsrbvC6LNhQ8uca4rwkJdqmV5gpnH1pilVwXhAsVX0ZTmL1dqBua+ZlCVtYvtkWCyEP3jfAt7/5KPd+/zlmJ8Wp/7wGsg0ujMJqgy8lU+crlEy2kQfuPcHv/uZX+K3ffSfbdrRsVeE7gb9T1V86A8aaF57OzbXE6kyG2a497FYPt+5bC4hosmbHzdnUPHkz8AlV/YlUknXK6Ae617tgOQnf7HgRJE4Mn2jzBC1HTHzza+jkBL4XrBSfACvTeLnfCyK0XnstTRdd7sbS+lKRBzbZ3NPFJZwne+DmIQiWaG6McC4E66FG8bI+QVMT5JzAfPX3H9NdVUoT42ih6N6picPoGMh0dWOydfmRPSAiz2zZYz3PcE4Hl6peCvwH4F9xivYpSyWg7v/T43luv+1ZvvkvD/H0wSGKJY/Ay+L7BhGDqnE5+VYKVFP8ACCXaeXJRyf56G9+kd/4nXexo3/LSFYn8FeqmhORv9qKAmM7j6tOpwxB0YkpyiOjiJY5szQLwKCeB+2tSNbH5HxMLgemkZVS4pVYR7r1FuC/xyQ29TDcPC5lA+9pQdGpGcrDJxFbpP6x4jZcmpvx9+4HPBSLxaAzcwz+2V/QPHwC8bw1Kq6GZ3awiFpGxbD3d3+Xposux40KyzphVs6YHeQauPIs13eKqJ1RArZE4bEnmPrOLcwcfISFoQlM6FFWS1NnK827d9BwxWV03PhSgn37wVsePsuFIgpHhnns134NRoYwXhbEEIURc31N3PB7f0Bu14F6GvdPW/qozzOcE4IVR17/MPDzbHAiW7OM+GwtNbKnw89Ocfstj3Hrt59m8NgcYBCvAd9XRBIp1+kRqZSGnTtUl29LEGR57KGT/NF//zq/89/fRUt7wFrJuDeJLPBnqsoWkaweNhd5eQmSMX7k619m7G8+RaN3KuqfTdZpDdbzKQWGTEsT2e4uGrf10XDgAlr3HyC46AJMRxeKj8VtxRGCFyseNsCHgHuBT57hx3g+4mLWUdQlY+X47d9g+M8+ThN2ibBpowDGUamAf8OLueL3/jieSREGQ374BLnFMjlrMCKrCrBYcWA1mCiitaOVpp3b4vYJTt246iOUgaPrNnDrsQ7BUqwqRswpHWe2co4uSRg2N8mJT/0d41/4Es3jkzSLocXEAgNVzNgI0ZMHmbnlVkY++4+0vvAael72MhquvBKzow/EkS2dneTEX36clieepMkWMWSxIpQjpfMlV5HbXteSNY9zEkixBs4qwVLVRuAncVKri065nHjIJwFBn3h0lK/90wPcfcfTTI6V8b1GPJNFjDqjXGtjw0CzpJQ627zk95RgnX2s3Bjc79lMMw98b5CP/dl3+OVffx2eDxV7kdMnWX+uqioif306BQGXcxo5MQUgDCkffJKOqUkazakt+JtCpQJBTyj2qcNopCyIMNnZSLF/G703vJyOl7+c7MWXQTZXsQmpGsCvKccywO+q6u0icuxMP8rzDOuqmp1fj1J+7hDNoyO0+5klhFedsemaKEZF2pqbcRapWjmOLgwdRxYXwJgNnACX2fGoEHS00dDbHge2WffwMw4Mrfd8WwlVbQXWj/siAGUkVGejK3b9bSN5tIpaxYDJuJ+ngUSEYCcnOPbHf0jx6zfTicH4WWeYrvFblnj+GUMj0HhynOif/4XRm79Nftd2mq69itYXXI1aw+y3biW8515yEuGZAFFLJLDYnGHXG94Cfl0KpceAx0/r4Z7nOCsEKzZgfwsu5P9pB1tMBtxDDwzwlS9+j/vuPs7sjCXwsmSCDE6NkYzyDfX+KX6AEQQNfP2fHuLCS/t427tegGI3b4K9OjLA/1XVUEROR9pyIacxzwRgfpbgxDDGeKd8ot58pVX4zseASBVvvkju4GHmH3+GyS/8I7lrX8L2d72bhuuvAS9ANFYDrW+jtQv4VVX9ZRE544/zfECsal5p4F4DAVhcwDs2QNZ4IGbpa1iuxVt2b8lkyO67oGoorx6IMnfsMLacR4xho5ANCQEThbJYwvY2gs5tNfV7xOqE5RjHpVs5W9iBS5C8KhJiWhgZ4rm//Gv82WnUE8yq4rulDpZGhUgtpc52Lv/lj+A3nvL5KrFph0KBgf/7Zyx8/Zu042OWh6pZVaoInu/j2ZDgyACl544w+eWvIgqBzZI1YDFYsSAeeWvxX3EDLddfW+8K+qU0a8P6OFsSrP+JC/u/hgJ/dSy1kHKvPCzDow8f40uf/z73fu84xaLBMxlyOR8iAyZyxutiUCJsDcFKl/LnH0SEyDbwqb+9h6tfsJvd+zvdiW5rOHUG+FNVnRCRr51iGZv3IFyGUn6WaGSEbCW/5dmFqJNNGZfPCCMQ+D656QLRN77Dsfvupfk9b6X//T8NbT1U7WzWxb/GRXw/VPulqu4EtuFc6PfiNsJenKq1mdVNChSXjLqIk4KcxCWlHgKGnieSsgx1OEuE5QVKQ0M04DmpPVpfPlJVwlwOe8GFCJ4jWCJQKpIZHCAUixhZP1QD1AguhVCU1v5+8J2NpI2N5tdozlNApKoX4RxOdgEHcOr1bpx9ZOsq983hyNk4bgw8gwtWOiQiR9Zp6U7WSTfkZG1KYWiI8je/RevcIioeuk5fJlPTj4QCIcFVl2BO19IwPq+MfvlLLPzzV2k11jkeWNz7gHXfr6oFA+JBVj0XfgNBvBJWvNgWGSIi8p3tHPjAT0O2GYvFW38Oz+FiCG4IVe3DzeW9uJA1u3HzeRuQW3b5OM4z8ThwBJf8exB4TkTK9dR3PuFsEayrWJdcrc6XK56mgKrw/bsP88XPfZ9HHxwjvxji+Rl8k4kHmEW9Mi4YcOCqM+uLxFM8DyDge1mGh+b4zCfv5iO/+RaMt6WxylqAv1bVt2w2f2EcLPfUVeHODozo5Ch2Zg7DuUiY5mIhKYpnndpIVbAiiBgyvsGbnWfxrz/B0eeOs+sj/xFvez8b6JLAkaVfinM0vhB4KXAZbmPtYWtSYVlgWFWP4FL4fC3+7ndYubCf78iwjr2qxq5/0egY0fg0niT5VeuDquK1tpHp6HKbciIpLRaxg0MkFo4b0ebEEktRSkbp3LkzriAmIGsPixfhbPP2rPec9T4OMKKqh3Cb9D5WVtu7USFWlejYUUx5gUwQoHjr8ksFVJRADCWUxr27MNnTD7pfeOoxxj/xSXKlMsYTRDW2djQYCzYJpCLVNUPVoBLbpCqAuJhmAp5qLOk3qBGwIYue0Pu+H6fp0quIKlqAdfFd4IlV+0G1CWffdhPwKtwa2M8mIgQswwQwoKr34nKe3veDklz+bBGsdZhnNXpL8qt1HqNuQ7Hw0IPH+MI/fI/7vneYUsHH91sJMlkgwkkoHcFyRN5DpWrcuYTdp8LM5xW0OmQIgiZu/c6zvPYtA1x73Z6tMXevohcXwuENIrKZVB4tOBusU4MzZKF45DhBmDzJuaBYbupEXo22SQGiilSiRQLy3/4OI36Onf/lN6CltbKpytp89+dw+dbOVIR7g5NU7AReBvw6Trq17QzVd+4gTnlXGhwmKFZT16+F5YJQVUumu4tcdw/JpqwCUX6BhZExmrVOO6L4ZYsYSp6P3bsn/n7DFO372Lp0UoKTfG0HXn6qhRjx0IFBclEEmqnMx7UrtRi1WDGU1OLv2w/BKRKseM5oqciJT/8VuZERsr4hqolfFdmQGeNhsxl0sURjyZJB8D2DGCVkmUNCHJYoEjAYPBW8KGSSiKb3voe+D3wAPAMabbRwRsBfLPcEjqMC/DjwNuAFbN0ptyv+XINzjHtOVW/BOcrccz6bGZwHMUBqt8EIRDC409OjDw/wD5++i/vuGSGfL5PNNhBkPCCMvSaSMmJj9hUGByme9xAFiogYinnli5+7h6uv2YXnJxvCllGtS4CPq+rbRSRf5z27WEcNsTHcQjp57BhEIZiqfOBcQZP/16xpiiGUgAbPY/qWWzGXXcD2n/ogVgSDouI8x2Sl/ONsS5GE5yO5ojq6pwcGsKXCRjZwS8wlVCDEotu6kA4XeNnEtxdPDKKFvPMerGfYqTjpV2TxWlvxd+xY0cYfGGjI3OEBspFgfdho3rnArIJVC5ksTbt2b75KlvbTwkP3kb/jPlrFQxRUBBWPMIyILr2Q7vf8KF7fTkonBpGHH2Ps4ceRsRM0FvME4uOLic1l4pmrCsb9VrbKnAloftfb2PWLv4JkmlBVrHgY7CqtqeAO4BtxeT4u9+jP4rJirKbG3WociD8fBG5W1f8L3Hw+2oOdBwSrFu6M88xTJ/nc39/FXXceZmHGEAQ5cpksqNbEr9pSNVCKH1SoxDZ2iucFPPj9IZ4+eJLLr9weLyohblytMtQ3P4ReB/w34FfqvP4anEfiqaNUhIGjGA1BfIxdnyzWzozTmR3r0bjVyk5UVM1imfj8P9L16teT2Z1IEhN/33TOnlGUS4RHDmGiMsb3iTbo6yQ0syqEHnTu3+2MdVSdxxyWhaFjSLnorC2sjcfEOsQNg1ogKhO0d9LQs6Em7vzF7ATRyEmCmKDEFk9rXi6x5E9U8ZpaaNm1d9NVVlRzIlAuMvzVL5OZm0XUw4QBWRHyUkZfeBV7fvs38WtjVb3rPXQMDzJ9//eY//Z3mH3qEDI9g0QlfOtEkqKCVQgDD/bsovfd76Hr7e9AGppjfzCJfbDXlFiWgf8mImVVfSlOKvwW6s6etaXw4rrfBPyTqv4PEbn3HLRjTZxXBOvk8Dxf+Nz3+cZXn2RmymJME5nAgERuwsf5Aat6h5RopQDUWYioWhbm4bbvPMzlVyZxXBIJVs3lyTey4pt68O9U9V4R+Wwd156y/VUFcwv4IyNu2G/ow1VFlSBJ1Y6xEkihSndIrlFiW6tTgduQIwFPIRge5eSt32TXT32wangDZ2aaptO/ikKBYHiEyNQr43QuaqJC6HnkYmlTdazA+MgQ07aECXxEnNfZenupqAdYQgTb003QtQWZaCphD9b6YzU4yFYiPzJCNDePxNvkRn2qcV9aFGlpwu85BWFprMYTgcLICHNPHiTblEM1QI0PVtD2dvZ9+N/h79qPjSPQCWCNwd+5l+6de+h+01spHT3MxFPPED37LN74KFiLtYrX0opcdjndL7kBv39PHL4jBDGgxjmzrH1M+1vgEVX9Pzip1WnZSi51ZDtlGFyaoZtU9Y+A/7UJLcMZxbmN5E51apRL8Af/7Z+57+5j+EE7fpBIqmI1b22y5ZqowU5srfFaka60P5yI7T5wgQEfuGeE2Z8KaW33cbbBKy9XlKhQxMv4iCdufNXnehgAf6iqD66XIiIOTXJaHoQCFKenKY5O0Cg+YLCbtMGysbGyrxrbNoaIeihCJE7lYDSZOYKK3dDbdqV0S+LTsfs+Y0MW7rkL3vkeTGv7Fmo0l+m1lvzFtUpW3xR+KFBeXGTxxEmaxMREaClkWXgEd/RwjgvlbJZo957qtQpoSO+119Hct4NA62Rt4iz2rLWYbbtO3QZpBbTS5qXylSSL7FIT/K0YBfNDo5Sm5hE1LqGtxAnZZGVXJL9bgaJCblcf0tS45mxdtX269J9eWxt7f/U/4dtS/O6cdNBr7iJ72SWgYU3EC4mVem4emEwDmYuuYMdFVzmbqnK52kDPQBC4ddDGrUnUiDXlrdLCE8BzwK2cjm0p1TUkOdIlu31S5ynS5Q6chuFVqvphEXn2dNq4FThvJFilsjI3U8LzchhJ/KVWm9GpbVWKtWDwPGFwYIaBY2Nc0b6d1aaqk14J08cOUZ4aZfsNr9rsatwP/JGqvktEwjWuacG5JJ8WymMnCOfn3ZZyCiEajCY5DyzGeiA+qCJiMUsEwG5BlRURuTcP33hER4cojo2SbW2vOpycVqmgWnOEklqi5yRzP5y0qopofJxodg6DVx8Nj422NbTkuneQ6+pmyaEWn84X3EjnCzbbs4lbxOakruu3UyqtqpbujLXPmHftieNIsRgf3TYev4nkqYSQ27UTyWROaSYl8yVobaHzxpdCnC+hCoEoqsoaYpblUUtSpHKQRDwIvFoWs7LDxB1REwK7xrN2Ab/LFvCGqlRdlny3RXgt8M+q+gERuW/rit08zoXedA24E0K6TKY4VYgYjPEoF+HxR4+7L1ezzI2HWNeuHZy8907yx57hFKb323BGlmuhmy1QES4ceY4gtE51oBvbcCbOHhWnD6OURJgLcsxnfOaCDDPZLLNBjnk/y7wELKpPwULZWhfaxMj6nw0kfZ4ITExTmpiMv1nrsFQ/FBJBGVaokEYhij/xBT/E68fikUMEpTA2tt64v6WiDFC8ri6Czjg6QuU9G+fGbSOwNrbBcp/V/nN/cwEAQgwREVDeojOxYmJpm4fiYfGwlbdtln1OexSoUn7uWTIa1oz3jcaxs3MKTYZo3wVgnLH4ah9HZaxbn2pEwrVnqAhDVOlLKh8L4Hm4UERx4NaYFJnKf7JUYpmQq/WMKzdGli0UyiSkVWt+30JcDHxeVa/b2mI3h/NGgrUEKxKHpkhRP1ThiceeA65bc6VVQJo66N5/CUc+/Sku+/X/SpxrhzoDZQL8pqreLCJHV/nbRayqn9wMImYGj2JKJfAzGNGaLaU+lKMy8rKXsOuXfsWRNAE1Tq4g+SI6PU04M8HskaOUHnmcwmOP0VAq45vk+ZP6NjcfNYowExNxCSuNaDZrOiUA5UVYmAMbB5KNU2DR0Ig2ttQXUHNpKzfZivMMcfOTp5gaPIofhS4O0iq2h2sWo5ZM/y6kpaPmHhtLRhLVXNWibz0a6yiI8wRHA0ck1onTseRB1kQEWoTZeYhiUmIVPIG2NvAy1J8ea137oprLLAvHB8nGDECIVazr1uE8Lf2mZlr6d7E+m6kpSyyJraKrrsY1RJ1dV3LAWDLEjXueigF+aGF+DmwJtAwSS6QzOWhuX/Wxk/PnVuy2m57TYRHmZjA2HmueD62d9Zpq1IM9OJL1FhFZNWbXmcZ5SLBSYpXi9CAinBjMU1iAXNPKyZqcQz2EnS++iRNf/mcm7ryVrpteF/+9bkueHcBv4OI5Lce19RWxDLWbZn6B6MhzBAYwglFDovCrF2VR2i+9iMYLqsK0lWbBShsW5uaZfuARhv/iz8k89RQ5L+ds00x5c08SqybC/EL8uxM9JWaUiRxuxZYYFglHhpkfHsKeGMKcOE5+fIri9Bz5qRkCG0K+gEauBLGxgjCTQbNZTEOWph3dZLfvJNq1j6Y9e2jcsxsaY8/x2ogdukb4CLWwOA9RCTT2qPM2CCCtgAagFs1lkFwz6xH0Su64Uh7KC2BjJdeGL1Zce4wBLwtxChYFJCygRw9jozKRCTBOIVDXWCn5QsfO7VTeRuJEVJ6HfJFEerTkZenq5MFFEjHg5SDXBmbZFrNELZ1UZ5FYBaYTI8wPDFI6egyOD1IYO8ni2Di6mIfFQpxVgErIgiiTIdPcRFNPJ8G2Huy+A2R376V5/15obkPxVvCKClVUHEkvzDobJZMwKYOeHMcfm8ZXcQeSOsa/oJhIyTRkaGkJYO7kyi4ygm3sQEym2i4rtTozDHFQUGRpcNZ8geLJQRZPTmBPnECPH6U8M0Z5dJbC9IIrolBEbZgYVzn4AeVsA7m2Zhp7Osj1dRPuPUDDvr007d0HmUYiiIlaoh7cHF1Sqv7aqMWrzb9YLlEaOc784UOYw4coDp5gYWQUXcwj+QIaxfTdGMLGJrIdLbT2b0P29hNcdCnNey9CWtpXtGrlsW1V7AH+UlXfLCIzdT/QFuH8I1gC5yKYYornEUSZnS0zPjZHf1PLij8b3AHYAqa7h64XXsXRT/0tnS98UeUUvwk+8X5V/WsRuWfZ95tXDy5TFdiZGYKxCTzf5YSrKw5RDQQo+j4N8WlaFSSO0u3qqJUoGUxLG+2vfAUNvR0M/Kf/jD12omK7sVlbAiMG/MQqxMSbhI0lcIrnthGiuXHmDx6ifM99zDx9kOjkMPnRk5AvkNEIzyqBCi0ak4tlb0YSFiGKFSVvYFoNJS8g09FJsGsnHS++nuY3vJ7s3v3JpbFUYqVjQ7Q4z8Hf+31yzx5F4tXRbKCaVXEEslgK6f3gT9P1hjeve31S5cAXv8TiV79KFh9L5CKor3VP3G4FwqhM5jWvZd/Pf6jyN+bm8EfG4jhJWv85VaEYBGR29btfnZYLo/DcJz9P6ZbvkDOeK3NJoasTLEGx5SJ6/Q1c8B8+UhUG13A3Rx60QmxlYZbZR+9n7pbbmHv8KaKRGez8Ir6WyEpETi1eFKeLNi64hAKaSLMEFoylbDwKfoDX0obp76PjumtpfdVraLz8CsCPx388t00E4mGnJzj8ex9FT0w4ihf3nZcv4g+Pu/qknq089tD1wJ+fZuijf0jZeCDVY4RVS7GziQO/9p/I7b5wVcluMjiTvtHJk8w++SSFh+5n+vEnKJ8YJ5qYgVKBQMtkUYJQaGL1OGXJW8qoRY1QMIYZlILn43d2EezfS9dLXkTLG15HZvueJW9yMxCo+Fo6SV9I4blnyd97P5P33Et07Djl0ZMEhUUCVTLubVZS9YA7eACEKJOeM23Q9l5yO/pove5qml79apoufQH4PsnY2zh4BgA3Ah8FfmlTD7UFOM8I1rnJtZbi+QXf81lYyDM3N4ezNV8JIZHKG3a88lWMfemLjH7zZvp+9H3O+Du5aGPkgI+o6ruTyMaq2sgGiXnXhFR/FGYXKI/NkMUtzrBJ+a4qpWwDunc/7sSuTrpQc2BOapPqLWQvu5K2V72CuU/9HQ2iseXL5hdcryFLUplWvgUfoXT0GWa++U0mvnsPpcNHyeQXabSWjBEaTWyML54zM4nVlTaKVqe+4ohWoBCgNFoLNsSOnaQ8OsLoIw8y9NV/Yvt7f4ye97wXcg0Vlczyg7odHaX8vftoGz6JBhZjNyLbgliDNRFFBS+TjftxvRyAAuUyxYceJvvgE/iBF0vT1n67Jsl6IhBpSOOrXrakr4vzBconp8gk8oe6pFfOg7SUa0Z3762UBUC+AI88Suaxg2S9hGAlWE/15VEIS+RecJ0TZy2XVuHokcFAaYHp229l/B+/wNzBg+QWijSHBp8AMSbePGOSkkkKcQTdkSRFrSNYvsTvPgyRqXFKkyeZefQJxr74L7S//ia2v/99+P0XUNG1xY0qz8xi73+I3OQcfkLWVVEjLrl13eoqwWIQAb8QYQ8eJqPWeezFdYVRSPnCPWQzcWSDSmbsms6JY/rlnz3IzDduYebOeygMHiNTWKQxUprx4vORgAR4eFg/UeOubiMmOFWhxO8jo9CqYEcnCEfHGL/vPk7+89fZ/oGfoOMtb4Mgu6oAa6NlUQCiPIsPPsj0N25h+q77YHSERhuS0wBjBPEyjh5XQi3V3B/bo/nWkhGhEdCpGaLJCSaffJLRr9xM5w030PGj76Lhhde4tUAtInUd/35eVb8iIt+u5+KtwnlDsKRGOJkixenAGCEsRZRKpTWvqeXxDXsupPPAAU5+7rN03/QyvO5dm63yzbhT0p3x7+2cQtoPG6/vJpbUlI+PYOYKiQ2rS0uzwa4pWk2PYW1EY28P2e4+bEyTKkSrBqHE6rqaRbVxWx8zXoAXuXo3yvG7tBFusbft7dQoZADBjo0w+pUvM/7Vr+MfH6axHNFqDJExiGdc+01iKu3UJ2I13ovWbkRVreFUaFaFKACD0h4aysdGGfvTP6c0OMDOX/qlWH20sj/1+HGaCvOYjCDGx4hxUqE1CIUgGOthbETQ20bDtq76lLhzM2RPDiIZwSQeXivuS8Q+4FlDZJzMo+wHNO91UcLj+K4Uh0exU3Mu4jpKZEA2UgSooJGQ6+4jaO9YUqtdzOONnMDPeIh4yIrCVi9cMZSCBvou2FvzCNWBJbHkonzkKYY+9tcsfPe7NM/n6fECVDJ4vkU0rNxTG0at0hm1DZWETDrLJSM+qorxMjQjMDXL/Ge/xLOPPM7+X/sI2Wte5EpQRUQpj40SzRaQjIdEFl+F0AhlD3wrG863WiQuWtYI4gUV1aKT2LmQKdGufUj3tqqkmoqM1/Hu4RHG/+FznPzGN2gcO0kuKtEsnouu7gme2tgz2J2VIkKqBg0rJZcGKs4MauP7kvnsCxkNaLCQPzTA4B/+AXNDA+z+uQ9BkFve8UtfwSp/Kj31CKNf+DKTt9xK49QUHWpQY7AmE7/HCAUsPirxjq9amYSqsbZcXDBUUTBGsSYgh+LNTJP/2teYvud7tL7lzWz/8ffi7dy1rGGs1e4M8P+o6h0isvbGsMU4j7wIqyLCFClOBRXvOYRSKSJfcPNotVG1ZN1ubqPlxTdSPjLA5He+zSmoqLPAh2p+308dyWRXomqgIsDiscN41iIidfp9CGBc3B5AbURu+w4ynV0VNZp7cKEmgA5eTbnJFIxsGatu00o2iMonlhit9VEU09JE0NEel+kW/8J993LoV3+d6f/fx2gbGKJFBM/3sKZ6yiZupqmuu0D8+7p1EqvHxG0e4myrBAExBJkM7RFM/+OXGfn8Z3AWI9VOTarKj4yQWcjjgi562Ipn8+ofV29IRETU2UK2q5t6pH2luTnKw2MYz7jo/NaA9ZZ9jPuohxUXx8wqFHNZJI5ZlayZ+aOHMaVy/MRL61fVOJjkyvES2ZDG/m1kujqovaA0N0vpxLBLdURU2aWTslRl1Q9WKWcDZHf/kppstUbyD9zGc7/xnyl97Rt0LIRkvIzz/tMIxWLFJRi3ohWZTPLxLPg2GYfx+4/ftCfxTzyMintmT2g2Po1PHuLo73yU4uMPu95xXhIUB485yZX1UfGI4vdt4tNObd1WFau1fVDtL4mpA7H6k5hcJePas4KoR8uBC8H3KmrcODsgEDJzx2088x9/hblPfoLuk2M0IRjPeQlWhVwGK25cJkkdkv5abV4gzvaxouJ1hmeIuk/S31mBzvkis5/8NNM3/5N710sGUXX5cAcoSGSutpBn/LOf5tmP/Dr5L36FjtlZcr7B+mA9jYXYzsM0Vgy6T0xyK96sJiHhNc8AeDa2pBMh5/m0TU2z8Pef4eiv/z8U7vt+ZXyFVJ0x1sCrcCEczhrOG4KlFdFtihSnCTVEEYRRbeScpVh+MO58yY1kMg3MfO2b2JmJU6n1zaqa2F1ddSoFVFzMBbDK7LEjS9zt15+sy57IQiiCv6sfsg2Vxbai7avRuFQWzqQYIiaOHMGEYQ3dW6O6VT6RKqavh1xbu3MYEJi/7RaO/tZvkXnsEZpxaj8l2YSWCymk+qn5dc06k9uooUuSuMSDehFWIlSgVQ0Tn/1HSs89i1ANYWZjo+2FgUFMaBHxajZXpXajXfpRrEQUTUTY2w2t7axqDLMMxZGTlKYWkIqNzvoPqCbewFEy3d34HY4QJSqsmWOH8RJx1nKsQc6ttYSi2N5uyDa5L2NWGw0PoaUCaizW2JjgrdcPcRUqmNZWTK87X7iNuDqCpr/7LQ7/14+Se+owLX4GPOKtN4Hbfm08UIUaiQurS1MlvtJJx1yMNx/r5o46NpbxfILnjnH0L/4cOzNakfrMHB9EbRQ7kEiVYCVjb7XPqv2ZEG6p/Fql4K5brYGWPf3ErKjSO0aVyX/+Osd+56O0HHyKJs8gXuId6OGU6q4/rLg2WqlO5CXzc435UX1/SxuuWCKxRAY849FWsIz83WewQ4PLH5Dl1XhA+fgQhz76/zLyv/83LcMnaTDubOKIncT9kFCx6jv1NH5XWm2mSQ5WSdskIeaKsVTeCUZoViXz6GMM/L8fJX/v/TXhOdZdJQ3wb1TVW++ircR5Q7BSpNgqJCfszViFB7sP0HzBAeaeeoap759SOqt24J3xv68+lQK01oZiYZZwcGAJ8dj4+FHdhMUYihmflv7tSxQ0q8IJmFBRDJby4GFKDz9KTrw4inX9y4TgTpKFnbuRti4MsHDHrQz93u/RNDqMH0gsmFmp6qEiHmPFBrEVsm1RZxMRjEwxfusdJLugO9sJ5OcJR06472y9UkwnP1E8GnftgWx9+asLh55FbEil8zeCNU7No5bGbf0EbW3V+wp57MCxWLVcP0SgmAloju2vqm1Rpg4dQqJNS3KxUZnGnh6ybR1AVZskwOID32PkD/+EpuNTeCZ2gKijyUbF5UIEQlFCEolc/c8bqeL7PuX7H2DuO7e4FoUR84ODMTHdCseqlfPLVn4qYdYn07+zeqm4uTrxjW9w9I//hNbxaQJZvvevOH6sg3Wp77rtVvFRAnzfwxw+xNTd38UF+1j5jjSO41U6dJAj/+U30Jv/hTaNXPy7DerZ6K+yfO6v8RHP4GUCcseHGPzjPyI8eiyRV2+EV7EFAaDrRUqwUjzvkBgYi6l/eJumFvSaq6BUZPqWW9Bi8VSqfnts4H7hqdzs4FYzOzlJ4/gYnkid5MKRK5VY4mCVciYgc2DvhotOsgQLQLjI6Kc+gzw3ENedJFOpv/nWGDpfdD1kGig9d5jBP/7/aBwdxzM+vrpI8ppsaMIytQaImKVqjrhtSyRtpwoB31gmH3wIFgskMgYDlBfnmBkccEazm4rFI6h6dOy/EMRf0u7VYZk67NS/ilT7Yt0anPrLhhFm23akqdO9MwGdmSI7MpxoWDbRaihmszTu3lfzrQIhi8ePx84emytUbYS/fRumrasSFcMA0fBxTvzP/01ucJiccVI7BZeuae3SALClMsVigVkpsehZSrZMVA6dgXu9zyogxtJUyDN5191gCzA7Q8O4C8WAlDlT3usChDYi29lGrru7+qUIhYOPcvxP/5S2mWl831vSF9X5sHyOrBxfAqBJrtHNEiyLRxkRJ0v0NWL68UcgKtVQlqUyxuKhJzn8Ox8l99gjNHueU2OfagcteZC1zQBWPLsIGd/He+YgJz7xt2ihUE8NrcBrtqKp9eC8MXJPkWKrYAzkGgKaGjeRh1Sg44VXU/xsA+H9D1IePEzmgoupjQxTBy4GXoKLj3UKqNazODFGNDZORuqJyV17e3zEU4vX2Y3ZkSS9Xq4aqCFWEgvWC0VGP/NZ5r76DVqdAD9WzVWtkNauOk4lrSG2qYnWq68CtYx95u8xRw+7tCHqgQoiq0kfkhqc554KhKoQlSkGhrzvY8RDNCJTLpOJFOP7LhhjnUiCrKKKGT2JTs0gjc3VmicnKU+Mxam66oclNqjfs7pzRPKklYQrC4tEg0cIYhVJPY7TRmP7ksDQ2t8fk15XcnF6Cjs6RjbJJ7eZtjc2wo7euJ1xeID5OeT4sThiQZ0SthihB617+sHzq6Q9KnHiU5/EPvE0Gd/HxqqgRN3nsbwKAXFODtOewVx6OU1XX017fx+5IGBhdJzZRx4hfPgJmsJS7HjnlER2BSWojisEfC/L/NEhyhOjyEKEjpx0EerDMgYfi6CeqTiEbAYGILSESSQUiVyML2uIrBJ1duK3tFceVct5hv72E7QcP0HGz8TJ1pP3unafJ0cCsYZSWKAY+JSDADEeai2ZckjGOim2mIR8rddyFxPOhecQPHzGBkfQchRHmago3hEgGjrBkf/+R2SfeJKcl8Fa1x4rG0S6j70mLYpXyS3s3lgSvsORcsFTR743egNWIYvH7Le/ReG1r6Lhpa/Y4A4AXq+q/1dkMy4Mp4aUYKV43qAarsaSyUA2CNyvG9xnY0+W5gMHkJZGdHSKybtvZ9sFF6JxAMQ6l9pO4PW44HangGot5eND6HwBCTaRz0wFE5sXlG2J3I6dZHsSgrXyxOu+CaFYpvDUQYa/8I/kv3ELrWGEJtI/qdpJrdvyOB5RMQrxLrqQ3P79LDzzBNO33UKroWK4opXNcIXeIU487bbJ0IZMBxlyL3gB7S96EdsPHEB8H8IipcGjjH3/XoqPPklHVNk+3b68kZRLwFfIxBvukp4ZGKRhMb9pMVlkLaaniVxnrBZbNURDYlMj6PQMjWOjjuiJEiZODOvAxrtkIRfQtGfnkiaWh0dgoYARj2gzBMtamnp7CDob0Yr8wWDnFvBOnozVNZvbg/I5j77+qn+HQck//TgzX7+ZduOhhCAx0Vb3LlaDAnM+tL377fR98Ofw26tlNgLdcxOc/PjfsPiZz5C1No7pptW4aMTENYmQLuCS7AjFsQlKE/OoH7DQ3YntaMOTMp74BFGEHhvClIvOzknrJ9tlwNu5jXJzC6EoImWsKYNmKIUGueaF0NpeeXfzDzzE/PcfpNv33dhQoZ5cDSpKFJVZzOXIXH8d3S+6nuyOHdCYg4U8+cHjjN9zF+GTB2mNk3SvLwcXVJMoViDWRwu2GhG4chWQz3P4z/8M78FHyQaBG28C4qjpGu2lEsMNLGFUYsF6hH5AZIRAI3IRZI1HYqbu0j1tfPgQNRiB3Owsk7d8i5033LgyuO1KXIZLZTa20YWni5RgpXjewUaWhsaA5tamuq5PUmD4bV0079vL3PAY+TvugXe9L3bnd2e3OvfdH8fFxto0XPlO1Vc4dtTZqUjNH+va67TijedFEdP3PYgmETOdPs7JKcIImRyjNDnG5BNPED3yEJnxcVr9BqRGOpJEzt7o4UMxGIWSl6XjNa9GGppY/O7tmKlpjKlVfVQtwpY8kChChKcQEjHf1cT2D/48XW/7EaSxbcnVjYS0/dj7GPyjP6L41ZtpwE/CB63TTK3UU61TljRp5vhxsoWYAKzSxLUQYbE9PWQ6Ole9XGJdmcV5TRVmJgjHpskZQyQupIJn2ZjLKBQbm9D+/pqGWQpHj+CdQqyyyIZkd2zDa2l398YShsLMBKWxSRqMVHe4dVSelaGpSjGXQyo2XYBGnPzOrZiZeXwCVBW7IWNXrLXYfXvZ9pPvx2vvrXjtOecIoKWLvvd/gIG77iZ65jC+CZ0nXY2eKqZcgBAZl6/SWIOWhXB+gdarr+LK//k/cLG2QDwPjjzDwEd+m6BYWJpfeQOIKnNBhh0f+jdsu/4arLWVF6/GA/UxjS1UrHJsxOQtt9I0vYD6SUVVWed6fRMRMtPXwc5/+2E6Xvt6yLUuuaIR6HjvOzn2B/+d8r98lwaTw66Zk3556Q5GaseTVMKBjH3tyxS/9Q06vAz1qR+JJVRKqBF538O7+DJaXnANLfsvwGYCpLTA5N33MnXPQ7QsLrjQDFBZx9YlWXFsmcD4TD32GD3Dx8ns3LtRk/YAPaQEK0WKzSOySltHhq7uFpYmZlkbCphcE3bPHrzvPUT47FGKx4+TvaQ1ljvUjVOUXkEliU8UMnP0GA2xLVAibUrsLNZ8BgFrLKiPZwKihx5i4r4HnAt95SqDitvQDc69O2PAMz6e3xS75Vs2n3TdULYhHNhD16tfDeU8E3d8j6aiYrywGmG82lpWLs5OvTIdePT8/C/Q/e73x98m2dYc+TQI0tzFjje+nme/fQvZRYuPqcbRWQVJDyRqIxEqbvauX0PygwN4kSLB0mffKHxMpBHlvh3Q3ulI3nIyUqu6xVI6fgS7kMeIwVInuUKwVgi6t+F1dlf6Aizzg4NIpJuO0RyKJdi5E4ImKpFJBeyJQbzFvOvnSuiMhP4vbWit6tiq4nX1YHp3EFl1bvdT0xTuu5+c8WLPQN+piNdTgcV/zuUaMX4swZWlc1AB09GJ3HAdk4VZF5xWYyIQv68kwj5A6EUYBd8airkM0tyEBFm8vr2VkSXA/MEnsKUyuknplUYWr7eN3KUXY3p2YkiyGCyL1x4TwHDiJMWHH6KhXAKCZR2w7EXWtENRFpuy7PrVX6XttW9hST/WnF1MxzY6X/8mTt5yJ1EUbXo210oCE8lg+NyzjH3672gvgufZNelVMpYTyZVVi40iZjvb6fnJ99P39h/BdPRVGwu0vO2dTPzjFxj7kz+huWCJfImdbtZfe1Uc8zPWozR0guKRo/UQrAzQDzxZT1+cDlKCdc6w8WlFxFRVKzW3qVZ/VyWOZhy5jUAqFy2tza1SFWmNizESR/VONnFxkYjXtuNwIlwkqojez0dEkaWnp4HG5vraV5EceQGZA/sp+D7MzDP3+ENkL7msvkwZWwYDM7PoyZNxwtz66xWFSNyq5it4BHieYJaYn8bvzSgGF74gcflWiEX+Gm9y68uDBCreXZ4qiwZaf+SNmN7tFEcGKGUbCPYfIPRdzj9rakhu7KbthmW1feUoInfxxfS8+e2VikQS9YOP4rkchAaiQh4vsogRQpab4q/x0jQOGprLYTJBXD/o7ALRiZMEm3zHMf2gZdd+8DKrv65YeoVa0IjFY4Ou32KrdE83lgUoThXZsn0bQVMTlcSKcwvkDx+hUdUFaE3e3UZQJWrI0rt3H1RsYJx6dv7YUcRaPPHWbZMrx0mjVC3WhjRv20ampa1CEsoTkwQnhvFEiMRW45tt0M/G84iePsLIX/wVPR/4AP6u3ZUyK0TFBGz7mQ/S8563YiJwBkfxlqy6hGCpieJx7dGPT3Zbf6yKWvIYTB8bIiwU0BXefCuRqKNVXFBf6Wwj29cb2xrGMu9Y8pMIheNEChQW80w1N8AlF0KFjEFFzVmtxXnyCk5CaCMy119P202JnXatKwGVbFICyEIBI4FLim3Xl2DF2TBJCmlqaXV5E63FGIFynhOf/zTBkUF8r6ES3HVZj9T8WyslqsJcYxM7funf0v2OH8XNYxeFX8DN56CBrne/l9lHHyC8+Tay1rjAymajEZgcBZXGUogePgwve+UG9wAuVuEZR0qwzgEkWdVJftZuDS5nnBOJW5QQa0tYDYlshOcJDbksfuDR0JDBDwzZLATZgNbWFhBoamqisTEXk6pqjg1VYX5ujlIpIgxDZmfnKRQiioWQ/GJIoWCJ4hhAnvExnlfxYnG7hO/IFc6Ksx4D2LMfPFYw4nPgwkSQVJ+cPzFAbuzfzmw2wMwtMvHIY3S/IwQ/2PD+rYMhHJ/Gnxh3i1C8+dabssO3jiZBzeax5IpkBVbcudaLpRNVQqLIuq9VESIjeLHATVAWowL+ddfQ+9Z3gEC2s4er/sfvLq12RUHxbmtqEx0bl8S4sTlJMxefYWukTcZSHj7BwD98mcxCERtkYhKTtKcqxUj6zcaG66IGq0qxpRlanWrFAmG+RGFkjECr6Vfq7XNjMnTsqXp+L+cPGtciIhB6TD19zBEi3OYcsbEEKwnREOzqh2wjSYwknVygYXISIy4gaFKO1BFao5BtwN9edYJIloupQ4NkFAymkl2g0o4aV0WX10/iPoByOSLYtRepyf9ZXCwQzebJxn1gNLZ5XO9540WnsVym9IUvceT799HwxtfR/YqXkdvVDy3tSGxnk+nogY6e+AnqPwMlMlEAFRsnKVbs0JALT+FV892tW4q4+RCKxe/tQZo642dY5eqKkFBp2rGLa//n/8GzCfFOxJgJI0tgqqwsjmOljc0Q5GJbv5r1TarVhs8dZvQLXyYb2rrGcTIWJe6HXG8XkvVjfiqUn32a+e/eRouXiTMtrDzCJK4BtVub4GwM2975Trp/5D0ke50gSUrUar1Bhu7X/wgnvnUnuQgMHqrRupPDWCeNDlEyNqI4VrfWr6PeC08HKcE6S3Bh8JIlwM00pYZIaUhk84gomaxHLhfQ1OzT0dlOT3cHnd3NbN/RQFdPlpaWZppaGmhtbaaxKaCzuwFVcbme1phLtV9bBcSikTI9VWJmZpHpyTkmJ2Y5PpDnyKExBo6NMjlRZH5+gahs8fwMRmKpWnVpOoM9dopQxXiWF1wbe3TVveq6JaNxx06kIYc3n6f89DNEk1N4vacQlP1UEK80C+OjRNMz8WZ2Jvp4yZa54q8bbvY4tZZnib39Qoo97ez++Q9iOnrcFZkm6G6qGHyvRrTdOUOqv8iyFlX+EY+5hUXKc9NM33sn45//IsHTz2I8F/ahXiWuxCEJszt3QhBUqo2Gh5y0xfjLbXs3gFL2PMye/rj8xI+thohUfjMwM40ZGXEpbcQZr294SAfUWsJshvb+fiepidlnYWoCOz5FYFZxHFivPFVKzc2wY2dNGwUW59GRYRfdfwNRUxId3GUOEMKMT/POPqhRL9lyhCmHeJ4hVAjjFD71SNkEyBiDNzhE+WN/y7HP/j0NF1xIad8+Wi/cR9tFl+Hv2ovX2gKZXFUSXfNjCetI2k3V4aD2WxYX8IdH8MS9x43U5ComjkRuKYqhe+/6QpHqOBDEz+B1dm3cCauUUfl37WKvwOI8pZkJJu64i9nP/SOZo0cIDFjdWBKZ/N2iFEVp2bUNPGfZR1Ri9Jbv4J+YxPOzdY8yYw3WWkrb2tn9I292LHy1Z6p5jobLL8Tb0UU4OI4fH6zWqy85Rxjx8MKQ0sJina07VU/vzSElWFuEesXyVouEUQnBkgkCGhoytLb6bNvRw979fezd30R7VzPdPe3s3tNFrsF3xs7rHEiTIahLvlvWvtrvxS2mxhO6uhvo6m6AA9XJbhXKpYiBozM898wADz1wgsceOcroSJ5SQcgEDW4RNXZZ6ecekY3o7g3o392GkwbGy+gGTbSJtr9rO7a5kWBimtLwSUrjozT09m7qdHzqcJtkOHQMr1DEmMwZr3GziA/S7oAgAlHEYkOGvp/7WZpedAOoEIkzv3EpMBI1iawoaIncLPlzqYQtFgnzizA9RmFslPkTIxSGhpBDT1M6dhxOjtNYjgg8qXo7bgCjscGxChDQuucAtfZtduAImUJ+w011OSK1BL1d5LraKv2zEtWzfjQ9jjc+4tQum4FAsbGBxp0uB6HEVv3lkWPI4mJFolMvrLU07dyF39pe0U8aA3ZqkszYMF68oImaWI0TSwKNXVaSgkRuI23M0rxrZ80zC2IEU1EN26qgpk4o4HkQqBIsLlJ45BHsg48y53lMNQT423vxD1yAd8XldF12OdkLLsQ0tTpPsoTkxeWs7Qbg+jIq5JkfGCIrFiM2JibrjS9Hpz2xlP0cZn/94e9qTN7qH3FWoVRESwXC/CI6PkphbIyFoRHyA4Nkho4zf+gwjE3QaCN8Y7C6dJ6tB1EnTS0HQrBnF4lcX8dGmPruHXTYRKkvq+VqriA5MFgLZQvZF12Pf+DiutZQL5cl29OFHRh1EruNQqbEAj9nNqPMzc3X8aTACsO3M4OUYG0plp6IrLVEURmrEUFgyGQ9unub2bu/kwMXdNK/q4M9e7fTv7sN3/cJsvW9jlrx7PJ661u7khJk2TcW4ozwQdbjwos7ufDiTt74thcwOZnnofsPcestT/PQfceZn7UEQXZTwTzPGMSSGIJGUcSlV/SxbXsbViNknWU1gbMWcFd6XpZMVzv26AmChTLFQ8/ScNkVZ/wRgHglDFkcOIYXBw08H2SETm2wPEWJEhKRz3l0/ey/ovNHfwyl6n2XEPlEmZDkjKvAhmipSDQ3TTg0wOzgAIWBYbzjx4mGT1IYHSNaWIRyGQlLGGvxVcnGdoKeJ2jF/mRzZCX0DA179oAL3QkYZo4PEhRLqAmodxaBi9UV9Xbitzq12Er5VdJdsXRyYgydnqrUXDfUESz6+5NfESwLx466vHJ1jJXaQ5YjWDvwGptqm8fi1BQ6OYlvpI5u0Dinnrsw35BFdu2oViKQa2rEb2uhPD2NR9U/dTPj2uXMU1R8skT4XoSxlmjBEh06TvjsMeZv+Q6zDS1kd+0ieMEVtL30lbReeiXS0byMYKxFawQ7NYWdmnEWUXU1UEAtqhGmqRWvEnNuA6iL8eRG3srsVQBEJTSMiKanKB0fZOb4AMXBATIDx7BDg8yPTSOFEAkttlTCRCGI0KSASCzZBZd3cLnacZUmiVZCHngtzTR1dFX6bPGJgwQnTqCB56TR4tofpxlcE9ZAyc/Qe8NN4Ad1tUPUi80i7MbkCvBiphqKEhmhpatuzV/doq7TQUqw6oXE+pDk1+RMlBhRqqBWiWwRJcTzLW3tjezes51LLu9i7/5eLr6kn97tTTQ15qjDhvK0sNbQrKgClk0N941Z8nstOjsbeM3rr+RVr72Sp54c5kufe4A7bn2aQt4nk2lAKccE4RwRLrGV5fulN10en2qkrv1XcPF4nFGlJdfZRR5DpBFTzz1D+3JjgzOJMGL+6Ama4hCa5wXBUkcbIjGoCGKVYpgn39FE70//DN0/8X7wcjFlSEIlSiUHnRFBtIydmWL6qadYfO4YwdEj5I8dZmHkJMHcIizm0WIZDyVAaIrVYAoYzwAeYqpxqt37jdtH/VKRkBJhaw6vtyd5OohCwsFhMpGwmSxlghMq2O390NJCrTpj5VBx86J89Ajky4jxY6eU+gaVtZaGHb347VkUly4mUMvU4eM01K0edWovsS4URtC/E/wkzlssoTo+CIsW8dzaJrpe2Un8LsGKkunrw2tfqvYK+npZ3N1PZnISzzNL1tB64Tz6XBBMwUfFONWQGDyr+OLjRxYzv0D5iScpPPkk+a98lakLL6bhja+n+9U34fX0OVMMUXRV4YVQPDaMWSzgslh6G1tvqnW2b1om091MY0+dm7vGlqHifgpuLEcTo8w+9QT5557DP3qE4sAAi0MnnFQ3X0CLRUq4xMxNaqqhFETAr67sDjaWEusauROXP75WmW97O7ne7Y4A2gJzD9xLZjEPQaZyadLu9WBCCHf10HzZhSTH2A1tYq17R5Fxyaw3qkMxsSY7JDSCn6s7uPTJei88HaQEq17EkYKdwaFL0qkaEoWCVfC8Mm0dAfsv7ObiS3u4+LJ9XHr5dlrbmsnlzh8V2unCGOWyK7Zz6eVv5b43Xc4n/+oOHntknCDwY9fecyfRCqMS23Y28oIX7ou/qbPfExWJABkfae8kim9dGB6GMITg7EwVnRgnHBquGoyeB0giuhsUtZa8DSledID+D3+ItptejZpgiTzGEZ44zYeNmD/4OAvf/hbTd3+faOgkJr+IRCEZEVpF8MXDiBB51WCHQMVua7naTpb8axN2R7HULdPbTa69rVKGnZ0lHB4mcyo2b8ajdc9+8OoIamtDZuOYVZslzoqlpb8f05jDxfsWmJ8nGh5GqDOdkhqc3M9ic420xtKwKiIWjh7BxzivR9nYPsxJGRQblmjZsROvubXWoAfa22h586uZeeIRstYlHN7sGqGxPrnat6bC06qyeEc0As8jA8hCnvChR5l47HEm/+mL9H3gx+h8zRvAb3RNW8U1eGrwGFouxv1Zx1gQRxpUFa+zA7+nZ93LE4ohohhxKZyxEXMP38/8t77NzPcfxg6P4xcWMVoiJxEtWCLjHIpMHB2/trzafy2V1S//az0QrFVKXd3Q3eNSHM3PM3voWRqMWWKSV4+dZtlazIF9+DviuG11ZkjwFaI4+faGK7jGnrPixmJDe3tddQAT9V54OkgJ1hpYOYDchLO2SDksACEtbQ3s3dfHhZe288IX7uXiy7fT2dVGsEocneSwuiS+yCnyrs0rRbYC1UbbWOVz/Q37uPTynXzsz2/nq195BAmza9kxnuGmuUqtDXnxS/vp7WuMt/s6N8zazjQBYSZDSEhWPXInx2F2BunavEHqqSAcGyc7PYFXWY7PLdFyQklLaC2l0FDqaKPlbW9i73vfh9+/q3LN8qVdBMrPHWLsHz7LzC23k5uaoEUjjPGdsY8fn4ZjA/hIqEhzKiVtoEMzSWWbQISl3LcdaeusKItKs7Pkx4bJeDYWjW08uyqJPoyhde8+wHOBMNdKbSQKpRKLg4M0GmebtCTU0AYoY2nY1Q9+K4LgAdHYONnxUXzRJWryteBkiy7ESqmlndz2ZQTLRswcPULWRuDVKvPW7g9VAyYi0ohs/x4k1xx70bgaUdj5hrdSuOt+Zr79Xdoy3pZKZitejXbp704CZWhXKB08zPHf/WPMfET7j763IoGpkAUAW6Z0YgCxIaZONXEcgokIn+yOPZDbKLCxAlFMSg2LBx9j4u//nvydd+LNzNNqfSeVSiLTirMB83D5QJenrFz+VtaKGRWLA9ZtmbEGUcUqNO3aAxl3YIiOD7J4ZICG5FCwapaC1Z+07CmdV14GQbY6pza41YYh4dyiC+dhqcTOk7VsFo2Nc3kKEQbbUHeM5+l6LzwdpARrA6gqURRibYTxLL3bmjhw0S6uv+EiXnDNbnq3ddHYVPvyEx+kOLaQOrWJi2lVjYj7g4daC47Qbf/q09Ka4Zc/8jp6elv424/fgWpj7Gl4NmFQG9HYlOUtb78eERJlSN0laKJiFI9yRwcRFg8ff3ISOzeLOUsEa25sFG92Bk9WZKqoD5WcFKulaV3jllW+s2qJVF1IAyvozn4aX3UT/W98I7nLLnOn0RUmUE5tLigT3/gKQ3/5MRqOnKBTA4xksV5IZBLiGLdw/f17CVZPQVM/BLeBBDt3Q66hUnU0OEg4M4mpUTuuDxc1SBXIZDB79kNFjrTauIt/nxhFTpyII9tvgmKoEjX4NO7YSW1nLU5MINMT+OI4zUYlahxBXa0lbG2F7ctshmZmYeQEnkksgza2EjO4sAthzqd9Zz8rDjUq0NrF7v/4nxk0MHv73TRZi2fWU4HXO3LXgBoQRazioXjG4i0WOPHXn6L5imvwL7nQrcWVmgTyebLDQ5ScqzRGbR1Ja5yktuAZOvbuYa3DUHJMcpEHfSTKM/7lL3Dirz5F64mTtACRs+QnoowV6yTAiYrWMUBHZpfXX/vYUhUM6PKL6jpnuvnetncv+LE0eWCAhoVFjNTG6qoPNvDIXnyxa8Aa83y5rCEslSlOzdOA61vdUPcfryfqEXk+3gqp7JoYrffC08EPN8GKjwSV2DHio9aFTYisEoYlMlll34WtXHblNl72ikvZe2Ab27Y3Lylm6dhZqmMWMSuuOAfipy2FSYTR4gij58H7//VLmJiY5Uufe5ggaI4feRO752lAFErleV79+su48OKeuI1SXTzrKaP2X0EWZ+0Bi9OzFBbmadzyVq8GpXxsAC1GWD9Alwmw7EYLZTyQIwuqFiNgNInwvXY/RJEjLpaYWInFtjRR3rmT4MIL6LvhRlquegF+/06qxoO6hFzFVlpIocDYpz/DyCf+ipb5BbISYEWxyRm6kjRwtTZZjFYjaGs8PyuhBFWJbBlU8SQL4mE3IAG1hExUMV6Gjn37cBTcAj7R8QGCcimWC22kbpOK2qxsQzJ9O8i01aQqSSKYr4L5oSH8kQk8q/VqS4BY7tHcTPP22sTdSun4ILJYxngeiBBtpM4jOelHNO3qw2vKLq1jegozPeVeT6zOs2Z9Ax4hQmxE2JyjJTHyXuX5g/7t7P+tjzLyla8z/vl/IHN8mEa1eD5YkUr6HLEgsV2NM3SObYmWSRXVrt4m931saxXv61bANz65oWEmv3UzvRcfIIqdLiReo6LFEtHAMIF1m7q6s8IGcGR60XjYvbvX6aP4cICgi9MMffzjTHz2s7TnC+RMjkhq6sQ9fxLIOXEsqWQiqCFRthIR32DUItZSipxqzQ9ciASJ1+iNYMViVLFehLfHSWQB5sam8UohqgEmOfItD2KdPGciPYxJvHZ0ktu2tuF/HFFxCQmJho6hcwsYSV7A0vWraoWZtBt3XaT4vZ349RGsRVKCdRaQnAi0uniHUZlMBvZf2MlVL9zFS19xARdfupPmlmpXLVfc1HHOqffCHwAItSQyUXOIwM/+/Cs5fGiURx4cJfAbY6vxU5LDbApRVKapOcs7330dxpyqpKPmjapBYyPS0mKeUrl0dghWWGT+2FE88WO7p02e4NXZjhQ62wg7WjES4iyPPNZzeyjnmrENOVq299Dc20G2fy/Z3Xtp2LkDenogzmWoWmsNJUvGs1sOPUY/8xnG/vLjdJRKiIkjpSckaPnjLJ8WKu5aUVSchMjFYXRpcMpERN2tZLb3YQ4PYxaLm9SgCpHxadi3l9iyDIDZwePkyoqaejajOJcjgAjS3YHXGBMVu7p42q0XysLDj+LPLeAHmU1JsFSVYksLssRLrUzh6GECSUIJ1COicG7vaqFl13YkV408L8D8+CTh+DTxW8DKxj64NpYWl5qbkZ2rhBZKKlCgtZ1tP/kTdN54PcNf+yrTt91G5vggGVtCjEHw4gNpWH0aoSJtSn7dGFoR5Ebi1itRQ07LLD7+OCzMIM1t7sp4rQgnJiiMjdOsjtjXY4uvIlgL2a4uMv0717zOSYZCxIac/JtPMPPJz9IZhfgSxG4hWmPrZpaYpxhiobS4PnDexVGsQUg+QhQpJWOItvXQ3NFGceAoQTEmPKYmos6abQRLhGlrorG7aku2OD2HlCzWXz9gxXJEqpjObjI1qvjl60Xl+ajOkYmHH8YUSyBSI0Fce2xrkl1Ei+R27iTX01dP84ZIbbC2Hkv3K6fjDsMSUZQnyBr6dzdww0uv4EUv3svlV+6msXmpNCpZFDeVme6HAMn62dKa4f0/9VKeevIfsSUnPdgqW4v1ag+jEm978xVcckUflYjZp1ViHFZAhEAFmZ7bmqZuAFsqsnB8kFbjTvQqcSqJTQgCyyg973svHW9/G5RKuCVsHc8dI9Da4eyiAhP7jVeXBXdedJGu1+tXD5i9526GP/kJOkoFxNTvjqfEh+LYHisMIyINCRsNYVs75e4+Og9cSNtll9J6xVVE48c58p9/iw6J1n+2ZbCqmLZWTE9fvHj7UMxTOnaMIIpcUl7ZyNPNxilYlEiVqKcDGhsrh4zVIChMjTH53dtpY/PhN9Ragm07kZaO6pelkLmBYzQmkuR6SoxVx6F4mP7dII7oJZHydWiAIF+O7YCqksP1yxTUCtltOzGrGRjLyp+ZCy9gz7//ZeyPvoOZRx9l6vt3sPDUUzSMTGBmZzFi8MQHSfizqbn/FOe2RBgDxYlJSvNzBM1t1CoJiwNHMFFUqae+KeeMwnPdPeQ6Ole9wpUjGAPT3/gG4//wWboAY4KYxK7fy0kKI9E4qIgNKWsZq4rNNVLuaKPU0UXnxRfTcfXVtF15NdHjD3Hs93+fIAmNIklL1q8otEK5u49sT3f8nYXpKcTaTbtlqCphRzvS6ojscnYmLDXKNwDzixQffsyRSBOr4Td4CYlkM2+g+YrLkMbm9W9wOERKsDaLeJt3dH/dK90iXka8kJ5tDVx7/aW84lWXcsWVu2hpS4I7rhxQm1E5nTls7ix3NqHAi168lyuv3sED90wQBNkN7zldRFHIth3NvOcnXozx6jQ22AjxkT6R1Zn5MxkypeZ9Tk6hYyfxPa/ajk2Z6ijl5hyNl14KPdurJjTrDJXKVNFaSZKtkS4l8avWUAskX+TnOfnpT9A0M4nvZeOAnuvqM5eUoqrkbYliRwfBJRfSdsmlBBdeTNvOfkx/P6alGXxnwDrxyXsw5Qg1fr1GUwCEasls7yHT3I7Ljug8CKPJcbKbCIuhYnFBgwWvrxeCYN21QRCm7vk+PPMMxo899Oqcv4Ijhm27d2NyjVS2/qlpwqHjsQTZeXlu1HhBUQtRawtNO/pZGj2rzMKxw3jWguf6wq/Lx0KwkdC6aw+maeXmtmScyNL7zK69dOzaS8cbXo+dGKd05Dnmjx6m/MxRJg4+i4yOkJmdxisUCLzAhes4RYKlYlFjKWsRSxGJR0BS3tzQcbS8fr6+VWEtpqsH09HJek4pdmaC8X/4PG0LC3gm48iVKceBeNc+JDiNuuJHUJCQfHcL3mVX0HrxJTRedAktO7YjO7YhzS3guxAF49+6GcmXIXBR1zeSXoEbo5FA1NsHbe1YFFPOE05PYDwwieF9HUiGYhhkKs4sa1HW2m8XDx0hfOxJWjKeM3CvY54YdUm8F3NNbH/xi+pqH/CQyHKXgTOD5xHBAvcy4txF6lf19qJYGxFZSxiWaG33uOqyXm56zWVc/5JL2b5jWeyMNY+jm5jcS1aW2sxXSyqJv106kMrliLAcEkUR46MhUxN5FgtTvOSlF+F55x+xghoxryfc9OoruP+eW0C2OlhuDQuoqOLzvP+nX0H/7vb4+1MzsK99A2ZxFg+LSJzGyBZPp9Hr1gluFChgh07QNLcAYvGsi/2zPBfc+gUq+cYG/O19NSqWtZP/JidjTZYySUbjapLaasdbaiPauO9n7v8+xYefoF1MTK7WQ6KUUlC36c83NtLw+rey661vIHvJJUhDc2X8JBIuAYhKzD72BI0lRbz6+0Zwiifd0YtpaYwj90N5YZrS/CwNYuqe3SqKNYKGERp7FFc2a3V5FVWSXjXI6Agjn/ks2VII6xp3L63DGTlDiKVhZ7/LjxcPVD05SXZyBqsWrIvZFNU1VkLC1nZyfdvdmIvb4ocR04eP0izipJr1dqw6r2JvRz942RqKUf3Xhtukn8P09ZPp66f7JTdBqUjf3CzhsQGmnzvEwlNPMf3UY8jx47TMWwIVnP3spuSAbpZpcgD3qeZpjZg9+iy+LSPe5rInlDzIbOsGLxM3p7pG1bo7TH3nNooHn6RZAlTjRMZ1KN1cBgKYa2mi4a1v4MAb30jmgouRhgYS1X1SG6rYxQnmHnmYrC7t+40134oVaN67G7KZ2PNdsBg8We2ItTG8/DwszldixMUVVaukhkSFEcNf+xpmbh7ruZC09ZhIGGMJCyENV11O9pJL623a7Zt7klPH84xggZswHsS6WSUiikoYU6ZvZ5aXvvxqXvnaK7j08m34/hoLwClxmGWDYZkP9soihVIpZGYmYuDYOKMnJ5ianOPY0SnGR2eYnl5gemKeQt5QLpfZscvnuhdfgOedv68sWS72HeijqTlDYTHE+GcioqojrGE5z8tfdYA3vf3yrSs6CvHmZzA4I1+1EQvTY7RufOcpIxl/s8MnMVNzEHuZbXYYqoJ0dkBnK0qExkan61GHpVQq+f8qjhmroPLXqMz4t2/DXyi6MAzu/FvfvaosBj7tH3gf237mZyvBDFfMnfigUnjmCfKPPkELBmt1UzZYKkKwexcEPsY6rWhpcZ7y3MKmJCOiQiRK4Hvo08/C3CS0JPZRzoPNvb8I5qc49jcfR594kowErJYFby0Y1BGoxixNcUiM5Ma50XHM1Dwmdk6ovrONSiwTtrdCX1/legWYnkOGxypjpV45eURE1BDQFBsXi7VgI7AlYuUYYlaRDCUHpIpbtYcJMi6EQSaL6eoh09VD7wuvBRsRzk5QfvZphj/1OaI772GzsnFRg7HqEtibDFWSD+TnyYyNbNq0QDVZiZIYEZX/LUUUMn3bd8mVQtTPVdN3aVDxMl+73TCf9en+8AfpfPePQ5wLUHVpTckBqfjoYxQPPkWjv7l9QuODU/O+PSCxG1OQI9PURKh1qotr2uIZQ2Z4hGh2Cq+lhSX9Xft88XeFxx8ivPXbNHh+zd65cZ2RwmLGp/t1r14R5HYNDAKP1P0wp4nzd7c+RYi6R7JRSDnKk20oc+XVO3jNa6/mxpdfRGd31Vx5q+yDasupsnJqJqzT1Y+NFjl6ZITjxyZ4+ukhBgZGGTs5z/ycpZAPKZfAMz6e54w9Bad6KIdl2tvbXSTk8x7Krv5O2tsDRvJbbeBenXhqLdv7c/zMh15BkAQBPY0XWiULirdYRIwhEvA9D1s6M9LkpE4LiCrlY8eQQhlp9AlNLOHUepYZpx6M1NJaUSXJOpLY00PixJ8gnJkhevRxGmPbMS8OHb32oqyYOBVQSS26dxd973mHS76skZs3ssqCbEtMfe0rmLFRkJyLKl9PwhkFG0WQCWhLUuTERUdT80SLRWczplpXcAKjQogLnr3wyLOMff6f6H7vjzk1TexlaWyB6PBzDH/i71j49u20mcBFoZIoztG4rFBZXodriYhQammlYZtLOG6txRgIB48ixTKSCQiN87Ty6+mKSGjeuRtpdOtg8h6jsWEyc7MrDoYbQcRSbGmiccc2N9zCIoc/82nsU8/gGQ8V8CizVqEumr2hHAnetdew/0ffg+I2WSdFUYzx8Nt78a/rZd+efTz9q7+KefwZgk3Y+QGUAOnoxG9odlrweIiFs4vYsWkya8UwW/vpyeCTv/teivfeTeaqKxAvcBJAI7HBfkA4cpzC4aO0EcQa7fprKauSufwyOt76TvCyaERN9NfqCqIoUsxz8p+/RtPsAuIttaPcCAqEfkDQv8+VbN0JxLY0uZA3xkNt/eu5J4bo5BhTDz9E987dVHIxLanTrRN2Zozhv/0rMuMT8YFc17HyWfo05Ugp7N9D800vQ11myI2a9l3OkgchPA8IVnWhchuStSGRXaSlTXjpTZfwujdcwxUv2E42u7Lja40cTweJFCBZrARBLQwOznH0yBBPPTnMow8/x/BQntmpkEJeMcagxtm3GOPhSYCfc/RMxNSodgRCn+bmpmp8qa1p9hmAUzq0d2RobLAuQexW2UXV1KBWCTIhH/rFN7P/wPrRk+suFGJtQUQ4PRs77AsWSyZXd/C6zcGFdXZe7eU8+RPH8Xxv5UF4ne6r9K7E+eX27EdyzZUQVWplld389FE7BMPhUez4CF4cfSySRB2zjuRMbEV133zJJUinkwCJGJSIRPYGpiLkmL/rTia+/i3akLUiIayKJAeaBhka9l5ILcEyc3kyVKV8GzkUCG5/83AbRnMIY3/5N8x8737arroc3dEHi3kWnn6W/COPkR06SaeT4xCJunx8GifLTT4rqhRHsEQJI0u5owvpcwTLiIGowMLxo3HvuDQ19a0HBrVCy+79SCZbCYwqwMzYSXRiHG+ThNzaMqWWZti+PdaQlVi4/z6av3MvgVFCI/HGt/oYFASxBqKQyBN457scSaF6QFWNvQKNRXq3k73qGqLHniHDJjSZCmWFhr0H8Fo6YvGT67fSXIHSySkaNhm/T3BEyj9xgiO//hvoJRfQ3NsHvkdklNzOnfS888dYGDiCmZ7BxUF0pEKSF4+u+RACFDWi6/IrkKbWiuxTl2Va1shJ5ma+8S2Kt99Js/Hid1G/N3UYgdfVTq4nlsTG97Xu3MaE760ZFmOdziETRYx+5u/puPxKvL37V7lEsKMnOPq//wh7x/dpNgFRTfS/2mWr+hhL95OFjEffO34Ev3+vS1fHhs/8jyJnYEFcAz+YBMtZwFHrh+CSKhfYtrOJm159PW9661XsPeA8O6rn6dV2rVNkKzW3JdNyarLAc88O8+jDx3nogWc5MbDI5OQCUeiiKhnPx5gsuQaDakSk1XjjS9qU7LAqJOl5uro7qBzYzktyVQN16TicrfRWqAiTLchlt4t0gff+xI3c9NqLtqBslsxZWy5TODlOTh0Bt6pk286QgrDmJBrlC8ycGKLTc4tjJTZb7TBdteluXIcKNuPh9+8giYtzethIrVidAoWhQaRcAlzCWDWgur5NkxUwok7FlS+AdV58cfZCaieYCBQfeZShP/4/tEzlEd8j1Aij3oYaQok3mXIU4nU0Y3p7q/0ruGCU1m3mumGgsWrbK88v0FYOie59mMXv3U9IiEXxTUCryeDiVmrV30Dda482mMNKrCIkomHndqS5yeU6RfCKJWaOH6fFcyljNghVVdtyyp7GCZmlRqhgCYdHMIUyEhtGL2lILWTpP9UqjTv6kaYmwIKxmIyPFwgZMXhGsLVrgCxvrCDWwxhLYXQUCkVoyiXNJX5JYCyCD8U59Ngxl7apjieujFVjWMwKe666wr00m4iwFDk5RJAvYtY/EyzthtirQEXxRWmemcTefR/gYlfZsMzJvh563vx2Fo4PEeUXUa/WarHOehCixTwuArxgvfgwoNXnE89j7q67GP6/f0njYj62YdrkJqEg3d1k2tsq6kaA5n37ONGQIbdQwmwkMVz2QjII+sSTPPfR32bb+3+K1osvha52KBWx45PMPvkMo5/7HDz2EE1+1hHtyuGKVTtKYvs7FSiFEd6119L1treSBD1Ts27/PgvcVk93bBV+wAhWNbloEogtikLEW2TH7mbe8JZX8Ia3XEFvX/Oyu5JJLUu+rR+1i0K82ce3nzgxw8HHh7j3e09z8LGTnBhcpJC3+H7gJFOmBT8Tn8TiE4vV2GpirYzPy1a4KIzYtr192R/PV5bljuWSxCc5raJqVpH4FFwszfOWd1zGT/7sS1Z2xel0SXyvnZ8jGp10RuaeJVJnu3PmEI/Lk+OYk2Nrnr6SvHzLmyyJMYYqtqOBlr5YohenfJEV4rCVv67E+hdI0u54h54bH0LKyV809oxavwZjXf7BwBhm77+fmZu/Stvr3wx+wBLj37k5Zu68neE//2uCwQE837j5o8m83uhJ1CU3Vou3s4tgRSqN2BXcunltNzCeXkINEhJsBN+A4JOhVjVj4xpiGhy/Q7th/2scKFMRVZp3bkeCAOIoGkzNIMMnCWKJi9H6ZlpkLWFbjqYdNYmuAbRMaWCIwK408F4+7paPz8gKLXv2YnJNgILvk+toJZSQsglYYSQXz+UlZXgWDNgnn2bmlu/Q9iPvcOO3slnGZSzOMv73n6Z8/0M0rWOLWuPbVDk7hWEJc9nFNLz4OuLBExMsS3F0gKAckvCWulLBLCGKBhGD7zvtoDPvt7Tv6MVrzGJHR1FNpKT1C08UyIjPxG130HLjbTS84pVUcpHF5JD5aSa++S3GP/5JGkeGET+JAE/tNrVhrWVVytu2QXsjLq6U63xzyaWwfy/66FNsFBJlxVgBGkyWxYce4/hT/wnTt4OwvQMNC+QmZ7Anx8hEEb7nQxzgOPFZXqunXHSsiLxa5nv7uOjnfgHT3gM2Ti2/iiqyBn8nItMbdMWW4geIYCVdLogGhFGZSOc5cGEbb3nHDbzytZfR1eXsChQn/a2qEVYK4evD0pNWUsbJ4Vkee2iQu+46yOMPjzA6OgOaxTM5PNNANmvjwHlr1bkZCaVifKWjY3PeLecOhvGTeRbniSVupyGNrayUbtMr5Bd52asu4Bf/wxvxM4lH1dbqS+3YKFIuOy8fcWodbT7zYUaj4SGy+XytLHxDKMR5zSwWS6G5lWB7EvDwLBDwuIooDCuvwh3sN6rbqcBEPVQsDVPTjPz+HzBx9110veR6pH8PRD7hoWcYv/M76P0PkS0pnm/iRbz+87kVg4clBNi5GxqWvsuop4ui70E5wnqxCfopDtmt1DuoOMVaJIamPXuclxqOapRGTuDPzi7ZQOuBxVJudx6EUHNvqczU0QFavM3HrQsBb88eFztNFSRHtncbi6GQzRiWabNWhbEgBOQWiwz86f+ie2CA1htfSqa7Dc8YCot5CgefYOLmb8DDj9FQWj/WXe25TABPYd54bHvLW/F7tsd2fgrqbL3mJ6fjrAGnZ+OquOCsiqWE0Ni/B3wPKYcQk+XkunphjJAdn2Doo/+NzEtvpf2lL4PePigvEh56monb7iJ66AmarEX8wHFH3WwtLhhp857d4GXc7IrXVb+jh56bXsvU40/RsonyVKo0OhAPv6TI0QEkGsBinbMBxLaPK5e9taItGbVYq8w1NdL3b/4NmWuvhRpp/TrzYQz4xCYeYUvwg0OwxIL6hFEJ1Rn27O/ibe+8kTe86VJa2mtOpepE8rJi4z2VDad6z9REnsceGuD27xzkiUdHGB1ZxOJjjBD4HTX2AiCx3dHmwgmuDlUlmxMam5NQEltLKLYSia3MyPA0c7MRW5P52fVhqZTnBS/q51d+/S00NgeJ+RJb1xfuRDp77AhaLCA4LYL1DNJ+hlSENbvP1NAAsjCHyGaItHOjJk6zoV190BVLJ86alQFkMxlKyYm6LiiRWEwcvlB8aFycovz1mxn5l2+h+HhW8NWSwZIxgpo4wGeNT+TGRCgJGhqhJsDr3wNxdPmkjFxPF7nOTvT4STROsSLrxCU6LYgSR1Td+FLctbapidyOXfGzuFtnR4YwczNxYukaD7YNkz1H2O4e6I7VpMn/J6awJ05s+rDiAre20LyjP7b5EUQCmi+4mLmGRiQMWS82lEkaHrc7A3RMTrP41x9n7tOfxnQ0EQQZitNz2PlFcmrxvQCDV33utZ41nlpWlTwh5obr6HjjG1EEKwZD6NSrCFExJDLCaWvVMXEQ2oiSGvz+PZBtwmY8jE0yFCRiz/rE7gZD1hPszAzRV/+Zka9/LZ4Him+VpgjEQGQ8UFOJQm82FpNWIIDxPVr27CWRUinxudYIXW94A+P/8nXss0cwdXonVqSHcTN8GxuuB4pvXWaJRBBS2wVGl/976UsJNWLK9+j5qZ+k9+1vr77ojQ+mHxORgboav4U4zwmWIfESimxEVC6ybWcTb3379bzhrZfR2+c4taojsUv3c2fFsDY2VrWVipannxrh1m89wvfvOsHxY7OAh+cb/KAJt0SVY4PdZKHYSFpWt8GEu1qVbMbQ2lZ78j4/CVbSrqefOcncYonAP1Xj8FrBtlAsLnLNtTv5rf/2Lrp7ctgzxTE1ZOG5Z7GlkktNoUquuZls42bObpuFgA2Jhk8QRNGpzUhVRKBzVz9yRsJirI+W3m3MGYNYlzvOqNlQZaUSYU0E1nNJXY1Pg3hk483DCPgIYRzrR6xxYTOcroUkYc1GM6mywOcaaF8lT1mup5vGnTsJB4dPU35RW2eSqsVVHqkTKwQVDrqxtMSoYq2l3NZMY98296W4e6MTQ3hhCbz655fTIistO3chmSxJlkeDEI1PkJmZZbNOPy6wbQONPXHctVii1n7JFYzs60efOhQHB7VxLoBYnZmo4yvP6hKEWwOGDDmxGGsraXuarIdIACZyY2AVZq0iFbsoxBKJksFQiMqUL9jNvl/6d0hnDzbJCYhXCc5qPKm0PZGc1CN5W6VHYgmShUyGzE6XuzPo78c3/mZi4lZg1BFj8Sy+ZwjAScI0lv4YJ5k0cWBOlRKKYbMxCCNj8PfsrTxGha9Eit+/i53vfR8jv/8HtFoQ40LYiJrKiFnuS5NIDiMgik0q3Vs2WHFxrlyHLFXJJ+XIkpLiYBg2YioT0Pmvf4ptP/2TqOfX+46OAn+yqQ7ZIpzHBCsmLGopRwWaWwPe/LZreOd7X8T2nQmxqvauuFweS+/fENVzXO0dI8OL3Hn7Y9x563M8c3CCxcUyxvhkMrnqfWLj++ITdaUpW+vSryjZbJbW1nb32/ID0HkCZ7uglMvKbd9+DMHfpMVBLYxLSScRoV3ghddt4z//1rvo7nFSvM14kNULBSgUCJ477PL1GoNai9/aTqahfesrrNQraLFE4egg/iYlJ4JirKIiRFFE0769kMmcaqdvEtWXkLn4Cmx3D/bkGKK2zuoNiY2YUd+pC01V9afGEMartxVLYC02KseJewPE1qPOqnozlrIBDQcOrLykqYPmG65j6t77abFmc4Fd16gzIQA2gpK15Ls76L3sUqLv3ev6x0BiQ7puOWoJu9qRnp7qGlUqkh8YxDdx3r5Eiy5suIFHkdKyZx/4fkxNHemZGxvHzM2fQhwopdDZgfb2xNLryP3cvoOON76eqWeepU3ijdZYjF2NECuYau5BxREHVfCS7anCRWOnBjGxeKV6j4pBrI8QoRLiYVnQMqX9e9n7H3+N7EVXx/20THIkQuuuXeStsx1UqZjjnZKqWFCQCK/Zp7Xf5cXrvOhKptrb0fEZxFva9o1gY7skY1euDdVsmm4eRVEZa8EL/E0N4kgj/N5ucl1dVauLhDnF/dX2trcyf/QIi//wRVpt0b1U3HxJnDZqW+jUpXGfqFOdVvtcUaK4j3XJu1cBv6LmFFAPJGLBlim2NdP1s7/A9h9/D/jZGgegDV/Wb4vIWQvNUIvzlGCJi3YblfD8Ei+/aR/v+8nrueyK2L5kQ0n7RovX0msEKJctBx8f5tZbnuSu7x5meGQGTwI8k8EPGtgqld9moQoNTUJXV7Zif3L+Qrj9O89y8OAYnp9FN+OWUynCiSNFPMphkZe98gJ+/TffTEvbmU+7E87PkX/uGI7GKRYl7GjDNJ0ZG6yka6KFOeZHRug6jSCypcDH27mTzeTm2ypk+rbT8urXsPDpT9MiHhG1x5b1kSyqq47reN30FBZsiPZvp+3C/cze/QCBiJOWbeQ+roaSifDa2zAdvSv/rJb217yO0a/dTPHQIXzjb37MLoNLTaKUVVnobGXPf/gQdnyGxVvvoDHwwMqGjhMah6fIbd8BDdVME1FhkZnBAbrMZgNJApks2QsO4Dau5KSmlAaOYcpl8DZK7VxboBJZS+vefZjm5sptNpbk97z1Hcw9+iCz376LNpNDBDy1TpqxDPVS8gqWERRBCaIQGx96LTAXKcF117HvP/5HGi68bBXLiuT4JzRc9gKi7i5Ko5N4Inh6qiTbMRNVQXNN5OIE0Jn9F9F23Uso3vwvZDcda6u2vVUpG8SeuCpEIsxpBBfvp7Org/n7HiHYxFYRqeJ1dRK0tDh5kllaDwrS0Ej/h/4tA1jGP//3NFknTVNxCae9TT1U9eLaqWZwr7bypOpI/AIhxQM72fkLv0Dba96IireZXfBLwKc307qtxHlFsEQETzxsOSIMi1x4aSc/+cGXcMMrLiIT1MaHFiex2jTZqEqskjvnZgvcc+dTfOtfnuaxh09QyCue75P1XbBGF+wuqsagOttQpaFRaW71a9p9/hEtAaYmi3zmU7cTlYI4KOrmpXlqDUoIOsePvPsqfu7Dr6Wl1dlcVLzLzsDjC0L+ySfQ6enYAFaJUIp9vdDWtLT+Lasz7qGTJ2BsnHojfCdQnHEqFrS9labubVvavrrhe/T92HsYePQxFh5/jJyJV93TeElinFqgUFKKxid46Yvo/8VfIHrmOcbvvp8cURz3Zn24hNkGmZ3Hjp6A5Ul5RfB37WHbBz7A0O//Ph2F4ibfwip1asRCVMTuv5C9v/jLtL7iRo7/3h9uUuzsDpmte/ZDkKm2aHwcGRs/hTZZbFsrjZ0dlfIBKBUoHj+CJ9ZJhurcKK110vSm3XuhknPUOfaogOnpYfd/+AhHSxHTdz1Mg9o4n12sfzrFlFarQeLwsNZGLNgIr7uHrne9nd4fex/S1U0lJc6SN1v9d2b3btre/Fbm//YTtODUjb66/WYzK5jgVLslK2T6diBJZPGMT9fP/QzPDRxCn3yKBs8/5aO6hUpoDg0jSijz2YDGV7+G/g//Avm7bie672EylWC9G8PDQyfniGam8No7a5562TtqaWb3h/8dJ/s6mPr0Z2F8hlzkVUjWKUt/E1mHuncpVgnVUtCIcmsTja95Lbv+9QfJ7N7nVLzqvH2dDd26q8xh4FdE5BSSTG4NzhuCJSKoWorFRVraM/z4e17Ie3/iJbR3ZuNTSc2CvWaPVifSGrVU/n9iaJpvf/NRbrn5CQaPzqI2g28aYg2LrRG3n32pVS1Ulbb2RkRil/tTUeSfMVSVq2Fo+cs//zqHDo4TBK24E8jmNyprS2Qay/zUz7yC9/7Ei/F9sIRLyc2Z4JZWmXzoEaJSATWJrYjQumcPeKeWuqZeRMePkysU3TtWXSLx3lhN4WJ1Ra1tBD19Z6iFG8Pv72f3b/0ah//oD5l+4CFaYyVxLc1a2oO69J81z6wopSik6Hlw0QF63v1OOt/0OqSli9mDhwit4lnjFvX15qe4lcOzPjo9zvDnP8uej/wXyOZq6o1QLJ1veT3h1AijH/8bmuZLZI1XxzirKsHd1HSR6RdyAf51N7D/Fz5E5pIr0PwkswNP02YMoTgXfrTqjl551zXlOgNtIbNn35KNMho+QcP8AstNE5aaKayEtUrY2Uamy236lalZKDIzMECL2Ni2ZyOJYEwkADwhu2fPqj1igWD3fi7877/NyX/4HJOf/yqLUydpQMnEIUI3M6OklkxXxorTKpRtSIEI291B7oYb6XvPj9Nw2WVgAio5ZaS6VtW2VlURz7DtZ36aE9PDzN98G6YUkYU42HBSIesf7OIsAAYX0LRx105MkIwzS3b/Xvb8xkcY/b3/yeLjT5Ex4Es1VduqRa5SXTLOwsiy6Bu8S/ex/T3/iq7XvB4aG1m48w6XJUFrSlbWfa++QH5oiLF//hrbfuEXED/JNFqz5qpTfWpzI30/9UFaL7uUob/+K/IPPUW2UCZrDF7dcutqs8C9GhMP4CiKKKmy2JKj4foX0Peu99L8wpdAQxZVF7w6jkCDwcb9sypZnwM+JCJHN9GkLcc5JVhLqJBVonCRq6/r5cO/9Douv6p/2XUbTcYyblonXlirJxt98rHjfPNrB7n7jsOMjMzhe1mMaQM/GYfLxM9rSK5Wi0t0JqAonV3tsQF/bIN1rgywKh2ZUN7Y0NgKf/Ox2/jG154myLTGBq26dF1bFwa1EaVSgR27s/y7X3kTL7/JBRF10aaXTfYthgI6PwUPP0iDVTzxMEaw1tK2dzfJyfxM1GuA8eEx5sshke9jtLyEl9b2X3WLqF7ga0hRodTbB13thKr452R8CP6BS7jgf3yUkS99ldnPfwWGR8iqkFU3j5YGKnQpo1WskxREio0sJWMoNTWQu/Ry2l77OrpfexNeVx+oM9q1nse8ZwlEUa1ugStaUyEegniKaMDYzbfS8vLX0HnTK+PbJCbuAp6h91/9FN6OHYx/8u8oPXGYXBQixuVVw1ZDoIoKqh5iDCERpTAkUks5myW47CK63vlO+l77RqTRxePTqWlmTp5EPBf7y7EUU33PVQZa05tCqa2J7p07a/4oTI6MsbBYJDIe/rI4erqObU9ZLaVt26GzPbZvdPVGU/Msjk5ixMOT9R1/FYhQfOt+ltua6Ovpru11oDpTVUHadrDt536R9pe/iqGvf5X83d+jfHwIr1zGIHjGjyXGNeO64s3oPp4FzxpCA6GNiFSJFMqi0Jwjs2cHwY03sOO1ryK771LI5Ja1ae35kNhkmdYO+n/9N5m+4VWMf/c28k8fpDw0jB+BRM4eTjQJtZAMLuPeY2I47wmRwDyG3K79cZwa5z2IQtNl17Dnj/43g5/9DBP/9GUaJ+YIgEAEL85J6CL8O+qgBiJ1B5UoighDp8aNmlvwrrqS7te9lo6X3ojXU5Vcl32fGSOUKg+oFVVbtW9Xwiqc/OLnaXzJ9bS+6LrYGUCX2awlpN6j4bqXc+CyK5i68x4m/+Vm5h56hGBuDlHBYPCNiYWV1fpqBFVOWi9CWS1RGLl2ZTKwZwdy4/XsvulVNF12JcRzCBtnAzAS893agbpi0Frg10Tkm2u++LOEs7ISq+rNwBvWuYJiPuL79xzk2usupanFJxnG9W2ryZkJWBLg0t1dKsEjDxzhn774KA/ce4jZmRK5TCuen0HV1lhX1Uye8wCiUCwWeN9Pv5AP/fIriTSKU2jIOeNYAFVLSKFUtPzNX97OZz91P2gDxjc1RKAOgiUuT1wUFnnJy/bw4V9+DXv2d1brSQo6w1h86F4Gf/GXaVksYzBYlOkGw96PfZzGy648Q7W6Ppw/8jTR4GEC8dfVqirLjUPcgm8VTO8OGi67AmuETVkobCksSoiglI4cZ/i221m860545mm8+cU4v56DQfFQIqMUM4Zsdzfhzt1kr76K7htvpGn/hdDSBmisfwD1DOHoIAvPPOaydK4nIa2MnWTTFyLAP3AJuZ39LN94E18mDyUaH2bsO99h9o7vEj3zFEzN4FmDiTyMKoaQUAwFY/DbWjA7tyMXX0TvjTfSds01ULH1cmk/dGGWuUcfwyuU4w1LVx3Tta9WNYQgQ+PV12Kaq2FCFo8dITzyNJ6sjJS/3qFP1UJvL82XXglxKBAVkNlZFp54CCkuxLGs1jlzqziVaxS/6Qaf1iuvxjS3r/ostTTJwUWMn3nwHkYffRh55lm8kVGiyWlMGTzruaTMJnLhBirVOslI6AleayNBbw+L2/tpv+IyOq6+moY9+6CrC/DdHImlHJvWtyfXl0pE42MUxifQyXEWR4cpTI0SFgtI4mGjkMk1k21qx/qGps42CNy6Eakld9FVZLftppLdLH43CWEpHHyC0du+w9z378EcGSCYmcezccBVY+K+duMyynr4fb2EO3bRedVVtL34peQuvACaHfmIUIQyBp/i8ACFZ57Cl4Al+1llbKyxwIglshG5y68l070d1DnPrOb4kBDKSn8t5skfOsT0Pd9l8vFH8IeGKA2fIChqRYpdG37BAtYI+B6Zrk4Wd+6k6cor6LnqKnKXXIm3rRcq0u/4GRJbSzEbCRgU+K/A757NlDhr4fwgWOoMA6up9jZrjeLuSvix82gR5maL3PPdw3ztnx7hyceHCEs+nhcs8ziszsSNMpufbRgxFPLz/OKvvZL3vO9FWI2j1dbkUjv7SHJFeYydXODP/+Q2vvONJ/D8nDtVmOrJfD2CJbHBfhiVaWxW3v3j1/CvfvrlZLK1lPpsMSxl8E//N+Ff/R0Nxgc88lGZ0uW7ufD/+wukq/cMya+Wj/TTreUsMtJV6yXORlhzwpyZYXF4hPD4URaHB7GqhGFIVjyyfha/twft66Ohowu/q7uyaSgk8WUruWBCkUqOyLXqX4la8ZAs++cSSuN+05q5NT9P6eQghZFBCidGCOcX8aIIG4WYpjayO3eT3bGLht5O6Ohg5XGwdjPbrAR2PX3UqaI6Ppz88PRUGLUqnrUIoyyrtwoLk1OUJicpTo2h4xPkxyeRcoRISLGUJ5N1tmcm00SmYxva0062vZlsW7uzpfOzS8tMVGGVJX1z/bf+7AlZeeRfywIomdtmaWagFaQvQqcmWTgxgj1+lMLoCIQhGgoS+KgvNPb2wo6dNLZ34nV2Q2MjxImCxFpn9yQAZTwbwIYJsDfY3+I+VCd+WlWDUymhEi4ieSgLhUXCiTEWpyeRkyMUTo6ipRATKaVSCZPxiAJDc08n3rZ+Gjp34HV1Q1NTpWypKd7Nxdp5tC5zVuC/ishHN+iEs4bzwwYrPtWpdSTr1NQxToCdpC349s2P8JlP3sPh5+bQqAHfb8XPAJWl5XyGANYRPjH09jXVfL+RMcCZhhP93nrLU3zi43dw5LlpgkxDnHuufh18ZC3Wlrj86l4+/Muv4fKrdgDx6WhJ0Moz/5x2cpLC3feQFUGNYKxLGmovvAjT0U3sFLalpLbqYu3eZbIRVQjXWh25og2ndhzZWlS30QjBYPGSB2hro7GtDS65mNZ632miWog7PBKXLNmz1SS5Sd8t3cDXKzLu55i0qbdynRFIUpq5Q3NzM0HzpWQOXFLT9mTuLd14okp73NtYqiyj6vWsVZf2tXrByeMFE+ferMSQqr1grX19NUjtP7Tyr8QNxdTcuuHKUiMlc3RjHfXbGo1yq6/BdHaR6ewiw4WA0rqEDNe2ZP3xYrUmdItI1RR3U5zW9bp7g2bZrbVxy5Y5AsR5JpdyqGROx/dV2maXXKF4aEcPTR09yOVXsiSc8ZKHqmmlE+ZWvPasKAZFrEtZbuMqTSX55fIS1u/LhFitd63bnSxWrAvGo14cksEguWb8nc207twHl7NK9HddVlL168RIXiC2I9MVl62DIvA7IvI/6rr6LOH8IFiwqihyc1i6mH3tnx/kqSdmyDU0YwJvScLVpW/s3G5Na0KcTUou69PYmIQJOBeSiaV1Pvv0JJ/99N3c+q3Hico5fD/rDjISSwBWjUWWGF0KViPCsERHl8d7/9WN/MiPXk9zcxI9eOPJvfVQZh98ADlynAATb8BK5AldV14FxrjFaoubI2v8f9kFdZW0dRKw04NAjYSpZrNUQBLqUfnWbfeqNXYaplpQ8lNqz6xS+VvlSXXJb8v+WP1lCcdYcvPK8ZYIP0zNN4lUfAXiE39iI8kadLf21vXIVVK/+f+3999hlhzXYTf8q+rumybnsDObASxyBpYACIIEQIIESBBgAkmQsqTXMmW/tCQ60LZe+7M+yZ/4yLJo2ZIlmpQ+mhTFTJAgQBKBBJGxiLvAAlhszrszOznd1F31/lHd9/a9c8PsYsMdoH/PszszfftWV1dXV50659Q5hXIq7F2t9uWqmcHDNy0WHSkvsnYvEgU5rb6SqLLEJ2t97Bca3L0uiJvVa7loyDghd01R8bmVX6vS4cU6FVH6YYVyil1QFyT7cA5FhChqcHRpCbLQ8NJ39C6a8moOn0sYHoJFcr0lm/FeFARmp9KzF2uLiyKzCN1jad2Kgn7QuQJnQVl64mLGgd8TQpyxcAzVaBgBCzgJ84MEPDLzmvlZG9tJmI4aBDYrPKgGFarK0ZpY3KKp6UQjoh/3Bct+L4ZF2L5thHt/9BKP/XoX05Ma22nHcrySlW9o/gsdFeBH4HbdHHYsy7vfcw53/+a7WH92l/ExMMr0ui/1ySLQ+WhA5DOMPfggTjqL9LdPK6Hw2lpovegicw+1E4ieMJUEq2XSM6uyeLitIBSFfhdVnWWKx+SiXxaXXR9R6UeNq5b9Xm0BKAAR3g9Xp0J1q1zUMFWauusWHipj8fGll7BYy19B+H2zVBE+in8V9bz4f1e68KIjJ1y3avq42gJ89cuVf1JBDBPFDinKTiv/u+RP/7+Crk2Imt85HpbyVTNaW6G/q5RQTQit+jqFpcklScrPAv9SCLFpKSefbhpLwHpTFGXkufk0szNpf8fOcsSX9SXEU5qWVl+DVbaaOZFSw4SH4+JKMdgxJ0jP59jy0iEefmALm57ewey0xLKacRyBxg8tIszyxCzgS9tb+p+5rgYxz4WX9vPJz17HxmvXIaXRXpho3qfbab8YPiC9bRtq03PELKtwLKtcnHPWER/oxwOs5dqNIupwsh5s1EFODVG71uet3EY1720O+GvgT4UQ06enPsfPchOwpoE4UEGlE/hWOczOLDA3nza5mpYxnufR3GLR0Zkq5A4znPhLVakEI04JwCKXczm4d44nHn+VJ369jX17x8jnJNKK4cRswEMrE54BEQgllSPa5PMuaJc16zu442Pv4ub3n08yZTw/dNiDo+jNeFootIHrMXrfvciZKaRM+scVnhQk33EFornDiJzyrT2MRURERCwjngf+CHgUEFrrjuP8viuEmD351VrMchKw/gH4HvANKgpYEEyDmXSezEIlX4g3of45rRhVldbQ1OTQ3BKkH4FSB9BKKtTFeirjtKtLzBzhlpmfzfPGGyNse/Ugzzy9gz07x5mZcpHSwbKSSKkJPEc0GCcA3yZYFK4CHwCbfD6DlFlWr2vnttsv46ZbLqC9IxaqXZkK+TRLL8Hlcju2kXnkMVLSAhWkFvfIt7cyeMVV4Cvgz+SWgoiI4+ek2+8iIhqJDcDXMEEvT6SjPqW1vv10RHhfDgLWfuA/YQSrK6F0s0WRorCRTmfIphWxmCScP7A8iGhjIvCzwSIEtLe3BEeN+U6E96xUc2bVi44WNEQa0gseu3aMsn/fCC9vPsDrr+5nbCTHwrxAWMIkto47BOkstL8dF0RJcEQROCgL0EqQy+WxnSwbzu/gfR+4mBtuOpeOzvjiugT1PBkOAyeABqTyOHLvj4kdHcO2HTwBttbkPI/4BRtIrNmAafGIiHqczP7bKMLRW6mMiOXDkp53s//vRKkiQ5x8Gl3A+jHwH4QQrwNorQeouUfEPJyD+6d8wap8u+9ywQggSivaOppKPjHhEEJmroq3VjyYmdPs2TfG5MQUO7Yf5fVX93Po4BgTxzzmZjy01liWZYSqmG02Ktcw15W40QqN65q8kc0tNpdvXMlNt1zIxnespqXVKZynlUZW2HJ8phBA+pWXmX7wATosiefHFVPKYyFm033DDZBMFc6NiIiIiHjLcNo0LY0qYI1horF+VQiRDx0fqnI+4alwamIBY2ILogEvF9MgmC2qJi6AAHr724GQVFm+w0TBzLRienqW6alZpiYXOHo4zd7dhzlyZIzJsSxjE1nmZnJkcx7SAstysEScmCP8a3l+1ODK8QhKTIvapG3Iuy6IBQaH2tj4jou5+ZYL2XB+P7bfo7SCYNv4mY7SVI5Ip9n/jX8gNj6Jtm2EEoDClS6sGqbjmmvPdBUjIiIiIpY5jShgPYjJI7Slwmc1BKwiExNTSCExfkNLDqHXQPiu4xrSc5JXtx4il80xMZ5majLL6OgYc7NppqfmmZ7MMDOtmZvNsLCQI7OgUEpiSQut/dxiQmBZSVJJjNAjNKggYJ5JUSr8y+qwCdJXlQU5YD2VQ+kMLa02l52/iuvedTYbr11Df39Lodae1mb3YEHVJRqu2acffwz3yadJWXG0cBF4CK1JC2i+4Z1Y/SvqFxIREREREVGDRhKwZoAvAf9dCJGuck4NAau4P25yYg50HIENwqOYcLPBZvqqGNOmHYvxkx+9wg+/9yL5XM6EMyhEBg72/pkQByaZbgzH8V3h/dxNQorFWjwtwkVQdJgXIPIEMbC0grybxXWztHYkWb22jWvfeRXvuHYNa9b2FxLD6lBRlgir2BpHuAruXh05xOGvf5XmTBrbskCbBLyu5+H2dtNzywcKoScape4REREREcuPRhGwngK+sIRgYYPVP/IDGSiL0ZEJcvk80jKCh0CANkmgTZDc5TFzCiFQHkjhEI/HQp9UN3kW9HX1btEXOLUCISxMmgWF8gSeyoPI0dnVzKo1fVy58RwuvryPszesIOb3mHDCoYZvzSBjRT7PoW9+Hfv17diW357aCKDzWpJ6z7uIrz+bkxEOIyIiIiLi7c2ZFrAywF8A/1UIMVXrRK21A/TULs6Y1d5762Ukm3Zx+OA8czNZ5uayftRZo3KxLQsRqF9qCStneH4NR8pdqh9ZYZNfPSFS+ZnfvQxau3jKpbk5SXdPK+ect5orrh7m3AtWMLyqk1ACefPV0LWWiwgigJlHHmbm3p/QVtbtPaXI9XWz+sN3gHAwGs/Ts38w2IYREREREfHmaLQ56UwKWJsxvlYPLfH8VqCp+sfG30paijs/cTV3fPxqJscz7N51hP17x9n++iTbtx3k2LE5Muks2UwOIWykdAr+WsbUBotDHYQf2Zs5fjJZHI6hcLmQ/5MOwiugQCuUMsKUlJKmphjtXTHWrl/JhRcPse6sXtafM0Rra7FbGBd45UfAsspicQX1aKQuXQEJ7u7t7P/qV0ilswjLMc1kAoSxgEfHu28gcda5xbSup8xEWKodWwatFxERERFxApwJAcsF/hb4/wohjh3H97qpG78itNtNQGd3gs7uNVxx9Ro0kMu6jI6k2bv7EPv3HmPvrhn27xthdHSazHyeTNbDzXlYlu/4LSykjIOQCG2ZHXfSNR7fgd+S9gUd4RX9vLQs1qXgNF5NIFH+OWX3EST3KxD+viRIjSnC2e01KOWihcLzXJQy0eDj8RiJJCRTgoHBbtasG+Tss1tYu34Fq9Z009TiVFXYyML/OnSsgcWCkGBUMJfOHGP3X/41sZ37sS2BRiO1SZOqyOH29dLz4bvA8htBnarQ7SYdkVaSvXsmSS/k634jIiIiIqIGwrj+xGOC1Wu7sO0zbZgrcrpqEkxXbwD/Tgjx4xMoowNoWfqlDGFzVjxuM7yyheGVG4ANZpdeJk9mPsfevWMcPjzK6JEp9u+dY3RklsmJeaYm58l7EuVKPFeglIcQ+P5dAoHj+3lpP5Cpn1dPB+qw8sjrYe8lH10u3QhKBDFdqqvSGrRyUbig8iitsS0LyxJYjkdTc5zO7k56eltYu76V3v52Vq7uZ+XqTpqaY8RiNsWwVH4mwnCOzYrmxQYVqMoJNbmQgJdj5O//D/qJTSSlje1/pgVYWjAvHHo+cgeJs1cXZbNTZh1UgMVrL4/w/3zxO+SyshDKIiIiIiLiOPH9eFzXo39FnC//r8/Q2fVmYpCeXE6XgKUwaW7+tRDiwAmW0YbJQ3hcVEwmo33xRUAq6ZBKOnR2N3EZq/wTIJ3Ok8vnmZvNcPDANCNHppgcn2V6aoGxsRyTE9PMz2eYm3GZm0vjuZp83kMKB8/zEEKitUb50eOlEHhK+2EPyoWV4iRr2RZaKT+0goXWyghDwkR2j8fjpJJJmtvitLYl6Oxy6O6x6exqo7u3jf7+LlasaPW1VjEsp/zuFZq8H4g10JRZ/u+Ns+vvTSE0WgiEVkzc80PGfvgD2shjYSOVRgnQQpB3PWIXXUTPnR9G+RmdT21qHImb1/zwu5uYGMthW8nT5usVERER8dZDgrbxlCbvChptuXq6BKx/C7wmhPDeRBl9b74aIQ1TxSk0iP8kSKYckji0taVYMdRZcpbnGXNj3nXJ5fLkcy65jGJ0NE06nWNyYhLXVczNzZFeSCOljVKKmZkZE4BTlyehNmqVRDJJMpEEYaKrt7e3gYCOzhTJphj9/a0kkzbSsknEHZy4QzwuC+ESKqG1hqJnkX/fVR57YOJc7kKWNsLozKO/4MjffoWmhawx93qgpWtkSW0z3eyw6u6PIroGjCLplEZnMCVvfn4/Tzz2Bo4TR2uFsCIBKyIiIuKE0AAeUmikjJ35nWllnBYBSwjxykko5iREf6w3mS1+OFobbUj4DMuCZMomiU047/SqtW++hidKoJEKAoYWXMBEKMbVUmis/rkENCiBlqYNjGZSMvfSMxz8i7+kaXwa7TgoDUpqlLCQnmZW5ej84Edov+nGkE/dSb59v1iFQiDIZfN86xuPk8tY2DFhns0i83BERERExNIIBRIvydPbGDSON1h9ToIG6wQo80dqpG31YVfzknQ0i6SEZSc1HQdGuPIIvNwEcy89z74//lOa9o3g2HE8BVIHab8FOa3wzjuL/n/yGbAcNKosJMZJq5qPRiB5+Bevs/nFgzjxJt/nLRKuIiIiIt48wWazM12PUpaTgLWkNDknmwZ7XiU0ct1OJ4H2zkIw/9IW9v/xn9C0ew9Jy8HTRotnLJ8a5XrMdDaz+l98Dqd/FYo8xfATJx9jnJWMHUvznW88j9ApP8C9jXmCpRsetG4U8T0iIiIi4s2wnASs/jN14UiQaTAqOErZwMKLm9n/J39Mcs8Bo7nSEtfXQFraxAKbjWn677qD1muv97WREivsonaSEX5lv/etZzmwd4ZYPFbcYFD4FwlVEREREW81loWA5Udx7z3T9YhoAArClR9TCm3Mgi88y/4//RJNe/bjSButBUoEAg5IrZjVWZLvuYHuT/2GH7Edjss/7ThQgNYelrB45eUj/OzeLUjb8RNnl2uuIgErIiIi4q3GshCwMEFGU2e6EhENQGGnn+f/lMw+8jB7vvzntB08QlzG/E80QhvfNC0grbOoi89n1ed/D9ncYcpSxiH+VOwc1GikEOSyLl//6qPMzno4toMm5JQZEREREfGWZbkIWL0sKchoxNuBgseVt8DEPT/m6N9+lebxMRzLRumw678RtDxPsTA8xJrf+32sletAe2g/VMapivNpUowLfvi9F3l+014cu51IqIqIiIh4+7BcBKwOIHmmKxHRGEhAz0xy9JtfZ/S736VtLo0jHTQW2g/cGaQQ0spjrqOdVX/w+zRdegUuri/8nFoEsP31Eb7zraeRoonI1yoiIiLi7cVyEbC6gUUxySPeRhRSMWrye7ey96//N9lfPUG7kFgi5p9jUs8IX4BSnsdMIsHA5/4ZLTe8GxPlVpoo76eskqbkXNbj7/72CaYm8jhWEngzMXYjIiIiIpYby0XA6jnTFYg4/QQx6INQqXg5Zp55hEN//T+IbTtIi2UhlTlTA0KY5NdSa7R2mU3EGfid36Hjwx8B4ZRGujrpElY45yR85x+e4vln9uHIhDkeRMmPiIiIiHhbsFwErJMQxT1iWVDQVJWmyNbTY4x865tMfP+HJKbnsKVdCMAe5JYsflkxk3Do+53fpePuu9GOfVpDbWx+7iDf/YfnASfkDhYJVxERERFvJ5aLgHXGYmBFnF4COaSgtQIWXnyGfV/7GvLZl2hSAiEtNCC1Oack4IFSzMTj9H3uc3R8+rN4tjiN8dIF46ML/NVf/Jz0nMS2bVTkdxUR8TYkvNEm4u1KwwtY2gQNijRYb2kqB0pQ4+Mc/cF3Gf/+90kcmyTu2GhLIj1V+I4MfhOglGa6Jc7gP/1ndHzqblxLIMIqsVOKwPM0X/mrh9i5fQonlkRptSjVUsTblfKJNuoXb23KBazTNQ5FNBINL2BhlBSRgPVWpMKYIwC8LNOPPsrRv/861quv02EJpGNMggrlyywCLSR5AUJplNLMtrcx+IXfofO2j6CF5WuuTuWAFniImWv84DubePDnr2LHUmgsEBoVuryMFrNvY6KJ9a1LODRMgKpwTsTbjeUgYCWAzjNdiYhTgAhkLFEIzp7etpXD3/4W7q8eJzaTJubEEMIlcLgqCClCIJXpwFntMTvcx/Dvf562G25CC+c0KOYDBzABAp55cjff+NoTSJFEINFaNmR294jTSVgzGzJUa33qArBFnAGCDS5QFLZKHRNMkJbweRFvB5aDgNWLEbIi3oIE04+7Zy9H7/8Rsz/9Gc7oOM3CxrIlSvjhDYRAiKIruwBcFAvCw7noXNZ+8V+ROvcSVOlewVNfcwEH9ozxV3/+MxZmY9i244+hp2PFWrxGtflaF9IKndrrVL5unVKrlqWXXEaxrPonn5y2qFWHwm9obbSqaIUWGq09irHQBAKJlOZfsW6nsl5Lu4Au3ES5d+NxXTF03ROh9Isnp21q9+E32zeE9teAKLTWCCHwPBetTRsKIVBaY1sxRGEs0ydwb8f/LkL9NgyXVXru8bfJUt5Fc523vrC5HASsfqI0OW9RNOroIY498CDT99yD2reflJRYlg2BORCFDuULNHOFRruKWcciccP1rP6Xn0esGMZMYacwc3MFm+bEWIYv/cl9HNg/RyzWAlqY+p5y7ZUmm5tHK+PnVVXAAqS0icWOY40ilD/KCrR2yeayFHSNSxWwlnKZKkdt2xdUl4jr5vHcXN2ragGxWAIpbN6cAFFesAYhQYOnPFwvgxMTpJrjNDcn6exqJtXkoLUmm8syM73A/Gye2dkM2ayHwMay46fMXS/vZnHdXG0ZS5u0UbF4HCHs4sHjXCh4Xp58Puv3yePs/xqEFEhhIaWNFJbxYRQnEqQ3+I4il8uhtFv8pEJRCnBsB9uJme/pSmNI0BaBsCwKz9xzM9gOJJKCRDJGR3sbqVSSWNyUNz8/x9zcAlOTWTILilxOYdtxhBV4kdbCP0crsvmFotB2PAJWnc/DTzksbFqWg2MHMb7rX9Dzcri5fCglWIVr+R3djjlYMla3zOXMchCwuoD4ma5ExPETdu8U6OLKRoB35AAjv7if+fsfRO05QFwpHBlDB1oqAWiJLom6rpBak/U8cu0tdH/mbvo+8Wl0qgVVduapwQOKdcpm4K/++0NsfXmceKLVt2JqBAolAk1FuEFOhsClQVtYdp47PrKB5nYbVHWBUkjN6JEsv/7VQbQKTBf1BAtfA4GmtR1ufv+F2I4oVl/Xv5V6Im5lY4lAWoLXtkzy8ubDCOlQaYIPJhaBxNN5zr+wh0uv7EUpD62qV0wrwS8f3MXEuIcQQQ1PJBOlaUNRSDiucPMKIWFwRTuXXLmWiy9dwYbz1tDR5WA7lr9o0Cit8TyXbFqxe9c421/bxzNP7ebVrWPk88KfzN6cKcnveggkrkpz8eW9nHdRZ927dPMWD/x8B7NTXkGztlSkkLh5j3VndXL1db2+mfz4BFitNVNTeSbGcowfm2fs2AyzMznyeYUlHSw76WsCl9I+RsCSQnHje89icCiO51Wvj7QEWzeP8MrL41iWQ1FAW1RLtM4jsHHzLgiP/sE2Lr/qbDZc0Mn6s/sZGurCsm1syyoIE0opPM9j/FiW11/fz6ubj/DEY9sZG5vDshJIbF+QrFJHbROP57n9YxeQSIiS928pw0ot/09NBQ2X0AgLdm+f57mnRwDbuGpUINCQuirN2nVtbHznSrSqHlhZCJNL46kn97N35wLSOl1Wh9PPchCwejl1KomIk0lIwaNK/zSihvbI7d3PxEM/Y/ZXD5HZvZ8mV5MUNsKyFg1nWgRJbRRaeKA95pWHPudsBv/5P6f52uvR0qGCKHOKbk4WrqQ1/O+/foRfPfg68XjSn2pP/dZsrQVuPs9lVw3yB1/8IGIJY9OhAzM89+z/YWYKpKWqrM7DFzErZtfNs3JVN7/7e7ecjKovmf/55w/z4vN7cGK1tVgaRTLl8fkv3MyGC5YWi/jY6M946IFtxByHE39OgVBmhCUhXS65YoD3f/BiNl6znraOylm9iqKcTVMKOrtSXHHVMB/55Eae27SHb3/jWV59eQRLpBCyxmS7pCoKXNejrcvhC//uVlataV/S1/btm+Dpx44Sq9P2pQiU0rhulo3XDvN//fMbTqTGBVxXk8nkmJ3J8ca2Q2x5fg9PP3mIwwensW0HaQHaorZgrFBKkWqR/MZvv4fV65vqXvd//cWjbH5p1AjDVZFoJfH0PBsu7OfW2y/hmuvOoaunXiY3CdikmuIMr76A977/Aj7xmY38/L7N/ORHW5iZzGPbsQpynQY8PA96+lJ87vPvw4mdPof573/neZ56fF9Ii1UZjcayFHf/1vW8++azllS2tJ9gx+ubiFtv3Sx4y0HAimJgLUNK3DyzWeZf3czkA/cx/8Qm9MgYTVrTJCy0XTT/VRo2POULaJ5iLpYkftstrPytf4I1vOqkWXiWju8vo+Db33iGe773PJbloLVE+87upxqlPeIpzSc/ezXCAs8XZYVefPkgfEV7Z4rOziST43nkUiZubYHwcN08q9f2hUorL/1kI8lmFXv3TCItUV2J4JPL5bj51nPZcF4PRa1GrYcgWHd2N7/4mWsSb51AdP3AjOd5Gi+fYfX6Tj7zm1dy/XvOJRavJu0GWzkWHwOIxy2uu/4sLrlkDd/8+yf5wXc3oVXC1+CdeDsrneO9H7goJFxVbp+g5SQwNNSOEEeMkC2gbl/x21AATkzTO2CuZQz7J/ZC2LaguTlOc3OcgcEN3PCeDdx9bIEHfraF7//jc0xOprGtljqlGA1aWwf0r4hTux0Fbl6zd+84gK8F1YtMtlqD62ZpaVPcdfc7ufNjV5FqLhdEa/dBjfKXaILBoRZ++3PvZOM1Z/PfvvQzdm+fxbKSCJkPVS3QkroMDA5gUleUL5BO7rtoriDRGnbtGjXVkKriu2KsiRrPy/OO69dzzfXra9Sr9Ptr13Vi+dqrgn/tSdHyNw6RgBVRndBYsRTPpnCQUAG4oyNMb3qMmYcfIbN5C3p6hmbLxhY2CIkWJsxBrXlOADlPoVauoO83PkvHh+4A2ymM+yfinbF0woNlsYI//M4m/u4rjyBEkxkYAo3QqUiHIwJ/EN/Qql3ecd0aLrp0iLzyfIGp+pPRaJqaLAaHW9i5fRRYis+DKVMIyao1naFj5eecfObmFjh8aBxBvGZbesqjowc+/smr/dtf7B9XiaGVzSQSMZbea4JObSYYrY25x1UL3HjL2fyL37+Z7l7jIhqYwhfXo1Kdij5C5p9Nc6vN7/7eu2hqFfyf//0UWrcUJp6l19W8U57y6O6L86E7r6hTj1LWnbUSwWvmDz+351LaSqNw4po1a1b4V6r2vUpa3nqaX0lXT4pP/cY7uGrjOv70j+9h57Z5HCfhP5cK/V9LlFL0D3QRT9TTdsFCOs2hw4eQMnxuqbCUdzP0rbD59//pDi67fCVAwTUh5Hpe8zpB2ONi6yjOv6iPL/7hbfyHf/19pibK2sDfjeypDEPDnTh2JSH+1LyLC+k8B/aNGZ9FXdlXzCziBLGE5sMfvYh4XBSc/OvVq6urk2QyTs71EKdtc9LpJRKwIiqjS381imqwfFEriABFyPdJAMxNM/v6NtKPPc7kM0+R3buHprxLq5Ro20FIaRwgtS44rAs0ljLDlCcsJJ5x6NQes6kYLbfcxoq7fwN75Wp/BedXTprrnjr7cSBWFgeKX9y7la/9zRNo1Yxl2SjhT6la+37tJ1Pc8wU2LUHmUErTlLK46+5rsCyzylyk/QuNaWYKNwdWDHeBOIDGqT8cCw+tFfGEpqvHaApKI36dXMIalLmZHDNTOX+iCxPymRIKpbPcdMtFrF7f7k+DS+sF3T0dxJOS9JyHlAqxFBurj9ASpV0sZ45P330Nn/mtdxKLCd+ht9RXsNhatTVqi3qvgM/85vWMHp3l3h9uIR5rB7TZcVXTtCsR2vevUxIvn+b2O9/BiuE2TD+CpXjRr1zV7vvbacRSNmoYZ0mU8kg1a1asDPYjVbuW8OuiARewKG0DHSo3tGLDvFrrz+nl//NfPsZ/+uIP2LtjATvmlCxshA7MVR65nMuqlQN1hFTz7GamM0yN5RAiVlp3oZBCks9q+vrj/Oc/uYPzLxzwv3m8Wrqi8B3cnfaF7HPO7+G2Oy7i6195ili8TDunbRzbYvXapkI7lPj9nyLNeS6TY3RkEin9gMm67F3xhdtcPst7blzD5VevNYeXuCgYGOyktQNGj3rIioLj8qehBSyttQMMnul6vC0JXmD/ZQ6vtyXKqJCDEzPzLOzdR/rxx5l49lkWdu4iPjlFCmi2LHBioV0lpT+FL2gpEUxQCld5zFku6qLzGf70p2m7/l1gN4GnfZlDIE6dVBW606AhDPff+zL/878+QD7XgmXZxi+s5GsnW5cWTDDGsT2fz3DLredz7vm9hAXbJZTAhnNXg97M0jQ3Gq0VqSbJ0HD3CdX8ePC97ADBvr2TZNIKgV02cQSe7QrlefQNJPjoJ64rfH+prFzdS3tHgrmZtK/9q0cw4VtoJVDM8U//2Q3c9ZmNhc8XP4fwaj+sBaml0QkdEfCp37iWZ57aweSo8oXNpfct182wem0rt37wYv+ywkyGVa4Xprunhc6uZo4dy2EtaaIsmnb6+7tIpGIlx2t8hWI7lb9HpYua8PcUsHJVJ5//wq38p397D5kMpq+EvD4FoFQecFm3fmm+eTveGCGTDj9LVdTeex7SzvG7n7+N8y8cKDzRxc+9msRT9Nkr/yy8MLjpfRfz03teYnpCIa1w3DRJPO7Q29thvnOKBKpinQxHDmeYmlhAyASmfcuEIC1RyqOl1eKjn7iC+r7qpX24qzdGd2+M0ZHcqdH+NwANLWBhnmikwTpThPq7ed2NDgtAzY2T3rOLhRdeZPb5l8ls34UYHyehXLqkbXbiCEwMIFFFGPAXs1KbuDCe8sioLOmhfjo/8QkGbr0N2dHnv5YapEYI61Qt2EorFgx8/pj483u38ld//hD5TBwrJsxupnouPycNjedpurpTfOzTV6H9beJiKX5ffh0HB5pIpWLk8m59rY1QKO3R3hVjaLjtJN1D3SoCsHv3fjwPbGdRlkn/LA9PpfnAh95J/+DxR29parIYHOjg0N45hPZ3i9VsQ12YGz09x513XcHHP7XR9zEK9G5hFpuUyz8rmlCqM7iinWuuO5+ffP9lYjJw0K4lZBV90ISV5o6Pv5POvqZQIvSl0d7eTE9vgtGjCyCXYko11/Q8l8EVnSTiSw3yKyltu7AAWFZjHfrE/+jSK1Zw7btW8/OfvkHMaTOacFEUcDzPI56w6elvrVMPU+DB/SO4rqYkOoivRcvlFnj3zWdxw03nFLSVtcrKLCiymSxCahLJGLG4mWbNc9eUCirFsoaHW1i9toOXxiaKApYGz3NpSUlWrRlYUt85WezdM4rnxswOx0X3bNrG9eZ5z3sv4vwLh467fCFgeFUfW16cNL5Y9TbfLEMaXcBqBTrOdCXejoTNNgAoTW7kCO72N5jc/DKzW18hs2sXydk0ibxHs5QIqbEsjcZDBeYrgb+CrjDs+lYPrRQLXh63t5Xmmz/Eqjs+hrPubDwkrj/w2ghMAAR9GlLghMoX8Iv7tvI//vxB8pkklmOhC6aN0zDQCQXawvMy3Pi+K1iztoOKkl3tMZ/OrnY6u5o4cnh+0UK0ElrnGVzRTyxuWvtUDn1hbcDRo1N4SmBXFB7Nzrjh1S3c9uHLT+xaAlauXMGmJ/eyNH80Qz6/wJUbB/nt330Xwgr0sbVbZeTIHC8+t5cjhyeQluD8C4e4+NIhYvGl7dDb+I71/PSe54Cl7LIyWlfXddlw3gA33nIBQViRpfZTDSRSgqFVzWzdMkFR61IfISRr1/chJSH3gQrX0JDL5JmemsP1FFIK4vEYqaYU8UTIl7FgF6xUT42Ugg/cfgm/fngXSolFVdVoWlrj9PV1UTC1V6yUudaRQzMmMGwJxlE+kYzxwQ9fibTwXdQXjY6A8c27/0ev8NAvXmN2NgsiS3OLw3XvuoAPf/QK4glR0jaLqiNhaGU3L2waKx6SAiU03d0t9PbHS9VXVe+n6odLIvjmyMgYyrWR9uJ+LoTCU4qWNskH77zshC+3bv0wSr+G1i5oOxQ+5a1BowtYPVQZBQtmfELvTqhvFdMB+zFAwv3uraeJBAJ7PoRv0F8zV/lG0Erlq0kQbhZvcpLc/n3MvvYaCzt2Mf/aNvTIERLpBZIetAjHvBC2xBPKF4GM9kchjSkvqINWaOnvdtPChBD1PFxPMd/aSvJd17Hmzg8Tv/BCkAnfv0GFal7+kE8VGu2rwwVw309e4a+//EuymRi25WCMWSYA4qmuCYBA4akcK4aSfOyuywpHKZuDCk+8guylgY7OBH2DDocOqvqqfH/lblutbHlxHKXzhULNOycBC63z9PTFGVrZXndYz2Zd3nhtBK2d4kpVFx1+hQVeHnZtnzH1CwnkQZwdKW3yyuPDH72Sru64H2fJr5eo1BiVWX92G5CnXvxi4bsuua6is9PhX/ze+0ilHHRV86y5uFbw0x+9wj9+81FGjsyjlY1WgnjqSc6/sJ9/+YUPseaszqIvTRWGVjbT0uYwP639iae+NknaeT7yiStpaYuj8fxNGBA46CMWX7O8CitWdIHYDcoB4btxV6moQKMVODHo6+8slFe1dQT8+ldb+PKf3YvyYti2pL2jmb7+Xq68ei3v/cAFdPemio+xgrVN+GPB2Wf3MryynT27ZrGkVWJhEtqmpzfJiqFmwMP4FFQ23aUX8uzbO4NlOWX3KfBUjrXrWznvot462it48YV9fPnPHkCrJEKYsUKpOTa/8CBCKj7+qY11Fypd3W0lO+mU1mitScTb2fz8JMLKF2K9lXj9aUVbR4y167sqLGZL307X9dixbZJcTlPwhS1BADFee3mCgoAeOslErdfk8zPccetG1p/dU3ETVHAbFZ+jz+o17cRiFlpLTs+IenppdAFrkCpBRis/r6LEVVglhJYL4cf31pOxAv0OaEJmNH9yKhVOgtchmGld9MIs3uQces9upna+QXrPbma37SZ/9AhNC3PE8jnapHFSR9h+zyn6T1j+/9rX7EhtxK1i3aSJ7KxBe5KsVqQ7Wmm6+krW3n4bTVdcjbLjwRRVeNlCIlrJ/6cOXbjqP37jSb72t4+B1wyWhSdchB9o9LShJJ6X5tYPXk7/YOWt6SXDfgXhSgN2TNA/2ITSo5hXqvZg5tgJNj29nWc3vR5K8yKLApa2yeWn+K3P3cBnfvO6ugLW7h2T/Js/+BrKSyC0b3IIOc0KKRBC4nka2w58akrJZhc4e0MXN7/vIv/TCmLOEh5N/2CKRMpB5ZdgbtEWystw64evZs1ZnWaCX/Qd4+wbCBDf+Ydn+cr//CXoFmynFSyjhdRunBeeGee/felnfOm/f4zmltrxkxOpGM3NSeYmhXnvyn2VSm5bks/nuOiyfq67YYN/rNwUtTj0QMmnfkc697xhLOsZLGnCpJRolCriEY9rVq7qK5RViaCUvbtnmJ9NEo+1k81rjqY9Dh0Y5YVn9/GrX77EH/3pXQwOtRXGAspL9QtKJGKsGO5g985AECiilKZvsA3pVKtRsf/PzWUYHZnx41+FJwzjY9Q/lCTVZKPqbF6QQqA0SEsjpYt5AnFwJb9+eBeDKwZ83zBAlPZdpc0z3LNrumSDh9Yay7LZ/sZh/v2//iaQR/ljehC4WaDI59O874Pn82/+8M4qgnuxHScn5vmPX/xHZmc8Ko0DpnkTeK7ArqC9AtBKMzDYyh0fu4JAhi+nUIfCh4s1op3dLbS1J5iZVKfYr/bM0OgCVi8mYk1NhPb8Ph8asCmddHSJfuctJl4VbkxSsqPP/6U4PArIZVD5LHp6Bu/wUab37yGzbxfZA/tJ79xLYnIGa2YOC027SPjN6oBlg+9PtdRKCd8BRPsraNdT5IQm3dtF2zuvY+iWm4mfdy6kmpe8djn1T07i5uD//9Vf8d1vvQCqBWnJosh3GruOQOB6LqvXdfPBO6+uqp2RwJ69B1m1agWywgwaHFm5egDBriVeXaLcGEor0HqR6l4j8JRk1er2JZW2840x0nNxHCcV0jaFBADPJMiWFgjh+VpEivesNcJO85FPvpvW9nghPlp5e4wcmcK2JV091X1vens7aW9rYuKYV1d4UMqlbyDBbXdcXOPuilq05zft45t//xToZmzHKn4uPbSGeDzBa6/uY8cbB7n0inU1dSK2EyMWiwM56sWjUkpjxzzu+szVJJoqqyiFhKmJeTLZLP0DnYs/93/2D7TQ1GSzMJdf0i5LpQRtHTEGVtTSCBqtn6c0Bw/MmhQpQvtRvS2kVCgFr70ywQ++8yyf/8LN1Sdcv6LShv6BJlwvh23FS07wVJ6hoZ7FXyr524w6E2MLzExnESJchgZhTKxdXYFmTtYcpy65bBW/929u5nvfeZIjB+fw8gIp49i2ZNtrR/ij//CDgvazEMZAFNswyLTgOMUpz2ggFZ4r0DpGpelQaEUu67F69YBfZKV7FQTG24P7Z5iZ1LhurLJgBCBUKGdiqFX8RXLenee2D1/P0HBb1T68MJdhanqWwRXVNxr09LbR2RVjanwBKY8nuO3yoNEFrKpbmDzM6lMCR594hPyzL2GvWkWqs51kaxN2axuirRuakoiYhbAkQsZZkgPKcqNsphFagavQeRfGR3HHRkgfmyB9dIT8/n1M7d9LfmSUxNwscm4amc0RR9KEwLMkwrYRCCxlZjFdZYVSHe2/6EbIyruaWRvyq/vpvekmhm+8kfj6DeDE/XIDj5YzqGP0R4n5WZe/+vKD/OK+zVhWK4K4r/Crrj04ZVXSoMly2+3X0NaZQKtSjWzAzEyWb3/jV3z+Dz5JS0upQ2q4Fc/ZMIRlLV5FLkb41/fNvrLsO8KsYFtaknR1ty/pXg4cOIrQSaSIg3BZJC0WJgYPUCGHZbMRPu/muPiKAd5143n+UYpzZGgSeOyRrXS293DjB1oXXwNQHgysaKOnu4nJYzNIRE3RxVNprnnnxfQPtISbpgwzUWYzeb719adIz0nsQqT44F+gNbbJZ23Gj+WrF+fT1pYklUqhVbbGWaaUvLfAtdes5qp3rAsdL79/wY9/8Cxnn9NXUcAKaG5porOzjfnZufo71rQxifb2GVNPjRMRCLKZPAf3jyOFEbBM8/gZEoTGsVO8+PwbZLM3kExWTpUUvrfmJrOrtFQ7orEsyYbzegv3XauMvbtGcXMaWXGcs4jFYv7Z1cVhjdHC3vHxS7nhxnN4/LE32LplP3t2jXHo4DEyC5psWmLJuNkA5O9gDcY/7bejEMK3nIe1q77XaWExEPZ5AbQgFrcYXl1vx68pb++ew7iui5CWv0NcLL6tGg7nnqcYWtnJBz50UQ1zOTzz1DaOHj7Gp/7JzVVcFzStrRaDg03sen0eIS2j7X0LmQobXcAaqPZB8aGmyT/zFLN/923smMNcKoZOxPASSdKpNmR7C6neDuJtLbS2r0B1dCK62km2tRBLNiETCURzMzS3QNyPgSIBS7A4RsuZxFctewqUH1V3bhbm5vFm58nMTLMwO40YHyN7+DCzI+PMj44SmxjDmZsmNruASuewbIcWfy0mpUALC2EnsLQJMaUQBS2VFsLfBeiF/NRNe5jpIvzWaKOpkgKtTL6uvNbMp1pInreB/pveRfM1G4kNrTEpZZRVeOkW74Y71S9YICyZCSHYmXPk4Ax/8aX72PT0AWJ2S9EU6I9pIpgQKnIy6hwaPAW4+Tzrzu7mllsv8T82AzFQUMsL4IVNe3j+mSNMTWZpaam+Cuzta6atPcnszFJ3IoUG8pCPk1Ya5Wo62tsYXNFbqHk1PE9z9MgkaNv025IwBuFr5SGsuQoEEwWW7XHHRzaSTFpUS3E3N5Pn/ns28773vzNUZmnNpAVY0DvQxmtbx7BlHCrk3QvGl1hCc/17zimWF+qrxelOIZA89/R+try4D8tJGSGykM6l9F5tO8kzTx7Cinm+bcifpELWMK0hnxPMTru+gLu4hkEttPZINcEn7r4K2wmrrAMjm3neE2NpfnbvKwx+rrPYxOEW8r/T2pZkcEUr+/bMVWqaRdUwfkqDdQQsc5WJiQwTE/5kGtRB+D5+2kIIl4X5/BIiept7y+VYlDdRa02qyaGjKzCpV+ud5vi+vUfxPIEsuD2EVjEa30S+NMuHBjq6Unzojkv50B2XMjeXZXRknldfPszWl3ex9eXDHD4wjcrFcGzHJHsWApSoM9VoqoXaUJhQCT091YVmKI5zhw8eI+8qYjVT7pS/O8rvpxZKZfnoXTfQ3dNMWKgtfkeQz2nuv2cr/YP+DtiKmmJzbHhlP0KOIqUu8bp9K9DoAlZftQ9kYA738uiZKZKOTcyJoXManc2gpzM06Um0Ng9Na828C57Q5B2JbEqg4jG8WAIv1YTX1ISXTJJsbyPW0kSyrQkSTSSau1EJB93ZAbYk0dLkCwWSVHNLMYi3sEyEccfx+1GpI+8i/MSeOpOGfBYhNFopMpkMOpdHeZrM/AKW6yEmR8jPTZGfXSA/O0tueo6FiSmsuVmshQWs9AJibh4vncPS4GhBM4JWfxY2qxSJcJJm+hLBasg3m2pj/vJkEBQ7mEEKUlXhzyBpaOCebGQOgRbgaRfPgxlbYg320XH5lfTc8F6aLz4f0eKvrrRnyvB3HghkhfHvdGmvzAsuhGDLi0f48p/dy+7tC8Ri7SCUETSFr2nQQVOcqgFAFP8JF6EsIMdHP7mRljazghbCQ/mvbNBCeVfz0x+/yPwMTE/NMLyyueoVmpqa6epuYmpqxnfYr4b2ryfLDxVS8njapacnQUdHzGi6KtpzTPvOzmQ4sH8CaWk0+VATVjVQlN5jfoGrr13JddevD2RP8+2QsAnw9OPb2LV9gsmrJxeVVc76c1byq4e3m3prKobg8VyPNWt7OOucvqK8p0tkTf8OJUrDQw+8DCQRMoZW5Stx6c/XCst2eOiBrfz8/s3++1S5T2khiMXiFYIwFoUr0OSy81z/nrO5+LLhkEIviBBPYbPJQ7/YyoH9M0xMFpP2lgpZpjzLEaxc3cpTjx2ikoBYUkdtTGkrV7cG8kjNt3fvrqOkF1yEtFF4oZM1wnfKFyXR1KteGTRMjPtxwkLxKJTyaG236O2rLXCA0RodOjCJ53nYlh+NrWSnnmZ0dNw/u1ZJlQlS/qxd18kH77iAibE5tr48yi/ue5GXXthNJh3DthN+UNfy0CT+nYa0uSXX9A9rLWntSNI/2EmtnBtCCPJZzf692ULohUIcrnqJuf2UPa7rsX5DFze9P0iJE77zYm965cVDbH7+ENfesNL/JIjWHn6/ze9nbxhGiM0sds5f/jSsgKW1tqghYAEgQGVdMhMzxANhgpDduPQHwjY6C0spWMhDJo/yZoBweH9QKBZ84WMO8+qjBdoS0JzwLdkWiWQC1xZoJEJKRDyOZ1sFJ8TSqy+uOwjUwgIxN2vU0xqymQxWXmHlPPILGWwNWAKlPGxp7s1RJnaFZRxWUCiEFEgnhlAgtC68YlqX9tui46EuVGMp8d0KLzMglDH9CSHB83CBWcdC9fSTPGs9K999LamrryC+YiVFh+pam7dPNxKtvYLK/b4fvcRX/+YxpicV8XgcLTKcfu2lppCaRCtyOcUFl67g3TefU/lsf6J/ZfM+Xn5pHyiH+blM9bKB1rYYA0MpdrwxCTUFrDoIU8+Boc4686953gvzC0yMzyBl0BeW2A+0BbjE4i4f+cRVWDFpzCiLTyS7oPnJD59Da4vJyemS65efC4J161tMhOqK1RHmvdIuZ2/oo7UtgdK64FRc6S6PHJ7h9dcOmwWd55YJnCHtlP+LYztGg1Hj5avu7xgcV2jt0dJm8bFPvcMoQoL3OnR9iWB2JsdDv3gRy7I5cvhYSSmVGBzq9s3idaYILUimbAYH65mnTEMfOXyMXM7DlrpqN3Bi1Z3xi5gNEbMzOdBWiflOKUV3T4ruugmYYX4uz5HDC9hW5XfdkpJtW0cYOTpHX3/1xUs9As1MZ3cz17+nmXe+ey3PPr2bb/2fp3hl8zHsgv9RbYG2EhpN30CKppZ63xUsLGQ4emSqQqaEehfx5xo9w0fvei+tbUkWvzxWUCF+/tOXyWZgZjpNSOqvyPBwK6mkQy4f7PZ869CwAhZmZq4ZZFQDKqdwpxdI1dh7rst+hh1ZpGUXVK9SC8xm0eIuJi21L7j5K92ZwB9CoqfTRY8us3c11L2XouKWhbpo/2/H/1sCKcsGrVEClGNqJgQo25f+tUQisbSJ9Ky1MeVpUVwH+dUvqY0I/V/6WylCmolGowvKLEtp8p7HglRkRBy7t5vUuesZuPoakldeRWLVKogFzqJmFW9Eq8CPIMhvdgbRAiFs0gsuf/c3D3LvD17GdRM4jg0iC8JjUVqI04EwWgeNwI5luevuq0kk7NCaNIjobfyGpNb86LubyGdMWIv0QjUBy/QvyxIMDXUjOPTmAidrjVIua/1E0PXMjRNjWeZmswiROO5L5fNZbnr/WVy+cZW5Vsi6KEJmsE1P7uD1V0ewrSQTE7M1SjR17e5tpqk5TmZel20M8M1qSiOky/qzh83RQJlb5VYP7Bln5Mg8tpU8rkCQS9s0Uk3TJ8h7aW656SLOu8AMlaX3UmysJx7dye6dY1hWnInxGZTfqYIRq7hbN9AqDJk0QPUSqguLZFLS298eqmv5/XuFYwf2zuO5CjsuKC5ggnqaDAI9PW04JbvXKpUpyOXyHDk0hWVJEy/KP67RrBjuwrKgamBOv8iZ6TlGj874Y93i61iWzciRef7x60/w+1+8BYRiqWmZwixKqSPg6mvWcsFFw/z1l3/NL366BWnFQFjH/WIq5TK4IhBwa393bi7N+LHZ4xew0Hiux/kXreD6G8+q8HnxuttePcKTj20jHksyO62Zm/Zobq9+vfb2Fjo7mzl6ZBaqCLrLlUYWsBzMLsLKBOrgbJb0xCT14vWGMpmhBGgklpK+BkegpUJqgaVN/Cblm/Dw3Z1UYBvRxZelNORcWceuvkCriecLSGijiRKAJ0zcKIvg+sZ8pTHClxBgKWHug9LhSIUWz7pQ/1LTgCAk84QSCxt1GGil8dAsoBBOHNXXReqic+i4bCMtl15JYvVqYxot3HswexthSgrfTEtJAIkzh4Bd28f5H//tAV56YT8xO+Vv0VZGuFa+L95prqrwn6mbzbHxuiGuvmZV2dTq90kEloDXth7mxecOYFut5N20CSJarWz/DVh31kqkfMWY14Qv5NfZobYYiW3D+nM6/L9ra6Vef+0onmtz/OnGFPGU5vY7rylNHRL6HDRuzuLen7yA61pI22J6Nsf8fJ5kyq64qxKgu7edzu5mDs5lkOGNL4F5TYOQWTac1124Tq0NMjt3zKE8G1FlW3uxfU7O4kIgUUrT2Z3gE3dvrHiO9jUHuazH/fduQqkECJupqXljwhTV+rimqztBW3sTE2NZ4ydUBeV5dPUkWTEc7AQuL7A40mSzHvv3Tvq7xXTZOR4Ciecphof7sZbQWeZnMoyOTmI00qY/mx3LmpVrBoLLVsY/PjWRZm4uh23HUF6l8yWWTHH/j18l1eLwmd98F6lUKH5b3VpWR6Fpanb4/BduZGZ6lsd/vQenph9bZbR2Q5HUa9fo0P4ZFhZyCHF8WRCEkAhrgU999oM0NTk17/1n925mftbDcRzm511mZtI0t1fT/im6elL09CU5sH8cx2pkkeT4aeS7aQPaq37qjw1ibJRYLl1TBQmlU4gMJCURBG/QCGWEKu37BhUHC1+YKpjIdMXpaNGxE5ifC0OR/4uSRV2THXaw1sU0KcG6zQt8waqUKbXvleHfeiA0Sr8sC2MC0UrjegpPQtrKopw4OtVOYmAFnReeh7jkAjrOPw97cAhklTg+iwZu4aeHPp2rk8rpiZWCX/5iO1/5n48zNjpLLGYcYU07GS2lOt2ClS8VCylAW8QTik98amMhxUZpdVzMayu5/8evMTsliccEUlocOTRrnm3Qe0OaWqXNLqm16zuJxyVuPhDWjn/C9zxNS3uKlram+icDhw+Nm6CVxzXaCPJuhltv3cB5Fw2UfwQEE5zklZeP8MqWI9h2AqVhfm6B2ZkFmpqKQRvLtRgdnUn6+prYt2sWu4L2WylNW2eSjs4YwYhRy3pxbHTCOFtXtdIc53QcODTrxX04GB88L8vN77uIlas7jd6zoNkTJSLhpqd28/rWI0jRiqcVkxNp3JwmlpJVaiVobWtmYEU7Y6OHsKxqEe8FrucxMNhZNV6SKd20byaTZWRkyl/MFP05C/erQQrNhRcOl8VQEiXFaW02Bbz8ygHmF7JIO4H2E11rNEIqzjmnn7JvlhDc96uvHcbzr6ErxTfz3T+0jvOPX3+JVzYf4667r+bKjauJxyuNtuXa0HAtin+Lwv+QbJJ88jPXsPnFPWQXfA3qEruK1ppEUtLdu5Ro/7Br5wGUB44tK9xvQMi6oU293XyGy68Z4qp3rDT3UfB5Ky1jz+5jPPKr17FjMRQwN59hdnYGqCZgmc0F/StiyOcbYPF9kmlkAWuQGjGwgu2hamIKO5tHIk9gYiwTi0T96eZ4tLdvdq266Psn2P+C90BQjGovtDDJXJVGaBdXe2RQePE4XqoJu7eH1g1nI86/gI5zLiC+Zg0i1ezvrlQEWeDDgUBrc7qFKyiv17GjC3z9q4/zwP2v4nkOTjxemIBPt1C1CAFaS3K5NDe+72wuvnxVldMcBII9e8Z56oltOLYDwkMrzdixcTP5SBblnQ4EjI7OBJ1dTYweydYQBmqjlUdXdys9vR31z9VwaP/EcV9DKZf2Tos777oay64+ESgl+PGPniad9nDsBALF3FyW2Zk0/QNthedaSS8wNNzDc/Jw5XqjSSUTfhwqc6RWP5+YmKzSlsH7oXwhQpLPZ/2dabpQcvk3kBqnYG5cfG3Py9Hda/HRu64ETPiJckWTJcDTmvvufZFcNobjSKRQpBcUk5Mufanq2pJ4wqJ/IMnmF7waejuN1i6r1wwUhLpar9H42DxTUwtl8aaComy0ztPda3PhJWtqlFLsy69vPeDvhqPQiFJCqjlOR0c94d+06ZHDYygtkVUHAN/wKCSO3cwrL42z/dX7Of/ifq5/9xquuuZsevubcexa46AO/St/6cxDO+fcXoaGO9j+2vwSkiaHaqc8OjuS9PYuLSn7kYMTvjBa60lpjJuEAGGhNTjJLHd9+gbicdsfMysPHj+/dwvTk3niMRu0TTabYX6+uutCwNp1Qwi51Bh9y4dGFrD6qSlgmUfszszj5BVSSD/SbkQYgZnklJ+mRnuKPJqc1mA56EQcK9VCakUf9vqVNJ13Lh3nnEts5TAymQI7TtGo6A+hanGgvMZChky3hice2cnffeVRdu+YxrFjWNI4w56uxKlLQSlFU6vizk9cVVOoAPj5T7YyNpIhHkthJm/J5OSs7zskq6ZDaW5K0dPfyuEDh/ydhMcvYWk8BgaaaG2rP3zMTGY5Njq/aCt9zfI15N00H/jQlaw7q3qQQoHg1VdGeO6pvdgiZSLEa4+FOZf0fL3YUbDurEHQL1BtshBShkIk1BawpiYnjTlyUTGBWkmhtYtlKa59xzpa25qKoQh8rWlYlsrn4PlNe02wT6yimtPHY45bP/wO+vpb0SqklS9jy/P7eGnTYWJ2MwKF0oLMgsfY+DT9K3qrPn0hYXh1B7JiiIgiUiqGV9Vz0DDs3jmCu+ixCOOWIASuO88VV60tDVha5fILc3m2vDBqNk5oq7CKdN0cnV0purrbatZFIHBd2Lt7ApREWMEOt8VzSJCqSQjhCw6azS8cYssLh2hte5Zzzuvj0ssHOOvcFZx1zgCtrWbaWhwhqnpbSinp6W1n29a54xKwtNK0ddj09rXWTb3k5uHw4QxSVM6UUFYyRnsncHNzvOP6tVxy+Qo/qoioeDuHD83y8C92IEmilHFvySwoZqdzNa5jCll7Vl/9kCDLkEYXsKp2l2BgSE/NIrOuv6X+9FTsZGMcjou+MPUVC8L3xwx5PRS+YH7xtNnd52mNlhZIG2VZ2G3NJAd70UMrsFeto3P92cRWr8Vqa0emkibUBGXOCEGqFCELqp4q71jDEKwnR4/O862vP8ED979CLiuJOQnMyvvk+MK8aUIJNfNumltuPr9oEiub04M/jxye5MnHtpBqsrDsNGjj5JvNunie9uP5VEITSwhWrkrx8guCyvGo6qPx6K+7a8yYVyYnphgdmTwOB3eN53n09DVz6+2X1hRrtNY89MDzZDMuiVQMdA6EIpvLMjWdrnulNWtbceKW36dV8VkgKprmaiGE8HMjlt5L8W3WuG6O1Ws7+Lf/8f20d8SKudrK5DIB7Nw+ybNPv4pGIgh8Ao1DqOe5rFzVzu13Xlk0OlWoqtKaX9y/GaU84smsr+kELfLMzk5jUr2Wm7SKO6qHV/Zh237lCv56xZoa85RFT2/bojLKagJIDh08Si7v+dqe0B0L8FxNW1uKj37i+iU1++aX9rN39xi2TBAeMZWn6O1L0NZRP5H3zHSGsdEMQix+YTSe7+AvQqbXoI9os4NbO8zPwNOP7ePpJ3bR1BSjp7eVy65cxWVX9XHZlefQ3ByUXe+mdCi47tJRaFYMtROLF1qAataC6al5jhyeqhJSpRwTME55Hslmzcc/fQ22Lc3miCo8+qsXmZ6eIpFqQgiFEHm0VBw9Oln3an39LTS3ppid1VXivi1PGlnAqhpk1B8GAEF2aho766LjNm86WeQSJt2T/uiFcTYVKBPElmIOPx1ITsKkK1HKQ0gLzwWQaEujdN5sIZc22rLxpMB1JInudmRnB2pgCGdoNS2DK2haOYQ9MIiTaoJUs6+dqkSNJdRxN8DpfVkK0WQUPPrLV/n6V59m764ZHDuBFURPFoLyXGDyTMhbQSwyodCeprVD89G7rvJXcosHyqC+nZ1NfPmvP0MxAbDp+ZYlQulZqjO4qh1sf/fGCe3ozHHWObU2+AbaTsnYaIaF+Ry2lUIv6f3UQIYP3XEdK4bb/ZAnlXeCCSH4J799PZ/67HVlcoKmtTUF1N5U0dHRTHt7M5NjfgBSXdhaZ3brKo3y/B3GFb4fPjY81MPm58YRQQJroTDBRv2E6NpD4HHb7Vca4QoqKoA1eQQOzzy5j9k5z0xqvim+sIVFZvnQne+ks7vJPP1CRYLSlC8KCf7p//1ufvtzN4QuJEBr2vxt9qWbc0ShXQHWndVDPAnZrNmsIkK7/QTgKZNgeGi4N9QeldrbJJo+dGAa48wOxc00YBZ0WT752Ws467yeEmNaqfhnjioPHn7wRRayeRKxJEVfQtOfB4fazPMjPJr67z5FzfbUxAzjx9JYVgwhvMIVlefhxPM0NydKhJFFfVBo8lmFnnZAJ8hlXA7uy7Bvz4vc/xOXlWue5BOfvI6bbz2vhmYpeH81C3OVNT3VE22bug2vKvc30xX/mpldYHJyDix7ia4uinw+x00fOJcLLhoq9LXiV0sLef8HL+fGmy8uqa/Wmqbm+our5rZmunqbmJmeRgrnxHc4NxiNLGBVXSKbtpegXRgb4fhy5DUYWqKlcZ3XSqO09kMEmpQJWhRX056UKFtjtzUhm5sRra14rR2oznYSAz009fSTGhiC3j5ibU3YiQQ0NYMTrPKgKJoeB4vH4AbSXJU6s0tg985jfPsbT/HIw2/g5W1iTpxGqnE52pO4bpZbP3Q5a9Z1+Eerm1/jiRi9A/VX6NVYtXYI23oO5QWD9HGY1jU4ccnKVUuLCbRvzyRKWWbzSM1H4PdxL8fAYBMf+NAl/mRYHkahlGK07uNDa013Txu9vW2MHZ1AWn4f8t81KQRzcxlymVrmjSLX3XAO9927Ba2DXIsC8E1XWpPJprnqmlXcctsFBKICIRNSUQFtMbfg8quHNwOOmdG08gUSietmWH9OO++79by6dRJC01PHVFaL1vYE3T0tHDroa3nKhlitNV3dcbp643XGFEF6Ps+BvQsmRQ4Q+A3m82DHcvzG71zJXZ+9imqbU8wFjeC7443DPPXEbt88rjHBgE37COGxZt3wku7v2OgM8/PZktx/AK7rcfGl6/jXf/ge7BglmsYwUsKTj23jy196BFs0I4SFlBrLToJW7Hh9hr/6i59xwcWDDA6316iJyYwwPjZzHC4LRujVuKxbP2hK0YEwFtLEhji4f5L0Qq7GpoVwlUzeyPYu+MQn3xkSWKvTXnWnYN2L0dYeY2i4iZ3bJkwqobcIDSlgaeOFV1WDBf6Dzi7gzowhZHjNswzRgrwt8NatxWtK4jY1QSpFsr0FKxkj1dFLvL0Lu7kJ3dOFk2zGTsSxEylIpiAZpPtYvM4OT52i7Odbg+I9T09m+fEPnuMnP3qJsWNZ4rEWbEeEzIENJoT7k7nW0N0f40N3XEbp4Hhq+vSa1R0kUxbzs4ql5eYsigFKaTp7munoCpyIKwlnAqVNurWdOw4aH5m692Kek6vSfPST19HbZybPWsLVm0EA8ZSkf6CZrVtGUMrk3zQTlBE652azHBvNsmIVlO9DLqLRZLn86jV88PZLue/HL+MpieX79GitUDrPVdcM86/+/ftJ+lv8y8O6FJVQkkcefpndu45iySZ/di8KHMLKcMfH30NrW3yJoXvLR4Clt2dTc4LB4S727dtDTFqUP2ulFCuGu028qTplLcynOXR4BK3AJQda0dTicPG5g9z+sYt55w3rCgJM0ahaYpBECoFSmh9890nS8zYxO+ZvFhC+H5bCicG69Utz+N72+lETmsGBQm4KbQRHx5H0rWipG5bpyo3r6e59kpFDaWIxx0Sh1yYDkpQSJxb0hWqYm969Y4Jjxxb8HYtFvRtU0mD5xkDl0dQk6OsPh1yo/ox37jiI0hYWkmLanaDI8hu18dQcN91yAWvOajcLnVM0cwS+Y32DzSeoUW9cGlLAAmLU0GCBb4TIpEnPTJC0dNgMv+wQ2iObaGL9f/h3OGvXQDxp4koVtsKWphhYPFV5FEO2i+JJUEhgWnCj0KBkYPY4vgG3kQivc/M5xSMPv873//Fptm87hmO1EI83UQhi1rCdww/iSoYP3fkOVqxsDT3zU7dgSDUl6O5pY3Z6upANYKnto5Sip6ed9o6WElN9ORLbz0E4hckxV9uAr/HI53OsPbufm265uCCgnbLe6U9aK1cNIq2dSCHQyn/PfBOymxdsemYHl1w5WDPIiMDCiVn8y3/7PtZv6Ofxx7YwMZZFKUlHR4xrrj+XW2+/jFTKpvS5FktVKCSSsbE5vv3NxxDaQVilBrJ8Ps+Fl63gpvdecFq28zi2YHi4mSeVr0ETwU8NWqKUYu1aY54yfltQ+Ylpkk0xfudf3IRyzf22tMVYtbqL4ZU9ODHz/JUWSCEr9JPiAvqRh7fx6MN7cexm3+dNgoqD0CjPo60jtWRNyqGDI0gZPBODEALbsvA8jZsHES/KIoFisiD4aRhc0cEffPE2vvl3T7Jz+ziZjItAEUsIhla08H997hp6B+rX57nnXmd2NoMtlxD6xA/Zo7Uy99u5tE0Ghw8eQ+vSuaS0zOKiTilFb3+cj9x1rZlr1SmMAepXZ826YaT16nKdkirSqAJWAuOBWRUBqHSO/PQCKWGZGE6np24nF23MgqK5BYaH0c0dxfk1pJgLu8uUTmq+R0EQJ6uCmqogRhVWiMtFl1XqjRE2rEiMLPLCc/v4wbefY9PTO/BcC8duRkjhp8KhtJkahKLTrMb18vSvsPnghy8ODlb4xlKFrfCav4Jd1yeVchgYambn9lFMSCLJUpweBCZ9TP9gM8mUVRKhP1yD4DamJhYYGZ1D2lbN4oP2kFaOuz59JW3tTmi7R6V7LK1VVZbQbOvP6QGRQWsHIfx6CuPzZFsJNj29nbs+cxWt7YkK1w6ub0waTkxy+0cv5faPXMr0dB6tBO3tdjEtqR+/qbgG0ghymKWOSbH191/5FQf3pok5LaWxrTTEYh53fuQdxBN2haesMTHSHKpvXqhuPqrG8OoOf8etRkhNMMpqBdLyWLW23ZRWR0hvak5w2+2XVL+QLo9CX/zABCiW7Np9jP/9N78kl43jOKFr+V9TyqOnt5mO7tpmYwF4Hhw+OF2m3Tb3YFkWRw6NszCfoSOeCLmfiPDlzOYEDddev55LL1/Lrl2HGT06SywmaWtLMrSyn86uGEoH4WxCNxsqaWY6w8MPbkFIx2x2qvdotJm2Pc+jf7CZnt7mgk9f5fvVzM8rDh7IIEUQoqHC89KWcUzHxXNdPvDBaxgcbMLTtRzPg0VCtf5Wn6Btzjq7l1hcolzP33iwuI4NszlpiTSygFXTRAjg5vIsTE7TLqxGmj+PDz9Ke7ytFcuyinNCSUcSZZrTCpPnUr0CBZRvHm5cSgXBoN4aeG3rCD/63iaeeuwNFuYsHKcFYQfxuY5vEjljaIEQHh/5xHV0dqcqyAO6cN7SbqXSwLm4RMcWrF3XymO/1NT0eSkru5DYd1V/yfEKNUYAE+PTTE7NIq1UhXqVks/nuODiAd5944Y67/JxPNMlnDq8KklLa4zMgj94C9CYDAaW7bB75zSPPLSdD3/sYupJbOHBv63dKR7zfWNEMCP7lTMZQ4t9+8fff4UHfrqdmNNG2ONFAK6XZePGlVx7/Vk1bq26UF39eO17WrlqiHjcJp9X/iIt0FsqWlsTZbHQlnrNCmeIamY0gYXgwMEp/uxPfsrIoRxOLEUxh0/wvhuN2sBgiuZm28/9V72VRkdmGR1ZMBrcQjlFjo5M8fqrB7nmnevRWpQsUMO/BjtAk02SCy8agosqXa9cuCqt13e/9Sx7d6WxrKbQOdUQBCZspfMMDLZh20GCrertPz21wMSxNKJaLIRQLigvD6vWtPKhOy5FaRcham0VCcp782Nte3uK9vYUY8cyfgLvE9vl3Eg0qoDVBtTVlcqJUWJunoLta5EdufERQuAiiHe0Im27oHpeQrbTtyylr1TpsLF92yjf+/YzPPHo6yzMSRwnhm1X2iLf6AjyrmLNWc289/2XEKQKqay+N/m93th2BAun8jkEPjGtXHb16ipbvovfGxjqwYlZaC2XKHD758g8K1dXz2AVZnRkjrnZPHHboVbUeKU0lu3y8U9dQyJp15zypyczPP7oqyiv0gBsjmmRJx6XXHP92bS2Vdspa0g1p2hrT7Ew5yKtkLDpy4NSJPnuPzzDFVetZGhVR82ySnZPlR8rVLNUixS0/S8f2MHX/tcjoBMEQZQRGiklynOJJXLc8bHLiSdrCTE2Y6OzPPPkHrQnFysWgp1ySpNIwg03X4gTqz1mrhjqINXsMDMZ7CI0uG6elvb4koLNLoVwNYNMW9JvhwP7J/kvf3QPr708TtxpgipaDCFhxbDxLAn2XFZjYnyamckslhX4L4XC5Pjm4h9990WuuHI9sQQ1czKGn6jyQ1wEx0RQXuHc0uf3+K+38ePvP2d22Wrbz9Vaa2NFsCgCtGZ4VV09BADzs/NMTy8ggyCvFRfkCiEkSnnc9uHL6e5J4Oms70qyeGzKpl1+/cvtZDI5X7tplfqLaWn0jzrP1despW+grWYdm1tS9Pa1MHpk3n8Xl998Xk6jClhDLMX7dnIMx8ujhTJpX7zl90C01uQ9l3h7Mzj2opXS25WiYdA4jL629Qg/+cGzPPPETqYnNU6shZgjAQ+Nd2Yre5wYIVojrTwfvesGWtti+ImMKO32piNkMi5/9eVf8vJL+31hpTL5vMuV1wxyyWUrsWL+7FomrQS/rls/TCxmkc/KJQjzxamqKeWwZu3SfD527phAaJNqpppwJYBsLs/V157FxneuLQoXVcSsXz74Bl/+s5+ZnV8aSgZh/1pKp4klPIZW3s0FF6+sWcfOzib6Bzo5uO+ISfxehm1ZHD44zX//swf4wz+6g47u2gJb+L7qfRqYuX96zyt85X88QmbeQdrKhHjATMZKKXK5PDfddA6XX7061CqV2+f+e17ka199DFv6UeBLBCw/zZLyaG71OHvDWlatq25O00Bzq8PAim6mJkb9nI1ewUTV15eio9NeorN9fcLmZRNqRfDoL3fx1b95mIMH0iScFrTwML6V5WO9AuGxZu1gSVnVGDs2x/ycSyIhKL53fr8TYEub5zft40fff4a7PrOx7isSfBw2c1Y3Choee+R1/vK//pz0fBzbtlDaq3BWJcyzlxZ1wqUU2b9vmvSCi+OEU+oE19F+35Dk8x7nnNfNLR+8wL9KeHND6ZT8/HM7+LM/+alZpEkTkqQo+xa16Upl+C//7aN1BaxEUjK8uoVXNo/4ml+97PUMjSpgDbCEdzY/MYWdzyAQNVIdND4KQba1149LtVR/m7cixWFIAnlX8/KLB/j5fS/w1GPbmJ+zsO0EsXjYj0AsDpyngymoXHtwhigmsiyM4fl8hvMvbec97w2225ea6cID8ubn9/LqllFiTgcIt2rvkJbL+MQcnvLMTqFAE1OGRtPbm6S5xWEyG1ylViOZVbPW0N7ZTFPr0oKGHtw37gs91coWKAVNzYJPfnYj8bikVqysudksv7jvJWzRZmL1C12Y7AFf1nKRxMll0iwspAu1L16xFMsWDK/u5PlNhwn8GAVFwURrheOkeP7Zg3zpj+/hC1/8AH2DrTXe0irm6Qoy48jRBb7x94/y4M9eQ7tJpC2NJlbgm49NMuXmFskdH9logn6W3Un4atOTGX710HZiVrPZ6r7ItKxBW2g88jmPkaPTrFrX4gu1le7ERC4fGu5k65aDfuR/84lxgu4su+d6LL6K8mdkGYpLl88rXtlymHt/9DxPPbaHfDaG4+eZFGUR7YtFKxxHsXZ9V6hOgeZlcQ22v2FyLGq0rxEL/fNfRcuK8/WvPUVzU5Lb7rzYlKrrrUcq9wytzaYNgNmZHD/83rN871ubyCzY2Lbjt4OCJS4Wtda0tMXp6Q070Fd/z3buOORrr6pv1dBao8nx4Y9dSmtbDE+DwGRHKL+lXBbu+9FW8jmLRCLut6MsVYz5TeG6momxbOhg2b0QGHkFQ8OdaO3nlESbJN5UjoO3HGhkAasu6bFpYjkXKWNILU7LzpqTjUAjpIPdMwQiWC0sz8705jGvWTateO7p3dx/3/O8+Ow+0vMSx04Rs+3iwrWgs188iJYU11D4ZgjhEktqPvNb7yZRMPmUPvPgL6Xg3nteQHkSx7ZQvndJOUbzaTM9lSeXk8RqyEACQXt7kuFVvRwb2Y9j10sUa9rYzecZWtlLb29L4W4q3WHAgf1jZoAuaOcW43l5brjpLC6/OgiWWLk9AJ57eh87th3DtpuK5q7wFYURTrQWKE8yNuJWKamUNeuGENbLVJwcJaAltkjy7NP7+Hf/6lt85jffzXXXn0MsUa3kKuKXf2hhzuWXD73K97/zPPt2zxBz4n7+SM/M3lr4GhoP18tzxwcv5aJLK2sqwld59JHX2L9vBsdJAJ4/npSdLRRCCnI5wfTMvH+08n1I/7+zN7Tyi/td/75MhG+tFRdfvNY/78THKykEnoJcTrF3zxSvv7qbx379Oq9uPcbCjMJxkn7aKFXcpVjBFcTzXPoGOujuaVtSnfbvPVqcxIGwCVv7fhpCWuQyFn/5579k544xPvnZjfQNLC3BeTmWgKnJNE88uoOf3vMC214bx7YTSEsch+YqQOCpPD09LXR1tQP173fkyIQvrOAL3qWmapC4Xo4rNw7xvlsvLNTZr/2i8nbuGOGl54/6mTF8nbMoF9T90LNScPjQVNV7DC8tN5y/Ctt5Bq2kiZxftvBcbjSqgDW4lJMWpqew8wIRs/CEWpYhNLQG4djEO040SNtbh2MjOZ54dCcP/vwl3njtKK6rcZw48bifoqFE0bKMXrogVIQ2A5Xnuaxc0Y2Xhacf32HyBy76iknafGxkgZeeG8GJxdDCreI7gVl4C0k2LXn45zvpHbQRKhDo/EYLxlcNaIly4yhltDS1V4jSr3seL6d56vEdhZQZ5RHwlTAr9Zlpl7FjWbNLsTzmTgEP29GsWtXLU4/vKDzbijURxgkcEv4JVoWTJeiYmS4kPLfpMB1dCZQqVrJ8M5SWmoljed+ptrzAYG1tBDfHTrFnxwJ/+p/v5bwLB3nfLZdy8WVrGFyRophtpfKEkMtotr8xwubn9vPYo1vZ+cYUiAQxpwlErhhyILiuyKO1IJmI09XVxtNP7DQaBq1LxrnASurlBT//yetIkaReCiST1kez5YX9NLfYaK+yBksB0rIYPeIRc1JGsNEmrIETsxk9Os3Tj+/EU6ZO9d/I0jPm5ucYH8+yY/tRDuwd5ciReWamMkiZMJrqmO+jJkKprar0f60Ujt3Mi5sOEU8Wzy88jeAdlGZX4t7dU1iWv/u6kgbXfAlpaVA293xvM5uefp333XoJ77x+g4m1VdGJpbR+k+N5tm8/wAvPHuCZJ1/n4L5ZlBczm3Kku5TkIZXuFjTYVoxXXjyAsIoa8kVmSQleHnbvnMK2/Hh0Irw4DU42vpiDgwM8+/ROXL9PlD9XpTXSltz/0+fJZnPEYsnis6n84mJbNttePcKTv94Z0hKX1dN/POOjWRw7Qd4VZnPBMhrmK9GQ1ddafxO4u+ZJKs+B//iH6Hsfwo7FjG1+GQpYQmtmUgkG/3//mdbrb6ZWLqnlTPmjCTqeVrBt61F++dBWHv/1DkYOZ0BKpBXs3KlgAiwUKMN/NDBB/XwfFnImoauWJhWMVpR70Ro/HIlSFuCYeD01BKxi6AeBxkUp1yQ/Rhj/CIpf1crX9BDDOKbWaz/f3Kg1UuTxdK6gRSj/aiG9ExIpYxWMT+EvKLNK1RKlXHN/VE5bZDIaJBBWEFA3EH6ql6913pg3QsfKd1EpXNC2r2mTofRDhW/4P1zwzLW19nBdFyHydHWn6BuMsXb9ID297fQPNBNPGn+e6fE8Rw7PMDoyzr69xzhyaIHZmTyWFcex434T+ubXQsy24LJ5M+kJEEIVEpOXC+PB5GRMV3Yxr17hwSxuzCBZr1auP9nqiq+QKVugsf0EwbJQN6GNOUtjNIZLE7BK2z6byaEQSGkhLPPOmzACgcN5MBbWMWELjfENs/znHUrpE1bSaBshFZ7OAo5/LdA1zGYI5d+ryTKgVJ72jgQDg3FWre1ncEUfAwMtxGImIXcm43L48DTHRqc4fOgYo0fTjI1myGQkth0zzuACBMVAtPWoHGhUIsia98ava2Uzb9DnBYIYShttYOmz8LWmwsUS4Ok8QT7a4vutCwK+lA6eVtiWY97LWk8+2ESjPZR20X6qpPJFQhAKw5IxpLRxPQXa8TXVxQVg5TANxtVgcJXFX/7Np+jqqquseAy4UQjh1jvxzdJwGixtIqHV3aakswu401PY/tZnUdWToMHRGsuxSXR2ARqUXMZa0eJopgt/m+ejRWnEpKOHZ3npxb386oGtvPryMebnPCzL8jU1/oSjAnNAtcYoiGmn5G5OGoXRxPP/dPDyCoQfs0iX+10YvxBtpCA/Bk3lSbAcrSVa236EawDbDFR+sWYy9wctWFqCWeFSNMfZCCSFBHiVLGpoP6WfP3AXhKGyexQCrW1QEiEcI1RUKLNwSGoz4OrFGwFK20AhhDG9K6/s8zJfTa0lGg/L0kUtR1C/4I6CSSmQtYTAidmgbcbHXI4dy7J1y7aCJtA8A/8tEL5/kTTCUjyRKkyuxXaRLFJlaAeESZ+FsBDYJtZf2e0Ef5oA9L6gETajhYMPl39TOIQ1dIvOEPijqvLbvagpUUoYoVRIv2uWe89V6lelgkws0USox1DYDR5cL3CXWNJ2/aBNg59m3Ch+y/Wb3EJok4JI1x0+BKggEKnCdhIIEszOeExNZHjtlb3AHjzlx7kS5opaGzOslAJhSYS0iSdsgl11Jc+9nvBYFT9+ooj5GtrqE4ZWZkepUh6LE1v7dfHNcUpZvtaokoa4uFA0eV1FWdUr3YfJoWsEPAeT/qn0vOCt08LD056vHZd+H2jwsb0ODSdgASmgq95JKjtHfnqCuL9i1ifaT884GpwYsr2XZSpV+RQn0fBvEkGQWHl+zuPVVw7z619t5rmn9zI2mgNsLMsmHndC4pk/4Szn5ghT5jOiEX5qIzMhLr5Nf8gRUNxJ45v6qplICoeNQ6rUfiYAvDINoPC1EYEP0/HU3+TJDPtvVJLPROj/4gkVJkmjokTLYHu8Ff5maRUKv4Udgct1OX6dCiN2scxqGCNUIET5k3sJXqGeSEVJ6hqtsW2bYBgt1Y7pQgys8P0aDWO5T1qN5yD8YAWiGF642tma0EQbNsuW9RldrIx/RLI4qIHyfeQDbVd4gBW+ZqvYT4XWLH6+5UJjhTovMh8HalaLJQsgJZq/4PfyULVFaUoIG5PHbymLC694lh/MWkoLWTWpeqnUVhTigrci/ASrhy6pUyn/Z8zXUPlHqxUlzJtrlLPl7R1W8QGWW/K3KLg3mHODPJs6/J1FZZVf3+9Dde7VCKnhvEvL0au6lEYVsOomk8rPZcjNzJuHJ2uFlWtsFBqdSkBLaOv7cryRwsChKEb3EaTTHvv3HOPRX29j09N72bd7kkwmj2PFkFbcD8C4eAG/NJb/CwhVZaY3R3hxKcpm+ZNItbqLRX+VqwvKV7+hMmtesFSIqXy2ODWNWi4oL7pEBR3OkqpRpR9Xiet3KtaRi8qsW+96bbGUMupRV8VUlzfXDSo/l9rP9cyt8mvf64nVqziGnNw7q91fJIW0TMucRhWw+uqd5M5myM2Y5JgKjRIa68z17RPGQ5PsbMMKtmw0eJ+qJtLI0G8LC3mOHJrmqSde45kndrNz+xjptFn52ZZDIu4UXh5R5c0tCdq4zNIjnEzKA/cdxzeXzdbmU1nPpZV9YtevV3bjtn8tPVhEPRrludb2S2oclks9TwWNKGC1Y1Ll1EROTOC4LtpPErAcEYCnwWlrKah5G9/F3egKy4foufkc+/eMs+nJN3jh2QPs3DFKJu0hfBNgLGZ8ECjZelvUekWcTt5+A11ExNub5TlHLncaUcAaWspJYvwojpvxp2ZdcefRcsAVAtEzAPE4aI3ypf0zImSV2FlVyQeBWCX9mmkNk1NpXtt6kFde2stzmw5xaP8EmbTJXSVtG9uJ+0JVyHco5M9j/FROx41FFKmwRfsUULpaXT4PuVFX2UuqV1UbUYUNBidcVkC9RdFSzLRncmF1/GbkM9U3Tsp1K0W+r3RaxYDd5e+y7192spsjXMdyc3yDvpf1aEQBa0kxsHITk1ieAmlieyxX+dwF8p0dBadncQZjfxTTlEClCMjZXI5DB2Z4+aU9bH7uKG9sO8z4sXnSaRfbjmFJB8eJmbJKTFvL9elEREREREScGMtWwFqYnMHKeWBbSD/b+bIzNGnAEojuNoIIzlJzxgSsct8Cz9UcOTjHG6/tY+vWEV57dT+Hj0wyO51HeRLLMvFx4okYQVr5YnqaYvLUiIiIiIiItxuNKGDVMBEWPZQWpmaQngLLxFgquPYsJ7RGxmKkOlsJVNai1E5X7Yv+z+NUcVOq7C1XGnse7Ns7weH9o2x7dYotm/dw6OAk05ML5HMelh3HchJYIm4idIe3yy+KIRQJVxERERERb18aSsDSJllS5RhYYbnDnUNMjRYiElM9PFCDo8GxcVqbKeYgrCc8hcWkcmGs3E5thJ5wPOTw2fNzLvv3j3BgzzRbXjrM7h0jHD06xexMhnwOk+NRWghSxBK+hqukEIva14+IiIiIiHh70lACFhCjWgysgiylIb2AmpnDlrIYCO80VfBkorXGiidIdvZQlFzq3UlYSxT+jot5nIJywUsC6bTL6OgURw/P8NrWcXa8cYBDB6YYOzbL3GwWz5U4dgIhBNJKEYuX1rP4R6W6RERERERERIRpNAGrGeis9EFRCyPIzabJTc+RkI0d0KAmgcDoxLCauwHL5GPSJkxnLQNbYEgsZlAXgFP4fHoqw/TEPEcPz7J92zF27znI/r1zjI3NMzU1g/KEnzLBwpIS20piSbAsyySTLZObyn2zjicOzHLd/REREREREfFmaDQBq4kKAlZgFAuELC+dRy1kGibg2/FiUptpUAo70QRNHYBJkyYr2DoriZHCj5I+OzPH9OQC+/Zk2LF9H4cPHePA/kmmjmWZncmglMBTAikchCWxrCZsh4JTPSi09hAySOGxPNs0IiIiIiKikWg0AasF6Kj8kcnbpIHc9DTZmSmElMvPNCiCfGkC19LEeprBtghSVQcBR8NC1dx8nsx8htmZHDu2j3P40BiH9o9x9Ogsx0anGB+bwc3ZaG2hlclOL4WNlC1Ylkb6yW/9C5ufBad0WRRU9ds34m5ERERERMTJpNEErA6MH1YJRhQQBZOYNX6MmPYTry43LVZhpyDkhaSpdxAZj5PJeGQyWfJZl/175zm4/xjHxiY5eniGo4enGB2dZHoyTT5n4bkSz5VYtsS2bKTswLYg7CClhfaj3AcmvVoB5JZZG0ZERERERDQ4jSZgVQ3RUDCdCWBqEst1i+qeZYTUEqEFSnpgJ9i2L8uLf3wfe44sMDE2w8z0AnPzHm5O43mglcSyLON8LpsRAmwHLFshTayEClnKIw1URERERETEmaTRBKwVVT8JKVtyk2PYyqPRs/ZVQiCQ2kKjEHaSF18d474tr5BzWrGEhRBxtPCwLLCkBKTvG+UhQgKlUdypGrJUpJWKiIiIiIg4UywfAQsKMkPm2CSOq9CNVvslodDCaJ1ywmKKJHYsgbbsQsAJFdyoMAIWOux8blLqmEjpkaYqIiIiIiKiEWk0EaV6FHd/05sAFo5NEPM02tLLbiehFgqNBwhyWjJrxfG0DLmTidA9aYrR0mXo76UE91xe7RIREREREfFWomEELK21pFoUd3MGICCzgDU3t+wEqwDtC1BCK1wkWRFHa1lTHFr8WeB1pQu7EhefsjzbJyIiIiIi4nhpxBmvYQQsIEmVIKNgdsUB6PkpmJ5C5BUad/EuQj+Qudk413hN7gdpQGqPbF4zj4OrQStVv7rCjwimg1AKtQSsk1nriIiIiIiIRkSgFChXNNy810gCVivQVumDINCoBDKex/zgCpJYNUyEgnw2h5s3vkuNFtbJRJ9SzDsdtIlmVlqpgmYLavmt+59oTeGmIg1WRERERMTbFiNg9Q3GsKzG2vjWMLOw1vps4BkqBBpV/v8S0J5Cz2WQ5EBUiDyuzLFcNk8+6ztu6VqJZ84MWoC2JDqeBDtuhCpfaFIN81QiIiIiIiIaH8uSNDXHWIKM9RhwoxDCPdV1aiQNVhsmkvsiZOh/YUlEa7NxemexhBjoc2JUiFgaEREREREREXEaaCQBq5M69QmEp7D1S4Q/KP87kGQbRYEVKNTwo9NHgdQjIiIiIiLekjSSgDVc74Rw0mcohuAU5QKXKPmrYeKRBkEWAn8yy3fILzkhIiIiIiIi4sRoIIVFIwlYtYOMUmw3q/C7aKS2rIvA1H3RwUq/R0RERERERCxbGknAGqx3QrWUxcuJ5V7/iIiIiIiIiPo0iPEMWIKJMCIiIiIiIiJiOdAQApbW2gK6z3Q9IiIiIiIiIiJOBg0hYGGCjPac6UpEREREREREvKU5bRGcGsUHKw68DBw50xWJiIiIiIiIeMuymWjPfkRERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERExLLn/wXbRHvFIAd9agAAAABJRU5ErkJggg=="

    def _logo_from_b64(b64s: str, w_mm: float, h_mm: float):
        try:
            bio = BytesIO(base64.b64decode(b64s))
            im = PILImage.open(bio).convert("RGBA")
            out = BytesIO()
            im.save(out, format="PNG")
            out.seek(0)
            return Image(ImageReader(out), width=w_mm*mm, height=h_mm*mm)
        except Exception:
            return Spacer(1, 1)

    # pega férias cadastradas e filtra por overlap com mês
    rows_f = list_ferias(setor) or []
    df = pd.DataFrame(rows_f, columns=["Chapa", "Início", "Fim"]).copy() if rows_f else pd.DataFrame(columns=["Chapa","Início","Fim"])

    if not df.empty:
        df["Início"] = pd.to_datetime(df["Início"], errors="coerce").dt.date
        df["Fim"] = pd.to_datetime(df["Fim"], errors="coerce").dt.date
        df = df.dropna(subset=["Início", "Fim"])

        ini_mes = pd.Timestamp(year=int(ano), month=int(mes), day=1).date()
        fim_mes = (pd.Timestamp(year=int(ano), month=int(mes), day=1) + pd.offsets.MonthEnd(0)).date()
        df = df[(df["Início"] >= ini_mes) & (df["Início"] <= fim_mes)].copy()

        # Nome ao lado da chapa
        nome_by = {str(c.get("Chapa","")): str(c.get("Nome","") or "") for c in (colaboradores or [])}
        df["Nome"] = df["Chapa"].astype(str).map(nome_by).fillna("")

        # Dias do período total
        df["Dias"] = df.apply(lambda r: int((r["Fim"] - r["Início"]).days + 1), axis=1)

        # filtro por keyword (nome ou chapa)
        if keyword and str(keyword).strip():
            kw = str(keyword).strip().lower()
            df = df[
                df["Nome"].astype(str).str.lower().str.contains(kw, na=False)
                | df["Chapa"].astype(str).str.lower().str.contains(kw, na=False)
            ].copy()

        df = df[["Nome","Chapa","Início","Fim","Dias"]].sort_values(["Início","Nome","Chapa"])

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12*mm, rightMargin=12*mm, topMargin=10*mm, bottomMargin=10*mm
    )
    styles = getSampleStyleSheet()

    # Cabeçalho com 2 logos
    logo_esq = _logo_from_b64(LOGO_SAVEGNAGO_B64, w_mm=42, h_mm=18)
    logo_dir = _logo_from_b64(LOGO_PAULISTAO_B64, w_mm=52, h_mm=18)
    titulo = Paragraph(f"<b>Férias do mês</b> — Setor: {setor} — {int(mes):02d}/{int(ano)}", styles["Title"])

    header = Table([[logo_esq, titulo, logo_dir]], colWidths=[65*mm, 160*mm, 65*mm])
    header.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("ALIGN",(0,0),(0,0),"LEFT"),
        ("ALIGN",(1,0),(1,0),"CENTER"),
        ("ALIGN",(2,0),(2,0),"RIGHT"),
        ("BOTTOMPADDING",(0,0),(-1,-1),6),
    ]))

    elements = [header, Spacer(1, 8)]

    if df.empty:
        elements.append(Paragraph("Nenhum colaborador em férias neste mês.", styles["Normal"]))
    else:
        data = [["Nome", "Chapa", "Início", "Fim", "Dias"]] + df.astype(str).values.tolist()
        tbl = Table(data, repeatRows=1, colWidths=[110*mm, 28*mm, 32*mm, 32*mm, 18*mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,0), 10),
            ("FONTSIZE", (0,1), (-1,-1), 9),
            ("ALIGN", (1,1), (-1,-1), "CENTER"),
            ("ALIGN", (0,0), (0,-1), "LEFT"),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.lightgrey]),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOTTOMPADDING", (0,0), (-1,0), 6),
            ("TOPPADDING", (0,0), (-1,0), 6),
        ]))
        elements.append(tbl)

    # Assinaturas
    elements.append(Spacer(1, 20))
    assin = Table(
        [["______________________________", "______________________________"],
         ["RH", "Gerência"]],
        colWidths=[120*mm, 120*mm]
    )
    assin.setStyle(TableStyle([
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),6),
    ]))
    elements.append(assin)

    doc.build(elements)
    return buf.getvalue()


def _is_fixed_day(status: str) -> bool:
    # FIXO: balanço
    return str(status) == BALANCO_STATUS



def gerar_pdf_trabalhando_no_dia(setor: str, ano: int, mes: int, dia: int, hist_db: dict, colaboradores: list) -> bytes:
    """Gera um PDF simples (A4) listando apenas quem TRABALHA no dia escolhido, com horários."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    # Mapa rápido chapa->(nome, subgrupo)
    meta = {}
    for c in colaboradores:
        meta[str(c.get("Chapa", "")).strip()] = (str(c.get("Nome", "")).strip(), str(c.get("Subgrupo", "")).strip())

    rows = [["Chapa", "Nome", "Subgrupo", "Entrada", "Saída"]]
    for chapa, df in (hist_db or {}).items():
        if df is None or df.empty:
            continue
        try:
            linha = df.loc[df["Data"].dt.day == int(dia)].head(1)
        except Exception:
            # fallback: Data pode estar como string
            linha = df.loc[pd.to_datetime(df["Data"], errors="coerce").dt.day == int(dia)].head(1)
        if linha.empty:
            continue
        r = linha.iloc[0].to_dict()
        stt = str(r.get("Status", "")).strip()
        if stt not in WORK_STATUSES:
            continue
        ent = str(r.get("H_Entrada", "") or "").strip()
        sai = str(r.get("H_Saida", "") or "").strip()
        nome, subg_base = meta.get(str(chapa).strip(), ("", ""))
        subg = get_subgrupo_competencia_ou_base(setor, str(chapa).strip(), int(ano), int(mes), subg_base)
        rows.append([str(chapa).strip(), nome, subg, ent, sai])

    # Ordena por subgrupo e nome (mantendo cabeçalho)
    if len(rows) > 1:
        body = rows[1:]
        body.sort(key=lambda x: (x[2], x[1]))
        rows = [rows[0]] + body

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.2*cm,
        rightMargin=1.2*cm,
        topMargin=1.2*cm,
        bottomMargin=1.2*cm,
    )
    styles = getSampleStyleSheet()
    story = []

    titulo = f"Escala - Quem trabalha no dia {dia:02d}/{mes:02d}/{ano}"
    story.append(Paragraph(f"<b>{titulo}</b>", styles["Title"]))
    story.append(Paragraph(f"Setor: <b>{setor}</b>", styles["Normal"]))
    story.append(Spacer(1, 8))

    table = Table(rows, colWidths=[2.3*cm, 8.2*cm, 4.0*cm, 2.3*cm, 2.3*cm])
    table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 9),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),

        ("FONTNAME", (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,1), (-1,-1), 8),

        ("ALIGN", (0,0), (0,-1), "LEFT"),
        ("ALIGN", (3,1), (4,-1), "CENTER"),

        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.white]),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING", (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
    ]))
    story.append(table)

    doc.build(story)
    return buf.getvalue()



def is_work_status(status: str) -> bool:
    return str(status) in WORK_STATUSES

def _locked(locked_status: set[int] | None, idx: int) -> bool:
    return bool(locked_status and idx in locked_status)

def _ajustar_para_intersticio(ent_desejada: str, saida_anterior: str) -> str:
    """
    Entrada >= desejada respeitando 11:10 após saída anterior
    (considera dia seguinte quando necessário)
    """
    if not ent_desejada or not saida_anterior:
        return ent_desejada

    s = _to_min(saida_anterior)
    e_des = _to_min(ent_desejada)
    e_min = _to_min(_add_min(saida_anterior, INTERSTICIO_MIN))

    if e_des <= s:
        e_des += 1440
    if e_min <= s:
        e_min += 1440

    e_ok = max(e_des, e_min)
    return _min_to_hhmm(e_ok)

# =========================================================
# ✅ Proibir folga consecutiva AUTOMÁTICA (DOM+SEG etc.)
# Só permite se estiver travado (override/manual/"caixinha")
# =========================================================
def enforce_no_consecutive_folga(df: pd.DataFrame, locked_status: set[int] | None = None):
    """
    Proibir folga consecutiva automática (DOM+SEG etc.).
    Robustez: garante índice 0..N-1 e usa iloc para não dar KeyError.
    """
    df.reset_index(drop=True, inplace=True)
    for i in range(1, len(df)):
        if df.iloc[i - 1]["Status"] == "Folga" and df.iloc[i]["Status"] == "Folga":
            prev_locked = _locked(locked_status, i - 1)
            cur_locked = _locked(locked_status, i)

            # ambos travados => foi decisão manual, mantém
            if prev_locked and cur_locked:
                continue

            # prioriza manter o travado e desfazer o outro
            if not cur_locked:
                df.iloc[i, df.columns.get_loc("Status")] = "Trabalho"
            elif not prev_locked:
                df.iloc[i - 1, df.columns.get_loc("Status")] = "Trabalho"
# =========================================================
# DB
# =========================================================
def db_conn():
    _ensure_runtime_storage_ready()
    conn = _app_db_connect(DB_PATH)
    return conn

def _norm_setor(v: str) -> str:
    return str(v or "").strip().upper()

def _norm_chapa(v: str) -> str:
    return str(v or "").strip()

def hash_password(password: str, salt: str) -> str:
    password = (password or "").strip()
    salt = (salt or "").strip()
    return hashlib.sha256((salt + password).encode("utf-8")).hexdigest()

def hash_password_legacy(password: str, salt: str) -> str:
    password = (password or "").strip()
    salt = (salt or "").strip()
    return hashlib.sha256((password + salt).encode("utf-8")).hexdigest()

def verify_password_compat(password: str, senha_hash_db: str, salt_db: str) -> bool:
    password = (password or "").strip()
    senha_hash_db = (senha_hash_db or "").strip()
    salt_db = (salt_db or "").strip()

    if not senha_hash_db:
        return False

    # Formato atual
    if hash_password(password, salt_db) == senha_hash_db:
        return True

    # Compatibilidade com versão antiga (ordem invertida)
    if hash_password_legacy(password, salt_db) == senha_hash_db:
        return True

    # Compatibilidade com bases antigas onde a senha pode ter ficado salva sem hash
    if password == senha_hash_db:
        return True

    return False

def _safe_exec(cur, sql: str, params=None):
    try:
        if params is None:
            cur.execute(sql)
        else:
            cur.execute(sql, params)
    except Exception:
        pass

def _ensure_usuarios_sistema_security_columns(cur):
    try:
        cur.execute("ALTER TABLE usuarios_sistema ADD COLUMN forcar_troca_senha INTEGER NOT NULL DEFAULT 0")
    except Exception:
        pass
    try:
        cur.execute("ALTER TABLE usuarios_sistema ADD COLUMN is_ax_lider INTEGER NOT NULL DEFAULT 0")
    except Exception:
        pass


def gerar_senha_temporaria_colaborador(tamanho: int = 8) -> str:
    alfabeto = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return "".join(secrets.choice(alfabeto) for _ in range(max(6, int(tamanho))))


def db_init_fast_login():
    """Inicialização mínima para abrir a tela de login rápido sem alterar regras."""
    con = db_conn()
    cur = con.cursor()

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS setores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT UNIQUE NOT NULL
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS usuarios_sistema (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        senha_hash TEXT NOT NULL,
        salt TEXT NOT NULL,
        is_admin INTEGER NOT NULL DEFAULT 0,
        is_lider INTEGER NOT NULL DEFAULT 0,
        criado_em TEXT NOT NULL,
        UNIQUE(setor, chapa)
    )
    """)
    _ensure_usuarios_sistema_security_columns(cur)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS colaboradores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        subgrupo TEXT DEFAULT '',
        entrada TEXT DEFAULT '06:00',
        folga_sab INTEGER DEFAULT 0,
        criado_em TEXT NOT NULL,
        UNIQUE(setor, chapa)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS login_recent (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        ts TEXT NOT NULL
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS ax_lider_aprovacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        setor_alvo TEXT NOT NULL,
        chapa_alvo TEXT NOT NULL,
        nome_alvo TEXT,
        nome_novo TEXT,
        subgrupo_novo TEXT,
        perfil_novo TEXT,
        entrada_nova TEXT,
        folga_sab_nova INTEGER DEFAULT 0,
        criado_por_nome TEXT,
        criado_por_chapa TEXT,
        status TEXT NOT NULL DEFAULT 'PENDENTE',
        observacao TEXT,
        aprovado_por TEXT,
        aprovado_em TEXT,
        criado_em TEXT NOT NULL
    )
    """)

    _safe_exec(cur, "INSERT OR IGNORE INTO setores(nome) VALUES (?)", ("GERAL",))
    _safe_exec(cur, "INSERT OR IGNORE INTO setores(nome) VALUES (?)", ("ADMIN",))
    _safe_exec(cur, "INSERT OR IGNORE INTO setores(nome) VALUES (?)", ("GESTAO",))

    cur.execute("SELECT 1 FROM usuarios_sistema WHERE setor=? AND chapa=? LIMIT 1", ("ADMIN", "admin"))
    if cur.fetchone() is None:
        salt = secrets.token_hex(16)
        senha_hash = hash_password("123", salt)
        cur.execute("""
            INSERT INTO usuarios_sistema(nome, setor, chapa, senha_hash, salt, is_admin, is_lider, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, ("Administrador", "ADMIN", "admin", senha_hash, salt, 1, 1, datetime.now().isoformat()))

    con.commit()
    con.close()


def db_init():
    con = db_conn()
    cur = con.cursor()

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS setores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT UNIQUE NOT NULL
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS usuarios_sistema (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        senha_hash TEXT NOT NULL,
        salt TEXT NOT NULL,
        is_admin INTEGER NOT NULL DEFAULT 0,
        is_lider INTEGER NOT NULL DEFAULT 0,
        criado_em TEXT NOT NULL,
        UNIQUE(setor, chapa)
    )
    """)
    _ensure_usuarios_sistema_security_columns(cur)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS colaboradores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        subgrupo TEXT DEFAULT '',
        entrada TEXT DEFAULT '06:00',
        folga_sab INTEGER DEFAULT 0,
        criado_em TEXT NOT NULL,
        UNIQUE(setor, chapa)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS subgrupos_setor (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        nome TEXT NOT NULL,
        UNIQUE(setor, nome)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS subgrupo_regras (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        subgrupo TEXT NOT NULL,
        evitar_seg INTEGER NOT NULL DEFAULT 0,
        evitar_ter INTEGER NOT NULL DEFAULT 0,
        evitar_qua INTEGER NOT NULL DEFAULT 0,
        evitar_qui INTEGER NOT NULL DEFAULT 0,
        evitar_sex INTEGER NOT NULL DEFAULT 0,
        evitar_sab INTEGER NOT NULL DEFAULT 0,
        UNIQUE(setor, subgrupo)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS ferias (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        inicio TEXT NOT NULL,
        fim TEXT NOT NULL
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS estado_mes_anterior (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        consec_trab_final INTEGER NOT NULL,
        ultima_saida TEXT NOT NULL,
        ultimo_domingo_status TEXT,
        ano INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        UNIQUE(setor, chapa, ano, mes)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS escala_mes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        ano INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        chapa TEXT NOT NULL,
        dia INTEGER NOT NULL,
        data TEXT NOT NULL,
        dia_sem TEXT NOT NULL,
        status TEXT NOT NULL,
        h_entrada TEXT,
        h_saida TEXT,
        UNIQUE(setor, ano, mes, chapa, dia)
    )
    """)

    # --- MIGRAÇÃO defensiva (Streamlit Cloud pode manter DB antigo)
    # Garante que a tabela escala_mes tenha todas as colunas esperadas
    try:
        cur.execute("PRAGMA table_info(escala_mes)")
        cols = {r[1] for r in cur.fetchall()}  # r[1] = name

        expected = {"setor","ano","mes","chapa","dia","data","dia_sem","status","h_entrada","h_saida"}
        missing = expected - cols

        for c in sorted(missing):
            # tipos simples (compatível com SQLite)
            if c in ("ano","mes","dia"):
                cur.execute(f"ALTER TABLE escala_mes ADD COLUMN {c} INTEGER")
            else:
                cur.execute(f"ALTER TABLE escala_mes ADD COLUMN {c} TEXT")

        con.commit()
    except Exception:
        # Não interrompe a inicialização caso o DB já esteja OK ou a migração falhe
        pass

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS login_recent (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        ts TEXT NOT NULL
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS portal_assinaturas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        ano INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        tipo TEXT NOT NULL,
        versao_ref INTEGER NOT NULL DEFAULT 1,
        assinado_em TEXT NOT NULL,
        UNIQUE(setor, chapa, ano, mes, tipo)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS portal_solicitacoes_folga (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        data_solicitada TEXT NOT NULL,
        tipo TEXT NOT NULL,
        motivo TEXT,
        observacao TEXT,
        status TEXT NOT NULL DEFAULT 'Em análise',
        criado_em TEXT NOT NULL,
        atualizado_em TEXT NOT NULL
    )
    """)


    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS ax_lider_aprovacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        setor_alvo TEXT NOT NULL,
        chapa_alvo TEXT NOT NULL,
        nome_alvo TEXT,
        nome_novo TEXT,
        subgrupo_novo TEXT,
        perfil_novo TEXT,
        entrada_nova TEXT,
        folga_sab_nova INTEGER DEFAULT 0,
        criado_por_nome TEXT,
        criado_por_chapa TEXT,
        status TEXT NOT NULL DEFAULT 'PENDENTE',
        observacao TEXT,
        aprovado_por TEXT,
        aprovado_em TEXT,
        criado_em TEXT NOT NULL
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS rodizio_caixa_cfg (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        subgrupo_origem TEXT NOT NULL,
        subgrupo_destino TEXT NOT NULL,
        qtd_destino INTEGER NOT NULL DEFAULT 12,
        tolerancia_min INTEGER NOT NULL DEFAULT 20,
        ativo INTEGER NOT NULL DEFAULT 1,
        UNIQUE(setor, subgrupo_origem, subgrupo_destino)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS rodizio_caixa_hist (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        ano INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        ciclo_ref TEXT NOT NULL,
        chapa TEXT NOT NULL,
        nome TEXT NOT NULL,
        movimento TEXT NOT NULL,
        subgrupo_origem TEXT NOT NULL,
        subgrupo_destino TEXT NOT NULL,
        entrada_antiga TEXT,
        entrada_nova TEXT,
        compat_status TEXT NOT NULL DEFAULT 'IGUAL',
        observacao TEXT,
        criado_em TEXT NOT NULL,
        UNIQUE(setor, ano, mes, chapa, movimento)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS overrides (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        ano INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        chapa TEXT NOT NULL,
        dia INTEGER NOT NULL,
        campo TEXT NOT NULL,
        valor TEXT NOT NULL,
        UNIQUE(setor, ano, mes, chapa, dia, campo)
    )
    """)


    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS folga_fixa (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        dia_semana INTEGER NOT NULL,
        ativo INTEGER NOT NULL DEFAULT 1,
        criado_em TEXT NOT NULL,
        UNIQUE(setor, chapa, dia_semana)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS inventario_diario (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        ano INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        dia INTEGER NOT NULL,
        abertura INTEGER NOT NULL DEFAULT 0,
        intermediario INTEGER NOT NULL DEFAULT 0,
        fechamento INTEGER NOT NULL DEFAULT 0,
        criado_em TEXT NOT NULL,
        atualizado_em TEXT NOT NULL,
        UNIQUE(setor, ano, mes, dia)
    )
    """)


    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS competencia_status (
        setor TEXT NOT NULL,
        ano INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        status TEXT NOT NULL DEFAULT 'ABERTA',
        atualizado_em TEXT NOT NULL,
        PRIMARY KEY (setor, ano, mes)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS retificacoes_competencia (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        ano INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        chapa TEXT NOT NULL,
        nome TEXT,
        dia INTEGER NOT NULL,
        novo_status TEXT,
        nova_entrada TEXT,
        nova_saida TEXT,
        novo_subgrupo TEXT,
        motivo TEXT,
        usuario TEXT,
        criado_em TEXT NOT NULL,
        UNIQUE(setor, ano, mes, chapa, dia)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS subgrupo_competencia (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        ano INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        chapa TEXT NOT NULL,
        subgrupo TEXT NOT NULL,
        atualizado_em TEXT NOT NULL,
        UNIQUE(setor, ano, mes, chapa)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS assinaturas_retificacao (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        ano INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        chapa TEXT NOT NULL,
        assinado_em TEXT NOT NULL,
        UNIQUE(setor, ano, mes, chapa)
    )
    """)

    # V91.1 — restore do Supabase antes de semear dados mínimos de login.
    # Isso evita o cenário em que o app recria apenas ADMIN/GERAL/GESTAO e
    # sobe “vazio” após reboot antes de puxar o conteúdo real do remoto.
    con.commit()
    try:
        con.close()
    except Exception:
        pass

    restore_ok = False
    if SUPABASE_AUTO_RESTORE_IF_LOCAL_EMPTY:
        try:
            restore_ok = bool(_restore_from_supabase_if_local_empty(force=True))
        except Exception as e:
            _set_supabase_error(e)

    con = db_conn()
    cur = con.cursor()

    con.commit()

    # v97 force: nunca bloquear nem exibir guarda de restore no login
    _set_restore_guard(False, "")
    _safe_exec(cur, "INSERT OR IGNORE INTO setores(nome) VALUES (?)", ("GERAL",))
    _safe_exec(cur, "INSERT OR IGNORE INTO setores(nome) VALUES (?)", ("ADMIN",))
    _safe_exec(cur, "INSERT OR IGNORE INTO setores(nome) VALUES (?)", ("GESTAO",))
    con.commit()

    cur.execute("SELECT 1 FROM usuarios_sistema WHERE setor=? AND chapa=? LIMIT 1", ("ADMIN", "admin"))
    if cur.fetchone() is None:
        salt = secrets.token_hex(16)
        senha_hash = hash_password("123", salt)
        cur.execute("""
            INSERT INTO usuarios_sistema(nome, setor, chapa, senha_hash, salt, is_admin, is_lider, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, ("Administrador", "ADMIN", "admin", senha_hash, salt, 1, 1, datetime.now().isoformat()))
        con.commit()

    if SUPABASE_AUTO_RESTORE_IF_LOCAL_EMPTY:
        try:
            _restore_from_supabase_if_local_empty(force=False)
        except Exception:
            pass

    if SUPABASE_AUTO_BOOTSTRAP_AFTER_SCHEMA:
        try:
            _bootstrap_from_supabase_after_schema(force=False)
        except Exception:
            pass
    con.commit()

    con.close()


def is_past_competencia(ano: int, mes: int) -> bool:
    """Meses anteriores ao mês atual (no fuso do servidor)."""
    today = date.today()
    return (int(ano), int(mes)) < (int(today.year), int(today.month))


# =========================================================
# AUTH
# =========================================================
def system_user_exists(setor: str, chapa: str) -> bool:
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT 1 FROM usuarios_sistema WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=? LIMIT 1", (setor, chapa))
    ok = cur.fetchone() is not None
    con.close()
    return ok

def colaborador_lookup(setor: str, chapa: str):
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    con = db_conn()
    cur = con.cursor()
    cur.execute(
        """
        SELECT nome, setor, chapa
        FROM colaboradores
        WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?
        LIMIT 1
        """,
        (setor, chapa),
    )
    row = cur.fetchone()
    con.close()
    return row

def default_password_for_chapa(chapa: str) -> str:
    chapa = _norm_chapa(chapa)
    nums = re.sub(r"\D+", "", chapa)
    return nums or chapa or "123456"


def ensure_system_user_from_colaborador(nome: str, setor: str, chapa: str, senha_padrao: str | None = None, is_lider: int = 0, is_admin: int = 0, is_ax_lider: int = 0):
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    nome = (nome or "").strip() or chapa
    if system_user_exists(setor, chapa):
        return False
    senha_final = (senha_padrao or default_password_for_chapa(chapa)).strip()
    create_system_user(nome, setor, chapa, senha_final, is_lider=is_lider, is_admin=is_admin, is_ax_lider=is_ax_lider)
    return True


def create_system_user(nome: str, setor: str, chapa: str, senha: str, is_lider: int = 0, is_admin: int = 0, is_ax_lider: int = 0):
    nome = (nome or "").strip() or _norm_chapa(chapa)
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    senha = (senha or "").strip()
    salt = secrets.token_hex(16)
    senha_hash = hash_password(senha, salt)
    con = db_conn()
    cur = con.cursor()
    cur.execute("INSERT OR IGNORE INTO setores(nome) VALUES (?)", (setor,))
    cur.execute(
        """
        SELECT id FROM usuarios_sistema
        WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?
        LIMIT 1
        """,
        (setor, chapa),
    )
    row = cur.fetchone()
    if row:
        cur.execute(
            """
            UPDATE usuarios_sistema
            SET nome=?, setor=?, chapa=?, senha_hash=?, salt=?, is_admin=?, is_lider=?, is_ax_lider=?
            WHERE id=?
            """,
            (nome, setor, chapa, senha_hash, salt, int(is_admin), int(is_lider), int(is_ax_lider), int(row[0])),
        )
    else:
        cur.execute(
            """
            INSERT INTO usuarios_sistema(nome, setor, chapa, senha_hash, salt, is_admin, is_lider, is_ax_lider, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (nome, setor, chapa, senha_hash, salt, int(is_admin), int(is_lider), int(is_ax_lider), datetime.now().isoformat()),
        )
    con.commit()
    con.close()

def recover_system_user_from_colaborador(setor: str, chapa: str, senha: str, is_lider: int = 0, is_admin: int = 0, is_ax_lider: int = 0):
    row = colaborador_lookup(setor, chapa)
    if not row:
        return False
    nome, setor_db, chapa_db = row
    create_system_user(nome or chapa_db, setor_db, chapa_db, senha, is_lider=is_lider, is_admin=is_admin, is_ax_lider=is_ax_lider)
    return True

def verify_login(setor: str, chapa: str, senha: str):
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    con = db_conn()
    cur = con.cursor()
    try:
        try:
            cur.execute(
                """
                SELECT nome, senha_hash, salt, is_admin, is_lider, COALESCE(is_ax_lider,0), setor, chapa, COALESCE(forcar_troca_senha,0)
                FROM usuarios_sistema
                WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?
                LIMIT 1
                """,
                (setor, chapa),
            )
        except sqlite3.OperationalError:
            _ensure_usuarios_sistema_security_columns(cur)
            try:
                con.commit()
            except Exception:
                pass
            cur.execute(
                """
                SELECT nome, senha_hash, salt, is_admin, is_lider, COALESCE(is_ax_lider,0), setor, chapa, 0
                FROM usuarios_sistema
                WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?
                LIMIT 1
                """,
                (setor, chapa),
            )
        row = cur.fetchone()
    finally:
        con.close()
    if not row:
        return None
    nome, senha_hash, salt, is_admin, is_lider, is_ax_lider, setor_db, chapa_db, forcar_troca_senha = row
    if verify_password_compat(senha, senha_hash, salt):
        return {
            "nome": nome,
            "setor": _norm_setor(setor_db),
            "chapa": _norm_chapa(chapa_db),
            "is_admin": bool(is_admin),
            "is_lider": bool(is_lider),
            "is_ax_lider": bool(is_ax_lider),
            "forcar_troca_senha": bool(forcar_troca_senha),
        }
    return None

def is_lider_chapa(setor: str, chapa_lider: str) -> bool:
    setor = _norm_setor(setor)
    chapa_lider = _norm_chapa(chapa_lider)
    con = db_conn()
    cur = con.cursor()
    cur.execute(
        "SELECT is_lider FROM usuarios_sistema WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=? LIMIT 1",
        (setor, chapa_lider),
    )
    row = cur.fetchone()
    con.close()
    return bool(row and row[0] == 1)

def update_password(setor: str, chapa: str, nova_senha: str):
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    nova_senha = (nova_senha or "").strip()
    salt = secrets.token_hex(16)
    senha_hash = hash_password(nova_senha, salt)
    con = db_conn()
    cur = con.cursor()
    cur.execute(
        "UPDATE usuarios_sistema SET senha_hash=?, salt=? WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?",
        (senha_hash, salt, setor, chapa),
    )
    con.commit()
    con.close()

def set_force_change_password(setor: str, chapa: str, ativo: bool = True):
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute(
            "UPDATE usuarios_sistema SET forcar_troca_senha=? WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?",
            (1 if ativo else 0, setor, chapa),
        )
        con.commit()
    finally:
        con.close()


def upsert_usuario_sistema(nome: str, setor: str, chapa: str, senha: str, is_admin: bool = False, is_lider: bool = False, is_ax_lider: bool = False, forcar_troca_senha: bool = False):
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    nome = (nome or "").strip() or chapa
    salt = secrets.token_hex(16)
    senha_hash = hash_password((senha or "").strip(), salt)
    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute(
            """
            INSERT INTO usuarios_sistema(nome, setor, chapa, senha_hash, salt, is_admin, is_lider, is_ax_lider, criado_em, forcar_troca_senha)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(setor, chapa) DO UPDATE SET
                nome=excluded.nome,
                senha_hash=excluded.senha_hash,
                salt=excluded.salt,
                is_admin=excluded.is_admin,
                is_lider=excluded.is_lider,
                is_ax_lider=excluded.is_ax_lider,
                forcar_troca_senha=excluded.forcar_troca_senha
            """,
            (nome, setor, chapa, senha_hash, salt, 1 if is_admin else 0, 1 if is_lider else 0, 1 if is_ax_lider else 0, datetime.now().isoformat(), 1 if forcar_troca_senha else 0),
        )
        con.commit()
    finally:
        con.close()


def get_usuario_sistema_por_setor_chapa(setor: str, chapa: str):
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    con = db_conn()
    cur = con.cursor()
    try:
        try:
            cur.execute(
                """
                SELECT nome, setor, chapa, is_admin, is_lider, COALESCE(is_ax_lider,0), COALESCE(forcar_troca_senha,0)
                FROM usuarios_sistema
                WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?
                LIMIT 1
                """,
                (setor, chapa),
            )
        except sqlite3.OperationalError:
            _ensure_usuarios_sistema_security_columns(cur)
            try:
                con.commit()
            except Exception:
                pass
            cur.execute(
                """
                SELECT nome, setor, chapa, is_admin, is_lider, COALESCE(is_ax_lider,0), 0
                FROM usuarios_sistema
                WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?
                LIMIT 1
                """,
                (setor, chapa),
            )
        row = cur.fetchone()
    finally:
        con.close()
    if not row:
        return None
    return {
        "nome": str(row[0] or "").strip(),
        "setor": _norm_setor(row[1]),
        "chapa": _norm_chapa(row[2]),
        "is_admin": bool(row[3]),
        "is_lider": bool(row[4]),
        "is_ax_lider": bool(row[5]),
        "forcar_troca_senha": bool(row[6]),
    }
# =========================================================
# ADMIN
# =========================================================
@st.cache_data(show_spinner=False)


def registrar_solicitacao_ax_lider(setor_solicitante: str, setor_alvo: str, chapa_alvo: str, nome_alvo: str,
                                   nome_novo: str, subgrupo_novo: str, perfil_novo: str,
                                   entrada_nova: str, folga_sab_nova: bool,
                                   criado_por_nome: str, criado_por_chapa: str, observacao: str = '') -> int:
    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute("""
            INSERT INTO ax_lider_aprovacoes(
                setor, setor_alvo, chapa_alvo, nome_alvo, nome_novo, subgrupo_novo, perfil_novo,
                entrada_nova, folga_sab_nova, criado_por_nome, criado_por_chapa,
                status, observacao, criado_em
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'PENDENTE', ?, ?)
        """, (
            _norm_setor(setor_solicitante), _norm_setor(setor_alvo), _norm_chapa(chapa_alvo),
            str(nome_alvo or '').strip(), str(nome_novo or '').strip(), str(subgrupo_novo or '').strip(),
            _norm_subgrupo_label(perfil_novo or 'COLABORADOR'), str(entrada_nova or '06:00').strip() or '06:00',
            1 if bool(folga_sab_nova) else 0, str(criado_por_nome or '').strip(), _norm_chapa(criado_por_chapa),
            str(observacao or '').strip(), datetime.now().isoformat()
        ))
        con.commit()
        return int(cur.lastrowid)
    finally:
        con.close()

@st.cache_data(show_spinner=False, ttl=60)
def listar_solicitacoes_ax_lider(status: str | None = None, setor_alvo: str | None = None) -> pd.DataFrame:
    con = db_conn()
    params = []
    sql = """
        SELECT id, setor, setor_alvo, chapa_alvo, nome_alvo, nome_novo, subgrupo_novo, perfil_novo,
               entrada_nova, folga_sab_nova, criado_por_nome, criado_por_chapa, status,
               observacao, aprovado_por, aprovado_em, criado_em
        FROM ax_lider_aprovacoes
        WHERE 1=1
    """
    if status:
        sql += " AND UPPER(TRIM(status))=?"
        params.append(str(status).strip().upper())
    if setor_alvo:
        sql += " AND UPPER(TRIM(setor_alvo))=?"
        params.append(_norm_setor(setor_alvo))
    sql += " ORDER BY CASE UPPER(TRIM(status)) WHEN 'PENDENTE' THEN 0 ELSE 1 END, id DESC"
    try:
        return pd.read_sql_query(sql, con, params=params)
    finally:
        con.close()

def decidir_solicitacao_ax_lider(solicitacao_id: int, aprovador_nome: str, aprovar: bool, observacao_aprovador: str = ''):
    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute("""
            SELECT setor_alvo, chapa_alvo, COALESCE(nome_novo,''), COALESCE(subgrupo_novo,''),
                   COALESCE(perfil_novo,'COLABORADOR'), COALESCE(entrada_nova,'06:00'),
                   COALESCE(folga_sab_nova,0), COALESCE(status,'PENDENTE')
            FROM ax_lider_aprovacoes
            WHERE id=?
            LIMIT 1
        """, (int(solicitacao_id),))
        row = cur.fetchone()
        if not row:
            raise ValueError('Solicitação AX não encontrada.')
        setor_alvo, chapa_alvo, nome_novo, subgrupo_novo, perfil_novo, entrada_nova, folga_sab_nova, status_atual = row
        if str(status_atual or '').strip().upper() != 'PENDENTE':
            raise ValueError('Essa solicitação já foi decidida.')
        novo_status = 'APROVADO' if aprovar else 'REPROVADO'
        cur.execute("""
            UPDATE ax_lider_aprovacoes
            SET status=?, aprovado_por=?, aprovado_em=?, observacao=TRIM(COALESCE(observacao,'') || CASE WHEN ?<>'' THEN ' | ' || ? ELSE '' END)
            WHERE id=?
        """, (novo_status, str(aprovador_nome or '').strip(), datetime.now().isoformat(),
                str(observacao_aprovador or '').strip(), str(observacao_aprovador or '').strip(), int(solicitacao_id)))
        con.commit()
    finally:
        con.close()

    if aprovar:
        admin_update_funcionario(
            setor=str(setor_alvo),
            chapa_atual=str(chapa_alvo),
            nome_novo=str(nome_novo),
            subgrupo_novo=str(subgrupo_novo),
            perfil_novo=str(perfil_novo),
            entrada_nova=str(entrada_nova),
            folga_sab=bool(int(folga_sab_nova or 0)),
            criar_usuario_se_nao_existir=True,
        )
    try:
        st.cache_data.clear()
    except Exception:
        pass

@st.cache_data(show_spinner=False, ttl=300)
def admin_list_users():
    con = db_conn()
    df = pd.read_sql_query("""
        SELECT id, nome, setor, chapa, COALESCE(is_admin,0) AS is_admin, COALESCE(is_lider,0) AS is_lider, COALESCE(is_ax_lider,0) AS is_ax_lider, criado_em
        FROM usuarios_sistema
        ORDER BY setor ASC, nome ASC
    """, con)
    con.close()
    return df


@st.cache_data(show_spinner=False, ttl=300)
def admin_get_funcionarios_leve_all() -> pd.DataFrame:
    con = db_conn()
    try:
        return pd.read_sql_query(
            """
            SELECT
                COALESCE(nome,'') AS nome,
                UPPER(TRIM(COALESCE(setor,''))) AS setor,
                TRIM(COALESCE(chapa,'')) AS chapa,
                COALESCE(subgrupo,'') AS subgrupo,
                COALESCE(entrada,'06:00') AS entrada,
                COALESCE(folga_sab,0) AS folga_sab
            FROM colaboradores
            ORDER BY setor ASC, nome ASC
            """,
            con,
        )
    finally:
        con.close()


@st.cache_data(show_spinner=False, ttl=300)
def admin_get_logins_leve_all() -> pd.DataFrame:
    con = db_conn()
    try:
        return pd.read_sql_query(
            """
            SELECT
                UPPER(TRIM(COALESCE(setor,''))) AS setor,
                TRIM(COALESCE(chapa,'')) AS chapa,
                COALESCE(is_admin,0) AS is_admin,
                COALESCE(is_lider,0) AS is_lider,
                COALESCE(is_ax_lider,0) AS is_ax_lider
            FROM usuarios_sistema
            """,
            con,
        )
    finally:
        con.close()

def admin_reset_user_password(user_id: int, nova_senha: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT setor, chapa FROM usuarios_sistema WHERE id=?", (int(user_id),))
    row = cur.fetchone()
    if not row:
        con.close()
        return False
    setor, chapa = row
    con.close()
    update_password(setor, chapa, nova_senha)
    return True


def admin_update_funcionario(setor: str, chapa_atual: str, nome_novo: str, subgrupo_novo: str, perfil_novo: str, entrada_nova: str = '06:00', folga_sab: bool = False, criar_usuario_se_nao_existir: bool = True):
    """
    Atualiza cadastro do colaborador e sincroniza o usuário do sistema.
    Perfil aceito: COLABORADOR, AX_LIDER, LIDER, ADMIN.
    """
    setor = _norm_setor(setor)
    chapa_atual = _norm_chapa(chapa_atual)
    nome_novo = str(nome_novo or '').strip()
    subgrupo_novo = str(subgrupo_novo or '').strip()
    perfil_novo = _norm_subgrupo_label(perfil_novo or 'COLABORADOR')
    entrada_nova = str(entrada_nova or '06:00').strip() or '06:00'

    if not setor or not chapa_atual:
        raise ValueError('Setor e chapa são obrigatórios.')

    rec = get_colaborador_record(setor, chapa_atual)
    if not rec:
        raise ValueError('Colaborador não encontrado para este setor/chapa.')

    nome_final = nome_novo or str(rec.get('Nome') or chapa_atual).strip()
    subgrupo_final = subgrupo_novo or str(rec.get('Subgrupo') or '').strip()
    entrada_final = entrada_nova or str(rec.get('Entrada') or '06:00').strip() or '06:00'

    if perfil_novo not in ['COLABORADOR', 'AX_LIDER', 'LIDER', 'ADMIN']:
        perfil_novo = 'COLABORADOR'

    if perfil_novo in ['AX_LIDER', 'LIDER'] and not subgrupo_final:
        subgrupo_final = 'LIDERANÇA'
    if perfil_novo == 'ADMIN' and not subgrupo_final:
        subgrupo_final = 'ADMIN'

    update_colaborador_perfil(
        setor=setor,
        chapa_antiga=chapa_atual,
        chapa_nova=chapa_atual,
        nome_novo=nome_final,
        subgrupo=subgrupo_final,
        entrada=entrada_final,
        folga_sab=bool(folga_sab),
    )

    is_admin = 1 if perfil_novo == 'ADMIN' else 0
    is_lider = 1 if perfil_novo in ['LIDER', 'ADMIN'] else 0
    is_ax_lider = 1 if perfil_novo == 'AX_LIDER' else 0

    con = db_conn()
    cur = con.cursor()
    cur.execute(
        """
        SELECT id FROM usuarios_sistema
        WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?
        LIMIT 1
        """,
        (setor, chapa_atual),
    )
    row = cur.fetchone()
    con.close()

    if row:
        con = db_conn()
        cur = con.cursor()
        cur.execute(
            """
            UPDATE usuarios_sistema
            SET nome=?, is_admin=?, is_lider=?, is_ax_lider=?
            WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?
            """,
            (nome_final, int(is_admin), int(is_lider), int(is_ax_lider), setor, chapa_atual),
        )
        con.commit()
        con.close()
    elif criar_usuario_se_nao_existir:
        senha_padrao = default_password_for_chapa(chapa_atual)
        create_system_user(nome_final, setor, chapa_atual, senha_padrao, is_lider=int(is_lider), is_admin=int(is_admin), is_ax_lider=int(is_ax_lider))

    try:
        st.cache_data.clear()
    except Exception:
        pass

    return {
        'nome': nome_final,
        'setor': setor,
        'chapa': chapa_atual,
        'subgrupo': subgrupo_final,
        'perfil': perfil_novo,
        'entrada': entrada_final,
        'folga_sab': bool(folga_sab),
        'is_admin': bool(is_admin),
        'is_lider': bool(is_lider),
        'is_ax_lider': bool(is_ax_lider),
    }

# =========================================================
# COLABORADORES
# =========================================================

def admin_rename_setor_global(setor_atual: str, setor_novo: str) -> dict:
    """Renomeia um setor em todas as tabelas que possuem a coluna 'setor'."""
    setor_atual_norm = _norm_setor(setor_atual)
    setor_novo_norm = _norm_setor(setor_novo)

    if not setor_atual_norm or not setor_novo_norm:
        raise ValueError('Informe o setor atual e o novo nome do setor.')
    if setor_atual_norm == setor_novo_norm:
        raise ValueError('O novo nome do setor precisa ser diferente do setor atual.')

    con = db_conn()
    cur = con.cursor()
    atualizados = []
    try:
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name")
        tabelas = [str(r[0]) for r in cur.fetchall() if str(r[0]).strip()]

        for tabela in tabelas:
            try:
                cur.execute(f"PRAGMA table_info({tabela})")
                cols = [str(r[1]) for r in cur.fetchall()]
            except Exception:
                continue

            if 'setor' not in cols:
                continue

            try:
                cur.execute(
                    f"UPDATE {tabela} SET setor=? WHERE UPPER(TRIM(setor))=UPPER(TRIM(?))",
                    (setor_novo_norm, setor_atual_norm),
                )
                if int(cur.rowcount or 0) > 0:
                    atualizados.append((tabela, int(cur.rowcount or 0)))
            except Exception:
                raise

        con.commit()
    finally:
        con.close()

    try:
        st.cache_data.clear()
    except Exception:
        pass

    return {
        'setor_antigo': setor_atual_norm,
        'setor_novo': setor_novo_norm,
        'tabelas_atualizadas': atualizados,
        'total_tabelas': len(atualizados),
        'total_registros': sum(q for _, q in atualizados),
    }

def colaborador_exists(setor: str, chapa: str) -> bool:
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT 1 FROM colaboradores WHERE setor=? AND chapa=? LIMIT 1", (setor, chapa))
    ok = cur.fetchone() is not None
    con.close()
    return ok

def create_colaborador(nome: str, setor: str, chapa: str, subgrupo: str = "", entrada: str = "06:00", folga_sab: bool = False, criar_login: bool = True, senha_padrao: str | None = None):
    """
    Cria colaborador (se não existir) já com perfil completo.
    Mantém compatibilidade: parâmetros adicionais são opcionais.
    Se criar_login=True, também garante usuário em usuarios_sistema.
    """
    nome = (nome or "").strip()
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    subgrupo = (subgrupo or "").strip()
    entrada = (entrada or "06:00").strip()

    con = db_conn()
    cur = con.cursor()
    # cria (ou ignora) com os campos completos
    cur.execute(
        """
        INSERT OR IGNORE INTO colaboradores(nome, setor, chapa, subgrupo, entrada, folga_sab, criado_em)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (nome, setor, chapa, subgrupo, entrada, 1 if folga_sab else 0, datetime.now().isoformat()),
    )
    # se já existia, não força overwrite de tudo; mas se os campos estão vazios, completa
    cur.execute(
        """
        UPDATE colaboradores
        SET nome = COALESCE(NULLIF(nome,''), ?),
            subgrupo = CASE WHEN (subgrupo IS NULL OR TRIM(subgrupo)='') THEN ? ELSE subgrupo END,
            entrada = CASE WHEN (entrada IS NULL OR TRIM(entrada)='') THEN ? ELSE entrada END,
            folga_sab = CASE WHEN folga_sab IS NULL THEN ? ELSE folga_sab END
        WHERE setor=? AND chapa=?
        """,
        (nome, subgrupo, entrada, 1 if folga_sab else 0, setor, chapa),
    )
    con.commit()
    con.close()
    if criar_login:
        try:
            ensure_system_user_from_colaborador(nome or chapa, setor, chapa, senha_padrao=senha_padrao)
        except Exception:
            pass
    try:
        st.cache_data.clear()
    except Exception:
        pass


def upsert_colaborador_nome(setor: str, chapa: str, nome: str):
    """
    Garante que existe o colaborador (SEM senha) e atualiza o nome pelo que veio na base manual.
    - Se a chapa existir: atualiza nome.
    - Se não existir: cria.
    """
    nome = (nome or "").strip()
    chapa = (chapa or "").strip()
    if not chapa:
        return
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT 1 FROM colaboradores WHERE setor=? AND chapa=? LIMIT 1", (setor, chapa))
    if cur.fetchone() is None:
        cur.execute("INSERT INTO colaboradores(nome, setor, chapa, criado_em) VALUES (?, ?, ?, ?)",
                    (nome or chapa, setor, chapa, datetime.now().isoformat()))
    else:
        if nome:
            cur.execute("UPDATE colaboradores SET nome=? WHERE setor=? AND chapa=?", (nome, setor, chapa))
    con.commit()
    con.close()
    try:
        ensure_system_user_from_colaborador(nome or chapa, setor, chapa)
    except Exception:
        pass

def apply_manual_base_folgas(setor: str, ano: int, mes: int, base_rows: list[dict], limpar_overrides_mes: bool = False):
    """
    Aplica uma base manual criando overrides Status=Folga.
    - Opcional: limpar_overrides_mes = True remove TODOS os overrides do mês antes de aplicar.
    """
    con = db_conn()
    cur = con.cursor()
    if limpar_overrides_mes:
        cur.execute("DELETE FROM overrides WHERE setor=? AND ano=? AND mes=?", (setor, int(ano), int(mes)))
        con.commit()
    con.close()

    # garante colaboradores e aplica folgas como override
    for r in base_rows:
        ch = str(r.get("Chapa","")).strip()
        nm = str(r.get("Nome","")).strip()
        dias = r.get("Dias_Folga", []) or []
        upsert_colaborador_nome(setor, ch, nm)
        for d in dias:
            try:
                dd = int(d)
            except Exception:
                continue
            if dd <= 0:
                continue
            set_override(setor, int(ano), int(mes), ch, dd, "status", "Folga")

def delete_colaborador_total(setor: str, chapa: str):
    """
    Exclui colaborador e tudo do setor relacionado a ele:
    - colaboradores
    - ferias
    - overrides
    - escala_mes
    - estado_mes_anterior
    """
    con = db_conn()
    cur = con.cursor()
    cur.execute("DELETE FROM ferias WHERE setor=? AND chapa=?", (setor, chapa))
    cur.execute("DELETE FROM overrides WHERE setor=? AND chapa=?", (setor, chapa))
    cur.execute("DELETE FROM escala_mes WHERE setor=? AND chapa=?", (setor, chapa))
    cur.execute("DELETE FROM estado_mes_anterior WHERE setor=? AND chapa=?", (setor, chapa))
    cur.execute("DELETE FROM colaboradores WHERE setor=? AND chapa=?", (setor, chapa))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

def update_colaborador_perfil(setor: str, chapa_antiga: str, chapa_nova: str, nome_novo: str, subgrupo: str, entrada: str, folga_sab: bool):
    setor = _norm_setor(setor)
    chapa_antiga = _norm_chapa(chapa_antiga)
    chapa_nova = _norm_chapa(chapa_nova)
    nome_novo = str(nome_novo or "").strip()
    subgrupo = str(subgrupo or "").strip()
    entrada = str(entrada or BALANCO_DIA_ENTRADA).strip()

    if not chapa_antiga or not chapa_nova:
        raise ValueError("Chapa antiga/nova inválida.")
    if not nome_novo:
        raise ValueError("Nome do colaborador é obrigatório.")

    con = db_conn()
    cur = con.cursor()

    if chapa_antiga != chapa_nova:
        cur.execute(
            "SELECT 1 FROM colaboradores WHERE setor=? AND chapa=? LIMIT 1",
            (setor, chapa_nova),
        )
        if cur.fetchone() is not None:
            con.close()
            raise ValueError(f"Já existe colaborador com a chapa {chapa_nova} neste setor.")

    # Atualiza cadastro principal
    cur.execute("""
        UPDATE colaboradores
        SET nome=?, chapa=?, subgrupo=?, entrada=?, folga_sab=?
        WHERE setor=? AND chapa=?
    """, (nome_novo, chapa_nova, subgrupo, entrada, 1 if folga_sab else 0, setor, chapa_antiga))

    # Mantém tabelas relacionadas consistentes quando a chapa for alterada
    tabelas = [
        "usuarios_sistema",
        "ferias",
        "overrides",
        "escala_mes",
        "estado_mes_anterior",
    ]
    for tb in tabelas:
        try:
            cur.execute(f"UPDATE {tb} SET chapa=? WHERE setor=? AND chapa=?", (chapa_nova, setor, chapa_antiga))
        except Exception:
            pass

    # Atualiza nome também no usuário do sistema, se existir login
    try:
        cur.execute(
            "UPDATE usuarios_sistema SET nome=? WHERE setor=? AND chapa=?",
            (nome_novo, setor, chapa_nova),
        )
    except Exception:
        pass

    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

@st.cache_data(show_spinner=False, ttl=120)
def load_colaboradores_setor(setor: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT nome, chapa, subgrupo, entrada, folga_sab
        FROM colaboradores
        WHERE setor=?
        ORDER BY nome ASC
    """, (setor,))
    rows = cur.fetchall()
    con.close()
    return [{
        "Nome": r[0],
        "Chapa": r[1],
        "Subgrupo": (r[2] or "").strip(),
        "Entrada": (r[3] or "06:00").strip(),
        "Folga_Sab": bool(r[4]),
        "Setor": setor,
    } for r in rows]


def _hora_to_min(h: str) -> int | None:
    h = str(h or '').strip()
    if not re.match(r'^\d{2}:\d{2}$', h):
        return None
    try:
        hh, mm = h.split(':')
        return int(hh) * 60 + int(mm)
    except Exception:
        return None


def _classificar_compat_horario(h1: str, h2: str, tolerancia_min: int = 20) -> str:
    m1 = _hora_to_min(h1)
    m2 = _hora_to_min(h2)
    if m1 is None or m2 is None:
        return 'DIFERENTE'
    diff = abs(m1 - m2)
    if diff == 0:
        return 'IGUAL'
    if diff <= int(tolerancia_min):
        return 'QUASE IGUAL'
    return 'DIFERENTE'


def _mes_ref_str(ano: int, mes: int) -> str:
    return f"{int(ano):04d}-{int(mes):02d}"


RODIZIO_CAIXA_QTD_FIXA = 14
RODIZIO_CAIXA_COTAS_HORARIO = {
    '06:50': 4,
    '07:30': 1,
    '08:00': 1,
    '09:00': 1,
    '10:00': 1,
    '11:00': 1,
    '12:00': 1,
    '12:40': 4,
}


def _rodizio_caixa_cotas_ordenadas() -> list[tuple[str, int]]:
    return [(h, int(q)) for h, q in RODIZIO_CAIXA_COTAS_HORARIO.items() if int(q) > 0]


def _rodizio_caixa_qtd_total() -> int:
    return int(sum(q for _, q in _rodizio_caixa_cotas_ordenadas()))


def _rodizio_caixa_candidatos_por_horario(candidatos_origem: list[dict], subgrupo_destino: str) -> dict[str, list[dict]]:
    grupos: dict[str, list[dict]] = {}
    for c in candidatos_origem or []:
        h = str(c.get('Entrada') or '').strip()
        grupos.setdefault(h, []).append(c)
    for h in list(grupos.keys()):
        grupos[h] = _rodizio_rank_origem('', grupos[h], subgrupo_destino) if False else grupos[h]
    return grupos


def _rodizio_caixa_tem_aplicacao_no_mes(setor: str, ano: int, mes: int, subgrupo_origem: str, subgrupo_destino: str) -> bool:
    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute("""
            SELECT 1
            FROM rodizio_caixa_hist
            WHERE setor=? AND ano=? AND mes=? AND subgrupo_origem=? AND subgrupo_destino=?
            LIMIT 1
        """, (setor, int(ano), int(mes), subgrupo_origem, subgrupo_destino))
        return cur.fetchone() is not None
    finally:
        con.close()


@st.cache_data(show_spinner=False, ttl=300)
def get_rodizio_caixa_cfg(setor: str, subgrupo_origem: str = 'OPERADOR DE CAIXA 01', subgrupo_destino: str = 'OPERADOR DE CAIXA 02') -> dict:
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT subgrupo_origem, subgrupo_destino, qtd_destino, tolerancia_min, ativo
        FROM rodizio_caixa_cfg
        WHERE setor=? AND subgrupo_origem=? AND subgrupo_destino=?
        LIMIT 1
    """, (setor, subgrupo_origem, subgrupo_destino))
    row = cur.fetchone()
    if row is None:
        cur.execute("""
            INSERT OR IGNORE INTO rodizio_caixa_cfg(setor, subgrupo_origem, subgrupo_destino, qtd_destino, tolerancia_min, ativo)
            VALUES (?, ?, ?, 14, 20, 1)
        """, (setor, subgrupo_origem, subgrupo_destino))
        con.commit()
        row = (subgrupo_origem, subgrupo_destino, 14, 20, 1)
    con.close()
    return {
        'subgrupo_origem': str(row[0] or subgrupo_origem),
        'subgrupo_destino': str(row[1] or subgrupo_destino),
        'qtd_destino': int(row[2] or 14),
        'tolerancia_min': int(row[3] or 20),
        'ativo': bool(row[4]),
    }


def set_rodizio_caixa_cfg(setor: str, subgrupo_origem: str, subgrupo_destino: str, qtd_destino: int = 14, tolerancia_min: int = 20, ativo: bool = True):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        INSERT INTO rodizio_caixa_cfg(setor, subgrupo_origem, subgrupo_destino, qtd_destino, tolerancia_min, ativo)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(setor, subgrupo_origem, subgrupo_destino)
        DO UPDATE SET qtd_destino=excluded.qtd_destino, tolerancia_min=excluded.tolerancia_min, ativo=excluded.ativo
    """, (setor, subgrupo_origem, subgrupo_destino, int(qtd_destino), int(tolerancia_min), 1 if ativo else 0))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass


@st.cache_data(show_spinner=False, ttl=180)
def list_rodizio_caixa_hist(setor: str, limit: int = 120):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT ano, mes, ciclo_ref, chapa, nome, movimento, subgrupo_origem, subgrupo_destino,
               entrada_antiga, entrada_nova, compat_status, observacao, criado_em
        FROM rodizio_caixa_hist
        WHERE setor=?
        ORDER BY ano DESC, mes DESC, nome ASC
        LIMIT ?
    """, (setor, int(limit)))
    rows = cur.fetchall()
    con.close()
    return [{
        'Ano': int(r[0]), 'Mes': int(r[1]), 'Ciclo': str(r[2] or ''), 'Chapa': str(r[3] or ''), 'Nome': str(r[4] or ''),
        'Movimento': str(r[5] or ''), 'Origem': str(r[6] or ''), 'Destino': str(r[7] or ''),
        'Entrada antiga': str(r[8] or ''), 'Entrada nova': str(r[9] or ''), 'Compatibilidade': str(r[10] or ''),
        'Observação': str(r[11] or ''), 'Criado em': str(r[12] or ''),
    } for r in rows]


def _rodizio_last_move_map(setor: str, subgrupo_destino: str) -> dict[str, int]:
    con = db_conn()
    cur = con.cursor()
    out: dict[str, int] = {}
    try:
        cur.execute("""
            SELECT chapa, MAX(ano * 100 + mes)
            FROM rodizio_caixa_hist
            WHERE setor=? AND movimento='ENTRA_DESTINO' AND subgrupo_destino=?
            GROUP BY chapa
        """, (setor, subgrupo_destino))
        for ch, ym in cur.fetchall():
            out[str(ch or '').strip()] = int(ym or 0)
    finally:
        con.close()
    return out


def _rodizio_format_ym(ym: int) -> str:
    ym = int(ym or 0)
    if ym <= 0:
        return 'Nunca fez rodízio para o destino'
    ano = ym // 100
    mes = ym % 100
    if ano <= 0 or mes <= 0:
        return 'Nunca fez rodízio para o destino'
    return f'{mes:02d}/{ano}'


def _rodizio_rank_origem(setor: str, candidatos_origem: list[dict], subgrupo_destino: str):
    last_move = _rodizio_last_move_map(setor, subgrupo_destino)
    out = []
    for c in candidatos_origem:
        ch = str(c.get('Chapa') or '').strip()
        out.append((last_move.get(ch, 0), str(c.get('Nome') or '').upper(), c))
    out.sort(key=lambda x: (x[0], x[1]))
    return [x[2] for x in out]


def _rodizio_domingos_trabalhados_map(setor: str, ano: int, mes: int) -> dict[str, int]:
    hist_db = get_hist_mes_com_overrides_cached(setor, int(ano), int(mes)) or {}
    out: dict[str, int] = {}
    for chapa, df in (hist_db or {}).items():
        total = 0
        try:
            if df is None or df.empty:
                out[str(chapa)] = 0
                continue
            for _, row in df.iterrows():
                dt = pd.to_datetime(row.get('Data'), errors='coerce')
                if pd.isna(dt):
                    continue
                if int(dt.weekday()) != 6:
                    continue
                if str(row.get('Status') or '').strip() in WORK_STATUSES:
                    total += 1
        except Exception:
            total = 0
        out[str(chapa)] = int(total)
    return out


def _rodizio_domingos_detalhe_map(setor: str, ano: int, mes: int) -> dict[str, dict]:
    hist_db = get_hist_mes_com_overrides_cached(setor, int(ano), int(mes)) or {}
    out: dict[str, dict] = {}
    for chapa, df in (hist_db or {}).items():
        trab = []
        folga = []
        try:
            if df is None or df.empty:
                out[str(chapa)] = {'trab': tuple(), 'folga': tuple()}
                continue
            for _, row in df.iterrows():
                dt = pd.to_datetime(row.get('Data'), errors='coerce')
                if pd.isna(dt) or int(dt.weekday()) != 6:
                    continue
                dia_label = f"{int(dt.day):02d}"
                status = str(row.get('Status') or '').strip()
                if status in WORK_STATUSES:
                    trab.append(dia_label)
                elif status == 'Folga':
                    folga.append(dia_label)
        except Exception:
            trab = []
            folga = []
        out[str(chapa)] = {'trab': tuple(trab), 'folga': tuple(folga)}
    return out


def _rodizio_domingos_label(info: dict | None) -> str:
    info = info or {}
    trab = list(info.get('trab') or [])
    folga = list(info.get('folga') or [])
    partes = []
    partes.append('Trabalha: ' + (', '.join(trab) if trab else '-'))
    partes.append('Folga: ' + (', '.join(folga) if folga else '-'))
    return ' | '.join(partes)


def _rodizio_match_domingos(orig_info: dict | None, dest_info: dict | None) -> dict:
    orig_info = orig_info or {}
    dest_info = dest_info or {}
    trab_o = set(orig_info.get('trab') or [])
    trab_d = set(dest_info.get('trab') or [])
    folga_o = set(orig_info.get('folga') or [])
    folga_d = set(dest_info.get('folga') or [])
    trab_eq = sorted(trab_o & trab_d)
    folga_eq = sorted(folga_o & folga_d)
    return {
        'trab_iguais_qtd': int(len(trab_eq)),
        'folga_iguais_qtd': int(len(folga_eq)),
        'trab_iguais': trab_eq,
        'folga_iguais': folga_eq,
        'score': int(len(trab_eq) * 100 + len(folga_eq) * 10),
    }





def simular_rodizio_caixa_mes(
    setor: str,
    ano: int,
    mes: int,
    subgrupo_origem: str = 'OPERADOR DE CAIXA 01',
    subgrupo_destino: str = 'OPERADOR DE CAIXA 02',
    qtd_destino: int = 14,
    tolerancia_min: int = 20,
    aprovados_por_slot: dict | None = None,
    negados_chapas: list[str] | None = None,
):
    colaboradores = load_colaboradores_setor(setor)
    origem = [c for c in colaboradores if (c.get('Subgrupo') or '').strip().upper() == str(subgrupo_origem).strip().upper()]
    destino = [c for c in colaboradores if (c.get('Subgrupo') or '').strip().upper() == str(subgrupo_destino).strip().upper()]

    qtd_obrigatoria = int(_rodizio_caixa_qtd_total())
    qtd_cfg = int(qtd_destino or qtd_obrigatoria)
    qtd_alvo = int(max(qtd_obrigatoria, qtd_cfg))

    try:
        origem = _colaboradores_com_rodizio_aplicado_frentecaixa(setor, int(ano), int(mes), origem)
        destino = _colaboradores_com_rodizio_aplicado_frentecaixa(setor, int(ano), int(mes), destino)
    except Exception:
        pass

    destino_sorted = sorted(destino, key=lambda c: (str(c.get('Entrada') or ''), str(c.get('Nome') or '').upper()))
    # Usa todo o grupo de destino na simulação e deixa as cotas/seleção escolherem
    # quais 14 sairão do Caixa 02. Isso evita cortar horários importantes
    # (ex.: 12:40) quando existem mais de 14 pessoas no destino.
    destino_sel = list(destino_sorted)

    candidatos = []
    for c in origem:
        try:
            if is_de_ferias(setor, str(c.get('Chapa') or ''), date(int(ano), int(mes), 1)):
                continue
        except Exception:
            pass
        candidatos.append(c)

    domingos_map = _rodizio_domingos_trabalhados_map(setor, ano, mes)
    domingos_detalhe_map = _rodizio_domingos_detalhe_map(setor, ano, mes)
    last_move_map = _rodizio_last_move_map(setor, subgrupo_destino)
    aprovados_por_slot = {str(k): str(v) for k, v in (aprovados_por_slot or {}).items() if str(v or '').strip()}
    negados_set = {str(x).strip() for x in (negados_chapas or []) if str(x).strip()}

    def _pool_horario_proximo(base_lista: list[dict], horario_ref: str, usados_local: set[str], bloqueados_local: set[str], max_diff_min: int = 60) -> list[dict]:
        ref_min = _hora_to_min(horario_ref)
        if ref_min is None:
            return []
        ranked = []
        seen = set()
        for cand in base_lista or []:
            ch = str(cand.get('Chapa') or '').strip()
            ent = str(cand.get('Entrada') or '').strip()
            ent_min = _hora_to_min(ent)
            if not ch or ch in usados_local or ch in bloqueados_local or ent_min is None:
                continue
            diff = abs(int(ent_min) - int(ref_min))
            if diff <= int(max_diff_min):
                ranked.append((0 if ent == horario_ref else 1, diff, str(cand.get('Nome') or '').upper(), ch, cand))
        ranked.sort(key=lambda x: (x[0], x[1], x[2], x[3]))
        out = []
        for _prio, diff, _nome, ch, cand in ranked:
            if ch in seen:
                continue
            seen.add(ch)
            item = dict(cand)
            item['_diff_horario_ref_min'] = int(diff)
            out.append(item)
        return out

    usados_origem = set()
    usados_destino = set()
    pares = []
    slots = []
    alertas = []
    cotas_resumo = []

    cotas_processamento = []
    for horario_ref, qtd_horario in _rodizio_caixa_cotas_ordenadas():
        exatos_origem_tmp = [c for c in candidatos if str(c.get('Entrada') or '').strip() == horario_ref and str(c.get('Chapa') or '').strip() not in negados_set]
        exatos_destino_tmp = [d for d in destino_sel if str(d.get('Entrada') or '').strip() == horario_ref]
        pool_origem_tmp = _pool_horario_proximo(candidatos, horario_ref, set(), negados_set, 60)
        pool_destino_tmp = _pool_horario_proximo(destino_sel, horario_ref, set(), set(), 60)
        cotas_processamento.append({
            'horario_ref': horario_ref,
            'qtd_horario': int(qtd_horario),
            'exatos_origem': int(len(exatos_origem_tmp)),
            'exatos_destino': int(len(exatos_destino_tmp)),
            'pool_origem': int(len(pool_origem_tmp)),
            'pool_destino': int(len(pool_destino_tmp)),
        })

    # Processa primeiro os horários mais escassos para não consumir cedo
    # candidatos flexíveis que fecham as 14 vagas obrigatórias.
    cotas_processamento = sorted(
        cotas_processamento,
        key=lambda x: (x['pool_destino'], x['pool_origem'], x['exatos_destino'], x['exatos_origem'], -x['qtd_horario'], x['horario_ref'])
    )
    cont_slot_por_horario = {}

    for info_cota in cotas_processamento:
        horario_ref = str(info_cota.get('horario_ref') or '')
        qtd_horario = int(info_cota.get('qtd_horario') or 0)
        exatos_origem = [c for c in candidatos if str(c.get('Entrada') or '').strip() == horario_ref and str(c.get('Chapa') or '').strip() not in negados_set]
        exatos_destino = [d for d in destino_sel if str(d.get('Entrada') or '').strip() == horario_ref]
        pool_origem_inicio = _pool_horario_proximo(candidatos, horario_ref, set(), negados_set, 60)
        pool_destino_inicio = _pool_horario_proximo(destino_sel, horario_ref, set(), set(), 60)

        if len(exatos_origem) < int(qtd_horario):
            faltam = max(0, int(qtd_horario) - len(exatos_origem))
            exemplos = [str(c.get('Entrada') or '').strip() for c in pool_origem_inicio if str(c.get('Entrada') or '').strip() != horario_ref][:8]
            sug = ', '.join(exemplos) if exemplos else 'sem candidato em até 1 hora'
            alertas.append(
                f"Origem {subgrupo_origem}: horário {horario_ref} tem {len(exatos_origem)} candidato(s) exato(s) e a regra pede {int(qtd_horario)}. "
                f"Sugestão: completar {faltam} vaga(s) com horário até 1 hora a menos/mais ({sug})."
            )
        if len(exatos_destino) < int(qtd_horario):
            faltam_dest = max(0, int(qtd_horario) - len(exatos_destino))
            exemplos_dest = [str(c.get('Entrada') or '').strip() for c in pool_destino_inicio if str(c.get('Entrada') or '').strip() != horario_ref][:8]
            sug_dest = ', '.join(exemplos_dest) if exemplos_dest else 'sem pessoa no destino em até 1 hora'
            alertas.append(
                f"Destino {subgrupo_destino}: horário {horario_ref} tem {len(exatos_destino)} pessoa(s) exata(s) para saída e a regra pede {int(qtd_horario)}. "
                f"Sugestão: completar {faltam_dest} vaga(s) com saída até 1 hora a menos/mais ({sug_dest})."
            )

        qtd_slots = min(int(qtd_horario), len(pool_origem_inicio), len(pool_destino_inicio))
        cotas_resumo.append({
            'Horário': horario_ref,
            'Regra': int(qtd_horario),
            'Disponível origem': len(exatos_origem),
            'Disponível destino': len(exatos_destino),
            'Selecionados': int(qtd_slots),
        })

        for idx in range(qtd_slots):
            pool_destino = _pool_horario_proximo(destino_sel, horario_ref, usados_destino, set(), 60)
            if not pool_destino:
                continue

            d = pool_destino[0]
            dest_fallback = (str(d.get('Entrada') or '').strip() != horario_ref)
            idx_slot_horario = int(cont_slot_por_horario.get(horario_ref, 0))
            slot_key = f"{horario_ref}|{idx_slot_horario}|{str(d.get('Chapa') or '').strip()}"
            locked_chapa = aprovados_por_slot.get(slot_key, '').strip()

            disponiveis_pool = _pool_horario_proximo(candidatos, horario_ref, usados_origem, negados_set, 60)
            if not disponiveis_pool:
                continue

            chosen = None
            escolha_fallback = False

            dest_ch = str(d.get('Chapa') or '').strip()
            domingos_dest = int(domingos_map.get(dest_ch, 0) or 0)
            dest_dom_info = domingos_detalhe_map.get(dest_ch, {'trab': tuple(), 'folga': tuple()})
            scored = []
            for pos, cand in enumerate(disponiveis_pool):
                ch = str(cand.get('Chapa') or '').strip()
                ent = str(cand.get('Entrada') or '').strip()
                diff_hor = abs(int((_hora_to_min(ent) or 0)) - int((_hora_to_min(horario_ref) or 0)))
                domingos_orig = int(domingos_map.get(ch, 0) or 0)
                sunday_diff = abs(domingos_orig - domingos_dest)
                match_info = _rodizio_match_domingos(domingos_detalhe_map.get(ch), dest_dom_info)
                last_move = int(last_move_map.get(ch, 0) or 0)
                scored.append((
                    0 if ent == horario_ref else 1,
                    diff_hor,
                    -int(match_info.get('trab_iguais_qtd', 0) or 0),
                    -int(match_info.get('folga_iguais_qtd', 0) or 0),
                    sunday_diff,
                    last_move,
                    pos,
                    str(cand.get('Nome') or '').upper(),
                    cand,
                    match_info,
                    domingos_orig,
                ))
            scored.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[4], x[5], x[6], x[7]))

            alternativas_opcoes = []
            for item in scored[:20]:
                cand = item[8]
                match_item = item[9]
                domingos_item = int(item[10] or 0)
                last_move_item = int(last_move_map.get(str(cand.get('Chapa') or '').strip(), 0) or 0)
                alternativas_opcoes.append({
                    'chapa': str(cand.get('Chapa') or '').strip(),
                    'nome': str(cand.get('Nome') or ''),
                    'entrada': str(cand.get('Entrada') or '').strip(),
                    'diff_horario_ref_min': int(item[1] or 0),
                    'domingos': int(domingos_item),
                    'diff_domingos': int(item[4] or 0),
                    'domingos_trabalho_iguais_qtd': int(match_item.get('trab_iguais_qtd', 0) or 0),
                    'domingos_folga_iguais_qtd': int(match_item.get('folga_iguais_qtd', 0) or 0),
                    'ultimo_mes_destino_label': _rodizio_format_ym(last_move_item),
                })

            if locked_chapa:
                for item in scored:
                    cand = item[8]
                    if str(cand.get('Chapa') or '').strip() == locked_chapa:
                        chosen = cand
                        escolha_fallback = (str(cand.get('Entrada') or '').strip() != horario_ref)
                        break

            if chosen is None and scored:
                chosen = scored[0][8]
                escolha_fallback = (str(chosen.get('Entrada') or '').strip() != horario_ref)

            ch = str(chosen.get('Chapa') or '').strip()
            usados_origem.add(ch)
            dest_ch_final = str(d.get('Chapa') or '').strip()
            usados_destino.add(dest_ch_final)

            domingos_orig = int(domingos_map.get(ch, 0) or 0)
            domingos_dest = int(domingos_map.get(dest_ch_final, 0) or 0)
            sunday_diff = abs(domingos_orig - domingos_dest)
            origem_dom_info = domingos_detalhe_map.get(ch, {'trab': tuple(), 'folga': tuple()})
            destino_dom_info = domingos_detalhe_map.get(dest_ch_final, {'trab': tuple(), 'folga': tuple()})
            match_info = _rodizio_match_domingos(origem_dom_info, destino_dom_info)
            compat = _classificar_compat_horario(chosen.get('Entrada', ''), d.get('Entrada', ''), tolerancia_min=tolerancia_min)
            last_move_ym = int(last_move_map.get(ch, 0) or 0)
            diff_horario_ref = abs(int((_hora_to_min(str(chosen.get('Entrada') or '').strip()) or 0)) - int((_hora_to_min(horario_ref) or 0)))
            diff_dest_ref = abs(int((_hora_to_min(str(d.get('Entrada') or '').strip()) or 0)) - int((_hora_to_min(horario_ref) or 0)))

            obs_parts = [f'Regra fixa do horário {horario_ref}']
            if escolha_fallback:
                obs_parts.append(f"Sugestão origem por horário próximo: {str(chosen.get('Entrada') or '').strip()} ({diff_horario_ref} min da regra)")
            else:
                obs_parts.append("Origem selecionada no horário exato da regra")
            if dest_fallback:
                obs_parts.append(f"Saída do destino por horário próximo: {str(d.get('Entrada') or '').strip()} ({diff_dest_ref} min da regra)")
            else:
                obs_parts.append("Saída do destino no horário exato da regra")
            if int(match_info.get('trab_iguais_qtd', 0) or 0) > 0 or int(match_info.get('folga_iguais_qtd', 0) or 0) > 0:
                obs_parts.append(
                    f"Domingos parecidos por data: trabalha igual {int(match_info.get('trab_iguais_qtd', 0) or 0)} | folga igual {int(match_info.get('folga_iguais_qtd', 0) or 0)}"
                )
            else:
                obs_parts.append(f"Sem domingo igual por data; mantido o principal do horário {horario_ref}")
            obs_parts.append(f'Quantidade de domingos: entra {domingos_orig} / sai {domingos_dest}')
            obs_parts.append(f'Último mês no destino: {_rodizio_format_ym(last_move_ym)}')

            alternativas_mesmo_horario = sum(
                1 for cand in candidatos
                if str(cand.get('Entrada') or '').strip() == horario_ref
                and str(cand.get('Chapa') or '').strip() not in usados_origem
                and str(cand.get('Chapa') or '').strip() not in negados_set
            )
            alternativas_ate_1h = sum(
                1 for cand in _pool_horario_proximo(candidatos, horario_ref, usados_origem, negados_set, 60)
            )

            row = {
                'slot_key': slot_key,
                'origem_chapa': ch,
                'origem_nome': str(chosen.get('Nome') or ''),
                'origem_subgrupo': subgrupo_origem,
                'origem_entrada': str(chosen.get('Entrada') or ''),
                'origem_domingos': int(domingos_orig),
                'origem_ultimo_mes_destino': last_move_ym,
                'origem_ultimo_mes_destino_label': _rodizio_format_ym(last_move_ym),
                'destino_chapa': str(d.get('Chapa') or ''),
                'destino_nome': str(d.get('Nome') or ''),
                'destino_subgrupo': subgrupo_destino,
                'destino_entrada': str(d.get('Entrada') or ''),
                'destino_domingos': int(domingos_dest),
                'diff_domingos': int(sunday_diff),
                'origem_domingos_label': _rodizio_domingos_label(origem_dom_info),
                'destino_domingos_label': _rodizio_domingos_label(destino_dom_info),
                'domingos_trabalho_iguais_qtd': int(match_info.get('trab_iguais_qtd', 0) or 0),
                'domingos_folga_iguais_qtd': int(match_info.get('folga_iguais_qtd', 0) or 0),
                'domingos_trabalho_iguais_label': ', '.join(match_info.get('trab_iguais') or []) or '-',
                'domingos_folga_iguais_label': ', '.join(match_info.get('folga_iguais') or []) or '-',
                'compatibilidade': compat,
                'observacao': ' | '.join(obs_parts),
                'horario_ref': horario_ref,
                'alternativas_mesmo_horario': int(alternativas_mesmo_horario),
                'alternativas_ate_1h': int(alternativas_ate_1h),
                'fallback_horario_proximo': bool(escolha_fallback),
                'fallback_destino_horario_proximo': bool(dest_fallback),
                'diff_horario_ref_min': int(diff_horario_ref),
                'diff_destino_ref_min': int(diff_dest_ref),
                'aprovado_manual': bool(aprovados_por_slot.get(slot_key) == ch),
                'alternativas_opcoes': list(alternativas_opcoes),
            }
            slots.append(dict(row))
            pares.append(dict(row))
            cont_slot_por_horario[horario_ref] = int(cont_slot_por_horario.get(horario_ref, 0)) + 1

    # preenchimento final para fechar as 14 sugestões obrigatórias, usando qualquer horário até 1h a menos/mais
    while len(slots) < qtd_obrigatoria:
        melhor = None
        for horario_ref, _qtd in _rodizio_caixa_cotas_ordenadas():
            pool_destino = _pool_horario_proximo(destino_sel, horario_ref, usados_destino, set(), 60)
            pool_origem = _pool_horario_proximo(candidatos, horario_ref, usados_origem, negados_set, 60)
            if not pool_destino or not pool_origem:
                continue
            d = pool_destino[0]
            cand = pool_origem[0]
            diff_total = int(cand.get('_diff_horario_ref_min', 0) or 0) + int(d.get('_diff_horario_ref_min', 0) or 0)
            item = (diff_total, str(horario_ref), cand, d)
            if melhor is None or item < melhor:
                melhor = item

        if melhor is None:
            break

        _diff_total, horario_ref, chosen, d = melhor
        extras_pool = _pool_horario_proximo(candidatos, horario_ref, usados_origem, negados_set, 60)
        alternativas_opcoes = []
        for cand in extras_pool[:20]:
            ch_alt = str(cand.get('Chapa') or '').strip()
            ult_alt = int(last_move_map.get(ch_alt, 0) or 0)
            domingos_alt = int(domingos_map.get(ch_alt, 0) or 0)
            alternativas_opcoes.append({
                'chapa': ch_alt,
                'nome': str(cand.get('Nome') or ''),
                'entrada': str(cand.get('Entrada') or '').strip(),
                'diff_horario_ref_min': int(cand.get('_diff_horario_ref_min', 0) or 0),
                'domingos': int(domingos_alt),
                'diff_domingos': 0,
                'domingos_trabalho_iguais_qtd': 0,
                'domingos_folga_iguais_qtd': 0,
                'ultimo_mes_destino_label': _rodizio_format_ym(ult_alt),
            })

        locked_chapa = aprovados_por_slot.get(f"{horario_ref}|extra|{str(d.get('Chapa') or '').strip()}|{len(slots)}", '').strip()
        if locked_chapa:
            for cand in extras_pool:
                if str(cand.get('Chapa') or '').strip() == locked_chapa:
                    chosen = cand
                    break

        ch = str(chosen.get('Chapa') or '').strip()
        dest_ch_final = str(d.get('Chapa') or '').strip()
        usados_origem.add(ch)
        usados_destino.add(dest_ch_final)

        domingos_orig = int(domingos_map.get(ch, 0) or 0)
        domingos_dest = int(domingos_map.get(dest_ch_final, 0) or 0)
        sunday_diff = abs(domingos_orig - domingos_dest)
        origem_dom_info = domingos_detalhe_map.get(ch, {'trab': tuple(), 'folga': tuple()})
        destino_dom_info = domingos_detalhe_map.get(dest_ch_final, {'trab': tuple(), 'folga': tuple()})
        match_info = _rodizio_match_domingos(origem_dom_info, destino_dom_info)
        last_move_ym = int(last_move_map.get(ch, 0) or 0)
        diff_horario_ref = int(chosen.get('_diff_horario_ref_min', 0) or 0)
        diff_dest_ref = int(d.get('_diff_horario_ref_min', 0) or 0)
        slot_key = f"{horario_ref}|extra|{dest_ch_final}|{len(slots)}"

        row = {
            'slot_key': slot_key,
            'origem_chapa': ch,
            'origem_nome': str(chosen.get('Nome') or ''),
            'origem_subgrupo': subgrupo_origem,
            'origem_entrada': str(chosen.get('Entrada') or ''),
            'origem_domingos': int(domingos_orig),
            'origem_ultimo_mes_destino': last_move_ym,
            'origem_ultimo_mes_destino_label': _rodizio_format_ym(last_move_ym),
            'destino_chapa': str(d.get('Chapa') or ''),
            'destino_nome': str(d.get('Nome') or ''),
            'destino_subgrupo': subgrupo_destino,
            'destino_entrada': str(d.get('Entrada') or ''),
            'destino_domingos': int(domingos_dest),
            'diff_domingos': int(sunday_diff),
            'origem_domingos_label': _rodizio_domingos_label(origem_dom_info),
            'destino_domingos_label': _rodizio_domingos_label(destino_dom_info),
            'domingos_trabalho_iguais_qtd': int(match_info.get('trab_iguais_qtd', 0) or 0),
            'domingos_folga_iguais_qtd': int(match_info.get('folga_iguais_qtd', 0) or 0),
            'domingos_trabalho_iguais_label': ', '.join(match_info.get('trab_iguais') or []) or '-',
            'domingos_folga_iguais_label': ', '.join(match_info.get('folga_iguais') or []) or '-',
            'compatibilidade': _classificar_compat_horario(chosen.get('Entrada', ''), d.get('Entrada', ''), tolerancia_min=tolerancia_min),
            'observacao': (
                f'Regra completada com fallback de até 1 hora | '
                f'Origem {str(chosen.get("Entrada") or "").strip()} ({diff_horario_ref} min da regra {horario_ref}) | '
                f'Destino {str(d.get("Entrada") or "").strip()} ({diff_dest_ref} min da regra {horario_ref}) | '
                f'Domingos parecidos por data: trabalha igual {int(match_info.get("trab_iguais_qtd", 0) or 0)} | '
                f'folga igual {int(match_info.get("folga_iguais_qtd", 0) or 0)} | '
                f'Último mês no destino: {_rodizio_format_ym(last_move_ym)}'
            ),
            'horario_ref': horario_ref,
            'alternativas_mesmo_horario': 0,
            'alternativas_ate_1h': sum(1 for _ in _pool_horario_proximo(candidatos, horario_ref, usados_origem, negados_set, 60)),
            'fallback_horario_proximo': True,
            'fallback_destino_horario_proximo': True,
            'diff_horario_ref_min': int(diff_horario_ref),
            'diff_destino_ref_min': int(diff_dest_ref),
            'aprovado_manual': bool(aprovados_por_slot.get(slot_key) == ch),
            'alternativas_opcoes': list(alternativas_opcoes),
        }
        slots.append(dict(row))
        pares.append(dict(row))
        cont_slot_por_horario[horario_ref] = int(cont_slot_por_horario.get(horario_ref, 0)) + 1

    ordem_horarios = {h: i for i, (h, _q) in enumerate(_rodizio_caixa_cotas_ordenadas())}
    slots = sorted(slots, key=lambda x: (ordem_horarios.get(str(x.get('horario_ref') or ''), 999), str(x.get('origem_entrada') or ''), str(x.get('origem_nome') or '').upper(), str(x.get('origem_chapa') or '')))
    pares = sorted(pares, key=lambda x: (ordem_horarios.get(str(x.get('horario_ref') or ''), 999), str(x.get('origem_entrada') or ''), str(x.get('origem_nome') or '').upper(), str(x.get('origem_chapa') or '')))

    resumo_por_horario = {}
    for s in slots:
        h = str(s.get('horario_ref') or '')
        resumo_por_horario[h] = int(resumo_por_horario.get(h, 0)) + 1
    cotas_resumo = []
    for horario_ref, qtd_horario in _rodizio_caixa_cotas_ordenadas():
        exatos_origem = [c for c in candidatos if str(c.get('Entrada') or '').strip() == horario_ref and str(c.get('Chapa') or '').strip() not in negados_set]
        exatos_destino = [d for d in destino_sel if str(d.get('Entrada') or '').strip() == horario_ref]
        cotas_resumo.append({
            'Horário': horario_ref,
            'Regra': int(qtd_horario),
            'Disponível origem': len(exatos_origem),
            'Disponível destino': len(exatos_destino),
            'Selecionados': int(resumo_por_horario.get(horario_ref, 0)),
        })

    proximos = []
    for horario_ref, _qtd in _rodizio_caixa_cotas_ordenadas():
        fila_pool = _pool_horario_proximo(candidatos, horario_ref, usados_origem, negados_set, 60)
        for o in fila_pool:
            ch = str(o.get('Chapa') or '').strip()
            if ch in usados_origem or ch in negados_set:
                continue
            ult = int(last_move_map.get(ch, 0) or 0)
            diff = abs(int((_hora_to_min(str(o.get('Entrada') or '').strip()) or 0)) - int((_hora_to_min(horario_ref) or 0)))
            proximos.append({
                'Chapa': ch,
                'Nome': str(o.get('Nome') or ''),
                'Subgrupo atual': subgrupo_origem,
                'Entrada atual': str(o.get('Entrada') or ''),
                'Horário da fila': horario_ref,
                'Dif. p/ regra (min)': int(diff),
                'Último mês no destino': _rodizio_format_ym(ult),
                'Previsão': 'Próximo rodízio do mesmo horário' if diff == 0 else 'Sugestão por horário próximo (até 1h)',
            })

    if len(pares) < qtd_obrigatoria:
        alertas.append(
            f"Rodízio incompleto no mês: foram montadas {len(pares)} troca(s), mas a regra fixa exige {qtd_obrigatoria}. "
            f"O sistema tentou completar com horário até 1 hora a menos/mais."
        )

    return {
        'pares': pares,
        'slots': slots,
        'alertas': alertas,
        'qtd_destino_atual': len(destino),
        'qtd_troca': len(pares),
        'proximos': proximos,
        'subgrupo_origem': subgrupo_origem,
        'subgrupo_destino': subgrupo_destino,
        'qtd_destino_cfg': int(qtd_alvo),
        'qtd_destino_obrigatoria': int(qtd_obrigatoria),
        'tolerancia_min': int(tolerancia_min),
        'cotas_horario': cotas_resumo,
        'aprovados_por_slot': aprovados_por_slot,
        'negados_chapas': sorted(list(negados_set)),
    }

def aplicar_rodizio_caixa_mes(setor: str, ano: int, mes: int, simulacao: dict):
    pares = list((simulacao or {}).get('pares') or [])
    if not pares:
        return {'ok': False, 'msg': 'Nenhuma troca simulada para aplicar.'}

    subgrupo_origem = str((simulacao or {}).get('subgrupo_origem') or 'OPERADOR DE CAIXA 01')
    subgrupo_destino = str((simulacao or {}).get('subgrupo_destino') or 'OPERADOR DE CAIXA 02')
    if _rodizio_caixa_tem_aplicacao_no_mes(setor, int(ano), int(mes), subgrupo_origem, subgrupo_destino):
        return {'ok': False, 'msg': 'Este rodízio já foi aplicado nesta competência. Para manter a regra fixa, não é permitido reaplicar no mesmo mês.'}

    qtd_obrigatoria = int((simulacao or {}).get('qtd_destino_obrigatoria') or _rodizio_caixa_qtd_total())
    if len(pares) < qtd_obrigatoria:
        return {'ok': False, 'msg': f'Rodízio incompleto: {len(pares)} troca(s) simulada(s), mas a regra fixa exige {qtd_obrigatoria} no mês.'}

    chapas_afetadas = set()
    con = db_conn()
    cur = con.cursor()
    ciclo = _mes_ref_str(ano, mes)
    ts = datetime.now().isoformat()
    try:
        cur.execute('BEGIN')
        for p in pares:
            ch_origem = str(p.get('origem_chapa') or '').strip()
            ch_destino = str(p.get('destino_chapa') or '').strip()
            ent_origem = str(p.get('origem_entrada') or '').strip()
            ent_destino = str(p.get('destino_entrada') or '').strip()

            if ch_origem:
                chapas_afetadas.add(ch_origem)
                cur.execute(
                    "UPDATE colaboradores SET subgrupo=?, entrada=? WHERE setor=? AND chapa=?",
                    (simulacao.get('subgrupo_destino', 'OPERADOR DE CAIXA 02'), ent_destino, setor, ch_origem)
                )
            if ch_destino:
                chapas_afetadas.add(ch_destino)
                cur.execute(
                    "UPDATE colaboradores SET subgrupo=?, entrada=? WHERE setor=? AND chapa=?",
                    (simulacao.get('subgrupo_origem', 'OPERADOR DE CAIXA 01'), ent_origem, setor, ch_destino)
                )

            cur.execute("""
                INSERT OR REPLACE INTO rodizio_caixa_hist(setor, ano, mes, ciclo_ref, chapa, nome, movimento, subgrupo_origem, subgrupo_destino, entrada_antiga, entrada_nova, compat_status, observacao, criado_em)
                VALUES (?, ?, ?, ?, ?, ?, 'ENTRA_DESTINO', ?, ?, ?, ?, ?, ?, ?)
            """, (setor, int(ano), int(mes), ciclo, ch_origem, p['origem_nome'], simulacao.get('subgrupo_origem', 'OPERADOR DE CAIXA 01'), simulacao.get('subgrupo_destino', 'OPERADOR DE CAIXA 02'), ent_origem, ent_destino, p['compatibilidade'], p.get('observacao',''), ts))
            cur.execute("""
                INSERT OR REPLACE INTO rodizio_caixa_hist(setor, ano, mes, ciclo_ref, chapa, nome, movimento, subgrupo_origem, subgrupo_destino, entrada_antiga, entrada_nova, compat_status, observacao, criado_em)
                VALUES (?, ?, ?, ?, ?, ?, 'SAI_DESTINO', ?, ?, ?, ?, ?, ?, ?)
            """, (setor, int(ano), int(mes), ciclo, ch_destino, p['destino_nome'], simulacao.get('subgrupo_destino', 'OPERADOR DE CAIXA 02'), simulacao.get('subgrupo_origem', 'OPERADOR DE CAIXA 01'), ent_destino, ent_origem, p['compatibilidade'], p.get('observacao',''), ts))

        # IMPORTANTE:
        # limpa a competência alvo já salva para obrigar a próxima geração/leitura
        # a respeitar o cadastro atualizado pós-rodízio, sem apagar outras regras do sistema.
        cur.execute("DELETE FROM escala_mes WHERE setor=? AND ano=? AND mes=?", (setor, int(ano), int(mes)))

        # remove apenas overrides de horário/status dos colaboradores afetados nesta competência,
        # para não manter desenho antigo do mês depois da troca de subgrupo/entrada.
        if chapas_afetadas:
            marks = ",".join(["?"] * len(chapas_afetadas))
            params = [setor, int(ano), int(mes), *sorted(chapas_afetadas)]
            cur.execute(
                f"""
                DELETE FROM overrides
                WHERE setor=? AND ano=? AND mes=?
                  AND chapa IN ({marks})
                  AND campo IN ('H_Entrada', 'H_Saida', 'Status')
                """,
                params
            )

        con.commit()
    except Exception:
        con.rollback()
        raise
    finally:
        con.close()

    try:
        st.cache_data.clear()
    except Exception:
        pass

    return {
        'ok': True,
        'msg': (
            f'Rodízio aplicado com {len(pares)} troca(s). '
            f'A competência {int(mes):02d}/{int(ano)} foi limpa para regenerar já com o rodízio refletido.'
        )
    }



def sincronizar_subgrupos_base_rodizio_caixa(setor: str, ano: int, mes: int, subgrupo_origem: str = 'OPERADOR DE CAIXA 01', subgrupo_destino: str = 'OPERADOR DE CAIXA 02'):
    """
    Reaplica manualmente na base de colaboradores o resultado já aprovado/aplicado
    do rodízio do mês, sem permitir nova aplicação do rodízio.
    Usa o histórico do próprio rodízio como fonte da verdade.
    """
    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute(
            """
            SELECT chapa, nome, movimento, subgrupo_origem, subgrupo_destino, entrada_antiga, entrada_nova
            FROM rodizio_caixa_hist
            WHERE setor=? AND ano=? AND mes=?
              AND ((movimento='ENTRA_DESTINO' AND subgrupo_origem=? AND subgrupo_destino=?)
                   OR (movimento='SAI_DESTINO' AND subgrupo_origem=? AND subgrupo_destino=?))
            ORDER BY criado_em ASC, chapa ASC
            """,
            (setor, int(ano), int(mes), subgrupo_origem, subgrupo_destino, subgrupo_destino, subgrupo_origem)
        )
        rows = cur.fetchall() or []
        if not rows:
            return {'ok': False, 'msg': f'Nenhum histórico do rodízio foi encontrado para {int(mes):02d}/{int(ano)}.'}

        atualizacoes = {}
        for chapa, nome, movimento, sg_origem_hist, sg_destino_hist, entrada_antiga, entrada_nova in rows:
            ch = str(chapa or '').strip()
            if not ch:
                continue
            mov = str(movimento or '').strip().upper()
            if mov == 'ENTRA_DESTINO':
                novo_subgrupo = str(subgrupo_destino)
            elif mov == 'SAI_DESTINO':
                novo_subgrupo = str(subgrupo_origem)
            else:
                continue
            atualizacoes[ch] = {
                'nome': str(nome or '').strip(),
                'novo_subgrupo': novo_subgrupo,
                'nova_entrada': str(entrada_nova or '').strip(),
            }

        if not atualizacoes:
            return {'ok': False, 'msg': 'O histórico foi localizado, mas não houve chapas válidas para sincronizar.'}

        chapas_afetadas = sorted(atualizacoes.keys())
        cur.execute('BEGIN')
        for ch in chapas_afetadas:
            info = atualizacoes[ch]
            cur.execute(
                "UPDATE colaboradores SET subgrupo=?, entrada=? WHERE setor=? AND chapa=?",
                (info['novo_subgrupo'], info['nova_entrada'], setor, ch)
            )

        marks = ','.join(['?'] * len(chapas_afetadas))
        params_base = [setor, int(ano), int(mes), *chapas_afetadas]
        cur.execute(
            "DELETE FROM escala_mes WHERE setor=? AND ano=? AND mes=? AND chapa IN ({})".format(marks),
            params_base
        )
        cur.execute(
            """
            DELETE FROM overrides
            WHERE setor=? AND ano=? AND mes=?
              AND chapa IN ({})
              AND campo IN ('H_Entrada', 'H_Saida', 'Status')
            """.format(marks),
            params_base
        )
        con.commit()
    except Exception:
        con.rollback()
        raise
    finally:
        con.close()

    try:
        st.cache_data.clear()
    except Exception:
        pass

    return {
        'ok': True,
        'msg': f'Subgrupos base sincronizados manualmente para {len(chapas_afetadas)} colaborador(es) em {int(mes):02d}/{int(ano)}. Gere a escala novamente para refletir a troca.'
    }

# =========================================================
# SUBGRUPOS + REGRAS
# =========================================================
@st.cache_data(show_spinner=False, ttl=120)
def get_kpis_cached(setor: str, ano: int, mes: int):
    """KPI leve: usa agregações SQL diretas para evitar montar todo hist_db no topo da tela."""
    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute("SELECT COUNT(*) FROM colaboradores WHERE setor=?", (setor,))
        total_colab = int((cur.fetchone() or [0])[0] or 0)

        cur.execute(
            """
            SELECT
                COALESCE(SUM(CASE WHEN status='Folga' THEN 1 ELSE 0 END),0) AS folgas,
                COALESCE(SUM(CASE WHEN status='Férias' THEN 1 ELSE 0 END),0) AS ferias,
                COALESCE(SUM(CASE WHEN status IN ('Trabalho','Balanço') THEN 1 ELSE 0 END),0) AS trabalhos
            FROM escala_mes
            WHERE setor=? AND ano=? AND mes=?
            """,
            (setor, int(ano), int(mes)),
        )
        row = cur.fetchone() or (0, 0, 0)
        folgas_mes = int(row[0] or 0)
        ferias_mes = int(row[1] or 0)
        trabalhos_mes = int(row[2] or 0)
    finally:
        con.close()

    return {
        "total_colab": int(total_colab),
        "folgas_mes": int(folgas_mes),
        "ferias_mes": int(ferias_mes),
        "trabalhos_mes": int(trabalhos_mes),
    }

@st.cache_data(show_spinner=False, ttl=300)
def list_subgrupos(setor: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT nome FROM subgrupos_setor WHERE setor=? ORDER BY nome ASC", (setor,))
    rows = [r[0] for r in cur.fetchall()]
    con.close()
    return rows

def add_subgrupo(setor: str, nome: str):
    nome = (nome or "").strip()
    if not nome:
        return
    con = db_conn()
    cur = con.cursor()
    cur.execute("INSERT OR IGNORE INTO subgrupos_setor(setor, nome) VALUES (?, ?)", (setor, nome))
    cur.execute("""
        INSERT OR IGNORE INTO subgrupo_regras(setor, subgrupo, evitar_seg, evitar_ter, evitar_qua, evitar_qui, evitar_sex, evitar_sab)
        VALUES (?, ?, 0,0,0,0,0,0)
    """, (setor, nome))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

def delete_subgrupo(setor: str, nome: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("DELETE FROM subgrupos_setor WHERE setor=? AND nome=?", (setor, nome))
    cur.execute("DELETE FROM subgrupo_regras WHERE setor=? AND subgrupo=?", (setor, nome))
    cur.execute("UPDATE colaboradores SET subgrupo='' WHERE setor=? AND subgrupo=?", (setor, nome))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

@st.cache_data(show_spinner=False)
def get_subgrupo_regras(setor: str, subgrupo: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT evitar_seg, evitar_ter, evitar_qua, evitar_qui, evitar_sex, evitar_sab
        FROM subgrupo_regras
        WHERE setor=? AND subgrupo=?
        LIMIT 1
    """, (setor, subgrupo))
    row = cur.fetchone()
    con.close()
    if not row:
        return {"seg": 0, "ter": 0, "qua": 0, "qui": 0, "sex": 0, "sáb": 0}
    return {"seg": row[0], "ter": row[1], "qua": row[2], "qui": row[3], "sex": row[4], "sáb": row[5]}

def set_subgrupo_regras(setor: str, subgrupo: str, regras: dict):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        INSERT OR REPLACE INTO subgrupo_regras(setor, subgrupo, evitar_seg, evitar_ter, evitar_qua, evitar_qui, evitar_sex, evitar_sab)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        setor, subgrupo,
        int(regras.get("seg", 0)),
        int(regras.get("ter", 0)),
        int(regras.get("qua", 0)),
        int(regras.get("qui", 0)),
        int(regras.get("sex", 0)),
        int(regras.get("sáb", 0)),
    ))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

# =========================================================
# FÉRIAS
# =========================================================
def add_ferias(setor: str, chapa: str, inicio: date, fim: date):
    con = db_conn()
    cur = con.cursor()
    cur.execute("INSERT INTO ferias(setor, chapa, inicio, fim) VALUES (?, ?, ?, ?)",
                (setor, chapa, inicio.strftime("%Y-%m-%d"), fim.strftime("%Y-%m-%d")))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

def delete_ferias_row(setor: str, chapa: str, inicio: str, fim: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        DELETE FROM ferias
        WHERE setor=? AND chapa=? AND inicio=? AND fim=?
    """, (setor, chapa, inicio, fim))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

@st.cache_data(show_spinner=False, ttl=120)
def list_ferias(setor: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT chapa, inicio, fim FROM ferias WHERE setor=? ORDER BY date(inicio) ASC", (setor,))
    rows = cur.fetchall()
    con.close()
    return rows

def is_de_ferias(setor: str, chapa: str, data_obj: date) -> bool:
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT 1 FROM ferias
        WHERE setor=? AND chapa=?
          AND date(inicio) <= date(?) AND date(fim) >= date(?)
        LIMIT 1
    """, (setor, chapa, data_obj.strftime("%Y-%m-%d"), data_obj.strftime("%Y-%m-%d")))
    ok = cur.fetchone() is not None
    con.close()
    return ok

def get_ultimo_fim_ferias_antes_de(setor: str, chapa: str, data_obj: date) -> date | None:
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT fim FROM ferias
        WHERE setor=? AND chapa=? AND date(fim) < date(?)
        ORDER BY date(fim) DESC
        LIMIT 1
    """, (setor, chapa, data_obj.strftime("%Y-%m-%d")))
    row = cur.fetchone()
    con.close()
    if not row:
        return None
    try:
        return datetime.strptime(row[0], "%Y-%m-%d").date()
    except Exception:
        return None


def is_first_week_after_return(setor: str, chapa: str, data_obj: date) -> bool:
    """
    `data_obj` representa a SEGUNDA da semana analisada.
    A 1a semana de retorno é a semana (seg->dom) que contém o dia de retorno.
    """
    semana_inicio = data_obj
    semana_fim = semana_inicio + timedelta(days=6)
    fim = get_ultimo_fim_ferias_antes_de(setor, chapa, semana_fim + timedelta(days=1))
    if not fim:
        return False
    retorno = fim + timedelta(days=1)
    return semana_inicio <= retorno <= semana_fim


def _apply_regra_retorno_ferias_primeiro_domingo(hist_all: dict, colab_by_chapa: dict, locked_idx: dict, df_ref_cur: pd.DataFrame, setor: str) -> None:
    """
    Regra especial somente para a primeira semana de retorno de férias:
    - considera o último dia de férias como a folga da semana de retorno;
    - analisa somente o primeiro domingo após o retorno;
    - se esse domingo estiver entre os menos cobertos do setor, o colaborador trabalha;
      caso contrário, folga;
    - se trabalhar nesse domingo e esse for o 5o dia trabalhado desde o retorno,
      a segunda seguinte vira folga obrigatória (e conta na semana seguinte).
    """
    if df_ref_cur is None or len(df_ref_cur) == 0:
        return

    ref = df_ref_cur.reset_index(drop=True).copy()
    ref["Data"] = pd.to_datetime(ref["Data"], errors="coerce")
    sunday_idxs = [i for i in range(len(ref)) if str(ref.loc[i, "Dia"]) == "dom"]
    if not sunday_idxs:
        return

    def _set_status(df: pd.DataFrame, idx: int, status: str, entrada_padrao: str) -> bool:
        if idx < 0 or idx >= len(df):
            return False
        if idx in set(locked_idx.get(ch, set())):
            return False
        atual = str(df.loc[idx, "Status"])
        if atual == "Férias":
            return False
        if status == "Folga":
            df.loc[idx, "Status"] = "Folga"
            df.loc[idx, "H_Entrada"] = ""
            df.loc[idx, "H_Saida"] = ""
            return True
        df.loc[idx, "Status"] = "Trabalho"
        ent = str(df.loc[idx, "H_Entrada"] or "").strip() or str(entrada_padrao or "06:00")
        df.loc[idx, "H_Entrada"] = ent
        df.loc[idx, "H_Saida"] = _saida_from_entrada(ent)
        return True

    for ch, df in list((hist_all or {}).items()):
        fim = get_ultimo_fim_ferias_antes_de(setor, ch, ref.loc[len(ref)-1, "Data"].date() + timedelta(days=1))
        if not fim:
            continue
        retorno = fim + timedelta(days=1)
        # só aplica se o retorno cai em alguma semana coberta pelo mês atual
        if retorno > ref.loc[len(ref)-1, "Data"].date():
            continue

        sunday_after_return = None
        for i in sunday_idxs:
            d = ref.loc[i, "Data"].date()
            if d >= retorno:
                sunday_after_return = i
                break
        if sunday_after_return is None:
            continue

        # aplica só se esse domingo pertence à 1a semana de retorno
        semana_inicio = retorno - timedelta(days=retorno.weekday())
        semana_fim = semana_inicio + timedelta(days=6)
        sunday_date = ref.loc[sunday_after_return, "Data"].date()
        if not (semana_inicio <= sunday_date <= semana_fim):
            continue

        # cobertura dos domingos do setor no mês
        sunday_counts = {}
        for sidx in sunday_idxs:
            total = 0
            for _ch2, _df2 in hist_all.items():
                try:
                    if str(_df2.loc[sidx, "Status"]) in WORK_STATUSES:
                        total += 1
                except Exception:
                    pass
            sunday_counts[sidx] = total
        min_cov = min(sunday_counts.values()) if sunday_counts else 0
        precisa_trabalhar = sunday_counts.get(sunday_after_return, 0) <= min_cov

        entrada_padrao = str((colab_by_chapa.get(ch, {}) or {}).get("Entrada", "06:00") or "06:00")
        if precisa_trabalhar:
            _set_status(df, sunday_after_return, "Trabalho", entrada_padrao)
            # conta dias trabalhados desde o retorno até o domingo inclusive
            consec = 0
            for i in range(len(ref)):
                d = ref.loc[i, "Data"].date()
                if d < retorno or d > sunday_date:
                    continue
                if str(df.loc[i, "Status"]) in WORK_STATUSES:
                    consec += 1
                else:
                    consec = 0
            if consec >= 5:
                monday_date = sunday_date + timedelta(days=1)
                monday_idx = None
                for i in range(len(ref)):
                    if ref.loc[i, "Data"].date() == monday_date:
                        monday_idx = i
                        break
                if monday_idx is not None:
                    _set_status(df, monday_idx, "Folga", entrada_padrao)
        else:
            _set_status(df, sunday_after_return, "Folga", entrada_padrao)

        hist_all[ch] = df

# =========================================================
# ESTADO
# =========================================================
def save_estado_mes(setor: str, ano: int, mes: int, estado: dict):
    con = db_conn()
    cur = con.cursor()
    for chapa, stt in estado.items():
        cur.execute("""
            INSERT OR REPLACE INTO estado_mes_anterior(setor, chapa, consec_trab_final, ultima_saida, ultimo_domingo_status, ano, mes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            setor, chapa,
            int(stt.get("consec_trab_final", 0)),
            stt.get("ultima_saida", "") or "",
            stt.get("ultimo_domingo_status", None),
            int(ano), int(mes)
        ))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

@st.cache_data(show_spinner=False, ttl=30)
def load_estado_prev(setor: str, ano: int, mes: int):
    """
    Carrega estado do mês anterior (consecutivos, última saída e status do último domingo)
    para manter continuidade entre competências.

    ✅ Robustez extra:
    - Se a tabela estado_mes_anterior não tiver o domingo (None) ou não tiver registro do colaborador,
      tenta inferir o "ultimo_domingo_status" a partir da escala do mês anterior salva em escala_mes.
    """
    prev_ano, prev_mes = ano, mes - 1
    if prev_mes == 0:
        prev_mes = 12
        prev_ano -= 1

    def _infer_ultimo_domingo_status_from_escala(chapa: str) -> str | None:
        try:
            con2 = db_conn()
            dfp = pd.read_sql_query(
                """
                SELECT dia, status
                FROM escala_mes
                WHERE setor=? AND ano=? AND mes=? AND chapa=?
                ORDER BY dia ASC
                """,
                con2,
                params=(setor, int(prev_ano), int(prev_mes), str(chapa)),
            )
            con2.close()
            if dfp is None or dfp.empty:
                return None
            # pega último domingo (dia == 'dom') do mês anterior
            for i in range(len(dfp) - 1, -1, -1):
                if str(dfp.loc[i, "dia"]).strip().lower() in ("dom", "domingo"):
                    stt = str(dfp.loc[i, "status"] or "").strip()
                    if stt == "Folga":
                        return "Folga"
                    if stt in WORK_STATUSES:
                        return "Trabalho"
                    # se for férias/blank, continua procurando domingo anterior
            return None
        except Exception:
            return None

    con = db_conn()
    cur = con.cursor()
    cur.execute(
        """
        SELECT chapa, consec_trab_final, ultima_saida, ultimo_domingo_status
        FROM estado_mes_anterior
        WHERE setor=? AND ano=? AND mes=?
        """,
        (setor, int(prev_ano), int(prev_mes)),
    )
    rows = cur.fetchall()
    con.close()

    estado: dict[str, dict] = {}
    for chapa, consec, ultima_saida, ultimo_dom in rows:
        estado[str(chapa)] = {
            "consec_trab_final": int(consec),
            "ultima_saida": ultima_saida or "",
            "ultimo_domingo_status": ultimo_dom,
        }

    # fallback do domingo quando estiver ausente
    for chapa in list(estado.keys()):
        if not (estado[chapa].get("ultimo_domingo_status") in ("Trabalho", "Folga")):
            estado[chapa]["ultimo_domingo_status"] = _infer_ultimo_domingo_status_from_escala(chapa)

    return estado


def infer_ultimo_domingo_status_from_escala(setor: str, ano: int, mes: int, chapa: str) -> str | None:
    """
    Inferir o status do ÚLTIMO domingo do mês anterior para manter continuidade DOM 1x1.

    Fonte de verdade:
      1) escala_mes (mês anterior) + overrides do mês anterior (se houver)
      2) retorna "Folga" ou "Trabalho" (ou None se não achar)

    Observação importante:
    - No banco, 'dia' é o dia do mês (inteiro). O dia da semana está em 'dia_sem'.
    """
    prev_ano, prev_mes = int(ano), int(mes) - 1
    if prev_mes == 0:
        prev_mes = 12
        prev_ano -= 1

    try:
        con = db_conn()

        # Escala do mês anterior (por dia)
        dfp = pd.read_sql_query(
            """
            SELECT dia, dia_sem, status
            FROM escala_mes
            WHERE setor=? AND ano=? AND mes=? AND chapa=?
            ORDER BY dia ASC
            """,
            con,
            params=(str(setor), int(prev_ano), int(prev_mes), str(chapa)),
        )

        # Overrides do mês anterior (se existirem)
        ov = pd.read_sql_query(
            """
            SELECT dia, campo, valor
            FROM overrides
            WHERE setor=? AND ano=? AND mes=? AND chapa=?
            """,
            con,
            params=(str(setor), int(prev_ano), int(prev_mes), str(chapa)),
        )

        con.close()

        if dfp is None or dfp.empty:
            return None

        # aplica override de status (somente campo=status)
        if ov is not None and not ov.empty:
            try:
                ov_s = ov[ov["campo"] == "status"][["dia", "valor"]].copy()
                if not ov_s.empty:
                    ov_map = {int(r["dia"]): str(r["valor"]) for _, r in ov_s.iterrows()}
                    for i in range(len(dfp)):
                        d_int = int(dfp.loc[i, "dia"])
                        if d_int in ov_map:
                            dfp.loc[i, "status"] = ov_map[d_int]
            except Exception:
                pass

        # último domingo do mês anterior (de trás pra frente)
        for i in range(len(dfp) - 1, -1, -1):
            try:
                d_int = int(dfp.loc[i, "dia"])
                # calcula dia da semana real
                is_sun = (pd.Timestamp(year=int(prev_ano), month=int(prev_mes), day=int(d_int)).day_name() == "Sunday")
            except Exception:
                is_sun = False
            if is_sun:
                stt = str(dfp.loc[i, "status"] or "").strip()
                if stt == "Folga":
                    return "Folga"
                if stt in WORK_STATUSES:
                    return "Trabalho"
                # se cair em férias/blank, continua procurando domingo anterior

        return None
    except Exception:
        return None



def _norm_override_campo(campo: str) -> str:
    c = str(campo or '').strip().lower()
    mapa = {
        'status': 'status',
        'h_entrada': 'h_entrada',
        'entrada': 'h_entrada',
        'h_saida': 'h_saida',
        'saida': 'h_saida',
    }
    return mapa.get(c, c)


def set_override(setor: str, ano: int, mes: int, chapa: str, dia: int, campo: str, valor: str):
    """
    Cria/atualiza um override (UPSERT) na tabela overrides.
    Todos os campos são normalizados para o padrão usado pelo motor:
    status / h_entrada / h_saida.
    """
    campo = _norm_override_campo(campo)
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        INSERT INTO overrides(setor, ano, mes, chapa, dia, campo, valor)
        VALUES(?,?,?,?,?,?,?)
        ON CONFLICT(setor, ano, mes, chapa, dia, campo)
        DO UPDATE SET valor=excluded.valor
    """, (str(setor).strip().upper(), int(ano), int(mes), str(chapa).strip(), int(dia), campo, str(valor).strip()))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass


def delete_override(setor: str, ano: int, mes: int, chapa: str, dia: int, campo: str | None = None):
    con = db_conn()
    cur = con.cursor()
    if campo:
        campos = sorted({str(campo).strip(), _norm_override_campo(campo)})
        placeholders = ",".join(["?"] * len(campos))
        cur.execute(f"""
            DELETE FROM overrides
            WHERE setor=? AND ano=? AND mes=? AND chapa=? AND dia=? AND campo IN ({placeholders})
        """, (setor, int(ano), int(mes), chapa, int(dia), *campos))
    else:
        cur.execute("""
            DELETE FROM overrides
            WHERE setor=? AND ano=? AND mes=? AND chapa=? AND dia=?
        """, (setor, int(ano), int(mes), chapa, int(dia)))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass


def delete_overrides_mes(setor: str, ano: int, mes: int, keep_campos: set[str] | None = None):
    """
    Remove overrides do mês inteiro (útil para "Gerar do zero").
    Por padrão remove TUDO para o mês. Se keep_campos for informado,
    preserva overrides cujo campo esteja em keep_campos.
    """
    con = db_conn()
    cur = con.cursor()
    if keep_campos and len(keep_campos) > 0:
        # mantém alguns campos (ex.: se quiser preservar algo específico)
        placeholders = ",".join(["?"] * len(keep_campos))
        cur.execute(
            f"""
            DELETE FROM overrides
            WHERE setor=? AND ano=? AND mes=? AND campo NOT IN ({placeholders})
            """,
            (setor, int(ano), int(mes), *list(keep_campos)),
        )
    else:
        cur.execute(
            """
            DELETE FROM overrides
            WHERE setor=? AND ano=? AND mes=?
            """,
            (setor, int(ano), int(mes)),
        )
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

@st.cache_data(show_spinner=False, ttl=120)
def load_overrides(setor: str, ano: int, mes: int):
    con = db_conn()
    df = pd.read_sql_query("""
        SELECT chapa, dia, campo, valor
        FROM overrides
        WHERE setor=? AND ano=? AND mes=?
    """, con, params=(setor, int(ano), int(mes)))
    con.close()
    return df

WEEKDAY_LABELS = ["seg", "ter", "qua", "qui", "sex", "sáb", "dom"]
WEEKDAY_LABELS_LONG = {
    0: "Segunda-feira",
    1: "Terça-feira",
    2: "Quarta-feira",
    3: "Quinta-feira",
    4: "Sexta-feira",
    5: "Sábado",
    6: "Domingo",
}


def _parse_hora_min(v: str) -> tuple[int, int]:
    s = str(v or "").strip()
    if not re.fullmatch(r"\d{2}:\d{2}", s):
        return (99, 99)
    h, m = s.split(":", 1)
    return int(h), int(m)


def _classificar_turno_por_entrada(h_entrada: str) -> str:
    hh, mm = _parse_hora_min(h_entrada)
    total = hh * 60 + mm
    if total == 99 * 60 + 99:
        return ""
    if 6 * 60 <= total <= 10 * 60:
        return "Abertura"
    if 10 * 60 + 1 <= total <= 12 * 60 + 19:
        return "Intermediário"
    if total >= 12 * 60 + 20:
        return "Fechamento"
    return ""




def _ensure_folga_fixa_schema():
    con = db_conn()
    try:
        cur = con.cursor()
        _safe_exec(cur, """
        CREATE TABLE IF NOT EXISTS folga_fixa (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            setor TEXT NOT NULL,
            chapa TEXT NOT NULL,
            dia_semana INTEGER NOT NULL,
            ativo INTEGER NOT NULL DEFAULT 1,
            criado_em TEXT NOT NULL,
            UNIQUE(setor, chapa, dia_semana)
        )
        """)
        try:
            cols = {str(r[1]).strip().lower() for r in cur.execute("PRAGMA table_info(folga_fixa)").fetchall()}
        except Exception:
            cols = set()
        if 'ativo' not in cols:
            _safe_exec(cur, "ALTER TABLE folga_fixa ADD COLUMN ativo INTEGER NOT NULL DEFAULT 1")
        if 'criado_em' not in cols:
            _safe_exec(cur, "ALTER TABLE folga_fixa ADD COLUMN criado_em TEXT")
            _safe_exec(cur, "UPDATE folga_fixa SET criado_em = COALESCE(criado_em, ?) WHERE criado_em IS NULL OR TRIM(COALESCE(criado_em,'')) = ''", (datetime.now().isoformat(),))
        con.commit()
    finally:
        con.close()


def _ensure_inventario_diario_schema():
    con = db_conn()
    try:
        cur = con.cursor()
        _safe_exec(cur, """
        CREATE TABLE IF NOT EXISTS inventario_diario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            setor TEXT NOT NULL,
            ano INTEGER NOT NULL,
            mes INTEGER NOT NULL,
            dia INTEGER NOT NULL,
            abertura INTEGER NOT NULL DEFAULT 0,
            intermediario INTEGER NOT NULL DEFAULT 0,
            fechamento INTEGER NOT NULL DEFAULT 0,
            criado_em TEXT NOT NULL,
            atualizado_em TEXT NOT NULL,
            UNIQUE(setor, ano, mes, dia)
        )
        """)
        try:
            cols = {str(r[1]).strip().lower() for r in cur.execute("PRAGMA table_info(inventario_diario)").fetchall()}
        except Exception:
            cols = set()
        if 'abertura' not in cols:
            _safe_exec(cur, "ALTER TABLE inventario_diario ADD COLUMN abertura INTEGER NOT NULL DEFAULT 0")
        if 'intermediario' not in cols:
            _safe_exec(cur, "ALTER TABLE inventario_diario ADD COLUMN intermediario INTEGER NOT NULL DEFAULT 0")
        if 'fechamento' not in cols:
            _safe_exec(cur, "ALTER TABLE inventario_diario ADD COLUMN fechamento INTEGER NOT NULL DEFAULT 0")
        if 'criado_em' not in cols:
            _safe_exec(cur, "ALTER TABLE inventario_diario ADD COLUMN criado_em TEXT")
        if 'atualizado_em' not in cols:
            _safe_exec(cur, "ALTER TABLE inventario_diario ADD COLUMN atualizado_em TEXT")
        agora = datetime.now().isoformat()
        _safe_exec(cur, "UPDATE inventario_diario SET criado_em = COALESCE(criado_em, ?) WHERE criado_em IS NULL OR TRIM(COALESCE(criado_em,'')) = ''", (agora,))
        _safe_exec(cur, "UPDATE inventario_diario SET atualizado_em = COALESCE(atualizado_em, ?) WHERE atualizado_em IS NULL OR TRIM(COALESCE(atualizado_em,'')) = ''", (agora,))
        con.commit()
    finally:
        con.close()

def list_folga_fixa(setor: str, chapa: str | None = None) -> pd.DataFrame:
    _ensure_folga_fixa_schema()
    setor = _norm_setor(setor)
    con = db_conn()
    params = [setor]
    sql = """
        SELECT f.chapa AS Chapa,
               f.dia_semana AS DiaSemana,
               COALESCE(f.ativo, 1) AS Ativo,
               COALESCE(f.criado_em, '') AS CriadoEm
        FROM folga_fixa f
        WHERE UPPER(TRIM(f.setor)) = ?
    """
    if chapa is not None:
        sql += " AND TRIM(f.chapa) = ?"
        params.append(_norm_chapa(chapa))
    sql += " ORDER BY f.chapa, f.dia_semana"
    try:
        df = pd.read_sql_query(sql, con, params=tuple(params))
    finally:
        con.close()
    if df.empty:
        df["Nome"] = []
        df["Dia"] = []
        return df

    nomes = {}
    try:
        for c in load_colaboradores_setor(setor):
            nomes[str(c.get('Chapa', '') or '').strip()] = str(c.get('Nome', '') or '').strip()
    except Exception:
        nomes = {}

    df["Nome"] = df["Chapa"].apply(lambda ch: nomes.get(str(ch).strip(), str(ch).strip()))
    df["Dia"] = df["DiaSemana"].map(WEEKDAY_LABELS_LONG)
    df["Ativo"] = df["Ativo"].apply(lambda x: "Sim" if int(x or 0) else "Não")
    return df


def save_folga_fixa(setor: str, chapa: str, weekdays: list[int]):
    _ensure_folga_fixa_schema()
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    weekdays = sorted({int(x) for x in weekdays if int(x) in range(7)})
    agora = datetime.now().isoformat()
    con = db_conn()
    cur = con.cursor()
    cur.execute("DELETE FROM folga_fixa WHERE setor=? AND chapa=?", (setor, chapa))
    for wd in weekdays:
        cur.execute(
            "INSERT OR REPLACE INTO folga_fixa(setor, chapa, dia_semana, ativo, criado_em) VALUES(?,?,?,?,?)",
            (setor, chapa, int(wd), 1, agora),
        )
    con.commit()
    con.close()


def remove_folga_fixa(setor: str, chapa: str):
    _ensure_folga_fixa_schema()
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    con = db_conn()
    cur = con.cursor()
    cur.execute("DELETE FROM folga_fixa WHERE setor=? AND chapa=?", (setor, chapa))
    con.commit()
    con.close()


def get_folga_fixa_weekdays(setor: str, chapa: str) -> list[int]:
    df = list_folga_fixa(setor, chapa)
    if df.empty:
        return []
    return sorted({int(x) for x in df["DiaSemana"].tolist() if pd.notna(x)})


def _dias_mes_por_weekdays(ano: int, mes: int, weekdays: list[int]) -> list[int]:
    out = []
    sel = set(int(x) for x in weekdays)
    for d in range(1, calendar.monthrange(int(ano), int(mes))[1] + 1):
        if date(int(ano), int(mes), d).weekday() in sel:
            out.append(d)
    return out


def _folga_fixa_days_map(setor: str, ano: int, mes: int) -> dict[str, list[int]]:
    """Retorna os dias do mês que devem ficar travados como folga fixa por colaborador.

    A folga fixa não apaga nenhuma regra do sistema. Ela só entra como trava prioritária
    quando a geração é feita em 'respeita ajustes'.
    """
    out: dict[str, list[int]] = {}
    try:
        df = list_folga_fixa(setor)
    except Exception:
        df = pd.DataFrame()
    if df is None or df.empty:
        return out

    max_day = calendar.monthrange(int(ano), int(mes))[1]
    for _, r in df.iterrows():
        try:
            if int(r.get('Ativo', 1) or 1) == 0:
                continue
        except Exception:
            pass
        ch = _norm_chapa(r.get('Chapa', ''))
        if not ch:
            continue
        try:
            wd = int(r.get('DiaSemana'))
        except Exception:
            continue
        dias = [d for d in range(1, max_day + 1) if date(int(ano), int(mes), d).weekday() == wd]
        if dias:
            out.setdefault(ch, []).extend(dias)

    return {ch: sorted(set(int(d) for d in dias if 1 <= int(d) <= max_day)) for ch, dias in out.items()}


def _merge_folga_fixa_into_ovmap(setor: str, ano: int, mes: int, ovmap: dict | None) -> dict:
    """Injeta a folga fixa no mapa de overrides para a geração respeitando ajustes.

    Prioridade: férias continuam soberanas. Se já existir override manual de status, ele é mantido.
    A folga fixa só ocupa o dia quando não há outro status manual definido.
    """
    merged = {}
    try:
        for ch, dias in (ovmap or {}).items():
            merged[str(ch)] = {int(d): dict(payload or {}) for d, payload in (dias or {}).items()}
    except Exception:
        merged = {}

    ff_map = _folga_fixa_days_map(setor, int(ano), int(mes))
    if not ff_map:
        return merged

    for ch, dias in ff_map.items():
        for dia in dias:
            data_obj = date(int(ano), int(mes), int(dia))
            if is_de_ferias(setor, ch, data_obj):
                continue
            payload = merged.setdefault(str(ch), {}).setdefault(int(dia), {})
            status_existente = str(payload.get('status') or '').strip()
            if status_existente:
                continue
            payload['status'] = 'Folga'
    return merged


def _week_chunks_for_month(ano: int, mes: int) -> list[list[int]]:
    dias = _dias_mes(int(ano), int(mes))
    ref = pd.DataFrame({'Data': dias, 'Dia': [D_PT[d.day_name()] for d in dias]})
    weeks = []
    current = []
    for i in range(len(ref)):
        current.append(i)
        if str(ref.loc[i, 'Dia']) == 'dom':
            weeks.append(current)
            current = []
    if current:
        weeks.append(current)
    return weeks


def _simulate_folga_fixa_warnings(df_hist: pd.DataFrame, ano: int, mes: int, dias_fixos: list[int]) -> list[str]:
    warns = []
    if df_hist is None or len(df_hist) == 0:
        return ["Não existe escala gerada para validar a folga fixa nesta competência."]
    sim = df_hist.reset_index(drop=True).copy()
    dias_fixos = sorted(set(int(x) for x in dias_fixos))
    for dia in dias_fixos:
        idx = dia - 1
        if idx < 0 or idx >= len(sim):
            continue
        atual = str(sim.loc[idx, 'Status'])
        if atual == 'Férias':
            warns.append(f"Dia {dia:02d}: já está em Férias.")
            continue
        if atual == 'Afastamento':
            warns.append(f"Dia {dia:02d}: já está em Afastamento.")
        sim.loc[idx, 'Status'] = 'Folga'
    # folga consecutiva
    for i in range(1, len(sim)):
        if str(sim.loc[i-1, 'Status']) == 'Folga' and str(sim.loc[i, 'Status']) == 'Folga':
            warns.append(f"Folga consecutiva em {i:02d}/{mes:02d} e {i+1:02d}/{mes:02d}.")
    # semanas > 2 folgas
    weeks = _week_chunks_for_month(ano, mes)
    for w in weeks:
        qtd = sum(1 for i in w if str(sim.loc[i, 'Status']) == 'Folga')
        if qtd > 2:
            ini = int(sim.loc[w[0], 'Data'].day)
            fim = int(sim.loc[w[-1], 'Data'].day)
            warns.append(f"Semana {ini:02d}-{fim:02d} ficará com {qtd} folgas.")
    # domingo
    for dia in dias_fixos:
        if date(int(ano), int(mes), int(dia)).weekday() == 6:
            warns.append(f"Domingo {dia:02d}/{mes:02d} foi marcado como folga fixa; confira a alternância 1x1.")
    # conflitos com inventário
    return list(dict.fromkeys(warns))


def upsert_inventario_dia(setor: str, ano: int, mes: int, dia: int, abertura: int, intermediario: int, fechamento: int):
    setor = _norm_setor(setor)
    agora = datetime.now().isoformat()
    con = db_conn()
    cur = con.cursor()
    cur.execute(
        """
        INSERT INTO inventario_diario(setor, ano, mes, dia, abertura, intermediario, fechamento, criado_em, atualizado_em)
        VALUES(?,?,?,?,?,?,?,?,?)
        ON CONFLICT(setor, ano, mes, dia)
        DO UPDATE SET abertura=excluded.abertura,
                      intermediario=excluded.intermediario,
                      fechamento=excluded.fechamento,
                      atualizado_em=excluded.atualizado_em
        """,
        (setor, int(ano), int(mes), int(dia), int(abertura), int(intermediario), int(fechamento), agora, agora),
    )
    con.commit()
    con.close()


def get_inventario_mes(setor: str, ano: int, mes: int) -> pd.DataFrame:
    setor = _norm_setor(setor)
    if not setor or not ano or not mes:
        return pd.DataFrame(columns=['Dia', 'Abertura', 'Intermediario', 'Fechamento'])
    con = db_conn()
    try:
        df = pd.read_sql_query(
            """
            SELECT dia AS Dia, abertura AS Abertura, intermediario AS Intermediario, fechamento AS Fechamento,
                   criado_em AS CriadoEm, atualizado_em AS AtualizadoEm
            FROM inventario_diario
            WHERE UPPER(TRIM(setor))=? AND ano=? AND mes=?
            ORDER BY dia
            """,
            con,
            params=(setor, int(ano), int(mes)),
        )
    finally:
        con.close()
    return df


def build_historico_folgas_diario(setor: str, ano: int, mes: int, hist_db: dict | None = None) -> pd.DataFrame:
    setor = _norm_setor(setor)
    hist_db = hist_db or get_hist_mes_com_overrides_cached(setor, ano, mes)
    dias = _dias_mes(int(ano), int(mes))
    rows = []
    for idx, dt_ in enumerate(dias):
        folga = ferias = afast = trabalho = 0
        nomes_folga = []
        for ch, df in (hist_db or {}).items():
            if df is None or idx >= len(df):
                continue
            stt = str(df.loc[idx, 'Status'])
            if stt == 'Folga':
                folga += 1
                try:
                    nomes_folga.append(str(df.loc[idx, 'Nome']))
                except Exception:
                    nomes_folga.append(str(ch))
            elif stt == 'Férias':
                ferias += 1
            elif stt == 'Afastamento':
                afast += 1
            elif stt in WORK_STATUSES:
                trabalho += 1
        rows.append({
            'Dia': int(dt_.day),
            'Data': dt_.strftime('%d/%m/%Y'),
            'Semana': WEEKDAY_LABELS_LONG.get(dt_.weekday(), ''),
            'Folga': int(folga),
            'Férias': int(ferias),
            'Afastamento': int(afast),
            'Trabalho': int(trabalho),
            'Pessoas de folga': ', '.join(sorted([n for n in nomes_folga if n]))
        })
    return pd.DataFrame(rows)


def build_inventario_comparativo(setor: str, ano: int, mes: int, hist_db: dict | None = None) -> pd.DataFrame:
    hist_db = hist_db or get_hist_mes_com_overrides_cached(setor, ano, mes)
    inv = get_inventario_mes(setor, ano, mes)
    inv_map = {int(r['Dia']): r for _, r in inv.iterrows()} if inv is not None and not inv.empty else {}
    rows = []
    for d in range(1, calendar.monthrange(int(ano), int(mes))[1] + 1):
        ab = inter = fech = 0
        for _, df in (hist_db or {}).items():
            if df is None or d - 1 >= len(df):
                continue
            if str(df.loc[d - 1, 'Status']) not in WORK_STATUSES:
                continue
            turno = _classificar_turno_por_entrada(str(df.loc[d - 1, 'H_Entrada'] or ''))
            if turno == 'Abertura':
                ab += 1
            elif turno == 'Intermediário':
                inter += 1
            elif turno == 'Fechamento':
                fech += 1
        meta = inv_map.get(d, {})
        meta_ab = int(meta['Abertura']) if meta != {} else 0
        meta_in = int(meta['Intermediario']) if meta != {} else 0
        meta_fe = int(meta['Fechamento']) if meta != {} else 0
        rows.append({
            'Dia': d,
            'Data': date(int(ano), int(mes), d).strftime('%d/%m/%Y'),
            'Meta abertura': meta_ab,
            'Real abertura': ab,
            'Dif abertura': ab - meta_ab,
            'Meta intermediário': meta_in,
            'Real intermediário': inter,
            'Dif intermediário': inter - meta_in,
            'Meta fechamento': meta_fe,
            'Real fechamento': fech,
            'Dif fechamento': fech - meta_fe,
        })
    return pd.DataFrame(rows)


def build_cobertura_diaria_geral(setor: str, ano: int, mes: int, hist_db: dict | None = None) -> pd.DataFrame:
    hist_db = hist_db or get_hist_mes_com_overrides_cached(setor, ano, mes)
    dias = _dias_mes(int(ano), int(mes))
    rows = []
    for d in range(1, calendar.monthrange(int(ano), int(mes))[1] + 1):
        ab = inter = fech = total = folga = ferias = afast = 0
        for _, df in (hist_db or {}).items():
            if df is None or d - 1 >= len(df):
                continue
            stt = str(df.loc[d - 1, 'Status'])
            if stt in WORK_STATUSES:
                total += 1
                turno = _classificar_turno_por_entrada(str(df.loc[d - 1, 'H_Entrada'] or ''))
                if turno == 'Abertura':
                    ab += 1
                elif turno == 'Intermediário':
                    inter += 1
                elif turno == 'Fechamento':
                    fech += 1
            elif stt == 'Folga':
                folga += 1
            elif stt == 'Férias':
                ferias += 1
            elif stt == 'Afastamento':
                afast += 1
        dt_ = date(int(ano), int(mes), int(d))
        rows.append({
            'Dia': int(d),
            'Data': dt_.strftime('%d/%m/%Y'),
            'Semana': WEEKDAY_LABELS_LONG.get(dt_.weekday(), ''),
            'Abertura': int(ab),
            'Intermediário': int(inter),
            'Fechamento': int(fech),
            'Total trabalhando': int(total),
            'Folga': int(folga),
            'Férias': int(ferias),
            'Afastamento': int(afast),
        })
    return pd.DataFrame(rows)


def build_cobertura_por_subgrupo_no_dia(setor: str, ano: int, mes: int, dia: int, hist_db: dict | None = None) -> pd.DataFrame:
    hist_db = hist_db or get_hist_mes_com_overrides_cached(setor, ano, mes)
    colaboradores = load_colaboradores_setor(setor)
    colab_by = {str((c or {}).get('Chapa', '')).strip(): (c or {}) for c in (colaboradores or [])}
    resumo = {}
    idx = int(dia) - 1

    for ch, df in (hist_db or {}).items():
        if df is None or idx < 0 or idx >= len(df):
            continue
        ch_str = str(ch).strip()
        sg = (colab_by.get(ch_str, {}).get('Subgrupo', '') or '').strip() or 'SEM SUBGRUPO'
        item = resumo.setdefault(sg, {
            'Subgrupo': sg,
            'Abertura': 0,
            'Intermediário': 0,
            'Fechamento': 0,
            'Total trabalhando': 0,
            'Folga': 0,
            'Férias': 0,
            'Afastamento': 0,
        })
        stt = str(df.loc[idx, 'Status'])
        if stt in WORK_STATUSES:
            item['Total trabalhando'] += 1
            turno = _classificar_turno_por_entrada(str(df.loc[idx, 'H_Entrada'] or ''))
            if turno == 'Abertura':
                item['Abertura'] += 1
            elif turno == 'Intermediário':
                item['Intermediário'] += 1
            elif turno == 'Fechamento':
                item['Fechamento'] += 1
        elif stt == 'Folga':
            item['Folga'] += 1
        elif stt == 'Férias':
            item['Férias'] += 1
        elif stt == 'Afastamento':
            item['Afastamento'] += 1

    if not resumo:
        return pd.DataFrame(columns=['Subgrupo', 'Abertura', 'Intermediário', 'Fechamento', 'Total trabalhando', 'Folga', 'Férias', 'Afastamento'])

    out = pd.DataFrame(list(resumo.values()))
    out = out.sort_values(['Subgrupo']).reset_index(drop=True)
    return out


def _ov_map(setor: str, ano: int, mes: int):
    df = load_overrides(setor, ano, mes)
    ov = {}
    if df is None or df.empty:
        return ov
    for _, r in df.iterrows():
        ch = str(r["chapa"])
        dia = int(r["dia"])
        campo = _norm_override_campo(r["campo"])
        valor = str(r["valor"])
        ov.setdefault(ch, {}).setdefault(dia, {})[campo] = valor
    return ov

def _is_status_locked(ovmap: dict, chapa: str, data_ts: pd.Timestamp) -> bool:
    dia = int(pd.to_datetime(data_ts).day)
    return bool(ovmap.get(chapa, {}).get(dia, {}).get("status"))

def _apply_overrides_to_df_inplace(df: pd.DataFrame, setor: str, chapa: str, ovmap: dict):
    """Aplica ajustes manuais (overrides) no DataFrame.

    Regras de FÉRIAS:
    - Dias de férias são definidos SOMENTE pela tabela `ferias` (aba Férias).
    - Em dia de férias, força: Status='Férias', H_Entrada='', H_Saida='' (ignora qualquer override).
    - Override tentando marcar 'Férias' fora da tabela é ignorado.
    - Override tentando mudar o Status / horários em um dia que está em férias também é ignorado.
    """
    ovmap = (ovmap or {})
    if not ovmap:
        return df

    if "Data" in df.columns and not pd.api.types.is_datetime64_any_dtype(df["Data"]):
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce")

    for d_raw, payload in ovmap.items():
        if not payload:
            continue

        try:
            dia = int(d_raw)
        except Exception:
            dd = None
            if isinstance(d_raw, dt.datetime):
                dd = d_raw.date()
            elif isinstance(d_raw, dt.date):
                dd = d_raw
            elif isinstance(d_raw, str):
                try:
                    dd = dt.date.fromisoformat(d_raw[:10])
                except Exception:
                    dd = None
            if dd is None:
                continue
            mask = df["Data"].dt.date == dd
        else:
            mask = df["Data"].dt.day == dia
            dd = None
            if bool(mask.any()):
                dd = pd.to_datetime(df.loc[df.index[mask][0], "Data"]).date()

        if not bool(mask.any()):
            continue
        i = df.index[mask][0]
        if dd is None:
            dd = pd.to_datetime(df.loc[i, "Data"]).date()

        if is_de_ferias(setor, chapa, dd):
            df.loc[i, "Status"] = "Férias"
            df.loc[i, "H_Entrada"] = ""
            df.loc[i, "H_Saida"] = ""
            continue

        st_new = str(payload.get("status") or "").strip()
        if st_new and st_new.lower() not in ["férias", "ferias"]:
            df.loc[i, "Status"] = st_new
            if st_new.strip().upper() in ("FOLGA", "FOLG", "AFA", "AFASTAMENTO"):
                df.loc[i, "H_Entrada"] = ""
                df.loc[i, "H_Saida"] = ""

        ent_new = str(payload.get("h_entrada") or payload.get("entrada") or "").strip()
        if ent_new and str(df.loc[i, "Status"]).strip().upper() not in ("FOLGA", "FOLG", "FÉRIAS", "FERIAS", "FER", "AFA", "AFASTAMENTO"):
            df.loc[i, "H_Entrada"] = ent_new
            df.loc[i, "H_Saida"] = _saida_from_entrada(ent_new)

    return df


def save_escala_mes_db(setor: str, ano: int, mes: int, historico_df_por_chapa: dict[str, pd.DataFrame]):
    """Grava escala no banco de forma robusta.
    - Limpa a competência (setor/ano/mes) antes de gravar para evitar IntegrityError em DB antigo/corrompido.
    - Robustez contra NaT/NaN.
    """
    con = db_conn()
    cur = con.cursor()

    # Limpa o mês inteiro do setor antes de inserir (evita conflito/duplicidade em DB antigo)
    try:
        cur.execute("DELETE FROM escala_mes WHERE setor=? AND ano=? AND mes=?", (setor, int(ano), int(mes)))
        con.commit()
    except Exception:
        pass

    for chapa, df in historico_df_por_chapa.items():
        df2 = df.copy()
        df2.reset_index(drop=True, inplace=True)

        for j, row in df2.iterrows():
            dt = pd.to_datetime(row.get("Data", None), errors="coerce")
            max_day = calendar.monthrange(int(ano), int(mes))[1]

            if pd.isna(dt):
                dia = int(j) + 1
                if dia < 1: dia = 1
                if dia > max_day: dia = max_day
                dt = pd.Timestamp(year=int(ano), month=int(mes), day=int(dia))
            else:
                dia = int(getattr(dt, "day", 1) or 1)
                if dia < 1: dia = 1
                if dia > max_day:
                    dia = max_day
                    dt = pd.Timestamp(year=int(ano), month=int(mes), day=int(dia))

            dia_sem = row.get("Dia", "")
            if pd.isna(dia_sem): dia_sem = ""
            dia_sem = str(dia_sem)

            status = row.get("Status", "Trabalho")
            if pd.isna(status) or not str(status).strip():
                status = "Trabalho"
            status = str(status)

            h_ent = row.get("H_Entrada", "")
            h_sai = row.get("H_Saida", "")
            if pd.isna(h_ent): h_ent = ""
            if pd.isna(h_sai): h_sai = ""
            h_ent = str(h_ent or "")
            h_sai = str(h_sai or "")

            try:
                cur.execute("""
                    INSERT OR REPLACE INTO escala_mes(setor, ano, mes, chapa, dia, data, dia_sem, status, h_entrada, h_saida)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    setor, int(ano), int(mes), str(chapa), int(dia),
                    pd.to_datetime(dt).strftime("%Y-%m-%d"),
                    dia_sem,
                    status,
                    h_ent,
                    h_sai,
                ))
            except Exception:
                # não derruba o app por causa de uma linha ruim
                continue

    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

@st.cache_data(show_spinner=False, ttl=120)
def load_escala_mes_db(setor: str, ano: int, mes: int):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT chapa, data, dia_sem, status, h_entrada, h_saida
        FROM escala_mes
        WHERE setor=? AND ano=? AND mes=?
        ORDER BY chapa, date(data) ASC
    """, (setor, int(ano), int(mes)))
    rows = cur.fetchall()
    con.close()
    if not rows:
        return {}
    hist = {}
    for chapa, data_s, dia_sem, status, h_ent, h_sai in rows:
        ch = _norm_chapa(chapa)
        dt = pd.to_datetime(data_s)
        hist.setdefault(ch, []).append({
            "Data": dt, "Dia": dia_sem, "Status": status,
            "H_Entrada": h_ent or "", "H_Saida": h_sai or ""
        })
    return {_norm_chapa(ch): pd.DataFrame(items) for ch, items in hist.items()}


@st.cache_data(show_spinner=False, ttl=120)
def get_hist_mes_com_overrides_cached(setor: str, ano: int, mes: int):
    hist_db = load_escala_mes_db(setor, ano, mes)
    return apply_overrides_to_hist(setor, ano, mes, hist_db)

def apply_overrides_to_hist(setor: str, ano: int, mes: int, hist_db: dict[str, pd.DataFrame]):
    """
    Aplica overrides e retificações no histórico carregado do banco.
    REGRA GERAL:
    - "Férias" só existe se estiver na tabela ferias.
    - Se encontrar "Férias" no banco mas NÃO estiver na tabela, vira "Trabalho".
    - Retificações de competência fechada precisam aparecer no portal e em todas as leituras.
    """
    ov = load_overrides(setor, ano, mes)
    ff_map = _folga_fixa_days_map(setor, int(ano), int(mes))

    # aplica overrides (se houver)
    if ov is not None and not ov.empty and hist_db:
        for _, r in ov.iterrows():
            ch = str(r["chapa"])
            dia = int(r["dia"])
            campo = str(r["campo"])
            valor = str(r["valor"])
            if ch not in hist_db:
                continue

            df = hist_db[ch].copy()
            idx = dia - 1
            if idx < 0 or idx >= len(df):
                continue

            try:
                data_obj = pd.to_datetime(df.loc[idx, "Data"]).date()
            except Exception:
                data_obj = None

            if campo == "status":
                if valor == "Férias" and data_obj is not None and not is_de_ferias(setor, ch, data_obj):
                    pass
                else:
                    df.loc[idx, "Status"] = valor
                    if valor not in WORK_STATUSES:
                        df.loc[idx, "H_Entrada"] = ""
                        df.loc[idx, "H_Saida"] = ""

            elif campo == "h_entrada":
                df.loc[idx, "H_Entrada"] = valor
                if df.loc[idx, "Status"] in WORK_STATUSES:
                    df.loc[idx, "H_Saida"] = _saida_from_entrada(valor)

            elif campo == "h_saida":
                df.loc[idx, "H_Saida"] = valor

            hist_db[ch] = df

    # aplica folga fixa também na leitura do histórico
    if hist_db and ff_map:
        for ch, dias_fixos in ff_map.items():
            if ch not in hist_db:
                continue

            df = hist_db[ch].copy()
            for dia in sorted(set(int(x) for x in dias_fixos if int(x) > 0)):
                idx = int(dia) - 1
                if idx < 0 or idx >= len(df):
                    continue
                try:
                    data_obj = pd.to_datetime(df.loc[idx, "Data"]).date()
                except Exception:
                    continue
                if is_de_ferias(setor, ch, data_obj):
                    df.loc[idx, "Status"] = "Férias"
                    df.loc[idx, "H_Entrada"] = ""
                    df.loc[idx, "H_Saida"] = ""
                else:
                    df.loc[idx, "Status"] = "Folga"
                    df.loc[idx, "H_Entrada"] = ""
                    df.loc[idx, "H_Saida"] = ""
            hist_db[ch] = df

    # saneamento final de férias vindas do banco
    if hist_db:
        for ch, df in list(hist_db.items()):
            if df is None or df.empty:
                continue
            df2 = df.copy()
            try:
                ent_pad = get_colaborador_record(setor, ch).get("Entrada", "06:00")
            except Exception:
                ent_pad = "06:00"

            for i in range(len(df2)):
                try:
                    data_obj = pd.to_datetime(df2.loc[i, "Data"]).date()
                except Exception:
                    continue
                in_ferias = is_de_ferias(setor, ch, data_obj)

                if in_ferias:
                    df2.loc[i, "Status"] = "Férias"
                    df2.loc[i, "H_Entrada"] = ""
                    df2.loc[i, "H_Saida"] = ""
                else:
                    if df2.loc[i, "Status"] == "Férias":
                        df2.loc[i, "Status"] = "Trabalho"
                        df2.loc[i, "H_Entrada"] = ent_pad
                        df2.loc[i, "H_Saida"] = _saida_from_entrada(ent_pad)

            hist_db[ch] = df2

    # por último, aplica retificações e espelho mensal de subgrupo
    hist_db = _apply_retificacoes_to_hist(setor, ano, mes, hist_db)
    hist_db = _apply_subgrupo_competencia_to_hist(setor, ano, mes, hist_db)
    return hist_db

# =========================================================
# MOTOR
# =========================================================
def _dias_mes(ano: int, mes: int):
    qtd = calendar.monthrange(ano, mes)[1]
    return pd.date_range(start=f"{ano}-{mes:02d}-01", periods=qtd, freq="D")

def _nao_consecutiva_folga(df, idx):
    """
    Verifica se o índice 'idx' NÃO fica colado com outra folga (idx-1 ou idx+1).
    Usa iloc (posição) para evitar KeyError quando o índice do DF não é 0..N-1.
    """
    n = len(df)
    if n == 0:
        return True
    if idx > 0 and df.iloc[idx - 1]["Status"] == "Folga":
        return False
    if idx < n - 1 and df.iloc[idx + 1]["Status"] == "Folga":
        return False
    return True

def _set_trabalho(df, idx, ent_padrao, locked_status: set[int] | None = None):
    if _locked(locked_status, idx):
        return
    df.loc[idx, "Status"] = "Trabalho"
    if not (df.loc[idx, "H_Entrada"] or ""):
        df.loc[idx, "H_Entrada"] = ent_padrao
    df.loc[idx, "H_Saida"] = _saida_from_entrada(df.loc[idx, "H_Entrada"])

def _set_folga(df, idx, locked_status: set[int] | None = None):
    if _locked(locked_status, idx):
        return
    df.loc[idx, "Status"] = "Folga"
    df.loc[idx, "H_Entrada"] = ""
    df.loc[idx, "H_Saida"] = ""

def _set_ferias(df, idx, locked_status: set[int] | None = None):
    if _locked(locked_status, idx):
        return
    df.loc[idx, "Status"] = "Férias"
    df.loc[idx, "H_Entrada"] = ""
    df.loc[idx, "H_Saida"] = ""

def _set_balanco(df, idx, locked_status: set[int] | None = None):
    if _locked(locked_status, idx):
        return
    df.loc[idx, "Status"] = BALANCO_STATUS
    df.loc[idx, "H_Entrada"] = BALANCO_DIA_ENTRADA
    df.loc[idx, "H_Saida"] = BALANCO_DIA_SAIDA

def _semana_seg_dom_indices(datas: pd.DatetimeIndex, idx_any: int):
    """Retorna índices da semana SEG->DOM do item idx_any.
    Robustez: ignora NaT para evitar TypeError em comparações.
    """
    d = datas[idx_any]
    if pd.isna(d):
        return []
    monday = d - timedelta(days=int(d.weekday()))
    sunday = monday + timedelta(days=6)

    out = []
    for i, dd in enumerate(datas):
        if pd.isna(dd):
            continue
        if monday.date() <= dd.date() <= sunday.date():
            out.append(i)
    return out

def _all_weeks_seg_dom(datas: pd.DatetimeIndex):
    weeks, seen = [], set()
    for i in range(len(datas)):
        w = tuple(_semana_seg_dom_indices(datas, i))
        if w and w not in seen:
            seen.add(w)
            weeks.append(list(w))
    return weeks

def _sunday_indices_df(df: pd.DataFrame) -> set[int]:
    out = set()
    if df is None or len(df) == 0 or "Data" not in df.columns:
        return out
    for i in df.index.tolist():
        try:
            if pd.to_datetime(df.loc[i, "Data"]).day_name() == "Sunday":
                out.add(i)
        except Exception:
            pass
    return out

def _merge_locked_status(*parts) -> set[int]:
    merged = set()
    for part in parts:
        if not part:
            continue
        try:
            merged.update(int(x) for x in part)
        except Exception:
            try:
                for x in list(part):
                    try:
                        merged.add(int(x))
                    except Exception:
                        pass
            except Exception:
                pass
    return merged

# =========================================================
# ✅ DOMINGO 1x1 POR COLABORADOR (GLOBAL, RESPEITA LOCK/FÉRIAS)
# =========================================================
def enforce_sundays_1x1_for_employee(
    df: pd.DataFrame,
    ent_padrao: str,
    locked_status: set[int] | None = None,
    base_first: str | None = None
):
    domingos = [i for i in range(len(df)) if df.loc[i, "Data"].day_name() == "Sunday"]
    if not domingos:
        return

    def _normalize_dom_status(i: int) -> str | None:
        stt = df.loc[i, "Status"]
        if stt == "Férias":
            return None
        if stt == "Folga":
            return "Folga"
        if stt in WORK_STATUSES:
            return "Trabalho"
        return None

    def _force_dom(i: int, val: str):
        if _locked(locked_status, i):
            return
        if df.loc[i, "Status"] == "Férias":
            return
        if val == "Folga":
            _set_folga(df, i, locked_status=locked_status)
        else:
            df.loc[i, "H_Entrada"] = ent_padrao
            _set_trabalho(df, i, ent_padrao, locked_status=locked_status)

    first_idx = domingos[0]
    if not _locked(locked_status, first_idx) and df.loc[first_idx, "Status"] != "Férias":
        if base_first in ("Trabalho", "Folga"):
            _force_dom(first_idx, base_first)

    cur = None
    for i in domingos:
        norm = _normalize_dom_status(i)
        if norm in ("Trabalho", "Folga"):
            cur = norm
            break
    if cur is None:
        return

    for i in domingos:
        if df.loc[i, "Status"] == "Férias":
            continue

        if _locked(locked_status, i):
            norm = _normalize_dom_status(i)
            if norm in ("Trabalho", "Folga"):
                cur = norm
            continue

        _force_dom(i, cur)
        cur = "Folga" if cur == "Trabalho" else "Trabalho"

# =========================================================
# ✅ DESCANSO GLOBAL 11:10 (corrigido para NÃO criar folga consecutiva)
# =========================================================
def enforce_global_rest_keep_targets(df: pd.DataFrame, ent_padrao: str, locked_status: set[int] | None = None, ultima_saida_prev: str | None = None):
    # mantém horário fixo de balanço
    for i in range(len(df)):
        if df.loc[i, "Status"] == BALANCO_STATUS:
            df.loc[i, "H_Entrada"] = BALANCO_DIA_ENTRADA
            df.loc[i, "H_Saida"] = BALANCO_DIA_SAIDA

    last_saida = (ultima_saida_prev or "").strip()

    for i in range(len(df)):
        stt = df.loc[i, "Status"]

        # 🔥 REGRA SUPREMA (MANUAL): se este dia está TRAVADO por override,
        # NENHUMA regra automática pode mexer (nem status, nem horários).
        if _locked(locked_status, i):
            if stt not in WORK_STATUSES:
                df.loc[i, "H_Entrada"] = ""
                df.loc[i, "H_Saida"] = ""
                last_saida = ""
            else:
                if stt == BALANCO_STATUS:
                    df.loc[i, "H_Entrada"] = BALANCO_DIA_ENTRADA
                    df.loc[i, "H_Saida"] = BALANCO_DIA_SAIDA
                else:
                    ent_fix = (df.loc[i, "H_Entrada"] or "").strip() or ent_padrao
                    df.loc[i, "H_Entrada"] = ent_fix
                    if not (df.loc[i, "H_Saida"] or ""):
                        df.loc[i, "H_Saida"] = _saida_from_entrada(ent_fix)
                last_saida = (df.loc[i, "H_Saida"] or "").strip()
            continue

        if stt not in WORK_STATUSES:
            df.loc[i, "H_Entrada"] = ""
            df.loc[i, "H_Saida"] = ""
            last_saida = ""
            continue

        if stt == BALANCO_STATUS:
            last_saida = (df.loc[i, "H_Saida"] or "").strip()
            continue

        target = (df.loc[i, "H_Entrada"] or "").strip() or ent_padrao

        if not last_saida:
            df.loc[i, "H_Entrada"] = target
            df.loc[i, "H_Saida"] = _saida_from_entrada(target)
            last_saida = df.loc[i, "H_Saida"]
            continue

        min_ent = _add_min(last_saida, INTERSTICIO_MIN)

        s = _to_min(last_saida)
        e_t = _to_min(target)
        e_min = _to_min(min_ent)
        if e_t <= s: e_t += 1440
        if e_min <= s: e_min += 1440

        if e_t >= e_min:
            df.loc[i, "H_Entrada"] = target
            df.loc[i, "H_Saida"] = _saida_from_entrada(target)
            last_saida = df.loc[i, "H_Saida"]
            continue

        prev = i - 1
        if prev >= 0:
            # tenta ajustar o dia anterior (saída mais cedo) sem virar folga
            if (
                df.loc[prev, "Status"] == "Trabalho"
                and not _locked(locked_status, prev)
            ):
                saida_req = _sub_min(target, INTERSTICIO_MIN)
                ent_req = _sub_min(saida_req, DURACAO_JORNADA)
                df.loc[prev, "H_Entrada"] = ent_req
                df.loc[prev, "H_Saida"] = _saida_from_entrada(ent_req)
                last_saida = df.loc[prev, "H_Saida"]

                df.loc[i, "H_Entrada"] = target
                df.loc[i, "H_Saida"] = _saida_from_entrada(target)
                last_saida = df.loc[i, "H_Saida"]
                continue

            # plano B: folgar o dia anterior SÓ se NÃO gerar folga consecutiva
            if prev >= 0 and not _locked(locked_status, prev) and df.loc[prev, "Status"] != "Férias":
                if _nao_consecutiva_folga(df, prev):
                    _set_folga(df, prev, locked_status=locked_status)
                    last_saida = ""
                    df.loc[i, "H_Entrada"] = target
                    df.loc[i, "H_Saida"] = _saida_from_entrada(target)
                    last_saida = df.loc[i, "H_Saida"]
                    continue
                else:
                    # alternativa: empurra o dia atual (não cria folga seguida)
                    ent_ok = _ajustar_para_intersticio(target, last_saida)
                    df.loc[i, "H_Entrada"] = ent_ok
                    df.loc[i, "H_Saida"] = _saida_from_entrada(ent_ok)
                    last_saida = df.loc[i, "H_Saida"]
                    continue

        # fallback final: empurra entrada
        ent_ok = _ajustar_para_intersticio(target, last_saida)
        df.loc[i, "H_Entrada"] = ent_ok
        df.loc[i, "H_Saida"] = _saida_from_entrada(ent_ok)
        last_saida = df.loc[i, "H_Saida"]

# =========================================================
# ✅ 5x2: máxima sequência de trabalho = 5
# =========================================================
def enforce_max_5_consecutive_work(df, ent_padrao, pode_folgar_sabado: bool, initial_consec: int = 0, locked_status: set[int] | None = None):
    # Segurança: garante índice 0..N-1 (evita KeyError por índice quebrado)
    df.reset_index(drop=True, inplace=True)

    def can_make_folga(i):
        # Segurança (evita iloc out-of-bounds)
        if i is None or i < 0 or i >= len(df):
            return False
        # Só converte TRABALHO normal em folga (não mexe em Balanço)
        if _locked(locked_status, i):
            return False
        if df.iloc[i]["Status"] != "Trabalho":
            return False
        dia = df.iloc[i]["Dia"]
        if dia == "dom":
            return False
        if dia == "sáb" and not pode_folgar_sabado:
            return False
        if not _nao_consecutiva_folga(df, i):
            return False
        return True

    consec, i = int(initial_consec), 0
    while i < len(df):
        if df.iloc[i]["Status"] in WORK_STATUSES:
            consec += 1
            if consec > 5:
                block_start = i - (consec - 1)
                block_end = i
                # Evita range negativo quando initial_consec vem do mês anterior
                block_start = max(0, int(block_start))
                block_end = min(len(df) - 1, int(block_end))
                candidatos = []
                for j in range(block_start, block_end + 1):
                    if can_make_folga(j):
                        dia = df.iloc[j]["Dia"]
                        weekday_prio = 0 if dia in ["seg", "ter", "qua", "qui", "sex"] else 1
                        mid = (block_start + block_end) / 2
                        dist = abs(j - mid)
                        candidatos.append((weekday_prio, dist, j))
                if candidatos:
                    candidatos.sort()
                    escolhido = candidatos[0][2]
                    _set_folga(df, escolhido, locked_status=locked_status)
                    consec = 0
                    i = max(0, escolhido - 2)
                    continue
                else:
                    consec = 0
        else:
            consec = 0
        i += 1

TARGET_FOLGAS_POR_SEMANA = 2  # 5x2: 2 folgas totais por semana (SEG->DOM). Domingo conta se for folga.

def _is_folga_status(stt: str) -> bool:
    s = str(stt or "").strip().lower()
    return s in ("folga", "f", "folg")

def _is_work_status(stt: str) -> bool:
    s = str(stt or "").strip().lower()
    if s in ("trabalho", "t", "work"):
        return True
    try:
        return str(stt) in WORK_STATUSES
    except Exception:
        return False

def _week_start_monday(d: pd.Timestamp) -> pd.Timestamp:
    d = pd.to_datetime(d)
    return (d - pd.to_timedelta(int(d.weekday()), unit="D")).normalize()

def enforce_weekly_folga_targets(df: pd.DataFrame, df_ref=None, pode_folgar_sabado=None, locked_status=None, **kwargs) -> pd.DataFrame:
    """
    REGRA SEMANAL (SEG->DOM) — SEMANA CONTÍNUA (corrige virada de mês):
      - Semana sempre tem 2 folgas no total (5x2).
      - Domingo conta como folga quando é Folga.
      - Se domingo = Folga => precisa completar para 2 no total (então SEG-SÁB tende a 1, mas pode variar na virada).
      - Se domingo = Trabalho => precisa completar para 2 no total (SEG-SÁB tende a 2).
    Importante (virada de mês):
      - Se a semana começou no mês anterior, usamos df_ref (se fornecido) para contar as folgas já existentes
        nos dias anteriores (SEG..dia anterior ao 1º do mês).
      - Ajustamos APENAS os dias do mês atual (não mexe no mês anterior).
    """
    if df is None or df.empty:
        return df
    if "Data" not in df.columns or "Chapa" not in df.columns or "Status" not in df.columns:
        return df

    df = df.copy()
    df["Data_dt"] = pd.to_datetime(df["Data"])
    df["week_start"] = df["Data_dt"].apply(_week_start_monday)
    df.sort_values(["Chapa", "Data_dt"], inplace=True)

    # Data mínima do mês corrente dentro do df (para saber onde começa o "mês atual")
    min_cur = df["Data_dt"].min().normalize()

    # df_ref (se vier) deve conter pelo menos Data, Chapa, Status
    ref_ok = isinstance(df_ref, pd.DataFrame) and {"Data", "Chapa", "Status"}.issubset(set(df_ref.columns))
    if ref_ok:
        ref = df_ref.copy()
        ref["Data_dt"] = pd.to_datetime(ref["Data"])
    else:
        ref = None

    # Permissão de sábado: pode ser bool global ou dict por chapa; ou coluna do df
    sab_col = None
    for c in ("Pode_Folgar_Sabado", "pode_folgar_sabado", "folga_sabado", "Folga_Sabado", "pode_sabado"):
        if c in df.columns:
            sab_col = c
            break

    def _allow_sab(idx_row) -> bool:
        if pode_folgar_sabado is not None:
            try:
                if isinstance(pode_folgar_sabado, bool):
                    return pode_folgar_sabado
                chv = str(df.at[idx_row, "Chapa"])
                if isinstance(pode_folgar_sabado, dict) and chv in pode_folgar_sabado:
                    return bool(pode_folgar_sabado[chv])
            except Exception:
                pass
        if sab_col is None:
            return True
        try:
            v = df.at[idx_row, sab_col]
            return bool(int(v)) if str(v).strip() != "" else bool(v)
        except Exception:
            try:
                return bool(df.at[idx_row, sab_col])
            except Exception:
                return True

    # Lock (travado): pode vir como set/list de índices, Series/DataFrame booleana, ou coluna no df
    lock_col = None
    for c in ("Travado_Status", "travado_status", "Lock_Status", "lock_status", "Status_Travado", "status_travado"):
        if c in df.columns:
            lock_col = c
            break

    def _is_locked(idx_row) -> bool:
        if locked_status is not None:
            try:
                if isinstance(locked_status, (set, list, tuple)) and idx_row in locked_status:
                    return True
                if hasattr(locked_status, "get") and locked_status.get(idx_row, False):
                    return True
            except Exception:
                pass
        if lock_col is None:
            return False
        try:
            return bool(df.at[idx_row, lock_col])
        except Exception:
            return False

    # Helper: status combinado (df tem prioridade, senão df_ref)
    def _status_comb(chapa: str, day: pd.Timestamp) -> str:
        day_d = day.normalize()
        # procura no df (mês atual)
        rows = df[(df["Chapa"].astype(str) == str(chapa)) & (df["Data_dt"].dt.normalize() == day_d)]
        if not rows.empty:
            return rows.iloc[0]["Status"]
        # procura no df_ref (mês anterior / contexto), se existir
        if ref is not None:
            r2 = ref[(ref["Chapa"].astype(str) == str(chapa)) & (ref["Data_dt"].dt.normalize() == day_d)]
            if not r2.empty:
                return r2.iloc[0]["Status"]
        return ""

    # Para cada colaborador, para cada semana que intersecta o mês atual:
    for ch in df["Chapa"].astype(str).unique():
        sub = df[df["Chapa"].astype(str) == str(ch)]
        for ws in sub["week_start"].unique():
            ws = pd.to_datetime(ws).normalize()
            we = ws + pd.Timedelta(days=6)

            # domingo da semana
            sunday = we
            dom_status = _status_comb(ch, sunday)

            # V74: quando o domingo da semana cai no mês seguinte e ainda não existe no df atual,
            # ele MESMO ASSIM precisa contar na meta semanal da virada.
            # Regra: o domingo futuro segue a alternância do último domingo real anterior.
            if not str(dom_status or '').strip():
                prev_sundays = []
                for back in range(7, 70, 7):
                    prev_day = sunday - pd.Timedelta(days=back)
                    st_prev = _status_comb(ch, prev_day)
                    if str(st_prev or '').strip() in ('Trabalho', 'Folga'):
                        prev_sundays.append(str(st_prev or '').strip())
                        break
                if prev_sundays:
                    dom_status = 'Folga' if prev_sundays[0] == 'Trabalho' else 'Trabalho'

            dom_folga = 1 if _is_folga_status(dom_status) else 0

            # Conta folgas SEG-SÁB na parte "antes do mês" (se semana começou antes de min_cur)
            folgas_prev_seg_sab = 0
            if ws < min_cur:
                for d in range(0, 6):  # SEG..SÁB
                    day = ws + pd.Timedelta(days=d)
                    if day < min_cur:
                        if _is_folga_status(_status_comb(ch, day)):
                            folgas_prev_seg_sab += 1

            # Target de folgas total = 2
            # Folgas que precisam existir em SEG-SÁB no mês atual para fechar 2:
            needed_seg_sab_total = max(0, int(TARGET_FOLGAS_POR_SEMANA) - dom_folga)
            needed_seg_sab_current = max(0, needed_seg_sab_total - folgas_prev_seg_sab)

            # Linhas do df no intervalo SEG-SÁB desta semana (somente mês atual)
            wk_mask = (df["Chapa"].astype(str) == str(ch)) & (df["Data_dt"] >= ws) & (df["Data_dt"] <= we)
            wk = df[wk_mask].copy()
            seg_sab = wk[wk["Data_dt"].dt.weekday <= 5].copy()

            if seg_sab.empty:
                continue

            seg_sab = seg_sab[seg_sab["Data_dt"] >= min_cur]  # só ajusta no mês atual
            if seg_sab.empty:
                continue

            seg_sab_folgas = seg_sab["Status"].apply(_is_folga_status)
            folgas_seg_sab_current = int(seg_sab_folgas.sum())

            # candidatos (respeita lock e sábado permitido)
            cand_folga = [i for i in seg_sab[seg_sab_folgas].index.tolist()
                          if (not _is_locked(i)) and (_allow_sab(i) or df.at[i, "Data_dt"].weekday() != 5)]
            cand_trab = [i for i in seg_sab[~seg_sab_folgas].index.tolist()
                         if (not _is_locked(i)) and (_allow_sab(i) or df.at[i, "Data_dt"].weekday() != 5)]

            # Se excesso no mês atual (considerando o que já veio do mês anterior), remove
            if folgas_seg_sab_current > needed_seg_sab_current:
                excesso = folgas_seg_sab_current - needed_seg_sab_current
                # remove do fim da semana primeiro
                cand_sorted = sorted(cand_folga, key=lambda i: df.at[i, "Data_dt"], reverse=True)
                for i in cand_sorted[:excesso]:
                    df.at[i, "Status"] = "Trabalho"

            # Se falta no mês atual, adiciona folga
            elif folgas_seg_sab_current < needed_seg_sab_current:
                falta = needed_seg_sab_current - folgas_seg_sab_current

                # heurística: preferir quarta/quinta, depois terça/sexta, depois segunda/sábado
                def _prio(i):
                    wd = int(df.at[i, "Data_dt"].weekday())
                    return {2:0, 3:0, 1:1, 4:1, 0:2, 5:3}.get(wd, 9)

                cand_sorted = sorted(cand_trab, key=_prio)
                for i in cand_sorted[:falta]:
                    df.at[i, "Status"] = "Folga"

    return df

def _cap_total_folgas_por_semana(df: pd.DataFrame, target_total: int = 2, locked_status=None, df_ref=None) -> pd.DataFrame:
    """
    Segurança extra (SEMANA CONTÍNUA): garante que cada semana (SEG->DOM) tenha no máximo `target_total` folgas.
    - Considera folgas que já existam no mês anterior via df_ref (se fornecido).
    - Remove excesso APENAS no mês atual (df), preferindo SEG-SÁB e respeitando lock/travado.
    - NÃO mexe em 'Férias' (elas não contam como Folga no seu modelo se quiser tratar diferente; aqui só olha Status=Folga/F).
    """
    if df is None or df.empty:
        return df
    if "Data" not in df.columns or "Chapa" not in df.columns or "Status" not in df.columns:
        return df

    df = df.copy()
    df["Data_dt"] = pd.to_datetime(df["Data"])
    df["week_start"] = df["Data_dt"].apply(_week_start_monday)
    df.sort_values(["Chapa", "Data_dt"], inplace=True)

    min_cur = df["Data_dt"].min().normalize()

    ref_ok = isinstance(df_ref, pd.DataFrame) and {"Data", "Chapa", "Status"}.issubset(set(df_ref.columns))
    if ref_ok:
        ref = df_ref.copy()
        ref["Data_dt"] = pd.to_datetime(ref["Data"])
    else:
        ref = None

    lock_col = None
    for c in ("Travado_Status", "travado_status", "Lock_Status", "lock_status", "Status_Travado", "status_travado"):
        if c in df.columns:
            lock_col = c
            break

    def _is_locked(i):
        if locked_status is not None:
            try:
                if isinstance(locked_status, (set, list, tuple)) and i in locked_status:
                    return True
                if hasattr(locked_status, "get") and locked_status.get(i, False):
                    return True
            except Exception:
                pass
        if lock_col is None:
            return False
        try:
            return bool(df.at[i, lock_col])
        except Exception:
            return False

    def _status_comb(chapa: str, day: pd.Timestamp) -> str:
        day_d = pd.to_datetime(day).normalize()
        rows = df[(df["Chapa"].astype(str) == str(chapa)) & (df["Data_dt"].dt.normalize() == day_d)]
        if not rows.empty:
            return rows.iloc[0]["Status"]
        if ref is not None:
            r2 = ref[(ref["Chapa"].astype(str) == str(chapa)) & (ref["Data_dt"].dt.normalize() == day_d)]
            if not r2.empty:
                return r2.iloc[0]["Status"]
        return ""

    for ch in df["Chapa"].astype(str).unique():
        sub = df[df["Chapa"].astype(str) == str(ch)]
        for ws in sub["week_start"].unique():
            ws = pd.to_datetime(ws).normalize()
            we = ws + pd.Timedelta(days=6)

            # conta folgas total na semana (combinado)
            folgas_total = 0
            for d in range(0, 7):
                day = ws + pd.Timedelta(days=d)
                st_day = _status_comb(ch, day)

                # V74: se o domingo ainda está fora do mês atual, conta pela alternância esperada
                if (not str(st_day or '').strip()) and int(day.weekday()) == 6:
                    for back in range(7, 70, 7):
                        prev_day = day - pd.Timedelta(days=back)
                        st_prev = _status_comb(ch, prev_day)
                        if str(st_prev or '').strip() in ('Trabalho', 'Folga'):
                            st_day = 'Folga' if st_prev == 'Trabalho' else 'Trabalho'
                            break

                if _is_folga_status(st_day):
                    folgas_total += 1

            if folgas_total <= target_total:
                continue

            excesso = folgas_total - target_total

            # remove somente no mês atual: candidatos SEG-SÁB (dias >= min_cur)
            wk_mask = (df["Chapa"].astype(str) == str(ch)) & (df["Data_dt"] >= ws) & (df["Data_dt"] <= we)
            wk = df[wk_mask].copy()
            seg_sab = wk[(wk["Data_dt"].dt.weekday <= 5) & (wk["Data_dt"] >= min_cur)]
            cand = [i for i in seg_sab[seg_sab["Status"].apply(_is_folga_status)].index.tolist() if not _is_locked(i)]

            # remove do fim primeiro
            cand_sorted = sorted(cand, key=lambda i: df.at[i, "Data_dt"], reverse=True)
            for i in cand_sorted[:excesso]:
                df.at[i, "Status"] = "Trabalho"

    return df



def enforce_no_consecutive_folgas(df: pd.DataFrame, locked_status=None) -> pd.DataFrame:
    """
    Impede folgas consecutivas (ex.: DOM+SEG) criadas automaticamente.
    Respeita travas (locked_status ou coluna de travado no DF).
    """
    if df is None or df.empty:
        return df
    if "Data" not in df.columns or "Chapa" not in df.columns or "Status" not in df.columns:
        return df

    df = df.copy()
    df["Data_dt"] = pd.to_datetime(df["Data"])
    df.sort_values(["Chapa", "Data_dt"], inplace=True)

    lock_col = None
    for c in ("Travado_Status", "travado_status", "Lock_Status", "lock_status", "Status_Travado", "status_travado"):
        if c in df.columns:
            lock_col = c
            break

    def _is_locked(i):
        if locked_status is not None:
            try:
                if isinstance(locked_status, (set, list, tuple)) and i in locked_status:
                    return True
                if hasattr(locked_status, "get") and locked_status.get(i, False):
                    return True
            except Exception:
                pass
        if lock_col is None:
            return False
        try:
            return bool(df.at[i, lock_col])
        except Exception:
            return False

    for ch in df["Chapa"].astype(str).unique():
        sub = df[df["Chapa"].astype(str) == str(ch)].sort_values("Data_dt")
        idxs = sub.index.tolist()
        for a, b in zip(idxs, idxs[1:]):
            if _is_folga_status(df.at[a, "Status"]) and _is_folga_status(df.at[b, "Status"]):
                if not _is_locked(b):
                    df.at[b, "Status"] = "Trabalho"
    return df

def strict_weekly_5x2_never_break(df: pd.DataFrame, chapa: str, ent_padrao: str, pode_folgar_sabado: bool, locked_status=None, df_ref_prev: pd.DataFrame | None = None) -> pd.DataFrame:
    """
    REGRA INQUEBRÁVEL:
    - A unidade da escala é a semana real SEG->DOM
    - Cada semana deve fechar com EXATAMENTE 2 folgas no total
    - Se domingo = Folga => SEG-SÁB devem ter EXATAMENTE 1 folga
    - Se domingo = Trabalho => SEG-SÁB devem ter EXATAMENTE 2 folgas
    - Nunca reinicia pela virada do mês
    - Domingo não é alterado aqui
    - Sábado só entra se o colaborador puder folgar sábado
    """
    if df is None or df.empty:
        return df

    df = df.copy().reset_index(drop=True)
    df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
    df = df.sort_values("Data").reset_index(drop=True)
    min_cur = pd.to_datetime(df["Data"]).min().normalize()

    ref_prev = None
    if isinstance(df_ref_prev, pd.DataFrame) and (not df_ref_prev.empty) and {"Data", "Chapa", "Status"}.issubset(set(df_ref_prev.columns)):
        ref_prev = df_ref_prev.copy()
        ref_prev["Data"] = pd.to_datetime(ref_prev["Data"], errors="coerce")
        ref_prev = ref_prev[ref_prev["Chapa"].astype(str) == str(chapa)].copy()

    def _locked_local(i: int) -> bool:
        return _locked(locked_status, i)

    def _status_on(day_ts: pd.Timestamp) -> str:
        day_ts = pd.to_datetime(day_ts).normalize()
        rows = df[df["Data"].dt.normalize() == day_ts]
        if not rows.empty:
            return str(rows.iloc[0]["Status"] or "")
        if ref_prev is not None:
            rows2 = ref_prev[ref_prev["Data"].dt.normalize() == day_ts]
            if not rows2.empty:
                return str(rows2.iloc[0]["Status"] or "")
        return ""

    def _infer_sunday_status(sunday_ts: pd.Timestamp) -> str:
        st = _status_on(sunday_ts)
        if st in ("Folga", "Trabalho"):
            return st
        # procura domingo real anterior e alterna 1x1
        probe = pd.to_datetime(sunday_ts).normalize() - pd.Timedelta(days=7)
        for _ in range(12):
            st_prev = _status_on(probe)
            if st_prev in ("Folga", "Trabalho"):
                return "Trabalho" if st_prev == "Folga" else "Folga"
            probe -= pd.Timedelta(days=7)
        return "Trabalho"

    week_starts = sorted({(_week_start_monday(pd.to_datetime(d).normalize())) for d in df["Data"].dropna().tolist()})

    for ws in week_starts:
        ws = pd.to_datetime(ws).normalize()
        sunday_ts = ws + pd.Timedelta(days=6)
        sunday_status = _infer_sunday_status(sunday_ts)
        target_seg_sab = 1 if sunday_status == "Folga" else 2

        seg_sab_days = [ws + pd.Timedelta(days=k) for k in range(6)]

        def _current_seg_sab_idxs():
            idxs = []
            for i in range(len(df)):
                dd = pd.to_datetime(df.loc[i, "Data"]).normalize()
                if dd in seg_sab_days:
                    idxs.append(i)
            return idxs

        current_idxs = _current_seg_sab_idxs()

        prev_count = 0
        if ref_prev is not None:
            for day in seg_sab_days:
                if day < min_cur and _is_folga_status(_status_on(day)):
                    prev_count += 1

        current_count = sum(1 for i in current_idxs if _is_folga_status(df.loc[i, "Status"]))
        total_seg_sab = prev_count + current_count

        if total_seg_sab > target_seg_sab:
            excesso = total_seg_sab - target_seg_sab
            cand_remove = []
            for i in current_idxs:
                dia_nome = D_PT[pd.to_datetime(df.loc[i, "Data"]).day_name()]
                if dia_nome == "sáb" and not bool(pode_folgar_sabado):
                    continue
                if not _is_folga_status(df.loc[i, "Status"]):
                    continue
                if _locked_local(i):
                    continue
                cand_remove.append(i)
            cand_remove = sorted(cand_remove, key=lambda i: pd.to_datetime(df.loc[i, "Data"]), reverse=True)
            for i in cand_remove:
                if excesso <= 0:
                    break
                _set_trabalho(df, i, ent_padrao, locked_status=locked_status)
                excesso -= 1

        prev_count = 0
        if ref_prev is not None:
            for day in seg_sab_days:
                if day < min_cur and _is_folga_status(_status_on(day)):
                    prev_count += 1
        current_idxs = _current_seg_sab_idxs()
        current_count = sum(1 for i in current_idxs if _is_folga_status(df.loc[i, "Status"]))
        total_seg_sab = prev_count + current_count

        if total_seg_sab < target_seg_sab:
            falta = target_seg_sab - total_seg_sab
            cand_add = []
            for i in current_idxs:
                dia_nome = D_PT[pd.to_datetime(df.loc[i, "Data"]).day_name()]
                if dia_nome == "sáb" and not bool(pode_folgar_sabado):
                    continue
                if _locked_local(i):
                    continue
                if str(df.loc[i, "Status"]) != "Trabalho":
                    continue
                if not _nao_consecutiva_folga(df, i):
                    continue
                cand_add.append(i)
            cand_add = sorted(cand_add, key=lambda i: (1 if D_PT[pd.to_datetime(df.loc[i, "Data"]).day_name()] == "sáb" else 0, pd.to_datetime(df.loc[i, "Data"])))
            for i in cand_add:
                if falta <= 0:
                    break
                _set_folga(df, i, locked_status=locked_status)
                falta -= 1

    return df


def _counts_folgas_day_and_hour(hist_by_chapa: dict, colab_by_chapa: dict, chapas_grupo: list, idxs_semana: list, df_ref):
    counts_day = {i: 0 for i in idxs_semana}
    counts_day_hour = {}
    for ch in chapas_grupo:
        df = hist_by_chapa[ch]
        bucket = colab_by_chapa[ch].get("Entrada", "06:00")
        for i in idxs_semana:
            if df_ref.loc[i, "Dia"] == "dom":
                continue
            if df.loc[i, "Status"] == "Folga":
                counts_day[i] += 1
                counts_day_hour[(i, bucket)] = counts_day_hour.get((i, bucket), 0) + 1
    return counts_day, counts_day_hour

# =========================================================
# ✅ REBALANCE (corrigido): recebe estado_prev e respeita locked_idx
# =========================================================

def rebalance_folgas_dia(
    hist_by_chapa: dict,
    colab_by_chapa: dict,
    chapas_grupo: list,
    weeks: list,
    df_ref,
    estado_prev: dict | None = None,
    locked_idx: dict | None = None,
    past_flag: bool = False,
    max_iters=240
):
    """
    Rebalance do subgrupo por SEMANA REAL (seg->dom), com proteção anti-loop.

    # 🔒 REGRA: se o usuário alterou "Folgas manuais em grade", NÃO rebalancear
    try:
        import streamlit as st
        if st.session_state.get("manual_folga_lock", False):
            return
    except Exception:
        pass

    Objetivo:
    - reduzir dias leves/zerados e dias muito carregados
    - domingo nunca entra no balanceamento
    - sábado só recebe folga se a regra do colaborador permitir
    - nunca mexe em célula travada por override
    - não cria folga dupla automática
    - não altera a quantidade de folgas do colaborador na semana (troca 1x1)
    - só aceita movimentos que melhoram a pontuação da semana
    """
    estado_prev = estado_prev or {}
    locked_idx = locked_idx or {}
    _past = bool(past_flag)

    def is_dom(i: int) -> bool:
        return str(df_ref.loc[i, "Dia"]) == "dom"

    def is_sab(i: int) -> bool:
        return str(df_ref.loc[i, "Dia"]) == "sáb"

    def is_locked(ch: str, i: int) -> bool:
        return bool(i in (locked_idx.get(ch, set()) or set()))

    def _eligible_days_for_group(week: list[int]) -> list[int]:
        return [i for i in week if not is_dom(i)]

    def _week_state_signature(week: list[int]) -> tuple:
        sig = []
        for ch in chapas_grupo:
            df = hist_by_chapa[ch]
            row = tuple("F" if str(df.loc[i, "Status"]) == "Folga" else "T" for i in week)
            sig.append((ch, row))
        return tuple(sig)

    def _counts(week: list[int], eligible_days: list[int]) -> dict:
        counts = {i: 0 for i in eligible_days}
        for ch in chapas_grupo:
            df = hist_by_chapa[ch]
            for i in eligible_days:
                if str(df.loc[i, "Status"]) == "Folga":
                    counts[i] += 1
        return counts

    def _score_counts(counts: dict) -> float:
        if not counts:
            return 0.0
        vals = list(counts.values())
        total = sum(vals)
        n = max(1, len(vals))
        avg = total / float(n)
        low_target = math.floor(avg)
        high_target = math.ceil(avg)
        spread = max(vals) - min(vals)

        zeros = sum(1 for v in vals if v == 0)
        ones = sum(1 for v in vals if v == 1)
        very_heavy = sum(1 for v in vals if v >= high_target + 2)
        above_target = sum(max(0, v - high_target) for v in vals)
        below_target = sum(max(0, low_target - v) for v in vals)

        # penaliza forte dias zerados e sobrecarga alta; prefere ficar perto da meta da semana
        return (
            sum((v - avg) ** 2 for v in vals)
            + (spread ** 2) * 10.0
            + zeros * 90.0
            + ones * 10.0
            + very_heavy * 80.0
            + above_target * 40.0
            + below_target * 30.0
        )

    def _candidate_order(counts: dict) -> tuple[list[int], list[int]]:
        heavy = sorted(counts.keys(), key=lambda i: (counts[i], i), reverse=True)
        light = sorted(counts.keys(), key=lambda i: (counts[i], i))
        return heavy, light

    def can_swap(ch: str, i_from: int, i_to: int) -> bool:
        df = hist_by_chapa[ch]
        pode_sab = bool(colab_by_chapa[ch].get("Folga_Sab", False))

        if i_from == i_to:
            return False
        if is_dom(i_from) or is_dom(i_to):
            return False
        if is_locked(ch, i_from) or is_locked(ch, i_to):
            return False

        st_from = str(df.loc[i_from, "Status"])
        st_to = str(df.loc[i_to, "Status"])

        if st_from in {"Férias", "Afastamento"} or st_to in {"Férias", "Afastamento"}:
            return False
        if st_from != "Folga" or st_to != "Trabalho":
            return False
        if is_sab(i_to) and not pode_sab:
            return False

        # não permite criar folga dupla automática no destino
        if (i_to > 0 and str(df.loc[i_to - 1, "Status"]) == "Folga") or (i_to < len(df) - 1 and str(df.loc[i_to + 1, "Status"]) == "Folga"):
            return False

        # simulação conservadora para não quebrar regras duras do colaborador
        try:
            sim = df.copy()
            ent = colab_by_chapa[ch].get("Entrada", "06:00")
            _set_trabalho(sim, i_from, ent, locked_status=locked_idx.get(ch, set()))
            _set_folga(sim, i_to, locked_status=locked_idx.get(ch, set()))
            sim = strict_weekly_5x2_never_break(
                sim,
                ch,
                ent,
                bool(colab_by_chapa[ch].get("Folga_Sab", False)),
                locked_status=locked_idx.get(ch, set()),
                df_ref=df_ref
            )
            # swap aceito apenas se preservou os pontos desejados
            return str(sim.loc[i_from, "Status"]) == "Trabalho" and str(sim.loc[i_to, "Status"]) == "Folga"
        except Exception:
            return False

    def do_swap(ch: str, i_from: int, i_to: int) -> None:
        df = hist_by_chapa[ch]
        ent = colab_by_chapa[ch].get("Entrada", "06:00")
        pode_sab = bool(colab_by_chapa[ch].get("Folga_Sab", False))

        _set_trabalho(df, i_from, ent, locked_status=locked_idx.get(ch, set()))
        _set_folga(df, i_to, locked_status=locked_idx.get(ch, set()))

        # revalida regras duras no colaborador após a troca
        df = strict_weekly_5x2_never_break(
            df,
            ch,
            ent,
            pode_sab,
            locked_status=locked_idx.get(ch, set()),
            df_ref=df_ref
        )
        enforce_max_5_consecutive_work(
            df, ent, pode_sab,
            initial_consec=(0 if _past else int((estado_prev.get(ch, {}) or {}).get("consec_trab_final", 0))),
            locked_status=locked_idx.get(ch, set())
        )
        hist_by_chapa[ch] = df

    for week in weeks:
        eligible_days = _eligible_days_for_group(week)
        if len(eligible_days) <= 1:
            continue

        seen_states = set()
        no_improve_rounds = 0
        iter_count = 0

        while iter_count < int(max_iters):
            iter_count += 1
            sig = _week_state_signature(week)
            if sig in seen_states:
                break
            seen_states.add(sig)

            counts = _counts(week, eligible_days)
            current_score = _score_counts(counts)
            heavy_days, light_days = _candidate_order(counts)

            if not heavy_days or not light_days:
                break

            spread = counts[heavy_days[0]] - counts[light_days[0]]
            # se ainda existir dia zerado ou dia bem mais carregado, continua tentando
            if spread <= 1 and min(counts.values()) > 0:
                break

            best_move = None
            best_score = current_score

            # prioriza tirar do mais pesado e jogar no mais leve
            for i_to in light_days:
                for i_from in heavy_days:
                    if i_from == i_to:
                        continue
                    if counts[i_from] <= counts[i_to]:
                        continue
                    # tenta priorizar trocas que aliviam de verdade a diferença do dia
                    if (counts[i_from] - counts[i_to]) < 1:
                        continue

                    # candidatos do dia pesado para o dia leve
                    for ch in chapas_grupo:
                        df = hist_by_chapa[ch]
                        if str(df.loc[i_from, "Status"]) != "Folga":
                            continue
                        if str(df.loc[i_to, "Status"]) != "Trabalho":
                            continue
                        if not can_swap(ch, i_from, i_to):
                            continue

                        sim_counts = dict(counts)
                        sim_counts[i_from] -= 1
                        sim_counts[i_to] += 1
                        sim_score = _score_counts(sim_counts)

                        if sim_score + 1e-9 < best_score:
                            best_score = sim_score
                            best_move = (ch, i_from, i_to)

                # se já encontrou uma boa troca para o dia mais leve atual, usa
                if best_move is not None and counts[i_to] == min(counts.values()):
                    break

            if not best_move:
                no_improve_rounds += 1
                if no_improve_rounds >= 2:
                    break
                continue

            do_swap(*best_move)
            no_improve_rounds = 0

# =========================================================
# GERAR ESCALA — POR SUBGRUPO
# =========================================================
def gerar_escala_setor_por_subgrupo(setor: str, colaboradores: list[dict], ano: int, mes: int, respeitar_ajustes: bool = True, df_ref_prev: 'pd.DataFrame|None' = None):
    datas = _dias_mes(ano, mes)
    weeks = _all_weeks_seg_dom(datas)
    df_ref_cur = pd.DataFrame({'Data': datas, 'Dia': [D_PT[d.day_name()] for d in datas]})
    df_ref_use = df_ref_prev if isinstance(df_ref_prev, pd.DataFrame) and (not df_ref_prev.empty) else None
    # Meses passados: não aplicar continuidade/travamentos do mês anterior.
    _past = is_past_competencia(ano, mes)
    estado_prev = {} if _past else load_estado_prev(setor, ano, mes)

    ovmap = _ov_map(setor, int(ano), int(mes)) if respeitar_ajustes else {}
    if respeitar_ajustes:
        ovmap = _merge_folga_fixa_into_ovmap(setor, int(ano), int(mes), ovmap)

    grupos = {}
    for c in colaboradores:
        sg = (c.get("Subgrupo") or "").strip() or "SEM SUBGRUPO"
        grupos.setdefault(sg, []).append(c)

    regras_cache = {}
    for sg in grupos.keys():
        if sg == "SEM SUBGRUPO":
            regras_cache[sg] = {"seg": 0, "ter": 0, "qua": 0, "qui": 0, "sex": 0, "sáb": 0}
        else:
            regras_cache[sg] = get_subgrupo_regras(setor, sg)

    hist_all = {}
    colab_by_chapa = {c["Chapa"]: c for c in colaboradores}
    locked_idx = {}

    # base de cada colaborador
    for c in colaboradores:
        ch = c["Chapa"]
        df = df_ref_cur.copy()
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
        df["Status"] = "Trabalho"
        df["H_Entrada"] = ""
        df["H_Saida"] = ""

        # ✅ férias só via tabela ferias
        for i, d in enumerate(datas):
            if is_de_ferias(setor, ch, d.date()):
                df.loc[i, "Status"] = "Férias"
                df.loc[i, "H_Entrada"] = ""
                df.loc[i, "H_Saida"] = ""

        if respeitar_ajustes:
            _apply_overrides_to_df_inplace(df, setor, ch, ovmap)

        locked = set()
        if respeitar_ajustes:
            for i in range(len(df)):
                if _is_status_locked(ovmap, ch, pd.to_datetime(df.loc[i, "Data"])):
                    locked.add(i)
        locked_idx[ch] = locked
        hist_all[ch] = df

    # ✅ Domingo 1x1 por colaborador COM CONTINUIDADE ENTRE MESES
    for ch, df in hist_all.items():
        ent = colab_by_chapa[ch].get("Entrada", "06:00")
        locked = locked_idx.get(ch, set())

        if _past:
            base_first = None
        else:
            prev_dom = infer_ultimo_domingo_status_from_escala(setor, int(ano), int(mes), ch)
            if prev_dom not in ("Folga","Trabalho"):
                prev_dom = (estado_prev.get(ch, {}) or {}).get("ultimo_domingo_status", None)
            if prev_dom == "Folga":
                base_first = "Trabalho"
            elif prev_dom == "Trabalho":
                base_first = "Folga"
            else:
                base_first = None

        enforce_sundays_1x1_for_employee(df, ent, locked_status=locked, base_first=base_first)
        hist_all[ch] = df

    # =====================================================
    # ✅ REGRA ESPECIAL: RETORNO DE FÉRIAS (somente 1a semana)
    # =====================================================
    _apply_regra_retorno_ferias_primeiro_domingo(hist_all, colab_by_chapa, locked_idx, df_ref_cur, setor)

    # =====================================================
    # ✅ REGRA SEMANAL NOVA (SEG->DOM) DEPENDE DO DOMINGO
    # =====================================================
    for sg, membros in grupos.items():
        chapas = [m["Chapa"] for m in membros]
        if not chapas:
            continue

        pref = regras_cache.get(sg, {"seg": 0, "ter": 0, "qua": 0, "qui": 0, "sex": 0, "sáb": 0})

        for week in weeks:
            idxs_week = sorted(week, key=lambda i: df_ref_cur.loc[i, "Data"])
            domingos = [i for i in idxs_week if df_ref_cur.loc[i, "Dia"] == "dom"]
            dom_idx = domingos[0] if domingos else None

            for ch in chapas:
                df = hist_all[ch]
                locked = locked_idx.get(ch, set())
                pode_sab = bool(colab_by_chapa[ch].get("Folga_Sab", False))
                ent_bucket = colab_by_chapa[ch].get("Entrada", "06:00")

                segunda_idx = idxs_week[0]
                segunda_date = df_ref_cur.loc[segunda_idx, "Data"].date()
                if is_first_week_after_return(setor, ch, segunda_date):
                    continue

                # candidatos seg-sex e sábado só se permitido
                cand_days = []
                for i in idxs_week:
                    dia = df_ref_cur.loc[i, "Dia"]
                    if dia == "dom":
                        continue
                    if dia == "sáb" and not pode_sab:
                        continue
                    cand_days.append(i)

                if dom_idx is None:
                    target_folgas = 2
                else:
                    dom_status = df.loc[dom_idx, "Status"]
                    target_folgas = 1 if dom_status == "Folga" else 2

                folgas_sem = int((df.loc[cand_days, "Status"] == "Folga").sum()) if cand_days else 0

                while folgas_sem < target_folgas:
                    counts_day, counts_day_hour = _counts_folgas_day_and_hour(hist_all, colab_by_chapa, chapas, cand_days, df_ref_cur)

                    possiveis = []
                    for j in cand_days:
                        if j in locked:
                            continue
                        dia = df_ref_cur.loc[j, "Dia"]
                        if df.loc[j, "Status"] != "Trabalho":
                            continue
                        if dia == "sáb" and not pode_sab:
                            continue
                        if not _nao_consecutiva_folga(df, j):
                            continue
                        possiveis.append(j)

                    if not possiveis:
                        break

                    random.shuffle(possiveis)

                    def score(j):
                        dia = df_ref_cur.loc[j, "Dia"]
                        weekday_prio = 0 if dia in ["seg", "ter", "qua", "qui", "sex"] else 1
                        pref_pen = PREF_EVITAR_PENALTY if pref.get(dia, 0) == 1 else 0
                        return (
                            counts_day.get(j, 0),
                            counts_day_hour.get((j, ent_bucket), 0),
                            pref_pen,
                            weekday_prio,
                            random.random()
                        )

                    possiveis.sort(key=score)
                    pick = possiveis[0]
                    _set_folga(df, pick, locked_status=locked)
                    folgas_sem += 1
                    hist_all[ch] = df

    # Pós: aplica regras globais por colaborador
    for ch, df in hist_all.items():
        ent = colab_by_chapa[ch].get("Entrada", "06:00")
        locked = locked_idx.get(ch, set())
        pode_sab = bool(colab_by_chapa[ch].get("Folga_Sab", False))

        if _past:
            base_first = None
        else:
            prev_dom = infer_ultimo_domingo_status_from_escala(setor, int(ano), int(mes), ch)
            if prev_dom not in ("Folga","Trabalho"):
                prev_dom = (estado_prev.get(ch, {}) or {}).get("ultimo_domingo_status", None)
            if prev_dom == "Folga":
                base_first = "Trabalho"
            elif prev_dom == "Trabalho":
                base_first = "Folga"
            else:
                base_first = None
        enforce_sundays_1x1_for_employee(df, ent, locked_status=locked, base_first=base_first)
        locked_dom = _merge_locked_status(locked, _sunday_indices_df(df))

        # 1) Garante 5 dias seguidos antes de mexer em metas semanais
        enforce_max_5_consecutive_work(
            df, ent, pode_sab,
            initial_consec=(0 if _past else int((estado_prev.get(ch, {}) or {}).get('consec_trab_final', 0))),
            locked_status=locked_dom
        )
        enforce_no_consecutive_folga(df, locked_status=locked_dom)

        # 2) Metas semanais podem REMOVER folga => pode criar >5 de novo
        enforce_weekly_folga_targets(df, df_ref=(df_ref_use if df_ref_use is not None else df_ref_cur), pode_folgar_sabado=pode_sab, locked_status=locked_dom)

        # 3) Reforça novamente o limite de 5 depois das metas semanais
        enforce_max_5_consecutive_work(
            df, ent, pode_sab,
            initial_consec=(0 if _past else int((estado_prev.get(ch, {}) or {}).get('consec_trab_final', 0))),
            locked_status=locked_dom
        )
        enforce_no_consecutive_folga(df, locked_status=locked_dom)

        ultima_saida_prev = "" if _past else (estado_prev.get(ch, {}).get("ultima_saida", "") or "")
        enforce_global_rest_keep_targets(df, ent, locked_status=locked_dom, ultima_saida_prev=ultima_saida_prev)

        # limpeza
        enforce_no_consecutive_folga(df, locked_status=locked_dom)
        enforce_global_rest_keep_targets(df, ent, locked_status=locked_dom, ultima_saida_prev=ultima_saida_prev)

        if respeitar_ajustes:
            _apply_overrides_to_df_inplace(df, setor, ch, ovmap)

        hist_all[ch] = df

    # rebalance por grupo (com estado_prev e travas)
    for sg, membros in grupos.items():
        chapas = [m["Chapa"] for m in membros]
        if chapas:
            rebalance_folgas_dia(
                hist_all, colab_by_chapa, chapas, weeks, df_ref_cur,
                estado_prev=estado_prev,
                locked_idx=locked_idx,
                past_flag=_past,
                max_iters=2200
            )

    # ✅ Pós-rebalance: revalida regras por colaborador (evita semana com 3 folgas)
    for ch, df in hist_all.items():
        ent = colab_by_chapa[ch].get("Entrada", "06:00")
        locked = locked_idx.get(ch, set())
        pode_sab = bool(colab_by_chapa[ch].get("Folga_Sab", False))
        locked_dom = _merge_locked_status(locked, _sunday_indices_df(df))

        # 1) Limite 5 dias (pode ter sido quebrado pelo rebalance)
        enforce_max_5_consecutive_work(
            df, ent, pode_sab,
            initial_consec=(0 if _past else int((estado_prev.get(ch, {}) or {}).get('consec_trab_final', 0))),
            locked_status=locked_dom
        )

        # 2) Regra semanal SEG→DOM (remove excesso e completa falta)
        enforce_weekly_folga_targets(df, df_ref=(df_ref_use if df_ref_use is not None else df_ref_cur), pode_folgar_sabado=pode_sab, locked_status=locked_dom)

        # 3) Reforça 5 dias novamente (regra semanal pode remover folga e criar sequência >5)
        enforce_max_5_consecutive_work(
            df, ent, pode_sab,
            initial_consec=(0 if _past else int((estado_prev.get(ch, {}) or {}).get('consec_trab_final', 0))),
            locked_status=locked_dom
        )

        # 4) Regra semanal novamente (se o max_5 criou folga extra, normaliza para alvo)
        enforce_weekly_folga_targets(df, df_ref=(df_ref_use if df_ref_use is not None else df_ref_cur), pode_folgar_sabado=pode_sab, locked_status=locked_dom)

        # 5) Proíbe folga consecutiva automática (DOM+SEG etc.)
        enforce_no_consecutive_folga(df, locked_status=locked_dom)

        hist_all[ch] = df

    # Pós final (garantia)
    for ch, df in hist_all.items():
        ent = colab_by_chapa[ch].get("Entrada", "06:00")
        locked = locked_idx.get(ch, set())
        ultima_saida_prev = "" if _past else (estado_prev.get(ch, {}).get("ultima_saida", "") or "")

        if _past:
            base_first = None
        else:
            prev_dom = infer_ultimo_domingo_status_from_escala(setor, int(ano), int(mes), ch)
            if prev_dom not in ("Folga","Trabalho"):
                prev_dom = (estado_prev.get(ch, {}) or {}).get("ultimo_domingo_status", None)
            if prev_dom == "Folga":
                base_first = "Trabalho"
            elif prev_dom == "Trabalho":
                base_first = "Folga"
            else:
                base_first = None
        enforce_sundays_1x1_for_employee(df, ent, locked_status=locked, base_first=base_first)
        locked_dom = _merge_locked_status(locked, _sunday_indices_df(df))
        enforce_no_consecutive_folga(df, locked_status=locked_dom)
        enforce_weekly_folga_targets(df, df_ref=(df_ref_use if df_ref_use is not None else df_ref_cur), pode_folgar_sabado=bool(colab_by_chapa[ch].get('Folga_Sab', False)), locked_status=locked_dom)

        # ✅ garante 5 dias depois do weekly (porque weekly pode remover folga)
        enforce_max_5_consecutive_work(
            df, ent, bool(colab_by_chapa[ch].get('Folga_Sab', False)),
            initial_consec=(0 if _past else int((estado_prev.get(ch, {}) or {}).get('consec_trab_final', 0))),
            locked_status=locked_dom
        )
        enforce_no_consecutive_folga(df, locked_status=locked_dom)

        enforce_global_rest_keep_targets(df, ent, locked_status=locked, ultima_saida_prev=ultima_saida_prev)

        if respeitar_ajustes:
            _apply_overrides_to_df_inplace(df, setor, ch, ovmap)

        hist_all[ch] = df

    # Travamento global final de domingo 1x1 (todos os setores/subgrupos)
    locked_idx = _lock_and_fix_sundays_global(
        hist_all, colab_by_chapa, locked_idx, setor, ano, mes,
        estado_prev=estado_prev, past_flag=_past
    )
    try:
        enforce_max_two_folgas_per_week(hist_all, list(hist_all.keys()), df_ref_cur, setor, ano, mes, locked_idx_map=locked_idx)
    except Exception:
        pass

    # Estado do mês
    estado_out = {}
    for ch, df in hist_all.items():
        consec = 0
        for i in range(len(df) - 1, -1, -1):
            if df.loc[i, "Status"] in WORK_STATUSES:
                consec += 1
            else:
                break

        ultima_saida = ""
        for i in range(len(df) - 1, -1, -1):
            if df.loc[i, "Status"] in WORK_STATUSES and (df.loc[i, "H_Saida"] or ""):
                ultima_saida = df.loc[i, "H_Saida"]
                break

        ultimo_dom = None
        for i in range(len(df) - 1, -1, -1):
            if df.loc[i, "Dia"] == "dom":
                if df.loc[i, "Status"] == "Folga":
                    ultimo_dom = "Folga"
                    break
                if df.loc[i, "Status"] in WORK_STATUSES:
                    ultimo_dom = "Trabalho"
                    break

        estado_out[ch] = {"consec_trab_final": consec, "ultima_saida": ultima_saida, "ultimo_domingo_status": ultimo_dom}

    # ============================
    # GARANTIA FINAL CORRETA (TODOS OS COLABORADORES)
    # A versão anterior aplicava a garantia final só no último `df` da função.
    # Aqui roda para cada colaborador do hist_all, usando a referência do mês anterior
    # para a semana contínua SEG->DOM (virada de mês).
    # ============================
    for ch, df in list(hist_all.items()):
        ent = colab_by_chapa[ch].get("Entrada", "06:00")
        pode_sab = bool(colab_by_chapa[ch].get("Folga_Sab", False))
        locked = locked_idx.get(ch, set())
        locked_dom = _merge_locked_status(locked, _sunday_indices_df(df))

        try:
            enforce_max_5_consecutive_work(df, ent, pode_sab, locked_status=locked_dom)
        except Exception:
            pass

        try:
            df = enforce_weekly_folga_targets(
                df,
                df_ref=(df_ref_use if df_ref_use is not None else df_ref_cur),
                pode_folgar_sabado=pode_sab,
                locked_status=locked_dom
            )
        except Exception:
            pass

        try:
            df = enforce_no_consecutive_folgas(df, locked_status=locked_dom)
        except Exception:
            pass

        try:
            df = enforce_weekly_folga_targets(
                df,
                df_ref=(df_ref_use if df_ref_use is not None else df_ref_cur),
                pode_folgar_sabado=pode_sab,
                locked_status=locked_dom
            )
        except Exception:
            pass

        try:
            df = _cap_total_folgas_por_semana(
                df,
                target_total=2,
                locked_status=locked_dom,
                df_ref=(df_ref_use if df_ref_use is not None else df_ref_cur)
            )
        except Exception:
            pass

        try:
            df = enforce_weekly_folga_targets(
                df,
                df_ref=(df_ref_use if df_ref_use is not None else df_ref_cur),
                pode_folgar_sabado=pode_sab,
                locked_status=locked_dom
            )
        except Exception:
            pass

        hist_all[ch] = df




    # GARANTIA FINAL INQUEBRÁVEL: a unidade é a semana real SEG->DOM
    try:
        locked_idx = _lock_and_fix_sundays_global(
            hist_all, colab_by_chapa, locked_idx, setor, ano, mes,
            estado_prev=estado_prev, past_flag=_past
        )
        for ch, df in list(hist_all.items()):
            hist_all[ch] = strict_weekly_5x2_never_break(
                df,
                chapa=ch,
                ent_padrao=colab_by_chapa[ch].get("Entrada", "06:00"),
                pode_folgar_sabado=bool(colab_by_chapa[ch].get("Folga_Sab", False)),
                locked_status=_merge_locked_status(locked_idx.get(ch, set()), _sunday_indices_df(df)),
                df_ref_prev=df_ref_prev
            )
    except Exception:
        pass

    estado_out = _rebuild_estado_out(hist_all)
    return hist_all, estado_out

# =========================================================
# DASHBOARD / CALENDÁRIO / BANCO DE HORAS
# (resto do arquivo igual ao seu original — UI completa)
# =========================================================


    # ============================
    # GARANTIA FINAL (5x2 semanal)
    # - Semana SEG->DOM tem 2 folgas no total (domingo conta se for folga)
    # - Reaplica após todas as outras regras que podem adicionar/remover folga
    # - Cap final: nunca deixa 3 folgas na mesma semana (exceto se travado)
    # ============================
    try:
        df = enforce_weekly_folga_targets(df, df_ref=(df_ref_use if df_ref_use is not None else df_ref_cur), pode_folgar_sabado=pode_sab, locked_status=locked)
    except Exception:
        try:
            df = enforce_weekly_folga_targets(df)
        except Exception:
            pass

    try:
        enforce_max_5_consecutive_work(df, ent, pode_sab)
    except Exception:
        pass

    try:
        df = enforce_weekly_folga_targets(df, df_ref=(df_ref_use if df_ref_use is not None else df_ref_cur), pode_folgar_sabado=pode_sab, locked_status=locked)
    except Exception:
        try:
            df = enforce_weekly_folga_targets(df)
        except Exception:
            pass

    try:
        df = _cap_total_folgas_por_semana(df, target_total=2, locked_status=locked, df_ref=(df_ref_use if df_ref_use is not None else df_ref_cur))
    except Exception:
        try:
            df = _cap_total_folgas_por_semana(df, target_total=2)
        except Exception:
            pass


def banco_horas_df(hist_db: dict[str, pd.DataFrame], colab_by: dict, base_min: int):
    rows = []
    for ch, df in hist_db.items():
        nome = colab_by.get(ch, {}).get("Nome", ch)
        saldo = 0
        for _, r in df.iterrows():
            if r["Status"] not in WORK_STATUSES:
                continue
            ent = r.get("H_Entrada", "") or ""
            sai = r.get("H_Saida", "") or ""
            if not ent or not sai:
                continue
            dur = _to_min(sai) - _to_min(ent)
            if dur < 0:
                dur += 24 * 60
            saldo += (dur - base_min)
        rows.append({"Nome": nome, "Chapa": ch, "Saldo_min": saldo, "Saldo_h": round(saldo/60, 2)})
    return pd.DataFrame(rows).sort_values(["Saldo_min"], ascending=False)

def calendario_rh_df(hist_db: dict[str, pd.DataFrame], colab_by: dict):
    if not hist_db:
        return pd.DataFrame()
    any_df = next(iter(hist_db.values()))
    dias = [str(int(r.day)) for r in pd.to_datetime(any_df["Data"]).dt.date]
    cols = ["Nome", "Chapa", "Subgrupo"] + dias
    rows = []
    for ch, df in hist_db.items():
        nome = colab_by.get(ch, {}).get("Nome", ch)
        sg = (colab_by.get(ch, {}).get("Subgrupo", "") or "").strip() or "SEM SUBGRUPO"
        row = [nome, ch, sg]
        for i in range(len(df)):
            stt = df.loc[i, "Status"]
            if stt == "Folga":
                row.append("F")
            elif stt == "Férias":
                row.append("FER")
            else:
                row.append(df.loc[i, "H_Entrada"] or "")
        rows.append(row)
    out = pd.DataFrame(rows, columns=cols)
    return out.sort_values(["Subgrupo", "Nome"]).reset_index(drop=True)

def style_calendario(df: pd.DataFrame, mes: int, ano: int):
    if df.empty:
        return df
    dias_cols = df.columns[3:]
    qtd = calendar.monthrange(int(ano), int(mes))[1]
    dsem = {}
    for d in range(1, qtd + 1):
        ds = pd.Timestamp(year=int(ano), month=int(mes), day=int(d)).day_name()
        dsem[str(d)] = D_PT[ds]




# =========================================================
# MAPA ANUAL DE FÉRIAS (visual tipo "grade")
# =========================================================
MESES_PT = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

def _parse_date_ymd(s: str):
    try:
        return datetime.strptime(str(s), "%Y-%m-%d").date()
    except Exception:
        return None

def ferias_mapa_ano_df(setor: str, ano: int, colaboradores: list[dict]) -> pd.DataFrame:
    """
    DF:
      Nome | Chapa | Janeiro..Dezembro
    Marca "FER" quando houver QUALQUER dia de férias no mês.
    """
    rows = list_ferias(setor)  # [(chapa,inicio,fim), ...]
    fer_by = {}
    for chapa, ini, fim in rows:
        ini_d = _parse_date_ymd(ini)
        fim_d = _parse_date_ymd(fim)
        if not ini_d or not fim_d:
            continue
        fer_by.setdefault(str(chapa), []).append((ini_d, fim_d))

    colabs_sorted = sorted(colaboradores, key=lambda c: ((c.get("Nome") or ""), (c.get("Chapa") or "")))
    out = []
    for c in colabs_sorted:
        ch = str(c.get("Chapa") or "")
        nome = str(c.get("Nome") or ch)
        linha = {"Nome": nome, "Chapa": ch}
        periods = fer_by.get(ch, [])
        for m in range(1, 13):
            first = date(int(ano), m, 1)
            last = date(int(ano), m, calendar.monthrange(int(ano), m)[1])
            marcou = False
            for ini_d, fim_d in periods:
                if ini_d <= last and fim_d >= first:
                    marcou = True
                    break
            linha[MESES_PT[m-1]] = "FER" if marcou else ""
        out.append(linha)

    return pd.DataFrame(out, columns=["Nome","Chapa"] + MESES_PT)

def style_ferias_mapa(df: pd.DataFrame):
    if df is None or df.empty:
        return df
    meses = [c for c in df.columns if c in MESES_PT]

    def cell(v, col):
        if col in meses:
            if str(v) == "FER":
                return "background-color:#1F4E78; color:#FFFFFF; font-weight:800; text-align:center;"
            return "background-color:#F2F2F2; color:#000000; text-align:center;"
        if col == "Nome":
            return "font-weight:700;"
        return ""

    styles = pd.DataFrame("", index=df.index, columns=df.columns)
    for col in df.columns:
        styles[col] = df[col].apply(lambda v: cell(v, col))
    return df.style.apply(lambda _: styles, axis=None)


# =========================================================
# ÚLTIMAS FÉRIAS + ALERTA (1 ano e 11 meses) + DURAÇÃO
# =========================================================
def _months_between(d1: date, d2: date) -> int:
    """Meses inteiros aproximados entre datas (d2 >= d1)."""
    if not d1 or not d2:
        return 0
    if d2 < d1:
        d1, d2 = d2, d1
    return (d2.year - d1.year) * 12 + (d2.month - d1.month)

def get_ultima_ferias_info(setor: str, chapa: str):
    """
    Retorna dict com:
      - ultima_inicio (date|None)
      - ultima_fim (date|None)
      - dias_ultima (int|None)
      - meses_desde_ultima_fim (int|None)  # até hoje
    Considera o período com maior 'fim' como a última.
    """
    chapa = str(chapa or "").strip()
    if not chapa:
        return {"ultima_inicio": None, "ultima_fim": None, "dias_ultima": None, "meses_desde_ultima_fim": None}

    rows = list_ferias(setor)  # [(chapa,inicio,fim), ...]
    last = None  # (fim_date, ini_date)
    for ch, ini, fim in rows:
        if str(ch) != chapa:
            continue
        ini_d = _parse_date_ymd(ini)
        fim_d = _parse_date_ymd(fim)
        if not ini_d or not fim_d:
            continue
        if last is None or fim_d > last[0]:
            last = (fim_d, ini_d)

    if not last:
        return {"ultima_inicio": None, "ultima_fim": None, "dias_ultima": None, "meses_desde_ultima_fim": None}

    ultima_fim, ultima_ini = last[0], last[1]
    dias = (ultima_fim - ultima_ini).days + 1
    meses = _months_between(ultima_fim, date.today())
    return {"ultima_inicio": ultima_ini, "ultima_fim": ultima_fim, "dias_ultima": dias, "meses_desde_ultima_fim": meses}

def _classificar_duracao_ferias(qtd_dias: int) -> str:
    if qtd_dias == 15:
        return "15 dias"
    if qtd_dias == 30:
        return "30 dias"
    if qtd_dias and qtd_dias > 0:
        return f"{qtd_dias} dias"
    return "-"



def ferias_resumo_mensal_df(setor: str, ano: int) -> pd.DataFrame:
    """
    Resumo mensal:
      - Pessoas_em_ferias: qtd de colaboradores com QUALQUER dia de férias no mês
      - Lancamentos: qtd de períodos (linhas) de férias que encostam no mês
    """
    rows = list_ferias(setor)  # [(chapa,inicio,fim), ...]
    # map month -> set(chapa) and count launches touching month
    people = {m: set() for m in range(1, 13)}
    launches = {m: 0 for m in range(1, 13)}

    for chapa, ini, fim in rows:
        ini_d = _parse_date_ymd(ini)
        fim_d = _parse_date_ymd(fim)
        if not ini_d or not fim_d:
            continue

        for m in range(1, 13):
            first = date(int(ano), m, 1)
            last = date(int(ano), m, calendar.monthrange(int(ano), m)[1])
            if ini_d <= last and fim_d >= first:
                people[m].add(str(chapa))
                launches[m] += 1

    data = []
    for m in range(1, 13):
        data.append({
            "Mês": MESES_PT[m-1],
            "Pessoas_em_ferias": len(people[m]),
            "Lancamentos": int(launches[m])
        })
    return pd.DataFrame(data)


# =========================================================
# PDF UI helpers (filtro estilo "Impressão de Escala")
# =========================================================
def _filtrar_colaboradores(colaboradores: list[dict], subgrupos_sel: list[str] | None, busca: str | None):
    subgrupos_sel = subgrupos_sel or []
    busca = (busca or "").strip().lower()
    out = []
    for c in colaboradores:
        sg = (c.get("Subgrupo") or "").strip() or "SEM SUBGRUPO"
        nome = (c.get("Nome") or "").strip()
        ch = (c.get("Chapa") or "").strip()
        if subgrupos_sel and sg not in subgrupos_sel:
            continue
        if busca:
            key = f"{nome} {ch} {sg}".lower()
            if busca not in key:
                continue
        out.append(c)
    return out

    def cell_style(v, col):
        if col in dias_cols:
            dia_sem = dsem.get(col, "")
            if str(v) == "F":
                return "background-color:#FFF2CC; color:#000000; font-weight:700;"
            if str(v) == "FER":
                return "background-color:#92D050; color:#000000; font-weight:700;"
            if dia_sem == "dom":
                return "background-color:#BDD7EE; color:#000000;"
        return ""

    styles = pd.DataFrame("", index=df.index, columns=df.columns)
    for c in df.columns:
        styles[c] = df[c].apply(lambda v: cell_style(v, c))
    return df.style.apply(lambda _: styles, axis=None)



def _month_iter(start_d: date, end_d: date):
    y, m = start_d.year, start_d.month
    while (y, m) <= (end_d.year, end_d.month):
        yield y, m
        if m == 12:
            y += 1
            m = 1
        else:
            m += 1

def _load_hist_periodo(setor: str, data_ini: date, data_fim: date) -> dict[str, pd.DataFrame]:
    hist_all: dict[str, list[pd.DataFrame]] = {}
    for y, m in _month_iter(data_ini, data_fim):
        hist_mes = load_escala_mes_db(setor, int(y), int(m)) or {}
        if not hist_mes:
            continue
        hist_mes = apply_overrides_to_hist(setor, int(y), int(m), hist_mes)
        hist_mes = _apply_retificacoes_to_hist(setor, int(y), int(m), hist_mes)
        for ch, df in hist_mes.items():
            dfx = df.copy()
            dfx["Data"] = pd.to_datetime(dfx["Data"]).dt.date
            dfx = dfx[(dfx["Data"] >= data_ini) & (dfx["Data"] <= data_fim)]
            if dfx.empty:
                continue
            hist_all.setdefault(str(ch), []).append(dfx)
    out = {}
    for ch, parts in hist_all.items():
        dfc = pd.concat(parts, ignore_index=True).sort_values("Data").drop_duplicates(subset=["Data"], keep="last")
        out[ch] = dfc.reset_index(drop=True)
    return out

def _fmt_periodo_cell(row: pd.Series) -> str:
    stt = str(row.get("Status", "") or "").strip()
    ent = str(row.get("H_Entrada", "") or "").strip()
    sai = str(row.get("H_Saida", "") or "").strip()
    if stt == "Folga":
        return "F"
    if stt == "Férias":
        return "FER"
    if stt == "Afastamento":
        return "AFA"
    if stt in WORK_STATUSES:
        if ent and sai:
            return f"{ent}\n{sai}"
        if ent:
            return ent
        return "T"
    return ""

def gerar_pdf_periodo_panoramico(setor: str, data_ini: date, data_fim: date, hist_db: dict, colaboradores: list[dict]) -> bytes:
    from io import BytesIO
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.pagesizes import landscape, A1, A2, A3, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm

    total_days = (data_fim - data_ini).days + 1

    # Papel automático maior para não cortar nome nem esmagar os dias
    if total_days <= 20:
        page_size = landscape(A4)
    elif total_days <= 35:
        page_size = landscape(A3)
    elif total_days <= 50:
        page_size = landscape(A2)
    else:
        page_size = landscape(A1)

    W, H = page_size
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=8*mm,
        rightMargin=8*mm,
        topMargin=10*mm,
        bottomMargin=8*mm
    )

    styles = getSampleStyleSheet()
    font_small = 4.6 if total_days > 60 else (5.0 if total_days > 45 else 5.6)
    leading_small = font_small + 0.8
    normal = ParagraphStyle(
        'periodo_cell',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=font_small,
        leading=leading_small,
        alignment=1,
    )
    left_small = ParagraphStyle(
        'periodo_left',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=6.2 if total_days > 45 else 6.8,
        leading=7.0 if total_days > 45 else 7.6,
        alignment=0,
    )
    left_small_b = ParagraphStyle('periodo_left_b', parent=left_small, fontName='Helvetica-Bold')
    title_st = ParagraphStyle('periodo_title', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=13, leading=15, alignment=1)
    meta_st = ParagraphStyle('periodo_meta', parent=styles['BodyText'], fontName='Helvetica', fontSize=8, leading=10, alignment=1)

    datas = [data_ini + timedelta(days=i) for i in range(total_days)]

    usable_w = W - doc.leftMargin - doc.rightMargin

    # Coluna bem mais larga: Nome + Chapa + Subgrupo
    first_col_w = 78 * mm if total_days <= 35 else (88 * mm if total_days <= 50 else 98 * mm)

    # Todos os dias do período, sem pular
    day_w = max(5.2*mm, (usable_w - first_col_w) / max(1, total_days))

    def build_group_rows(group_name: str, items: list[dict]):
        rows = []
        rows.append([Paragraph(f'<b>SUBGRUPO: {group_name}</b>', left_small_b)] + [''] * total_days)
        for c in items:
            ch = str(c.get('Chapa'))
            nome = str(c.get('Nome') or '').strip()
            subgrupo = (c.get('Subgrupo') or '').strip() or 'SEM SUBGRUPO'
            df = hist_db.get(ch, pd.DataFrame())
            row_map = {}

            if not df.empty:
                tmp = df.copy()
                if 'Data' in tmp.columns:
                    tmp['Data'] = pd.to_datetime(tmp['Data']).dt.date
                    for _, r in tmp.iterrows():
                        row_map[r['Data']] = _fmt_periodo_cell(r)

            label = (
                f"<b>{nome}</b><br/>"
                f"<font size='5'>CHAPA: {ch}</font><br/>"
                f"<font size='5'>SUBGRUPO: {subgrupo}</font>"
            )
            row = [Paragraph(label, left_small)]
            for dd in datas:
                txt = row_map.get(dd, '') or ''
                row.append(Paragraph(str(txt).replace('\n', '<br/>'), normal))
            rows.append(row)
        return rows

    story = []
    title = f"Escala panorâmica - {setor} - {data_ini.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
    story.append(Paragraph(title, title_st))
    story.append(Paragraph('Visual panorâmico por período. Células exibem entrada e saída; F = Folga; FER = Férias; AFA = Afastamento.', meta_st))
    story.append(Spacer(1, 4*mm))

    filtered = [c for c in colaboradores if str(c.get('Chapa')) in set(hist_db.keys())]
    filtered.sort(key=lambda x: ((x.get('Subgrupo') or '').strip() or 'SEM SUBGRUPO', (x.get('Nome') or '').strip()))
    groups = {}
    for c in filtered:
        sg = (c.get('Subgrupo') or '').strip() or 'SEM SUBGRUPO'
        groups.setdefault(sg, []).append(c)

    header1 = [Paragraph('<b>COLABORADOR / CHAPA / SUBGRUPO</b>', left_small_b)] + [Paragraph(f"<b>{d.day}</b>", normal) for d in datas]
    header2 = [''] + [Paragraph(f"<b>{D_PT[pd.Timestamp(d).day_name()]}</b>", normal) for d in datas]

    all_rows = [header1, header2]
    for sg, items in groups.items():
        all_rows.extend(build_group_rows(sg, items))

    tbl = Table(all_rows, colWidths=[first_col_w] + [day_w] * total_days, repeatRows=2)
    ts = TableStyle([
        ('GRID', (0,0), (-1,-1), 0.30, colors.black),
        ('BACKGROUND', (0,0), (-1,1), colors.HexColor('#1f4e78')),
        ('TEXTCOLOR', (0,0), (-1,1), colors.white),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('LEFTPADDING', (0,0), (-1,-1), 2),
        ('RIGHTPADDING', (0,0), (-1,-1), 2),
        ('TOPPADDING', (0,0), (-1,-1), 1),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ('FONTNAME', (0,0), (-1,1), 'Helvetica-Bold'),
    ])

    row_idx = 2
    for sg, items in groups.items():
        ts.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#d9e2f3'))
        ts.add('SPAN', (0, row_idx), (-1, row_idx))
        row_idx += 1
        for _ in items:
            ts.add('BACKGROUND', (0, row_idx), (0, row_idx), colors.HexColor('#d9e2f3'))
            row_idx += 1

    for cidx, dd in enumerate(datas, start=1):
        if dd.weekday() == 6:
            ts.add('BACKGROUND', (cidx, 0), (cidx, -1), colors.HexColor('#fde9d9'))
        elif dd.weekday() == 5:
            ts.add('BACKGROUND', (cidx, 2), (cidx, -1), colors.HexColor('#f7f7f7'))

    for r in range(2, len(all_rows)):
        if isinstance(all_rows[r][0], Paragraph) and 'SUBGRUPO:' in getattr(all_rows[r][0], 'text', ''):
            continue
        for cidx in range(1, total_days+1):
            cell = all_rows[r][cidx]
            txt = cell.text if isinstance(cell, Paragraph) else str(cell)
            plain = txt.replace('<br/>', ' ').replace('<b>', '').replace('</b>', '').strip().upper()
            if plain == 'F':
                ts.add('BACKGROUND', (cidx, r), (cidx, r), colors.HexColor('#FFF2CC'))
                ts.add('FONTNAME', (cidx, r), (cidx, r), 'Helvetica-Bold')
            elif plain == 'FER':
                ts.add('BACKGROUND', (cidx, r), (cidx, r), colors.HexColor('#92D050'))
                ts.add('FONTNAME', (cidx, r), (cidx, r), 'Helvetica-Bold')
            elif plain == 'AFA':
                ts.add('BACKGROUND', (cidx, r), (cidx, r), colors.HexColor('#D9EAD3'))
                ts.add('FONTNAME', (cidx, r), (cidx, r), 'Helvetica-Bold')

    tbl.setStyle(ts)
    story.append(tbl)
    doc.build(story)
    return buffer.getvalue()

# =========================================================
# UI
# =========================================================
if "auth" not in st.session_state:
    st.session_state["auth"] = None
if "auth_force_change" not in st.session_state:
    st.session_state["auth_force_change"] = False
if "pwd_temp_last" not in st.session_state:
    st.session_state["pwd_temp_last"] = ""
if "pwd_temp_last_chapa" not in st.session_state:
    st.session_state["pwd_temp_last_chapa"] = ""
if "cfg_mes" not in st.session_state:
    st.session_state["cfg_mes"] = datetime.now().month
if "cfg_ano" not in st.session_state:
    st.session_state["cfg_ano"] = datetime.now().year
if "last_seed" not in st.session_state:
    st.session_state["last_seed"] = 0


def page_login():
    @st.cache_data(ttl=60, show_spinner=False)
    def _cache_login_setores():
        con = db_conn()
        try:
            setores_df = pd.concat([
                pd.read_sql_query("SELECT nome AS setor FROM setores", con),
                pd.read_sql_query("SELECT DISTINCT setor FROM usuarios_sistema", con),
                pd.read_sql_query("SELECT DISTINCT setor FROM colaboradores", con),
            ], ignore_index=True)
            return sorted({_norm_setor(x) for x in setores_df["setor"].dropna().tolist() if str(x).strip()})
        finally:
            con.close()

    @st.cache_data(ttl=30, show_spinner=False)
    def _cache_login_recent():
        con = db_conn()
        try:
            rec = pd.read_sql_query("SELECT setor, chapa, ts FROM login_recent ORDER BY ts DESC LIMIT 6", con)
            try:
                rec2 = rec.merge(pd.read_sql_query("SELECT setor, chapa, nome FROM usuarios_sistema", con), on=["setor","chapa"], how="left")
            except Exception:
                rec2 = rec.copy()
                rec2["nome"] = ""
            return rec2
        finally:
            con.close()

    st.title("ESCALA 5x2 DO FUTURO")

    if _RESTORE_GUARD_ACTIVE:
        st.warning(_RESTORE_GUARD_MESSAGE or "Base não restaurada automaticamente. O app liberou uma base mínima temporária para permitir o login.")
        st.info("Publique o latest_stable.db junto do projeto ou confirme que o Supabase tem os dados completos para recuperar a base real.")

    login_sec = st.radio("", ["Entrar", "Esqueci a senha"], horizontal=True, key="login_nav_fast", label_visibility="collapsed")

    if login_sec == "Entrar":
        setores = _cache_login_setores()
        setor_base = _norm_setor(st.session_state.get("lg_setor_txt", ""))
        opcoes_setor = setores[:] if setores else []
        if setor_base and setor_base not in opcoes_setor:
            opcoes_setor = [setor_base] + opcoes_setor
        if not opcoes_setor:
            opcoes_setor = [setor_base] if setor_base else [""]

        idx_setor = 0
        if setor_base in opcoes_setor:
            idx_setor = opcoes_setor.index(setor_base)

        setor_escolhido = st.selectbox("Setor:", opcoes_setor, index=idx_setor, key="lg_setor_sel")
        st.session_state["lg_setor_txt"] = setor_escolhido
        setor_norm = _norm_setor(setor_escolhido)

        chapa = st.text_input("Chapa:", value=st.session_state.get("lg_chapa",""), key="lg_chapa")
        senha = st.text_input("Senha:", type="password", key="lg_senha")

        if st.button("Entrar", key="lg_btn"):
            u = verify_login(setor_norm, chapa.strip(), senha)
            if u:
                try:
                    u["is_lider"] = bool(u.get("is_lider", False)) or colaborador_eh_lideranca(setor_norm, chapa.strip())
                except Exception:
                    pass

                st.session_state["auth"] = u
                st.session_state["auth_force_change"] = bool(u.get("forcar_troca_senha", False))
                st.success("Login efetuado!")
                st.rerun()
            else:
                if colaborador_lookup(setor_norm, chapa.strip()) and not system_user_exists(setor_norm, chapa.strip()):
                    st.error("Este colaborador existe, mas o login do sistema foi apagado ou ainda não foi criado. Peça para o ADMIN recuperar o usuário na aba Admin.")
                else:
                    st.error("Usuário ou senha inválidos.")

    elif login_sec == "Cadastro do Colaborador":
        st.subheader("Primeiro acesso do colaborador")
        st.info("Use setor + chapa para localizar sua escala. Nome, criação de setor e perfis de líder/admin continuam na aba Admin.")
        con = db_conn()
        try:
            setores_df = pd.concat([
                pd.read_sql_query("SELECT nome AS setor FROM setores", con),
                pd.read_sql_query("SELECT DISTINCT setor FROM colaboradores", con),
                pd.read_sql_query("SELECT DISTINCT setor FROM usuarios_sistema", con),
            ], ignore_index=True)
            setores = sorted({_norm_setor(x) for x in setores_df["setor"].dropna().tolist() if str(x).strip()})
        finally:
            con.close()

        setor = st.selectbox("Setor:", setores, key="cl_setor") if setores else st.text_input("Setor:", key="cl_setor_txt")
        chapa = st.text_input("Chapa:", key="cl_chapa")
        senha = st.text_input("Criar senha:", type="password", key="cl_senha")
        senha2 = st.text_input("Confirmar senha:", type="password", key="cl_senha2")

        if st.button("Criar acesso", key="cl_btn"):
            setor_n = _norm_setor(setor)
            chapa_n = _norm_chapa(chapa)
            if not setor_n or not chapa_n or not senha:
                st.error("Preencha setor, chapa e senha.")
            elif senha != senha2:
                st.error("Senhas não conferem.")
            else:
                row = colaborador_lookup(setor_n, chapa_n)
                if not row:
                    st.error("Colaborador não encontrado neste setor. Confira setor e chapa ou peça ajuste ao Admin.")
                else:
                    nome_db, setor_db, chapa_db = row
                    if system_user_exists(setor_db, chapa_db):
                        st.error("Este colaborador já possui acesso. Use Entrar ou Esqueci a senha.")
                    else:
                        create_system_user(nome_db or chapa_db, setor_db, chapa_db, senha, is_lider=0, is_admin=0)
                        st.success("Acesso criado com sucesso. Agora faça login com setor + chapa + senha.")
                        st.rerun()

    elif login_sec == "Esqueci a senha":
        st.subheader("Redefinir senha do colaborador")
        st.caption("Use a senha temporária recebida do líder ou admin para criar sua nova senha.")
        con = db_conn()
        setores_df = pd.concat([
            pd.read_sql_query("SELECT nome AS setor FROM setores", con),
            pd.read_sql_query("SELECT DISTINCT setor FROM usuarios_sistema", con),
            pd.read_sql_query("SELECT DISTINCT setor FROM colaboradores", con),
        ], ignore_index=True)
        setores = sorted({_norm_setor(x) for x in setores_df["setor"].dropna().tolist() if str(x).strip()})
        con.close()

        setor = st.selectbox("Setor:", setores, key="fp_setor") if setores else st.text_input("Setor:", key="fp_setor_txt")
        chapa = st.text_input("Chapa:", key="fp_chapa")
        senha_temp = st.text_input("Senha temporária:", type="password", key="fp_temp")
        nova = st.text_input("Nova senha:", type="password", key="fp_nova")
        nova2 = st.text_input("Confirmar senha:", type="password", key="fp_nova2")

        if st.button("Redefinir", key="fp_btn"):
            setor_n = _norm_setor(setor)
            chapa_n = _norm_chapa(chapa)
            senha_temp_n = (senha_temp or "").strip()
            if not setor_n or not chapa_n or not senha_temp_n or not nova or not nova2:
                st.error("Preencha setor, chapa, senha temporária, nova senha e confirmação.")
            elif nova != nova2:
                st.error("Senhas não conferem.")
            else:
                user = verify_login(setor_n, chapa_n, senha_temp_n)
                if not user:
                    st.error("Senha temporária inválida para este setor/chapa.")
                elif not bool(user.get("forcar_troca_senha", False)):
                    st.error("Este colaborador não está com redefinição por senha temporária ativa.")
                else:
                    try:
                        update_password(setor_n, chapa_n, nova)
                        set_force_change_password(setor_n, chapa_n, False)
                        st.success("Senha redefinida com sucesso. Agora faça login com a nova senha.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Falha ao redefinir senha: {e}")

def _regenerar_mes_inteiro(setor: str, ano: int, mes: int, seed: int = 0, respeitar_ajustes: bool = True):
    """
    Regera a escala do mês inteiro para TODO o setor.

    ✅ Garantias:
    - Se respeitar_ajustes=True, TODAS as folgas/alterações manuais (overrides) são reaplicadas
      no final e gravadas novamente no banco (escala_mes). Isso evita “sumir” folga manual ao gerar.
    """
    colaboradores = load_colaboradores_setor(setor)
    if not colaboradores:
        return False

    random.seed(int(seed))
    # ===== CONTEXTO SEMANA CONTÍNUA (mês anterior) =====
    df_ref = None
    try:
        ano_prev = int(ano)
        mes_prev = int(mes) - 1
        if mes_prev <= 0:
            mes_prev = 12
            ano_prev -= 1

        prev_obj = load_escala_mes_db(setor, ano_prev, mes_prev) if "load_escala_mes_db" in globals() else None

        # load_escala_mes_db retorna dict[chapa] -> DataFrame
        if isinstance(prev_obj, dict) and prev_obj:
            parts = []
            for ch, dfp in prev_obj.items():
                if dfp is None or getattr(dfp, "empty", True):
                    continue
                dfx = dfp.copy()
                dfx["Chapa"] = str(ch)
                # garante colunas
                if "Data" not in dfx.columns and "data" in dfx.columns:
                    dfx["Data"] = dfx["data"]
                if "Status" not in dfx.columns and "status" in dfx.columns:
                    dfx["Status"] = dfx["status"]
                parts.append(dfx[["Data", "Chapa", "Status"]].copy())
            if parts:
                df_ref = pd.concat(parts, ignore_index=True)

        # Caso antigo: se algum dia retornar DataFrame único
        elif isinstance(prev_obj, pd.DataFrame) and (not prev_obj.empty):
            prev = prev_obj.copy()
            if "Data" not in prev.columns:
                for c in ("data", "dia", "DataDia"):
                    if c in prev.columns:
                        prev["Data"] = prev[c]
                        break
            if "Chapa" not in prev.columns:
                for c in ("chapa", "CHAPA"):
                    if c in prev.columns:
                        prev["Chapa"] = prev[c]
                        break
            if "Status" not in prev.columns:
                for c in ("status", "STATUS"):
                    if c in prev.columns:
                        prev["Status"] = prev[c]
                        break
            df_ref = prev[["Data", "Chapa", "Status"]].copy()

    except Exception:
        df_ref = None
    # ===== df_ref_prev (mês anterior) para semana contínua SEG->DOM =====
    df_ref_prev = None
    try:
        ano_prev = int(ano)
        mes_prev = int(mes) - 1
        if mes_prev <= 0:
            mes_prev = 12
            ano_prev -= 1

        prev_obj = load_escala_mes_db(setor, ano_prev, mes_prev) if "load_escala_mes_db" in globals() else None

        # load_escala_mes_db normalmente retorna dict[chapa] -> DataFrame
        if isinstance(prev_obj, dict) and prev_obj:
            parts = []
            for ch_prev, dfp in prev_obj.items():
                if dfp is None or getattr(dfp, "empty", True):
                    continue
                dfx = dfp.copy()
                # garante colunas Data/Status
                if "Data" not in dfx.columns:
                    for c in ("data","dia","DataDia"):
                        if c in dfx.columns:
                            dfx["Data"] = dfx[c]
                            break
                if "Status" not in dfx.columns:
                    for c in ("status","STATUS"):
                        if c in dfx.columns:
                            dfx["Status"] = dfx[c]
                            break
                dfx["Chapa"] = str(ch_prev)
                if "Data" in dfx.columns and "Status" in dfx.columns:
                    parts.append(dfx[["Data","Chapa","Status"]].copy())
            if parts:
                df_ref_prev = pd.concat(parts, ignore_index=True)

        elif isinstance(prev_obj, pd.DataFrame) and (not prev_obj.empty):
            dfx = prev_obj.copy()
            if "Data" not in dfx.columns:
                for c in ("data","dia","DataDia"):
                    if c in dfx.columns:
                        dfx["Data"] = dfx[c]
                        break
            if "Chapa" not in dfx.columns:
                for c in ("chapa","CHAPA"):
                    if c in dfx.columns:
                        dfx["Chapa"] = dfx[c]
                        break
            if "Status" not in dfx.columns:
                for c in ("status","STATUS"):
                    if c in dfx.columns:
                        dfx["Status"] = dfx[c]
                        break
            if {"Data","Chapa","Status"}.issubset(set(dfx.columns)):
                df_ref_prev = dfx[["Data","Chapa","Status"]].copy()
    except Exception:
        df_ref_prev = None


    hist, estado_out = gerar_escala_setor_por_subgrupo(
        setor, colaboradores, int(ano), int(mes),
        respeitar_ajustes=bool(respeitar_ajustes),
        df_ref_prev=df_ref_prev
    )

    # 1) grava a geração
    save_escala_mes_db(setor, int(ano), int(mes), hist)
    save_estado_mes(setor, int(ano), int(mes), estado_out)

    # 2) “pós-fix”: reaplica overrides do banco e grava de novo
    if bool(respeitar_ajustes):
        hist_db = load_escala_mes_db(setor, int(ano), int(mes))
        hist_db = apply_overrides_to_hist(setor, int(ano), int(mes), hist_db)
        if hist_db:
            save_escala_mes_db(setor, int(ano), int(mes), hist_db)

    return True




def page_gestao_dashboard(ano: int, mes: int):
    st.title("📊 Gestão — Visão Geral (todos os setores)")
    st.caption("Indicadores de trabalho, folgas, férias e afastamentos. Use os filtros para cruzar setor e período.")

    con = db_conn()
    try:
        setores_all = pd.read_sql_query("SELECT nome FROM setores ORDER BY nome ASC", con)["nome"].tolist()
    except Exception:
        setores_all = []
    # remove setores técnicos
    setores_all = [s for s in setores_all if s and s.upper() not in ("ADMIN",)]
    if not setores_all:
        setores_all = ["GERAL"]

    c1, c2, c3 = st.columns([2,1,1])
    setores_sel = c1.multiselect("Setores", setores_all, default=setores_all, key="gest_setores")
    ano = int(c2.number_input("Ano", value=int(ano), step=1, key="gest_ano"))
    mes = int(c3.selectbox("Mês", list(range(1,13)), index=int(mes)-1, key="gest_mes"))

    if not setores_sel:
        st.warning("Selecione ao menos 1 setor.")
        return

    # Base: escala_mes
    q = """
        SELECT setor, chapa, dia, status
        FROM escala_mes
        WHERE ano=? AND mes=? AND setor IN ({})
    """.format(",".join(["?"]*len(setores_sel)))

    df = pd.read_sql_query(q, con, params=[ano, mes, *setores_sel])


    # --- Nomes (merge opcional com tabela colaboradores)
    try:
        qn = "SELECT setor, chapa, nome FROM colaboradores WHERE setor IN ({})".format(",".join(["?"]*len(setores_sel)))
        df_n = pd.read_sql_query(qn, con, params=[*setores_sel])
        df_n["chapa"] = df_n["chapa"].astype(str).str.strip()
        df["chapa"] = df["chapa"].astype(str).str.strip()
        df = df.merge(df_n.drop_duplicates(subset=["setor","chapa"]), on=["setor","chapa"], how="left")
    except Exception:
        df["nome"] = ""


    # Normalização de status
    df["status_norm"] = df["status"].fillna("").astype(str).str.strip().str.upper()
    # categorias
    is_fer = df["status_norm"].str.contains("F[ÉE]RIAS", regex=True)
    is_afa = df["status_norm"].isin(["AFA", "AFASTAMENTO"]) | df["status_norm"].str.contains("AFAST", regex=True)
    is_folga = df["status_norm"].str.contains("FOLG", regex=True) | df["status_norm"].isin(["FOLGA"])
    is_trab = ~(is_fer | is_afa | is_folga)

    df["cat"] = "TRABALHO"
    df.loc[is_folga, "cat"] = "FOLGA"
    df.loc[is_fer, "cat"] = "FÉRIAS"
    df.loc[is_afa, "cat"] = "AFASTAMENTO"

    # Resumo por setor
    pivot = (
        df.pivot_table(index="setor", columns="cat", values="dia", aggfunc="count", fill_value=0)
          .reset_index()
    )
    for col in ["TRABALHO","FOLGA","FÉRIAS","AFASTAMENTO"]:
        if col not in pivot.columns:
            pivot[col] = 0
    pivot["TOTAL_REGISTROS"] = pivot[["TRABALHO","FOLGA","FÉRIAS","AFASTAMENTO"]].sum(axis=1)

    st.subheader("Resumo por setor (mês)")
    st.dataframe(pivot.sort_values("setor"), use_container_width=True, hide_index=True)

    # Filtro detalhado
    st.subheader("Detalhe")
    sA, sB = st.columns([2,1])
    setor_det = sA.selectbox("Setor (detalhe)", setores_sel, key="gest_setor_det")
    modo = sB.selectbox("Visão", ["Por dia (contagem)", "Por colaborador (totais)"], key="gest_modo")

    df_det = df[df["setor"] == setor_det].copy()

    if modo.startswith("Por dia"):
        tabC, tabL = st.tabs(["📈 Contagem por dia", "👥 Listas do dia"])
        with tabC:
            by = df_det.groupby(["dia","cat"]).size().reset_index(name="qtd")
            piv = by.pivot_table(index="dia", columns="cat", values="qtd", fill_value=0).reset_index()
            for col in ["TRABALHO","FOLGA","FÉRIAS","AFASTAMENTO"]:
                if col not in piv.columns:
                    piv[col] = 0
            piv["TOTAL"] = piv[["TRABALHO","FOLGA","FÉRIAS","AFASTAMENTO"]].sum(axis=1)

            # Dia da semana (pt-br)
            DPT = {0:"Seg",1:"Ter",2:"Qua",3:"Qui",4:"Sex",5:"Sáb",6:"Dom"}
            piv["DIA_SEMANA"] = piv["dia"].apply(lambda d: DPT.get(dt.date(int(ano), int(mes), int(d)).weekday(), ""))
            piv = piv[["dia","DIA_SEMANA","TRABALHO","FOLGA","FÉRIAS","AFASTAMENTO","TOTAL"]]

            st.dataframe(piv.sort_values("dia"), use_container_width=True, hide_index=True)

        with tabL:
            last_day = calendar.monthrange(int(ano), int(mes))[1]
            dia_sel = st.selectbox("Dia para detalhar", list(range(1, last_day+1)), index=0, key="gest_dia_sel")
            dname = ["Segunda","Terça","Quarta","Quinta","Sexta","Sábado","Domingo"][dt.date(int(ano), int(mes), int(dia_sel)).weekday()]
            st.caption(f"Detalhe do dia **{dia_sel:02d}/{int(mes):02d}/{int(ano)}** — {dname}")

            df_day = df_det[df_det["dia"] == int(dia_sel)].copy()

            def _show_cat(title, cat, icon):
                sub = df_day[df_day["cat"] == cat].copy()
                sub["nome"] = sub.get("nome", "").fillna("")
                sub["status"] = sub.get("status", "").fillna("")
                sub = sub[["nome","chapa","status"]].rename(columns={"nome":"Nome","chapa":"Chapa","status":"Status"})
                st.markdown(f"#### {icon} {title} ({len(sub)})")
                st.dataframe(sub.sort_values(["Nome","Chapa"]), use_container_width=True, hide_index=True, height=280)

            cA, cB = st.columns(2)
            with cA:
                _show_cat("Trabalhando", "TRABALHO", "🟩")
                _show_cat("Férias", "FÉRIAS", "🟦")
            with cB:
                _show_cat("Folga", "FOLGA", "🟨")
                _show_cat("Afastamento", "AFASTAMENTO", "🟥")
    else:
        by = df_det.groupby(["chapa","cat"]).size().reset_index(name="qtd")
        piv = by.pivot_table(index="chapa", columns="cat", values="qtd", fill_value=0).reset_index()
        for col in ["TRABALHO","FOLGA","FÉRIAS","AFASTAMENTO"]:
            if col not in piv.columns:
                piv[col] = 0
        piv["TOTAL"] = piv[["TRABALHO","FOLGA","FÉRIAS","AFASTAMENTO"]].sum(axis=1)
        st.dataframe(piv.sort_values("TOTAL", ascending=False), use_container_width=True, hide_index=True)

    st.info("Dica: para o gerente, esta tela é a única exibida — as outras abas ficam ocultas para reduzir poluição visual.")




def _preview_cache_keys(setor: str, ano: int, mes: int):
    base = f"preview_cache::{setor}::{ano}::{mes}"
    return {
        "hist": base + "::hist",
        "cal": base + "::cal",
        "loaded": base + "::loaded",
    }

def _clear_preview_cache(setor: str, ano: int, mes: int):
    try:
        keys = _preview_cache_keys(setor, ano, mes)
        for k in keys.values():
            st.session_state.pop(k, None)
    except Exception:
        pass

def _ensure_preview_cache(setor: str, ano: int, mes: int, colaboradores: list):
    keys = _preview_cache_keys(setor, ano, mes)
    if keys["hist"] in st.session_state and keys["cal"] in st.session_state:
        return st.session_state[keys["hist"]], st.session_state[keys["cal"]]
    hist_db = get_hist_mes_com_overrides_cached(setor, int(ano), int(mes))
    colab_by = {c["Chapa"]: c for c in (colaboradores or [])}
    cal = calendario_rh_df(hist_db, colab_by) if hist_db else pd.DataFrame()
    st.session_state[keys["hist"]] = hist_db
    st.session_state[keys["cal"]] = cal
    st.session_state[keys["loaded"]] = True
    return hist_db, cal

def get_colaborador_record(setor: str, chapa: str):
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    con = db_conn()
    cur = con.cursor()
    cur.execute(
        """
        SELECT nome, chapa, subgrupo, entrada, folga_sab
        FROM colaboradores
        WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?
        LIMIT 1
        """,
        (setor, chapa),
    )
    row = cur.fetchone()
    con.close()
    if not row:
        return None
    return {
        "Nome": row[0],
        "Chapa": row[1],
        "Subgrupo": (row[2] or "").strip() or "SEM SUBGRUPO",
        "Entrada": (row[3] or "").strip() or "06:00",
        "Folga_Sab": bool(row[4]),
        "Setor": setor,
    }


@st.cache_data(show_spinner=False, ttl=120)
def get_subgrupo_competencia_ou_base(setor: str, chapa: str, ano: int, mes: int, base_subgrupo: str = "") -> str:
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    base_subgrupo = str(base_subgrupo or '').strip()

    # Regra correta para competência fechada/retificada:
    # 1) subgrupo_competencia do mês é a fonte oficial
    # 2) depois snapshot congelado
    # 3) depois base recebida / cadastro atual
    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute(
            """
            SELECT subgrupo
            FROM subgrupo_competencia
            WHERE UPPER(TRIM(setor))=? AND ano=? AND mes=? AND TRIM(chapa)=?
            LIMIT 1
            """,
            (setor, int(ano), int(mes), chapa),
        )
        row = cur.fetchone()
        if row and str(row[0] or '').strip():
            return str(row[0]).strip()
    finally:
        con.close()

    snap = get_colaborador_competencia_snapshot(setor, chapa, int(ano), int(mes))
    if snap and str(snap.get('Subgrupo') or '').strip():
        return str(snap.get('Subgrupo') or '').strip()

    if base_subgrupo:
        return base_subgrupo

    con = db_conn()
    cur = con.cursor()
    try:
        cur.execute(
            """
            SELECT subgrupo
            FROM colaboradores
            WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?
            LIMIT 1
            """,
            (setor, chapa),
        )
        row2 = cur.fetchone()
        return str((row2[0] if row2 else '') or '').strip() or 'SEM SUBGRUPO'
    finally:
        con.close()


def _apply_subgrupo_competencia_to_hist(setor: str, ano: int, mes: int, hist_db: dict[str, pd.DataFrame]):
    if not hist_db:
        return hist_db
    con = db_conn()
    try:
        df_sg = pd.read_sql_query(
            """
            SELECT chapa, subgrupo
            FROM subgrupo_competencia
            WHERE UPPER(TRIM(setor)) = UPPER(TRIM(?)) AND ano=? AND mes=?
            """,
            con,
            params=(setor, int(ano), int(mes)),
        )
    except Exception:
        try:
            con.close()
        except Exception:
            pass
        return hist_db
    finally:
        try:
            con.close()
        except Exception:
            pass
    if df_sg is None or df_sg.empty:
        return hist_db
    for _, r in df_sg.iterrows():
        ch = str(r.get('chapa') or '').strip()
        sg = str(r.get('subgrupo') or '').strip()
        if not ch or not sg or ch not in hist_db:
            continue
        df = hist_db[ch].copy()
        try:
            df.loc[:, 'Subgrupo'] = sg
        except Exception:
            pass
        hist_db[ch] = df
    return hist_db


@st.cache_data(show_spinner=False, ttl=120)
def get_escala_colaborador_mes(setor: str, chapa: str, ano: int, mes: int) -> pd.DataFrame:
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    hist_db = get_hist_mes_com_overrides_cached(setor, int(ano), int(mes)) or {}
    df_hist = hist_db.get(chapa)
    if df_hist is None:
        # compatibilidade extra caso alguma chave histórica tenha vindo sem normalização
        for k, v in (hist_db or {}).items():
            if _norm_chapa(k) == chapa:
                df_hist = v
                break
    if df_hist is not None and not df_hist.empty:
        out = df_hist.copy().reset_index(drop=True)

        # O histórico pode já vir com a coluna "Dia" contendo o dia da semana.
        # Renomeia antes de recriar a coluna numérica do dia do mês.
        if 'Dia' in out.columns:
            out = out.rename(columns={'Dia': 'Dia da semana'})
        elif 'dia_sem' in out.columns:
            out = out.rename(columns={'dia_sem': 'Dia da semana'})
        elif 'Dia da semana' not in out.columns:
            out['Dia da semana'] = ''

        out.insert(0, 'Dia', list(range(1, len(out) + 1)))

        if 'Data' in out.columns:
            try:
                out['Data'] = pd.to_datetime(out['Data'], errors='coerce')
            except Exception:
                pass
        else:
            out['Data'] = pd.NaT

        if 'Status' not in out.columns:
            out['Status'] = ''

        if 'H_Entrada' in out.columns and 'Entrada' not in out.columns:
            out = out.rename(columns={'H_Entrada': 'Entrada'})
        if 'H_Saida' in out.columns and 'Saída' not in out.columns:
            out = out.rename(columns={'H_Saida': 'Saída'})

        if 'Entrada' not in out.columns:
            out['Entrada'] = ''
        if 'Saída' not in out.columns:
            out['Saída'] = ''

        keep = ['Dia', 'Data', 'Dia da semana', 'Status', 'Entrada', 'Saída']
        for c in keep:
            if c not in out.columns:
                out[c] = ''
        return out[keep].copy()
    con = db_conn()
    try:
        df = pd.read_sql_query(
            """
            SELECT dia AS 'Dia', data AS 'Data', dia_sem AS 'Dia da semana', status AS 'Status',
                   COALESCE(h_entrada,'') AS 'Entrada', COALESCE(h_saida,'') AS 'Saída'
            FROM escala_mes
            WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=? AND ano=? AND mes=?
            ORDER BY dia ASC
            """,
            con,
            params=(setor, chapa, int(ano), int(mes)),
        )
    finally:
        con.close()
    return df
@st.cache_data(show_spinner=False, ttl=120)
def get_overrides_colaborador_mes(setor: str, chapa: str, ano: int, mes: int) -> pd.DataFrame:
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    con = db_conn()
    try:
        df_ov = pd.read_sql_query(
            """
            SELECT dia AS 'Dia', campo AS 'Campo', valor AS 'Novo valor', '' AS 'Motivo', '' AS 'Alterado em', id AS '_ord'
            FROM overrides
            WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=? AND ano=? AND mes=?
            """,
            con,
            params=(setor, chapa, int(ano), int(mes)),
        )
        df_ret = pd.read_sql_query(
            """
            SELECT dia AS 'Dia',
                   'retificacao' AS 'Campo',
                   TRIM(
                     COALESCE(NULLIF(novo_status,''), '') ||
                     CASE WHEN COALESCE(NULLIF(nova_entrada,''), '') <> '' THEN ' | Entrada: ' || nova_entrada ELSE '' END ||
                     CASE WHEN COALESCE(NULLIF(nova_saida,''), '') <> '' THEN ' | Saída: ' || nova_saida ELSE '' END ||
                     CASE WHEN COALESCE(NULLIF(novo_subgrupo,''), '') <> '' THEN ' | Subgrupo: ' || novo_subgrupo ELSE '' END
                   ) AS 'Novo valor',
                   COALESCE(motivo,'') AS 'Motivo',
                   COALESCE(criado_em,'') AS 'Alterado em',
                   (1000000 + id) AS '_ord'
            FROM retificacoes_competencia
            WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=? AND ano=? AND mes=?
            """,
            con,
            params=(setor, chapa, int(ano), int(mes)),
        )
    finally:
        con.close()
    df = pd.concat([df_ov, df_ret], ignore_index=True)
    if df is None or df.empty:
        return pd.DataFrame(columns=['Dia', 'Campo', 'Novo valor', 'Motivo', 'Alterado em', '_ord'])
    df = df.sort_values(['Dia', '_ord'], ascending=[False, False]).reset_index(drop=True)
    return df
def get_portal_version(setor: str, chapa: str, ano: int, mes: int) -> int:
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    con = db_conn()
    cur = con.cursor()
    cur.execute(
        "SELECT COUNT(*) FROM overrides WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=? AND ano=? AND mes=?",
        (setor, chapa, int(ano), int(mes)),
    )
    qtd = int((cur.fetchone() or [0])[0] or 0)
    con.close()
    return int(qtd + 1)


def get_assinatura_status(setor: str, chapa: str, ano: int, mes: int, tipo: str):
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    tipo = str(tipo or 'oficial').strip().lower()
    versao_atual = get_portal_version(setor, chapa, ano, mes)
    con = db_conn()
    cur = con.cursor()
    cur.execute(
        """
        SELECT versao_ref, assinado_em
        FROM portal_assinaturas
        WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=? AND ano=? AND mes=? AND tipo=?
        LIMIT 1
        """,
        (setor, chapa, int(ano), int(mes), tipo),
    )
    row = cur.fetchone()
    con.close()
    if not row:
        return {"status": "Pendente", "versao": versao_atual, "assinado_em": None}
    versao_ref, assinado_em = int(row[0] or 0), row[1]
    status = "Assinado" if versao_ref == versao_atual else "Reassinatura pendente"
    return {"status": status, "versao": versao_atual, "assinado_em": assinado_em, "versao_assinada": versao_ref}


def salvar_assinatura_portal(setor: str, chapa: str, ano: int, mes: int, tipo: str):
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    tipo = str(tipo or 'oficial').strip().lower()
    versao_ref = get_portal_version(setor, chapa, ano, mes)
    agora = datetime.now().isoformat()
    con = db_conn()
    cur = con.cursor()
    cur.execute(
        """
        INSERT INTO portal_assinaturas(setor, chapa, ano, mes, tipo, versao_ref, assinado_em)
        VALUES(?,?,?,?,?,?,?)
        ON CONFLICT(setor, chapa, ano, mes, tipo)
        DO UPDATE SET versao_ref=excluded.versao_ref, assinado_em=excluded.assinado_em
        """,
        (setor, chapa, int(ano), int(mes), tipo, int(versao_ref), agora),
    )
    con.commit()
    con.close()


def list_solicitacoes_colaborador(setor: str, chapa: str) -> pd.DataFrame:
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    con = db_conn()
    try:
        df = pd.read_sql_query(
            """
            SELECT id AS 'ID', data_solicitada AS 'Data solicitada', tipo AS 'Tipo',
                   COALESCE(motivo,'') AS 'Motivo', COALESCE(observacao,'') AS 'Observação',
                   status AS 'Status', criado_em AS 'Criado em', atualizado_em AS 'Atualizado em'
            FROM portal_solicitacoes_folga
            WHERE UPPER(TRIM(setor))=? AND TRIM(chapa)=?
            ORDER BY datetime(criado_em) DESC, id DESC
            """,
            con,
            params=(setor, chapa),
        )
    finally:
        con.close()
    return df


def criar_solicitacao_folga(setor: str, chapa: str, data_solicitada: str, tipo: str, motivo: str, observacao: str):
    setor = _norm_setor(setor)
    chapa = _norm_chapa(chapa)
    agora = datetime.now().isoformat()
    con = db_conn()
    cur = con.cursor()
    cur.execute(
        """
        INSERT INTO portal_solicitacoes_folga(setor, chapa, data_solicitada, tipo, motivo, observacao, status, criado_em, atualizado_em)
        VALUES(?,?,?,?,?,?,?,?,?)
        """,
        (setor, chapa, str(data_solicitada), str(tipo), str(motivo or '').strip(), str(observacao or '').strip(), 'Em análise', agora, agora),
    )
    con.commit()
    con.close()


def _norm_subgrupo_label(v: str) -> str:
    s = str(v or '').strip().upper()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def colaborador_eh_lideranca(setor: str, chapa: str) -> bool:
    rec = get_colaborador_record(setor, chapa)
    if not rec:
        return False

    sg = _norm_subgrupo_label((rec or {}).get('Subgrupo', ''))
    perfil = _norm_subgrupo_label((rec or {}).get('Perfil', ''))
    lider = _norm_subgrupo_label((rec or {}).get('Lider', ''))

    if sg in ['LIDERANCA', 'LIDER', 'LIDERES', 'LIDERANCA FLV']:
        return True

    if perfil in ['LIDER', 'LIDERANCA', 'ENCARREGADO', 'RESPONSAVEL']:
        return True

    if lider in ['SIM', 'S', '1', 'TRUE']:
        return True

    return False

def list_assinaturas_setor(setor: str, ano: int, mes: int) -> pd.DataFrame:
    setor = _norm_setor(setor)
    con = db_conn()
    q = pd.read_sql_query(
        """
        SELECT p.setor AS Setor, p.chapa AS Chapa, COALESCE(c.nome, p.chapa) AS Nome,
               COALESCE(c.subgrupo, 'SEM SUBGRUPO') AS Subgrupo,
               p.ano AS Ano, p.mes AS Mes, p.tipo AS Tipo, p.versao_ref AS 'Versão', p.assinado_em AS Assinado_em
        FROM portal_assinaturas p
        LEFT JOIN colaboradores c
          ON UPPER(TRIM(c.setor)) = UPPER(TRIM(p.setor))
         AND TRIM(c.chapa) = TRIM(p.chapa)
        WHERE UPPER(TRIM(p.setor)) = ? AND p.ano = ? AND p.mes = ?
        ORDER BY p.tipo, COALESCE(c.nome, p.chapa)
        """,
        con, params=(setor, int(ano), int(mes))
    )
    con.close()
    return q


def list_assinaturas_setor_todas(setor: str) -> pd.DataFrame:
    setor = _norm_setor(setor)
    con = db_conn()
    q = pd.read_sql_query(
        """
        SELECT p.setor AS Setor, p.chapa AS Chapa, COALESCE(c.nome, p.chapa) AS Nome,
               COALESCE(c.subgrupo, 'SEM SUBGRUPO') AS Subgrupo,
               p.ano AS Ano, p.mes AS Mes, p.tipo AS Tipo, p.versao_ref AS 'Versão', p.assinado_em AS Assinado_em
        FROM portal_assinaturas p
        LEFT JOIN colaboradores c
          ON UPPER(TRIM(c.setor)) = UPPER(TRIM(p.setor))
         AND TRIM(c.chapa) = TRIM(p.chapa)
        WHERE UPPER(TRIM(p.setor)) = ?
        ORDER BY p.ano DESC, p.mes DESC, p.tipo, COALESCE(c.nome, p.chapa)
        """,
        con, params=(setor,)
    )
    con.close()
    return q

def list_solicitacoes_setor(setor: str) -> pd.DataFrame:
    setor = _norm_setor(setor)
    con = db_conn()
    q = pd.read_sql_query(
        """
        SELECT s.id AS ID, s.setor AS Setor, s.chapa AS Chapa, COALESCE(c.nome, s.chapa) AS Nome,
               COALESCE(c.subgrupo, 'SEM SUBGRUPO') AS Subgrupo,
               s.data_solicitada AS Data, s.tipo AS Tipo, s.motivo AS Motivo, s.observacao AS 'Observação',
               s.status AS Status, s.criado_em AS Criado_em, s.atualizado_em AS Atualizado_em
        FROM portal_solicitacoes_folga s
        LEFT JOIN colaboradores c
          ON UPPER(TRIM(c.setor)) = UPPER(TRIM(s.setor))
         AND TRIM(c.chapa) = TRIM(s.chapa)
        WHERE UPPER(TRIM(s.setor)) = ?
        ORDER BY s.criado_em DESC, s.id DESC
        """,
        con, params=(setor,)
    )
    con.close()
    return q

def atualizar_status_solicitacao(solicitacao_id: int, novo_status: str):
    agora = datetime.now().isoformat()
    con = db_conn()
    cur = con.cursor()
    cur.execute(
        'UPDATE portal_solicitacoes_folga SET status=?, atualizado_em=? WHERE id=?',
        (str(novo_status), agora, int(solicitacao_id)),
    )
    con.commit()
    con.close()

def page_portal_colaborador(auth: dict, ano_cfg: int, mes_cfg: int):
    setor = _norm_setor(auth.get('setor', ''))
    chapa = _norm_chapa(auth.get('chapa', ''))
    nome = auth.get('nome', '-')
    colab = get_colaborador_competencia_snapshot(setor, chapa, int(ano_cfg or datetime.now().year), int(mes_cfg or datetime.now().month)) or get_colaborador_record(setor, chapa) or {
        'Nome': nome,
        'Chapa': chapa,
        'Subgrupo': 'SEM SUBGRUPO',
        'Entrada': '06:00',
        'Folga_Sab': False,
        'Setor': setor,
    }

    hoje = datetime.now()
    ano_ref = int(ano_cfg or hoje.year)
    mes_ref = int(mes_cfg or hoje.month)
    if mes_ref < 1 or mes_ref > 12:
        mes_ref = int(hoje.month)
    prox_mes = mes_ref + 1
    prox_ano = ano_ref
    if prox_mes > 12:
        prox_mes = 1
        prox_ano += 1

    st.markdown("### 👤 Portal do Colaborador — v0.01 beta premium")
    i1, i2, i3, i4 = st.columns(4)
    i1.info(f"**Nome**\n\n{colab.get('Nome','-')}")
    i2.info(f"**Setor**\n\n{setor}")
    subgrupo_portal = get_subgrupo_competencia_ou_base(setor, chapa, ano_ref, mes_ref, colab.get('Subgrupo','SEM SUBGRUPO'))
    i3.info(f"**Subgrupo**\n\n{subgrupo_portal}")
    i4.info(f"**Chapa**\n\n{chapa}")

    st.caption(
        f"Portal do colaborador mostrando a competência de referência {mes_ref:02d}/{ano_ref} e a pré-escala do próximo mês ({prox_mes:02d}/{prox_ano})."
    )

    df_oficial = get_escala_colaborador_mes(setor, chapa, ano_ref, mes_ref)
    df_pre = get_escala_colaborador_mes(setor, chapa, prox_ano, prox_mes)
    hist = get_overrides_colaborador_mes(setor, chapa, ano_ref, mes_ref)
    ass_escala = get_assinatura_status(setor, chapa, ano_ref, mes_ref, 'oficial')
    ass_mud = get_assinatura_status(setor, chapa, ano_ref, mes_ref, 'historico')

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        '📋 Escala Oficial',
        '🕒 Pré-Escala',
        '📝 Histórico de Mudanças',
        '✍️ Assinaturas',
        '🏖️ Férias',
        '⚙️ Ajustes',
    ])

    with tab1:
        st.markdown(f"#### Escala oficial — {mes_ref:02d}/{ano_ref}")
        c1, c2, c3 = st.columns(3)
        c1.metric('Versão atual', ass_escala.get('versao', 1))
        c2.metric('Status da assinatura', ass_escala.get('status', 'Pendente'))
        c3.metric('Mudanças registradas', len(hist) if hasattr(hist, '__len__') else 0)
        if ass_escala.get('assinado_em'):
            st.caption(f"Última assinatura da escala: {ass_escala.get('assinado_em')}")
        try:
            df_ret_mes = load_retificacoes_competencia(setor, ano_ref, mes_ref)
            if df_ret_mes is not None and not df_ret_mes.empty:
                st.warning('⚠️ Esta escala possui retificações após o fechamento.')
        except Exception:
            pass
        if df_oficial.empty:
            st.info('Ainda não há escala oficial carregada para o mês vigente.')
        else:
            st.dataframe(df_oficial, use_container_width=True, hide_index=True)
            if ass_escala.get('status') == 'Assinado':
                st.success('Escala do mês vigente já assinada. Botão ocultado automaticamente.')
            else:
                st.info('A assinatura da escala do mês vigente fica na aba Assinaturas.')

    with tab2:
        st.markdown(f"#### Pré-escala — {prox_mes:02d}/{prox_ano}")
        st.warning('Prévia do próximo mês. Ainda não é oficial, não pode ser assinada e pode ser alterada até a liberação do líder.')
        if df_pre.empty:
            st.info('Ainda não há pré-escala disponível para o próximo mês.')
        else:
            st.dataframe(df_pre, use_container_width=True, hide_index=True)
            st.caption('Assinatura bloqueada até o início do mês vigente correspondente.')

    with tab3:
        st.markdown(f"#### Histórico de mudanças — {mes_ref:02d}/{ano_ref}")
        m1, m2 = st.columns(2)
        m1.metric('Mudanças registradas no mês vigente', len(hist) if hasattr(hist, '__len__') else 0)
        m2.metric('Status do aceite das mudanças', ass_mud.get('status', 'Pendente'))
        if ass_mud.get('assinado_em'):
            st.caption(f"Último aceite das mudanças: {ass_mud.get('assinado_em')}")
        if hist.empty:
            st.info('Nenhuma mudança registrada no mês vigente para esta chapa.')
        else:
            hist_view = hist.copy()
            keep_cols = [c for c in ['Dia', 'Campo', 'Valor anterior', 'Novo valor', 'Motivo', 'Alterado em'] if c in hist_view.columns]
            if not keep_cols:
                keep_cols = list(hist_view.columns)
            st.dataframe(hist_view[keep_cols], use_container_width=True, hide_index=True)
            st.caption('A assinatura dessas mudanças fica concentrada na aba Assinaturas.')

    with tab4:
        st.markdown(f"#### Assinaturas — {mes_ref:02d}/{ano_ref}")
        sub1, sub2 = st.tabs(['🗓️ Assinatura da Escala do Mês', '🔁 Assinatura de Mudanças'])

        with sub1:
            a1, a2, a3 = st.columns(3)
            a1.metric('Competência', f'{mes_ref:02d}/{ano_ref}')
            a2.metric('Versão atual', ass_escala.get('versao', 1))
            a3.metric('Status', ass_escala.get('status', 'Pendente'))
            if ass_escala.get('assinado_em'):
                st.caption(f"Assinada em: {ass_escala.get('assinado_em')}")
            if df_oficial.empty:
                st.info('Sem escala oficial carregada para o mês vigente.')
            elif ass_escala.get('status') == 'Assinado':
                st.success('Escala do mês vigente já assinada. Botão ocultado.')
            else:
                st.warning('Assine aqui somente a escala do mês vigente.')
                if st.button('✅ Assinar escala do mês', key=f'ass_oficial_{setor}_{chapa}_{ano_ref}_{mes_ref}'):
                    salvar_assinatura_portal(setor, chapa, ano_ref, mes_ref, 'oficial')
                    st.success('Escala do mês assinada com sucesso.')
                    st.rerun()

        with sub2:
            b1, b2, b3 = st.columns(3)
            b1.metric('Mudanças registradas', len(hist) if hasattr(hist, '__len__') else 0)
            b2.metric('Versão das mudanças', ass_mud.get('versao', 1))
            b3.metric('Status', ass_mud.get('status', 'Pendente'))
            if ass_mud.get('assinado_em'):
                st.caption(f"Mudanças assinadas em: {ass_mud.get('assinado_em')}")
            if hist.empty:
                st.info('Não existem mudanças de horários, folgas ou ajustes no mês vigente.')
            else:
                hist_ass = hist.copy()
                hist_ass['Status da assinatura'] = ass_mud.get('status', 'Pendente')
                keep_cols = [c for c in ['Dia', 'Campo', 'Valor anterior', 'Novo valor', 'Motivo', 'Status da assinatura'] if c in hist_ass.columns]
                if not keep_cols:
                    keep_cols = list(hist_ass.columns)
                st.dataframe(hist_ass[keep_cols], use_container_width=True, hide_index=True)
                if ass_mud.get('status') == 'Assinado':
                    st.success('Mudanças do mês vigente já assinadas. Botão ocultado.')
                else:
                    if st.button('✍️ Assinar mudanças de horários e folgas', key=f'ass_hist_{setor}_{chapa}_{ano_ref}_{mes_ref}'):
                        salvar_assinatura_portal(setor, chapa, ano_ref, mes_ref, 'historico')
                        st.success('Mudanças do mês vigente assinadas com sucesso.')
                        st.rerun()

    with tab5:
        st.markdown(f"#### Férias — {mes_ref:02d}/{ano_ref}")
        rows_fer = [r for r in (list_ferias(setor) or []) if _norm_chapa(r[0]) == chapa]
        if not rows_fer:
            st.info('Nenhuma férias cadastrada para este colaborador.')
        else:
            dados_fer = []
            for _ch, ini, fim in rows_fer:
                try:
                    dt_ini = pd.to_datetime(str(ini)).date()
                    dt_fim = pd.to_datetime(str(fim)).date()
                    qtd = (dt_fim - dt_ini).days + 1
                except Exception:
                    dt_ini = str(ini)
                    dt_fim = str(fim)
                    qtd = None
                dados_fer.append({
                    'Chapa': _ch,
                    'Início': dt_ini.strftime('%d/%m/%Y') if hasattr(dt_ini, 'strftime') else dt_ini,
                    'Fim': dt_fim.strftime('%d/%m/%Y') if hasattr(dt_fim, 'strftime') else dt_fim,
                    'Dias': qtd if qtd is not None else '-',
                })
            df_fer = pd.DataFrame(dados_fer)
            fer_ativa = False
            hoje_date = hoje.date()
            for r in rows_fer:
                try:
                    ini_d = pd.to_datetime(str(r[1])).date()
                    fim_d = pd.to_datetime(str(r[2])).date()
                    if ini_d <= hoje_date <= fim_d:
                        fer_ativa = True
                        break
                except Exception:
                    pass
            cfa, cfb = st.columns(2)
            cfa.metric('Períodos cadastrados', len(df_fer))
            cfb.metric('Situação atual', 'Em férias' if fer_ativa else 'Programadas / encerradas')
            st.dataframe(df_fer, use_container_width=True, hide_index=True)
            st.caption('As férias exibidas aqui são somente do colaborador logado.')

    with tab6:
        st.markdown('#### Ajustes')
        suba, subb = st.tabs(['🌴 Sugestão de Folgas', '📨 Minhas solicitações'])

        with suba:
            st.caption('Envie sugestões de folgas e ajustes de forma organizada para o líder analisar.')
            c1, c2 = st.columns(2)
            data_padrao = hoje.date()
            data_sol = c1.date_input('Data desejada', value=data_padrao, key=f'data_folga_{setor}_{chapa}')
            tipo_sol = c2.selectbox('Tipo da sugestão', ['Folga', 'Troca de plantão', 'Ajuste de horário'], key=f'tipo_folga_{setor}_{chapa}')
            motivo = st.text_input('Motivo', key=f'motivo_folga_{setor}_{chapa}')
            observacao = st.text_area('Observação', key=f'obs_folga_{setor}_{chapa}')
            if st.button('📨 Enviar sugestão', key=f'env_folga_{setor}_{chapa}'):
                criar_solicitacao_folga(setor, chapa, str(data_sol), tipo_sol, motivo, observacao)
                st.success('Sugestão enviada para análise.')
                st.rerun()

        with subb:
            df_sol = list_solicitacoes_colaborador(setor, chapa)
            if df_sol.empty:
                st.info('Nenhuma sugestão enviada até agora.')
            else:
                st.dataframe(df_sol, use_container_width=True, hide_index=True)

def page_app():
    auth = st.session_state.get("auth") or {}
    setor = auth.get("setor", "GERAL")

    if st.session_state.get("auth_force_change", False):
        st.markdown("## 🔐 Troca obrigatória de senha")
        st.warning("Sua senha temporária precisa ser trocada antes de continuar.")
        nova1 = st.text_input("Nova senha", type="password", key="force_pwd_1")
        nova2 = st.text_input("Confirmar nova senha", type="password", key="force_pwd_2")
        c1, c2 = st.columns([1,1])
        if c1.button("Salvar nova senha", key="force_pwd_save"):
            if not (nova1 or "").strip():
                st.error("Digite a nova senha.")
                st.stop()
            if nova1 != nova2:
                st.error("A confirmação da senha não confere.")
                st.stop()
            try:
                update_password(setor, auth.get("chapa", ""), nova1)
                set_force_change_password(setor, auth.get("chapa", ""), False)
                st.session_state["auth_force_change"] = False
                if st.session_state.get("auth"):
                    st.session_state["auth"]["forcar_troca_senha"] = False
                st.success("Senha atualizada com sucesso.")
                st.rerun()
            except Exception as e:
                st.error(f"Falha ao atualizar senha: {e}")
        if c2.button("Sair", key="force_pwd_logout"):
            st.session_state["auth"] = None
            st.session_state["auth_force_change"] = False
            st.rerun()
        return

    # ---- Competência (mês/ano) compartilhada
    ano_cfg = int(st.session_state.get("cfg_ano", datetime.now().year))
    mes_cfg = int(st.session_state.get("cfg_mes", datetime.now().month))
    st.session_state["cfg_ano"] = ano_cfg
    st.session_state["cfg_mes"] = mes_cfg

    if "last_seed" not in st.session_state:
        st.session_state["last_seed"] = 0

    # =========================
    # SIDEBAR — Sessão + Competência
    # =========================
    with st.sidebar:
        st.title("👤 Sessão")
        st.caption("Acesso por setor (usuário / líder / admin)")
        st.caption(VERSAO_ACESSO_LIDER)

        _ano_sb = int(st.session_state.get('cfg_ano') or datetime.now().year)
        _mes_sb = int(st.session_state.get('cfg_mes') or datetime.now().month)
        _colab_sb = get_colaborador_competencia_snapshot(setor, auth.get('chapa',''), _ano_sb, _mes_sb) or get_colaborador_record(setor, auth.get('chapa',''))
        _subgrupo_auth = get_subgrupo_competencia_ou_base(setor, auth.get('chapa',''), _ano_sb, _mes_sb, (_colab_sb or {}).get('Subgrupo', 'SEM SUBGRUPO'))
        _lideranca_ok = bool(auth.get('is_lider', False)) or bool(auth.get('is_ax_lider', False)) or colaborador_eh_lideranca(setor, auth.get('chapa',''))
        _perfil_gestao = bool(auth.get('is_admin', False)) or _lideranca_ok

        cA, cB = st.columns([1, 1])
        perfil_label = 'ADMIN' if auth.get('is_admin', False) else ('AX LÍDER' if auth.get('is_ax_lider', False) else ('LÍDER' if _lideranca_ok else 'COLABORADOR'))
        cA.write(f"**Nome:** {auth.get('nome','-')}")
        cB.write(f"**Perfil:** {perfil_label}")

        st.write(f"**Setor:** {setor}")
        st.write(f"**Chapa:** {auth.get('chapa','-')}")
        st.write(f"**Subgrupo:** {_subgrupo_auth}")
        if bool(auth.get('is_lider', False)) and not _lideranca_ok and not bool(auth.get('is_admin', False)):
            st.warning('Perfil líder liberado somente para colaborador do subgrupo LIDERANÇA neste setor.')

        st.markdown("<div class='hr'></div>", unsafe_allow_html=True)

        st.subheader("🗓️ Competência")
        if not _perfil_gestao:
            hoje = datetime.now()
            st.session_state["cfg_mes"] = int(hoje.month)
            st.session_state["cfg_ano"] = int(hoje.year)
            st.write(f"**Mês vigente:** {hoje.month:02d}")
            st.write(f"**Ano vigente:** {hoje.year}")
            prox_mes = hoje.month + 1
            prox_ano = hoje.year
            if prox_mes > 12:
                prox_mes = 1
                prox_ano += 1
            st.write(f"**Pré-escala:** {prox_mes:02d}/{prox_ano}")
        else:
            m1, m2 = st.columns(2)
            mes_cfg = m1.selectbox("Mês", list(range(1, 13)), index=mes_cfg - 1, key="sb_mes")
            ano_cfg = m2.number_input("Ano", value=ano_cfg, step=1, key="sb_ano")
            st.session_state["cfg_mes"] = int(mes_cfg)
            st.session_state["cfg_ano"] = int(ano_cfg)

        st.markdown("<div class='hr'></div>", unsafe_allow_html=True)

        if st.button("🚪 Sair", use_container_width=True, key="logout_btn"):
            st.session_state["auth"] = None
            st.rerun()

    # =========================
    # PERFIL GESTÃO (GERENTE) — UI dedicada
    # =========================
    if str(setor).strip().upper() in ("GESTAO", "GERENCIA", "GERENTE"):
        page_gestao_dashboard(int(st.session_state["cfg_ano"]), int(st.session_state["cfg_mes"]))
        return

    _lideranca_ok = bool(auth.get('is_lider', False)) or bool(auth.get('is_ax_lider', False)) or colaborador_eh_lideranca(setor, auth.get('chapa',''))
    _perfil_gestao = bool(auth.get('is_admin', False)) or _lideranca_ok

    if not _perfil_gestao:
        page_portal_colaborador(auth, int(st.session_state["cfg_ano"]), int(st.session_state["cfg_mes"]))
        return

    # =========================
    # KPIs
    # =========================
    ano_k = int(st.session_state["cfg_ano"])
    mes_k = int(st.session_state["cfg_mes"])

    _kpi = get_kpis_cached(setor, ano_k, mes_k)
    total_colab = int(_kpi.get("total_colab", 0))
    folgas_mes = int(_kpi.get("folgas_mes", 0))
    ferias_mes = int(_kpi.get("ferias_mes", 0))
    trabalhos_mes = int(_kpi.get("trabalhos_mes", 0))

    k1, k2, k3, k4 = st.columns(4)
    k1.markdown(
        f"<div class='kpi-card'><div class='kpi-title'>Colaboradores</div>"
        f"<p class='kpi-value'>{total_colab}</p></div>",
        unsafe_allow_html=True
    )
    k2.markdown(
        f"<div class='kpi-card'><div class='kpi-title'>Dias de Folga (mês)</div>"
        f"<p class='kpi-value'>{folgas_mes}</p></div>",
        unsafe_allow_html=True
    )
    k3.markdown(
        f"<div class='kpi-card'><div class='kpi-title'>Dias de Férias (mês)</div>"
        f"<p class='kpi-value'>{ferias_mes}</p></div>",
        unsafe_allow_html=True
    )
    k4.markdown(
        f"<div class='kpi-card'><div class='kpi-title'>Dias de Trabalho (mês)</div>"
        f"<p class='kpi-value'>{trabalhos_mes}</p></div>",
        unsafe_allow_html=True
    )

    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)

    # =========================
    # ABAS
    # =========================
    tabs = ["👥 Colaboradores", "🚀 Gerar Escala", "⚙️ Ajustes", "🏖️ Férias", "🖨️ Impressão", "✍️ Assinaturas", "📨 Minhas solicitações"]
    is_admin_area = bool(auth.get("is_admin", False)) and setor == "ADMIN"
    if is_admin_area:
        tabs = ["🔒 Admin"]

    sec_main = st.radio("Navegação", tabs, horizontal=True, key="main_nav_radio_ultra_fast")

    # ------------------------------------------------------
    # ABA 1: Colaboradores
    # ------------------------------------------------------
    if sec_main == "👥 Colaboradores":
        sec_col = st.radio(
            "",
            (["👥 Colaboradores", "➕ Cadastrar colaborador", "🗑️ Excluir colaborador", "✏️ Editar perfil", "🔑 Alterar senha colaborador", "🛠️ Atualizar funcionário (AX/Líder)", "🧾 Aprovações AX"] + (["🔄 Rodízio Caixa"] if str(setor).strip().upper() == "FRENTECAIXA" else [])), 
            horizontal=True,
            key="sec_col_radio_real_speed",
            label_visibility="collapsed",
        )

        if sec_col == "👥 Colaboradores":
            st.markdown("### 👥 Colaboradores")
            colaboradores = load_colaboradores_setor(setor)
            if colaboradores:
                df_col = pd.DataFrame([{
                    "Nome": c["Nome"],
                    "Chapa": c["Chapa"],
                    "Subgrupo": c["Subgrupo"] or "SEM SUBGRUPO",
                    "Entrada": c["Entrada"],
                    "Folga Sábado": "Sim" if c["Folga_Sab"] else "Não",
                } for c in colaboradores])

                cbus1, cbus2, cbus3 = st.columns([2, 1, 1])
                termo = cbus1.text_input("Buscar nome/chapa/subgrupo", key="col_busca_fast")
                tam_pagina = cbus2.selectbox("Por página", [10, 15, 20, 30, 50], index=1, key="col_page_size_fast")

                if termo:
                    termo_n = str(termo).strip().lower()
                    mask = (
                        df_col["Nome"].astype(str).str.lower().str.contains(termo_n, na=False)
                        | df_col["Chapa"].astype(str).str.lower().str.contains(termo_n, na=False)
                        | df_col["Subgrupo"].astype(str).str.lower().str.contains(termo_n, na=False)
                    )
                    df_view = df_col.loc[mask].reset_index(drop=True)
                else:
                    df_view = df_col.reset_index(drop=True)

                total_regs = len(df_view)
                total_pag = max(1, (total_regs + int(tam_pagina) - 1) // int(tam_pagina))
                pagina = cbus3.number_input("Página", min_value=1, max_value=total_pag, value=1, step=1, key="col_page_fast")
                ini = (int(pagina) - 1) * int(tam_pagina)
                fim = ini + int(tam_pagina)
                st.caption(f"Mostrando {min(total_regs, 0 if total_regs == 0 else ini + 1)}–{min(total_regs, fim)} de {total_regs} registro(s).")
                st.dataframe(df_view.iloc[ini:fim], use_container_width=True, height=420)
            else:
                st.info("Sem colaboradores.")

            st.markdown("---")

        elif sec_col == "➕ Cadastrar colaborador":
            colaboradores = load_colaboradores_setor(setor)
            st.markdown("## ➕ Cadastrar colaborador (perfil completo + folgas do mês)")

            ano_cfg = int(st.session_state.get("cfg_ano", datetime.now().year))
            mes_cfg = int(st.session_state.get("cfg_mes", datetime.now().month))
            ndias_cfg = calendar.monthrange(ano_cfg, mes_cfg)[1]

            with st.form("form_add_colaborador", clear_on_submit=True):
                c1, c2 = st.columns(2)
                nome_n = c1.text_input("Nome:", key="col_nome")
                chapa_n = c2.text_input("Chapa:", key="col_chapa")

                c3, c4, c5 = st.columns([1.2, 1.2, 1])
                sg_opts_new = [""] + list_subgrupos(setor)
                subgrupo_n = c3.selectbox("Subgrupo:", sg_opts_new, index=0, key="col_subgrupo")
                entrada_n = c4.selectbox("Entrada:", HORARIOS_ENTRADA_PRESET, index=HORARIOS_ENTRADA_PRESET.index("06:00") if "06:00" in HORARIOS_ENTRADA_PRESET else 0, key="col_entrada")
                folga_sab_n = c5.checkbox("Permitir folga sábado", value=False, key="col_folga_sab")

                st.caption(f"Folgas do mês para já salvar como **Folga** (competência ativa: {mes_cfg:02d}/{ano_cfg}).")
                dias_folga = st.multiselect(
                    "Selecione os dias de folga (1..31):",
                    options=list(range(1, ndias_cfg + 1)),
                    default=[],
                    key="col_dias_folga",
                )

                submitted = st.form_submit_button("Cadastrar colaborador", use_container_width=True)

                if submitted:
                    if not nome_n or not chapa_n:
                        st.error("Preencha nome e chapa.")
                    elif colaborador_exists(setor, chapa_n.strip()):
                        st.error("Já existe essa chapa.")
                    else:
                        ch_new = chapa_n.strip()
                        create_colaborador(nome_n.strip(), setor, ch_new, subgrupo=subgrupo_n, entrada=entrada_n, folga_sab=folga_sab_n)
                        for d in dias_folga:
                            set_override(setor, ano_cfg, mes_cfg, ch_new, int(d), "status", "Folga")

                        st.success("Cadastrado! (perfil + folgas do mês salvos)")
                        st.rerun()

            st.markdown("---")

        elif sec_col == "🗑️ Excluir colaborador":
            colaboradores = load_colaboradores_setor(setor)
            st.markdown("## 🗑️ Excluir colaborador")
            if colaboradores:
                opts = []
                for c in colaboradores:
                    ch = str(c.get("Chapa","")).strip()
                    nm = str(c.get("Nome","") or "").strip()
                    label = f"{ch} — {nm}" if nm else ch
                    opts.append((label, ch))
                pick = st.selectbox("Escolha a chapa para excluir:", [o[0] for o in opts], key="del_chapa_label")
                ch_del = next((o[1] for o in opts if o[0] == pick), pick.split("—")[0].strip())
                st.warning("⚠️ Excluir remove também férias, ajustes, escala e estado desse colaborador no setor.")
                confirm = st.checkbox("Confirmo que quero excluir definitivamente", key="del_confirm")
                if st.button("Excluir colaborador", key="del_btn"):
                    if not confirm:
                        st.error("Marque a confirmação para excluir.")
                    else:
                        delete_colaborador_total(setor, ch_del)
                        st.success("Colaborador excluído!")
                        st.rerun()

            st.markdown("---")

        elif sec_col == "✏️ Editar perfil":
            colaboradores = load_colaboradores_setor(setor)
            st.markdown("## ✏️ Editar perfil do colaborador")
            if colaboradores:
                chapas = [c["Chapa"] for c in colaboradores]
                nome_by_chapa = {c["Chapa"]: c.get("Nome", "") for c in colaboradores}
                ch_sel = st.selectbox(
                    "Colaborador (Nome — Chapa):",
                    chapas,
                    key="pf_chapa",
                    format_func=lambda ch: f"{(nome_by_chapa.get(ch, ch) or ch)} — {ch}",
                )
                csel = next(x for x in colaboradores if x["Chapa"] == ch_sel)

                last = st.session_state.get("pf_last_chapa")
                if last != ch_sel:
                    _ent_atual = (csel.get("Entrada") or BALANCO_DIA_ENTRADA).strip()
                    st.session_state["pf_ent_sel"] = _ent_atual

                    _sg = (csel.get("Subgrupo") or "").strip()
                    _sg_opts = [""] + list_subgrupos(setor)
                    st.session_state["pf_sg"] = _sg if _sg in _sg_opts else ""

                    st.session_state["pf_sab"] = bool(csel.get("Folga_Sab"))
                    st.session_state["pf_last_chapa"] = ch_sel

                if st.session_state.get("pf_last_chapa_edit") != ch_sel:
                    st.session_state["pf_chapa_edit"] = ch_sel
                    st.session_state["pf_nome_edit"] = (csel.get("Nome") or "").strip()
                    st.session_state["pf_last_chapa_edit"] = ch_sel

                colp0, colp1 = st.columns(2)
                nome_edit = colp0.text_input("Nome:", key="pf_nome_edit")
                chapa_edit = colp1.text_input("Chapa:", key="pf_chapa_edit")

                ent_atual = (csel.get("Entrada") or BALANCO_DIA_ENTRADA).strip()
                opcoes_ent = list(HORARIOS_ENTRADA_PRESET)
                if ent_atual and ent_atual not in opcoes_ent:
                    opcoes_ent = opcoes_ent + [ent_atual]

                colp2, colp3, colp4 = st.columns(3)
                ent_sel = colp2.selectbox("Entrada:", options=opcoes_ent, key="pf_ent_sel")

                sg_opts = [""] + list_subgrupos(setor)
                sg_atual = (csel.get("Subgrupo") or "").strip()
                if sg_atual and sg_atual not in sg_opts:
                    sg_opts.append(sg_atual)
                sg = colp3.selectbox("Subgrupo:", sg_opts, key="pf_sg")
                sab = colp4.checkbox("Permitir folga sábado", key="pf_sab")

                if st.button("Salvar perfil", key="pf_save"):
                    if not (nome_edit or "").strip():
                        st.error("Preencha o nome.")
                    elif not (chapa_edit or "").strip():
                        st.error("Preencha a chapa.")
                    else:
                        try:
                            update_colaborador_perfil(setor, ch_sel, chapa_edit, nome_edit, sg, ent_sel, sab)
                            st.success("Salvo!")
                            st.rerun()
                        except Exception as e:
                            st.error(str(e))

                st.markdown("---")

        elif sec_col == "🔑 Alterar senha colaborador":
            colaboradores = load_colaboradores_setor(setor)
            st.markdown("## 🔑 Alterar senha colaborador")
            if colaboradores:
                chapas = [c["Chapa"] for c in colaboradores]
                nome_by_chapa = {c["Chapa"]: c.get("Nome", "") for c in colaboradores}
                ch_sel_pwd = st.selectbox(
                    "Colaborador (Nome — Chapa):",
                    chapas,
                    key="pwd_chapa",
                    format_func=lambda ch: f"{(nome_by_chapa.get(ch, ch) or ch)} — {ch}",
                )
                csel_pwd = next(x for x in colaboradores if x["Chapa"] == ch_sel_pwd)
                user_pwd = get_usuario_sistema_por_setor_chapa(setor, ch_sel_pwd)

                colx1, colx2 = st.columns(2)
                colx1.text_input("Nome:", value=(csel_pwd.get("Nome") or "").strip(), disabled=True, key="pwd_nome_view")
                colx2.text_input("Chapa:", value=str(ch_sel_pwd or "").strip(), disabled=True, key="pwd_chapa_view")
                colx3, colx4 = st.columns(2)
                colx3.text_input("Setor:", value=str(setor or "").strip(), disabled=True, key="pwd_setor_view")
                perfil_view = "ADMIN" if (user_pwd and user_pwd.get("is_admin")) else "LÍDER" if (user_pwd and user_pwd.get("is_lider")) else "AX LÍDER" if (user_pwd and user_pwd.get("is_ax_lider")) else "COLABORADOR" if user_pwd else "SEM ACESSO"
                colx4.text_input("Perfil:", value=perfil_view, disabled=True, key="pwd_perfil_view")

                nova_senha = st.text_input("Nova senha", type="password", key="pwd_nova")
                confirma_senha = st.text_input("Confirmar nova senha", type="password", key="pwd_confirma")
                gerar_tmp = st.checkbox("Gerar senha temporária automática", value=False, key="pwd_auto_temp")

                if st.session_state.get("pwd_temp_last") and st.session_state.get("pwd_temp_last_chapa") == ch_sel_pwd:
                    st.success("🔑 Senha temporária criada com sucesso.")
                    st.code(st.session_state.get("pwd_temp_last"), language=None)
                    st.caption("Copie essa senha e envie ao colaborador. No próximo login ele será obrigado a trocar a senha.")

                if st.button("Salvar nova senha", key="pwd_save"):
                    senha_final = ""
                    forcar_troca = False
                    if gerar_tmp:
                        senha_final = gerar_senha_temporaria_colaborador(8)
                        forcar_troca = True
                    else:
                        senha_final = (nova_senha or "").strip()
                        if not senha_final:
                            st.error("Digite a nova senha ou marque a senha temporária automática.")
                            st.stop()
                        if senha_final != (confirma_senha or "").strip():
                            st.error("A confirmação da senha não confere.")
                            st.stop()
                    try:
                        if not user_pwd:
                            upsert_usuario_sistema(
                                nome=(csel_pwd.get("Nome") or "").strip(),
                                setor=setor,
                                chapa=ch_sel_pwd,
                                senha=senha_final,
                                is_admin=False,
                                is_lider=False,
                                forcar_troca_senha=forcar_troca,
                            )
                            msg_base = "Acesso criado e senha definida com sucesso."
                        else:
                            update_password(setor, ch_sel_pwd, senha_final)
                            set_force_change_password(setor, ch_sel_pwd, forcar_troca)
                            msg_base = "Senha alterada com sucesso."
                        if gerar_tmp:
                            st.session_state["pwd_temp_last"] = senha_final
                            st.session_state["pwd_temp_last_chapa"] = ch_sel_pwd
                            st.success(msg_base)
                            st.success("🔑 Senha temporária criada com sucesso.")
                            st.code(senha_final, language=None)
                            st.caption("Copie essa senha e envie ao colaborador. No próximo login ele será obrigado a trocar a senha.")
                        else:
                            st.session_state["pwd_temp_last"] = ""
                            st.session_state["pwd_temp_last_chapa"] = ""
                            st.success(msg_base)
                            st.rerun()
                    except Exception as e:
                        st.error(f"Falha ao alterar senha: {e}")
            else:
                st.info("Sem colaboradores para alterar senha.")

            st.markdown("---")

        elif sec_col == "🛠️ Atualizar funcionário (AX/Líder)":
            st.markdown("## 🛠️ Atualizar funcionário (AX/Líder)")
            eh_ax = bool(auth.get("is_ax_lider", False)) and not bool(auth.get("is_admin", False))
            st.caption("Perfil AX do Líder propõe alterações. Perfil Líder aprova na subaba de aprovações. Admin e Líder podem aplicar direto.")
            try:
                con = db_conn()
                df_func_ax = pd.read_sql_query(
                    """
                    SELECT nome, setor, chapa, COALESCE(subgrupo,'') AS subgrupo, COALESCE(entrada,'06:00') AS entrada, COALESCE(folga_sab,0) AS folga_sab
                    FROM colaboradores
                    ORDER BY setor, nome
                    """,
                    con,
                )
                df_login_ax = pd.read_sql_query(
                    """
                    SELECT setor, chapa, COALESCE(is_admin,0) AS is_admin, COALESCE(is_lider,0) AS is_lider, COALESCE(is_ax_lider,0) AS is_ax_lider
                    FROM usuarios_sistema
                    """,
                    con,
                )
                con.close()
            except Exception:
                df_func_ax = pd.DataFrame(columns=['nome','setor','chapa','subgrupo','entrada','folga_sab'])
                df_login_ax = pd.DataFrame(columns=['setor','chapa','is_admin','is_lider','is_ax_lider'])

            if df_func_ax.empty:
                st.info("Nenhum colaborador cadastrado para atualizar.")
            else:
                setores_ax = sorted({_norm_setor(x) for x in df_func_ax['setor'].dropna().tolist() if str(x).strip()})
                ax1, ax2 = st.columns([1, 1.7])
                with ax1:
                    setor_ax = st.selectbox("Setor do funcionário", setores_ax, key="ax_func_setor")
                df_func_setor_ax = df_func_ax[df_func_ax['setor'].astype(str).str.strip().str.upper() == _norm_setor(setor_ax)].copy()
                opts_ax = [f"{str(r['nome']).strip()} ({str(r['chapa']).strip()})" for _, r in df_func_setor_ax.iterrows()]
                with ax2:
                    pick_ax = st.selectbox("Funcionário", opts_ax, key="ax_func_pick") if opts_ax else None

                rec_ax = None
                chapa_ax = ""
                if pick_ax:
                    chapa_ax = pick_ax.rsplit("(", 1)[-1].replace(")", "").strip()
                    df_hit_ax = df_func_setor_ax[df_func_setor_ax['chapa'].astype(str).str.strip() == chapa_ax]
                    if not df_hit_ax.empty:
                        rec_ax = df_hit_ax.iloc[0].to_dict()

                if rec_ax:
                    login_hit_ax = df_login_ax[(df_login_ax['setor'].astype(str).str.strip().str.upper() == _norm_setor(setor_ax)) & (df_login_ax['chapa'].astype(str).str.strip() == chapa_ax)]
                    is_admin_cur_ax = bool(int(login_hit_ax.iloc[0]['is_admin'])) if not login_hit_ax.empty else False
                    is_lider_cur_ax = bool(int(login_hit_ax.iloc[0]['is_lider'])) if not login_hit_ax.empty else False
                    is_ax_cur_ax = bool(int(login_hit_ax.iloc[0]['is_ax_lider'])) if not login_hit_ax.empty else False
                    perfil_cur_ax = 'ADMIN' if is_admin_cur_ax else ('LIDER' if is_lider_cur_ax else ('AX_LIDER' if is_ax_cur_ax else 'COLABORADOR'))

                    st.write(f"Atualizando: **{str(rec_ax.get('nome') or '').strip()}** — chapa **{chapa_ax}**")
                    x1, x2, x3, x4 = st.columns([1.4, 1.2, 1.2, 1])
                    with x1:
                        nome_ax_novo = st.text_input("Nome", value=str(rec_ax.get('nome') or '').strip(), key='ax_func_nome')
                    with x2:
                        subgrupo_ax_novo = st.text_input("Subgrupo", value=str(rec_ax.get('subgrupo') or '').strip(), key='ax_func_subgrupo')
                    with x3:
                        entrada_ax_nova = st.text_input("Entrada padrão", value=str(rec_ax.get('entrada') or '06:00').strip() or '06:00', key='ax_func_entrada')
                    with x4:
                        folga_sab_ax = st.checkbox("Folga sábado", value=bool(int(rec_ax.get('folga_sab', 0) or 0)), key='ax_func_folga_sab')

                    perfil_ax_novo = st.selectbox("Perfil do sistema", ['COLABORADOR', 'AX_LIDER', 'LIDER', 'ADMIN'], index=['COLABORADOR', 'AX_LIDER', 'LIDER', 'ADMIN'].index(perfil_cur_ax), key='ax_func_perfil')
                    obs_ax = st.text_area("Observação / motivo da alteração", key="ax_func_obs", height=90)

                    if eh_ax:
                        if st.button("📨 Enviar alteração para aprovação do líder", key='ax_func_salvar'):
                            try:
                                rid = registrar_solicitacao_ax_lider(
                                    setor_solicitante=setor,
                                    setor_alvo=setor_ax,
                                    chapa_alvo=chapa_ax,
                                    nome_alvo=str(rec_ax.get('nome') or '').strip(),
                                    nome_novo=nome_ax_novo,
                                    subgrupo_novo=subgrupo_ax_novo,
                                    perfil_novo=perfil_ax_novo,
                                    entrada_nova=entrada_ax_nova,
                                    folga_sab_nova=bool(folga_sab_ax),
                                    criado_por_nome=str(auth.get('nome') or '').strip(),
                                    criado_por_chapa=str(auth.get('chapa') or '').strip(),
                                    observacao=obs_ax,
                                )
                                st.success(f"Solicitação enviada para aprovação do líder. Protocolo #{rid}.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Falha ao enviar solicitação: {e}")
                    else:
                        if st.button("💾 Salvar atualização do funcionário", key='ax_func_salvar'):
                            try:
                                res = admin_update_funcionario(
                                    setor=setor_ax,
                                    chapa_atual=chapa_ax,
                                    nome_novo=nome_ax_novo,
                                    subgrupo_novo=subgrupo_ax_novo,
                                    perfil_novo=perfil_ax_novo,
                                    entrada_nova=entrada_ax_nova,
                                    folga_sab=bool(folga_sab_ax),
                                    criar_usuario_se_nao_existir=True,
                                )
                                st.success(f"Funcionário atualizado com sucesso. Perfil final: {res['perfil']} | Subgrupo: {res['subgrupo'] or 'SEM SUBGRUPO'}")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Falha ao atualizar funcionário: {e}")

        elif sec_col == "🧾 Aprovações AX":
            st.markdown("## 🧾 Aprovações AX do Líder")
            eh_ax = bool(auth.get("is_ax_lider", False)) and not bool(auth.get("is_admin", False))
            df_ax = listar_solicitacoes_ax_lider()
            if df_ax.empty:
                st.info("Nenhuma solicitação AX cadastrada.")
            else:
                if eh_ax:
                    df_meu = df_ax[df_ax["criado_por_chapa"].astype(str).str.strip() == str(auth.get("chapa") or "").strip()].copy()
                    st.caption("Aqui você acompanha suas solicitações enviadas para aprovação.")
                    st.dataframe(df_meu, use_container_width=True, height=380)
                else:
                    st.caption("O líder/admin aprova ou reprova as alterações propostas pelo AX do Líder.")
                    pend = df_ax[df_ax["status"].astype(str).str.upper() == "PENDENTE"].copy()
                    hist = df_ax[df_ax["status"].astype(str).str.upper() != "PENDENTE"].copy()
                    if pend.empty:
                        st.success("Não há pendências para aprovação no momento.")
                    else:
                        for _, r in pend.iterrows():
                            with st.container(border=True):
                                st.write(f"**Solicitação #{int(r['id'])}** — {str(r['setor_alvo'])} / {str(r['nome_alvo'])} ({str(r['chapa_alvo'])})")
                                st.write(f"**AX:** {str(r['criado_por_nome'])} ({str(r['criado_por_chapa'])})")
                                st.write(f"**Novo nome:** {str(r['nome_novo'])} | **Novo subgrupo:** {str(r['subgrupo_novo'])} | **Novo perfil:** {str(r['perfil_novo'])}")
                                st.write(f"**Entrada:** {str(r['entrada_nova'])} | **Folga sábado:** {'Sim' if bool(int(r['folga_sab_nova'] or 0)) else 'Não'}")
                                if str(r.get('observacao') or '').strip():
                                    st.caption(f"Observação: {str(r.get('observacao') or '').strip()}")
                                ap1, ap2 = st.columns(2)
                                if ap1.button("✅ Aprovar", key=f"ax_aprov_{int(r['id'])}"):
                                    try:
                                        decidir_solicitacao_ax_lider(int(r['id']), str(auth.get('nome') or '').strip(), True)
                                        st.success("Solicitação aprovada e aplicada.")
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Falha ao aprovar: {e}")
                                if ap2.button("❌ Reprovar", key=f"ax_reprov_{int(r['id'])}"):
                                    try:
                                        decidir_solicitacao_ax_lider(int(r['id']), str(auth.get('nome') or '').strip(), False)
                                        st.warning("Solicitação reprovada.")
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Falha ao reprovar: {e}")
                    if not hist.empty:
                        st.markdown("### Histórico")
                        st.dataframe(hist, use_container_width=True, height=280)

        elif sec_col == "🔄 Rodízio Caixa":
            st.markdown("## 🔄 Rodízio mensal Caixa 01 ↔ Caixa 02")
            if str(setor).strip().upper() != "FRENTECAIXA":
                st.info("Rodízio disponível somente para o setor FRENTECAIXA.")
            else:
                cfg = get_rodizio_caixa_cfg(setor)
                c1, c2, c3, c4 = st.columns([1.4, 1.4, 1, 1])
                if 'rod_caixa_origem' not in st.session_state:
                    st.session_state['rod_caixa_origem'] = str(cfg.get('subgrupo_origem') or 'OPERADOR DE CAIXA 01')
                if 'rod_caixa_destino' not in st.session_state:
                    st.session_state['rod_caixa_destino'] = str(cfg.get('subgrupo_destino') or 'OPERADOR DE CAIXA 02')
                if 'rod_caixa_qtd' not in st.session_state:
                    st.session_state['rod_caixa_qtd'] = int(cfg.get('qtd_destino', 14))
                if 'rod_caixa_tol' not in st.session_state:
                    st.session_state['rod_caixa_tol'] = int(cfg.get('tolerancia_min', 20))

                subgrupo_origem = c1.text_input("Subgrupo origem", key='rod_caixa_origem')
                subgrupo_destino = c2.text_input("Subgrupo destino", key='rod_caixa_destino')
                qtd_destino = int(c3.number_input("Qtd fixa no destino", min_value=1, max_value=100, step=1, key='rod_caixa_qtd'))
                tolerancia = int(c4.number_input("Tolerância (min)", min_value=0, max_value=120, step=5, key='rod_caixa_tol'))

                bcfg1, bcfg2, _bcfg3 = st.columns([1, 1, 4])
                if bcfg1.button("Salvar configuração do rodízio", key='rod_caixa_save_cfg', use_container_width=True, disabled=(_status_comp_rod == 'FECHADA')):
                    set_rodizio_caixa_cfg(setor, subgrupo_origem, subgrupo_destino, qtd_destino, tolerancia, True)
                    st.success("Configuração salva.")
                    st.rerun()
                if bcfg2.button("Voltar do zero", key='rod_caixa_reset_cfg', use_container_width=True, disabled=(_status_comp_rod == 'FECHADA')):
                    st.session_state['rod_caixa_origem'] = 'OPERADOR DE CAIXA 01'
                    st.session_state['rod_caixa_destino'] = 'OPERADOR DE CAIXA 02'
                    st.session_state['rod_caixa_qtd'] = 14
                    st.session_state['rod_caixa_tol'] = 20
                    set_rodizio_caixa_cfg(setor, 'OPERADOR DE CAIXA 01', 'OPERADOR DE CAIXA 02', 14, 20, True)
                    st.success('Configuração resetada para o padrão.')
                    st.rerun()

                ano_r = int(st.session_state.get('cfg_ano', datetime.now().year))
                mes_r = int(st.session_state.get('cfg_mes', datetime.now().month))
                _status_comp_rod = get_status_competencia(setor, ano_r, mes_r)
                if _status_comp_rod == 'FECHADA':
                    st.error(f'🔒 Competência {mes_r:02d}/{ano_r} fechada: o rodízio deste mês fica somente para consulta.')
                state_base = f"rod_caixa_aprov::{setor}::{ano_r}::{mes_r}::{subgrupo_origem}::{subgrupo_destino}"
                aprov_key = state_base + "::aprovados"
                neg_key = state_base + "::negados"
                if aprov_key not in st.session_state:
                    st.session_state[aprov_key] = {}
                if neg_key not in st.session_state:
                    st.session_state[neg_key] = []

                sim = simular_rodizio_caixa_mes(
                    setor,
                    ano_r,
                    mes_r,
                    subgrupo_origem,
                    subgrupo_destino,
                    qtd_destino,
                    tolerancia,
                    aprovados_por_slot=st.session_state.get(aprov_key, {}),
                    negados_chapas=st.session_state.get(neg_key, []),
                )
                st.caption(f"Competência ativa: {mes_r:02d}/{ano_r}. Regra fixa do rodízio: 14 trocas por mês, respeitando as cotas por horário do Caixa 01.")
                st.info(f"{subgrupo_destino}: {sim.get('qtd_destino_atual', 0)} pessoa(s) hoje. Rodízio planejado: {sim.get('qtd_troca', 0)} troca(s). Quantidade obrigatória: {sim.get('qtd_destino_obrigatoria', 14)}.")
                st.caption("Na sugestão do mês, o sistema prioriza: 1) horário fixo da cota, 2) domingo mais parecido, 3) quem está há mais tempo sem ir para o Caixa 02.")

                slots = sim.get('slots') or []
                aprovados_atuais = st.session_state.get(aprov_key, {})
                aprovados_validos = sum(1 for s in slots if aprovados_atuais.get(s.get('slot_key')) == s.get('origem_chapa'))
                top1, top2, top3 = st.columns(3)
                top1.metric('Sugestões montadas', len(slots))
                top2.metric('Aprovadas', aprovados_validos)
                top3.metric('Pendentes', max(0, len(slots) - aprovados_validos))

                a1, a2 = st.columns([1, 1])
                if a1.button('Limpar aprovações e negativas', key='rod_caixa_clear_aprov', use_container_width=True):
                    st.session_state[aprov_key] = {}
                    st.session_state[neg_key] = []
                    st.session_state.pop(state_base + "::aplicado", None)
                    st.rerun()
                if a2.button('Aprovar todas as sugestões atuais', key='rod_caixa_aprov_all', use_container_width=True):
                    st.session_state[aprov_key] = {str(s.get('slot_key')): str(s.get('origem_chapa')) for s in slots}
                    st.session_state.pop(state_base + "::aplicado", None)
                    st.rerun()

                aplic_key = state_base + "::aplicado"
                qtd_obrigatoria = int(sim.get('qtd_destino_obrigatoria', 14) or 14)
                pronto_aplicar = bool(slots) and int(aprovados_validos) >= int(qtd_obrigatoria) and int(max(0, len(slots) - aprovados_validos)) == 0

                if pronto_aplicar:
                    st.success(f"Todas as {qtd_obrigatoria} sugestões foram aprovadas. Agora falta aplicar o rodízio no mês {mes_r:02d}/{ano_r}.")
                    if st.button('🔁 Aplicar mudança de subgrupos agora (antes da escala)', key='rod_caixa_apply_now', use_container_width=True, disabled=(_status_comp_rod == 'FECHADA')):
                        res_apply = aplicar_rodizio_caixa_mes(setor, ano_r, mes_r, sim)
                        if res_apply.get('ok'):
                            st.session_state[aplic_key] = True
                            st.success(res_apply.get('msg', 'Rodízio aplicado com sucesso.'))
                            st.rerun()
                        else:
                            st.error(res_apply.get('msg', 'Não foi possível aplicar o rodízio.'))
                elif st.session_state.get(aplic_key):
                    st.success(f"Rodízio já aplicado na competência {mes_r:02d}/{ano_r}. Gere a escala novamente para refletir a troca.")
                elif slots:
                    st.info(f"Para aplicar de verdade no mês {mes_r:02d}/{ano_r}, todas as {qtd_obrigatoria} sugestões precisam estar aprovadas e depois você deve clicar em 'Aplicar mudança de subgrupos agora (antes da escala)'.")

                if slots:
                    st.markdown('### Aprovação das 14 pessoas sugeridas')
                    resumo_aprov = pd.DataFrame([{
                        'Status': 'APROVADO' if aprovados_atuais.get(s.get('slot_key')) == s.get('origem_chapa') else 'PENDENTE',
                        'Nome sugerido': s.get('origem_nome', ''),
                        'Chapa': s.get('origem_chapa', ''),
                        'Horário Caixa 01': s.get('origem_entrada', ''),
                        'Domingos origem': s.get('origem_domingos_label', ''),
                        'Última vez que foi para o Caixa 02': s.get('origem_ultimo_mes_destino_label', ''),
                        'Sai do Caixa 02': s.get('destino_nome', ''),
                        'Domingos destino': s.get('destino_domingos_label', ''),
                        'Domingos iguais trabalho': int(s.get('domingos_trabalho_iguais_qtd', 0) or 0),
                        'Domingos iguais folga': int(s.get('domingos_folga_iguais_qtd', 0) or 0),
                        'Dif. domingos': int(s.get('diff_domingos', 0) or 0),
                        'Alternativas no mesmo horário': int(s.get('alternativas_mesmo_horario', 0) or 0),
                    } for s in slots])
                    st.dataframe(resumo_aprov, use_container_width=True, height=340)

                    for i, s in enumerate(slots, start=1):
                        slot_key = str(s.get('slot_key') or '')
                        aprovado = aprovados_atuais.get(slot_key) == s.get('origem_chapa')
                        with st.container(border=True):
                            cinfo1, cinfo2, cinfo3 = st.columns([3.2, 2.1, 2.2])
                            cinfo1.markdown(
                                f"**{i}. {s.get('origem_nome', '-') }**  \n"
                                f"Chapa: `{s.get('origem_chapa', '-')}` | Horário Caixa 01: **{s.get('origem_entrada', '-') }** | Domingos: **{int(s.get('origem_domingos', 0) or 0)}**"
                            )
                            cinfo2.markdown(
                                f"**Sai do Caixa 02:** {s.get('destino_nome', '-')}  \n"
                                f"Horário destino: **{s.get('destino_entrada', '-')}** | Domingos: **{int(s.get('destino_domingos', 0) or 0)}**"
                            )
                            cinfo3.markdown(
                                f"**Última vez que entrou no Caixa 02:** {s.get('origem_ultimo_mes_destino_label', '-')}  \n"
                                f"Dif. domingos: **{int(s.get('diff_domingos', 0) or 0)}** | Alternativas restantes: **{int(s.get('alternativas_mesmo_horario', 0) or 0)}**"
                            )
                            st.caption(s.get('observacao') or '-')

                            alternativas_slot = list(s.get('alternativas_opcoes') or [])
                            mapa_alt = {}
                            opcoes_alt = []
                            for alt in alternativas_slot:
                                ch_alt = str(alt.get('chapa') or '').strip()
                                if not ch_alt or ch_alt in mapa_alt:
                                    continue
                                label_alt = (
                                    f"{str(alt.get('nome') or '-') } | chapa {ch_alt} | horário {str(alt.get('entrada') or '-')} | "
                                    f"dif. regra {int(alt.get('diff_horario_ref_min', 0) or 0)} min | domingos {int(alt.get('domingos', 0) or 0)} | "
                                    f"último Caixa 02 {str(alt.get('ultimo_mes_destino_label') or '-')}"
                                )
                                mapa_alt[ch_alt] = label_alt
                                opcoes_alt.append(ch_alt)

                            chapa_aprovada_slot = str(aprovados_atuais.get(slot_key) or '').strip()
                            chapa_padrao_slot = chapa_aprovada_slot if chapa_aprovada_slot in opcoes_alt else str(s.get('origem_chapa') or '').strip()
                            idx_padrao_slot = opcoes_alt.index(chapa_padrao_slot) if chapa_padrao_slot in opcoes_alt else 0

                            if opcoes_alt:
                                st.selectbox(
                                    'Escolha quem está mais próximo para este rodízio:',
                                    options=opcoes_alt,
                                    index=idx_padrao_slot,
                                    key=f'rod_caixa_pick_{slot_key}',
                                    format_func=lambda ch: mapa_alt.get(ch, ch),
                                )

                            bcol1, bcol2, bcol3 = st.columns([1, 1, 3])
                            if bcol1.button('✅ Aprovar seleção', key=f'rod_caixa_ok_{slot_key}', use_container_width=True, disabled=aprovado and chapa_aprovada_slot == str(s.get('origem_chapa') or '').strip()):
                                chapa_sel = str(st.session_state.get(f'rod_caixa_pick_{slot_key}', str(s.get('origem_chapa') or '')) or '').strip()
                                tmp = dict(st.session_state.get(aprov_key, {}))
                                tmp[slot_key] = chapa_sel or str(s.get('origem_chapa') or '')
                                st.session_state[aprov_key] = tmp
                                st.rerun()
                            if bcol2.button('❌ Negar e chamar próximo da fila', key=f'rod_caixa_no_{slot_key}', use_container_width=True):
                                negs = list(st.session_state.get(neg_key, []))
                                chapa_neg = str(s.get('origem_chapa') or '').strip()
                                if chapa_neg and chapa_neg not in negs:
                                    negs.append(chapa_neg)
                                tmp = dict(st.session_state.get(aprov_key, {}))
                                tmp.pop(slot_key, None)
                                st.session_state[aprov_key] = tmp
                                st.session_state[neg_key] = negs
                                st.rerun()
                            if aprovado:
                                bcol3.success('Aprovado para aplicação quando todas as 14 estiverem aprovadas.')
                            else:
                                bcol3.warning('Pendente de aprovação manual.')
                else:
                    st.warning("Nenhuma troca encontrada para aplicar neste mês.")

                pares = sim.get('pares') or []
                if pares:
                    df_pares = pd.DataFrame([{
                        'Entra no ' + subgrupo_destino: p['origem_nome'],
                        'Chapa entra': p['origem_chapa'],
                        'Horário atual entra': p['origem_entrada'],
                        'Domingos entra': p.get('origem_domingos_label', ''),
                        'Última vez no Caixa 02': p.get('origem_ultimo_mes_destino_label', ''),
                        'Sai do ' + subgrupo_destino: p['destino_nome'],
                        'Chapa sai': p['destino_chapa'],
                        'Horário atual sai': p['destino_entrada'],
                        'Domingos sai': p.get('destino_domingos_label', ''),
                        'Domingos iguais trabalho': int(p.get('domingos_trabalho_iguais_qtd', 0) or 0),
                        'Domingos iguais folga': int(p.get('domingos_folga_iguais_qtd', 0) or 0),
                        'Dif. domingos': int(p.get('diff_domingos', 0) or 0),
                        'Compatibilidade': p['compatibilidade'],
                        'Observação': p['observacao'] or '-',
                    } for p in pares])
                    st.markdown("### Simulação consolidada do mês")
                    st.dataframe(df_pares, use_container_width=True, height=380)

                cotas_horario = sim.get('cotas_horario') or []
                if cotas_horario:
                    st.markdown("### Regra fixa por horário")
                    st.dataframe(pd.DataFrame(cotas_horario), use_container_width=True, height=240)

                alertas = sim.get('alertas') or []
                if alertas:
                    st.markdown("### Alertas para liderança")
                    for a in alertas:
                        st.warning(a)

                proximos = sim.get('proximos') or []
                if proximos:
                    st.markdown("### Próximos da fila para o próximo mês")
                    st.dataframe(pd.DataFrame(proximos[:50]), use_container_width=True, height=260)

                todos_aprovados = bool(slots) and aprovados_validos == len(slots) and len(slots) >= int(sim.get('qtd_destino_obrigatoria', 14))
                b1, b2 = st.columns([1, 2])
                if not todos_aprovados:
                    b2.info('Para aplicar o rodízio, aprove manualmente todas as 14 sugestões atuais.')
                else:
                    b2.success('As 14 aprovações já estão prontas. Use o botão principal acima para aplicar. Se a base não refletir, use a sincronização manual abaixo.')
                if b1.button("🛠️ Sincronizar subgrupos base manualmente", key='rod_caixa_sync_manual', use_container_width=True):
                    try:
                        res = sincronizar_subgrupos_base_rodizio_caixa(setor, ano_r, mes_r, subgrupo_origem, subgrupo_destino)
                        if res.get('ok'):
                            st.success(res.get('msg', 'Subgrupos sincronizados com sucesso.'))
                            st.rerun()
                        else:
                            st.warning(res.get('msg', 'Nenhum dado para sincronizar.'))
                    except Exception as e:
                        st.error(str(e))

                hist = list_rodizio_caixa_hist(setor, limit=120)
                if hist:
                    st.markdown("### Relatório de trocas já aplicadas")
                    st.dataframe(pd.DataFrame(hist), use_container_width=True, height=320)

    elif sec_main == "🚀 Gerar Escala":
        st.subheader("🚀 Gerar escala")
        st.caption(f"Competência ativa: **{int(st.session_state['cfg_mes']):02d}/{int(st.session_state['cfg_ano'])}**")
        _status_comp_ger = get_status_competencia(setor, int(st.session_state['cfg_ano']), int(st.session_state['cfg_mes']))
        if _status_comp_ger == 'FECHADA':
            st.error('🔒 Competência fechada: geração e readequação ficam bloqueadas. Use a retificação pontual em Ajustes quando precisar corrigir algo.')

        with st.container(border=True):
            c1, c2, c3 = st.columns([1, 1, 2])
            # v9.3 UI: mês/ano vêm somente da Competência (sidebar)
            mes = int(st.session_state["cfg_mes"])
            ano = int(st.session_state["cfg_ano"])
            c1.markdown(f"**Mês/Ano:** {mes:02d}/{ano}")
            c2.caption("Alterar em 🗓️ Competência (sidebar)")
            seed = c3.number_input("Semente", min_value=0, max_value=999999, value=int(st.session_state.get("last_seed", 0)), key="gen_seed")


        colaboradores = load_colaboradores_setor(setor)
        if not colaboradores:
            st.warning("Cadastre colaboradores.")
        else:
            b1, b2, b3, _ = st.columns([1, 1, 1, 5])
            if b1.button("🚀 Gerar agora (respeita ajustes)", use_container_width=True, key="gen_btn", disabled=(_status_comp_ger == 'FECHADA')):
                _clear_preview_cache(setor, int(ano), int(mes))
                st.session_state["last_seed"] = int(seed)
                ok = _regenerar_mes_inteiro(setor, int(ano), int(mes), seed=int(seed), respeitar_ajustes=True)
                if ok:
                    st.success("Escala gerada (ajustes/travas preservados)!")
                else:
                    st.warning("Sem colaboradores.")
                st.rerun()

            if b2.button("🔄 Recarregar do banco", use_container_width=True, key="gen_reload_btn"):
                _clear_preview_cache(setor, int(ano), int(mes))
                st.rerun()

            # 🧹 Gerar do zero: ignora travas/ajustes (recalcula o mês totalmente)
            # -> pede confirmação antes de apagar os overrides do mês.
            if b3.button("🧹 Gerar do zero (ignorar ajustes)", use_container_width=True, key="gen_zero_btn", disabled=(_status_comp_ger == 'FECHADA')):
                st.session_state["confirm_gen_zero"] = True

            if st.session_state.get("confirm_gen_zero", False):
                st.warning(f"Tem certeza que deseja **zerar a escala {mes:02d}/{ano}**? Isso apaga ajustes/travas (overrides) desse mês.", icon="⚠️")
                cy, cn, _sp = st.columns([1, 1, 5])
                if cy.button("✅ Sim", use_container_width=True, key="gen_zero_yes"):
                    st.session_state["confirm_gen_zero"] = False
                    # Importante: se existirem overrides antigos no mês, eles podem "forçar" Folga/Trabalho e aparentar que o motor não funcionou.
                    # Ao gerar do zero, limpamos overrides do mês selecionado (não mexe em meses anteriores).
                    delete_overrides_mes(setor, int(ano), int(mes))
                    _clear_preview_cache(setor, int(ano), int(mes))
                    st.session_state["last_seed"] = int(seed)
                    ok = _regenerar_mes_inteiro(setor, int(ano), int(mes), seed=int(seed), respeitar_ajustes=False)
                    if ok:
                        st.success("Escala gerada do zero (ajustes ignorados)!")
                    else:
                        st.warning("Sem colaboradores.")
                    st.rerun()

                if cn.button("❌ Não", use_container_width=True, key="gen_zero_no"):
                    st.session_state["confirm_gen_zero"] = False
                    st.info("Ação cancelada.")
                    st.rerun()


            preview_key = f"gerar_preview_loaded_{setor}_{ano}_{mes}"
            if preview_key not in st.session_state:
                st.session_state[preview_key] = False

            cprev1, cprev2 = st.columns([1, 5])
            if cprev1.button("📅 Carregar calendário", use_container_width=True, key=f"btn_load_preview_{setor}_{ano}_{mes}"):
                st.session_state[preview_key] = True
                st.rerun()
            if cprev2.button("🧹 Ocultar visualização", use_container_width=True, key=f"btn_hide_preview_{setor}_{ano}_{mes}"):
                st.session_state[preview_key] = False
                st.rerun()

            if st.session_state.get(preview_key, False):
                with st.spinner("Carregando calendário do mês..."):
                    hist_db, cal = _ensure_preview_cache(setor, int(ano), int(mes), colaboradores)

                if hist_db:
                    st.markdown("### 📅 Calendário RH (visual por colaborador)")
                    show_color = st.checkbox("🎨 Mostrar cores no calendário (pode deixar lento)", value=False, key="cal_color")
                    if show_color:
                        st.dataframe(style_calendario(cal, int(mes), int(ano)), use_container_width=True)
                    else:
                        st.dataframe(cal, use_container_width=True)

                    st.markdown("---")
                    st.markdown("### 👤 Visualizar colaborador (detalhado)")
                    ch_view = st.selectbox("Chapa:", list(hist_db.keys()), key="view_ch")
                    st.dataframe(hist_db[ch_view], use_container_width=True, height=420)
                else:
                    st.info("Sem escala no mês. Clique em **Gerar agora**.")
            else:
                st.info("Visualização pesada ficou sob demanda para deixar a navegação mais rápida. Clique em **Carregar calendário** quando quiser ver a escala do mês.")

    # ------------------------------------------------------
    # ABA 3: Ajustes
    # ------------------------------------------------------
    elif sec_main == "⚙️ Ajustes":
        st.subheader("⚙️ Ajustes (travas) — sempre entram na geração")

        with st.container(border=True):
            c1, c2, c3 = st.columns([1, 1, 2])
            # v9.3 UI: mês/ano vêm somente da Competência (sidebar)
            mes = int(st.session_state["cfg_mes"])
            ano = int(st.session_state["cfg_ano"])
            c1.markdown(f"**Mês/Ano:** {mes:02d}/{ano}")
            c2.caption("Alterar em 🗓️ Competência (sidebar)")
            c3.caption("Ajustes aplicam na competência ativa.")

        sec_aj = st.radio("", ["🧩 Folgas manuais em grade", "📊 Contagens por dia", "🔁 Troca de horários", "✅ Preferência por subgrupo", "📌 Subgrupos (editável)", "✏️ Retificar folga, horário e subgrupo"], horizontal=True, key="ajustes_nav_fast", label_visibility="collapsed")

        status_comp = get_status_competencia(setor, ano, mes)
        _is_admin_auth = bool((auth or {}).get('is_admin', False))
        cst1, cst2, cst3 = st.columns([1,1,3])
        cst1.metric('Status da competência', status_comp)
        if cst2.button('🔒 Fechar competência', key=f'fechar_comp::{setor}::{ano}::{mes}', disabled=(status_comp == 'FECHADA')):
            set_status_competencia(setor, ano, mes, 'FECHADA')
            st.success('Competência fechada. Visualização preservada.')
            st.rerun()
        if _is_admin_auth:
            if cst3.button('🔓 Reabrir competência', key=f'reabrir_comp::{setor}::{ano}::{mes}', disabled=(status_comp == 'ABERTA')):
                set_status_competencia(setor, ano, mes, 'ABERTA')
                st.warning('Competência reaberta para edição.')
                st.rerun()
        else:
            cst3.caption('🔓 Reabrir competência: disponível somente para admin.')
        if status_comp == 'FECHADA' and sec_aj != '✏️ Retificar folga, horário e subgrupo':
            st.error('🔒 Competência fechada: nesta área a visualização é apenas leitura. Use a subaba de retificação para correções pontuais.')

        _ajustes_precisam_escala = sec_aj in ("🧩 Folgas manuais em grade", "📊 Contagens por dia", "🔁 Troca de horários")
        hist_db = {}
        colaboradores = []
        colab_by = {}

        if _ajustes_precisam_escala:
            _aj_load_key = f"ajustes_loaded::{setor}::{ano}::{mes}::{sec_aj}"
            if _aj_load_key not in st.session_state:
                st.session_state[_aj_load_key] = False
            c_load1, c_load2, c_load3 = st.columns([1, 1, 3])
            if c_load1.button("📥 Carregar dados dos ajustes", key=f"btn_{_aj_load_key}"):
                st.session_state[_aj_load_key] = True
            if c_load2.button("🧹 Limpar cache desta tela", key=f"clear_{_aj_load_key}"):
                st.session_state.pop(_aj_load_key, None)
                st.rerun()
            c_load3.caption("Para deixar leve, a grade só carrega quando você clicar no botão.")

            if not st.session_state.get(_aj_load_key, False):
                st.info("Esta aba carrega sob demanda. Clique em 📥 Carregar dados dos ajustes para abrir a grade.")
            else:
                with st.spinner("Carregando dados dos ajustes..."):
                    hist_db = get_hist_mes_com_overrides_cached(setor, ano, mes)
                    colaboradores = load_colaboradores_setor(setor)
                    colab_by = {c["Chapa"]: c for c in colaboradores}

                if not hist_db:
                    st.info("Gere a escala primeiro na aba 🚀 Gerar Escala.")
                    return

                if sec_aj == "🧩 Folgas manuais em grade":
                    st.markdown("### 🧩 Folgas manuais em grade (por colaborador)")
                    st.caption("Marque/desmarque as folgas do mês. Isso cria/remove travas (overrides) de Status=Folga. Domingo é editável aqui (manual é soberano).")
                    # --- filtro de colaboradores (para facilitar)
                    # Regra v8.4:
                    # - Se você selecionar 1+ colaboradores, a grade mostra SOMENTE os selecionados (mesmo se "Mostrar todos" estiver marcado).
                    # - Se não selecionar ninguém, a grade respeita o checkbox (todos ou nenhum).
                    show_all = st.checkbox("👥 Mostrar todos os colaboradores", value=True, key="grid_show_all")

                    labels_opts = [f'{c["Nome"]} ({c["Chapa"]})' for c in colaboradores]
                    inv_label = {f'{c["Nome"]} ({c["Chapa"]})': str(c["Chapa"]) for c in colaboradores}

                    sel_labels = st.multiselect(
                        "Selecionar colaboradores para editar (se selecionar, a grade mostra somente eles):",
                        options=labels_opts,
                        default=st.session_state.get("grid_sel_labels", []),
                        key="grid_sel_labels"
                    )
                    sel_chapas = [inv_label[l] for l in sel_labels if l in inv_label]

                    if sel_chapas:
                        colaboradores = [c for c in colaboradores if str(c["Chapa"]) in set(sel_chapas)]
                        st.caption(f"Mostrando {len(colaboradores)} colaborador(es) selecionado(s).")
                    else:
                        colaboradores = colaboradores if show_all else []
                        if not show_all:
                            st.info("Marque 'Mostrar todos' ou selecione 1+ colaboradores acima.")


                    qtd = calendar.monthrange(int(ano), int(mes))[1]
                    dias = list(range(1, qtd + 1))

                    # pega overrides existentes
                    ovdf = load_overrides(setor, ano, mes)
                    ov_status = {}
                    if ovdf is not None and not ovdf.empty:
                        od = ovdf[ovdf["campo"] == "status"]
                        for _, r in od.iterrows():
                            if str(r["valor"]) == "Folga":
                                ov_status.setdefault(str(r["chapa"]), set()).add(int(r["dia"]))

                    # monta grade
                    rows = []
                    for c in colaboradores:
                        chg = str(c["Chapa"])
                        row = {"Nome": c["Nome"], "Chapa": chg}
                        dfh = hist_db.get(chg)
                        for d in dias:
                            if dfh is not None and len(dfh) >= d:
                                if dfh.loc[d - 1, "Status"] == "Férias":
                                    row[str(d)] = False
                                else:
                                    row[str(d)] = (dfh.loc[d - 1, "Status"] == "Folga") or (d in ov_status.get(chg, set()))
                            else:
                                row[str(d)] = False
                        rows.append(row)

                    df_grid = pd.DataFrame(rows)
                    edited = st.data_editor(
                        df_grid,
                        use_container_width=True,
                        hide_index=True,
                        num_rows="fixed",
                        column_config={str(d): st.column_config.CheckboxColumn(str(d), width="small") for d in dias},
                        key="grid_editor"
                    )

                    auto_readequar = st.checkbox("🔄 Readequar escala ao salvar (somente se você quiser)", value=False, key="grid_auto_regen")

                    if st.button("💾 Salvar folgas manuais (e readequar mês)", key="grid_save", disabled=(status_comp == 'FECHADA')):
                        set_folga = 0
                        set_trab = 0
                        for _, r in edited.iterrows():
                            chg = str(r["Chapa"])
                            dfh = hist_db.get(chg)
                            ent_pad_local = colab_by.get(chg, {}).get("Entrada", "06:00")
                            for d in dias:
                                want_folga = bool(r[str(d)])
                                if dfh is not None and len(dfh) >= d:
                                    if dfh.loc[d - 1, "Status"] == "Férias":
                                        continue

                                if want_folga:
                                    set_override(setor, ano, mes, chg, d, "status", "Folga")
                                    set_folga += 1
                                else:
                                    # ✅ regra pedida: desmarcado = TRABALHO (travado)
                                    set_override(setor, ano, mes, chg, d, "status", "Trabalho")
                                    # mantém horário padrão no banco via geração/descanso global; se quiser travar horário também,
                                    # descomente as linhas abaixo:
                                    # set_override(setor, ano, mes, chg, d, "h_entrada", ent_pad_local)
                                    # set_override(setor, ano, mes, chg, d, "h_saida", _saida_from_entrada(ent_pad_local))
                                    set_trab += 1

                        if auto_readequar:
                            _regenerar_mes_inteiro(setor, ano, mes, seed=int(st.session_state.get("last_seed", 0)), respeitar_ajustes=True)

                        st.success(f"Salvo! Folgas travadas: {set_folga} | Trabalhos travados: {set_trab}.")
                        st.rerun()

                elif sec_aj == "🧷 Folga fixa":
                    st.markdown("### 🧷 Folga fixa por colaborador")
                    st.caption("Escolha a pessoa, marque os dias da semana de folga fixa e valide antes de salvar. Se quebrar regra, o sistema avisa e você decide se quer salvar mesmo assim.")

                    if not colaboradores:
                        colaboradores = load_colaboradores_setor(setor)
                        colab_by = {c["Chapa"]: c for c in colaboradores}
                    labels_ff = [f'{c["Nome"]} ({c["Chapa"]})' for c in colaboradores]
                    inv_ff = {f'{c["Nome"]} ({c["Chapa"]})': str(c["Chapa"]) for c in colaboradores}
                    label_sel_ff = st.selectbox("Colaborador:", options=labels_ff, key="folga_fixa_colab")
                    chapa_ff = inv_ff.get(label_sel_ff, "")
                    atuais_ff = get_folga_fixa_weekdays(setor, chapa_ff)
                    dias_sel_ff = st.multiselect(
                        "Dias da semana de folga fixa:",
                        options=list(range(7)),
                        default=atuais_ff,
                        format_func=lambda x: WEEKDAY_LABELS_LONG.get(int(x), str(x)),
                        key=f"folga_fixa_days::{chapa_ff}::{ano}::{mes}",
                    )

                    dias_mes_fixos = _dias_mes_por_weekdays(ano, mes, dias_sel_ff)
                    st.caption("Dias da competência afetados: " + (", ".join(f"{d:02d}" for d in dias_mes_fixos) if dias_mes_fixos else "nenhum"))

                    if hist_db:
                        df_hist_ff = hist_db.get(chapa_ff)
                    else:
                        df_hist_ff = get_hist_mes_com_overrides_cached(setor, ano, mes).get(chapa_ff)
                    warnings_ff = _simulate_folga_fixa_warnings(df_hist_ff, ano, mes, dias_mes_fixos) if chapa_ff else []
                    if warnings_ff:
                        st.warning("Validação da folga fixa encontrou estes pontos:")
                        for msg in warnings_ff:
                            st.write(f"- {msg}")
                    else:
                        st.success("Nenhuma quebra visível encontrada para esta folga fixa na competência ativa.")

                    force_ff = st.checkbox("Salvar mesmo se houver alerta de regra", value=False, key="folga_fixa_force")
                    col_ff1, col_ff2, col_ff3 = st.columns([1,1,2])
                    if col_ff1.button("💾 Salvar folga fixa", key="folga_fixa_salvar"):
                        if warnings_ff and not force_ff:
                            st.error("Há alertas de regra. Marque a opção para salvar mesmo assim, se quiser forçar.")
                        else:
                            save_folga_fixa(setor, chapa_ff, dias_sel_ff)
                            # aplica como trava manual do mês ativo
                            if dias_mes_fixos:
                                for dia in dias_mes_fixos:
                                    set_override(setor, ano, mes, chapa_ff, dia, "status", "Folga")
                            st.success("Folga fixa salva e aplicada como trava manual na competência ativa.")
                            st.rerun()
                    if col_ff2.button("🗑️ Remover folga fixa", key="folga_fixa_remover"):
                        remove_folga_fixa(setor, chapa_ff)
                        st.success("Folga fixa removida. As travas já gravadas no mês atual continuam até você alterar manualmente a grade ou regenerar.")
                        st.rerun()
                    folga_fixa_df = list_folga_fixa(setor)
                    if not folga_fixa_df.empty:
                        st.markdown("#### Folgas fixas cadastradas")
                        st.dataframe(folga_fixa_df[["Nome", "Chapa", "Dia", "Ativo", "CriadoEm"]], use_container_width=True, hide_index=True)
                    else:
                        st.info("Nenhuma folga fixa cadastrada ainda.")

                elif sec_aj == "🗂️ Inventário":
                    st.markdown("### 🗂️ Inventário")
                    st.caption("Escolha o dia e informe quantas pessoas você quer em abertura, intermediário e fechamento. A tabela mensal continua abaixo para conferência rápida.")
                    qtd_inv = calendar.monthrange(int(ano), int(mes))[1]
                    inv_atual = get_inventario_mes(setor, ano, mes)
                    inv_map = {int(r["Dia"]): r for _, r in inv_atual.iterrows()} if not inv_atual.empty else {}

                    dia_inv = st.selectbox(
                        "Dia do inventário:",
                        options=list(range(1, qtd_inv + 1)),
                        key=f"inventario_dia_foco::{setor}::{ano}::{mes}",
                    )
                    base_inv = inv_map.get(int(dia_inv), {})
                    data_inv = date(int(ano), int(mes), int(dia_inv))
                    st.info("Aqui você define quantas pessoas quer no dia do balanço em cada faixa: abertura, intermediário e fechamento.")
                    st.caption(f"Data escolhida: {data_inv.strftime('%d/%m/%Y')} — {WEEKDAY_LABELS_LONG[data_inv.weekday()]}")

                    ci1, ci2, ci3 = st.columns(3)
                    meta_ab = ci1.number_input(
                        "Abertura",
                        min_value=0,
                        step=1,
                        value=int(base_inv["Abertura"]) if base_inv != {} else 0,
                        key=f"meta_abertura::{setor}::{ano}::{mes}::{dia_inv}",
                    )
                    meta_in = ci2.number_input(
                        "Intermediário",
                        min_value=0,
                        step=1,
                        value=int(base_inv["Intermediario"]) if base_inv != {} else 0,
                        key=f"meta_intermediario::{setor}::{ano}::{mes}::{dia_inv}",
                    )
                    meta_fe = ci3.number_input(
                        "Fechamento",
                        min_value=0,
                        step=1,
                        value=int(base_inv["Fechamento"]) if base_inv != {} else 0,
                        key=f"meta_fechamento::{setor}::{ano}::{mes}::{dia_inv}",
                    )

                    csave1, csave2 = st.columns([1, 3])
                    if csave1.button("💾 Salvar dia selecionado", key=f"inventario_salvar_dia::{setor}::{ano}::{mes}::{dia_inv}"):
                        upsert_inventario_dia(setor, ano, mes, int(dia_inv), int(meta_ab), int(meta_in), int(meta_fe))
                        st.success(f"Inventário salvo para o dia {int(dia_inv):02d}/{int(mes):02d}/{int(ano)}.")
                        st.rerun()
                    csave2.caption("Use esta área para cadastrar a necessidade do dia. Isso entra na geração da escala quando houver inventário configurado.")

                    rows_inv = []
                    for dia in range(1, qtd_inv + 1):
                        base = inv_map.get(dia, {})
                        rows_inv.append({
                            "Dia": dia,
                            "Data": date(int(ano), int(mes), dia).strftime("%d/%m/%Y"),
                            "Semana": WEEKDAY_LABELS_LONG[date(int(ano), int(mes), dia).weekday()],
                            "Abertura": int(base["Abertura"]) if base != {} else 0,
                            "Intermediário": int(base["Intermediario"]) if base != {} else 0,
                            "Fechamento": int(base["Fechamento"]) if base != {} else 0,
                        })
                    df_inv_view = pd.DataFrame(rows_inv)
                    st.markdown("#### Inventário do mês")
                    st.dataframe(df_inv_view, use_container_width=True, hide_index=True)

                    comp_inv = build_inventario_comparativo(setor, ano, mes, hist_db if hist_db else None)
                    if not comp_inv.empty:
                        st.markdown("#### Comparativo meta x escala atual")
                        st.dataframe(comp_inv, use_container_width=True, hide_index=True)
                    else:
                        st.info("Cadastre as metas do mês para acompanhar o comparativo depois.")

                elif sec_aj == "✏️ Retificar folga, horário e subgrupo":
                    st.markdown("### ✏️ Retificar folga, horário e subgrupo")
                    st.caption("Use esta subaba para corrigir competência fechada sem descongelar o mês inteiro. A alteração aparece nas leituras da escala e no portal do colaborador.")
                    colaboradores_ret = load_colaboradores_setor(setor)
                    if not colaboradores_ret:
                        st.info("Cadastre colaboradores primeiro.")
                    else:
                        labels_ret = [f"{c['Nome']} ({c['Chapa']})" for c in colaboradores_ret]
                        inv_ret = {f"{c['Nome']} ({c['Chapa']})": c for c in colaboradores_ret}
                        colr1, colr2, colr3 = st.columns([2,1,1])
                        label_ret = colr1.selectbox("Funcionário", options=labels_ret, key=f"ret_func::{setor}::{ano}::{mes}")
                        colab_ret = inv_ret.get(label_ret) or {}
                        chapa_ret = str(colab_ret.get('Chapa') or '').strip()
                        qtd_ret = calendar.monthrange(int(ano), int(mes))[1]
                        dia_ret = int(colr2.selectbox("Dia", options=list(range(1, qtd_ret + 1)), key=f"ret_dia::{setor}::{ano}::{mes}"))
                        hist_ret = get_hist_mes_com_overrides_cached(setor, ano, mes) or {}
                        df_ret_hist = hist_ret.get(chapa_ret)
                        base_status = ''
                        base_ent = str(colab_ret.get('Entrada') or '06:00').strip()
                        base_sai = _saida_from_entrada(base_ent)
                        if df_ret_hist is not None and len(df_ret_hist) >= dia_ret:
                            base_status = str(df_ret_hist.loc[dia_ret - 1, 'Status'] or '').strip()
                            base_ent = str(df_ret_hist.loc[dia_ret - 1, 'H_Entrada'] or '').strip()
                            base_sai = str(df_ret_hist.loc[dia_ret - 1, 'H_Saida'] or '').strip()
                        colra, colrb, colrc, colrd = st.columns([1,1,1,1])
                        novo_status = colra.selectbox("Novo status", options=['', 'Trabalho', 'Folga', 'Férias', 'Afastamento'], index=0, key=f"ret_status::{setor}::{ano}::{mes}")
                        nova_entrada = colrb.text_input("Nova entrada", value=base_ent, key=f"ret_ent::{setor}::{ano}::{mes}")
                        nova_saida = colrc.text_input("Nova saída", value=base_sai, key=f"ret_sai::{setor}::{ano}::{mes}")
                        novo_subgrupo = colrd.selectbox("Novo subgrupo", options=['', 'OPERADOR DE CAIXA 01', 'OPERADOR DE CAIXA 02'] + sorted({str(c.get('Subgrupo') or '').strip() for c in colaboradores_ret if str(c.get('Subgrupo') or '').strip()}), index=0, key=f"ret_sub::{setor}::{ano}::{mes}")
                        motivo_ret = st.text_area("Motivo da retificação", key=f"ret_motivo::{setor}::{ano}::{mes}")
                        if st.button("💾 Salvar retificação", key=f"ret_save::{setor}::{ano}::{mes}", use_container_width=True):
                            if not chapa_ret:
                                st.warning('Selecione um funcionário válido.')
                            else:
                                salvar_retificacao_competencia(
                                    setor, ano, mes, chapa_ret, dia_ret,
                                    novo_status=novo_status or base_status,
                                    novo_entrada=nova_entrada,
                                    novo_saida=nova_saida,
                                    novo_subgrupo=novo_subgrupo,
                                    motivo=motivo_ret,
                                    usuario=str(st.session_state.get('auth_nome') or st.session_state.get('auth_chapa') or '')
                                )
                                st.success('Retificação salva com sucesso.')
                                st.rerun()
                        df_ret_list = load_retificacoes_competencia(setor, ano, mes)
                        if df_ret_list is not None and not df_ret_list.empty:
                            st.markdown('#### Retificações já registradas nesta competência')
                            st.dataframe(df_ret_list[[c for c in ['dia','nome','chapa','novo_status','nova_entrada','nova_saida','novo_subgrupo','motivo','usuario','criado_em'] if c in df_ret_list.columns]], use_container_width=True, hide_index=True)

                elif sec_aj == "📊 Contagens por dia":
                    st.markdown("### 📊 Contagens por dia")
                    st.caption("Mostra as contagens do dia escolhido, a visão Excel do mês e as contagens por subgrupo.")
                    qtd_cov = calendar.monthrange(int(ano), int(mes))[1]
                    dia_cov = st.selectbox(
                        "Dia para análise:",
                        options=list(range(1, qtd_cov + 1)),
                        key=f"contagens_dia_foco::{setor}::{ano}::{mes}",
                    )
                    data_cov = date(int(ano), int(mes), int(dia_cov))
                    st.caption(f"Data escolhida: {data_cov.strftime('%d/%m/%Y')} — {WEEKDAY_LABELS_LONG[data_cov.weekday()]}")

                    hist_view_inv = hist_db or get_hist_mes_com_overrides_cached(setor, ano, mes)
                    if hist_view_inv:
                        df_cov_geral = build_cobertura_diaria_geral(setor, ano, mes, hist_view_inv)
                        row_cov = df_cov_geral[df_cov_geral["Dia"] == int(dia_cov)]
                        if not row_cov.empty:
                            row_cov = row_cov.iloc[0]
                            st.markdown("#### Contagens do dia selecionado")
                            m1, m2, m3, m4 = st.columns(4)
                            m1.metric("Abertura", int(row_cov["Abertura"]))
                            m2.metric("Intermediário", int(row_cov["Intermediário"]))
                            m3.metric("Fechamento", int(row_cov["Fechamento"]))
                            m4.metric("Total trabalhando", int(row_cov["Total trabalhando"]))
                            m5, m6, m7 = st.columns(3)
                            m5.metric("Folga", int(row_cov["Folga"]))
                            m6.metric("Férias", int(row_cov["Férias"]))
                            m7.metric("Afastamento", int(row_cov["Afastamento"]))

                        df_cov_sub = build_cobertura_por_subgrupo_no_dia(setor, ano, mes, int(dia_cov), hist_view_inv)
                        st.markdown("#### Contagens por dia — visão Excel (geral)")
                        st.dataframe(df_cov_geral, use_container_width=True, hide_index=True)
                        if not df_cov_sub.empty:
                            st.markdown("#### Contagens por subgrupo no dia selecionado")
                            st.dataframe(df_cov_sub, use_container_width=True, hide_index=True)
                    else:
                        st.info("Gere a escala para visualizar as contagens por dia e por subgrupo.")

                elif sec_aj == "📝 Histórico":
                    st.markdown("### 📝 Histórico")
                    st.caption("Mostra quantas pessoas estarão de folga em cada dia da competência e quem são elas.")
                    hist_view = hist_db or get_hist_mes_com_overrides_cached(setor, ano, mes)
                    if not hist_view:
                        st.info("Gere a escala primeiro para visualizar o histórico.")
                    else:
                        df_hist_dia = build_historico_folgas_diario(setor, ano, mes, hist_view)
                        st.dataframe(df_hist_dia, use_container_width=True, hide_index=True)
                        dias_hist = df_hist_dia["Dia"].tolist()
                        dia_sel_hist = st.selectbox("Ver detalhes do dia:", options=dias_hist, key="hist_dia_sel")
                        row_hist = df_hist_dia[df_hist_dia["Dia"] == int(dia_sel_hist)].iloc[0]
                        st.info(f"{row_hist['Data']} — Folga: {row_hist['Folga']} | Férias: {row_hist['Férias']} | Afastamento: {row_hist['Afastamento']} | Trabalho: {row_hist['Trabalho']}")
                        nomes_folga = str(row_hist.get('Pessoas de folga', '') or '').strip()
                        if nomes_folga:
                            st.write("**Pessoas de folga no dia:**")
                            st.write(nomes_folga)
                        else:
                            st.write("**Pessoas de folga no dia:** nenhuma")
                        st.text_area("Pessoas de folga no dia selecionado", value=str(row_hist["Pessoas de folga"] or ""), height=140, key="hist_pessoas_folga", disabled=True)

                elif sec_aj == "🔁 Troca de horários":
                                st.markdown("### 🔁 Troca de horários em grade (por colaborador)")
                                st.caption("Escolha o horário e marque (quadradinhos) os dias em que ele deve valer. **Folga/Férias sempre prevalecem**: se o dia estiver como Folga/Férias/AFA, o sistema NÃO aplica horário nesse dia.")

                                qtd2 = calendar.monthrange(int(ano), int(mes))[1]
                                dias2 = list(range(1, qtd2 + 1))

                                # --- filtro/seleção de colaboradores (mesmo layout da grade de folgas)
                                show_all_th = st.checkbox("👥 Mostrar todos os colaboradores", value=True, key="th_show_all")

                                labels_opts_th = [f'{c["Nome"]} ({c["Chapa"]})' for c in colaboradores]
                                inv_label_th = {f'{c["Nome"]} ({c["Chapa"]})': str(c["Chapa"]) for c in colaboradores}

                                sel_labels_th = st.multiselect(
                                    "Selecionar colaboradores para editar (se selecionar, a grade mostra somente eles):",
                                    options=labels_opts_th,
                                    default=st.session_state.get("th_sel_labels", []),
                                    key="th_sel_labels"
                                )
                                sel_chapas_th = [inv_label_th[l] for l in sel_labels_th if l in inv_label_th]

                                if sel_chapas_th:
                                    colaboradores = [c for c in colaboradores if str(c["Chapa"]) in set(sel_chapas_th)]
                                    st.caption(f"Mostrando {len(colaboradores)} colaborador(es) selecionado(s).")
                                else:
                                    colaboradores = colaboradores if show_all_th else []
                                    if not colaboradores:
                                        st.info("Selecione colaboradores acima ou marque 'Mostrar todos'.")
                                        # evita montar grade vazia que confunde
                                        st.stop()

                                # ação a aplicar (horário/folga/afastamento)
                                acao_th = st.selectbox(
                                    "Ação para aplicar nos dias marcados:",
                                    options=["Horário", "Folga", "Afastamento"],
                                    index=0,
                                    key="th_acao_sel"
                                )

                                horario_sel = None
                                if acao_th == "Horário":
                                    horario_sel = st.selectbox(
                                        "Horário (Entrada) para aplicar nos dias marcados:",
                                        options=HORARIOS_ENTRADA_PRESET,
                                        index=HORARIOS_ENTRADA_PRESET.index(BALANCO_DIA_ENTRADA) if BALANCO_DIA_ENTRADA in HORARIOS_ENTRADA_PRESET else 0,
                                        key="th_horario_sel"
                                    )
                                elif acao_th == "Folga":
                                    st.info("Dias marcados serão salvos como **Folga**. (Folga sempre prevalece sobre horário.)")
                                else:
                                    st.info("Dias marcados serão salvos como **Afastamento (AFA)**. Após acabar, a escala volta a seguir as regras normalmente.")
    # overrides do mês (para respeitar folgas/férias)
                                ovmap = _ov_map(setor, ano, mes)

                                # monta grade: SOMENTE Nome, Chapa e dias (checkbox)
                                rows = []
                                for c in colaboradores:
                                    ch = str(c["Chapa"])
                                    nm = c.get("Nome","")
                                    row = {"Nome": nm, "Chapa": ch}
                                    # pré-preenche conforme a ação selecionada
                                    for d in dias2:
                                        cur = (ovmap.get(ch, {}).get(d, {}) or {})
                                        if acao_th == "Horário":
                                            row[str(d)] = (cur.get("h_entrada") == horario_sel)
                                        elif acao_th == "Folga":
                                            row[str(d)] = str(cur.get("status") or "").strip().upper() in ("FOLGA","FOLG")
                                        else:
                                            row[str(d)] = str(cur.get("status") or "").strip().upper() in ("AFASTAMENTO","AFA")
                                    rows.append(row)

                                df_th = pd.DataFrame(rows)

                                edited_th = st.data_editor(
                                    df_th,
                                    use_container_width=True,
                                    hide_index=True,
                                    num_rows="fixed",
                                    column_config={str(d): st.column_config.CheckboxColumn(str(d), width="small") for d in dias2},
                                    key="th_grid_editor"
                                )

                                auto_readequar_th = st.checkbox("🔄 Readequar escala ao salvar", value=True, key="th_auto_regen")

                                if st.button("💾 Salvar troca de horários (aplicar nos dias marcados)", key="th_save"):
                                    applied = 0
                                    skipped = 0
                                    for _, r in edited_th.iterrows():
                                        ch = str(r["Chapa"])
                                        dfh = hist_db.get(ch)
                                        # horário padrão para fallback
                                        ent_pad = (colab_by.get(ch, {}) or {}).get("Entrada", BALANCO_DIA_ENTRADA)

                                        for d in dias2:
                                            want = bool(r[str(d)])

                                            # status do dia (já com overrides)
                                            status_dia = None
                                            if dfh is not None and len(dfh) >= d:
                                                status_dia = str(dfh.loc[d - 1, "Status"])
                                            st_ov = (ovmap.get(ch, {}).get(d, {}) or {}).get("status")
                                            if st_ov:
                                                status_dia = str(st_ov)

                                            st_norm = str(status_dia or "").strip().upper()

                                            if acao_th == "Horário":
                                                # ✅ regra: Folga/Férias/Afastamento sempre prevalecem (não aplicar horário)
                                                if st_norm in ("FOLGA","FOLG","FÉRIAS","FERIAS","FER","AFA","AFASTAMENTO"):
                                                    if want:
                                                        skipped += 1
                                                    continue

                                                if want:
                                                    set_override(setor, ano, mes, ch, d, "h_entrada", horario_sel)
                                                    applied += 1
                                                else:
                                                    # desmarcado: remove override de horário (limpa h_entrada do dia)
                                                    del_override(setor, ano, mes, ch, d, "h_entrada")

                                            elif acao_th == "Folga":
                                                # Folga sobrepõe qualquer horário: salva status e remove h_entrada
                                                if st_norm in ("FER","FÉRIAS","FERIAS"):
                                                    if want:
                                                        skipped += 1
                                                    continue
                                                if want:
                                                    set_override(setor, ano, mes, ch, d, "status", "Folga")
                                                    del_override(setor, ano, mes, ch, d, "h_entrada")
                                                    applied += 1
                                                else:
                                                    del_override(setor, ano, mes, ch, d, "status")

                                            else:  # Afastamento
                                                if st_norm in ("FER","FÉRIAS","FERIAS"):
                                                    if want:
                                                        skipped += 1
                                                    continue
                                                if want:
                                                    set_override(setor, ano, mes, ch, d, "status", "AFA")
                                                    del_override(setor, ano, mes, ch, d, "h_entrada")
                                                    applied += 1
                                                else:
                                                    del_override(setor, ano, mes, ch, d, "status")

                                    if auto_readequar_th:
                                        _regenerar_mes_inteiro(setor, ano, mes, seed=int(st.session_state.get("last_seed", 0)), respeitar_ajustes=True)

                                    st.success(f"Salvo! Ação: {acao_th}. Aplicados: {applied}. Ignorados (por conflito com Folga/Férias): {skipped}.")
                                    st.rerun()

        if sec_aj == "✏️ Retificar folga, horário e subgrupo":
            st.markdown("### ✏️ Retificar folga, horário e subgrupo")
            st.caption("Use esta subaba para corrigir competência fechada sem descongelar o mês inteiro. A alteração aparece nas leituras da escala e no portal do colaborador.")
            colaboradores_ret = load_colaboradores_setor(setor)
            if not colaboradores_ret:
                st.info("Cadastre colaboradores primeiro.")
            else:
                labels_ret = [f"{c['Nome']} ({c['Chapa']})" for c in colaboradores_ret]
                inv_ret = {f"{c['Nome']} ({c['Chapa']})": c for c in colaboradores_ret}
                colr1, colr2, colr3 = st.columns([2, 1, 1])
                label_ret = colr1.selectbox("Funcionário", options=labels_ret, key=f"ret_func_live::{setor}::{ano}::{mes}")
                colab_ret = inv_ret.get(label_ret) or {}
                chapa_ret = str(colab_ret.get('Chapa') or '').strip()
                qtd_ret = calendar.monthrange(int(ano), int(mes))[1]
                dia_ret = int(colr2.selectbox("Dia", options=list(range(1, qtd_ret + 1)), key=f"ret_dia_live::{setor}::{ano}::{mes}"))

                hist_ret = get_hist_mes_com_overrides_cached(setor, ano, mes) or {}
                df_ret_hist = hist_ret.get(chapa_ret)
                base_status = ''
                base_ent = str(colab_ret.get('Entrada') or '06:00').strip()
                base_sai = _saida_from_entrada(base_ent)
                base_sub = str(colab_ret.get('Subgrupo') or '').strip()
                if df_ret_hist is not None and len(df_ret_hist) >= dia_ret:
                    base_status = str(df_ret_hist.loc[dia_ret - 1, 'Status'] or '').strip()
                    base_ent = str(df_ret_hist.loc[dia_ret - 1, 'H_Entrada'] or '').strip()
                    base_sai = str(df_ret_hist.loc[dia_ret - 1, 'H_Saida'] or '').strip()
                    try:
                        if 'Subgrupo' in df_ret_hist.columns:
                            base_sub = str(df_ret_hist.loc[dia_ret - 1, 'Subgrupo'] or '').strip() or base_sub
                    except Exception:
                        pass

                st.info(f"Base do dia {dia_ret:02d}/{int(mes):02d}/{int(ano)} → Status: {base_status or '-'} | Entrada: {base_ent or '-'} | Saída: {base_sai or '-'} | Subgrupo: {base_sub or '-'}")
                colra, colrb, colrc, colrd = st.columns([1, 1, 1, 1])
                status_opts = ['', 'Trabalho', 'Folga', 'Férias', 'Afastamento']
                idx_status = status_opts.index(base_status) if base_status in status_opts else 0
                novo_status = colra.selectbox("Novo status", options=status_opts, index=idx_status, key=f"ret_status_live::{setor}::{ano}::{mes}")
                nova_entrada = colrb.text_input("Nova entrada", value=base_ent, key=f"ret_ent_live::{setor}::{ano}::{mes}")
                nova_saida = colrc.text_input("Nova saída", value=base_sai, key=f"ret_sai_live::{setor}::{ano}::{mes}")
                sub_opts = [''] + sorted({str(c.get('Subgrupo') or '').strip() for c in colaboradores_ret if str(c.get('Subgrupo') or '').strip()})
                if 'OPERADOR DE CAIXA 01' not in sub_opts:
                    sub_opts.append('OPERADOR DE CAIXA 01')
                if 'OPERADOR DE CAIXA 02' not in sub_opts:
                    sub_opts.append('OPERADOR DE CAIXA 02')
                sub_opts = [''] + sorted({x for x in sub_opts if x})
                idx_sub = sub_opts.index(base_sub) if base_sub in sub_opts else 0
                novo_subgrupo = colrd.selectbox("Novo subgrupo", options=sub_opts, index=idx_sub, key=f"ret_sub_live::{setor}::{ano}::{mes}")
                motivo_ret = st.text_area("Motivo da retificação", key=f"ret_motivo_live::{setor}::{ano}::{mes}")

                if st.button("💾 Salvar retificação", key=f"ret_save_live::{setor}::{ano}::{mes}", use_container_width=True):
                    if not chapa_ret:
                        st.warning('Selecione um funcionário válido.')
                    else:
                        salvar_retificacao_competencia(
                            setor, ano, mes, chapa_ret, dia_ret,
                            novo_status=novo_status or base_status,
                            novo_entrada=nova_entrada,
                            novo_saida=nova_saida,
                            novo_subgrupo=novo_subgrupo or base_sub,
                            motivo=motivo_ret,
                            usuario=str(st.session_state.get('auth_nome') or st.session_state.get('auth_chapa') or '')
                        )
                        st.success('Retificação salva com sucesso.')
                        st.rerun()

                df_ret_list = load_retificacoes_competencia(setor, ano, mes)
                if df_ret_list is not None and not df_ret_list.empty:
                    st.markdown('#### Retificações já registradas nesta competência')
                    cols_view = [c for c in ['dia', 'nome', 'chapa', 'novo_status', 'nova_entrada', 'nova_saida', 'novo_subgrupo', 'motivo', 'usuario', 'criado_em'] if c in df_ret_list.columns]
                    st.dataframe(df_ret_list[cols_view], use_container_width=True, hide_index=True)

        if sec_aj == "✅ Preferência por subgrupo":
            st.markdown("### ✅ Preferência por subgrupo (Evitar folga se possível)")
            subgrupos = list_subgrupos(setor)
            if subgrupos:
                sg_sel = st.selectbox("Escolha o subgrupo:", subgrupos, key="pref_sg_sel")
                regras = get_subgrupo_regras(setor, sg_sel)

                p1, p2, p3 = st.columns(3)
                ev_seg = p1.checkbox("Evitar SEG", value=bool(regras["seg"]), key=f"ev_seg_{sg_sel}")
                ev_ter = p1.checkbox("Evitar TER", value=bool(regras["ter"]), key=f"ev_ter_{sg_sel}")
                ev_qua = p2.checkbox("Evitar QUA", value=bool(regras["qua"]), key=f"ev_qua_{sg_sel}")
                ev_qui = p2.checkbox("Evitar QUI", value=bool(regras["qui"]), key=f"ev_qui_{sg_sel}")
                ev_sex = p3.checkbox("Evitar SEX", value=bool(regras["sex"]), key=f"ev_sex_{sg_sel}")
                ev_sab = p3.checkbox("Evitar SÁB", value=bool(regras["sáb"]), key=f"ev_sab_{sg_sel}")

                if st.button("Salvar preferência do subgrupo (e readequar mês)", key="pref_save"):
                    set_subgrupo_regras(setor, sg_sel, {
                        "seg": int(ev_seg), "ter": int(ev_ter), "qua": int(ev_qua),
                        "qui": int(ev_qui), "sex": int(ev_sex), "sáb": int(ev_sab)
                    })
                    _regenerar_mes_inteiro(setor, ano, mes, seed=int(st.session_state.get("last_seed", 0)), respeitar_ajustes=True)
                    st.success("Preferência salva e escala readequada!")
                    st.rerun()
            else:
                st.info("Crie pelo menos 1 subgrupo na aba 👥 Colaboradores.")

        elif sec_aj == "📌 Subgrupos (editável)":
                st.markdown("## 📌 Subgrupos (editável)")
                subgrupos = list_subgrupos(setor)

                cA, cB = st.columns([1, 1])
                with cA:
                    novo_sub = st.text_input("Novo subgrupo:", key="sg_new")
                    if st.button("Adicionar subgrupo", key="sg_add"):
                        if novo_sub.strip():
                            add_subgrupo(setor, novo_sub.strip())
                            st.success("Subgrupo adicionado!")
                            st.rerun()
                        else:
                            st.error("Digite o nome do subgrupo.")

                with cB:
                    if subgrupos:
                        del_sel = st.selectbox("Remover subgrupo:", ["(nenhum)"] + subgrupos, key="sg_del")
                        if del_sel != "(nenhum)" and st.button("Remover", key="sg_del_btn"):
                            delete_subgrupo(setor, del_sel)
                            _regenerar_mes_inteiro(setor, ano, mes, seed=int(st.session_state.get("last_seed", 0)), respeitar_ajustes=True)
                            st.success("Subgrupo removido e escala readequada!")
                            st.rerun()
                    else:
                        st.caption("Nenhum subgrupo cadastrado.")

    # ------------------------------------------------------
    # ABA 4: Férias
    # ------------------------------------------------------
    elif sec_main == "🏖️ Férias":
        _status_comp_fer = get_status_competencia(setor, int(st.session_state['cfg_ano']), int(st.session_state['cfg_mes']))
        if _status_comp_fer == 'FECHADA':
            st.error('🔒 Competência fechada: lançamento normal de férias fica bloqueado nesta competência. Use retificação pontual se precisar corrigir histórico.')
        st.subheader("🏖️ Controle de Férias")

        st.markdown("---")
        st.markdown("---")
        colaboradores = load_colaboradores_setor(setor)
        if not colaboradores:
            st.warning("Sem colaboradores cadastrados.")
        else:
            sec_fer = st.radio("", ["🗺️ Mapa anual de férias", "➕ Lançar Férias", "📊 Controle (histórico)", "📋 Férias cadastradas", "❌ Remover férias"], horizontal=True, key="ferias_nav_fast", label_visibility="collapsed")

            # ---------------------------
            # TAB 1 — MAPA ANUAL
            # ---------------------------
            if sec_fer == "🗺️ Mapa anual de férias":
                st.markdown("## 🗺️ Mapa anual de férias (visual)")
                col_map1, col_map2 = st.columns([1, 3])
                ano_mapa = col_map1.number_input("Ano do mapa", value=int(st.session_state.get("cfg_ano", datetime.now().year)), step=1, key="fer_mapa_ano")
                col_map2.caption("Mostra em quais meses cada colaborador tem férias cadastradas (qualquer dia no mês marca o mês).")
                df_mapa = ferias_mapa_ano_df(setor, int(ano_mapa), colaboradores)
                show_chapa = st.checkbox("Mostrar coluna Chapa no mapa", value=False, key="fer_mapa_show_chapa")
                df_mapa_show = df_mapa if show_chapa else df_mapa.drop(columns=["Chapa"])
                st.dataframe(style_ferias_mapa(df_mapa_show), use_container_width=True, height=420)

            # ---------------------------
            # TAB 2 — LANÇAR
            # ---------------------------
            elif sec_fer == "➕ Lançar Férias":
                st.markdown("### ➕ Lançar Férias")
                opts = []
                for c in colaboradores:
                    chp = str(c.get("Chapa","")).strip()
                    nm = str(c.get("Nome","") or "").strip()
                    label = f"{chp} — {nm}" if nm else chp
                    opts.append((label, chp))
                pick = st.selectbox("Colaborador (chapa — nome):", [o[0] for o in opts], key="fer_pick")
                ch = next((o[1] for o in opts if o[0] == pick), pick.split("—")[0].strip())
                nome_sel = next((x.get("Nome","") for x in colaboradores if str(x.get("Chapa","")) == str(ch)), "")
                st.write(f"**Colaborador:** {nome_sel}  \n**Chapa:** {ch}")

                info_ult = get_ultima_ferias_info(setor, ch)
                ult_fim = info_ult.get("ultima_fim")
                meses_sem = info_ult.get("meses_desde_ultima_fim")
                if ult_fim:
                    st.write(
                        f"**Últimas férias:** {info_ult.get('ultima_inicio').strftime('%d/%m/%Y')} até {ult_fim.strftime('%d/%m/%Y')}  \n"
                        f"**Duração:** {_classificar_duracao_ferias(int(info_ult.get('dias_ultima') or 0))}  \n"
                        f"**Tempo desde o fim:** {int(meses_sem)} mês(es)"
                    )
                else:
                    st.warning("⚠️ Este colaborador ainda NÃO tem férias cadastradas.")

                c1, c2, c3 = st.columns(3)
                ini = c1.date_input("Início", value=datetime.now().date(), key="fer_ini")
                fim = c2.date_input("Fim", value=datetime.now().date(), key="fer_fim")
                if c3.button("Salvar férias (e readequar mês)", key="fer_add_btn"):
                    if fim < ini:
                        st.error("Data final não pode ser menor que a inicial.")
                    else:
                        add_ferias(setor, ch, ini, fim)
                        _regenerar_mes_inteiro(setor, int(st.session_state["cfg_ano"]), int(st.session_state["cfg_mes"]), seed=int(st.session_state.get("last_seed", 0)), respeitar_ajustes=True)
                        st.success("Férias adicionadas e escala readequada!")
                        st.rerun()

            # ---------------------------
            # TAB 3 — CONTROLE / HISTÓRICO
            # ---------------------------
            elif sec_fer == "📊 Controle (histórico)":
                st.markdown("### 📊 Controle de Férias (histórico por mês)")
                ano_ref = st.number_input(
                    "Ano para análise:",
                    min_value=2000, max_value=2100,
                    value=int(st.session_state.get("cfg_ano", datetime.now().year)),
                    step=1,
                    key="fer_hist_ano"
                )
                rows_all = list_ferias(setor)
                if not rows_all:
                    st.info("Nenhuma férias cadastrada para este setor.")
                else:
                    df_all = pd.DataFrame(rows_all, columns=["Chapa", "Início", "Fim"]).copy()
                    def _to_date(x):
                        try:
                            return pd.to_datetime(x).date()
                        except Exception:
                            return None
                    df_all["Início"] = df_all["Início"].apply(_to_date)
                    df_all["Fim"] = df_all["Fim"].apply(_to_date)
                    df_all = df_all.dropna(subset=["Início", "Fim"])
                    nome_by_hist = {str(c.get("Chapa","")): str(c.get("Nome","") or "") for c in (colaboradores or [])}
                    df_all["Nome"] = df_all["Chapa"].astype(str).map(nome_by_hist).fillna("")
                    resumo = []
                    for mes_i in range(1, 13):
                        ini_mes = pd.Timestamp(year=int(ano_ref), month=int(mes_i), day=1).date()
                        fim_mes = (pd.Timestamp(year=int(ano_ref), month=int(mes_i), day=1) + pd.offsets.MonthEnd(0)).date()
                        inter = df_all[(df_all["Fim"] >= ini_mes) & (df_all["Início"] <= fim_mes)].copy()
                        if inter.empty:
                            resumo.append({"Mês": mes_i, "Colaboradores em férias": 0, "Dias de férias (soma)": 0, "Períodos iniciados no mês": 0})
                            continue
                        dias_soma = 0
                        for _, r in inter.iterrows():
                            s = max(r["Início"], ini_mes)
                            e = min(r["Fim"], fim_mes)
                            dias_soma += max(0, int((e - s).days + 1))
                        iniciados = df_all[(df_all["Início"] >= ini_mes) & (df_all["Início"] <= fim_mes)]
                        resumo.append({
                            "Mês": mes_i,
                            "Colaboradores em férias": int(inter["Chapa"].nunique()),
                            "Dias de férias (soma)": int(dias_soma),
                            "Períodos iniciados no mês": int(iniciados.shape[0]),
                        })
                    df_res = pd.DataFrame(resumo)
                    try:
                        df_res["Mês (nome)"] = df_res["Mês"].apply(lambda m: pd.Timestamp(year=2000, month=int(m), day=1).strftime("%b").upper())
                        df_res = df_res[["Mês", "Mês (nome)", "Colaboradores em férias", "Dias de férias (soma)", "Períodos iniciados no mês"]]
                    except Exception:
                        pass
                    st.dataframe(df_res, use_container_width=True, height=360)
                    with st.expander("🔎 Ver detalhes de um mês"):
                        mes_det = st.selectbox("Mês:", list(range(1, 13)), index=0, key="fer_hist_mes_det")
                        ini_mes = pd.Timestamp(year=int(ano_ref), month=int(mes_det), day=1).date()
                        fim_mes = (pd.Timestamp(year=int(ano_ref), month=int(mes_det), day=1) + pd.offsets.MonthEnd(0)).date()
                        det = df_all[(df_all["Fim"] >= ini_mes) & (df_all["Início"] <= fim_mes)].copy()
                        if det.empty:
                            st.info("Nenhuma férias nesse mês.")
                        else:
                            det = det[["Chapa", "Nome", "Início", "Fim"]].sort_values(["Nome","Chapa"])
                            st.dataframe(det, use_container_width=True, height=360)

            # ---------------------------
            # TAB 4 — CADASTRADAS
            # ---------------------------
            elif sec_fer == "📋 Férias cadastradas":
                st.markdown("### 📋 Férias cadastradas")
                rows = list_ferias(setor)
                if rows:
                    df_f = pd.DataFrame(rows, columns=["Chapa", "Início", "Fim"])
                    nome_by = {str(c.get("Chapa","")): str(c.get("Nome","") or "") for c in (colaboradores or [])}
                    df_f.insert(1, "Nome", df_f["Chapa"].astype(str).map(nome_by).fillna(""))
                    st.dataframe(df_f, use_container_width=True, height=420)
                else:
                    st.info("Nenhuma férias cadastrada.")

            # ---------------------------
            # TAB 5 — REMOVER
            # ---------------------------
            elif sec_fer == "❌ Remover férias":
                st.markdown("### ❌ Remover férias")
                rows = list_ferias(setor)
                if not rows:
                    st.info("Nenhuma férias cadastrada.")
                else:
                    df_f = pd.DataFrame(rows, columns=["Chapa", "Início", "Fim"])
                    nome_by = {str(c.get("Chapa","")): str(c.get("Nome","") or "") for c in (colaboradores or [])}
                    df_f.insert(1, "Nome", df_f["Chapa"].astype(str).map(nome_by).fillna(""))
                    st.dataframe(df_f, use_container_width=True, height=260)
                    rem_idx = st.number_input("Linha para remover (1,2,3...)", min_value=1, max_value=len(df_f), value=1, key="fer_rem_idx")
                    if st.button("Remover linha (e readequar mês)", key="fer_rem_btn"):
                        r = df_f.iloc[int(rem_idx) - 1]
                        delete_ferias_row(setor, r["Chapa"], r["Início"], r["Fim"])
                        _regenerar_mes_inteiro(setor, int(st.session_state["cfg_ano"]), int(st.session_state["cfg_mes"]), seed=int(st.session_state.get("last_seed", 0)), respeitar_ajustes=True)
                        st.success("Férias removidas e escala readequada!")
                        st.rerun()

    elif sec_main == "🖨️ Impressão":
        sec_imp = st.radio("", ["📊 Excel modelo", "🗓️ Quem trabalha no dia", "📅 Escala", "🖨️ Imprimir escala parede"], horizontal=True, key="impressao_nav_fast", label_visibility="collapsed")

        # V94.2 — lazy load da impressão:
        # evita carregar escala + colaboradores + overrides logo ao abrir a aba.
        ano = int(st.session_state["cfg_ano"])
        mes = int(st.session_state["cfg_mes"])
        hist_db = {}
        colaboradores = []

        # V94.3 — estado/cache leve para exportação/impressão
        imp_state = st.session_state.setdefault("imp_state", {})
        excel_cache = imp_state.setdefault("excel_cache", {})
        dia_cache = imp_state.setdefault("dia_cache", {})
        parede_cache = imp_state.setdefault("parede_cache", {})

        if sec_imp == "📊 Excel modelo":
            st.subheader("📊 Excel modelo RH (separado por subgrupo)")
            st.caption("Geração pesada ficou sob demanda para deixar a aba Impressão rápida.")
            excel_key = f"{setor}|{ano}|{mes}"
            if st.button("📊 Gerar Excel", key="xls_btn"):
                st.session_state.pop("xls_cached_bytes", None)
                colaboradores = load_colaboradores_setor(setor)
                hist_db = load_escala_mes_db(setor, ano, mes) or {}
                colab_by = {str(c.get("Chapa", "")).strip(): c for c in colaboradores}
                if not hist_db:
                    st.info("Gere a escala.")
                else:
                    hist_db = apply_overrides_to_hist(setor, ano, mes, hist_db)
                    hist_db = _apply_retificacoes_to_hist(setor, ano, mes, hist_db)
                    from openpyxl import Workbook
                    output = io.BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Escala Mensal"

                    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", patternType="solid")
                    fill_dom = PatternFill(start_color="C00000", end_color="C00000", patternType="solid")
                    fill_folga = PatternFill(start_color="FFF2CC", end_color="FFF2CC", patternType="solid")
                    fill_nome = PatternFill(start_color="D9E1F2", end_color="D9E1F2", patternType="solid")
                    fill_ferias = PatternFill(start_color="92D050", end_color="92D050", patternType="solid")
                    fill_group = PatternFill(start_color="BDD7EE", end_color="BDD7EE", patternType="solid")

                    font_header = Font(color="FFFFFF", bold=True)
                    font_dom = Font(color="FFFFFF", bold=True)
                    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
                    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    ch0 = list(hist_db.keys())[0]
                    df_ref_xls = hist_db[ch0].copy().reset_index(drop=True)
                    total_dias = len(df_ref_xls)

                    c = ws.cell(1, 1, "COLABORADOR")
                    c.fill = fill_header
                    c.font = font_header
                    c.alignment = center
                    c.border = border
                    c = ws.cell(2, 1, "")
                    c.fill = fill_header
                    c.alignment = center
                    c.border = border

                    for i in range(total_dias):
                        data_i = pd.to_datetime(df_ref_xls.iloc[i].get("Data"), errors="coerce")
                        dia_num = int(data_i.day) if pd.notna(data_i) else (i + 1)
                        dia_sem = str(df_ref_xls.iloc[i].get("Dia", ""))
                        cA = ws.cell(1, i + 2, dia_num)
                        cB = ws.cell(2, i + 2, dia_sem)
                        if dia_sem == "dom":
                            cA.fill = fill_dom
                            cB.fill = fill_dom
                            cA.font = font_dom
                            cB.font = font_dom
                        else:
                            cA.fill = fill_header
                            cB.fill = fill_header
                            cA.font = font_header
                            cB.font = font_header
                        cA.alignment = center
                        cB.alignment = center
                        cA.border = border
                        cB.border = border
                        ws.column_dimensions[get_column_letter(i + 2)].width = 7
                    ws.column_dimensions["A"].width = 36

                    subgrupo_map = {}
                    for chx in hist_db.keys():
                        ch_str = str(chx).strip()
                        df_sg = hist_db.get(chx)
                        sg = ""
                        try:
                            if df_sg is not None and "Subgrupo" in df_sg.columns:
                                vals_sg = [str(v).strip() for v in df_sg["Subgrupo"].astype(str).tolist() if str(v).strip()]
                                if vals_sg:
                                    sg = vals_sg[-1]
                        except Exception:
                            sg = ""
                        sg = sg or get_subgrupo_competencia_ou_base(
                            setor, ch_str, int(ano), int(mes),
                            (colab_by.get(ch_str, {}).get("Subgrupo", "") or "").strip()
                        ) or "SEM SUBGRUPO"
                        subgrupo_map.setdefault(sg, []).append(chx)

                    row_idx = 3
                    total_linhas_gravadas = 0
                    resumo_cobertura = {
                        "abertura": [0] * total_dias,
                        "intermediario": [0] * total_dias,
                        "fechamento": [0] * total_dias,
                        "total_trabalhando": [0] * total_dias,
                    }

                    def _minutos_hora_excel(v):
                        s = str(v or "").strip()
                        if not s or ":" not in s:
                            return None
                        try:
                            hh, mm = s.split(":", 1)
                            return int(hh) * 60 + int(mm)
                        except Exception:
                            return None

                    for sg in sorted(subgrupo_map.keys()):
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_dias + 1)
                        t = ws.cell(row_idx, 1, f"SUBGRUPO: {sg}")
                        t.fill = fill_group
                        t.font = Font(bold=True)
                        t.alignment = Alignment(horizontal="left", vertical="center")
                        t.border = border
                        row_idx += 1

                        chapas_sg = sorted(subgrupo_map[sg], key=lambda chx: str(colab_by.get(str(chx), {}).get("Nome", chx)))
                        resumo_sg = {
                            "abertura": [0] * total_dias,
                            "intermediario": [0] * total_dias,
                            "fechamento": [0] * total_dias,
                            "total_trabalhando": [0] * total_dias,
                        }

                        for chx in chapas_sg:
                            df_f = hist_db[chx].copy().reset_index(drop=True)
                            nome = str(colab_by.get(str(chx), {}).get("Nome", chx))
                            c_nome = ws.cell(row=row_idx, column=1, value=f"{nome}\nCHAPA: {chx}")
                            c_nome.fill = fill_nome
                            c_nome.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                            c_nome.border = border
                            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx + 1, end_column=1)
                            for i, row in df_f.iterrows():
                                dia_sem = str(row.get("Dia", ""))
                                status = str(row.get("Status", ""))
                                if status == "Férias":
                                    v1, v2 = "FÉRIAS", ""
                                elif status == "Folga":
                                    v1, v2 = "F", ""
                                else:
                                    v1 = str(row.get("H_Entrada", "") or "")
                                    v2 = str(row.get("H_Saida", "") or "")
                                cell1 = ws.cell(row_idx, i + 2, v1)
                                cell2 = ws.cell(row_idx + 1, i + 2, v2)
                                cell1.alignment = center
                                cell2.alignment = center
                                cell1.border = border
                                cell2.border = border
                                if status == "Férias":
                                    cell1.fill = fill_ferias
                                    cell2.fill = fill_ferias
                                elif status == "Folga":
                                    if dia_sem == "dom":
                                        cell1.fill = fill_dom
                                        cell2.fill = fill_dom
                                    else:
                                        cell1.fill = fill_folga
                                        cell2.fill = fill_folga
                                else:
                                    ent_min = _minutos_hora_excel(v1)
                                    if ent_min is not None:
                                        resumo_cobertura["total_trabalhando"][i] += 1
                                        resumo_sg["total_trabalhando"][i] += 1
                                        if 360 <= ent_min <= 600:
                                            resumo_cobertura["abertura"][i] += 1
                                            resumo_sg["abertura"][i] += 1
                                        elif 601 <= ent_min <= 739:
                                            resumo_cobertura["intermediario"][i] += 1
                                            resumo_sg["intermediario"][i] += 1
                                        elif ent_min >= 740:
                                            resumo_cobertura["fechamento"][i] += 1
                                            resumo_sg["fechamento"][i] += 1
                            total_linhas_gravadas += 1
                            row_idx += 2

                        resumo_rows_sg = [
                            (f"ABERTURA — {sg}", resumo_sg["abertura"]),
                            (f"INTERMEDIÁRIO — {sg}", resumo_sg["intermediario"]),
                            (f"FECHAMENTO — {sg}", resumo_sg["fechamento"]),
                            (f"TOTAL TRABALHANDO — {sg}", resumo_sg["total_trabalhando"]),
                        ]
                        for titulo_resumo, valores_resumo in resumo_rows_sg:
                            c0 = ws.cell(row_idx, 1, titulo_resumo)
                            c0.fill = fill_group
                            c0.font = Font(bold=True)
                            c0.alignment = Alignment(horizontal="left", vertical="center")
                            c0.border = border
                            for i, valor_resumo in enumerate(valores_resumo):
                                c1 = ws.cell(row_idx, i + 2, int(valor_resumo))
                                c1.alignment = center
                                c1.border = border
                                if str(df_ref_xls.iloc[i].get("Dia", "")) == "dom":
                                    c1.fill = fill_folga
                            row_idx += 1
                        row_idx += 1

                    row_idx += 1
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_dias + 1)
                    t_res = ws.cell(row_idx, 1, "RESUMO DE COBERTURA — TODOS OS SUBGRUPOS")
                    t_res.fill = fill_header
                    t_res.font = font_header
                    t_res.alignment = Alignment(horizontal="left", vertical="center")
                    t_res.border = border
                    row_idx += 1

                    resumo_rows = [
                        ("ABERTURA (06:00 até 10:00)", resumo_cobertura["abertura"]),
                        ("INTERMEDIÁRIO (10:01 até 12:19)", resumo_cobertura["intermediario"]),
                        ("FECHAMENTO (a partir de 12:20)", resumo_cobertura["fechamento"]),
                        ("TOTAL TRABALHANDO", resumo_cobertura["total_trabalhando"]),
                    ]
                    for titulo_resumo, valores_resumo in resumo_rows:
                        c0 = ws.cell(row_idx, 1, titulo_resumo)
                        c0.fill = fill_group
                        c0.font = Font(bold=True)
                        c0.alignment = Alignment(horizontal="left", vertical="center")
                        c0.border = border
                        for i, valor_resumo in enumerate(valores_resumo):
                            c1 = ws.cell(row_idx, i + 2, int(valor_resumo))
                            c1.alignment = center
                            c1.border = border
                            if str(df_ref_xls.iloc[i].get("Dia", "")) == "dom":
                                c1.fill = fill_folga
                        row_idx += 1

                    try:
                        rows_f = list_ferias(setor) or []
                        if rows_f:
                            ws_f = wb.create_sheet("Férias do mês")
                            headers_f = ["Chapa", "Nome", "Sai de férias", "Volta ao trabalho", "Início", "Fim", "Dias de férias no mês"]
                            for col_idx, head in enumerate(headers_f, start=1):
                                c = ws_f.cell(1, col_idx, head)
                                c.fill = fill_header
                                c.font = font_header
                                c.border = border
                                c.alignment = center
                            df_fer = pd.DataFrame(rows_f, columns=["Chapa", "Início", "Fim"]).copy()
                            df_fer["Início"] = pd.to_datetime(df_fer["Início"], errors="coerce").dt.date
                            df_fer["Fim"] = pd.to_datetime(df_fer["Fim"], errors="coerce").dt.date
                            df_fer = df_fer.dropna(subset=["Início", "Fim"])
                            ini_mes = pd.Timestamp(year=int(ano), month=int(mes), day=1).date()
                            fim_mes = (pd.Timestamp(year=int(ano), month=int(mes), day=1) + pd.offsets.MonthEnd(0)).date()
                            df_fer = df_fer[(df_fer["Fim"] >= ini_mes) & (df_fer["Início"] <= fim_mes)].copy()
                            if not df_fer.empty:
                                nome_by = {str(c.get("Chapa", "")): str(c.get("Nome", "") or "") for c in (colaboradores or [])}
                                df_fer["Nome"] = df_fer["Chapa"].astype(str).map(nome_by).fillna("")
                                df_fer["Sai de férias"] = df_fer["Início"]
                                df_fer["Volta ao trabalho"] = df_fer["Fim"].apply(lambda d: (pd.Timestamp(d) + pd.Timedelta(days=1)).date())
                                def _dias_no_mes(r):
                                    s = max(r["Início"], ini_mes)
                                    e = min(r["Fim"], fim_mes)
                                    return max(0, int((e - s).days + 1))
                                df_fer["Dias de férias no mês"] = df_fer.apply(_dias_no_mes, axis=1)
                                df_fer = df_fer[["Chapa", "Nome", "Sai de férias", "Volta ao trabalho", "Início", "Fim", "Dias de férias no mês"]].sort_values(["Nome", "Chapa"])
                                for row_excel, vals in enumerate(df_fer.itertuples(index=False, name=None), start=2):
                                    for col_excel, val in enumerate(vals, start=1):
                                        c = ws_f.cell(row_excel, col_excel, val)
                                        c.border = border
                    except Exception:
                        pass

                    ws.freeze_panes = "B3"
                    wb.active = wb.sheetnames.index("Escala Mensal")
                    wb.save(output)
                    output.seek(0)
                    excel_bytes = output.getvalue()
                    if total_linhas_gravadas > 0 and excel_bytes and len(excel_bytes) > 2000:
                        excel_cache[excel_key] = excel_bytes
                        st.session_state["xls_cached_bytes"] = excel_bytes
                        st.success(f"Excel gerado com {total_linhas_gravadas} colaborador(es).")
                    else:
                        st.error("Excel gerado vazio. A escala do mês não trouxe linhas para exportação.")
            if st.session_state.get("xls_cached_bytes"):
                st.download_button(
                    "📥 Baixar Excel",
                    data=st.session_state["xls_cached_bytes"],
                    file_name=f"escala_{setor}_{mes:02d}_{ano}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="xls_down"
                )
        elif sec_imp == "🗓️ Quem trabalha no dia":
            # --- Lista (e PDF) de quem TRABALHA no dia escolhido ---
            st.markdown("### 🗓️ Quem trabalha no dia (para impressão)")
            st.caption("Carregamento sob demanda para deixar a aba Impressão rápida.")
            try:
                dias_mes = calendar.monthrange(int(ano), int(mes))[1]
            except Exception:
                dias_mes = 31
            dia_sel = st.number_input("Dia do mês", min_value=1, max_value=int(dias_mes), value=1, step=1)
            dia_key = f"{setor}|{ano}|{mes}|{int(dia_sel)}"
            carregar_dia = st.button("🔎 Carregar lista do dia", key="dia_trabalho_load_btn")

            if carregar_dia and dia_key not in dia_cache:
                colaboradores = load_colaboradores_setor(setor)
                hist_db = load_escala_mes_db(setor, ano, mes) or {}
                if hist_db:
                    hist_db = apply_overrides_to_hist(setor, ano, mes, hist_db)
                    hist_db = _apply_retificacoes_to_hist(setor, ano, mes, hist_db)
                linhas = []
                for _chapa, _df in (hist_db or {}).items():
                    if _df is None or _df.empty:
                        continue
                    try:
                        _linha = _df.loc[_df["Data"].dt.day == int(dia_sel)].head(1)
                    except Exception:
                        _linha = _df.loc[pd.to_datetime(_df["Data"], errors="coerce").dt.day == int(dia_sel)].head(1)
                    if _linha.empty:
                        continue
                    _r = _linha.iloc[0].to_dict()
                    _stt = str(_r.get("Status", "")).strip()
                    if _stt not in WORK_STATUSES:
                        continue
                    _ent = str(_r.get("H_Entrada", "") or "").strip()
                    _sai = str(_r.get("H_Saida", "") or "").strip()
                    _nome = ""
                    _subg_base = ""
                    for c in colaboradores:
                        if str(c.get("Chapa", "")).strip() == str(_chapa).strip():
                            _nome = str(c.get("Nome", "")).strip()
                            _subg_base = str(c.get("Subgrupo", "")).strip()
                            break
                    _subg = get_subgrupo_competencia_ou_base(setor, str(_chapa).strip(), int(ano), int(mes), _subg_base)
                    linhas.append({"Chapa": str(_chapa).strip(), "Nome": _nome, "Subgrupo": _subg, "Entrada": _ent, "Saída": _sai})
                df_dia = pd.DataFrame(linhas).sort_values(["Subgrupo", "Nome"]) if linhas else pd.DataFrame(columns=["Chapa","Nome","Subgrupo","Entrada","Saída"])
                dia_cache[dia_key] = {"df": df_dia, "hist": hist_db, "colaboradores": colaboradores}

            payload_dia = dia_cache.get(dia_key, {"df": pd.DataFrame(columns=["Chapa","Nome","Subgrupo","Entrada","Saída"]), "hist": {}, "colaboradores": []})
            df_dia = payload_dia.get("df") if isinstance(payload_dia.get("df"), pd.DataFrame) else pd.DataFrame(columns=["Chapa","Nome","Subgrupo","Entrada","Saída"])
            st.dataframe(df_dia, use_container_width=True, hide_index=True)

            colp1, colp2 = st.columns([1, 2])
            pdf_day_key = f"pdf::{dia_key}"
            with colp1:
                if st.button("📄 Gerar PDF (quem trabalha no dia)"):
                    if df_dia.empty:
                        st.warning("Não há colaboradores trabalhando nesse dia (ou ainda não foi gerado para este mês).")
                    else:
                        pdf_bytes_dia = gerar_pdf_trabalhando_no_dia(setor, int(ano), int(mes), int(dia_sel), payload_dia.get("hist", {}), payload_dia.get("colaboradores", []))
                        st.session_state[pdf_day_key] = pdf_bytes_dia
                        st.success("PDF pronto.")
            with colp2:
                if st.session_state.get(pdf_day_key):
                    st.download_button(
                        "⬇️ Baixar PDF (quem trabalha no dia)",
                        data=st.session_state[pdf_day_key],
                        file_name=f"escala_trabalhando_dia_{int(dia_sel):02d}_{int(mes):02d}_{int(ano)}.pdf",
                        mime="application/pdf",
                    )

        elif sec_imp == "📅 Escala":
            st.subheader("📅 Escala")
            st.markdown("---")
            st.markdown("### 🏖️ Férias do mês (PDF)")
            cfx1, cfx2 = st.columns([1, 2])
            pdf_fer_busca = cfx2.text_input("Filtro (nome ou chapa) — opcional:", value="", key="pdf_fer_busca")
            btn_fer_pdf = cfx1.button("📄 Gerar PDF — Férias do mês", use_container_width=True, key="pdf_fer_btn")
            cfx2.caption("Gera um relatório A4 com Nome, Chapa, Início, Fim e Dias. Considera quem tem férias que encostam no mês selecionado.")
            if btn_fer_pdf:
                colabs_all = load_colaboradores_setor(setor) or []
                # aplica filtro simples
                if pdf_fer_busca.strip():
                    kw = pdf_fer_busca.strip().lower()
                    colabs_all = [c for c in colabs_all if kw in str(c.get("Nome","")).lower() or kw in str(c.get("Chapa","")).lower()]
                pdf_bytes = gerar_pdf_ferias_mes(setor, int(ano), int(mes), load_colaboradores_setor(setor) or [], keyword=pdf_fer_busca)
                st.download_button(
                    "⬇️ Baixar PDF (Férias do mês)",
                    data=pdf_bytes,
                    file_name=f"ferias_{setor}_{int(mes):02d}_{int(ano)}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="pdf_fer_dl"
                )
        elif sec_imp == "🖨️ Imprimir escala parede":
            st.subheader("🖨️ Imprimir escala parede")

            colaboradores = load_colaboradores_setor(setor)
            all_subgrupos = sorted({((c.get("Subgrupo") or "").strip() or "SEM SUBGRUPO") for c in colaboradores})
            cfx1, cfx2, cfx3 = st.columns([1.2, 1.2, 1.6])
            loja_txt = cfx1.text_input("Loja:", value=str(setor), key="pdf_loja_txt")
            secoes_sel = cfx2.multiselect("Seções (Subgrupo):", options=all_subgrupos, default=[], key="pdf_secoes_sel")
            busca_txt = cfx3.text_input("Filtro (nome/chapa/subgrupo):", value="", key="pdf_busca")

            modo_pdf = st.radio(
                "Formato de impressão:",
                options=["Modelo oficial do mês", "Panorâmico por período"],
                horizontal=True,
                key="pdf_modo_impressao"
            )

            cols_dates = st.columns([1, 1, 2])
            data_ini = cols_dates[0].date_input("Dia inicial:", value=date(int(ano), int(mes), 1), key="pdf_dt_ini")
            data_fim = cols_dates[1].date_input("Dia final:", value=date(int(ano), int(mes), calendar.monthrange(int(ano), int(mes))[1]), key="pdf_dt_fim")
            if modo_pdf == "Panorâmico por período":
                cols_dates[2].caption("Use qualquer período contínuo, inclusive dois meses juntos (ex.: 01/03/2026 até 30/04/2026).")
            else:
                cols_dates[2].caption("Obs.: o PDF segue o modelo oficial do mês. Aqui o filtro é para escolher colaboradores/Seções como no sistema.")

            colabs_filtrados = _filtrar_colaboradores(colaboradores, secoes_sel, busca_txt)

            opcoes = [
                f"{(c.get('Nome') or '').strip()} — Chapa: {str(c.get('Chapa') or '').strip()} — {((c.get('Subgrupo') or '').strip() or 'SEM SUBGRUPO')}"
                for c in colabs_filtrados
            ]
            mapa_idx = {opcoes[i]: colabs_filtrados[i] for i in range(len(opcoes))}

            st.markdown("### 👥 Colaboradores")
            sel = st.multiselect(
                "Selecione (se vazio, imprime TODOS do filtro):",
                options=opcoes,
                default=[],
                key="pdf_colabs_sel"
            )

            cbtn1, cbtn2 = st.columns([1, 3])
            gerar = cbtn1.button("🖨️ Imprimir (gerar PDF)", key="pdf_print_btn", use_container_width=True)
            cbtn2.caption("Dica: selecione uma seção, depois marque os colaboradores. Se não marcar nenhum, imprime todos os filtrados.")

            pdf_parede_key = f"{setor}|{ano}|{mes}|{loja_txt.strip()}|{modo_pdf}|{data_ini}|{data_fim}|{','.join(sorted(chapas_sel)) if 'chapas_sel' in locals() else ''}|{','.join(sorted(secoes_sel))}|{busca_txt.strip()}"
            if gerar:
                if data_fim < data_ini:
                    st.warning("O dia final precisa ser maior ou igual ao dia inicial.")
                else:
                    if sel:
                        chapas_sel = [str(mapa_idx[x].get("Chapa")) for x in sel if x in mapa_idx]
                    else:
                        chapas_sel = [str(c.get("Chapa")) for c in colabs_filtrados]
                    pdf_parede_key = f"{setor}|{ano}|{mes}|{loja_txt.strip()}|{modo_pdf}|{data_ini}|{data_fim}|{','.join(sorted(chapas_sel))}|{','.join(sorted(secoes_sel))}|{busca_txt.strip()}"
                    if pdf_parede_key in parede_cache:
                        st.session_state["pdf_parede_bytes"] = parede_cache[pdf_parede_key]
                    elif modo_pdf == "Panorâmico por período":
                        hist_db_pdf = _load_hist_periodo(setor, data_ini, data_fim)
                        hist_db_pdf = {ch: df for ch, df in hist_db_pdf.items() if ch in set(chapas_sel)}
                        if not hist_db_pdf:
                            st.warning("Nenhum colaborador com escala salva no período informado.")
                        else:
                            pdf_bytes = gerar_pdf_periodo_panoramico(loja_txt.strip() or str(setor), data_ini, data_fim, hist_db_pdf, colaboradores)
                            parede_cache[pdf_parede_key] = pdf_bytes
                            st.session_state["pdf_parede_bytes"] = pdf_bytes
                            st.download_button(
                                "⬇️ Baixar PDF panorâmico",
                                data=st.session_state["pdf_parede_bytes"],
                                file_name=f"escala_panoramica_{(loja_txt.strip() or str(setor))}_{data_ini.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.pdf",
                                mime="application/pdf",
                                key="pdf_down_pan"
                            )
                    else:
                        hist_db_pdf = load_escala_mes_db(setor, ano, mes)
                        if not hist_db_pdf:
                            st.warning("Gere a escala antes na aba 🚀 Gerar Escala.")
                        else:
                            hist_db_pdf = apply_overrides_to_hist(setor, ano, mes, hist_db_pdf)
                            hist_db_pdf = _apply_retificacoes_to_hist(setor, ano, mes, hist_db_pdf)
                            hist_db_pdf = {ch: df for ch, df in hist_db_pdf.items() if ch in set(chapas_sel)}
                            if not hist_db_pdf:
                                st.warning("Nenhum colaborador para imprimir com os filtros atuais.")
                            else:
                                pdf_bytes = gerar_pdf_modelo_oficial(loja_txt.strip() or str(setor), ano, mes, hist_db_pdf, colaboradores)
                                st.download_button(
                                    "⬇️ Baixar PDF",
                                    data=pdf_bytes,
                                    file_name=f"escala_{(loja_txt.strip() or str(setor))}_{mes:02d}_{ano}.pdf",
                                    mime="application/pdf",
                                    key="pdf_down"
                                )


    # ------------------------------------------------------
    # ABA 6: Assinaturas (líder/admin)
    # ------------------------------------------------------
    elif sec_main == "✍️ Assinaturas":
        ano = int(st.session_state["cfg_ano"])
        mes = int(st.session_state["cfg_mes"])
        hoje_ass = datetime.now()
        ano_vig_ass = int(hoje_ass.year)
        mes_vig_ass = int(hoje_ass.month)

        st.subheader(f"✍️ Assinaturas do setor — {setor}")

        df_ass_sel = list_assinaturas_setor(setor, ano, mes)
        df_ass_vig = list_assinaturas_setor(setor, ano_vig_ass, mes_vig_ass)
        df_ass_all = list_assinaturas_setor_todas(setor)
        colaboradores_setor = load_colaboradores_setor(setor) or []
        total_colabs_setor = len({str((c or {}).get("Chapa", "")).strip() for c in colaboradores_setor if str((c or {}).get("Chapa", "")).strip()})

        def _filtrar_assinatura_escala_mes(df_src: pd.DataFrame) -> pd.DataFrame:
            if df_src is None or df_src.empty:
                return pd.DataFrame(columns=getattr(df_src, 'columns', []))
            if 'Tipo' not in df_src.columns:
                return df_src.copy()
            tipo_norm = df_src['Tipo'].astype(str).str.strip().str.lower()
            df_oficial = df_src[tipo_norm.isin(['oficial', 'assinatura da escala do mês'])].copy()
            return df_oficial if not df_oficial.empty else df_src.copy()

        df_ass_sel_escala = _filtrar_assinatura_escala_mes(df_ass_sel)
        df_ass_vig_escala = _filtrar_assinatura_escala_mes(df_ass_vig)
        chapas_ass_sel = {str(x).strip() for x in df_ass_sel_escala.get('Chapa', pd.Series(dtype=str)).astype(str).tolist() if str(x).strip()}
        chapas_ass_vig = {str(x).strip() for x in df_ass_vig_escala.get('Chapa', pd.Series(dtype=str)).astype(str).tolist() if str(x).strip()}
        faltam_sel = max(0, total_colabs_setor - len(chapas_ass_sel))
        faltam_vig = max(0, total_colabs_setor - len(chapas_ass_vig))

        c_ass1, c_ass2, c_ass3 = st.columns(3)
        c_ass1.metric("Competência selecionada", f"{mes:02d}/{ano}", delta=f"{len(df_ass_sel)} assinatura(s)")
        c_ass2.metric("Mês vigente", f"{mes_vig_ass:02d}/{ano_vig_ass}", delta=f"{len(df_ass_vig)} assinatura(s)")
        c_ass3.metric("Total do setor", len(df_ass_all))

        c_ass4, c_ass5, c_ass6 = st.columns(3)
        c_ass4.metric("Colaboradores do setor", total_colabs_setor)
        c_ass5.metric("Assinaram escala do mês vigente", len(chapas_ass_vig))
        c_ass6.metric("Faltam assinar mês vigente", faltam_vig)

        escopo_opts = [
            f"Competência selecionada ({mes:02d}/{ano})",
            f"Mês vigente ({mes_vig_ass:02d}/{ano_vig_ass})",
            "Todas do setor",
        ]
        if not df_ass_sel.empty:
            escopo_default = 0
        elif not df_ass_vig.empty:
            escopo_default = 1
        else:
            escopo_default = 2

        escopo_ass = st.radio(
            "Visualizar",
            escopo_opts,
            index=escopo_default,
            horizontal=True,
            key="ass_setor_escopo",
        )

        if escopo_ass == escopo_opts[0]:
            df_ass = df_ass_sel.copy()
            st.caption(f"Mostrando a competência selecionada na lateral: {mes:02d}/{ano}.")
            if df_ass.empty and not df_ass_vig.empty and (ano != ano_vig_ass or mes != mes_vig_ass):
                st.warning(
                    f"Não há assinaturas em {mes:02d}/{ano}. Existem assinatura(s) no mês vigente {mes_vig_ass:02d}/{ano_vig_ass}."
                )
        elif escopo_ass == escopo_opts[1]:
            df_ass = df_ass_vig.copy()
            st.caption(f"Mostrando o mês vigente do portal do colaborador: {mes_vig_ass:02d}/{ano_vig_ass}.")
        else:
            df_ass = df_ass_all.copy()
            st.caption("Mostrando todas as assinaturas do setor, sem apagar nem alterar a lógica existente.")

        if df_ass.empty:
            st.info("Nenhuma assinatura encontrada para o filtro selecionado.")
        else:
            m1, m2, m3 = st.columns(3)
            m1.metric("Assinaturas", len(df_ass))
            m2.metric("Colaboradores com assinatura", int(df_ass['Chapa'].astype(str).nunique()))
            m3.metric("Tipos assinados", int(df_ass['Tipo'].astype(str).nunique()))

            tipo_opts = ["Todos"] + sorted(df_ass['Tipo'].astype(str).dropna().unique().tolist())
            tipo_sel = st.selectbox("Filtrar tipo", tipo_opts, key="ass_setor_tipo")
            df_view = df_ass.copy()
            if tipo_sel != "Todos":
                df_view = df_view[df_view['Tipo'].astype(str) == str(tipo_sel)].copy()

            comp_opts = ["Todas"] + sorted(
                {f"{int(a):04d}-{int(m):02d}" for a, m in zip(df_view['Ano'].fillna(0), df_view['Mes'].fillna(0))}
            ) if {'Ano', 'Mes'}.issubset(df_view.columns) else ["Todas"]
            comp_sel = st.selectbox("Filtrar competência", comp_opts, key="ass_setor_competencia")
            if comp_sel != "Todas" and {'Ano', 'Mes'}.issubset(df_view.columns):
                ano_sel, mes_sel = comp_sel.split('-')
                df_view = df_view[
                    (df_view['Ano'].astype(int) == int(ano_sel)) &
                    (df_view['Mes'].astype(int) == int(mes_sel))
                ].copy()

            if 'Tipo' in df_view.columns:
                tipo_map = {
                    'oficial': 'Assinatura da Escala do Mês',
                    'historico': 'Assinatura de Mudanças',
                }
                df_view['Tipo'] = df_view['Tipo'].astype(str).map(lambda x: tipo_map.get(str(x).strip().lower(), x))

            if 'Assinado_em' in df_view.columns:
                try:
                    df_view['Assinado_em'] = pd.to_datetime(df_view['Assinado_em'], errors='coerce').dt.strftime('%d/%m/%Y %H:%M')
                except Exception:
                    pass
            st.dataframe(df_view, use_container_width=True, hide_index=True)

            faltantes_vig = [
                c for c in colaboradores_setor
                if str((c or {}).get('Chapa', '')).strip() and str((c or {}).get('Chapa', '')).strip() not in chapas_ass_vig
            ]
            faltantes_sel = [
                c for c in colaboradores_setor
                if str((c or {}).get('Chapa', '')).strip() and str((c or {}).get('Chapa', '')).strip() not in chapas_ass_sel
            ]

            with st.expander('📋 Conferência de quem ainda falta assinar', expanded=False):
                st.caption('Sem apagar nada da lógica existente: aqui o sistema compara os colaboradores cadastrados no setor com as assinaturas já registradas.')
                t1, t2 = st.columns(2)
                with t1:
                    st.markdown(f"**Competência selecionada ({mes:02d}/{ano})**")
                    st.write(f"Faltam assinar: **{faltam_sel}**")
                    if faltantes_sel:
                        st.dataframe(
                            pd.DataFrame([
                                {
                                    'Chapa': str((c or {}).get('Chapa', '')).strip(),
                                    'Nome': str((c or {}).get('Nome', '')).strip(),
                                    'Subgrupo': str((c or {}).get('Subgrupo', '')).strip(),
                                }
                                for c in faltantes_sel
                            ]),
                            use_container_width=True,
                            hide_index=True,
                        )
                    else:
                        st.success('Todos os colaboradores do setor já assinaram a escala desta competência.')
                with t2:
                    st.markdown(f"**Mês vigente ({mes_vig_ass:02d}/{ano_vig_ass})**")
                    st.write(f"Faltam assinar: **{faltam_vig}**")
                    if faltantes_vig:
                        st.dataframe(
                            pd.DataFrame([
                                {
                                    'Chapa': str((c or {}).get('Chapa', '')).strip(),
                                    'Nome': str((c or {}).get('Nome', '')).strip(),
                                    'Subgrupo': str((c or {}).get('Subgrupo', '')).strip(),
                                }
                                for c in faltantes_vig
                            ]),
                            use_container_width=True,
                            hide_index=True,
                        )
                    else:
                        st.success('Todos os colaboradores do setor já assinaram a escala do mês vigente.')

    # ------------------------------------------------------
    # ABA 7: Minhas solicitações (líder/admin)
    # ------------------------------------------------------
    elif sec_main == "📨 Minhas solicitações":
        st.subheader(f"📨 Solicitações recebidas do setor — {setor}")
        df_sol_setor = list_solicitacoes_setor(setor)
        if df_sol_setor.empty:
            st.info("Nenhuma solicitação enviada para este setor até agora.")
        else:
            pend = int((df_sol_setor['Status'].astype(str) == 'Em análise').sum()) if 'Status' in df_sol_setor.columns else 0
            aprov = int((df_sol_setor['Status'].astype(str) == 'Aprovado').sum()) if 'Status' in df_sol_setor.columns else 0
            rec = int((df_sol_setor['Status'].astype(str) == 'Recusado').sum()) if 'Status' in df_sol_setor.columns else 0
            s1, s2, s3 = st.columns(3)
            s1.metric('Em análise', pend)
            s2.metric('Aprovadas', aprov)
            s3.metric('Recusadas', rec)

            status_opts = ['Todos'] + sorted(df_sol_setor['Status'].astype(str).dropna().unique().tolist())
            filtro_status = st.selectbox('Filtrar status', status_opts, key='sol_setor_status')
            df_view = df_sol_setor.copy()
            if filtro_status != 'Todos':
                df_view = df_view[df_view['Status'].astype(str) == str(filtro_status)].copy()

            for c in ['Criado_em', 'Atualizado_em']:
                if c in df_view.columns:
                    try:
                        df_view[c] = pd.to_datetime(df_view[c], errors='coerce').dt.strftime('%d/%m/%Y %H:%M')
                    except Exception:
                        pass

            st.dataframe(df_view, use_container_width=True, hide_index=True)

            pendentes = df_sol_setor[df_sol_setor['Status'].astype(str) == 'Em análise'].copy() if 'Status' in df_sol_setor.columns else pd.DataFrame()
            if pendentes.empty:
                st.success('Não há solicitações pendentes no momento.')
            else:
                ids_pend = [int(x) for x in pendentes['ID'].tolist()]
                sol_id = st.selectbox('Selecionar solicitação pendente', ids_pend, key='sol_pendente_id')
                sol_row = pendentes[pendentes['ID'].astype(int) == int(sol_id)].head(1)
                if not sol_row.empty:
                    r = sol_row.iloc[0]
                    st.caption(f"{r.get('Nome','-')} • Chapa {r.get('Chapa','-')} • Data {r.get('Data','-')} • Tipo {r.get('Tipo','-')}")
                    mot = str(r.get('Motivo', '') or '').strip()
                    obs = str(r.get('Observação', '') or '').strip()
                    if mot:
                        st.write(f"**Motivo:** {mot}")
                    if obs:
                        st.write(f"**Observação:** {obs}")
                a1, a2 = st.columns(2)
                if a1.button('✅ Aprovar solicitação', key='aprovar_solicitacao_btn'):
                    atualizar_status_solicitacao(int(sol_id), 'Aprovado')
                    st.success('Solicitação aprovada com sucesso.')
                    st.rerun()
                if a2.button('❌ Recusar solicitação', key='recusar_solicitacao_btn'):
                    atualizar_status_solicitacao(int(sol_id), 'Recusado')
                    st.success('Solicitação recusada com sucesso.')
                    st.rerun()


    # ------------------------------------------------------
    # ABA 6: Admin (somente ADMIN)
    # ------------------------------------------------------
    elif is_admin_area and sec_main == "🔒 Admin":
            st.subheader("🔒 Admin do Sistema (somente ADMIN)")

            st.markdown("## 🛠️ ATUALIZAR FUNCIONÁRIO DE QUALQUER SETOR")
            st.warning("NOVO BLOCO ADMIN ATIVO: aqui você altera subgrupo e perfil do sistema do funcionário.")
            st.caption("Aqui o ADMIN pode alterar nome, subgrupo e perfil do colaborador em qualquer setor. O perfil sincroniza o login do sistema.")
            try:
                df_func_adm = admin_get_funcionarios_leve_all()
                df_login_adm = admin_get_logins_leve_all()
            except Exception:
                df_func_adm = pd.DataFrame(columns=['nome','setor','chapa','subgrupo','entrada','folga_sab'])
                df_login_adm = pd.DataFrame(columns=['setor','chapa','is_admin','is_lider','is_ax_lider'])

            if df_func_adm.empty:
                st.info("Nenhum colaborador cadastrado para atualizar.")
            else:
                setores_func = sorted({_norm_setor(x) for x in df_func_adm['setor'].dropna().tolist() if str(x).strip()})
                admf1, admf2 = st.columns([1, 1.7])
                with admf1:
                    setor_func = st.selectbox("Setor do funcionário", setores_func, key="adm_func_setor")
                df_func_setor = df_func_adm[df_func_adm['setor'].astype(str).str.strip().str.upper() == _norm_setor(setor_func)].copy()
                opts_func = [f"{str(r['nome']).strip()} ({str(r['chapa']).strip()})" for _, r in df_func_setor.iterrows()]
                with admf2:
                    pick_func = st.selectbox("Funcionário", opts_func, key="adm_func_pick") if opts_func else None

                rec_func = None
                chapa_func = ""
                if pick_func:
                    chapa_func = pick_func.rsplit("(", 1)[-1].replace(")", "").strip()
                    df_hit = df_func_setor[df_func_setor['chapa'].astype(str).str.strip() == chapa_func]
                    if not df_hit.empty:
                        rec_func = df_hit.iloc[0].to_dict()

                if rec_func:
                    login_hit = df_login_adm[(df_login_adm['setor'].astype(str).str.strip().str.upper() == _norm_setor(setor_func)) & (df_login_adm['chapa'].astype(str).str.strip() == chapa_func)]
                    login_row = login_hit.iloc[0] if not login_hit.empty else {}
                    is_admin_cur = bool(int(login_row.get('is_admin', 0) or 0)) if hasattr(login_row, 'get') else False
                    is_lider_cur = bool(int(login_row.get('is_lider', 0) or 0)) if hasattr(login_row, 'get') else False
                    is_ax_cur = bool(int(login_row.get('is_ax_lider', 0) or 0)) if hasattr(login_row, 'get') else False
                    perfil_cur = 'ADMIN' if is_admin_cur else ('LIDER' if is_lider_cur else ('AX_LIDER' if is_ax_cur else ('LIDER' if _norm_subgrupo_label(rec_func.get('subgrupo','')) == 'LIDERANCA' else 'COLABORADOR')))

                    st.write(f"Atualizando: **{str(rec_func.get('nome') or '').strip()}** — chapa **{chapa_func}**")
                    af1, af2, af3, af4 = st.columns([1.4, 1.2, 1.2, 1])
                    with af1:
                        nome_func_novo = st.text_input("Nome", value=str(rec_func.get('nome') or '').strip(), key='adm_func_nome')
                    with af2:
                        subgrupo_func_novo = st.text_input("Subgrupo", value=str(rec_func.get('subgrupo') or '').strip(), key='adm_func_subgrupo')
                    with af3:
                        entrada_func_nova = st.text_input("Entrada padrão", value=str(rec_func.get('entrada') or '06:00').strip() or '06:00', key='adm_func_entrada')
                    with af4:
                        folga_sab_func = st.checkbox("Folga sábado", value=bool(int(rec_func.get('folga_sab', 0) or 0)), key='adm_func_folga_sab')

                    perfil_func_novo = st.selectbox("Perfil do sistema", ['COLABORADOR', 'AX_LIDER', 'LIDER', 'ADMIN'], index=['COLABORADOR', 'AX_LIDER', 'LIDER', 'ADMIN'].index(perfil_cur), key='adm_func_perfil')
                    criar_login_func = st.checkbox("Criar login do sistema se não existir", value=True, key='adm_func_criar_login')

                    if st.button("Salvar atualização do funcionário", key='adm_func_salvar'):
                        try:
                            res = admin_update_funcionario(
                                setor=setor_func,
                                chapa_atual=chapa_func,
                                nome_novo=nome_func_novo,
                                subgrupo_novo=subgrupo_func_novo,
                                perfil_novo=perfil_func_novo,
                                entrada_nova=entrada_func_nova,
                                folga_sab=bool(folga_sab_func),
                                criar_usuario_se_nao_existir=bool(criar_login_func),
                            )
                            st.success(f"Funcionário atualizado com sucesso. Perfil final: {res['perfil']} | Subgrupo: {res['subgrupo'] or 'SEM SUBGRUPO'}")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Falha ao atualizar funcionário: {e}")

            st.markdown("---")
            dfu = admin_list_users()
            st.dataframe(dfu, use_container_width=True, height=420)

            st.markdown("### Resetar senha de um usuário")
            if not dfu.empty:
                uid = st.selectbox("ID do usuário", dfu["id"].tolist(), key="adm_uid")
                newp = st.text_input("Nova senha", type="password", key="adm_newp")
                if st.button("Resetar senha", key="adm_reset"):
                    if not newp:
                        st.error("Digite a senha.")
                    else:
                        ok = admin_reset_user_password(int(uid), newp)
                        st.success("Senha resetada!" if ok else "Falha.")
                        st.rerun()

            st.markdown("---")
            st.subheader("♻️ Recuperar usuário do sistema")
            st.caption("Use esta área quando o colaborador existe, mas sumiu do login. Se não existir colaborador, use o cadastro manual logo abaixo.")
            con = db_conn()
            df_colabs_adm = pd.read_sql_query("SELECT nome, setor, chapa FROM colaboradores ORDER BY setor, nome", con)
            con.close()
            if df_colabs_adm.empty:
                st.info("Nenhum colaborador cadastrado para recuperar. Use o cadastro manual de usuário abaixo.")
            else:
                colr1, colr2, colr3 = st.columns([1.1, 1.2, 1.0])
                with colr1:
                    setores_rec = sorted({_norm_setor(x) for x in df_colabs_adm["setor"].dropna().tolist() if str(x).strip()})
                    setor_rec = st.selectbox("Setor do colaborador", setores_rec, key="adm_rec_setor")
                df_setor_rec = df_colabs_adm[df_colabs_adm["setor"].astype(str).str.strip().str.upper() == _norm_setor(setor_rec)].copy()
                opts_rec = [f"{str(r['nome']).strip()} ({str(r['chapa']).strip()})" for _, r in df_setor_rec.iterrows()]
                with colr2:
                    pick_rec = st.selectbox("Colaborador", opts_rec, key="adm_rec_pick") if opts_rec else None
                with colr3:
                    senha_rec = st.text_input("Nova senha do usuário", type="password", key="adm_rec_pwd")
                if st.button("Recuperar / recriar usuário", key="adm_rec_btn"):
                    if not pick_rec or not senha_rec:
                        st.error("Selecione o colaborador e digite a senha.")
                    else:
                        chapa_rec = pick_rec.rsplit("(", 1)[-1].replace(")", "").strip()
                        ok = recover_system_user_from_colaborador(setor_rec, chapa_rec, senha_rec)
                        if ok:
                            try:
                                st.cache_data.clear()
                            except Exception:
                                pass
                            st.success("Usuário recuperado com sucesso.")
                            st.rerun()
                        else:
                            st.error("Não encontrei esse colaborador para recuperar.")

            st.markdown("### ➕ Cadastro manual de usuário do sistema")
            cman1, cman2, cman3 = st.columns([1, 1, 1])
            with cman1:
                setor_man = st.text_input("Setor do usuário", value="FLV", key="adm_man_setor")
            with cman2:
                chapa_man = st.text_input("Chapa do usuário", key="adm_man_chapa")
            with cman3:
                nome_man = st.text_input("Nome do usuário", key="adm_man_nome")
            senha_man = st.text_input("Senha do usuário", type="password", key="adm_man_pwd", help="Se deixar em branco, a senha padrão será a própria chapa sem símbolos.")
            cman4, cman5, cman6 = st.columns([1, 1, 1])
            with cman4:
                lider_man = st.checkbox("É líder", value=False, key="adm_man_lider")
            with cman5:
                admin_man = st.checkbox("É admin", value=False, key="adm_man_admin")
            with cman6:
                criar_colab_man = st.checkbox("Criar colaborador junto", value=True, key="adm_man_colab")
            if st.button("Salvar usuário manualmente", key="adm_man_btn"):
                setor_norm = _norm_setor(setor_man)
                chapa_norm = _norm_chapa(chapa_man)
                nome_final = (nome_man or chapa_norm).strip()
                senha_final = (senha_man or default_password_for_chapa(chapa_norm)).strip()
                if not setor_norm or not chapa_norm:
                    st.error("Digite setor e chapa.")
                else:
                    try:
                        if criar_colab_man and not colaborador_exists(setor_norm, chapa_norm):
                            create_colaborador(nome_final, setor_norm, chapa_norm, criar_login=False)
                        create_system_user(nome_final, setor_norm, chapa_norm, senha_final, is_lider=int(lider_man), is_admin=int(admin_man), is_ax_lider=0)
                        try:
                            st.cache_data.clear()
                        except Exception:
                            pass
                        st.success(f"Usuário salvo com sucesso. Senha ativa: {senha_final}")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Falha ao salvar usuário: {e}")

            st.markdown("---")
            st.subheader("🏷️ Renomear setor")
            st.caption("Use esta subárea para trocar o nome de um setor em todo o sistema sem precisar editar tabela por tabela.")
            try:
                setores_ren = listar_setores_db()
            except Exception:
                setores_ren = []
            rr1, rr2 = st.columns([1.2, 1.4])
            with rr1:
                setor_ren_atual = st.selectbox("Setor atual", setores_ren, key="adm_ren_setor_atual") if setores_ren else st.text_input("Setor atual", value="FLV", key="adm_ren_setor_atual_txt")
            with rr2:
                setor_ren_novo = st.text_input("Novo nome do setor", value=str(setor_ren_atual or ''), key="adm_ren_setor_novo")
            st.caption("Isso atualiza o nome do setor nas tabelas que possuem a coluna setor, incluindo colaboradores, usuários, escala, retificações, assinaturas e competências.")
            if st.button("Renomear setor", key="adm_ren_setor_btn"):
                try:
                    res = admin_rename_setor_global(str(setor_ren_atual), str(setor_ren_novo))
                    st.success(f"Setor renomeado de {res['setor_antigo']} para {res['setor_novo']}. Tabelas afetadas: {res['total_tabelas']} | Registros atualizados: {res['total_registros']}")
                    if res['tabelas_atualizadas']:
                        st.dataframe(pd.DataFrame(res['tabelas_atualizadas'], columns=['Tabela', 'Registros atualizados']), use_container_width=True, hide_index=True)
                    st.rerun()
                except Exception as e:
                    st.error(f"Falha ao renomear setor: {e}")

            st.markdown("---")
            st.subheader("🧊 Competência do setor (fechar / reabrir)")
            st.caption("Use este painel do ADMIN para congelar ou descongelar a competência de qualquer setor.")
            setores_comp = listar_setores_db()
            ac1, ac2, ac3 = st.columns([1.4, 1, 1])
            with ac1:
                setor_comp_admin = st.selectbox("Setor da competência", setores_comp, key="adm_comp_setor") if setores_comp else st.text_input("Setor da competência", value="FLV", key="adm_comp_setor_txt")
            with ac2:
                ano_comp_admin = st.number_input("Ano da competência", value=int(st.session_state.get("cfg_ano", datetime.now().year)), step=1, key="adm_comp_ano")
            with ac3:
                mes_comp_admin = st.selectbox("Mês da competência", list(range(1, 13)), index=max(0, int(st.session_state.get("cfg_mes", datetime.now().month)) - 1), key="adm_comp_mes")

            status_comp_admin = get_status_competencia(str(setor_comp_admin), int(ano_comp_admin), int(mes_comp_admin))
            s1, s2, s3 = st.columns([1.2, 1, 1])
            s1.metric("Status atual", status_comp_admin)
            if s2.button("🔒 Fechar competência", key="adm_comp_fechar", disabled=(status_comp_admin == "FECHADA")):
                try:
                    set_status_competencia(str(setor_comp_admin), int(ano_comp_admin), int(mes_comp_admin), "FECHADA")
                    st.success(f"Competência {int(mes_comp_admin):02d}/{int(ano_comp_admin)} do setor {str(setor_comp_admin).strip()} fechada com sucesso.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Falha ao fechar competência: {e}")
            if s3.button("🔓 Reabrir competência", key="adm_comp_reabrir", disabled=(status_comp_admin == "ABERTA")):
                try:
                    set_status_competencia(str(setor_comp_admin), int(ano_comp_admin), int(mes_comp_admin), "ABERTA")
                    st.success(f"Competência {int(mes_comp_admin):02d}/{int(ano_comp_admin)} do setor {str(setor_comp_admin).strip()} reaberta com sucesso.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Falha ao reabrir competência: {e}")

            st.markdown("---")
            st.subheader("🗄️ Backup / Restauração (escala.db)")

            c1, c2 = st.columns([1, 2])
            with c1:
                if st.button("Criar backup agora", key="adm_backup_now"):
                    try:
                        p = create_backup_now(prefix="manual")
                        st.success(f"Backup criado: {os.path.basename(p)}")
                    except Exception as e:
                        st.error(f"Falha ao criar backup: {e}")

            bks = list_backups()
            bk_sel = st.selectbox("Backups disponíveis", bks, key="adm_bk_sel") if bks else None
            if bk_sel:
                bk_path = os.path.join(BACKUP_DIR, bk_sel)
                with open(bk_path, "rb") as f:
                    st.download_button("⬇️ Baixar backup selecionado", data=f, file_name=bk_sel, mime="application/octet-stream", key="adm_bk_dl")

            st.markdown("### Restaurar um backup")
            up = st.file_uploader("Envie um arquivo .db (backup do escala.db)", type=["db"], key="adm_bk_up")
            if up is not None:
                if st.button("Restaurar este backup", key="adm_bk_restore"):
                    try:
                        restore_backup_from_bytes(up.getvalue())
                        st.success("Backup restaurado! Recarregando...")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Falha ao restaurar: {e}")

            st.caption(f"Backup automático (1x/dia) após {AUTO_BACKUP_HOUR:02d}:00. Pasta: {BACKUP_DIR}/")

            st.markdown("---")
            st.subheader("🧪 Teste Supabase")

            ok_conn, msg_conn = _supabase_test_connection_detail()
            if "sb_diag_last_error" not in st.session_state:
                st.session_state["sb_diag_last_error"] = ""
            if "sb_diag_last_msg" not in st.session_state:
                st.session_state["sb_diag_last_msg"] = msg_conn

            d1, d2, d3, d4 = st.columns(4)
            d1.metric("Sync habilitado", "Sim" if SUPABASE_SYNC_ENABLED else "Não")
            d2.metric("Conexão", "OK" if ok_conn else "Falha")
            d3.metric("Último push", _fmt_ts_br(_SUPABASE_LAST_PUSH_TS))
            d4.metric("Último pull", _fmt_ts_br(_SUPABASE_LAST_PULL_TS))

            info_df = pd.DataFrame([
                {"Campo": "URL", "Valor": SUPABASE_URL or "—"},
                {"Campo": "Schema", "Valor": SUPABASE_SCHEMA or "public"},
                {"Campo": "Key", "Valor": _mask_secret(SUPABASE_KEY)},
                {"Campo": "Auto pull ao abrir", "Valor": "Sim" if SUPABASE_AUTO_PULL_ON_START else "Não"},
                {"Campo": "Auto push no commit", "Valor": "Sim" if SUPABASE_AUTO_PUSH_ON_COMMIT else "Não"},
                {"Campo": "Auto push ao fechar", "Valor": "Sim" if SUPABASE_AUTO_PUSH_ON_CLOSE else "Não"},
                {"Campo": "Auto bootstrap pós-schema", "Valor": "Sim" if SUPABASE_AUTO_BOOTSTRAP_AFTER_SCHEMA else "Não"},
                {"Campo": "Auto restore se local vazio", "Valor": "Sim" if SUPABASE_AUTO_RESTORE_IF_LOCAL_EMPTY else "Não"},
                {"Campo": "Lock de sync (s)", "Valor": str(SUPABASE_SYNC_LOCK_TIMEOUT_SEC)},
                {"Campo": "Status atual", "Valor": msg_conn},
                {"Campo": "Último erro", "Valor": (_SUPABASE_LAST_ERROR or st.session_state.get("sb_diag_last_error", "")) or "—"},
            ])
            st.dataframe(info_df, use_container_width=True, hide_index=True)

            db_runtime = _db_runtime_summary()
            st.caption(f"Banco local: {db_runtime.get('db_path','')} | Existe: {db_runtime.get('db_exists')} | Melhor candidato: {db_runtime.get('best_candidate','')}")

            sb1, sb2, sb3 = st.columns(3)
            if sb1.button("Testar conexão", key="sb_test_conn_admin"):
                ok_now, msg_now = _supabase_test_connection_detail()
                st.session_state["sb_diag_last_msg"] = msg_now
                if ok_now:
                    st.success(msg_now)
                else:
                    st.session_state["sb_diag_last_error"] = msg_now
                    st.error(msg_now)

            if sb2.button("Forçar push", key="sb_force_push_admin"):
                try:
                    ok_push = _supabase_push_all_from_sqlite(force=True)
                    if ok_push:
                        st.session_state["sb_diag_last_error"] = ""
                        st.success("Push executado com sucesso.")
                    else:
                        st.warning(_SUPABASE_LAST_ERROR or "Push não executado.")
                except Exception as e:
                    st.session_state["sb_diag_last_error"] = str(e)
                    st.error(f"Falha no push: {e}")

            if sb3.button("Forçar pull", key="sb_force_pull_admin"):
                try:
                    ok_pull = _supabase_pull_all_to_sqlite(force=True)
                    if ok_pull:
                        st.session_state["sb_diag_last_error"] = ""
                        st.success("Pull executado com sucesso.")
                    else:
                        st.warning(_SUPABASE_LAST_ERROR or "Pull não trouxe dados.")
                except Exception as e:
                    st.session_state["sb_diag_last_error"] = str(e)
                    st.error(f"Falha no pull: {e}")

            with st.expander("Comparar tabelas local x Supabase", expanded=False):
                st.dataframe(_supabase_compare_tables_snapshot(), use_container_width=True, hide_index=True, height=360)

            st.markdown("---")
            st.subheader("🏷️ Setores (criar / listar)")
            setores = listar_setores_db()
            st.info("Setores cadastrados: " + ", ".join(setores))

            novo_setor = st.text_input("Novo setor (ex: FLV)", key="adm_new_setor")
            if st.button("➕ Criar setor", key="adm_create_setor"):
                try:
                    criar_setor_db(novo_setor)
                    st.success("Setor criado/garantido.")
                    st.rerun()
                except Exception as e:
                    st.error(str(e))

            st.markdown("---")

            st.subheader("📄 Importar escala a partir de PDF (automático — ESCALA_PONTO_NEW)")

            st.caption("Importa 100% automático: Nome + Chapa + Entrada (1ª linha) + FOLG/FER/AFA. Aplica no mês detectado do PDF como overrides (e pode cadastrar férias).")


            colA, colB, colC, colD = st.columns([1.3, 1, 1, 1])

            with colA:

                setor_dest = st.selectbox("Setor destino:", list_setores(), key="pdf_setor_dest")

            with colB:

                criar_colabs = st.checkbox("Criar/atualizar colaboradores", value=True, key="pdf_criar_colabs")

            with colC:

                limpar_mes = st.checkbox("Limpar overrides do mês antes", value=False, key="pdf_limpar_mes")

            with colD:

                cadastrar_ferias = st.checkbox("Cadastrar férias (FER)", value=True, key="pdf_cad_ferias")


            map_afa = st.checkbox("Tratar AFA como Folga", value=False, key="pdf_map_afa")
            auto_gerar_pdf = st.checkbox("Após importar, gerar mês automaticamente respeitando ajustes", value=True, key="pdf_auto_gerar")


            pdf = st.file_uploader("Enviar PDF da escala (ESCALA_PONTO_NEW)", type=["pdf"], key="adm_pdf_auto")

            if pdf is not None:

                try:

                    import PyPDF2

                    reader = PyPDF2.PdfReader(pdf)

                    parts = []

                    for page in reader.pages:

                        parts.append(page.extract_text() or "")

                    pdf_bytes = pdf.getvalue()
                    extracted = "\n".join(parts).strip()

                    if not extracted and not pdf_bytes:

                        st.warning("Não consegui extrair texto desse PDF (provável PDF imagem). Converta para PDF pesquisável ou envie CSV/Excel. OCR exige tesseract+poppler no servidor.")

                    else:

                        ano, mes, items, erros = _parse_escala_ponto_new_pdf_bytes(pdf_bytes) if pdf_bytes else (None, None, [], [])
                        if not items:
                            ano, mes, items, erros_txt = _parse_escala_ponto_new_pdf_text(extracted)
                            erros = (erros or []) + (erros_txt or [])


                        if erros:

                            st.warning("Encontrei divergências na leitura (ainda dá para aplicar, mas recomendo revisar):")

                            st.write(erros[:20])

                            if len(erros) > 20:

                                st.caption(f"... +{len(erros)-20} avisos")


                        if not items:

                            st.error("Não consegui identificar funcionários/entradas nesse PDF.")

                        else:

                            st.success(f"Modelo reconhecido ✅  Mês detectado: {mes:02d}/{ano} | Funcionários no PDF: {len(items)}")


                            with st.expander("Prévia (primeiros 3 funcionários)"):

                                for it in items[:3]:

                                    st.markdown(f"**{it.get('nome','')}**  — Chapa: `{it.get('chapa','')}`")

                                    st.write(it.get("tokens", [])[:10], " ...")


                            if st.button("✅ Aplicar escala do PDF no sistema (1 clique)", key="btn_apply_pdf"):

                                _apply_pdf_import_to_db(

                                    setor_destino=setor_dest,

                                    ano=int(ano),

                                    mes=int(mes),

                                    items=items,

                                    criar_colabs=bool(criar_colabs),

                                    limpar_mes_antes=bool(limpar_mes),

                                    map_afa_para_folga=bool(map_afa),

                                    cadastrar_ferias=bool(cadastrar_ferias),

                                )

                                if bool(auto_gerar_pdf):

                                    try:

                                        hist_pdf, estado_pdf = _build_hist_from_pdf_items(

                                            setor_dest, int(ano), int(mes), items,

                                            map_afa_para_folga=bool(map_afa)

                                        )

                                        if hist_pdf:

                                            save_escala_mes_db(setor_dest, int(ano), int(mes), hist_pdf)

                                            save_estado_mes(setor_dest, int(ano), int(mes), estado_pdf)

                                            st.session_state["ano"] = int(ano)

                                            st.session_state["mes"] = int(mes)

                                            st.success("PDF importado com sucesso! Para este mês, o PDF virou a fonte da verdade: folgas, férias e AFA foram salvos exatamente como estão no PDF. As regras do aplicativo voltam a valer normalmente na geração do mês seguinte.")

                                        else:

                                            st.warning("PDF importado, mas não consegui montar a escala final do mês a partir dos itens lidos.")

                                    except Exception as e_auto:

                                        st.warning(f"PDF importado, mas falhou ao salvar a escala final exatamente como veio no PDF: {e_auto}")

                                else:

                                    st.success("Importação aplicada com sucesso! Agora clique em 'Gerar agora (respeita ajustes)' para montar a escala do mês com folgas, AFA e férias do PDF.")

                except Exception as e:

                    st.error(f"Falha ao ler/importar PDF: {e}")





def _fast_restore_bundled_latest_before_start() -> None:
    """
    Restore mínimo e rápido antes de qualquer inicialização pesada:
    - se data/escala.db não existir ou vier zerado
    - e existir latest_stable.db ao lado do main.py
    - copia diretamente para data/escala.db
    """
    try:
        app_dir = Path(__file__).resolve().parent
        data_dir = app_dir / "data"
        db_path = data_dir / "escala.db"
        backup_candidates = [
            app_dir / "latest_stable.db",
            app_dir / "backups" / "latest_stable.db",
            app_dir / "data" / "latest_stable.db",
        ]
        data_dir.mkdir(parents=True, exist_ok=True)
        db_ok = db_path.exists() and db_path.stat().st_size > 0
        if db_ok:
            return
        for backup in backup_candidates:
            if backup.exists() and backup.stat().st_size > 0:
                shutil.copy2(backup, db_path)
                try:
                    latest_local = Path(BACKUP_DIR) / "latest_stable.db"
                    latest_local.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(backup, latest_local)
                except Exception:
                    pass
                break
    except Exception:
        pass


# =========================================================
# MAIN
# =========================================================
_fast_restore_bundled_latest_before_start()
validar_contrato_sistema()

if st.session_state["auth"] is None and QUICK_LOGIN_BOOT:
    db_init_fast_login()
    page_login()
else:
    if not st.session_state.get("_full_boot_done", False):
        db_init()
        if not FAST_BOOT_SKIP_STARTUP_AUTO_BACKUP:
            auto_backup_if_due()
        st.session_state["_full_boot_done"] = True
    if st.session_state["auth"] is None:
        page_login()
    else:
        page_app()
