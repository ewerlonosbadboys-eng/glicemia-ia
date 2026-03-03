# app.py
# =========================================================
# ESCALA 5x2 OFICIAL — COMPLETO (SUBGRUPO = REGRAS)
# + Preferência "Evitar folga" por subgrupo
# + Persistência real (SQLite) de ajustes (overrides)
# + Calendário RH visual + Banco de Horas
# + Admin (somente setor ADMIN e is_admin)
# + Gerar respeitando ajustes (overrides) OU ignorando
#
# ✅ CORREÇÕES ATIVAS:
# 1) DESCANSO GLOBAL 11:10 (INTERSTÍCIO) PARA A ESCALA INTEIRA
# 2) DOMINGO 1x1 (POR COLABORADOR) GLOBAL
# 3) PROIBIR FOLGAS CONSECUTIVAS AUTOMÁTICAS (ex.: DOM+SEG)
#    - Só fica folga consecutiva se estiver TRAVADO por override (manual / "caixinha")
# 4) enforce_global_rest_keep_targets NÃO PODE criar folga consecutiva “por acidente”
# 5) enforce_max_5_consecutive_work conta WORK_STATUSES como trabalho para sequência
#
# ✅ REGRAS GERAIS (ATUALIZAÇÃO):
# 6) FÉRIAS: só entra "Férias" se estiver cadastrada na ABA 🏖️ Férias (tabela ferias).
#    - Override "Férias" sem estar na tabela é ignorado.
#    - Se o banco tiver "Férias" sem estar na tabela, é corrigido para "Trabalho".
# 7) REGRA SEMANAL (SEG→DOM):
#    - Semana inicia SEG e termina DOM.
#    - Domingo 1x1 permanece.
#    - Se o colaborador FOLGA no domingo => 1 folga no período SEG–SÁB (SÁB só se permitir).
#    - Se o colaborador TRABALHA no domingo => 2 folgas no período SEG–SÁB (SÁB só se permitir).
#
# ✅ ALTERAÇÃO PEDIDA ANTES:
# - Removido tudo relacionado a "Balanço Madrugada" e ciclo "saída tarde"
#   (status, horários, funções e ações)
#
# ✅ ATUALIZAÇÃO DE HOJE (REGRA CRÍTICA):
# 8) PROIBIR TRABALHAR MAIS DE 5 DIAS DIRETO (GLOBAL, GARANTIA FINAL):
#    - Reaplica enforce_max_5_consecutive_work após funções que podem desfazer folgas:
#      enforce_weekly_folga_targets e rebalance_folgas_dia e no "pós final (garantia)".
# =========================================================

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, date
import io
import random
import calendar
import sqlite3
import hashlib
import secrets
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

# =========================================================
# PDF (Modelo Oficial) — ReportLab
# =========================================================
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
st.set_page_config(page_title="Escala 5x2 Oficial", layout="wide")



# =========================================================
# UI THEME (CSS) — só visual
# =========================================================
st.markdown("""
<style>
/* layout geral */
.block-container { padding-top: .6rem; padding-bottom: 2rem; max-width: 1600px; }
h1, h2, h3 { letter-spacing: -0.2px; }

/* KPI cards (topo) */
.kpi-card{
  border: 1px solid rgba(255,255,255,0.10);
  border-radius: 16px;
  padding: 12px 14px;
  background: rgba(255,255,255,0.06);
  box-shadow: 0 6px 18px rgba(0,0,0,0.18);
  backdrop-filter: blur(6px);
}
.kpi-card:hover{ transform: translateY(-1px); transition: 120ms ease; border-color: rgba(255,255,255,0.18); }
.kpi-title{ font-size: .78rem; opacity: .72; margin: 0 0 4px 0; text-transform: uppercase; letter-spacing: .4px; }
.kpi-value{ font-size: 1.35rem; font-weight: 800; margin: 0; line-height: 1.05; }

/* divisória */
.hr{ height:1px; background: rgba(255,255,255,0.08); margin: 14px 0; }

/* Tabs (menu superior) */
div[data-testid="stTabs"] { margin-top: .25rem; }
div[data-testid="stTabs"] button {
  font-size: .92rem;
  padding: 10px 14px;
  border-radius: 12px;
}
div[data-testid="stTabs"] button[aria-selected="true"]{
  background: rgba(255,255,255,0.07);
  border-bottom: 2px solid rgba(255,255,255,0.35);
}
div[data-testid="stTabs"] button:hover{
  background: rgba(255,255,255,0.06);
}

/* sidebar mais limpa */
section[data-testid="stSidebar"] .block-container { padding-top: 1rem; }

/* dataframe: arredondar */
div[data-testid="stDataFrame"] { border-radius: 12px; overflow: hidden; }
</style>
""", unsafe_allow_html=True)

DB_PATH = "escala.db"

# ---- Regras fixas
INTERSTICIO_MIN = timedelta(hours=11, minutes=10)   # 11:10
DURACAO_JORNADA = timedelta(hours=9, minutes=58)    # 9:58

PREF_EVITAR_PENALTY = 1000

BALANCO_STATUS = "Balanço"
WORK_STATUSES = {"Trabalho", BALANCO_STATUS}

BALANCO_DIA_ENTRADA = "06:00"
BALANCO_DIA_SAIDA = "11:50"

# Presets de horários (facilita seleção no app) — mais completo
HORARIOS_ENTRADA_PRESET = sorted({
    # Madrugada / manhã cedo
    "04:00","04:30","05:00","05:30",
    "06:00","06:10","06:20","06:30","06:40","06:45","06:50",
    "07:00","07:10","07:20","07:30","07:40","07:50",
    "08:00","08:10","08:20","08:30","08:40","08:50",
    # Manhã / meio-dia
    "09:00","09:10","09:20","09:30","09:40","09:50",
    "10:00","10:10","10:20","10:30","10:40","10:50",
    "11:00","11:10","11:20","11:30","11:40","11:50",
    "12:00","12:10","12:20","12:30","12:40","12:45","12:50",
    # Tarde
    "13:00","13:10","13:20","13:30","13:40","13:50",
    "14:00","14:10","14:20","14:30","14:40","14:50",
    "15:00","15:10","15:20","15:30","15:40","15:50",
    "16:00","16:10","16:20","16:30","16:40","16:50",
    "17:00","17:10","17:20","17:30","17:40","17:50",
    # Noite (se precisar)
    "18:00","18:10","18:20","18:30","18:40","18:50",
    "19:00","19:10","19:20","19:30","00:10",
})

D_PT = {
    "Monday": "seg",
    "Tuesday": "ter",
    "Wednesday": "qua",
    "Thursday": "qui",
    "Friday": "sex",
    "Saturday": "sáb",
    "Sunday": "dom",
}


# =========================================================
# ESCALA MANUAL (BASE) — Fevereiro/2026 (DSR)
# - Esta base serve para "iniciar" o mês com folgas pré-definidas.
# - Ao clicar em "Aplicar base", o app cria overrides (Status=Folga) nesses dias.
# - Depois, "Gerar agora (respeitando ajustes)" completa o restante mantendo as folgas travadas.
# =========================================================
MANUAL_BASES = {
    (2026, 2): [
        {"Chapa": "020.0823", "Nome": "ALEXANDRE ROBERTO ALMEIDA DOS REIS", "Dias_Folga": [1,4,6,9,15,18,20,23]},
        {"Chapa": "020.1447", "Nome": "ANA CAROLINA THEODORO PADILHA", "Dias_Folga": [1,3,6,11,15,18,20,26]},
        {"Chapa": "020.1733", "Nome": "BEATRIZ VITORIA DOS SANTOS LOPES", "Dias_Folga": [3,8,11,13,16,22,24,26]},
        {"Chapa": "020.1751", "Nome": "BRUNA SILVA MARTINS", "Dias_Folga": [1,4,6,9,15,17,19,23]},
        {"Chapa": "020.2288", "Nome": "CRISTIANE ALVES DOS SANTOS", "Dias_Folga": [2,8,11,13,16,22,24,26]},
        {"Chapa": "020.0265", "Nome": "DECIO EPAMINONDAS DE ALMEIDA NETO", "Dias_Folga": [4,8,10,12,18,22,24,26]},
        {"Chapa": "020.1839", "Nome": "DEYBSON JOSE DA SILVA", "Dias_Folga": [2,8,10,13,19,22,24,26]},
        {"Chapa": "020.1884", "Nome": "DISNEI OLIVEIRA ADORNO", "Dias_Folga": [1,4,6,10,15,17,20,23]},
        {"Chapa": "020.2192", "Nome": "EDILENE MARTINS DE MIRANDA", "Dias_Folga": [3,8,10,12,17,22,24,26]},
        {"Chapa": "020.2144", "Nome": "ELIS MIRIAN MARQUES OLIVEIRA", "Dias_Folga": [1,4,6,12,15,18,20,26]},
        {"Chapa": "020.1750", "Nome": "ELIZANGELA BARBOSA MOREIRA", "Dias_Folga": [22,25,27]},
        {"Chapa": "020.1984", "Nome": "EWERLON DE JESUS DA SILVA E SILVA", "Dias_Folga": [1,3,6,9,15,17,20,23]},
        {"Chapa": "020.2139", "Nome": "FABIANA SOUZA SILVA", "Dias_Folga": [3,8,11,13,18,22,24,26]},
        {"Chapa": "020.2450", "Nome": "GABRIEL CAMELO PINTO", "Dias_Folga": [3,8,10,12,18,22,25,27]},
        {"Chapa": "020.0748", "Nome": "IVANILDO FIGUEIREDO DA VERA CRUZ", "Dias_Folga": [16,22,25,27]},
        {"Chapa": "020.2299", "Nome": "JAIRON MACHADO DE ALMEIDA", "Dias_Folga": [2,8,11,13,16,22,24,26]},
        {"Chapa": "020.1649", "Nome": "JOAO VICTOR DE SOUZA SAMPAIO", "Dias_Folga": [1,3,5,9,15,17,20,25]},
        {"Chapa": "020.2274", "Nome": "JOSE FERNANDO OLIVEIRA DO NASCIMENTO", "Dias_Folga": [1,4,6,10,15,18,20,25]},
        {"Chapa": "020.2143", "Nome": "LUCAS EDUARDO DOS SANTOS SANTILLO", "Dias_Folga": [8,10,12,16,22,24,26]},
        {"Chapa": "020.1639", "Nome": "LUCIMARA EMILIA MARQUES", "Dias_Folga": [1,3,5,9,15,18,20,25]},
        {"Chapa": "020.2050", "Nome": "LUIZ FERNANDO DE TULIO", "Dias_Folga": [1,3,5,11,15,17,19,23]},
        {"Chapa": "020.1628", "Nome": "MACICLEIDE CONCEICAO DOS SANTOS", "Dias_Folga": [1,5,8,10,13,19,22,25,27]},
        {"Chapa": "020.0463", "Nome": "MARIA EDUARDA GONCALVES NUNES", "Dias_Folga": [2,8,10,12,16,22,24,26]},
        {"Chapa": "020.1854", "Nome": "MARIANA MABILLE DE MORAES", "Dias_Folga": []},
        {"Chapa": "020.1128", "Nome": "MARIVALDO RODRIGUES DA SILVA", "Dias_Folga": [1,4,6,12,15,18,20,23]},
        {"Chapa": "020.2309", "Nome": "MAURICIO DAVI DA SILVA NEIVAS ARAUJO", "Dias_Folga": [1,3,5,9,15,17,20,23]},
        {"Chapa": "020.2348", "Nome": "NATALIA CRISTINA GIMENES DE OLIVEIRA", "Dias_Folga": [1,4,6,12,15,17,19,23,27]},
        {"Chapa": "020.1856", "Nome": "RIQUELME CABRAL DE JESUS", "Dias_Folga": [3,8,11,13,18,22,24,26]},
        {"Chapa": "020.2388", "Nome": "RUTH PEREIRA DA SILVA", "Dias_Folga": [2,8,11,13,19,22,25,27]},
        {"Chapa": "020.1906", "Nome": "SHAIAN RUAN BARBOSA ALVES", "Dias_Folga": [4,8]},
        {"Chapa": "020.2203", "Nome": "TATIANE APARECIDA CABECA", "Dias_Folga": [1,4,6,9,15,18,20,25]},
        {"Chapa": "020.0994", "Nome": "VERA LUCIA BENEDITO ARRUDA", "Dias_Folga": [1,4,8,11,15,18,22,25]},
        {"Chapa": "020.1559", "Nome": "VIVIANE NASCIMENTO LIMA LEMOS", "Dias_Folga": [1,3,5,11,15,17,19,23]},
        {"Chapa": "020.1980", "Nome": "YASMIM STEFHANNY BATA SANTOS", "Dias_Folga": [5,8,10,12,17,22,24,26]},
    ]
}

# =========================================================
# Helpers de hora (minutos)
# =========================================================
def _to_min(hhmm: str) -> int:
    if not hhmm:
        return 0
    h, m = map(int, str(hhmm).split(":"))
    return h * 60 + m

def _min_to_hhmm(x: int) -> str:
    x %= (24 * 60)
    return f"{x//60:02d}:{x%60:02d}"

def _add_min(hhmm: str, delta: timedelta) -> str:
    return _min_to_hhmm(_to_min(hhmm) + int(delta.total_seconds() // 60))

def _sub_min(hhmm: str, delta: timedelta) -> str:
    return _min_to_hhmm(_to_min(hhmm) - int(delta.total_seconds() // 60))

def _saida_from_entrada(ent: str) -> str:
    return _add_min(ent, DURACAO_JORNADA)



# =========================================================
# PDF helpers (modelo de escala/ponto)
# - Linha "Horas Trab." do modelo costuma ser 08:48 (jornada 9:58 com 1:10 de intervalo)
# =========================================================
DURACAO_TRABALHADA = timedelta(hours=8, minutes=48)   # 08:48 (modelo)

def _hhmm_add(hhmm: str, minutes: int) -> str:
    if not hhmm:
        return ""
    h, m = map(int, hhmm.split(":"))
    total = (h * 60 + m + int(minutes)) % (24 * 60)
    return f"{total//60:02d}:{total%60:02d}"

def _montar_batidas_modelo(h_entrada: str):
    """
    Retorna (entrada1, saida_ref, entrada_ref, saida, horas_trab)

    Modelo igual ao do PDF:
      - Jornada (entrada->saída) = 9:58  (DURACAO_JORNADA)
      - Intervalo refeição = 1:10
      - Primeira parte (entrada -> saída refeição) = 5:10
      - Resultado "Horas Trab." = 08:48 quando é jornada padrão.
    """
    h_entrada = (h_entrada or "").strip()
    if not h_entrada:
        return "", "", "", "", ""

    # Parte 1 = 5h10, Refeição = 1h10
    parte1 = 5 * 60 + 10
    refeicao = 1 * 60 + 10

    saida_ref = _hhmm_add(h_entrada, parte1)
    ent_ref = _hhmm_add(saida_ref, refeicao)
    saida = _hhmm_add(h_entrada, int(DURACAO_JORNADA.total_seconds() // 60))  # 9:58

    # Horas trabalhadas no modelo = (9:58 - 1:10) = 8:48
    horas = "08:48"
    return h_entrada, saida_ref, ent_ref, saida, horas

def gerar_pdf_modelo_oficial(setor: str, ano: int, mes: int, hist_db: dict, colaboradores: list[dict]) -> bytes:
    """
    Gera PDF (A4 paisagem) com **4 colaboradores por página** (como o modelo enviado).
    - Folga: "FOLG" com destaque amarelo.
    - Férias: "FER" (sem destaque).
    - Manual supremo: o PDF reflete exatamente o que está salvo em hist_db.
    """
    from io import BytesIO
    from reportlab.pdfgen import canvas
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    import re

    # -----------------------------
    # Canvas com contagem total de páginas (X / Y)
    # -----------------------------
    class _NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            canvas.Canvas.__init__(self, *args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            num_pages = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self._draw_page_number(num_pages)
                canvas.Canvas.showPage(self)
            canvas.Canvas.save(self)

        def _draw_page_number(self, page_count):
            # no topo direito
            self.setFont("Helvetica", 7)
            self.drawRightString(landscape(A4)[0] - 12*mm, landscape(A4)[1] - 10*mm, f"Página: {self._pageNumber} / {page_count}")

    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    normal.fontName = "Helvetica"
    normal.fontSize = 7
    normal.leading = 8

    # Ordena colaboradores pelo nome (igual na tela)
    colab_by = {c["Chapa"]: c for c in colaboradores}
    chapas = sorted([ch for ch in hist_db.keys()], key=lambda ch: (colab_by.get(ch, {}).get("Nome", ch) or ch))

    # Config páginas
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10*mm,
        rightMargin=10*mm,
        topMargin=14*mm,
        bottomMargin=10*mm,
        title=f"Escala DSR {setor} {mes:02d}/{ano}"
    )

    W, H = landscape(A4)
    usable_w = W - doc.leftMargin - doc.rightMargin

    # ----- helpers
    def _pt_weekday(ts: pd.Timestamp) -> str:
        # usa D_PT já definido no app: Monday->seg...
        return {
            "seg": "Seg", "ter": "Ter", "qua": "Qua", "qui": "Qui", "sex": "Sex", "sáb": "Sáb", "dom": "Dom"
        }.get(D_PT[ts.day_name()], D_PT[ts.day_name()])

    def _format_mes():
        return f"Mês: {mes:02d}/{ano}"

    def _hhmm_norm(h: str) -> str:
        h = (h or "").strip()
        if not h:
            return ""
        h = h.replace(".", ":")
        if re.fullmatch(r"\d{1,2}:\d{2}", h):
            hh, mm_ = h.split(":")
            return f"{int(hh):02d}:{int(mm_):02d}"
        if re.fullmatch(r"\d{3,4}", h):
            h = h.zfill(4)
            return f"{h[:2]}:{h[2:]}"
        return h

    def _hhmm_diff_min(h1: str, h2: str) -> int:
        try:
            h1n = _hhmm_norm(h1); h2n = _hhmm_norm(h2)
            if not h1n or not h2n:
                return 0
            t1 = datetime.strptime(h1n, "%H:%M")
            t2 = datetime.strptime(h2n, "%H:%M")
            return int((t2 - t1).total_seconds() // 60)
        except Exception:
            return 0

    def _sum_total_horas(df: pd.DataFrame) -> str:
        # soma horas trabalhadas no modelo (primeira parte + segunda parte), respeitando horários reais quando existirem.
        total_min = 0
        for _, r in df.iterrows():
            stt = str(r.get("Status", ""))
            if stt not in WORK_STATUSES:
                continue
            ent = (r.get("H_Entrada") or "").strip()
            sai = (r.get("H_Saida") or "").strip()
            if not ent or not sai:
                continue
            # tenta modelo com refeição
            ent1, sref, entref, sai2, _ = _montar_batidas_modelo(ent)
            if sai2 == sai and sref and entref:
                # 8:48 padrão
                total_min += 8*60 + 48
            else:
                # fallback: duração bruta
                dur = _hhmm_diff_min(ent, sai)
                if dur > 0:
                    total_min += dur
        return f"{total_min//60}:{total_min%60:02d}"

    def _make_block(ch: str) -> list:
        df = hist_db[ch].copy()
        nome = colab_by.get(ch, {}).get("Nome", ch)
        sg = (colab_by.get(ch, {}).get("Subgrupo", "") or "").strip() or "COLABORADOR"
        sg_title = str(sg).upper()

        # tabela por dia
        qtd = len(df)
        dias_nums = [str(int(d.day)) for d in pd.to_datetime(df["Data"])]
        dias_sem = [_pt_weekday(pd.to_datetime(d)) for d in pd.to_datetime(df["Data"])]

        # constrói matriz 7 x (1+qtd)
        data = []
        data.append(["Data / Dia"] + dias_nums)
        data.append(["Dia / Semana"] + dias_sem)

        row_ent = ["Entrada"]
        row_sref = ["Saída Refeição"]
        row_entref = ["Entrada"]
        row_sai = ["Saída"]
        row_h = ["Horas Trab."]

        folg_cols = []
        for i in range(qtd):
            stt = str(df.loc[i, "Status"])
            ent = (df.loc[i, "H_Entrada"] or "").strip()
            sai = (df.loc[i, "H_Saida"] or "").strip()

            if stt == "Folga":
                row_ent.append("FOLG")
                row_sref.append("FOLG")
                row_entref.append("FOLG")
                row_sai.append("FOLG")
                row_h.append("")
                folg_cols.append(i+1)  # +1 por causa do label col
            elif stt == "Férias":
                row_ent.append("FER")
                row_sref.append("FER")
                row_entref.append("FER")
                row_sai.append("FER")
                row_h.append("")
            elif stt in WORK_STATUSES:
                if stt == BALANCO_STATUS:
                    row_ent.append(ent)
                    row_sref.append("")
                    row_entref.append("")
                    row_sai.append(sai)
                    # horas brutas
                    dm = _hhmm_diff_min(ent, sai) if ent and sai else 0
                    row_h.append(f"{dm//60:02d}:{dm%60:02d}" if dm else "")
                else:
                    ent1, sref, entref, saida2, horas = _montar_batidas_modelo(ent or colab_by.get(ch, {}).get("Entrada", "06:00"))
                    # respeita saída real do DF se diferente
                    if sai and saida2 and _hhmm_norm(sai) != _hhmm_norm(saida2):
                        # se alterado manualmente, mantém o do DF e deixa refeição em branco
                        row_ent.append(ent or "")
                        row_sref.append("")
                        row_entref.append("")
                        row_sai.append(sai)
                        dm = _hhmm_diff_min(ent, sai) if ent and sai else 0
                        row_h.append(f"{dm//60:02d}:{dm%60:02d}" if dm else "")
                    else:
                        row_ent.append(ent1)
                        row_sref.append(sref)
                        row_entref.append(entref)
                        row_sai.append(saida2)
                        row_h.append(horas)
            else:
                # status desconhecido
                row_ent.append("")
                row_sref.append("")
                row_entref.append("")
                row_sai.append("")
                row_h.append("")

        data += [row_ent, row_sref, row_entref, row_sai, row_h]

        label_w = 34*mm
        day_w = (usable_w - label_w) / max(1, qtd)

        tbl = Table(
            data,
            colWidths=[label_w] + [day_w]*qtd,
            rowHeights=[10, 10, 10, 10, 10, 10, 10]
        )

        ts = TableStyle([
            ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
            ("FONTSIZE", (0,0), (-1,-1), 7),
            ("ALIGN", (0,0), (0,-1), "LEFT"),
            ("ALIGN", (1,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("GRID", (0,0), (-1,-1), 0.5, colors.black),
            ("BACKGROUND", (0,0), (0,-1), colors.whitesmoke),
            ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ])

        # destaque FOLG (linhas 2..5)
        for c in folg_cols:
            for r in [2,3,4,5]:
                ts.add("BACKGROUND", (c, r), (c, r), colors.HexColor("#FFE699"))
                ts.add("FONTNAME", (c, r), (c, r), "Helvetica-Bold")

        tbl.setStyle(ts)

        # Barra cinza (cargo)
        bar = Table([[sg_title]], colWidths=[usable_w], rowHeights=[10])
        bar.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#D9D9D9")),
            ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOX", (0,0), (-1,-1), 0.8, colors.black),
        ]))

        # Linha Nome / Mês / Cliente
        header2 = Table(
            [[f"{nome} ({ch})", _format_mes(), "CLIENTE:"]],
            colWidths=[usable_w*0.55, usable_w*0.20, usable_w*0.25],
            rowHeights=[10]
        )
        header2.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 7),
            ("ALIGN", (0,0), (0,0), "LEFT"),
            ("ALIGN", (1,0), (1,0), "CENTER"),
            ("ALIGN", (2,0), (2,0), "LEFT"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOX", (0,0), (-1,-1), 0.8, colors.black),
        ]))

        # Rodapé do bloco
        total_horas = _sum_total_horas(df)
        footer = Table(
            [["É DE RESPONSABILIDADE DE CADA FUNCIONÁRIO CUMPRIR RIGOROSAMENTE ESTA ESCALA.", f"TOTAL DE HORAS : {total_horas}"]],
            colWidths=[usable_w*0.78, usable_w*0.22],
            rowHeights=[10]
        )
        footer.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 7),
            ("ALIGN", (0,0), (0,0), "LEFT"),
            ("ALIGN", (1,0), (1,0), "RIGHT"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOX", (0,0), (-1,-1), 0.8, colors.black),
        ]))

        return [bar, header2, tbl, footer, Spacer(1, 6)]

    # Cabeçalho de página (desenhado pelo onPage)
    emissao = datetime.now().strftime("%d/%m/%Y %H:%M")

    def _draw_header(canv, doc_):
        canv.saveState()
        canv.setStrokeColor(colors.black)
        canv.setFillColor(colors.black)

        # caixas do topo
        y = H - 12*mm
        canv.setFont("Helvetica-Bold", 9)
        canv.drawString(doc.leftMargin, y, f"Loja: {setor}")
        canv.drawCentredString(W/2, y, "Escala de DSR e Horário de Trabalho - Mês : {:02d}/{:04d}".format(mes, ano))
        canv.setFont("Helvetica", 7)
        canv.drawRightString(W - doc.rightMargin, y, f"Emissão: {emissao}")

        # título grande
        canv.setFont("Helvetica-Bold", 10)
        canv.drawString(doc.leftMargin, y - 10, "ESCALA_PONTO_NEW")

        # linha separadora
        canv.setLineWidth(1)
        canv.line(doc.leftMargin, y - 12, W - doc.rightMargin, y - 12)

        canv.restoreState()

    # Monta story: 4 blocos por página
    story = []
    per_page = 4
    for i, ch in enumerate(chapas):
        story += _make_block(ch)
        if (i+1) % per_page == 0 and (i+1) < len(chapas):
            story.append(PageBreak())

    doc.build(story, onFirstPage=_draw_header, onLaterPages=_draw_header, canvasmaker=_NumberedCanvas)
    return buffer.getvalue()

def _is_fixed_day(status: str) -> bool:
    # FIXO: balanço
    return str(status) == BALANCO_STATUS



def gerar_pdf_trabalhando_no_dia(setor: str, ano: int, mes: int, dia: int, hist_db: dict, colaboradores: list) -> bytes:
    """Gera um PDF simples (A4) listando apenas quem TRABALHA no dia escolhido, com horários."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    # Mapa rápido chapa->(nome, subgrupo)
    meta = {}
    for c in colaboradores:
        meta[str(c.get("Chapa", "")).strip()] = (str(c.get("Nome", "")).strip(), str(c.get("Subgrupo", "")).strip())

    rows = [["Chapa", "Nome", "Subgrupo", "Entrada", "Saída"]]
    for chapa, df in (hist_db or {}).items():
        if df is None or df.empty:
            continue
        try:
            linha = df.loc[df["Data"].dt.day == int(dia)].head(1)
        except Exception:
            # fallback: Data pode estar como string
            linha = df.loc[pd.to_datetime(df["Data"], errors="coerce").dt.day == int(dia)].head(1)
        if linha.empty:
            continue
        r = linha.iloc[0].to_dict()
        stt = str(r.get("Status", "")).strip()
        if stt not in WORK_STATUSES:
            continue
        ent = str(r.get("H_Entrada", "") or "").strip()
        sai = str(r.get("H_Saida", "") or "").strip()
        nome, subg = meta.get(str(chapa).strip(), ("", ""))
        rows.append([str(chapa).strip(), nome, subg, ent, sai])

    # Ordena por subgrupo e nome (mantendo cabeçalho)
    if len(rows) > 1:
        body = rows[1:]
        body.sort(key=lambda x: (x[2], x[1]))
        rows = [rows[0]] + body

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.2*cm,
        rightMargin=1.2*cm,
        topMargin=1.2*cm,
        bottomMargin=1.2*cm,
    )
    styles = getSampleStyleSheet()
    story = []

    titulo = f"Escala - Quem trabalha no dia {dia:02d}/{mes:02d}/{ano}"
    story.append(Paragraph(f"<b>{titulo}</b>", styles["Title"]))
    story.append(Paragraph(f"Setor: <b>{setor}</b>", styles["Normal"]))
    story.append(Spacer(1, 8))

    table = Table(rows, colWidths=[2.3*cm, 8.2*cm, 4.0*cm, 2.3*cm, 2.3*cm])
    table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 9),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),

        ("FONTNAME", (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,1), (-1,-1), 8),

        ("ALIGN", (0,0), (0,-1), "LEFT"),
        ("ALIGN", (3,1), (4,-1), "CENTER"),

        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.white]),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING", (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
    ]))
    story.append(table)

    doc.build(story)
    return buf.getvalue()



def is_work_status(status: str) -> bool:
    return str(status) in WORK_STATUSES

def _locked(locked_status: set[int] | None, idx: int) -> bool:
    return bool(locked_status and idx in locked_status)

def _ajustar_para_intersticio(ent_desejada: str, saida_anterior: str) -> str:
    """
    Entrada >= desejada respeitando 11:10 após saída anterior
    (considera dia seguinte quando necessário)
    """
    if not ent_desejada or not saida_anterior:
        return ent_desejada

    s = _to_min(saida_anterior)
    e_des = _to_min(ent_desejada)
    e_min = _to_min(_add_min(saida_anterior, INTERSTICIO_MIN))

    if e_des <= s:
        e_des += 1440
    if e_min <= s:
        e_min += 1440

    e_ok = max(e_des, e_min)
    return _min_to_hhmm(e_ok)

# =========================================================
# ✅ Proibir folga consecutiva AUTOMÁTICA (DOM+SEG etc.)
# Só permite se estiver travado (override/manual/"caixinha")
# =========================================================
def enforce_no_consecutive_folga(df: pd.DataFrame, locked_status: set[int] | None = None):
    """
    Proibir folga consecutiva automática (DOM+SEG etc.).
    Robustez: garante índice 0..N-1 e usa iloc para não dar KeyError.
    """
    df.reset_index(drop=True, inplace=True)
    for i in range(1, len(df)):
        if df.iloc[i - 1]["Status"] == "Folga" and df.iloc[i]["Status"] == "Folga":
            prev_locked = _locked(locked_status, i - 1)
            cur_locked = _locked(locked_status, i)

            # ambos travados => foi decisão manual, mantém
            if prev_locked and cur_locked:
                continue

            # prioriza manter o travado e desfazer o outro
            if not cur_locked:
                df.iloc[i, df.columns.get_loc("Status")] = "Trabalho"
            elif not prev_locked:
                df.iloc[i - 1, df.columns.get_loc("Status")] = "Trabalho"
# =========================================================
# DB
# =========================================================
def db_conn():
    return sqlite3.connect(DB_PATH, check_same_thread=False)

def hash_password(password: str, salt: str) -> str:
    return hashlib.sha256((salt + password).encode("utf-8")).hexdigest()

def _safe_exec(cur, sql: str, params=None):
    try:
        if params is None:
            cur.execute(sql)
        else:
            cur.execute(sql, params)
    except Exception:
        pass

def db_init():
    con = db_conn()
    cur = con.cursor()

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS setores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT UNIQUE NOT NULL
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS usuarios_sistema (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        senha_hash TEXT NOT NULL,
        salt TEXT NOT NULL,
        is_admin INTEGER NOT NULL DEFAULT 0,
        is_lider INTEGER NOT NULL DEFAULT 0,
        criado_em TEXT NOT NULL,
        UNIQUE(setor, chapa)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS colaboradores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        subgrupo TEXT DEFAULT '',
        entrada TEXT DEFAULT '06:00',
        folga_sab INTEGER DEFAULT 0,
        criado_em TEXT NOT NULL,
        UNIQUE(setor, chapa)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS subgrupos_setor (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        nome TEXT NOT NULL,
        UNIQUE(setor, nome)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS subgrupo_regras (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        subgrupo TEXT NOT NULL,
        evitar_seg INTEGER NOT NULL DEFAULT 0,
        evitar_ter INTEGER NOT NULL DEFAULT 0,
        evitar_qua INTEGER NOT NULL DEFAULT 0,
        evitar_qui INTEGER NOT NULL DEFAULT 0,
        evitar_sex INTEGER NOT NULL DEFAULT 0,
        evitar_sab INTEGER NOT NULL DEFAULT 0,
        UNIQUE(setor, subgrupo)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS ferias (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        inicio TEXT NOT NULL,
        fim TEXT NOT NULL
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS estado_mes_anterior (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        chapa TEXT NOT NULL,
        consec_trab_final INTEGER NOT NULL,
        ultima_saida TEXT NOT NULL,
        ultimo_domingo_status TEXT,
        ano INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        UNIQUE(setor, chapa, ano, mes)
    )
    """)

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS escala_mes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        ano INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        chapa TEXT NOT NULL,
        dia INTEGER NOT NULL,
        data TEXT NOT NULL,
        dia_sem TEXT NOT NULL,
        status TEXT NOT NULL,
        h_entrada TEXT,
        h_saida TEXT,
        UNIQUE(setor, ano, mes, chapa, dia)
    )
    """)

    # --- MIGRAÇÃO defensiva (Streamlit Cloud pode manter DB antigo)
    # Garante que a tabela escala_mes tenha todas as colunas esperadas
    try:
        cur.execute("PRAGMA table_info(escala_mes)")
        cols = {r[1] for r in cur.fetchall()}  # r[1] = name
        # colunas esperadas (além das já existentes)
        expected = {"setor","ano","mes","chapa","dia","data","dia_sem","status","h_entrada","h_saida"}
        missing = expected - cols
        for c in sorted(missing):
            # tipos simples (compatível com SQLite)
            if c in ("ano","mes","dia"):
                cur.execute(f"ALTER TABLE escala_mes ADD COLUMN {c} INTEGER")
            else:
                cur.execute(f"ALTER TABLE escala_mes ADD COLUMN {c} TEXT")
        con.commit()
    except Exception:
        pass

    _safe_exec(cur, """
    CREATE TABLE IF NOT EXISTS overrides (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        setor TEXT NOT NULL,
        ano INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        chapa TEXT NOT NULL,
        dia INTEGER NOT NULL,
        campo TEXT NOT NULL,
        valor TEXT NOT NULL,
        UNIQUE(setor, ano, mes, chapa, dia, campo)
    )
    """)

    con.commit()
    _safe_exec(cur, "INSERT OR IGNORE INTO setores(nome) VALUES (?)", ("GERAL",))
    _safe_exec(cur, "INSERT OR IGNORE INTO setores(nome) VALUES (?)", ("ADMIN",))
    con.commit()

    cur.execute("SELECT 1 FROM usuarios_sistema WHERE setor=? AND chapa=? LIMIT 1", ("ADMIN", "admin"))
    if cur.fetchone() is None:
        salt = secrets.token_hex(16)
        senha_hash = hash_password("123", salt)
        cur.execute("""
            INSERT INTO usuarios_sistema(nome, setor, chapa, senha_hash, salt, is_admin, is_lider, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, ("Administrador", "ADMIN", "admin", senha_hash, salt, 1, 1, datetime.now().isoformat()))
        con.commit()

    con.close()


def is_past_competencia(ano: int, mes: int) -> bool:
    """Meses anteriores ao mês atual (no fuso do servidor)."""
    today = date.today()
    return (int(ano), int(mes)) < (int(today.year), int(today.month))


# =========================================================
# AUTH
# =========================================================
def system_user_exists(setor: str, chapa: str) -> bool:
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT 1 FROM usuarios_sistema WHERE setor=? AND chapa=? LIMIT 1", (setor, chapa))
    ok = cur.fetchone() is not None
    con.close()
    return ok

def create_system_user(nome: str, setor: str, chapa: str, senha: str, is_lider: int = 0, is_admin: int = 0):
    salt = secrets.token_hex(16)
    senha_hash = hash_password(senha, salt)
    con = db_conn()
    cur = con.cursor()
    cur.execute("INSERT OR IGNORE INTO setores(nome) VALUES (?)", (setor,))
    cur.execute("""
        INSERT INTO usuarios_sistema(nome, setor, chapa, senha_hash, salt, is_admin, is_lider, criado_em)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (nome, setor, chapa, senha_hash, salt, int(is_admin), int(is_lider), datetime.now().isoformat()))
    con.commit()
    con.close()

def verify_login(setor: str, chapa: str, senha: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT nome, senha_hash, salt, is_admin, is_lider
        FROM usuarios_sistema
        WHERE setor=? AND chapa=?
        LIMIT 1
    """, (setor, chapa))
    row = cur.fetchone()
    con.close()
    if not row:
        return None
    nome, senha_hash, salt, is_admin, is_lider = row
    if hash_password(senha, salt) == senha_hash:
        return {"nome": nome, "setor": setor, "chapa": chapa, "is_admin": bool(is_admin), "is_lider": bool(is_lider)}
    return None

def is_lider_chapa(setor: str, chapa_lider: str) -> bool:
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT is_lider FROM usuarios_sistema WHERE setor=? AND chapa=? LIMIT 1", (setor, chapa_lider))
    row = cur.fetchone()
    con.close()
    return bool(row and row[0] == 1)

def update_password(setor: str, chapa: str, nova_senha: str):
    salt = secrets.token_hex(16)
    senha_hash = hash_password(nova_senha, salt)
    con = db_conn()
    cur = con.cursor()
    cur.execute("UPDATE usuarios_sistema SET senha_hash=?, salt=? WHERE setor=? AND chapa=?",
                (senha_hash, salt, setor, chapa))
    con.commit()
    con.close()

# =========================================================
# ADMIN
# =========================================================
@st.cache_data(show_spinner=False)
def admin_list_users():
    con = db_conn()
    df = pd.read_sql_query("""
        SELECT id, nome, setor, chapa, is_admin, is_lider, criado_em
        FROM usuarios_sistema
        ORDER BY setor ASC, nome ASC
    """, con)
    con.close()
    return df

def admin_reset_user_password(user_id: int, nova_senha: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT setor, chapa FROM usuarios_sistema WHERE id=?", (int(user_id),))
    row = cur.fetchone()
    if not row:
        con.close()
        return False
    setor, chapa = row
    con.close()
    update_password(setor, chapa, nova_senha)
    return True

# =========================================================
# COLABORADORES
# =========================================================
def colaborador_exists(setor: str, chapa: str) -> bool:
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT 1 FROM colaboradores WHERE setor=? AND chapa=? LIMIT 1", (setor, chapa))
    ok = cur.fetchone() is not None
    con.close()
    return ok

def create_colaborador(nome: str, setor: str, chapa: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("INSERT OR IGNORE INTO colaboradores(nome, setor, chapa, criado_em) VALUES (?, ?, ?, ?)",
                (nome, setor, chapa, datetime.now().isoformat()))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass


def upsert_colaborador_nome(setor: str, chapa: str, nome: str):
    """
    Garante que existe o colaborador (SEM senha) e atualiza o nome pelo que veio na base manual.
    - Se a chapa existir: atualiza nome.
    - Se não existir: cria.
    """
    nome = (nome or "").strip()
    chapa = (chapa or "").strip()
    if not chapa:
        return
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT 1 FROM colaboradores WHERE setor=? AND chapa=? LIMIT 1", (setor, chapa))
    if cur.fetchone() is None:
        cur.execute("INSERT INTO colaboradores(nome, setor, chapa, criado_em) VALUES (?, ?, ?, ?)",
                    (nome or chapa, setor, chapa, datetime.now().isoformat()))
    else:
        if nome:
            cur.execute("UPDATE colaboradores SET nome=? WHERE setor=? AND chapa=?", (nome, setor, chapa))
    con.commit()
    con.close()

def apply_manual_base_folgas(setor: str, ano: int, mes: int, base_rows: list[dict], limpar_overrides_mes: bool = False):
    """
    Aplica uma base manual criando overrides Status=Folga.
    - Opcional: limpar_overrides_mes = True remove TODOS os overrides do mês antes de aplicar.
    """
    con = db_conn()
    cur = con.cursor()
    if limpar_overrides_mes:
        cur.execute("DELETE FROM overrides WHERE setor=? AND ano=? AND mes=?", (setor, int(ano), int(mes)))
        con.commit()
    con.close()

    # garante colaboradores e aplica folgas como override
    for r in base_rows:
        ch = str(r.get("Chapa","")).strip()
        nm = str(r.get("Nome","")).strip()
        dias = r.get("Dias_Folga", []) or []
        upsert_colaborador_nome(setor, ch, nm)
        for d in dias:
            try:
                dd = int(d)
            except Exception:
                continue
            if dd <= 0:
                continue
            set_override(setor, int(ano), int(mes), ch, dd, "status", "Folga")

def delete_colaborador_total(setor: str, chapa: str):
    """
    Exclui colaborador e tudo do setor relacionado a ele:
    - colaboradores
    - ferias
    - overrides
    - escala_mes
    - estado_mes_anterior
    """
    con = db_conn()
    cur = con.cursor()
    cur.execute("DELETE FROM ferias WHERE setor=? AND chapa=?", (setor, chapa))
    cur.execute("DELETE FROM overrides WHERE setor=? AND chapa=?", (setor, chapa))
    cur.execute("DELETE FROM escala_mes WHERE setor=? AND chapa=?", (setor, chapa))
    cur.execute("DELETE FROM estado_mes_anterior WHERE setor=? AND chapa=?", (setor, chapa))
    cur.execute("DELETE FROM colaboradores WHERE setor=? AND chapa=?", (setor, chapa))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

def update_colaborador_perfil(setor: str, chapa: str, subgrupo: str, entrada: str, folga_sab: bool):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        UPDATE colaboradores
        SET subgrupo=?, entrada=?, folga_sab=?
        WHERE setor=? AND chapa=?
    """, (subgrupo or "", entrada, 1 if folga_sab else 0, setor, chapa))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

@st.cache_data(show_spinner=False)
def load_colaboradores_setor(setor: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT nome, chapa, subgrupo, entrada, folga_sab
        FROM colaboradores
        WHERE setor=?
        ORDER BY nome ASC
    """, (setor,))
    rows = cur.fetchall()
    con.close()
    return [{
        "Nome": r[0],
        "Chapa": r[1],
        "Subgrupo": (r[2] or "").strip(),
        "Entrada": (r[3] or "06:00").strip(),
        "Folga_Sab": bool(r[4]),
        "Setor": setor,
    } for r in rows]

# =========================================================
# SUBGRUPOS + REGRAS
# =========================================================
@st.cache_data(show_spinner=False)
def list_subgrupos(setor: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT nome FROM subgrupos_setor WHERE setor=? ORDER BY nome ASC", (setor,))
    rows = [r[0] for r in cur.fetchall()]
    con.close()
    return rows

def add_subgrupo(setor: str, nome: str):
    nome = (nome or "").strip()
    if not nome:
        return
    con = db_conn()
    cur = con.cursor()
    cur.execute("INSERT OR IGNORE INTO subgrupos_setor(setor, nome) VALUES (?, ?)", (setor, nome))
    cur.execute("""
        INSERT OR IGNORE INTO subgrupo_regras(setor, subgrupo, evitar_seg, evitar_ter, evitar_qua, evitar_qui, evitar_sex, evitar_sab)
        VALUES (?, ?, 0,0,0,0,0,0)
    """, (setor, nome))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

def delete_subgrupo(setor: str, nome: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("DELETE FROM subgrupos_setor WHERE setor=? AND nome=?", (setor, nome))
    cur.execute("DELETE FROM subgrupo_regras WHERE setor=? AND subgrupo=?", (setor, nome))
    cur.execute("UPDATE colaboradores SET subgrupo='' WHERE setor=? AND subgrupo=?", (setor, nome))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

@st.cache_data(show_spinner=False)
def get_subgrupo_regras(setor: str, subgrupo: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT evitar_seg, evitar_ter, evitar_qua, evitar_qui, evitar_sex, evitar_sab
        FROM subgrupo_regras
        WHERE setor=? AND subgrupo=?
        LIMIT 1
    """, (setor, subgrupo))
    row = cur.fetchone()
    con.close()
    if not row:
        return {"seg": 0, "ter": 0, "qua": 0, "qui": 0, "sex": 0, "sáb": 0}
    return {"seg": row[0], "ter": row[1], "qua": row[2], "qui": row[3], "sex": row[4], "sáb": row[5]}

def set_subgrupo_regras(setor: str, subgrupo: str, regras: dict):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        INSERT OR REPLACE INTO subgrupo_regras(setor, subgrupo, evitar_seg, evitar_ter, evitar_qua, evitar_qui, evitar_sex, evitar_sab)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        setor, subgrupo,
        int(regras.get("seg", 0)),
        int(regras.get("ter", 0)),
        int(regras.get("qua", 0)),
        int(regras.get("qui", 0)),
        int(regras.get("sex", 0)),
        int(regras.get("sáb", 0)),
    ))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

# =========================================================
# FÉRIAS
# =========================================================
def add_ferias(setor: str, chapa: str, inicio: date, fim: date):
    con = db_conn()
    cur = con.cursor()
    cur.execute("INSERT INTO ferias(setor, chapa, inicio, fim) VALUES (?, ?, ?, ?)",
                (setor, chapa, inicio.strftime("%Y-%m-%d"), fim.strftime("%Y-%m-%d")))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

def delete_ferias_row(setor: str, chapa: str, inicio: str, fim: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        DELETE FROM ferias
        WHERE setor=? AND chapa=? AND inicio=? AND fim=?
    """, (setor, chapa, inicio, fim))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

@st.cache_data(show_spinner=False)
def list_ferias(setor: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("SELECT chapa, inicio, fim FROM ferias WHERE setor=? ORDER BY date(inicio) ASC", (setor,))
    rows = cur.fetchall()
    con.close()
    return rows

def is_de_ferias(setor: str, chapa: str, data_obj: date) -> bool:
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT 1 FROM ferias
        WHERE setor=? AND chapa=?
          AND date(inicio) <= date(?) AND date(fim) >= date(?)
        LIMIT 1
    """, (setor, chapa, data_obj.strftime("%Y-%m-%d"), data_obj.strftime("%Y-%m-%d")))
    ok = cur.fetchone() is not None
    con.close()
    return ok

def is_first_week_after_return(setor: str, chapa: str, data_obj: date) -> bool:
    ontem = data_obj - timedelta(days=1)
    if is_de_ferias(setor, chapa, data_obj):
        return False
    if is_de_ferias(setor, chapa, ontem):
        return True
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT fim FROM ferias
        WHERE setor=? AND chapa=? AND date(fim) < date(?)
        ORDER BY date(fim) DESC
        LIMIT 1
    """, (setor, chapa, data_obj.strftime("%Y-%m-%d")))
    row = cur.fetchone()
    con.close()
    if not row:
        return False
    fim = datetime.strptime(row[0], "%Y-%m-%d").date()
    retorno = fim + timedelta(days=1)
    return retorno <= data_obj <= (retorno + timedelta(days=6))

def get_last_ferias_fim(setor: str, chapa: str, ate_data: date) -> date | None:
    """Retorna a data de FIM das últimas férias do colaborador antes (ou no) ate_data.
    Usa apenas para regras de retorno/encaixe. Retorna None se não houver registro.
    """
    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute(
            """
            SELECT inicio, fim
            FROM ferias
            WHERE setor = ? AND chapa = ? AND date(fim) <= date(?)
            ORDER BY date(fim) DESC
            LIMIT 1
            """,
            (setor, str(chapa), ate_data.isoformat()),
        )
        row = cur.fetchone()
    except Exception:
        row = None
    finally:
        try:
            conn.close()
        except Exception:
            pass

    if not row:
        return None
    try:
        fim = datetime.strptime(row[1], "%Y-%m-%d").date()
        return fim
    except Exception:
        return None


def balance_primeiro_domingo_pos_ferias(df, hist_all, df_ref, setor, ch, chapas_grupo, locked_idx, ate_data):
    """
    Ajuste específico: após retorno de FÉRIAS, decidir o 1º DOMINGO (trabalha/folga)
    com base no balanceamento de cobertura do DOMINGO dentro do mesmo grupo (subgrupo/setor).

    - Identifica o retorno: primeiro dia NÃO-FÉRIAS após um bloco de FÉRIAS.
    - Pega o primeiro domingo >= retorno.
    - Se o domingo estiver "cheio" (muita gente trabalhando), força FOLGA no colaborador.
      Se estiver "precisando" (pouca gente), mantém TRABALHO.
    - Não sobrescreve dia travado por override (locked_idx).
    """
    if df is None or df_ref is None or len(df_ref) == 0:
        return

    # --- localizar data de retorno (primeiro dia não-FÉRIAS após FÉRIAS)
    retorno_idx = None
    for i in range(1, len(df_ref)):
        prev_status = str(df.loc[i-1, "Status"]) if i-1 in df.index else ""
        cur_status  = str(df.loc[i, "Status"])   if i in df.index else ""
        if prev_status == "Férias" and cur_status != "Férias":
            retorno_idx = i
            break

    if retorno_idx is None:
        return

    retorno_data = pd.to_datetime(df_ref.loc[retorno_idx, "Data"]).date()

    # --- achar primeiro domingo em/apos retorno
    dom_i = None
    for i in range(retorno_idx, len(df_ref)):
        d = pd.to_datetime(df_ref.loc[i, "Data"]).date()
        if d > ate_data:
            break
        dia = str(df_ref.loc[i, "Dia"]).lower()
        if dia.startswith("dom"):
            dom_i = i
            break

    if dom_i is None:
        return

    # não mexer se estiver travado por override
    if dom_i in (locked_idx.get(ch, set()) or set()):
        return

    # se no próprio domingo ainda é férias, não mexe
    if str(df.loc[dom_i, "Status"]) == "Férias":
        return

    # --- contar cobertura do domingo no grupo
    work_count = 0
    avail_count = 0
    for ch2 in chapas_grupo:
        df2 = df if ch2 == ch else hist_all.get(ch2)
        if df2 is None:
            continue
        st2 = str(df2.loc[dom_i, "Status"])
        if st2 == "Férias":
            continue
        avail_count += 1
        if st2 in WORK_STATUSES:
            work_count += 1

    if avail_count <= 1:
        return

    # alvo simples: ~50% trabalhando
    target_work = int(math.ceil(avail_count / 2))

    # --- decidir status do colaborador no 1º domingo
    desejado = "Folga" if work_count > target_work else None  # None = manter como está (trabalho)
    if desejado == "Folga":
        # evitar criar folga consecutiva automática (DOM+SEG), se SEG não está travado
        if dom_i + 1 < len(df_ref):
            prox_status = str(df.loc[dom_i + 1, "Status"])
            prox_locked = (dom_i + 1) in (locked_idx.get(ch, set()) or set())
            if (prox_status == "Folga") and (not prox_locked):
                desejado = None  # não força folga; deixa trabalhar
        # também evita folga consecutiva com sábado anterior
        if dom_i - 1 >= 0:
            ant_status = str(df.loc[dom_i - 1, "Status"])
            ant_locked = (dom_i - 1) in (locked_idx.get(ch, set()) or set())
            if (ant_status == "Folga") and (not ant_locked):
                desejado = None

    if desejado == "Folga":
        _set_folga(df, dom_i, locked=False, reason="auto_balance_1o_dom_pos_ferias")


def load_estado_prev(setor: str, ano: int, mes: int):
    prev_ano, prev_mes = ano, mes - 1
    if prev_mes == 0:
        prev_mes = 12
        prev_ano -= 1
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT chapa, consec_trab_final, ultima_saida, ultimo_domingo_status
        FROM estado_mes_anterior
        WHERE setor=? AND ano=? AND mes=?
    """, (setor, prev_ano, prev_mes))
    rows = cur.fetchall()
    con.close()
    estado = {}
    for chapa, consec, ultima_saida, ultimo_dom in rows:
        estado[chapa] = {"consec_trab_final": int(consec), "ultima_saida": ultima_saida or "", "ultimo_domingo_status": ultimo_dom}
    return estado

# =========================================================
# OVERRIDES
# =========================================================
def set_override(setor: str, ano: int, mes: int, chapa: str, dia: int, campo: str, valor: str):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        INSERT OR REPLACE INTO overrides(setor, ano, mes, chapa, dia, campo, valor)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (setor, int(ano), int(mes), chapa, int(dia), campo, str(valor)))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

def delete_override(setor: str, ano: int, mes: int, chapa: str, dia: int, campo: str | None = None):
    con = db_conn()
    cur = con.cursor()
    if campo:
        cur.execute("""
            DELETE FROM overrides
            WHERE setor=? AND ano=? AND mes=? AND chapa=? AND dia=? AND campo=?
        """, (setor, int(ano), int(mes), chapa, int(dia), campo))
    else:
        cur.execute("""
            DELETE FROM overrides
            WHERE setor=? AND ano=? AND mes=? AND chapa=? AND dia=?
        """, (setor, int(ano), int(mes), chapa, int(dia)))
    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

@st.cache_data(show_spinner=False)
def load_overrides(setor: str, ano: int, mes: int):
    con = db_conn()
    df = pd.read_sql_query("""
        SELECT chapa, dia, campo, valor
        FROM overrides
        WHERE setor=? AND ano=? AND mes=?
    """, con, params=(setor, int(ano), int(mes)))
    con.close()
    return df

def _ov_map(setor: str, ano: int, mes: int):
    df = load_overrides(setor, ano, mes)
    ov = {}
    if df is None or df.empty:
        return ov
    for _, r in df.iterrows():
        ch = str(r["chapa"])
        dia = int(r["dia"])
        campo = str(r["campo"])
        valor = str(r["valor"])
        ov.setdefault(ch, {}).setdefault(dia, {})[campo] = valor
    return ov

def _is_status_locked(ovmap: dict, chapa: str, data_ts: pd.Timestamp) -> bool:
    dia = int(pd.to_datetime(data_ts).day)
    return bool(ovmap.get(chapa, {}).get(dia, {}).get("status"))

def _apply_overrides_to_df_inplace(df: pd.DataFrame, setor: str, chapa: str, ovmap: dict):
    """Aplica ajustes manuais (overrides) no DataFrame.

    Regras de FÉRIAS:
    - Dias de férias são definidos SOMENTE pela tabela `ferias` (aba Férias).
    - Em dia de férias, força: Status='Férias', H_Entrada='', H_Saida='' (ignora qualquer override).
    - Override tentando marcar 'Férias' fora da tabela é ignorado.
    - Override tentando mudar o Status / horários em um dia que está em férias também é ignorado.
    """
    ovmap = (ovmap or {})
    if not ovmap:
        return df

    # Garante tipo datetime na coluna Data
    if "Data" in df.columns and not pd.api.types.is_datetime64_any_dtype(df["Data"]):
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce")

    for d_raw, payload in ovmap.items():
        if not payload:
            continue

        # Normaliza data (aceita date/datetime/str ISO)
        dd = None
        if isinstance(d_raw, dt.datetime):
            dd = d_raw.date()
        elif isinstance(d_raw, dt.date):
            dd = d_raw
        elif isinstance(d_raw, str):
            try:
                dd = dt.date.fromisoformat(d_raw[:10])
            except Exception:
                dd = None
        if dd is None:
            continue

        mask = df["Data"].dt.date == dd
        if not bool(mask.any()):
            continue
        i = df.index[mask][0]

        # Se é férias (tabela), férias sempre vence
        if is_de_ferias(setor, chapa, dd):
            df.loc[i, "Status"] = "Férias"
            df.loc[i, "H_Entrada"] = ""
            df.loc[i, "H_Saida"] = ""
            continue

        # Status (exceto tentar marcar Férias)
        st_new = str(payload.get("status") or "").strip()
        if st_new and st_new.lower() not in ["férias", "ferias"]:
            df.loc[i, "Status"] = st_new

        # Entrada / Saída
        ent_new = str(payload.get("h_entrada") or payload.get("entrada") or "").strip()
        if ent_new:
            df.loc[i, "H_Entrada"] = ent_new
            df.loc[i, "H_Saida"] = _saida_from_entrada(ent_new)

    return df


def save_escala_mes_db(setor: str, ano: int, mes: int, historico_df_por_chapa: dict[str, pd.DataFrame]):
    """Grava escala no banco de forma robusta.
    - Limpa a competência (setor/ano/mes) antes de gravar para evitar IntegrityError em DB antigo/corrompido.
    - Robustez contra NaT/NaN.
    """
    con = db_conn()
    cur = con.cursor()

    # Limpa o mês inteiro do setor antes de inserir (evita conflito/duplicidade em DB antigo)
    try:
        cur.execute("DELETE FROM escala_mes WHERE setor=? AND ano=? AND mes=?", (setor, int(ano), int(mes)))
        con.commit()
    except Exception:
        pass

    for chapa, df in historico_df_por_chapa.items():
        df2 = df.copy()
        df2.reset_index(drop=True, inplace=True)

        for j, row in df2.iterrows():
            dt = pd.to_datetime(row.get("Data", None), errors="coerce")
            max_day = calendar.monthrange(int(ano), int(mes))[1]

            if pd.isna(dt):
                dia = int(j) + 1
                if dia < 1: dia = 1
                if dia > max_day: dia = max_day
                dt = pd.Timestamp(year=int(ano), month=int(mes), day=int(dia))
            else:
                dia = int(getattr(dt, "day", 1) or 1)
                if dia < 1: dia = 1
                if dia > max_day:
                    dia = max_day
                    dt = pd.Timestamp(year=int(ano), month=int(mes), day=int(dia))

            dia_sem = row.get("Dia", "")
            if pd.isna(dia_sem): dia_sem = ""
            dia_sem = str(dia_sem)

            status = row.get("Status", "Trabalho")
            if pd.isna(status) or not str(status).strip():
                status = "Trabalho"
            status = str(status)

            h_ent = row.get("H_Entrada", "")
            h_sai = row.get("H_Saida", "")
            if pd.isna(h_ent): h_ent = ""
            if pd.isna(h_sai): h_sai = ""
            h_ent = str(h_ent or "")
            h_sai = str(h_sai or "")

            try:
                cur.execute("""
                    INSERT OR REPLACE INTO escala_mes(setor, ano, mes, chapa, dia, data, dia_sem, status, h_entrada, h_saida)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    setor, int(ano), int(mes), str(chapa), int(dia),
                    pd.to_datetime(dt).strftime("%Y-%m-%d"),
                    dia_sem,
                    status,
                    h_ent,
                    h_sai,
                ))
            except Exception:
                # não derruba o app por causa de uma linha ruim
                continue

    con.commit()
    con.close()
    try:
        st.cache_data.clear()
    except Exception:
        pass

def load_escala_mes_db(setor: str, ano: int, mes: int):
    con = db_conn()
    cur = con.cursor()
    cur.execute("""
        SELECT chapa, data, dia_sem, status, h_entrada, h_saida
        FROM escala_mes
        WHERE setor=? AND ano=? AND mes=?
        ORDER BY chapa, date(data) ASC
    """, (setor, int(ano), int(mes)))
    rows = cur.fetchall()
    con.close()
    if not rows:
        return {}
    hist = {}
    for chapa, data_s, dia_sem, status, h_ent, h_sai in rows:
        dt = pd.to_datetime(data_s)
        hist.setdefault(chapa, []).append({
            "Data": dt, "Dia": dia_sem, "Status": status,
            "H_Entrada": h_ent or "", "H_Saida": h_sai or ""
        })
    return {ch: pd.DataFrame(items) for ch, items in hist.items()}

def apply_overrides_to_hist(setor: str, ano: int, mes: int, hist_db: dict[str, pd.DataFrame]):
    """
    Aplica overrides no histórico carregado do banco.
    REGRA GERAL:
    - "Férias" só existe se estiver na tabela ferias.
    - Se encontrar "Férias" no banco mas NÃO estiver na tabela, vira "Trabalho".
    """
    ov = load_overrides(setor, ano, mes)
    if (ov is None or ov.empty) and not hist_db:
        return hist_db

    # aplica overrides (se houver)
    if ov is not None and not ov.empty and hist_db:
        for _, r in ov.iterrows():
            ch = str(r["chapa"])
            dia = int(r["dia"])
            campo = str(r["campo"])
            valor = str(r["valor"])
            if ch not in hist_db:
                continue

            df = hist_db[ch].copy()
            idx = dia - 1
            if idx < 0 or idx >= len(df):
                continue

            data_obj = pd.to_datetime(df.loc[idx, "Data"]).date()

            if campo == "status":
                if valor == "Férias" and not is_de_ferias(setor, ch, data_obj):
                    pass
                else:
                    df.loc[idx, "Status"] = valor
                    if valor not in WORK_STATUSES:
                        df.loc[idx, "H_Entrada"] = ""
                        df.loc[idx, "H_Saida"] = ""

            elif campo == "h_entrada":
                df.loc[idx, "H_Entrada"] = valor
                if df.loc[idx, "Status"] in WORK_STATUSES:
                    df.loc[idx, "H_Saida"] = _saida_from_entrada(valor)

            elif campo == "h_saida":
                df.loc[idx, "H_Saida"] = valor

            hist_db[ch] = df

    # ✅ SANITIZA: força férias SOMENTE pela tabela ferias
    if hist_db:
        colaboradores = load_colaboradores_setor(setor)
        colab_by = {c["Chapa"]: c for c in colaboradores}

        for ch, df in list(hist_db.items()):
            ent_pad = colab_by.get(ch, {}).get("Entrada", "06:00")
            df2 = df.copy()
            for i in range(len(df2)):
                data_obj = pd.to_datetime(df2.loc[i, "Data"]).date()
                in_ferias = is_de_ferias(setor, ch, data_obj)

                if in_ferias:
                    df2.loc[i, "Status"] = "Férias"
                    df2.loc[i, "H_Entrada"] = ""
                    df2.loc[i, "H_Saida"] = ""
                else:
                    if df2.loc[i, "Status"] == "Férias":
                        df2.loc[i, "Status"] = "Trabalho"
                        df2.loc[i, "H_Entrada"] = ent_pad
                        df2.loc[i, "H_Saida"] = _saida_from_entrada(ent_pad)

            hist_db[ch] = df2

    return hist_db

# =========================================================
# MOTOR
# =========================================================
def _dias_mes(ano: int, mes: int):
    qtd = calendar.monthrange(ano, mes)[1]
    return pd.date_range(start=f"{ano}-{mes:02d}-01", periods=qtd, freq="D")

def _nao_consecutiva_folga(df, idx):
    """
    Verifica se o índice 'idx' NÃO fica colado com outra folga (idx-1 ou idx+1).
    Usa iloc (posição) para evitar KeyError quando o índice do DF não é 0..N-1.
    """
    n = len(df)
    if n == 0:
        return True
    if idx > 0 and df.iloc[idx - 1]["Status"] == "Folga":
        return False
    if idx < n - 1 and df.iloc[idx + 1]["Status"] == "Folga":
        return False
    return True

def _set_trabalho(df, idx, ent_padrao, locked_status: set[int] | None = None):
    if _locked(locked_status, idx):
        return
    df.loc[idx, "Status"] = "Trabalho"
    if not (df.loc[idx, "H_Entrada"] or ""):
        df.loc[idx, "H_Entrada"] = ent_padrao
    df.loc[idx, "H_Saida"] = _saida_from_entrada(df.loc[idx, "H_Entrada"])

def _set_folga(df, idx, locked_status: set[int] | None = None):
    if _locked(locked_status, idx):
        return
    df.loc[idx, "Status"] = "Folga"
    df.loc[idx, "H_Entrada"] = ""
    df.loc[idx, "H_Saida"] = ""

def _set_ferias(df, idx, locked_status: set[int] | None = None):
    if _locked(locked_status, idx):
        return
    df.loc[idx, "Status"] = "Férias"
    df.loc[idx, "H_Entrada"] = ""
    df.loc[idx, "H_Saida"] = ""

def _set_balanco(df, idx, locked_status: set[int] | None = None):
    if _locked(locked_status, idx):
        return
    df.loc[idx, "Status"] = BALANCO_STATUS
    df.loc[idx, "H_Entrada"] = BALANCO_DIA_ENTRADA
    df.loc[idx, "H_Saida"] = BALANCO_DIA_SAIDA

def _semana_seg_dom_indices(datas: pd.DatetimeIndex, idx_any: int):
    """Retorna índices da semana SEG->DOM do item idx_any.
    Robustez: ignora NaT para evitar TypeError em comparações.
    """
    d = datas[idx_any]
    if pd.isna(d):
        return []
    monday = d - timedelta(days=int(d.weekday()))
    sunday = monday + timedelta(days=6)

    out = []
    for i, dd in enumerate(datas):
        if pd.isna(dd):
            continue
        if monday.date() <= dd.date() <= sunday.date():
            out.append(i)
    return out

def _all_weeks_seg_dom(datas: pd.DatetimeIndex):
    weeks, seen = [], set()
    for i in range(len(datas)):
        w = tuple(_semana_seg_dom_indices(datas, i))
        if w and w not in seen:
            seen.add(w)
            weeks.append(list(w))
    return weeks

# =========================================================
# ✅ DOMINGO 1x1 POR COLABORADOR (GLOBAL, RESPEITA LOCK/FÉRIAS)
# =========================================================
def enforce_sundays_1x1_for_employee(
    df: pd.DataFrame,
    ent_padrao: str,
    locked_status: set[int] | None = None,
    base_first: str | None = None
):
    domingos = [i for i in range(len(df)) if df.loc[i, "Data"].day_name() == "Sunday"]
    # 🔥 REGRA SUPREMA: se existir domingo travado por override, não aplicar domingo 1x1.
    # Manual manda (pode trabalhar/folgar quantos domingos quiser).
    if locked_status and any(i in locked_status for i in domingos):
        return
    if not domingos:
        return

    def _normalize_dom_status(i: int) -> str | None:
        stt = df.loc[i, "Status"]
        if stt == "Férias":
            return None
        if stt == "Folga":
            return "Folga"
        if stt in WORK_STATUSES:
            return "Trabalho"
        return None

    def _force_dom(i: int, val: str):
        if _locked(locked_status, i):
            return
        if df.loc[i, "Status"] == "Férias":
            return
        if val == "Folga":
            _set_folga(df, i, locked_status=locked_status)
        else:
            df.loc[i, "H_Entrada"] = ent_padrao
            _set_trabalho(df, i, ent_padrao, locked_status=locked_status)

    first_idx = domingos[0]
    if not _locked(locked_status, first_idx) and df.loc[first_idx, "Status"] != "Férias":
        if base_first in ("Trabalho", "Folga"):
            _force_dom(first_idx, base_first)

    cur = None
    for i in domingos:
        norm = _normalize_dom_status(i)
        if norm in ("Trabalho", "Folga"):
            cur = norm
            break
    if cur is None:
        return

    for i in domingos:
        if df.loc[i, "Status"] == "Férias":
            continue

        if _locked(locked_status, i):
            norm = _normalize_dom_status(i)
            if norm in ("Trabalho", "Folga"):
                cur = norm
            continue

        _force_dom(i, cur)
        cur = "Folga" if cur == "Trabalho" else "Trabalho"

# =========================================================
# ✅ DESCANSO GLOBAL 11:10 (corrigido para NÃO criar folga consecutiva)
# =========================================================
def enforce_global_rest_keep_targets(df: pd.DataFrame, ent_padrao: str, locked_status: set[int] | None = None, ultima_saida_prev: str | None = None):
    # mantém horário fixo de balanço
    for i in range(len(df)):
        if df.loc[i, "Status"] == BALANCO_STATUS:
            df.loc[i, "H_Entrada"] = BALANCO_DIA_ENTRADA
            df.loc[i, "H_Saida"] = BALANCO_DIA_SAIDA

    last_saida = (ultima_saida_prev or "").strip()

    for i in range(len(df)):
        stt = df.loc[i, "Status"]

        # 🔥 REGRA SUPREMA (MANUAL): se este dia está TRAVADO por override,
        # NENHUMA regra automática pode mexer (nem status, nem horários).
        if _locked(locked_status, i):
            if stt not in WORK_STATUSES:
                df.loc[i, "H_Entrada"] = ""
                df.loc[i, "H_Saida"] = ""
                last_saida = ""
            else:
                if stt == BALANCO_STATUS:
                    df.loc[i, "H_Entrada"] = BALANCO_DIA_ENTRADA
                    df.loc[i, "H_Saida"] = BALANCO_DIA_SAIDA
                else:
                    ent_fix = (df.loc[i, "H_Entrada"] or "").strip() or ent_padrao
                    df.loc[i, "H_Entrada"] = ent_fix
                    if not (df.loc[i, "H_Saida"] or ""):
                        df.loc[i, "H_Saida"] = _saida_from_entrada(ent_fix)
                last_saida = (df.loc[i, "H_Saida"] or "").strip()
            continue

        if stt not in WORK_STATUSES:
            df.loc[i, "H_Entrada"] = ""
            df.loc[i, "H_Saida"] = ""
            last_saida = ""
            continue

        if stt == BALANCO_STATUS:
            last_saida = (df.loc[i, "H_Saida"] or "").strip()
            continue

        target = (df.loc[i, "H_Entrada"] or "").strip() or ent_padrao

        if not last_saida:
            df.loc[i, "H_Entrada"] = target
            df.loc[i, "H_Saida"] = _saida_from_entrada(target)
            last_saida = df.loc[i, "H_Saida"]
            continue

        min_ent = _add_min(last_saida, INTERSTICIO_MIN)

        s = _to_min(last_saida)
        e_t = _to_min(target)
        e_min = _to_min(min_ent)
        if e_t <= s: e_t += 1440
        if e_min <= s: e_min += 1440

        if e_t >= e_min:
            df.loc[i, "H_Entrada"] = target
            df.loc[i, "H_Saida"] = _saida_from_entrada(target)
            last_saida = df.loc[i, "H_Saida"]
            continue

        prev = i - 1
        if prev >= 0:
            # tenta ajustar o dia anterior (saída mais cedo) sem virar folga
            if (
                df.loc[prev, "Status"] == "Trabalho"
                and not _locked(locked_status, prev)
            ):
                saida_req = _sub_min(target, INTERSTICIO_MIN)
                ent_req = _sub_min(saida_req, DURACAO_JORNADA)
                df.loc[prev, "H_Entrada"] = ent_req
                df.loc[prev, "H_Saida"] = _saida_from_entrada(ent_req)
                last_saida = df.loc[prev, "H_Saida"]

                df.loc[i, "H_Entrada"] = target
                df.loc[i, "H_Saida"] = _saida_from_entrada(target)
                last_saida = df.loc[i, "H_Saida"]
                continue

            # plano B: folgar o dia anterior SÓ se NÃO gerar folga consecutiva
            if prev >= 0 and not _locked(locked_status, prev) and df.loc[prev, "Status"] != "Férias":
                if _nao_consecutiva_folga(df, prev):
                    _set_folga(df, prev, locked_status=locked_status)
                    last_saida = ""
                    df.loc[i, "H_Entrada"] = target
                    df.loc[i, "H_Saida"] = _saida_from_entrada(target)
                    last_saida = df.loc[i, "H_Saida"]
                    continue
                else:
                    # alternativa: empurra o dia atual (não cria folga seguida)
                    ent_ok = _ajustar_para_intersticio(target, last_saida)
                    df.loc[i, "H_Entrada"] = ent_ok
                    df.loc[i, "H_Saida"] = _saida_from_entrada(ent_ok)
                    last_saida = df.loc[i, "H_Saida"]
                    continue

        # fallback final: empurra entrada
        ent_ok = _ajustar_para_intersticio(target, last_saida)
        df.loc[i, "H_Entrada"] = ent_ok
        df.loc[i, "H_Saida"] = _saida_from_entrada(ent_ok)
        last_saida = df.loc[i, "H_Saida"]

# =========================================================
# ✅ 5x2: máxima sequência de trabalho = 5
# =========================================================
def enforce_max_5_consecutive_work(df, ent_padrao, pode_folgar_sabado: bool, initial_consec: int = 0, locked_status: set[int] | None = None):
    # Segurança: garante índice 0..N-1 (evita KeyError por índice quebrado)
    df.reset_index(drop=True, inplace=True)

    def can_make_folga(i):
        # Só converte TRABALHO normal em folga (não mexe em Balanço)
        if _locked(locked_status, i):
            return False
        if df.iloc[i]["Status"] != "Trabalho":
            return False
        dia = df.iloc[i]["Dia"]
        if dia == "dom":
            return False
        if dia == "sáb" and not pode_folgar_sabado:
            return False
        if not _nao_consecutiva_folga(df, i):
            return False
        return True

    consec, i = int(initial_consec), 0
    while i < len(df):
        if df.iloc[i]["Status"] in WORK_STATUSES:
            consec += 1
            if consec > 5:
                block_start = i - (consec - 1)
                block_end = i
                candidatos = []
                for j in range(block_start, block_end + 1):
                    if can_make_folga(j):
                        dia = df.iloc[j]["Dia"]
                        weekday_prio = 0 if dia in ["seg", "ter", "qua", "qui", "sex"] else 1
                        mid = (block_start + block_end) / 2
                        dist = abs(j - mid)
                        candidatos.append((weekday_prio, dist, j))
                if candidatos:
                    candidatos.sort()
                    escolhido = candidatos[0][2]
                    _set_folga(df, escolhido, locked_status=locked_status)
                    consec = 0
                    i = max(0, escolhido - 2)
                    continue
                else:
                    consec = 0
        else:
            consec = 0
        i += 1

def enforce_weekly_folga_targets(df: pd.DataFrame, df_ref: pd.DataFrame, pode_folgar_sabado: bool, locked_status: set[int] | None = None, setor: str | None = None, ch: str | None = None):
    """
    SEMANA SEG->DOM (regra geral):
      - Se DOM = Folga => 1 folga SEG-SÁB
      - Se DOM = Trabalho/Balanço => 2 folgas SEG-SÁB
      - Sábado só se permitido
      - Não cria folga consecutiva (exceto travado)
    Ajusta semana para cumprir o alvo (se outras regras mexerem depois).
    """
    datas = pd.to_datetime(df["Data"])
    weeks = _all_weeks_seg_dom(pd.DatetimeIndex(datas))

    def is_dom(i): return df_ref.loc[i, "Dia"] == "dom"

    def target_for_week(week):
        doms = [i for i in week if is_dom(i)]
        if not doms:
            return 2
        stt = df.loc[doms[0], "Status"]
        return 1 if stt == "Folga" else 2

    def can_turn_folga(i):
        if _locked(locked_status, i): return False
        if is_dom(i): return False
        if df.loc[i, "Status"] != "Trabalho": return False
        if df_ref.loc[i, "Dia"] == "sáb" and not pode_folgar_sabado: return False
        if not _nao_consecutiva_folga(df, i): return False
        return True

    def can_turn_trabalho(i):
        if _locked(locked_status, i): return False
        if is_dom(i): return False
        return df.loc[i, "Status"] == "Folga"

    for week in weeks:
        week = list(week)

        week_start = pd.to_datetime(df.loc[week[0], "Data"]).date()
        if setor and ch and is_first_week_after_return(setor, ch, week_start):
            # Semana de retorno de férias: não força folgas adicionais aqui.
            # O balanceamento do 1º domingo pós-férias é tratado separadamente.
            continue

        weekdays = [i for i in week if not is_dom(i)]
        t = target_for_week(week)

        cur = int((df.loc[weekdays, "Status"] == "Folga").sum())

        # excesso => remove
        if cur > t:
            cands = [i for i in weekdays if can_turn_trabalho(i)]
            def pr(i):
                return (0 if df_ref.loc[i, "Dia"] == "sáb" else 1, i)
            cands.sort(key=pr)
            for i in cands:
                if cur <= t: break
                _set_trabalho(df, i, ent_padrao="", locked_status=locked_status)  # entrada será re-setada depois pelo descanso global
                cur -= 1

        # falta => adiciona
        if cur < t:
            cands = [i for i in weekdays if can_turn_folga(i)]
            def pr2(i):
                return (0 if df_ref.loc[i, "Dia"] in ["seg","ter","qua","qui","sex"] else 1, i)
            cands.sort(key=pr2)
            for i in cands:
                if cur >= t: break
                _set_folga(df, i, locked_status=locked_status)
                cur += 1

    enforce_no_consecutive_folga(df, locked_status=locked_status)

def _counts_folgas_day_and_hour(hist_by_chapa: dict, colab_by_chapa: dict, chapas_grupo: list, idxs_semana: list, df_ref):
    counts_day = {i: 0 for i in idxs_semana}
    counts_day_hour = {}
    for ch in chapas_grupo:
        df = hist_by_chapa[ch]
        bucket = colab_by_chapa[ch].get("Entrada", "06:00")
        for i in idxs_semana:
            if df_ref.loc[i, "Dia"] == "dom":
                continue
            if df.loc[i, "Status"] == "Folga":
                counts_day[i] += 1
                counts_day_hour[(i, bucket)] = counts_day_hour.get((i, bucket), 0) + 1
    return counts_day, counts_day_hour

# =========================================================
# ✅ REBALANCE (corrigido): recebe estado_prev e respeita locked_idx
# =========================================================
def rebalance_folgas_dia(
    hist_by_chapa: dict,
    colab_by_chapa: dict,
    chapas_grupo: list,
    weeks: list,
    df_ref,
    estado_prev: dict | None = None,
    locked_idx: dict | None = None,
    past_flag: bool = False,
    max_iters=2200
):
    """
    Correções:
    - NÃO usa variável global: estado_prev é parâmetro (evita NameError)
    - Não faz swap em células travadas por override (locked_idx)
    """
    estado_prev = estado_prev or {}
    locked_idx = locked_idx or {}

    _past = bool(past_flag)

    def is_dom(i): return df_ref.loc[i, "Dia"] == "dom"

    def is_locked(ch, i):
        return bool(i in (locked_idx.get(ch, set()) or set()))

    def can_swap(ch, i_from, i_to):
        df = hist_by_chapa[ch]
        pode_sab = bool(colab_by_chapa[ch].get("Folga_Sab", False))

        if is_dom(i_from) or is_dom(i_to): return False
        if is_locked(ch, i_from) or is_locked(ch, i_to): return False

        if df.loc[i_from, "Status"] == "Férias" or df.loc[i_to, "Status"] == "Férias": return False
        if df.loc[i_from, "Status"] != "Folga": return False
        if df.loc[i_to, "Status"] != "Trabalho": return False
        if df_ref.loc[i_to, "Dia"] == "sáb" and not pode_sab: return False
        if (i_to > 0 and df.loc[i_to - 1, "Status"] == "Folga") or (i_to < len(df)-1 and df.loc[i_to + 1, "Status"] == "Folga"):
            return False
        return True

    def do_swap(ch, i_from, i_to):
        df = hist_by_chapa[ch]
        ent = colab_by_chapa[ch].get("Entrada", "06:00")
        pode_sab = bool(colab_by_chapa[ch].get("Folga_Sab", False))

        _set_trabalho(df, i_from, ent, locked_status=locked_idx.get(ch, set()))
        _set_folga(df, i_to, locked_status=locked_idx.get(ch, set()))

        enforce_max_5_consecutive_work(
            df, ent, pode_sab,
            initial_consec=(0 if _past else int((estado_prev.get(ch, {}) or {}).get('consec_trab_final', 0))),
            locked_status=locked_idx.get(ch, set())
        )
        hist_by_chapa[ch] = df

    it = 0
    for week in weeks:
        week_idxs = [i for i in week if not is_dom(i)]
        if not week_idxs:
            continue
        while it < max_iters:
            it += 1
            counts = {i: 0 for i in week_idxs}
            for ch in chapas_grupo:
                df = hist_by_chapa[ch]
                for i in week_idxs:
                    if df.loc[i, "Status"] == "Folga":
                        counts[i] += 1
            mx = max(counts, key=lambda x: counts[x])
            mn = min(counts, key=lambda x: counts[x])
            if counts[mx] - counts[mn] <= 1:
                break
            candidates = [ch for ch in chapas_grupo if hist_by_chapa[ch].loc[mx, "Status"] == "Folga" and hist_by_chapa[ch].loc[mn, "Status"] == "Trabalho"]
            random.shuffle(candidates)
            moved = False
            for ch in candidates:
                if can_swap(ch, mx, mn):
                    do_swap(ch, mx, mn)
                    moved = True
                    break
            if not moved:
                break

# =========================================================
# GERAR ESCALA — POR SUBGRUPO
# =========================================================
def gerar_escala_setor_por_subgrupo(setor: str, colaboradores: list[dict], ano: int, mes: int, respeitar_ajustes: bool = True):
    datas = _dias_mes(ano, mes)
    weeks = _all_weeks_seg_dom(datas)
    df_ref = pd.DataFrame({"Data": datas, "Dia": [D_PT[d.day_name()] for d in datas]})
    # Meses passados: não aplicar continuidade/travamentos do mês anterior.
    _past = is_past_competencia(ano, mes)
    estado_prev = {} if _past else load_estado_prev(setor, ano, mes)

    ovmap = _ov_map(setor, int(ano), int(mes)) if respeitar_ajustes else {}

    grupos = {}
    for c in colaboradores:
        sg = (c.get("Subgrupo") or "").strip() or "SEM SUBGRUPO"
        grupos.setdefault(sg, []).append(c)

    regras_cache = {}
    for sg in grupos.keys():
        if sg == "SEM SUBGRUPO":
            regras_cache[sg] = {"seg": 0, "ter": 0, "qua": 0, "qui": 0, "sex": 0, "sáb": 0}
        else:
            regras_cache[sg] = get_subgrupo_regras(setor, sg)

    hist_all = {}
    colab_by_chapa = {c["Chapa"]: c for c in colaboradores}
    locked_idx = {}

    # base de cada colaborador
    for c in colaboradores:
        ch = c["Chapa"]
        df = df_ref.copy()
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
        df["Status"] = "Trabalho"
        df["H_Entrada"] = ""
        df["H_Saida"] = ""

        # ✅ férias só via tabela ferias
        for i, d in enumerate(datas):
            if is_de_ferias(setor, ch, d.date()):
                df.loc[i, "Status"] = "Férias"
                df.loc[i, "H_Entrada"] = ""
                df.loc[i, "H_Saida"] = ""

        if respeitar_ajustes:
            _apply_overrides_to_df_inplace(df, setor, ch, ovmap)

        locked = set()
        if respeitar_ajustes:
            for i in range(len(df)):
                if _is_status_locked(ovmap, ch, pd.to_datetime(df.loc[i, "Data"])):
                    locked.add(i)
        locked_idx[ch] = locked
        hist_all[ch] = df

    chapas_all = list(hist_all.keys())  # lista de chapas no setor (para balanceamento pós-férias)
    # ✅ Domingo 1x1 por colaborador COM CONTINUIDADE ENTRE MESES
    for ch, df in hist_all.items():
        ent = colab_by_chapa[ch].get("Entrada", "06:00")
        locked = locked_idx.get(ch, set())

        if _past:
            base_first = None
        else:
            prev_dom = (estado_prev.get(ch, {}) or {}).get("ultimo_domingo_status", None)
            if prev_dom == "Folga":
                base_first = "Trabalho"
            elif prev_dom == "Trabalho":
                base_first = "Folga"
            else:
                base_first = None

        enforce_sundays_1x1_for_employee(df, ent, locked_status=locked, base_first=base_first)

        balance_primeiro_domingo_pos_ferias(df, hist_all, df_ref, setor, ch, chapas_all, locked_idx, ate_data=pd.to_datetime(df_ref.loc[len(df_ref)-1, "Data"]).date())
        hist_all[ch] = df

    # =====================================================
    # ✅ REGRA SEMANAL NOVA (SEG->DOM) DEPENDE DO DOMINGO
    # =====================================================
    for sg, membros in grupos.items():
        chapas = [m["Chapa"] for m in membros]
        if not chapas:
            continue

        pref = regras_cache.get(sg, {"seg": 0, "ter": 0, "qua": 0, "qui": 0, "sex": 0, "sáb": 0})

        for week in weeks:
            idxs_week = sorted(week, key=lambda i: df_ref.loc[i, "Data"])
            domingos = [i for i in idxs_week if df_ref.loc[i, "Dia"] == "dom"]
            dom_idx = domingos[0] if domingos else None

            for ch in chapas:
                df = hist_all[ch]
                locked = locked_idx.get(ch, set())
                pode_sab = bool(colab_by_chapa[ch].get("Folga_Sab", False))
                ent_bucket = colab_by_chapa[ch].get("Entrada", "06:00")

                segunda_idx = idxs_week[0]
                segunda_date = df_ref.loc[segunda_idx, "Data"].date()
                if is_first_week_after_return(setor, ch, segunda_date):
                    continue

                # candidatos seg-sex e sábado só se permitido
                cand_days = []
                for i in idxs_week:
                    dia = df_ref.loc[i, "Dia"]
                    if dia == "dom":
                        continue
                    if dia == "sáb" and not pode_sab:
                        continue
                    cand_days.append(i)

                if dom_idx is None:
                    target_folgas = 2
                else:
                    dom_status = df.loc[dom_idx, "Status"]
                    target_folgas = 1 if dom_status == "Folga" else 2

                folgas_sem = int((df.loc[cand_days, "Status"] == "Folga").sum()) if cand_days else 0

                while folgas_sem < target_folgas:
                    counts_day, counts_day_hour = _counts_folgas_day_and_hour(hist_all, colab_by_chapa, chapas, cand_days, df_ref)

                    possiveis = []
                    for j in cand_days:
                        if j in locked:
                            continue
                        dia = df_ref.loc[j, "Dia"]
                        if df.loc[j, "Status"] != "Trabalho":
                            continue
                        if dia == "sáb" and not pode_sab:
                            continue
                        if not _nao_consecutiva_folga(df, j):
                            continue
                        possiveis.append(j)

                    if not possiveis:
                        break

                    random.shuffle(possiveis)

                    def score(j):
                        dia = df_ref.loc[j, "Dia"]
                        weekday_prio = 0 if dia in ["seg", "ter", "qua", "qui", "sex"] else 1
                        pref_pen = PREF_EVITAR_PENALTY if pref.get(dia, 0) == 1 else 0
                        return (
                            counts_day.get(j, 0),
                            counts_day_hour.get((j, ent_bucket), 0),
                            pref_pen,
                            weekday_prio,
                            random.random()
                        )

                    possiveis.sort(key=score)
                    pick = possiveis[0]
                    _set_folga(df, pick, locked_status=locked)
                    folgas_sem += 1
                    hist_all[ch] = df

    # Pós: aplica regras globais por colaborador
    for ch, df in hist_all.items():
        ent = colab_by_chapa[ch].get("Entrada", "06:00")
        locked = locked_idx.get(ch, set())
        pode_sab = bool(colab_by_chapa[ch].get("Folga_Sab", False))

        if _past:
            base_first = None
        else:
            prev_dom = (estado_prev.get(ch, {}) or {}).get("ultimo_domingo_status", None)
            if prev_dom == "Folga":
                base_first = "Trabalho"
            elif prev_dom == "Trabalho":
                base_first = "Folga"
            else:
                base_first = None
        enforce_sundays_1x1_for_employee(df, ent, locked_status=locked, base_first=base_first)

        balance_primeiro_domingo_pos_ferias(df, hist_all, df_ref, setor, ch, chapas_all, locked_idx, ate_data=pd.to_datetime(df_ref.loc[len(df_ref)-1, "Data"]).date())

        # 1) Garante 5 dias seguidos antes de mexer em metas semanais
        enforce_max_5_consecutive_work(
            df, ent, pode_sab,
            initial_consec=(0 if _past else int((estado_prev.get(ch, {}) or {}).get('consec_trab_final', 0))),
            locked_status=locked
        )
        enforce_no_consecutive_folga(df, locked_status=locked)

        # 2) Metas semanais podem REMOVER folga => pode criar >5 de novo
        enforce_weekly_folga_targets(df, df_ref=df_ref, pode_folgar_sabado=pode_sab, locked_status=locked)

        # 3) Reforça novamente o limite de 5 depois das metas semanais
        enforce_max_5_consecutive_work(
            df, ent, pode_sab,
            initial_consec=(0 if _past else int((estado_prev.get(ch, {}) or {}).get('consec_trab_final', 0))),
            locked_status=locked
        )
        enforce_no_consecutive_folga(df, locked_status=locked)

        ultima_saida_prev = "" if _past else (estado_prev.get(ch, {}).get("ultima_saida", "") or "")
        enforce_global_rest_keep_targets(df, ent, locked_status=locked, ultima_saida_prev=ultima_saida_prev)

        # limpeza
        enforce_no_consecutive_folga(df, locked_status=locked)
        enforce_global_rest_keep_targets(df, ent, locked_status=locked, ultima_saida_prev=ultima_saida_prev)

        if respeitar_ajustes:
            _apply_overrides_to_df_inplace(df, setor, ch, ovmap)

        hist_all[ch] = df

    # rebalance por grupo (com estado_prev e travas)
    for sg, membros in grupos.items():
        chapas = [m["Chapa"] for m in membros]
        if chapas:
            rebalance_folgas_dia(
                hist_all, colab_by_chapa, chapas, weeks, df_ref,
                estado_prev=estado_prev,
                locked_idx=locked_idx,
                past_flag=_past,
                max_iters=2200
            )

    # ✅ Pós-rebalance: re-garante regra dos 5 dias por colaborador
    for ch, df in hist_all.items():
        ent = colab_by_chapa[ch].get("Entrada", "06:00")
        locked = locked_idx.get(ch, set())
        pode_sab = bool(colab_by_chapa[ch].get("Folga_Sab", False))

        enforce_max_5_consecutive_work(
            df, ent, pode_sab,
            initial_consec=(0 if _past else int((estado_prev.get(ch, {}) or {}).get('consec_trab_final', 0))),
            locked_status=locked
        )
        enforce_no_consecutive_folga(df, locked_status=locked)
        hist_all[ch] = df

    # Pós final (garantia)
    for ch, df in hist_all.items():
        ent = colab_by_chapa[ch].get("Entrada", "06:00")
        locked = locked_idx.get(ch, set())
        ultima_saida_prev = "" if _past else (estado_prev.get(ch, {}).get("ultima_saida", "") or "")

        if _past:
            base_first = None
        else:
            prev_dom = (estado_prev.get(ch, {}) or {}).get("ultimo_domingo_status", None)
            if prev_dom == "Folga":
                base_first = "Trabalho"
            elif prev_dom == "Trabalho":
                base_first = "Folga"
            else:
                base_first = None
        enforce_sundays_1x1_for_employee(df, ent, locked_status=locked, base_first=base_first)

        balance_primeiro_domingo_pos_ferias(df, hist_all, df_ref, setor, ch, chapas_all, locked_idx, ate_data=pd.to_datetime(df_ref.loc[len(df_ref)-1, "Data"]).date())
        enforce_no_consecutive_folga(df, locked_status=locked)
        enforce_weekly_folga_targets(df, df_ref=df_ref, pode_folgar_sabado=bool(colab_by_chapa[ch].get('Folga_Sab', False)), locked_status=locked)

        # ✅ garante 5 dias depois do weekly (porque weekly pode remover folga)
        enforce_max_5_consecutive_work(
            df, ent, bool(colab_by_chapa[ch].get('Folga_Sab', False)),
            initial_consec=(0 if _past else int((estado_prev.get(ch, {}) or {}).get('consec_trab_final', 0))),
            locked_status=locked
        )
        enforce_no_consecutive_folga(df, locked_status=locked)

        enforce_global_rest_keep_targets(df, ent, locked_status=locked, ultima_saida_prev=ultima_saida_prev)

        if respeitar_ajustes:
            _apply_overrides_to_df_inplace(df, setor, ch, ovmap)

        hist_all[ch] = df

    # Estado do mês
    estado_out = {}
    for ch, df in hist_all.items():
        consec = 0
        for i in range(len(df) - 1, -1, -1):
            if df.loc[i, "Status"] in WORK_STATUSES:
                consec += 1
            else:
                break

        ultima_saida = ""
        for i in range(len(df) - 1, -1, -1):
            if df.loc[i, "Status"] in WORK_STATUSES and (df.loc[i, "H_Saida"] or ""):
                ultima_saida = df.loc[i, "H_Saida"]
                break

        ultimo_dom = None
        for i in range(len(df) - 1, -1, -1):
            if df.loc[i, "Dia"] == "dom":
                if df.loc[i, "Status"] == "Folga":
                    ultimo_dom = "Folga"
                    break
                if df.loc[i, "Status"] in WORK_STATUSES:
                    ultimo_dom = "Trabalho"
                    break

        estado_out[ch] = {"consec_trab_final": consec, "ultima_saida": ultima_saida, "ultimo_domingo_status": ultimo_dom}

    return hist_all, estado_out

# =========================================================
# DASHBOARD / CALENDÁRIO / BANCO DE HORAS
# (resto do arquivo igual ao seu original — UI completa)
# =========================================================

def banco_horas_df(hist_db: dict[str, pd.DataFrame], colab_by: dict, base_min: int):
    rows = []
    for ch, df in hist_db.items():
        nome = colab_by.get(ch, {}).get("Nome", ch)
        saldo = 0
        for _, r in df.iterrows():
            if r["Status"] not in WORK_STATUSES:
                continue
            ent = r.get("H_Entrada", "") or ""
            sai = r.get("H_Saida", "") or ""
            if not ent or not sai:
                continue
            dur = _to_min(sai) - _to_min(ent)
            if dur < 0:
                dur += 24 * 60
            saldo += (dur - base_min)
        rows.append({"Nome": nome, "Chapa": ch, "Saldo_min": saldo, "Saldo_h": round(saldo/60, 2)})
    return pd.DataFrame(rows).sort_values(["Saldo_min"], ascending=False)

def calendario_rh_df(hist_db: dict[str, pd.DataFrame], colab_by: dict):
    if not hist_db:
        return pd.DataFrame()
    any_df = next(iter(hist_db.values()))
    dias = [str(int(r.day)) for r in pd.to_datetime(any_df["Data"]).dt.date]
    cols = ["Nome", "Chapa", "Subgrupo"] + dias
    rows = []
    for ch, df in hist_db.items():
        nome = colab_by.get(ch, {}).get("Nome", ch)
        sg = (colab_by.get(ch, {}).get("Subgrupo", "") or "").strip() or "SEM SUBGRUPO"
        row = [nome, ch, sg]
        for i in range(len(df)):
            stt = df.loc[i, "Status"]
            if stt == "Folga":
                row.append("F")
            elif stt == "Férias":
                row.append("FER")
            else:
                row.append(df.loc[i, "H_Entrada"] or "")
        rows.append(row)
    out = pd.DataFrame(rows, columns=cols)
    return out.sort_values(["Subgrupo", "Nome"]).reset_index(drop=True)

def style_calendario(df: pd.DataFrame, mes: int, ano: int):
    if df.empty:
        return df
    dias_cols = df.columns[3:]
    qtd = calendar.monthrange(int(ano), int(mes))[1]
    dsem = {}
    for d in range(1, qtd + 1):
        ds = pd.Timestamp(year=int(ano), month=int(mes), day=int(d)).day_name()
        dsem[str(d)] = D_PT[ds]




# =========================================================
# MAPA ANUAL DE FÉRIAS (visual tipo "grade")
# =========================================================
MESES_PT = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

def _parse_date_ymd(s: str):
    try:
        return datetime.strptime(str(s), "%Y-%m-%d").date()
    except Exception:
        return None

def ferias_mapa_ano_df(setor: str, ano: int, colaboradores: list[dict]) -> pd.DataFrame:
    """
    DF:
      Nome | Chapa | Janeiro..Dezembro
    Marca "FER" quando houver QUALQUER dia de férias no mês.
    """
    rows = list_ferias(setor)  # [(chapa,inicio,fim), ...]
    fer_by = {}
    for chapa, ini, fim in rows:
        ini_d = _parse_date_ymd(ini)
        fim_d = _parse_date_ymd(fim)
        if not ini_d or not fim_d:
            continue
        fer_by.setdefault(str(chapa), []).append((ini_d, fim_d))

    colabs_sorted = sorted(colaboradores, key=lambda c: ((c.get("Nome") or ""), (c.get("Chapa") or "")))
    out = []
    for c in colabs_sorted:
        ch = str(c.get("Chapa") or "")
        nome = str(c.get("Nome") or ch)
        linha = {"Nome": nome, "Chapa": ch}
        periods = fer_by.get(ch, [])
        for m in range(1, 13):
            first = date(int(ano), m, 1)
            last = date(int(ano), m, calendar.monthrange(int(ano), m)[1])
            marcou = False
            for ini_d, fim_d in periods:
                if ini_d <= last and fim_d >= first:
                    marcou = True
                    break
            linha[MESES_PT[m-1]] = "FER" if marcou else ""
        out.append(linha)

    return pd.DataFrame(out, columns=["Nome","Chapa"] + MESES_PT)

def style_ferias_mapa(df: pd.DataFrame):
    if df is None or df.empty:
        return df
    meses = [c for c in df.columns if c in MESES_PT]

    def cell(v, col):
        if col in meses:
            if str(v) == "FER":
                return "background-color:#1F4E78; color:#FFFFFF; font-weight:800; text-align:center;"
            return "background-color:#F2F2F2; color:#000000; text-align:center;"
        if col == "Nome":
            return "font-weight:700;"
        return ""

    styles = pd.DataFrame("", index=df.index, columns=df.columns)
    for col in df.columns:
        styles[col] = df[col].apply(lambda v: cell(v, col))
    return df.style.apply(lambda _: styles, axis=None)


# =========================================================
# ÚLTIMAS FÉRIAS + ALERTA (1 ano e 11 meses) + DURAÇÃO
# =========================================================
def _months_between(d1: date, d2: date) -> int:
    """Meses inteiros aproximados entre datas (d2 >= d1)."""
    if not d1 or not d2:
        return 0
    if d2 < d1:
        d1, d2 = d2, d1
    return (d2.year - d1.year) * 12 + (d2.month - d1.month)

def get_ultima_ferias_info(setor: str, chapa: str):
    """
    Retorna dict com:
      - ultima_inicio (date|None)
      - ultima_fim (date|None)
      - dias_ultima (int|None)
      - meses_desde_ultima_fim (int|None)  # até hoje
    Considera o período com maior 'fim' como a última.
    """
    chapa = str(chapa or "").strip()
    if not chapa:
        return {"ultima_inicio": None, "ultima_fim": None, "dias_ultima": None, "meses_desde_ultima_fim": None}

    rows = list_ferias(setor)  # [(chapa,inicio,fim), ...]
    last = None  # (fim_date, ini_date)
    for ch, ini, fim in rows:
        if str(ch) != chapa:
            continue
        ini_d = _parse_date_ymd(ini)
        fim_d = _parse_date_ymd(fim)
        if not ini_d or not fim_d:
            continue
        if last is None or fim_d > last[0]:
            last = (fim_d, ini_d)

    if not last:
        return {"ultima_inicio": None, "ultima_fim": None, "dias_ultima": None, "meses_desde_ultima_fim": None}

    ultima_fim, ultima_ini = last[0], last[1]
    dias = (ultima_fim - ultima_ini).days + 1
    meses = _months_between(ultima_fim, date.today())
    return {"ultima_inicio": ultima_ini, "ultima_fim": ultima_fim, "dias_ultima": dias, "meses_desde_ultima_fim": meses}

def _classificar_duracao_ferias(qtd_dias: int) -> str:
    if qtd_dias == 15:
        return "15 dias"
    if qtd_dias == 30:
        return "30 dias"
    if qtd_dias and qtd_dias > 0:
        return f"{qtd_dias} dias"
    return "-"



def ferias_resumo_mensal_df(setor: str, ano: int) -> pd.DataFrame:
    """
    Resumo mensal:
      - Pessoas_em_ferias: qtd de colaboradores com QUALQUER dia de férias no mês
      - Lancamentos: qtd de períodos (linhas) de férias que encostam no mês
    """
    rows = list_ferias(setor)  # [(chapa,inicio,fim), ...]
    # map month -> set(chapa) and count launches touching month
    people = {m: set() for m in range(1, 13)}
    launches = {m: 0 for m in range(1, 13)}

    for chapa, ini, fim in rows:
        ini_d = _parse_date_ymd(ini)
        fim_d = _parse_date_ymd(fim)
        if not ini_d or not fim_d:
            continue

        for m in range(1, 13):
            first = date(int(ano), m, 1)
            last = date(int(ano), m, calendar.monthrange(int(ano), m)[1])
            if ini_d <= last and fim_d >= first:
                people[m].add(str(chapa))
                launches[m] += 1

    data = []
    for m in range(1, 13):
        data.append({
            "Mês": MESES_PT[m-1],
            "Pessoas_em_ferias": len(people[m]),
            "Lancamentos": int(launches[m])
        })
    return pd.DataFrame(data)


# =========================================================
# PDF UI helpers (filtro estilo "Impressão de Escala")
# =========================================================
def _filtrar_colaboradores(colaboradores: list[dict], subgrupos_sel: list[str] | None, busca: str | None):
    subgrupos_sel = subgrupos_sel or []
    busca = (busca or "").strip().lower()
    out = []
    for c in colaboradores:
        sg = (c.get("Subgrupo") or "").strip() or "SEM SUBGRUPO"
        nome = (c.get("Nome") or "").strip()
        ch = (c.get("Chapa") or "").strip()
        if subgrupos_sel and sg not in subgrupos_sel:
            continue
        if busca:
            key = f"{nome} {ch} {sg}".lower()
            if busca not in key:
                continue
        out.append(c)
    return out

    def cell_style(v, col):
        if col in dias_cols:
            dia_sem = dsem.get(col, "")
            if str(v) == "F":
                return "background-color:#FFF2CC; color:#000000; font-weight:700;"
            if str(v) == "FER":
                return "background-color:#92D050; color:#000000; font-weight:700;"
            if dia_sem == "dom":
                return "background-color:#BDD7EE; color:#000000;"
        return ""

    styles = pd.DataFrame("", index=df.index, columns=df.columns)
    for c in df.columns:
        styles[c] = df[c].apply(lambda v: cell_style(v, c))
    return df.style.apply(lambda _: styles, axis=None)

# =========================================================
# UI
# =========================================================
if "auth" not in st.session_state:
    st.session_state["auth"] = None
if "cfg_mes" not in st.session_state:
    st.session_state["cfg_mes"] = datetime.now().month
if "cfg_ano" not in st.session_state:
    st.session_state["cfg_ano"] = datetime.now().year
if "last_seed" not in st.session_state:
    st.session_state["last_seed"] = 0


db_init()

def page_login():
    st.title("🔐 Login por Setor (Usuário / Líder / Admin)")
    tab_login, tab_cadastrar, tab_esqueci = st.tabs(["Entrar", "Cadastrar Usuário do Sistema", "Esqueci a senha"])

    with tab_login:
        con = db_conn()
        setores = pd.read_sql_query("SELECT nome FROM setores ORDER BY nome ASC", con)["nome"].tolist()
        con.close()

        setor = st.selectbox("Setor:", setores, key="lg_setor")
        chapa = st.text_input("Chapa:", key="lg_chapa")
        senha = st.text_input("Senha:", type="password", key="lg_senha")

        if st.button("Entrar", key="lg_btn"):
            u = verify_login(setor, chapa, senha)
            if u:
                st.session_state["auth"] = u
                st.success("Login efetuado!")
                st.rerun()
            else:
                st.error("Usuário ou senha inválidos.")

        st.caption("Admin padrão: setor ADMIN | chapa admin | senha 123")

    with tab_cadastrar:
        st.subheader("Cadastrar usuário do sistema (com senha)")
        st.info("⚠️ Somente usuário do sistema tem senha. Colaborador é SEM senha.")
        nome = st.text_input("Nome:", key="cl_nome")
        setor = st.text_input("Setor:", key="cl_setor").strip().upper()
        chapa = st.text_input("Chapa:", key="cl_chapa")
        senha = st.text_input("Senha:", type="password", key="cl_senha")
        senha2 = st.text_input("Confirmar senha:", type="password", key="cl_senha2")
        is_admin = st.checkbox("Admin?", key="cl_admin")
        is_lider = st.checkbox("Líder?", value=False, key="cl_lider")

        if st.button("Criar usuário", key="cl_btn"):
            if not nome or not setor or not chapa or not senha:
                st.error("Preencha tudo.")
            elif senha != senha2:
                st.error("Senhas não conferem.")
            elif system_user_exists(setor, chapa):
                st.error("Já existe.")
            else:
                create_system_user(nome.strip(), setor, chapa.strip(), senha, is_lider=1 if is_lider else 0, is_admin=1 if is_admin else 0)
                st.success("Criado! Faça login.")
                st.rerun()

    with tab_esqueci:
        st.subheader("Redefinir senha (com chapa do líder do setor)")
        con = db_conn()
        setores = pd.read_sql_query("SELECT nome FROM setores ORDER BY nome ASC", con)["nome"].tolist()
        con.close()

        setor = st.selectbox("Setor:", setores, key="fp_setor")
        chapa = st.text_input("Sua chapa (usuário do sistema):", key="fp_chapa")
        chapa_lider = st.text_input("Chapa do líder:", key="fp_lider")
        nova = st.text_input("Nova senha:", type="password", key="fp_nova")
        nova2 = st.text_input("Confirmar:", type="password", key="fp_nova2")

        if st.button("Redefinir", key="fp_btn"):
            if not chapa or not chapa_lider or not nova:
                st.error("Preencha.")
            elif nova != nova2:
                st.error("Senhas não conferem.")
            elif not system_user_exists(setor, chapa):
                st.error("Usuário não encontrado.")
            elif not is_lider_chapa(setor, chapa_lider):
                st.error("Chapa do líder inválida.")
            else:
                update_password(setor, chapa, nova)
                st.success("Senha alterada.")
                st.rerun()

def _regenerar_mes_inteiro(setor: str, ano: int, mes: int, seed: int = 0, respeitar_ajustes: bool = True):
    """
    Regera a escala do mês inteiro para TODO o setor.

    ✅ Garantias:
    - Se respeitar_ajustes=True, TODAS as folgas/alterações manuais (overrides) são reaplicadas
      no final e gravadas novamente no banco (escala_mes). Isso evita “sumir” folga manual ao gerar.
    """
    colaboradores = load_colaboradores_setor(setor)
    if not colaboradores:
        return False

    random.seed(int(seed))
    hist, estado_out = gerar_escala_setor_por_subgrupo(
        setor, colaboradores, int(ano), int(mes),
        respeitar_ajustes=bool(respeitar_ajustes)
    )

    # 1) grava a geração
    save_escala_mes_db(setor, int(ano), int(mes), hist)
    save_estado_mes(setor, int(ano), int(mes), estado_out)

    # 2) “pós-fix”: reaplica overrides do banco e grava de novo
    if bool(respeitar_ajustes):
        hist_db = load_escala_mes_db(setor, int(ano), int(mes))
        hist_db = apply_overrides_to_hist(setor, int(ano), int(mes), hist_db)
        if hist_db:
            save_escala_mes_db(setor, int(ano), int(mes), hist_db)

    return True



def page_app():
    auth = st.session_state.get("auth") or {}
    setor = auth.get("setor", "GERAL")

    # ---- Competência (mês/ano) compartilhada
    ano_cfg = int(st.session_state.get("cfg_ano", datetime.now().year))
    mes_cfg = int(st.session_state.get("cfg_mes", datetime.now().month))
    st.session_state["cfg_ano"] = ano_cfg
    st.session_state["cfg_mes"] = mes_cfg

    if "last_seed" not in st.session_state:
        st.session_state["last_seed"] = 0

    # =========================
    # SIDEBAR — Sessão + Competência
    # =========================
    with st.sidebar:
        st.title("👤 Sessão")
        st.caption("Acesso por setor (usuário / líder / admin)")

        cA, cB = st.columns([1, 1])
        cA.write(f"**Nome:** {auth.get('nome','-')}")
        cB.write(f"**Perfil:** {'ADMIN' if auth.get('is_admin', False) else ('LÍDER' if auth.get('is_lider', False) else 'USUÁRIO')}")

        st.write(f"**Setor:** {setor}")
        st.write(f"**Chapa:** {auth.get('chapa','-')}")

        st.markdown("<div class='hr'></div>", unsafe_allow_html=True)

        st.subheader("🗓️ Competência")
        m1, m2 = st.columns(2)
        mes_cfg = m1.selectbox("Mês", list(range(1, 13)), index=mes_cfg - 1, key="sb_mes")
        ano_cfg = m2.number_input("Ano", value=ano_cfg, step=1, key="sb_ano")
        st.session_state["cfg_mes"] = int(mes_cfg)
        st.session_state["cfg_ano"] = int(ano_cfg)

        st.markdown("<div class='hr'></div>", unsafe_allow_html=True)

        if st.button("🚪 Sair", use_container_width=True, key="logout_btn"):
            st.session_state["auth"] = None
            st.rerun()

        

    # =========================
    # KPIs
    # =========================
    ano_k = int(st.session_state["cfg_ano"])
    mes_k = int(st.session_state["cfg_mes"])

    colaboradores_k = load_colaboradores_setor(setor)
    total_colab = len(colaboradores_k)

    hist_db_kpi = load_escala_mes_db(setor, ano_k, mes_k)
    if hist_db_kpi:
        hist_db_kpi = apply_overrides_to_hist(setor, ano_k, mes_k, hist_db_kpi)

    folgas_mes = ferias_mes = trabalhos_mes = 0
    if hist_db_kpi:
        for _, dfk in hist_db_kpi.items():
            folgas_mes += int((dfk["Status"] == "Folga").sum())
            ferias_mes += int((dfk["Status"] == "Férias").sum())
            trabalhos_mes += int(dfk["Status"].isin(WORK_STATUSES).sum())

    k1, k2, k3, k4 = st.columns(4)
    k1.markdown(
        f"<div class='kpi-card'><div class='kpi-title'>Colaboradores</div>"
        f"<p class='kpi-value'>{total_colab}</p></div>",
        unsafe_allow_html=True
    )
    k2.markdown(
        f"<div class='kpi-card'><div class='kpi-title'>Dias de Folga (mês)</div>"
        f"<p class='kpi-value'>{folgas_mes}</p></div>",
        unsafe_allow_html=True
    )
    k3.markdown(
        f"<div class='kpi-card'><div class='kpi-title'>Dias de Férias (mês)</div>"
        f"<p class='kpi-value'>{ferias_mes}</p></div>",
        unsafe_allow_html=True
    )
    k4.markdown(
        f"<div class='kpi-card'><div class='kpi-title'>Dias de Trabalho (mês)</div>"
        f"<p class='kpi-value'>{trabalhos_mes}</p></div>",
        unsafe_allow_html=True
    )

    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)

    # =========================
    # ABAS
    # =========================
    tabs = ["👥 Colaboradores", "🚀 Gerar Escala", "⚙️ Ajustes", "🏖️ Férias", "📥 Excel"]
    is_admin_area = bool(auth.get("is_admin", False)) and setor == "ADMIN"
    if is_admin_area:
        tabs.append("🔒 Admin")

    abas = st.tabs(tabs)

    # ------------------------------------------------------
    # ABA 1: Colaboradores
    # ------------------------------------------------------
    with abas[0]:
        st.subheader("👥 Colaboradores (SEM senha)")
        colaboradores = load_colaboradores_setor(setor)

        if colaboradores:
            st.dataframe(pd.DataFrame([{
                "Nome": c["Nome"],
                "Chapa": c["Chapa"],
                "Subgrupo": c["Subgrupo"] or "SEM SUBGRUPO",
                "Entrada": c["Entrada"],
                "Folga Sábado": "Sim" if c["Folga_Sab"] else "Não",
            } for c in colaboradores]), use_container_width=True, height=420)
        else:
            st.info("Sem colaboradores.")

        st.markdown("---")
        with st.form("form_add_colaborador", clear_on_submit=True):
            c1, c2 = st.columns(2)
            nome_n = c1.text_input("Nome:", key="col_nome")
            chapa_n = c2.text_input("Chapa:", key="col_chapa")
            submitted = st.form_submit_button("Cadastrar colaborador", use_container_width=True)
            if submitted:
                if not nome_n or not chapa_n:
                    st.error("Preencha nome e chapa.")
                elif colaborador_exists(setor, chapa_n.strip()):
                    st.error("Já existe essa chapa.")
                else:
                    create_colaborador(nome_n.strip(), setor, chapa_n.strip())
                    st.success("Cadastrado!")
                    st.rerun()

        st.markdown("---")
        st.markdown("## 🗑️ Excluir colaborador")
        if colaboradores:
            ch_del = st.selectbox("Escolha a chapa para excluir:", [c["Chapa"] for c in colaboradores], key="del_chapa")
            st.warning("⚠️ Excluir remove também férias, ajustes, escala e estado desse colaborador no setor.")
            confirm = st.checkbox("Confirmo que quero excluir definitivamente", key="del_confirm")
            if st.button("Excluir colaborador", key="del_btn"):
                if not confirm:
                    st.error("Marque a confirmação para excluir.")
                else:
                    delete_colaborador_total(setor, ch_del)
                    st.success("Colaborador excluído!")
                    st.rerun()

        st.markdown("---")
        st.markdown("## ✏️ Editar perfil do colaborador")
        if colaboradores:
            chapas = [c["Chapa"] for c in colaboradores]
            nome_by_chapa = {c["Chapa"]: c.get("Nome", "") for c in colaboradores}
            ch_sel = st.selectbox(
                "Colaborador (Nome — Chapa):",
                chapas,
                key="pf_chapa",
                format_func=lambda ch: f"{(nome_by_chapa.get(ch, ch) or ch)} — {ch}",
            )
            csel = next(x for x in colaboradores if x["Chapa"] == ch_sel)

            colp1, colp2, colp3 = st.columns(3)
            # Entrada: usar presets (inclui 06:50 e 12:40) para facilitar
            ent_atual = (csel.get("Entrada") or BALANCO_DIA_ENTRADA).strip()
            if ent_atual not in HORARIOS_ENTRADA_PRESET:
                opcoes_ent = HORARIOS_ENTRADA_PRESET + [ent_atual]
            else:
                opcoes_ent = HORARIOS_ENTRADA_PRESET
            ent_sel = colp1.selectbox(
                "Entrada:",
                options=opcoes_ent,
                index=opcoes_ent.index(ent_atual),
                key="pf_ent_sel",
            )
            sg_opts = [""] + list_subgrupos(setor)
            idx_default = sg_opts.index(csel["Subgrupo"]) if csel["Subgrupo"] in sg_opts else 0
            sg = colp2.selectbox("Subgrupo:", sg_opts, index=idx_default, key="pf_sg")
            sab = colp3.checkbox("Permitir folga sábado", value=bool(csel["Folga_Sab"]), key="pf_sab")

            if st.button("Salvar perfil", key="pf_save"):
                update_colaborador_perfil(setor, ch_sel, sg, ent_sel, sab)
                st.success("Salvo!")
                st.rerun()

    # ------------------------------------------------------
    # ABA 2: Gerar Escala
    # ------------------------------------------------------
    with abas[1]:
        st.subheader("🚀 Gerar escala")
        st.caption(f"Competência ativa: **{int(st.session_state['cfg_mes']):02d}/{int(st.session_state['cfg_ano'])}**")

        with st.container(border=True):
            c1, c2, c3 = st.columns([1, 1, 2])
            mes = c1.selectbox("Mês:", list(range(1, 13)), index=int(st.session_state["cfg_mes"]) - 1, key="gen_mes")
            ano = c2.number_input("Ano:", value=int(st.session_state["cfg_ano"]), step=1, key="gen_ano")
            seed = c3.number_input("Semente", min_value=0, max_value=999999, value=int(st.session_state.get("last_seed", 0)), key="gen_seed")

        st.session_state["cfg_mes"] = int(mes)
        st.session_state["cfg_ano"] = int(ano)

        colaboradores = load_colaboradores_setor(setor)
        if not colaboradores:
            st.warning("Cadastre colaboradores.")
        else:
            b1, b2, _ = st.columns([1, 1, 6])
            if b1.button("🚀 Gerar agora (respeita ajustes)", use_container_width=True, key="gen_btn"):
                st.session_state["last_seed"] = int(seed)
                ok = _regenerar_mes_inteiro(setor, int(ano), int(mes), seed=int(seed), respeitar_ajustes=True)
                if ok:
                    st.success("Escala gerada (ajustes/travas preservados)!")
                else:
                    st.warning("Sem colaboradores.")
                st.rerun()

            if b2.button("🔄 Recarregar do banco", use_container_width=True, key="gen_reload_btn"):
                st.rerun()

            hist_db = load_escala_mes_db(setor, int(ano), int(mes))
            hist_db = apply_overrides_to_hist(setor, int(ano), int(mes), hist_db)

            if hist_db:
                colab_by = {c["Chapa"]: c for c in colaboradores}
                st.markdown("### 📅 Calendário RH (visual por colaborador)")
                cal = calendario_rh_df(hist_db, colab_by)
                show_color = st.checkbox("🎨 Mostrar cores no calendário (pode deixar lento)", value=False, key="cal_color")
                if show_color:
                    st.dataframe(style_calendario(cal, int(mes), int(ano)), use_container_width=True)
                else:
                    st.dataframe(cal, use_container_width=True)

                st.markdown("---")
                st.markdown("### 👤 Visualizar colaborador (detalhado)")
                ch_view = st.selectbox("Chapa:", list(hist_db.keys()), key="view_ch")
                st.dataframe(hist_db[ch_view], use_container_width=True, height=420)
            else:
                st.info("Sem escala no mês. Clique em **Gerar agora**.")

    # ------------------------------------------------------
    # ABA 3: Ajustes
    # ------------------------------------------------------
    with abas[2]:
        st.subheader("⚙️ Ajustes (travas) — sempre entram na geração")

        with st.container(border=True):
            c1, c2, c3 = st.columns([1, 1, 2])
            mes = c1.selectbox("Mês (ajustes)", list(range(1, 13)), index=int(st.session_state["cfg_mes"]) - 1, key="adj_mes")
            ano = c2.number_input("Ano (ajustes)", value=int(st.session_state["cfg_ano"]), step=1, key="adj_ano")
            c3.caption("Dica: deixe o mês/ano aqui igual ao mês/ano da aba 🚀 Gerar Escala.")

        st.session_state["cfg_mes"] = int(mes)
        st.session_state["cfg_ano"] = int(ano)

        hist_db = load_escala_mes_db(setor, ano, mes)
        colaboradores = load_colaboradores_setor(setor)
        colab_by = {c["Chapa"]: c for c in colaboradores}

        if not hist_db:
            st.info("Gere a escala primeiro na aba 🚀 Gerar Escala.")
        else:
            hist_db = apply_overrides_to_hist(setor, ano, mes, hist_db)

            tgrid, t2, t3, t4 = st.tabs([
                "🧩 Folgas manuais em grade",
                "📅 Trocar horário mês inteiro",
                "✅ Preferência por subgrupo",
                "📌 Subgrupos (editável)"
            ])

            with tgrid:
                st.markdown("### 🧩 Folgas manuais em grade (por colaborador)")
                st.caption("Marque/desmarque as folgas do mês. Isso cria/remove travas (overrides) de Status=Folga. Domingo é editável aqui (manual é soberano).")

                qtd = calendar.monthrange(int(ano), int(mes))[1]
                dias = list(range(1, qtd + 1))

                # pega overrides existentes
                ovdf = load_overrides(setor, ano, mes)
                ov_status = {}
                if ovdf is not None and not ovdf.empty:
                    od = ovdf[ovdf["campo"] == "status"]
                    for _, r in od.iterrows():
                        if str(r["valor"]) == "Folga":
                            ov_status.setdefault(str(r["chapa"]), set()).add(int(r["dia"]))

                # monta grade
                rows = []
                for c in colaboradores:
                    chg = str(c["Chapa"])
                    row = {"Nome": c["Nome"], "Chapa": chg}
                    dfh = hist_db.get(chg)
                    for d in dias:
                        if dfh is not None and len(dfh) >= d:
                            if dfh.loc[d - 1, "Status"] == "Férias":
                                row[str(d)] = False
                            else:
                                row[str(d)] = (dfh.loc[d - 1, "Status"] == "Folga") or (d in ov_status.get(chg, set()))
                        else:
                            row[str(d)] = False
                    rows.append(row)

                df_grid = pd.DataFrame(rows)
                edited = st.data_editor(
                    df_grid,
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed",
                    column_config={str(d): st.column_config.CheckboxColumn(str(d), width="small") for d in dias},
                    key="grid_editor"
                )

                auto_readequar = st.checkbox("🔄 Readequar escala ao salvar", value=True, key="grid_auto_regen")

                if st.button("💾 Salvar folgas manuais (e readequar mês)", key="grid_save"):
                    set_folga = 0
                    set_trab = 0
                    for _, r in edited.iterrows():
                        chg = str(r["Chapa"])
                        dfh = hist_db.get(chg)
                        ent_pad_local = colab_by.get(chg, {}).get("Entrada", "06:00")
                        for d in dias:
                            want_folga = bool(r[str(d)])
                            if dfh is not None and len(dfh) >= d:
                                if dfh.loc[d - 1, "Status"] == "Férias":
                                    continue

                            if want_folga:
                                set_override(setor, ano, mes, chg, d, "status", "Folga")
                                set_folga += 1
                            else:
                                # ✅ regra pedida: desmarcado = TRABALHO (travado)
                                set_override(setor, ano, mes, chg, d, "status", "Trabalho")
                                # mantém horário padrão no banco via geração/descanso global; se quiser travar horário também,
                                # descomente as linhas abaixo:
                                # set_override(setor, ano, mes, chg, d, "h_entrada", ent_pad_local)
                                # set_override(setor, ano, mes, chg, d, "h_saida", _saida_from_entrada(ent_pad_local))
                                set_trab += 1

                    if auto_readequar:
                        _regenerar_mes_inteiro(setor, ano, mes, seed=int(st.session_state.get("last_seed", 0)), respeitar_ajustes=True)

                    st.success(f"Salvo! Folgas travadas: {set_folga} | Trabalhos travados: {set_trab}.")
                    st.rerun()

            with t2:
                ch2 = st.selectbox("Chapa:", list(hist_db.keys()), key="adjm_ch")
                dfm = hist_db[ch2].copy()
                ent_pad2 = colab_by.get(ch2, {}).get("Entrada", "06:00")
                pode_sab2 = bool(colab_by.get(ch2, {}).get("Folga_Sab", False))
                subgrupo2 = (colab_by.get(ch2, {}).get("Subgrupo", "") or "").strip()

                nova_ent_mes = st.time_input("Nova entrada:", value=datetime.strptime(ent_pad2, "%H:%M").time(), key="adjm_ent")

                if st.button("Aplicar mês inteiro (e readequar)", key="adjm_apply"):
                    e = nova_ent_mes.strftime("%H:%M")
                    s = _saida_from_entrada(e)

                    for i in range(len(dfm)):
                        stt = dfm.loc[i, "Status"]
                        dia_num = int(pd.to_datetime(dfm.loc[i, "Data"]).day)
                        if stt in WORK_STATUSES:
                            dfm.loc[i, "Status"] = "Trabalho"
                            dfm.loc[i, "H_Entrada"] = e
                            dfm.loc[i, "H_Saida"] = s
                            set_override(setor, ano, mes, ch2, dia_num, "status", "Trabalho")
                            set_override(setor, ano, mes, ch2, dia_num, "h_entrada", e)
                            set_override(setor, ano, mes, ch2, dia_num, "h_saida", s)
                        else:
                            dfm.loc[i, "H_Entrada"] = ""
                            dfm.loc[i, "H_Saida"] = ""

                    update_colaborador_perfil(setor, ch2, subgrupo2, e, bool(pode_sab2))
                    save_escala_mes_db(setor, ano, mes, {ch2: dfm})
                    _regenerar_mes_inteiro(setor, ano, mes, seed=int(st.session_state.get("last_seed", 0)), respeitar_ajustes=True)
                    st.success("Horário do mês inteiro FORÇADO e escala readequada.")
                    st.rerun()

                st.dataframe(dfm, use_container_width=True, height=420)

            with t3:
                st.markdown("### ✅ Preferência por subgrupo (Evitar folga se possível)")
                subgrupos = list_subgrupos(setor)
                if subgrupos:
                    sg_sel = st.selectbox("Escolha o subgrupo:", subgrupos, key="pref_sg_sel")
                    regras = get_subgrupo_regras(setor, sg_sel)

                    p1, p2, p3 = st.columns(3)
                    ev_seg = p1.checkbox("Evitar SEG", value=bool(regras["seg"]), key=f"ev_seg_{sg_sel}")
                    ev_ter = p1.checkbox("Evitar TER", value=bool(regras["ter"]), key=f"ev_ter_{sg_sel}")
                    ev_qua = p2.checkbox("Evitar QUA", value=bool(regras["qua"]), key=f"ev_qua_{sg_sel}")
                    ev_qui = p2.checkbox("Evitar QUI", value=bool(regras["qui"]), key=f"ev_qui_{sg_sel}")
                    ev_sex = p3.checkbox("Evitar SEX", value=bool(regras["sex"]), key=f"ev_sex_{sg_sel}")
                    ev_sab = p3.checkbox("Evitar SÁB", value=bool(regras["sáb"]), key=f"ev_sab_{sg_sel}")

                    if st.button("Salvar preferência do subgrupo (e readequar mês)", key="pref_save"):
                        set_subgrupo_regras(setor, sg_sel, {
                            "seg": int(ev_seg), "ter": int(ev_ter), "qua": int(ev_qua),
                            "qui": int(ev_qui), "sex": int(ev_sex), "sáb": int(ev_sab)
                        })
                        _regenerar_mes_inteiro(setor, ano, mes, seed=int(st.session_state.get("last_seed", 0)), respeitar_ajustes=True)
                        st.success("Preferência salva e escala readequada!")
                        st.rerun()
                else:
                    st.info("Crie pelo menos 1 subgrupo na aba 👥 Colaboradores.")

            with t4:
                st.markdown("## 📌 Subgrupos (editável)")
                subgrupos = list_subgrupos(setor)

                cA, cB = st.columns([1, 1])
                with cA:
                    novo_sub = st.text_input("Novo subgrupo:", key="sg_new")
                    if st.button("Adicionar subgrupo", key="sg_add"):
                        if novo_sub.strip():
                            add_subgrupo(setor, novo_sub.strip())
                            st.success("Subgrupo adicionado!")
                            st.rerun()
                        else:
                            st.error("Digite o nome do subgrupo.")

                with cB:
                    if subgrupos:
                        del_sel = st.selectbox("Remover subgrupo:", ["(nenhum)"] + subgrupos, key="sg_del")
                        if del_sel != "(nenhum)" and st.button("Remover", key="sg_del_btn"):
                            delete_subgrupo(setor, del_sel)
                            _regenerar_mes_inteiro(setor, ano, mes, seed=int(st.session_state.get("last_seed", 0)), respeitar_ajustes=True)
                            st.success("Subgrupo removido e escala readequada!")
                            st.rerun()
                    else:
                        st.caption("Nenhum subgrupo cadastrado.")

    # ------------------------------------------------------
    # ABA 4: Férias
    # ------------------------------------------------------
    with abas[3]:
        st.subheader("🏖️ Controle de Férias")

        st.markdown("---")
        st.markdown("## 🗺️ Mapa anual de férias (visual)")
        col_map1, col_map2 = st.columns([1, 3])
        ano_mapa = col_map1.number_input("Ano do mapa", value=int(st.session_state.get("cfg_ano", datetime.now().year)), step=1, key="fer_mapa_ano")
        col_map2.caption("Mostra em quais meses cada colaborador tem férias cadastradas (qualquer dia no mês marca o mês).")
        df_mapa = ferias_mapa_ano_df(setor, int(ano_mapa), colaboradores)
        show_chapa = st.checkbox("Mostrar coluna Chapa no mapa", value=False, key="fer_mapa_show_chapa")
        df_mapa_show = df_mapa if show_chapa else df_mapa.drop(columns=["Chapa"])
        st.dataframe(style_ferias_mapa(df_mapa_show), use_container_width=True, height=420)
        st.markdown("---")
        colaboradores = load_colaboradores_setor(setor)

        if not colaboradores:
            st.warning("Sem colaboradores cadastrados.")
        else:
            chapas = [c["Chapa"] for c in colaboradores]
            st.markdown("### ➕ Lançar Férias")
            ch = st.selectbox("Chapa:", chapas, key="fer_ch")

            nome_sel = next((x.get("Nome","") for x in colaboradores if str(x.get("Chapa","")) == str(ch)), "")
            st.write(f"**Colaborador:** {nome_sel}  \n**Chapa:** {ch}")

            info_ult = get_ultima_ferias_info(setor, ch)
            ult_fim = info_ult.get("ultima_fim")
            meses_sem = info_ult.get("meses_desde_ultima_fim")

            if ult_fim:
                st.write(
                    f"**Últimas férias:** {info_ult.get('ultima_inicio').strftime('%d/%m/%Y')} até {ult_fim.strftime('%d/%m/%Y')}  \n"
                    f"**Duração:** {_classificar_duracao_ferias(int(info_ult.get('dias_ultima') or 0))}  \n"
                    f"**Tempo desde o fim:** {int(meses_sem)} mês(es)"
                )
            else:
                st.warning("⚠️ Este colaborador ainda NÃO tem férias cadastradas no sistema.")

            if meses_sem is not None and int(meses_sem) >= 23:
                st.error("🚨 ALERTA: colaborador está sem férias há 1 ano e 11 meses (ou mais). Priorize agendamento!")

            col1, col2 = st.columns(2)
            ini = col1.date_input("Início:", key="fer_ini")
            fim = col2.date_input("Fim:", key="fer_fim")

            try:
                qtd_dias_sel = (fim - ini).days + 1
            except Exception:
                qtd_dias_sel = 0
            st.info(f"📌 Duração selecionada: **{_classificar_duracao_ferias(int(qtd_dias_sel))}**")

            if st.button("Adicionar férias (e readequar mês)", key="fer_add"):
                if fim < ini:
                    st.error("Data final não pode ser menor que a inicial.")
                else:
                    add_ferias(setor, ch, ini, fim)
                    _regenerar_mes_inteiro(setor, int(st.session_state["cfg_ano"]), int(st.session_state["cfg_mes"]), seed=int(st.session_state.get("last_seed", 0)), respeitar_ajustes=True)
                    st.success("Férias adicionadas e escala readequada!")
                    st.rerun()

            st.markdown("---")
            st.markdown("### 📋 Férias cadastradas")

            rows = list_ferias(setor)

            if rows:
                df_f = pd.DataFrame(rows, columns=["Chapa", "Início", "Fim"])
                st.dataframe(df_f, use_container_width=True, height=420)

                st.markdown("### ❌ Remover férias")
                rem_idx = st.number_input("Linha para remover (1,2,3...)", min_value=1, max_value=len(df_f), value=1, key="fer_rem_idx")

                if st.button("Remover linha (e readequar mês)", key="fer_rem_btn"):
                    r = df_f.iloc[int(rem_idx) - 1]
                    delete_ferias_row(setor, r["Chapa"], r["Início"], r["Fim"])
                    _regenerar_mes_inteiro(setor, int(st.session_state["cfg_ano"]), int(st.session_state["cfg_mes"]), seed=int(st.session_state.get("last_seed", 0)), respeitar_ajustes=True)
                    st.success("Férias removidas e escala readequada!")
                    st.rerun()
            else:
                st.info("Nenhuma férias cadastrada.")

    # ------------------------------------------------------
    # ABA 5: Excel
    # ------------------------------------------------------
    with abas[4]:
        st.subheader("📥 Excel modelo RH (separado por subgrupo)")
        ano = int(st.session_state["cfg_ano"])
        mes = int(st.session_state["cfg_mes"])
        hist_db = load_escala_mes_db(setor, ano, mes)
        colaboradores = load_colaboradores_setor(setor)
        colab_by = {c["Chapa"]: c for c in colaboradores}

        if not hist_db:
            st.info("Gere a escala.")
        else:
            hist_db = apply_overrides_to_hist(setor, ano, mes, hist_db)

            if st.button("📊 Gerar Excel", key="xls_btn"):
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    wb = writer.book
                    ws = wb.create_sheet("Escala Mensal", index=0)

                    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", patternType="solid")
                    fill_dom = PatternFill(start_color="C00000", end_color="C00000", patternType="solid")
                    fill_folga = PatternFill(start_color="FFF2CC", end_color="FFF2CC", patternType="solid")
                    fill_nome = PatternFill(start_color="D9E1F2", end_color="D9E1F2", patternType="solid")
                    fill_ferias = PatternFill(start_color="92D050", end_color="92D050", patternType="solid")
                    fill_group = PatternFill(start_color="BDD7EE", end_color="BDD7EE", patternType="solid")

                    font_header = Font(color="FFFFFF", bold=True)
                    font_dom = Font(color="FFFFFF", bold=True)

                    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
                    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    ch0 = list(hist_db.keys())[0]
                    df_ref_xls = hist_db[ch0]
                    total_dias = len(df_ref_xls)

                    ws.cell(1, 1, "COLABORADOR").fill = fill_header
                    ws.cell(1, 1).font = font_header
                    ws.cell(1, 1).alignment = center
                    ws.cell(1, 1).border = border
                    ws.cell(2, 1, "").fill = fill_header
                    ws.cell(2, 1).alignment = center
                    ws.cell(2, 1).border = border

                    for i in range(total_dias):
                        dia_num = df_ref_xls.iloc[i]["Data"].day
                        dia_sem = df_ref_xls.iloc[i]["Dia"]
                        cA = ws.cell(1, i + 2, dia_num)
                        cB = ws.cell(2, i + 2, dia_sem)

                        if dia_sem == "dom":
                            cA.fill = fill_dom
                            cB.fill = fill_dom
                            cA.font = font_dom
                            cB.font = font_dom
                        else:
                            cA.fill = fill_header
                            cB.fill = fill_header
                            cA.font = font_header
                            cB.font = font_header

                        cA.alignment = center
                        cB.alignment = center
                        cA.border = border
                        cB.border = border
                        ws.column_dimensions[get_column_letter(i + 2)].width = 7

                    ws.column_dimensions["A"].width = 36

                    subgrupo_map = {}
                    for chx in hist_db.keys():
                        sg = (colab_by.get(chx, {}).get("Subgrupo", "") or "").strip() or "SEM SUBGRUPO"
                        subgrupo_map.setdefault(sg, []).append(chx)

                    subgrupos_ordenados = sorted(subgrupo_map.keys())
                    row_idx = 3

                    for sg in subgrupos_ordenados:
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_dias + 1)
                        t = ws.cell(row_idx, 1, f"SUBGRUPO: {sg}")
                        t.fill = fill_group
                        t.font = Font(bold=True)
                        t.alignment = Alignment(horizontal="left", vertical="center")
                        t.border = border
                        row_idx += 1

                        chapas_sg = sorted(subgrupo_map[sg], key=lambda chx: colab_by.get(chx, {}).get("Nome", chx))
                        for chx in chapas_sg:
                            df_f = hist_db[chx]
                            nome = colab_by.get(chx, {}).get("Nome", chx)

                            c_nome = ws.cell(row_idx, 1, f"{nome}\nCHAPA: {chx}")
                            c_nome.fill = fill_nome
                            c_nome.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                            c_nome.border = border
                            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx + 1, end_column=1)

                            for i, row in df_f.iterrows():
                                dia_sem = row["Dia"]
                                status = row["Status"]
                                if status == "Férias":
                                    v1, v2 = "FÉRIAS", ""
                                elif status == "Folga":
                                    v1, v2 = "F", ""
                                else:
                                    v1, v2 = row["H_Entrada"], row["H_Saida"]

                                cell1 = ws.cell(row_idx, i + 2, v1)
                                cell2 = ws.cell(row_idx + 1, i + 2, v2)

                                cell1.alignment = center
                                cell2.alignment = center
                                cell1.border = border
                                cell2.border = border

                                if status == "Férias":
                                    cell1.fill = fill_ferias
                                    cell2.fill = fill_ferias
                                elif status == "Folga":
                                    if dia_sem == "dom":
                                        cell1.fill = fill_dom
                                        cell2.fill = fill_dom
                                    else:
                                        cell1.fill = fill_folga
                                        cell2.fill = fill_folga

                            row_idx += 2
                        row_idx += 1

                    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                        wb.remove(wb["Sheet"])

                st.download_button(
                    "📥 Baixar Excel",
                    data=output.getvalue(),
                    file_name=f"escala_{setor}_{mes:02d}_{ano}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="xls_down"
                )


        # --- Lista (e PDF) de quem TRABALHA no dia escolhido ---
        st.markdown("### 🗓️ Quem trabalha no dia (para impressão)")
        try:
            dias_mes = calendar.monthrange(int(ano), int(mes))[1]
        except Exception:
            dias_mes = 31
        dia_sel = st.number_input("Dia do mês", min_value=1, max_value=int(dias_mes), value=1, step=1)

        # Monta tabela para visualização
        linhas = []
        for _chapa, _df in (hist_db or {}).items():
            if _df is None or _df.empty:
                continue
            try:
                _linha = _df.loc[_df["Data"].dt.day == int(dia_sel)].head(1)
            except Exception:
                _linha = _df.loc[pd.to_datetime(_df["Data"], errors="coerce").dt.day == int(dia_sel)].head(1)
            if _linha.empty:
                continue
            _r = _linha.iloc[0].to_dict()
            _stt = str(_r.get("Status", "")).strip()
            if _stt not in WORK_STATUSES:
                continue
            _ent = str(_r.get("H_Entrada", "") or "").strip()
            _sai = str(_r.get("H_Saida", "") or "").strip()
            # metadados do colaborador
            _nome = ""
            _subg = ""
            for c in colaboradores:
                if str(c.get("Chapa", "")).strip() == str(_chapa).strip():
                    _nome = str(c.get("Nome", "")).strip()
                    _subg = str(c.get("Subgrupo", "")).strip()
                    break
            linhas.append({"Chapa": str(_chapa).strip(), "Nome": _nome, "Subgrupo": _subg, "Entrada": _ent, "Saída": _sai})

        df_dia = pd.DataFrame(linhas).sort_values(["Subgrupo", "Nome"]) if linhas else pd.DataFrame(columns=["Chapa","Nome","Subgrupo","Entrada","Saída"])
        st.dataframe(df_dia, use_container_width=True, hide_index=True)

        colp1, colp2 = st.columns([1, 2])
        with colp1:
            if st.button("📄 Gerar PDF (quem trabalha no dia)"):
                if df_dia.empty:
                    st.warning("Não há colaboradores trabalhando nesse dia (ou ainda não foi gerado para este mês).")
                else:
                    pdf_bytes_dia = gerar_pdf_trabalhando_no_dia(setor, int(ano), int(mes), int(dia_sel), hist_db, colaboradores)
                    st.session_state["pdf_dia_trabalho_bytes"] = pdf_bytes_dia
                    st.success("PDF pronto.")
        with colp2:
            if st.session_state.get("pdf_dia_trabalho_bytes"):
                st.download_button(
                    "⬇️ Baixar PDF (quem trabalha no dia)",
                    data=st.session_state["pdf_dia_trabalho_bytes"],
                    file_name=f"escala_trabalhando_dia_{int(dia_sel):02d}_{int(mes):02d}_{int(ano)}.pdf",
                    mime="application/pdf",
                )


    

            

            st.markdown("---")
            st.markdown("## 🖨️ Impressão de Escala (PDF)")

            all_subgrupos = sorted({((c.get("Subgrupo") or "").strip() or "SEM SUBGRUPO") for c in colaboradores})
            cfx1, cfx2, cfx3 = st.columns([1.2, 1.2, 1.6])
            loja_txt = cfx1.text_input("Loja:", value=str(setor), key="pdf_loja_txt")
            secoes_sel = cfx2.multiselect("Seções (Subgrupo):", options=all_subgrupos, default=[], key="pdf_secoes_sel")
            busca_txt = cfx3.text_input("Filtro (nome/chapa/subgrupo):", value="", key="pdf_busca")

            cols_dates = st.columns([1,1,2])
            data_ini = cols_dates[0].date_input("Dia inicial:", value=date(int(ano), int(mes), 1), key="pdf_dt_ini")
            data_fim = cols_dates[1].date_input("Dia final:", value=date(int(ano), int(mes), calendar.monthrange(int(ano), int(mes))[1]), key="pdf_dt_fim")
            cols_dates[2].caption("Obs.: o PDF segue o modelo oficial do mês. Aqui o filtro é para escolher colaboradores/Seções como no sistema.")

            colabs_filtrados = _filtrar_colaboradores(colaboradores, secoes_sel, busca_txt)

            opcoes = [
                f"{(c.get('Nome') or '').strip()} — Chapa: {str(c.get('Chapa') or '').strip()} — {((c.get('Subgrupo') or '').strip() or 'SEM SUBGRUPO')}"
                for c in colabs_filtrados
            ]
            mapa_idx = {opcoes[i]: colabs_filtrados[i] for i in range(len(opcoes))}

            st.markdown("### 👥 Colaboradores")
            sel = st.multiselect(
                "Selecione (se vazio, imprime TODOS do filtro):",
                options=opcoes,
                default=[],
                key="pdf_colabs_sel"
            )

            cbtn1, cbtn2 = st.columns([1, 3])
            gerar = cbtn1.button("🖨️ Imprimir (gerar PDF)", key="pdf_print_btn", use_container_width=True)
            cbtn2.caption("Dica: selecione uma seção, depois marque os colaboradores. Se não marcar nenhum, imprime todos os filtrados.")

            if gerar:
                hist_db_pdf = load_escala_mes_db(setor, ano, mes)
                if not hist_db_pdf:
                    st.warning("Gere a escala antes na aba 🚀 Gerar Escala.")
                else:
                    hist_db_pdf = apply_overrides_to_hist(setor, ano, mes, hist_db_pdf)

                    if sel:
                        chapas_sel = [str(mapa_idx[x].get("Chapa")) for x in sel if x in mapa_idx]
                    else:
                        chapas_sel = [str(c.get("Chapa")) for c in colabs_filtrados]

                    hist_db_pdf = {ch: df for ch, df in hist_db_pdf.items() if ch in set(chapas_sel)}

                    if not hist_db_pdf:
                        st.warning("Nenhum colaborador para imprimir com os filtros atuais.")
                    else:
                        pdf_bytes = gerar_pdf_modelo_oficial(loja_txt.strip() or str(setor), ano, mes, hist_db_pdf, colaboradores)
                        st.download_button(
                            "⬇️ Baixar PDF",
                            data=pdf_bytes,
                            file_name=f"escala_{(loja_txt.strip() or str(setor))}_{mes:02d}_{ano}.pdf",
                            mime="application/pdf",
                            key="pdf_down"
                        )

    # ------------------------------------------------------
    # ABA 6: Admin (somente ADMIN)
    # ------------------------------------------------------
    if is_admin_area:
        with abas[5]:
            st.subheader("🔒 Admin do Sistema (somente ADMIN)")
            dfu = admin_list_users()
            st.dataframe(dfu, use_container_width=True, height=420)

            st.markdown("### Resetar senha de um usuário")
            if not dfu.empty:
                uid = st.selectbox("ID do usuário", dfu["id"].tolist(), key="adm_uid")
                newp = st.text_input("Nova senha", type="password", key="adm_newp")
                if st.button("Resetar senha", key="adm_reset"):
                    if not newp:
                        st.error("Digite a senha.")
                    else:
                        ok = admin_reset_user_password(int(uid), newp)
                        st.success("Senha resetada!" if ok else "Falha.")
                        st.rerun()


# =========================================================
# MAIN
# =========================================================
db_init()

if st.session_state["auth"] is None:
    page_login()
else:
    page_app()
